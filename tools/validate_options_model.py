"""Validate the institutional options workbook contract."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from workbook_engineering import audit_workbook

REQUIRED_SHEETS = {
    "Cover", "Assumptions", "European Pricer", "Implied Vol", "Greeks",
    "Vol Surface", "Portfolio", "Scenario P&L", "Checks", "Sources",
    "RefreshLog",
}
REQUIRED_LABELS = {
    "Put-call parity residual",
    "Converged implied volatility",
    "Portfolio total",
    "Delta-Gamma-Vega Scenario P&L",
    "Implied volatility convergence",
}


def validate(path: Path) -> list[str]:
    errors: list[str] = []
    workbook = load_workbook(path, data_only=False, keep_links=True)
    missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing:
        errors.append(f"missing sheets: {missing}")
    values = {
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    missing_labels = sorted(REQUIRED_LABELS - values)
    if missing_labels:
        errors.append(f"missing labels: {missing_labels}")
    formula_count = sum(
        1
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    if formula_count < 120:
        errors.append(f"formula depth below contract: {formula_count} < 120")
    summary, findings = audit_workbook(path)
    if summary["external_links"]:
        errors.append("external links present")
    for finding in findings:
        if finding.severity == "error":
            errors.append(f"{finding.code}:{finding.sheet}:{finding.cell}:{finding.message}")
    checks = workbook["Checks"] if "Checks" in workbook.sheetnames else None
    if checks is not None:
        check_formulas = [
            cell.value for row in checks.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if len(check_formulas) < 6:
            errors.append("insufficient explicit model checks")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    errors = validate(args.workbook)
    if errors:
        for error in errors:
            print(f"ERROR: {error}")
        return 1
    print(f"validated {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
