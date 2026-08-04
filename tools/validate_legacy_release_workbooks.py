"""Build and validate the six legacy release-grade workbooks."""
from __future__ import annotations

import argparse
import importlib
import json
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from tools.workbook_engineering import audit_workbook
except ModuleNotFoundError:
    from workbook_engineering import audit_workbook

ROOT = Path(__file__).resolve().parents[1]

SPECS: dict[str, dict[str, Any]] = {
    "01": {
        "domain": "Investment Banking",
        "module": "tools.builders.build_investment_banking_release",
        "filename": "investment_banking_release.xlsx",
        "required_sheets": [
            "Transaction Analysis",
            "Accretion Dilution",
            "Valuation Reconciliation",
            "Decision & Checks",
        ],
        "required_formulas": {
            "Transaction Analysis": ["C21", "C29", "D30"],
            "Accretion Dilution": ["C18", "D18"],
            "Decision & Checks": ["C5", "D7", "C12"],
        },
        "minimum_formulas": 60,
    },
    "02": {
        "domain": "Corporate Finance",
        "module": "tools.builders.build_corporate_finance_release",
        "filename": "corporate_finance_release.xlsx",
        "required_sheets": [
            "Treasury & Liquidity",
            "Capital Allocation",
            "Credit Metrics",
            "Decision & Checks",
        ],
        "required_formulas": {
            "Treasury & Liquidity": ["C21", "D24", "D26"],
            "Capital Allocation": ["C13", "D14"],
            "Decision & Checks": ["C5", "D7", "C12"],
        },
        "minimum_formulas": 50,
    },
    "05": {
        "domain": "Private Credit",
        "module": "tools.builders.build_private_credit_release",
        "filename": "private_credit_release.xlsx",
        "required_sheets": [
            "Portfolio & Concentration",
            "Amendment Economics",
            "Decision & Checks",
        ],
        "required_formulas": {
            "Portfolio & Concentration": ["C12", "D13", "D14"],
            "Amendment Economics": ["D23", "D25", "D28"],
            "Decision & Checks": ["C5", "D7", "C13"],
        },
        "minimum_formulas": 90,
    },
    "06": {
        "domain": "Debt Finance",
        "module": "tools.builders.build_debt_finance_release",
        "filename": "debt_finance_release.xlsx",
        "required_sheets": [
            "Maturity Ladder",
            "Refinancing & Rates",
            "Decision & Checks",
        ],
        "required_formulas": {
            "Maturity Ladder": ["C15", "C17", "C18"],
            "Refinancing & Rates": ["D26", "D28", "D32"],
            "Decision & Checks": ["C5", "D7", "C12"],
        },
        "minimum_formulas": 100,
    },
    "07": {
        "domain": "Public Finance",
        "module": "tools.builders.build_public_finance_release",
        "filename": "public_finance_release.xlsx",
        "required_sheets": [
            "Debt Sustainability",
            "Revenue Stress",
            "Decision & Checks",
        ],
        "required_formulas": {
            "Revenue Stress": ["D19", "D20", "D23"],
            "Decision & Checks": ["C5", "D7", "C12"],
        },
        "minimum_formulas": 100,
    },
    "13": {
        "domain": "Venture Capital",
        "module": "tools.builders.build_venture_capital_release",
        "filename": "venture_capital_release.xlsx",
        "required_sheets": [
            "Ownership & Dilution",
            "Reserves & Follow-ons",
            "Exit Waterfall",
            "Decision & Checks",
        ],
        "required_formulas": {
            "Ownership & Dilution": ["D15", "D17", "D20"],
            "Reserves & Follow-ons": ["D11", "D13"],
            "Exit Waterfall": ["D12", "D13"],
            "Decision & Checks": ["C5", "D7", "C13"],
        },
        "minimum_formulas": 65,
    },
}


def build_workbook(model_id: str, output: Path) -> Path:
    spec = SPECS[model_id]
    module = importlib.import_module(spec["module"])
    output.parent.mkdir(parents=True, exist_ok=True)
    module.build(output)
    if not output.exists():
        raise FileNotFoundError(output)
    return output


def formula_count(workbook) -> int:
    return sum(
        1
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


def validate_workbook(model_id: str, path: Path) -> dict[str, Any]:
    spec = SPECS[model_id]
    errors: list[str] = []
    workbook = load_workbook(path, data_only=False)
    missing = [name for name in spec["required_sheets"] if name not in workbook.sheetnames]
    if missing:
        errors.append(f"missing required sheets {missing}")
    formulas = formula_count(workbook)
    if formulas < spec["minimum_formulas"]:
        errors.append(
            f"formula depth {formulas} below minimum {spec['minimum_formulas']}"
        )
    for sheet_name, cells in spec["required_formulas"].items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for coordinate in cells:
            value = worksheet[coordinate].value
            if not isinstance(value, str) or not value.startswith("="):
                errors.append(f"{sheet_name}!{coordinate} must contain a formula")
    if workbook._external_links:
        errors.append("workbook contains external links")
    decision = workbook["Decision & Checks"] if "Decision & Checks" in workbook.sheetnames else None
    if decision is not None:
        overall_formulas = [
            cell.value
            for row in decision.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
            and cell.value.startswith("=")
            and "COUNTIF" in cell.value.upper()
        ]
        if not overall_formulas:
            errors.append("Decision & Checks lacks an aggregate status formula")
    summary, findings = audit_workbook(path)
    audit_errors = [item.__dict__ for item in findings if item.severity == "error"]
    if audit_errors:
        errors.append(f"engineering audit errors: {audit_errors}")
    return {
        "model_id": model_id,
        "domain": spec["domain"],
        "path": str(path),
        "sheets": len(workbook.sheetnames),
        "formulas": formulas,
        "audit_summary": summary,
        "errors": errors,
        "status": "PASS" if not errors else "FAIL",
    }


def build_and_validate(output_dir: Path) -> dict[str, Any]:
    results = []
    errors: list[str] = []
    for model_id, spec in SPECS.items():
        path = output_dir / spec["filename"]
        try:
            build_workbook(model_id, path)
            result = validate_workbook(model_id, path)
        except Exception as exc:
            result = {
                "model_id": model_id,
                "domain": spec["domain"],
                "path": str(path),
                "errors": [f"build or validation failed: {exc}"],
                "status": "FAIL",
            }
        results.append(result)
        errors.extend(f"{model_id}: {item}" for item in result["errors"])
    return {
        "schema_version": "1.0",
        "status": "PASS" if not errors else "FAIL",
        "models": len(SPECS),
        "results": results,
        "errors": errors,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument(
        "--report", type=Path, default=ROOT / "legacy-release-workbook-report.json"
    )
    args = parser.parse_args()
    if args.output_dir:
        output_dir = args.output_dir
        output_dir.mkdir(parents=True, exist_ok=True)
        report = build_and_validate(output_dir)
    else:
        with tempfile.TemporaryDirectory(prefix="legacy-release-workbooks-") as temp_name:
            report = build_and_validate(Path(temp_name))
    report_path = args.report if args.report.is_absolute() else ROOT / args.report
    report_path.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0 if report["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
