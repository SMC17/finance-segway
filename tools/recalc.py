"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice
"""

import contextlib
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula

MACRO_FILENAME = "Module1.xba"
SOFFICE_MISSING = "soffice not found on PATH; LibreOffice is required to recalculate"


def get_soffice_env() -> dict:
    """Headless LibreOffice needs a real (non-GUI) VCL backend to run in CI/containers."""
    env = os.environ.copy()
    env["SAL_USE_VCLPLUGIN"] = "svp"
    return env


def run_soffice(args, **kwargs) -> subprocess.CompletedProcess:
    args = list(args)
    if not any(str(a).startswith("-env:UserInstallation") for a in args):
        with tempfile.TemporaryDirectory(prefix="lo_profile_", ignore_cleanup_errors=True) as profile:
            args = [f"-env:UserInstallation={Path(profile).as_uri()}"] + args
            return subprocess.run(["soffice"] + args, env=get_soffice_env(), **kwargs)
    return subprocess.run(["soffice"] + args, env=get_soffice_env(), **kwargs)

MAX_LOCATIONS = 100

EXTERNAL_REF_RE = re.compile(r"""(?<![\w"\[])'?\[\d+\][^!"\[\]]*'?!""")

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def has_gtimeout():
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def _stamp(path):
    st = os.stat(path)
    return st.st_mtime_ns, st.st_size


def setup_libreoffice_macro(profile_dir: Path, timeout=30):
    url = profile_dir.as_uri()
    try:
        run_soffice(
            ["--headless", "--terminate_after_init", f"-env:UserInstallation={url}"],
            capture_output=True,
            timeout=timeout,
        )
    except FileNotFoundError:
        return None, SOFFICE_MISSING
    except subprocess.TimeoutExpired:
        return None, "LibreOffice timed out creating its profile; formulas were NOT recalculated"

    macro_dir = profile_dir / "user" / "basic" / "Standard"
    if not macro_dir.exists():
        return None, "LibreOffice did not create a usable profile; formulas were NOT recalculated"

    try:
        (macro_dir / MACRO_FILENAME).write_text(RECALCULATE_MACRO)
    except OSError as e:
        return None, f"Could not install the recalculation macro: {e}"

    return url, None


def _names_pattern(names):
    if not names:
        return None
    return re.compile(r"\b(" + "|".join(re.escape(n) for n in sorted(names)) + r")\b")


def external_links_at_risk(filename):
    try:
        with zipfile.ZipFile(filename) as archive:
            names = archive.namelist()
    except (zipfile.BadZipFile, OSError):
        return []
    if not any(n.startswith("xl/externalLinks/") for n in names):
        return []

    with contextlib.ExitStack() as stack:
        formulas = load_workbook(filename, data_only=False)
        stack.callback(formulas.close)
        values = load_workbook(filename, data_only=True)
        stack.callback(values.close)

        defined_names = list(formulas.defined_names.items())
        for scope_ws in formulas.worksheets:
            defined_names.extend(scope_ws.defined_names.items())
        name_texts = [
            (name, dn.value)
            for name, dn in defined_names
            if isinstance(getattr(dn, "value", None), str)
        ]
        external_names = {name for name, text in name_texts if EXTERNAL_REF_RE.search(text)}
        while True:
            name_re = _names_pattern(external_names)
            if name_re is None:
                break
            added = {
                name
                for name, text in name_texts
                if name not in external_names and name_re.search(text)
            }
            if not added:
                break
            external_names |= added
        name_re = _names_pattern(external_names)

        at_risk = []
        for sheet in formulas.sheetnames:
            ws = formulas[sheet]
            if not hasattr(ws, "iter_rows"):  
                continue
            cached = values[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, ArrayFormula):
                        v = v.text
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    reaches_out = EXTERNAL_REF_RE.search(v) or (name_re and name_re.search(v))
                    if reaches_out and cached[cell.coordinate].value is None:
                        at_risk.append(f"{sheet}!{cell.coordinate}")
        return at_risk


def recalc(filename, timeout=30, force=False):
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not os.access(abs_path, os.W_OK):
        return {"error": f"{filename} is not writable; recalculation rewrites the file in place"}

    try:
        get_soffice_env()
    except Exception as e:  
        return {"error": f"Could not prepare the LibreOffice environment: {e}"}

    if not force:
        try:
            at_risk = external_links_at_risk(filename)
        except Exception as e:  
            return {"error": f"Could not inspect {filename} for external links: {e}"}
        if at_risk:
            shown = at_risk[:MAX_LOCATIONS]
            return {
                "error": (
                    "Refusing to recalculate: this workbook links to another workbook, and "
                    f"{len(at_risk)} linked cell(s) have lost their cached value (openpyxl strips "
                    "these on save). Recalculating would resolve them to #NAME? and delete the "
                    "external links for good. Copy those cells' values from the original file "
                    "before saving, or pass --force to accept the loss. Charts and conditional "
                    "formats can hold external references too, so this list may not be exhaustive."
                ),
                "external_link_cells": shown,
                "external_link_cells_truncated": max(0, len(at_risk) - len(shown)),
            }

    with tempfile.TemporaryDirectory(
        prefix="recalc-lo-profile-", ignore_cleanup_errors=True
    ) as profile_dir:
        return _recalc_with_profile(filename, abs_path, timeout, Path(profile_dir))


def _recalc_with_profile(filename, abs_path, timeout, profile_dir: Path):
    started = time.monotonic()
    deadline = started + timeout
    profile_url, err = setup_libreoffice_macro(profile_dir, timeout=timeout)
    if err:
        return {"error": err}

    remaining = max(5, int(deadline - time.monotonic()))
    # Preserve enough of the caller's budget for a conversion-based fallback.
    # Some container builds initialize correctly but hang on Basic macro dispatch.
    macro_timeout = max(5, min(30, remaining // 2 if remaining >= 10 else remaining))

    before = _stamp(abs_path)

    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        f"-env:UserInstallation={profile_url}",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    if platform.system() == "Linux" and shutil.which("timeout"):
        cmd = ["timeout", str(macro_timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(macro_timeout)] + cmd

    timed_out = f"LibreOffice macro dispatch timed out after {macro_timeout}s"

    macro_error = None
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            env=get_soffice_env(),
            timeout=macro_timeout + 15,
        )
    except subprocess.TimeoutExpired:
        macro_error = timed_out
    except FileNotFoundError:
        return {"error": SOFFICE_MISSING}
    else:
        if result.returncode == 124:
            macro_error = timed_out
        elif result.returncode != 0:
            detail = (result.stderr or "").strip() or f"soffice exited {result.returncode}"
            macro_error = f"LibreOffice macro recalculation failed: {detail}"
        elif _stamp(abs_path) == before:
            macro_error = "LibreOffice macro exited without rewriting the workbook"

    method = "basic_macro"
    if macro_error:
        remaining = max(0, int(deadline - time.monotonic()))
        fallback_error = _recalc_via_conversion(
            abs_path,
            timeout=remaining,
            profile_dir=profile_dir / "conversion-profile",
        )
        if fallback_error:
            return {
                "error": (
                    f"{macro_error}; conversion fallback failed: {fallback_error}. "
                    "Formulas were NOT recalculated."
                )
            }
        method = "conversion_fallback"

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not hasattr(ws, "iter_rows"):  
                continue
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                entry = {"count": len(locations), "locations": locations[:MAX_LOCATIONS]}
                if len(locations) > MAX_LOCATIONS:
                    entry["locations_truncated"] = len(locations) - MAX_LOCATIONS
                result["error_summary"][err_type] = entry

        wb.close()

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            if not hasattr(ws, "iter_rows"):  
                continue
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, ArrayFormula):
                        v = v.text
                    if isinstance(v, str) and v.startswith("="):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count
        result["recalculation_method"] = method

        return result

    except Exception as e:
        return {"error": str(e)}


def _recalc_via_conversion(abs_path: str, timeout: int, profile_dir: Path) -> str | None:
    """Recalculate an .xlsx by opening and exporting it with a clean profile."""

    source = Path(abs_path)
    if source.suffix.lower() != ".xlsx":
        return f"conversion fallback is restricted to .xlsx files, got {source.suffix}"
    if timeout < 5:
        return "no timeout budget remained"
    output_dir = profile_dir.parent / "converted"
    output_dir.mkdir(parents=True, exist_ok=True)
    profile_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / source.name
    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        f"-env:UserInstallation={profile_dir.as_uri()}",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        str(source),
    ]
    if platform.system() == "Linux" and shutil.which("timeout"):
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            env=get_soffice_env(),
            timeout=timeout + 15,
        )
    except subprocess.TimeoutExpired:
        return f"LibreOffice conversion timed out after {timeout}s"
    except FileNotFoundError:
        return SOFFICE_MISSING
    if result.returncode == 124:
        return f"LibreOffice conversion timed out after {timeout}s"
    if result.returncode != 0:
        detail = (result.stderr or "").strip() or f"soffice exited {result.returncode}"
        return detail
    if not output.exists() or output.stat().st_size == 0:
        return "LibreOffice conversion produced no workbook"
    # The LibreOffice profile normally lives under /tmp while a repository
    # can be mounted on another filesystem. os.replace cannot cross that
    # boundary, so stage a sibling copy and retain an atomic final replace.
    staged: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{source.stem}-recalc-",
            suffix=source.suffix,
            dir=source.parent,
            delete=False,
        ) as staged_handle:
            staged = Path(staged_handle.name)
            with output.open("rb") as converted_handle:
                shutil.copyfileobj(converted_handle, staged_handle)
            staged_handle.flush()
            os.fsync(staged_handle.fileno())
        os.replace(staged, source)
    except OSError as error:
        if staged is not None:
            with contextlib.suppress(OSError):
                staged.unlink()
        return f"could not install recalculated workbook: {error}"
    return None


def main():
    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv[1:]

    if not args:
        print("Usage: python recalc.py <excel_file> [timeout_seconds] [--force]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        print("\nOn any failure the JSON has an 'error' key and no 'status'.")
        print("--force recalculates even when it would destroy external links.")
        sys.exit(1)

    filename = args[0]
    try:
        timeout = int(args[1]) if len(args) > 1 else 30
    except ValueError:
        result = {"error": f"timeout must be an integer number of seconds, got {args[1]!r}"}
        print(json.dumps(result, indent=2))
        sys.exit(1)

    result = recalc(filename, timeout, force=force)
    print(json.dumps(result, indent=2))
    sys.exit(1 if "error" in result else 0)


if __name__ == "__main__":
    main()
