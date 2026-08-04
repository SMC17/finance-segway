"""Validate the institutional fixed-income and rates workbook contract."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from workbook_engineering import audit_workbook

REQUIRED_SHEETS = {
    "Cover", "Assumptions", "Zero Curve", "Bond Analytics", "Portfolio",
    "Key Rate Risk", "Carry & Roll", "Scenarios", "P&L Explain", "Checks",
    "Sources", "RefreshLog",
}
REQUIRED_LABELS = {
    "Zero Curve and Discount Factors",
    "Numerical modified duration",
    "Numerical convexity",
    "Portfolio key-rate DV01",
    "One-Year Carry and Roll-Down",
    "Curve and Spread Scenarios",
    "Unexplained P&L",
}


def validate(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    errors: list[str] = []
    missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing:
        errors.append(f"missing sheets: {missing}")
    values = {
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row if cell.value is not None
    }
    missing_labels = sorted(REQUIRED_LABELS - values)
    if missing_labels:
        errors.append(f"missing labels: {missing_labels}")
    formulas = [
        cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    if len(formulas) < 150:
        errors.append(f"formula depth below contract: {len(formulas)} < 150")
    portfolio = workbook["Portfolio"] if "Portfolio" in workbook.sheetnames else None
    if portfolio is not None:
        expected_headers = [
            "Security", "Market value", "Coupon", "Yield", "Maturity",
            "Spread duration", "Price", "Mod duration", "Convexity", "DV01",
        ]
        actual = [portfolio.cell(4, column).value for column in range(2, 12)]
        if actual != expected_headers:
            errors.append(f"portfolio column contract mismatch: {actual}")
        if any(portfolio.cell(row, 12).value is not None for row in range(5, 13)):
            errors.append("uncontracted duplicate data remain in portfolio column L")
    summary, findings = audit_workbook(path)
    if summary["external_links"]:
        errors.append("external links present")
    errors.extend(
        f"{finding.code}:{finding.sheet}:{finding.cell}:{finding.message}"
        for finding in findings if finding.severity == "error"
    )
    checks = workbook["Checks"] if "Checks" in workbook.sheetnames else None
    if checks is not None:
        check_formulas = [
            cell.value for row in checks.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if len(check_formulas) < 8:
            errors.append("insufficient fixed-income checks")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    errors = validate(args.workbook)
    if errors:
        for error in errors:
            print(f"ERROR: {error}")
        return 1
    print(f"validated {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
