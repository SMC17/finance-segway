"""Measure how much of each public case's template is real vs. still template default.

A manifest can honestly declare "observed"/"derived" for a handful of cells
and stop there -- the workbook still looks fully populated because every
template default is already filled in with a plausible-looking number. The
only way to tell "real, sourced financial model" from "three real cells
wearing a fully-built template" is to compare what the template actually
offers as an editable input against what the manifest actually overrides.

This script does that for every case in standards/public_cases/index.json:

  1. Scan the case's template workbook for "input cells" -- cells styled
     with the blue-font/yellow-fill signature every builder in this repo
     uses for a hardcoded, modeler-editable input (see
     tools/builders/institutional_helpers.py:input_cell). These are the
     cells a real instance COULD override with sourced data.
  2. Exclude cells in a column literally labeled "Downside" in the nearest
     header row above them -- a stress-case lever is definitionally
     hypothetical, not a fact to source, and every real case in this repo
     only ever overrides the Base column. Also exclude the Cover sheet's
     fixed "Blue text / yellow fill" legend cell, which uses the same
     styling to demonstrate the convention, not as an actual input.
  3. Compare that candidate set against the manifest's actual "observed"/
     "derived" inputs (standards/public_cases/<case_id>.json) to compute a
     real-data coverage ratio.

Coverage is NOT a pass/fail gate: plenty of candidate cells are legitimately
permanent template defaults (an LBO's illustrative entry multiple, a stress
case's downside markdown) that will never be a disclosed fact for a
hypothetical transaction. What this script gives is honest visibility --
which cases are thin, and thin on what -- so nothing ships looking deeper
than it is, and thin cases are a visible, prioritizable punch list instead
of something a reviewer has to discover cell-by-cell.

Usage:
    python tools/verify_template_exhaustion.py
    python tools/verify_template_exhaustion.py --report report.json
    python tools/verify_template_exhaustion.py --case-id pe-public-home-depot-2023
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INDEX = ROOT / "standards" / "public_cases" / "index.json"

INPUT_FONT_RGB = "000000FF"
LEGEND_EXCLUDE_TEXT = "Blue text / yellow fill"

# Governance surfaces, not model-input surfaces. These carry blue-font cells
# (the repo's input marker) but are provenance and audit apparatus rather
# than places a modeler enters decision-driving data: the Sources sheet is
# auto-populated from the manifest's own sources list by
# model_instances._append_sources, RefreshLog is an append-only change log,
# and the institutional-surface trio is generated commentary. Counting them
# as sourceable inputs inflates every case's denominator with cells no
# amount of real-data work is supposed to fill by hand.
NON_INPUT_SHEETS = {
    "Sources",
    "RefreshLog",
    "Institutional Surface",
    "Challenge Log",
    "Lineage Map",
}


def _rgb(color: Any) -> str | None:
    try:
        return color.rgb if color else None
    except (AttributeError, ValueError, TypeError):
        return None


def find_candidate_cells(template_path: Path) -> dict[str, set[str]]:
    """Return {sheet_name: {cell_coord, ...}} of real-data-eligible cells.

    The signal is blue font color (000000FF) -- this repo's universal
    modeling-convention marker for "hardcoded input," per every template's
    own Cover-sheet legend ("Blue text / yellow fill" = input; black =
    same-sheet formula; green = cross-sheet formula). Fill color is not
    part of the detection: institutional_helpers.input_cell() uses a pale
    yellow (FFF2CC) while the older legacy_frontier_release archetype uses
    pure yellow (FFFF00), and some legacy cells carry no fill at all --
    checking fill would silently miss real input cells in whichever
    convention isn't matched. Verified empirically to produce zero false
    positives from formula cells across both conventions.

    Both Base and Downside columns count as candidates (no column-label
    filtering): an adversarial/stress case's whole point is that its
    "Downside" column often holds real, disclosed distress-scenario data,
    not a hypothetical modeler lever -- excluding it undercounted real,
    sourced cells for exactly the cases meant to be the most real.

    A styled cell counts even when it holds no value. A table that reserves
    N rows of capacity (an ETF holdings list, a comps peer set) but ships
    illustrative defaults for only the first few still *offers* all N rows
    for real data -- and a real instance fills them. Skipping blank cells
    made those reserved slots invisible, so real sourced data landing in
    them registered as "outside the candidate set" rather than as coverage.
    """
    workbook = load_workbook(template_path, data_only=False)
    result: dict[str, set[str]] = {}
    for sheet in workbook.worksheets:
        if sheet.title in NON_INPUT_SHEETS:
            continue
        candidates: set[str] = set()
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value) == LEGEND_EXCLUDE_TEXT:
                    continue
                if _rgb(cell.font.color if cell.font else None) != INPUT_FONT_RGB:
                    continue
                candidates.add(cell.coordinate)
        if candidates:
            result[sheet.title] = candidates
    return result


def _cover_real_cells(manifest: dict[str, Any], template_path: Path) -> set[tuple[str, str]]:
    """Cover-sheet cells the manifest's own "cover" dict overrides.

    model_instances.py applies these through a separate code path
    (_set_cover, matched by row label) from the "inputs" list, so without
    this they're invisible to the coverage scanner even though they are
    real, case-specific facts (e.g. the actual company/deal name). Two
    labels are excluded deliberately: "Last refreshed:" and "Active
    scenario:" are unconditionally populated for every case regardless of
    how much real sourcing went into it (an as_of date and a Base/Downside
    toggle, not a disclosed fact about the subject), so counting them
    would inflate every case's coverage by the same fixed, meaningless
    amount rather than reflecting real work.
    """
    always_populated = {"Last refreshed:", "Active scenario:"}
    cover = {
        label: value for label, value in manifest.get("cover", {}).items()
        if label not in always_populated
    }
    if not cover:
        return set()
    workbook = load_workbook(template_path, data_only=False)
    if "Cover" not in workbook.sheetnames:
        return set()
    sheet = workbook["Cover"]
    cells: set[tuple[str, str]] = set()
    for label in cover:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.column == 2 and str(cell.value or "").strip() == label:
                    cells.add(("Cover", sheet.cell(cell.row, 3).coordinate))
    return cells


def case_coverage(case: dict[str, Any]) -> dict[str, Any]:
    manifest_path = ROOT / case["manifest"]
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    template_path = ROOT / manifest["template"]

    candidates_by_sheet = find_candidate_cells(template_path)
    total_candidates = sum(len(cells) for cells in candidates_by_sheet.values())

    real_cells: set[tuple[str, str]] = {
        (item["sheet"], item["cell"])
        for item in manifest.get("inputs", [])
        if item.get("input_kind") in {"observed", "derived"}
    }
    real_cells |= _cover_real_cells(manifest, template_path)
    real_in_candidates = {
        (sheet, cell) for sheet, cell in real_cells if cell in candidates_by_sheet.get(sheet, set())
    }
    real_outside_candidates = sorted(
        f"{sheet}!{cell}" for sheet, cell in real_cells - real_in_candidates
    )

    # Cells that are not point observations but scenario drivers. The
    # modeling standard is that a model never contains a hardcoded number
    # that isn't real: where no disclosed value exists (Home Depot was never
    # the subject of a leveraged buyout, so there is no real entry multiple),
    # the cell must be a driver whose range is grounded in real comparable
    # data -- not an invented point estimate wearing a rationale.
    #
    # So a declaration only earns credit when it carries a `basis` naming
    # the real data that grounds the range. A declaration without one is
    # reported separately as `unsourced_driver`: it is known to be a driver,
    # but its range has not been grounded yet. That is visible remaining
    # work, deliberately NOT counted as accounted-for.
    #
    # Two forms are recognised:
    #   1. `driver_declarations` entries (sheet, cell, driver_type, basis)
    #   2. a manifest input typed `modeler_assumption` whose source carries a
    #      real url -- the pre-existing convention for the same idea
    sourced_drivers: set[tuple[str, str]] = set()
    unsourced_drivers: set[tuple[str, str]] = set()

    for entry in manifest.get("driver_declarations", []):
        key = (entry["sheet"], entry["cell"])
        basis = entry.get("basis") or {}
        has_real_basis = bool(str(basis.get("source_url", "")).strip()) or bool(
            str(basis.get("source_name", "")).strip()
        )
        (sourced_drivers if has_real_basis else unsourced_drivers).add(key)

    for item in manifest.get("inputs", []):
        if item.get("input_kind") != "modeler_assumption":
            continue
        key = (item["sheet"], item["cell"])
        if key in sourced_drivers or key in unsourced_drivers:
            continue
        source = item.get("source") or {}
        url = str(source.get("url", "")).strip()
        # A repo:// self-reference is the manifest pointing at itself, which
        # grounds nothing. Only an external source counts.
        if url and not url.startswith("repo://"):
            sourced_drivers.add(key)
        else:
            unsourced_drivers.add(key)

    def _in_candidates(keys: set[tuple[str, str]]) -> set[tuple[str, str]]:
        return {
            key for key in keys if key[1] in candidates_by_sheet.get(key[0], set())
        } - real_in_candidates

    sourced_driver_cells = _in_candidates(sourced_drivers)
    unsourced_driver_cells = _in_candidates(unsourced_drivers) - sourced_driver_cells

    all_candidates = {
        (sheet, cell) for sheet, cells in candidates_by_sheet.items() for cell in cells
    }
    untouched = (
        all_candidates - real_in_candidates - sourced_driver_cells - unsourced_driver_cells
    )
    unaccounted = sorted(
        f"{sheet}!{cell}" for sheet, cell in untouched | unsourced_driver_cells
    )

    coverage = len(real_in_candidates) / total_candidates if total_candidates else 0.0
    accounted = len(real_in_candidates) + len(sourced_driver_cells)
    accounted_ratio = accounted / total_candidates if total_candidates else 0.0
    by_sheet = {
        sheet: {
            "candidates": len(cells),
            "real": len([1 for s, c in real_in_candidates if s == sheet]),
            "sourced_driver": len([1 for s, c in sourced_driver_cells if s == sheet]),
            "unsourced_driver": len([1 for s, c in unsourced_driver_cells if s == sheet]),
        }
        for sheet, cells in candidates_by_sheet.items()
    }

    return {
        "case_id": case["case_id"],
        "model_id": case["model_id"],
        "domain": case["domain"],
        "case_type": case["case_type"],
        "total_candidate_cells": total_candidates,
        "real_cells": len(real_in_candidates),
        "sourced_driver_cells": len(sourced_driver_cells),
        "unsourced_driver_cells": len(unsourced_driver_cells),
        "untouched_cells": len(untouched),
        "unaccounted_cells": len(unaccounted),
        "coverage": round(coverage, 4),
        "accounted_ratio": round(accounted_ratio, 4),
        "real_cells_outside_candidate_set": real_outside_candidates,
        "unaccounted": unaccounted,
        "by_sheet": by_sheet,
    }


def run(case_id: str | None = None) -> dict[str, Any]:
    index = json.loads(INDEX.read_text(encoding="utf-8"))
    cases = index["cases"]
    if case_id:
        cases = [item for item in cases if item["case_id"] == case_id]
        if not cases:
            raise ValueError(f"unknown case_id {case_id!r}")

    results = [case_coverage(case) for case in cases]
    results.sort(key=lambda item: (item["accounted_ratio"], item["coverage"]))

    anomalies = [
        f"{item['case_id']}: real inputs outside the candidate set: {item['real_cells_outside_candidate_set']}"
        for item in results
        if item["real_cells_outside_candidate_set"]
    ]

    return {
        "cases_measured": len(results),
        "mean_coverage": round(sum(item["coverage"] for item in results) / len(results), 4) if results else 0.0,
        "mean_accounted": round(
            sum(item["accounted_ratio"] for item in results) / len(results), 4
        )
        if results
        else 0.0,
        "total_unaccounted_cells": sum(item["unaccounted_cells"] for item in results),
        "fully_accounted_cases": [
            item["case_id"] for item in results if item["unaccounted_cells"] == 0
        ],
        "anomalies": anomalies,
        "results": results,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--report", type=Path, default=None)
    parser.add_argument("--case-id", type=str, default=None)
    args = parser.parse_args()

    report = run(args.case_id)

    header = (
        f"{'case_id':<45} {'type':<13} {'real':>6} {'drv+':>5} {'drv-':>5} "
        f"{'none':>5} {'of':>5} {'real%':>7} {'acct%':>7}"
    )
    print(header)
    print("-" * len(header))
    for item in report["results"]:
        print(
            f"{item['case_id']:<45} {item['case_type']:<13} "
            f"{item['real_cells']:>6} {item['sourced_driver_cells']:>5} "
            f"{item['unsourced_driver_cells']:>5} {item['untouched_cells']:>5} "
            f"{item['total_candidate_cells']:>5} "
            f"{item['coverage'] * 100:>6.1f}% {item['accounted_ratio'] * 100:>6.1f}%"
        )
    print("-" * len(header))
    print(
        f"Cases: {report['cases_measured']}  "
        f"Mean real: {report['mean_coverage'] * 100:.1f}%  "
        f"Mean accounted: {report['mean_accounted'] * 100:.1f}%  "
        f"Cells still needing work: {report['total_unaccounted_cells']}  "
        f"Fully accounted: {len(report['fully_accounted_cases'])}/{report['cases_measured']}"
    )
    print(
        "\nStates, per the modeling standard that a model never holds a hardcoded "
        "number that isn't real:\n"
        "  real  = sourced from disclosed data (an observation or a derivation of one)\n"
        "  drv+  = scenario driver whose range is grounded in named real data\n"
        "  drv-  = known to be a driver, range NOT yet grounded -- remaining work\n"
        "  none  = neither sourced nor declared -- remaining work\n"
        "acct% counts real + drv+ only. drv- deliberately earns no credit: a driver "
        "without a grounded range\nis still an ungrounded number in a model."
    )
    if report["anomalies"]:
        print("\nAnomalies (real inputs the scanner didn't recognize as template input cells):")
        for line in report["anomalies"]:
            print(f"  {line}")

    if args.case_id and report["results"] and report["results"][0]["unaccounted"]:
        print(f"\nUnaccounted cells for {args.case_id}:")
        for coordinate in report["results"][0]["unaccounted"]:
            print(f"  {coordinate}")

    if args.report:
        args.report.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
