"""Release-grade wrapper around manifest-driven model instance generation."""
from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:  # Package import used by tests and module consumers.
    from tools.model_instances import apply_manifest as apply_base_manifest
except ModuleNotFoundError:  # Direct CLI execution with tools/ on PYTHONPATH.
    from model_instances import apply_manifest as apply_base_manifest


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    digest.update(path.read_bytes())
    return digest.hexdigest()


def apply_manifest(manifest_path: Path, root: Path) -> dict[str, Any]:
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("classification") == "external_historical_case":
        benchmark_prefix = "repo://standards/benchmark_cases/"
        contaminated = [
            f"{item.get('sheet')}!{item.get('cell')}"
            for item in manifest.get("inputs", [])
            if str((item.get("source") or {}).get("url", "")).startswith(
                benchmark_prefix
            )
        ]
        if contaminated:
            raise ValueError(
                "public manifest contains synthetic benchmark lineage: "
                + ", ".join(contaminated)
            )
    receipt = apply_base_manifest(manifest_path, root)
    output = (root / manifest["output"]).resolve()
    workbook = load_workbook(output, data_only=False, keep_links=True)

    if "RefreshLog" in workbook.sheetnames:
        refresh = workbook["RefreshLog"]
        headers = {
            str(refresh.cell(4, column).value or "").strip().lower(): column
            for column in range(2, refresh.max_column + 1)
        }
        row = 5
        while any(
            refresh.cell(row, column).value not in (None, "")
            for column in range(2, refresh.max_column + 1)
        ):
            row += 1
        entry = manifest.get("refresh") or {}
        values = {
            "date": entry.get("date", manifest["as_of"]),
            "trigger": entry.get("trigger", "Manifest generation"),
            "source snapshot": entry.get(
                "source_snapshot",
                f"repo://standards/public_cases/{manifest['id']}.json",
            ),
            "what changed": entry.get(
                "what_changed", f"Applied manifest {manifest['id']}"
            ),
            "reviewer / challenge": entry.get(
                "reviewer_notes", "Generated reproducibly"
            ),
            "reviewer notes": entry.get(
                "reviewer_notes", "Generated reproducibly"
            ),
            "next check": entry.get(
                "next_check", "On builder or contract change"
            ),
        }
        # The base generator may have populated a legacy row. Remove any row
        # whose trigger matches before writing the canonical six-column entry.
        for candidate in range(5, row):
            trigger_column = headers.get("trigger")
            if trigger_column and refresh.cell(candidate, trigger_column).value == values["trigger"]:
                for column in range(2, refresh.max_column + 1):
                    refresh.cell(candidate, column).value = None
                row = candidate
                break
        for header_name, column in headers.items():
            if header_name in values:
                refresh.cell(row, column, values[header_name])

    workbook.save(output)
    receipt_path = output.with_suffix(".receipt.json")
    receipt["workbook_sha256"] = _sha256(output)
    receipt["refresh_schema"] = "source-addressed-v2"
    receipt["classification"] = manifest.get("classification")
    receipt["counts_toward_M4"] = manifest.get("counts_toward_M4")
    if receipt["counts_toward_M4"] is not False:
        raise ValueError("released public evidence must explicitly exclude M4 credit")
    receipt_path.write_text(
        json.dumps(receipt, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    return receipt
