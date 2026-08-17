"""L2 data fabric: pull Aswath Damodaran's (NYU Stern) public industry
cost-of-capital / multiples datasets and the global country equity risk
premium table, and emit structured facts + provenance.

These are the standard academic/practitioner source for exactly the inputs
this repo's valuation-heavy domains (Investment Banking, Corporate Finance,
Private Equity, Merchant Banking, Asset Management, ...) need and cannot get
from a single company's own filings: industry-average beta, cost of capital,
margins, and trading multiples, plus per-country equity risk premiums for
cross-border WACC. Damodaran updates these datasets himself, dated, in a
fixed public location -- no API key, no scraping workaround needed.

Two shapes are handled:
  - "industry" files (wacc, betas, margin, pedata, pbvdata, psdata, ...):
    an "Industry Averages" sheet, header row starting "Industry Name".
  - "country premium" (ctryprem): an "ERPs by country" sheet, header row
    starting "Country".

Legacy .xls files need `xlrd` (not `openpyxl`, which only reads the modern
OOXML format); .xlsx files (ctryprem's current form) use `openpyxl`, same as
the rest of this repo.

Rules (repo policy):
- Public data only.
- Every fact lands with as-of, retrieval date, transformation, and snapshot pointer.
- Domain builders remain the calculation engine; this layer only supplies inputs.

Examples:
  python tools/data_fabric/damodaran_public_facts.py --dataset wacc
  python tools/data_fabric/damodaran_public_facts.py --dataset pedata --industry Advertising
  python tools/data_fabric/damodaran_public_facts.py --dataset ctryprem --country "United States"
"""
from __future__ import annotations

import argparse
import csv
import json
import subprocess
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "tools" / "data_fabric" / "out"
OUT_DIR.mkdir(parents=True, exist_ok=True)
DOWNLOAD_DIR = OUT_DIR / "_damodaran_raw"

BASE_URL = "https://pages.stern.nyu.edu/~adamodar/pc/datasets"

# US-market industry-average files. All share the "Industry Averages" sheet
# / "Industry Name" header-row shape. Regional variants (Europe/Japan/China/
# emerg/India/Global/Rest) exist under the same stem + region suffix and
# work with --dataset <stem><Region>, e.g. --dataset waccEurope.
INDUSTRY_DATASETS = {
    "wacc": ("wacc.xls", "Cost of capital by industry (beta, cost of equity/debt, D/E, WACC)"),
    "betas": ("betas.xls", "Unlevered/levered beta by industry"),
    "margin": ("margin.xls", "Operating and net margin by industry"),
    "pedata": ("pedata.xls", "PE ratios (current/trailing/forward) by industry"),
    "pbvdata": ("pbvdata.xls", "Price-to-book by industry"),
    "psdata": ("psdata.xls", "Price-to-sales by industry"),
    "capex": ("capex.xls", "Capital expenditure by industry"),
}

COUNTRY_PREMIUM_DATASET = ("ctryprem.xlsx", "Country equity risk premiums")


def _download(filename: str) -> Path:
    """Fetch via curl (same TLS-edge rationale as treasury_public_facts.py
    and resolve_from_fred.py: some public-data hosts reject Python's HTTP
    stack at the TLS layer). Cached under out/_damodaran_raw/ by filename;
    Damodaran updates these files in place at a fixed URL, so a cached copy
    from an earlier run in the same day is fine, but re-running always
    re-fetches -- these are small (tens of KB) and the point is freshness."""
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    dest = DOWNLOAD_DIR / filename
    url = f"{BASE_URL}/{filename}"
    result = subprocess.run(
        ["curl", "-sSgL", "-m", "30", "-o", str(dest), url],
        capture_output=True, text=True,
    )
    if result.returncode != 0 or not dest.exists() or dest.stat().st_size == 0:
        raise SystemExit(f"failed to fetch {url}: {result.stderr.strip()}")
    return dest


def rows_from_grid(grid: list[list[Any]], first_col_label: str) -> list[dict[str, Any]]:
    """Locate the header row (the row whose first cell equals `first_col_label`
    exactly -- both Damodaran sheet shapes have several metadata/notes rows
    above the real table, at a row offset that has moved before and will
    again) and turn every row below it into a dict keyed by that header row's
    column labels. Blank-key columns (Damodaran's sheets carry stray notes/
    blank columns past the real table) are dropped; rows with a blank first
    cell (spacer rows, sheet end) are skipped.

    Pure function over an in-memory grid so it's testable without a real
    binary .xls/.xlsx fixture -- xlrd's open_workbook and openpyxl's
    load_workbook each hand this function a `list[list]` view of their sheet.
    """
    header_row = next(
        (r for r, row in enumerate(grid) if str(row[0] if row else "").strip() == first_col_label),
        None,
    )
    if header_row is None:
        raise SystemExit(f"no header row starting {first_col_label!r} found ({len(grid)} rows scanned)")
    headers = [str(c).strip() if c is not None else "" for c in grid[header_row]]
    rows: list[dict[str, Any]] = []
    for raw in grid[header_row + 1:]:
        if not raw or not str(raw[0] if raw[0] is not None else "").strip():
            continue
        rows.append({headers[c]: raw[c] for c in range(min(len(headers), len(raw))) if headers[c]})
    return rows


def parse_industry_xls(path: Path) -> list[dict[str, Any]]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    sheet = wb.sheet_by_name("Industry Averages")
    grid = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    return rows_from_grid(grid, "Industry Name")


def parse_country_premium_xlsx(path: Path) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["ERPs by country"]
    grid = [[cell.value for cell in row] for row in ws.iter_rows(min_row=1, max_row=ws.max_row)]
    return rows_from_grid(grid, "Country")


def write_outputs(
    dataset: str,
    label: str,
    source_filename: str,
    rows: list[dict[str, Any]],
    key_field: str,
    filter_value: str | None,
) -> tuple[Path, Path]:
    today = date.today().isoformat()
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

    if filter_value:
        selected = [r for r in rows if str(r.get(key_field, "")).strip().lower() == filter_value.strip().lower()]
        if not selected:
            available = ", ".join(sorted({str(r.get(key_field, "")) for r in rows})[:10])
            raise SystemExit(f"no {key_field!r} matching {filter_value!r}; examples: {available}")
        payload_rows = selected
    else:
        payload_rows = rows

    payload = {
        "dataset": dataset,
        "dataset_label": label,
        "source_filename": source_filename,
        "retrieved_utc": stamp,
        "row_count": len(payload_rows),
        "rows": payload_rows,
    }

    facts_path = OUT_DIR / f"damodaran_{dataset}.json"
    (OUT_DIR / f"damodaran_{dataset}_{stamp}.json").write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    facts_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")

    reg_path = OUT_DIR / f"damodaran_{dataset}_source_register.csv"
    with reg_path.open("w", newline="", encoding="utf-8") as f:
        fieldnames = [
            "source_name",
            "document_or_dataset",
            "publication_date",
            "as_of_date",
            "retrieval_date",
            "unit_currency",
            "transformation",
            "workbook_destination",
            "license_or_restriction",
            "checksum_or_snapshot",
        ]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in payload_rows:
            w.writerow(
                {
                    "source_name": "Aswath Damodaran, NYU Stern (pages.stern.nyu.edu/~adamodar)",
                    "document_or_dataset": f"{label} ({row.get(key_field, '')})",
                    "publication_date": "",
                    "as_of_date": today,
                    "retrieval_date": today,
                    "unit_currency": "ratio/percent per Damodaran's own convention for this dataset",
                    "transformation": "as-published, no re-derivation",
                    "workbook_destination": "Assumptions / cost of capital (IB, Corporate Finance, PE, Merchant Banking, AM)",
                    "license_or_restriction": "Publicly posted academic dataset; attribute to Aswath Damodaran / NYU Stern",
                    "checksum_or_snapshot": str(facts_path.relative_to(ROOT)),
                }
            )

    print(f"Wrote {facts_path}")
    print(f"Wrote {reg_path}")
    print(f"Rows: {len(payload_rows)}")
    return facts_path, reg_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--dataset", required=True,
                        help="wacc, betas, margin, pedata, pbvdata, psdata, capex, "
                             "a regional variant (e.g. waccEurope), or ctryprem")
    parser.add_argument("--industry", default=None, help="filter industry-average datasets to one industry")
    parser.add_argument("--country", default=None, help="filter ctryprem to one country")
    args = parser.parse_args()

    if args.dataset == "ctryprem":
        filename, label = COUNTRY_PREMIUM_DATASET
        path = _download(filename)
        rows = parse_country_premium_xlsx(path)
        write_outputs("ctryprem", label, filename, rows, key_field="Country", filter_value=args.country)
        return 0

    if args.dataset in INDUSTRY_DATASETS:
        filename, label = INDUSTRY_DATASETS[args.dataset]
    else:
        # Regional variant: stem + Region, e.g. waccEurope -> wacc.xls stem "wacc" + "Europe".
        matched = next((stem for stem in INDUSTRY_DATASETS if args.dataset.startswith(stem)
                         and args.dataset != stem), None)
        if not matched:
            raise SystemExit(f"unknown dataset {args.dataset!r}; known stems: "
                              f"{sorted(INDUSTRY_DATASETS)} or ctryprem")
        filename = f"{args.dataset}.xls"
        label = f"{INDUSTRY_DATASETS[matched][1]} ({args.dataset[len(matched):]} region)"

    path = _download(filename)
    rows = parse_industry_xls(path)
    write_outputs(args.dataset, label, filename, rows, key_field="Industry Name", filter_value=args.industry)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
