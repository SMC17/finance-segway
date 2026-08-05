"""L3 tool: private_credit_underwrite

Fail-closed agent-facing interface for the 05 Private Credit flagship.
Does not invent math: every number in the output comes from
tools/builders/build_private_credit_release.py's formulas, recalculated for
real by LibreOffice, never computed in this file. This tool's own job is
narrowly the six steps in docs/AGENT_TOOL_CONTRACT.md: validate provenance,
call the governed instance-generation path (tools/model_instance_release.py,
the same manifest-driven machinery that produced every public_*.xlsx
instance in this repo), recalculate, read Checks, and report status --
never re-implement CFADS, leverage, or covenant formulas here.

Usage (CLI):
  python tools/agents/private_credit_underwrite.py --demo
  python tools/agents/private_credit_underwrite.py --use-ares-fixture
  python tools/agents/private_credit_underwrite.py --inputs path/to/inputs.json

Contract: docs/AGENT_TOOL_CONTRACT.md
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DOMAIN = ROOT / "05_Private_Credit"
TEMPLATE = DOMAIN / "_template_CREDIT.xlsx"

sys.path.insert(0, str(ROOT / "tools"))
sys.path.insert(0, str(ROOT))
from model_instance_release import apply_manifest  # noqa: E402
from recalc import recalc  # noqa: E402


@dataclass
class Provenance:
    source_name: str
    as_of_date: str
    retrieval_date: str
    source_url: str | None = None
    transformation: str = "none"


@dataclass
class ToolInput:
    instance_slug: str
    borrower_name: str
    as_of: str
    facility: str = "Unitranche"
    scenario: str = "Base"
    # facts maps an Assumptions-sheet row label (e.g. "Revenue (LTM)") to a
    # Base-column override. Label-driven, not coordinate-driven -- matches
    # db/README.md's extraction convention -- so a template row reorder
    # breaks loudly (KeyError on an unknown label) rather than silently
    # writing the wrong cell.
    facts: dict[str, float] = field(default_factory=dict)
    provenance: dict[str, Provenance] = field(default_factory=dict)


@dataclass
class ToolOutput:
    ok: bool
    checks_status: str  # PASS | REVIEW | FAIL | NOT_RUN
    workbook_path: str | None
    headline: dict[str, Any]
    sources_written: list[dict[str, str]]
    message: str
    refresh_log_entry: str | None = None


def _assumption_rows() -> dict[str, int]:
    """Label -> row number on the Assumptions sheet, read from the live
    template rather than hardcoded, so a builder change is reflected here
    automatically instead of silently going stale."""
    workbook = openpyxl.load_workbook(TEMPLATE, data_only=False)
    sheet = workbook["Assumptions"]
    rows: dict[str, int] = {}
    for row in range(5, sheet.max_row + 1):
        label = sheet.cell(row, 2).value
        if label:
            rows[str(label).strip()] = row
    return rows


def validate_provenance(inp: ToolInput) -> list[str]:
    errors = []
    if not inp.borrower_name or not inp.borrower_name.strip():
        errors.append("missing required fact: borrower_name")
    if not inp.facts:
        errors.append("at least one material Assumptions fact is required")
        return errors
    known_labels = _assumption_rows()
    for label in inp.facts:
        if label not in known_labels:
            errors.append(
                f"unknown Assumptions row label: {label!r} "
                f"(not present on {TEMPLATE.relative_to(ROOT)})"
            )
        if label not in inp.provenance:
            errors.append(f"missing provenance for material fact: {label}")
    return errors


def _build_manifest(inp: ToolInput, instances_dir: Path) -> dict[str, Any]:
    rows = _assumption_rows()
    inputs = []
    for label, value in inp.facts.items():
        prov = inp.provenance[label]
        inputs.append({
            "sheet": "Assumptions",
            "cell": f"C{rows[label]}",
            "value": value,
            "source": {
                "name": prov.source_name,
                "url": prov.source_url,
                "as_of": prov.as_of_date,
                "notes": prov.transformation,
            },
        })
    return {
        "schema_version": "1.0",
        "id": inp.instance_slug,
        # Agent-drafted, not yet independently reviewed -- must never be
        # silently counted as M4 evidence regardless of whether the
        # underlying facts are real or a demo fixture. Human review
        # (Issue #7) is required to promote a draft instance further.
        "classification": "agent_tool_draft",
        "counts_toward_M4": False,
        "template": str(TEMPLATE.relative_to(ROOT)),
        "output": str((instances_dir / f"{inp.instance_slug}.xlsx").relative_to(ROOT)),
        "as_of": inp.as_of,
        "scenario": inp.scenario,
        "cover": {
            "Borrower / issuer:": inp.borrower_name,
            "Facility:": inp.facility,
            "Active scenario:": inp.scenario,
        },
        "inputs": inputs,
        "refresh": {
            "date": inp.as_of,
            "trigger": "L3 agent tool: private_credit_underwrite",
            "what_changed": f"Applied agent-submitted facts for {inp.borrower_name}",
            "reviewer_notes": "Generated by tools/agents/private_credit_underwrite.py -- not yet human-reviewed",
            "next_check": "On next material fact refresh",
        },
    }


def _read_checks(workbook_path: Path) -> tuple[str, dict[str, str]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook["Checks"]
    statuses: dict[str, str] = {}
    overall = "NOT_RUN"
    for row in range(5, sheet.max_row + 1):
        label = sheet.cell(row, 2).value
        status = sheet.cell(row, 3).value
        if not label:
            continue
        statuses[str(label)] = str(status)
        if str(label) == "Overall":
            overall = str(status)
    return overall, statuses


def _read_headline(workbook_path: Path, scenario: str) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    assumptions = workbook["Assumptions"]
    labels = {
        str(assumptions.cell(row, 2).value).strip(): row
        for row in range(5, assumptions.max_row + 1)
        if assumptions.cell(row, 2).value
    }

    def active(label: str) -> Any:
        row = labels.get(label)
        return assumptions.cell(row, 5).value if row else None

    debt_schedule = workbook["Debt Schedule"]
    covenants = workbook["Covenants"]
    return {
        "scenario": scenario,
        "opening_gross_debt": active("Opening gross debt"),
        "maximum_leverage": active("Maximum leverage"),
        "minimum_dscr": active("Minimum DSCR"),
        "yr5_ending_debt": debt_schedule.cell(15, 8).value,
        "covenant_breach_count": sum(
            1
            for row in covenants.iter_rows(min_row=14, max_row=14, min_col=3, max_col=7)
            for cell in row
            if cell.value == "BREACH"
        ),
    }


def run(inp: ToolInput, *, instances_dir: Path | None = None) -> ToolOutput:
    """Execute the tool: validate -> governed instance write -> real
    LibreOffice recalc -> read Checks. Fails closed on missing provenance
    or on the recalculated workbook showing FAIL.

    instances_dir defaults to 05_Private_Credit/instances/ (production).
    Tests must override it to a scratch directory under the repo root --
    never write agent-drafted or demo instances into the real evidence
    corpus directory, which is reserved for source-addressed public cases
    (see tests/test_real_data_only.py)."""
    instances_dir = instances_dir or (DOMAIN / "instances")
    errors = validate_provenance(inp)
    if errors:
        return ToolOutput(
            ok=False,
            checks_status="FAIL",
            workbook_path=None,
            headline={},
            sources_written=[],
            message="fail-closed: " + "; ".join(errors),
        )

    manifest = _build_manifest(inp, instances_dir)
    manifest_path = instances_dir / f"{inp.instance_slug}.manifest.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

    receipt = apply_manifest(manifest_path, ROOT)
    output_path = ROOT / manifest["output"]

    recalc_result = recalc(str(output_path), timeout=60)
    if recalc_result.get("status") != "success" or recalc_result.get("total_errors", 0):
        return ToolOutput(
            ok=False,
            checks_status="FAIL",
            workbook_path=str(output_path.relative_to(ROOT)),
            headline={},
            sources_written=[asdict(v) for v in inp.provenance.values()],
            message=f"fail-closed: recalculation did not succeed cleanly: {recalc_result}",
        )

    overall, statuses = _read_checks(output_path)
    headline = _read_headline(output_path, inp.scenario)
    headline["checks_detail"] = statuses

    return ToolOutput(
        ok=overall != "FAIL",
        checks_status=overall,
        workbook_path=str(output_path.relative_to(ROOT)),
        headline=headline,
        sources_written=[
            {"field": label, **asdict(prov)} for label, prov in inp.provenance.items()
        ],
        message=(
            "instance generated and recalculated via the governed builder path; "
            f"Checks Overall = {overall}. Not a client deliverable until Checks "
            "is PASS and stakeholder sign-off is recorded (Issue #7)."
        ),
        refresh_log_entry=receipt.get("as_of"),
    )


SCRATCH_DIR = ROOT / ".agent-tool-scratch" / "05_Private_Credit"


def demo() -> ToolOutput:
    """Illustrative run with a fully fictional borrower and no real source
    URLs. Writes to SCRATCH_DIR, not the real 05_Private_Credit/instances/
    corpus -- a demo instance must never be mistaken for, or accidentally
    committed alongside, a source-addressed public case."""
    today = date.today().isoformat()
    inp = ToolInput(
        instance_slug="demo_unitranche_stub",
        borrower_name="Demo Borrower LLC",
        as_of=today,
        facility="Unitranche",
        scenario="Base",
        facts={
            "Revenue (LTM)": 480.0,
            "EBITDA margin": 0.19,
            "Opening gross debt": 340.0,
            "Base rate": 0.045,
            "Cash spread": 0.06,
            "Maximum leverage": 6.0,
            "Minimum DSCR": 1.05,
        },
        provenance={
            label: Provenance(
                source_name="demo_fixture",
                as_of_date=today,
                retrieval_date=today,
                source_url=None,
                transformation="demo only -- not for client use",
            )
            for label in (
                "Revenue (LTM)", "EBITDA margin", "Opening gross debt",
                "Base rate", "Cash spread", "Maximum leverage", "Minimum DSCR",
            )
        },
    )
    return run(inp, instances_dir=SCRATCH_DIR)


ARCC_COMPANYFACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK0001287750.json"
ARCC_FACTS_PATH = ROOT / "tools" / "data_fabric" / "out" / "ARCC_facts_selected.json"


def ares_fixture() -> ToolInput:
    """Real reference instance: Ares Capital Corporation (ARCC, CIK
    0001287750), a large publicly reporting BDC. Maps portfolio-level
    balance-sheet facts from its Q1 2026 10-Q (via
    tools/data_fabric/edgar_company_facts.py's companyfacts extract) onto
    the CREDIT template's Assumptions rows.

    Deliberately a portfolio-level proxy, NOT a single-borrower unitranche
    commitment memo -- only facts with a real EDGAR-sourced number are
    submitted (opening gross debt, opening cash). Revenue, EBITDA margin,
    and covenant-driving assumptions are not overridden because ARCC's
    public filings report portfolio-level investment income, not a
    single-credit CFADS build -- inventing a mapping would violate this
    tool's fail-closed provenance requirement. Expect Checks to show
    REVIEW or FAIL against the template's default revenue/EBITDA
    assumptions: that is the correct, honest signal for a proxy instance,
    not a defect (see
    05_Private_Credit/instances/public_ares_capital_2024.thesis.md).
    """
    if not ARCC_FACTS_PATH.exists():
        raise FileNotFoundError(
            f"{ARCC_FACTS_PATH.relative_to(ROOT)} not found -- run "
            "tools/data_fabric/edgar_company_facts.py --ticker ARCC --cik 1287750 first"
        )
    raw_facts = {
        row["concept"]: row for row in json.loads(ARCC_FACTS_PATH.read_text(encoding="utf-8"))
    }
    as_of = raw_facts["LongTermDebt"]["end"]
    retrieval_date = date.today().isoformat()

    def prov(concept: str, notes: str) -> Provenance:
        return Provenance(
            source_name=f"SEC EDGAR companyfacts: {concept} ({raw_facts[concept]['form']})",
            as_of_date=raw_facts[concept]["end"],
            retrieval_date=retrieval_date,
            source_url=ARCC_COMPANYFACTS_URL,
            transformation=notes,
        )

    return ToolInput(
        instance_slug="public_ares_capital_2024",
        borrower_name="Ares Capital Corporation (public BDC portfolio proxy)",
        as_of=as_of,
        facility="Portfolio (BDC aggregate, not a single facility)",
        scenario="Base",
        facts={
            "Opening gross debt": raw_facts["LongTermDebt"]["value"] / 1_000_000,
            "Opening cash": raw_facts["CashAndCashEquivalentsAtCarryingValue"]["value"] / 1_000_000,
        },
        provenance={
            "Opening gross debt": prov(
                "LongTermDebt",
                "Portfolio-aggregate LongTermDebt from XBRL companyfacts, "
                "converted from USD to the template's $ millions scale; "
                "mapped onto the single-facility 'Opening gross debt' row "
                "as a portfolio-level proxy, not a single-borrower balance.",
            ),
            "Opening cash": prov(
                "CashAndCashEquivalentsAtCarryingValue",
                "Portfolio-aggregate cash from XBRL companyfacts, converted "
                "from USD to the template's $ millions scale.",
            ),
        },
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--demo", action="store_true", help="fictional borrower, writes to .agent-tool-scratch/")
    parser.add_argument(
        "--use-ares-fixture", action="store_true",
        help="real Ares Capital (ARCC) portfolio-proxy instance from EDGAR companyfacts, "
        "writes to 05_Private_Credit/instances/public_ares_capital_2024.xlsx",
    )
    parser.add_argument("--inputs", type=Path, help="JSON file matching ToolInput shape")
    parser.add_argument(
        "--output-dir", type=Path, default=None,
        help="override instance output directory (default: 05_Private_Credit/instances/ for --inputs/--use-ares-fixture)",
    )
    args = parser.parse_args()

    if args.demo:
        out = demo()
    elif args.use_ares_fixture:
        out = run(ares_fixture(), instances_dir=args.output_dir)
    elif args.inputs:
        raw = json.loads(args.inputs.read_text())
        provenance = {
            label: Provenance(**value) if isinstance(value, dict) else value
            for label, value in raw.get("provenance", {}).items()
        }
        inp = ToolInput(
            instance_slug=raw["instance_slug"],
            borrower_name=raw["borrower_name"],
            as_of=raw["as_of"],
            facility=raw.get("facility", "Unitranche"),
            scenario=raw.get("scenario", "Base"),
            facts=raw.get("facts", {}),
            provenance=provenance,
        )
        out = run(inp, instances_dir=args.output_dir)
    else:
        parser.error("provide --demo, --use-ares-fixture, or --inputs")
        return 2

    print(json.dumps(asdict(out), indent=2))
    return 0 if out.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
