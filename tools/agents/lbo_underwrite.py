"""L3 tool: lbo_underwrite

Fail-closed agent-facing interface for the 03 Private Equity flagship.
Does not invent math: every number in the output comes from
03_Private_Equity/_template_LBO.xlsx's formulas, recalculated for real by
LibreOffice, never computed in this file. This tool's own job is narrowly
the six steps in docs/AGENT_TOOL_CONTRACT.md: validate provenance, call the
governed instance-generation path (tools/model_instance_release.py, the
same manifest-driven machinery that produced every public_*.xlsx instance
in this repo), recalculate, read Checks, and report status -- never
re-implement Sources & Uses, the debt schedule, or the returns waterfall
here.

This is the second tool on docs/AGENT_TOOL_CONTRACT.md's implementation
order (private_credit_underwrite shipped first), and follows its exact
shape: ToolInput/ToolOutput dataclasses, label-driven Assumptions row
lookup (a template reorder breaks loudly, not silently), a demo fixture
that writes to scratch, and a real fixture that reuses this repo's own
already-sourced public-case facts rather than re-deriving them.

Usage (CLI):
  python tools/agents/lbo_underwrite.py --demo
  python tools/agents/lbo_underwrite.py --use-home-depot-fixture
  python tools/agents/lbo_underwrite.py --inputs path/to/inputs.json

Contract: docs/AGENT_TOOL_CONTRACT.md
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DOMAIN = ROOT / "03_Private_Equity"
TEMPLATE = DOMAIN / "_template_LBO.xlsx"

sys.path.insert(0, str(ROOT / "tools"))
sys.path.insert(0, str(ROOT))
from model_instance_release import apply_manifest  # noqa: E402
from recalc import recalc  # noqa: E402


@dataclass
class Provenance:
    source_name: str
    as_of_date: str
    retrieval_date: str
    source_url: str | None = None
    transformation: str = "none"


@dataclass
class ToolInput:
    instance_slug: str
    target_name: str
    as_of: str
    scenario: str = "Base"
    # facts maps an Assumptions-sheet row label (e.g. "LTM revenue") to a
    # Base-column override. Label-driven, not coordinate-driven -- matches
    # db/README.md's extraction convention and private_credit_underwrite's
    # own convention -- so a template row reorder breaks loudly (KeyError
    # on an unknown label) rather than silently writing the wrong cell.
    facts: dict[str, float] = field(default_factory=dict)
    provenance: dict[str, Provenance] = field(default_factory=dict)


@dataclass
class ToolOutput:
    ok: bool
    checks_status: str  # PASS | REVIEW | FAIL | NOT_RUN
    workbook_path: str | None
    headline: dict[str, Any]
    sources_written: list[dict[str, str]]
    message: str
    refresh_log_entry: str | None = None


def _assumption_rows() -> dict[str, int]:
    """Label -> row number on the Assumptions sheet, read from the live
    template rather than hardcoded, so a builder change is reflected here
    automatically instead of silently going stale."""
    workbook = openpyxl.load_workbook(TEMPLATE, data_only=False)
    sheet = workbook["Assumptions"]
    rows: dict[str, int] = {}
    for row in range(5, sheet.max_row + 1):
        label = sheet.cell(row, 2).value
        if label:
            rows[str(label).strip()] = row
    return rows


def validate_provenance(inp: ToolInput) -> list[str]:
    errors = []
    if not inp.target_name or not inp.target_name.strip():
        errors.append("missing required fact: target_name")
    if not inp.facts:
        errors.append("at least one material Assumptions fact is required")
        return errors
    known_labels = _assumption_rows()
    for label in inp.facts:
        if label not in known_labels:
            errors.append(
                f"unknown Assumptions row label: {label!r} "
                f"(not present on {TEMPLATE.relative_to(ROOT)})"
            )
        if label not in inp.provenance:
            errors.append(f"missing provenance for material fact: {label}")
    if "Exit year" in inp.facts:
        # Returns Waterfall only has year-1..7 columns (C:I). Writing an
        # out-of-range value doesn't fail loudly at write time -- it
        # produces #REF! errors in downstream formulas that reference the
        # exit-year column, which recalc() then reports as a generic
        # failure far from the actual cause. Reject it here instead, at
        # the one place that knows what "out of range" means for this
        # template, rather than silently clamping a bad input into the
        # workbook (found the hard way: CI's real LibreOffice recalc
        # caught what this session's sandbox-blocked recalc could not).
        try:
            exit_year = int(inp.facts["Exit year"])
        except (TypeError, ValueError):
            exit_year = None
        if exit_year is None or not (1 <= exit_year <= 7):
            errors.append(
                f"Exit year must be an integer from 1 to 7 (Returns Waterfall's "
                f"columns C:I), got {inp.facts['Exit year']!r}"
            )
    return errors


def _build_manifest(inp: ToolInput, instances_dir: Path) -> dict[str, Any]:
    rows = _assumption_rows()
    inputs = []
    for label, value in inp.facts.items():
        prov = inp.provenance[label]
        inputs.append({
            "sheet": "Assumptions",
            "cell": f"C{rows[label]}",
            "value": value,
            "source": {
                "name": prov.source_name,
                "url": prov.source_url,
                "as_of": prov.as_of_date,
                "notes": prov.transformation,
            },
        })
    return {
        "schema_version": "1.0",
        "id": inp.instance_slug,
        # Agent-drafted, not yet independently reviewed -- must never be
        # silently counted as M4 evidence regardless of whether the
        # underlying facts are real or a demo fixture. Human review
        # (Issue #7) is required to promote a draft instance further.
        "classification": "agent_tool_draft",
        "counts_toward_M4": False,
        "template": str(TEMPLATE.relative_to(ROOT)),
        "output": str((instances_dir / f"{inp.instance_slug}.xlsx").relative_to(ROOT)),
        "as_of": inp.as_of,
        "scenario": inp.scenario,
        "cover": {
            "Target / transaction:": inp.target_name,
            "Active scenario:": inp.scenario,
        },
        "inputs": inputs,
        "refresh": {
            "date": inp.as_of,
            "trigger": "L3 agent tool: lbo_underwrite",
            "what_changed": f"Applied agent-submitted facts for {inp.target_name}",
            "reviewer_notes": "Generated by tools/agents/lbo_underwrite.py -- not yet human-reviewed",
            "next_check": "On next material fact refresh",
        },
    }


def _read_checks(workbook_path: Path) -> tuple[str, dict[str, str]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook["Checks"]
    statuses: dict[str, str] = {}
    overall = "NOT_RUN"
    for row in range(5, sheet.max_row + 1):
        label = sheet.cell(row, 2).value
        status = sheet.cell(row, 3).value
        if not label:
            continue
        statuses[str(label)] = str(status)
        if str(label) == "Overall":
            overall = str(status)
    return overall, statuses


def _exit_year_column(exit_year: Any) -> str:
    """Returns Waterfall columns C..I are years 1..7. Fail closed on an
    out-of-range exit year rather than silently reading the wrong column."""
    try:
        year = int(exit_year)
    except (TypeError, ValueError):
        year = 1
    year = max(1, min(7, year))
    return chr(ord("C") + year - 1)


def _read_headline(workbook_path: Path, scenario: str) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    assumptions = workbook["Assumptions"]
    labels = {
        str(assumptions.cell(row, 2).value).strip(): row
        for row in range(5, assumptions.max_row + 1)
        if assumptions.cell(row, 2).value
    }

    def active(label: str) -> Any:
        row = labels.get(label)
        return assumptions.cell(row, 5).value if row else None

    exit_year = active("Exit year")
    col = _exit_year_column(exit_year)

    waterfall = workbook["Returns Waterfall"]
    waterfall_labels = {
        str(waterfall.cell(row, 2).value).strip(): row
        for row in range(5, waterfall.max_row + 1)
        if waterfall.cell(row, 2).value
    }

    def wf(label: str) -> Any:
        row = waterfall_labels.get(label)
        return waterfall[f"{col}{row}"].value if row else None

    covenants = workbook["Covenants"]
    breach_count = sum(
        1
        for row in covenants.iter_rows(min_row=13, max_row=13, min_col=3, max_col=9)
        for cell in row
        if cell.value == "BREACH"
    )

    sources_uses = workbook["Sources & Uses"]
    return {
        "scenario": scenario,
        "exit_year": exit_year,
        "sponsor_equity": sources_uses["F8"].value,
        "entry_ebitda_multiple": active("Entry EBITDA multiple"),
        "exit_ebitda_multiple": active("Exit EBITDA multiple"),
        "sponsor_moic": wf("Sponsor MOIC"),
        "sponsor_irr": wf("Sponsor IRR"),
        "gross_equity_value": wf("Gross equity value"),
        "covenant_breach_count": breach_count,
    }


def run(inp: ToolInput, *, instances_dir: Path | None = None) -> ToolOutput:
    """Execute the tool: validate -> governed instance write -> real
    LibreOffice recalc -> read Checks. Fails closed on missing provenance
    or on the recalculated workbook showing FAIL.

    instances_dir defaults to 03_Private_Equity/instances/ (production).
    Tests must override it to a scratch directory under the repo root --
    never write agent-drafted or demo instances into the real evidence
    corpus directory, which is reserved for source-addressed public cases
    (see tests/test_real_data_only.py)."""
    instances_dir = instances_dir or (DOMAIN / "instances")
    errors = validate_provenance(inp)
    if errors:
        return ToolOutput(
            ok=False,
            checks_status="FAIL",
            workbook_path=None,
            headline={},
            sources_written=[],
            message="fail-closed: " + "; ".join(errors),
        )

    manifest = _build_manifest(inp, instances_dir)
    manifest_path = instances_dir / f"{inp.instance_slug}.manifest.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

    receipt = apply_manifest(manifest_path, ROOT)
    output_path = ROOT / manifest["output"]

    recalc_result = recalc(str(output_path), timeout=60)
    if recalc_result.get("status") != "success" or recalc_result.get("total_errors", 0):
        return ToolOutput(
            ok=False,
            checks_status="FAIL",
            workbook_path=str(output_path.relative_to(ROOT)),
            headline={},
            sources_written=[asdict(v) for v in inp.provenance.values()],
            message=f"fail-closed: recalculation did not succeed cleanly: {recalc_result}",
        )

    overall, statuses = _read_checks(output_path)
    headline = _read_headline(output_path, inp.scenario)
    headline["checks_detail"] = statuses

    return ToolOutput(
        ok=overall != "FAIL",
        checks_status=overall,
        workbook_path=str(output_path.relative_to(ROOT)),
        headline=headline,
        sources_written=[
            {"field": label, **asdict(prov)} for label, prov in inp.provenance.items()
        ],
        message=(
            "instance generated and recalculated via the governed builder path; "
            f"Checks Overall = {overall}. Not a client deliverable until Checks "
            "is PASS and stakeholder sign-off is recorded (Issue #7)."
        ),
        refresh_log_entry=receipt.get("as_of"),
    )


SCRATCH_DIR = ROOT / ".agent-tool-scratch" / "03_Private_Equity"


def demo() -> ToolOutput:
    """Illustrative run with a fully fictional target and no real source
    URLs. Writes to SCRATCH_DIR, not the real 03_Private_Equity/instances/
    corpus -- a demo instance must never be mistaken for, or accidentally
    committed alongside, a source-addressed public case."""
    today = date.today().isoformat()
    facts = {
        "LTM revenue": 1200.0,
        "LTM EBITDA margin": 0.22,
        "Entry EBITDA multiple": 9.5,
        "TLB opening leverage": 4.0,
        "Second-lien opening leverage": 1.5,
        "Exit EBITDA multiple": 9.5,
        "Exit year": 5,
    }
    inp = ToolInput(
        instance_slug="demo_lbo_stub",
        target_name="Demo Target Inc.",
        as_of=today,
        scenario="Base",
        facts=facts,
        provenance={
            label: Provenance(
                source_name="demo_fixture",
                as_of_date=today,
                retrieval_date=today,
                source_url=None,
                transformation="demo only -- not for client use",
            )
            for label in facts
        },
    )
    return run(inp, instances_dir=SCRATCH_DIR)


HOME_DEPOT_MANIFEST_PATH = ROOT / "standards" / "public_cases" / "pe-public-home-depot-2023.json"


def home_depot_fixture() -> ToolInput:
    """Real reference instance: reuses this repo's own already-sourced
    pe-public-home-depot-2023 public case (a real, SEC-cited operating
    profile) rather than re-deriving from EDGAR -- the source of truth is
    the same 10-K either way, and re-fetching would just risk drift from
    the already-verified, already-hashed snapshot.

    Deliberately a real OPERATING profile, not a real LBO -- Home Depot has
    no sponsor, no acquisition debt, and no deal structure. Only the facts
    that describe the business itself (revenue, margin, growth, D&A/capex
    intensity, cash tax rate) are submitted with real provenance; entry/exit
    multiples, leverage, and covenant thresholds stay at the template's own
    illustrative defaults because they are deal-structuring choices no
    public 10-K discloses for a company that was never actually taken
    private. Expect Checks to reflect a plausible-but-illustrative deal
    wrapped around real operating economics, not a real transaction's
    real terms -- that is the honest, correct signal for this kind of
    proxy instance, matching private_credit_underwrite's ares_fixture()
    precedent.
    """
    if not HOME_DEPOT_MANIFEST_PATH.exists():
        raise FileNotFoundError(
            f"{HOME_DEPOT_MANIFEST_PATH.relative_to(ROOT)} not found"
        )
    manifest = json.loads(HOME_DEPOT_MANIFEST_PATH.read_text(encoding="utf-8"))
    row_labels = {v: k for k, v in _assumption_rows().items()}
    retrieval_date = date.today().isoformat()
    source = manifest["sources"][0]

    facts: dict[str, float] = {}
    provenance: dict[str, Provenance] = {}
    for item in manifest["inputs"]:
        if item["sheet"] != "Assumptions":
            continue
        row = int(item["cell"][1:])
        label = row_labels.get(row)
        if not label:
            continue
        facts[label] = item["value"]
        provenance[label] = Provenance(
            source_name=source["name"],
            as_of_date=manifest["as_of"],
            retrieval_date=retrieval_date,
            source_url=source["url"],
            transformation=(
                f"Reused from {HOME_DEPOT_MANIFEST_PATH.relative_to(ROOT)} "
                f"({item['input_kind']}); real operating fact, not a real LBO term."
            ),
        )

    return ToolInput(
        instance_slug="public_home_depot_lbo_proxy",
        target_name="The Home Depot, Inc. (public-company operating proxy, not a real LBO)",
        as_of=manifest["as_of"],
        scenario="Base",
        facts=facts,
        provenance=provenance,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--demo", action="store_true", help="fictional target, writes to .agent-tool-scratch/")
    parser.add_argument(
        "--use-home-depot-fixture", action="store_true",
        help="real Home Depot operating-profile proxy reused from the existing public case, "
        "writes to 03_Private_Equity/instances/public_home_depot_lbo_proxy.xlsx",
    )
    parser.add_argument("--inputs", type=Path, help="JSON file matching ToolInput shape")
    parser.add_argument(
        "--output-dir", type=Path, default=None,
        help="override instance output directory (default: 03_Private_Equity/instances/ for --inputs/--use-home-depot-fixture)",
    )
    args = parser.parse_args()

    if args.demo:
        out = demo()
    elif args.use_home_depot_fixture:
        out = run(home_depot_fixture(), instances_dir=args.output_dir)
    elif args.inputs:
        raw = json.loads(args.inputs.read_text())
        provenance = {
            label: Provenance(**value) if isinstance(value, dict) else value
            for label, value in raw.get("provenance", {}).items()
        }
        inp = ToolInput(
            instance_slug=raw["instance_slug"],
            target_name=raw["target_name"],
            as_of=raw["as_of"],
            scenario=raw.get("scenario", "Base"),
            facts=raw.get("facts", {}),
            provenance=provenance,
        )
        out = run(inp, instances_dir=args.output_dir)
    else:
        parser.error("provide --demo, --use-home-depot-fixture, or --inputs")
        return 2

    print(json.dumps(asdict(out), indent=2))
    return 0 if out.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
