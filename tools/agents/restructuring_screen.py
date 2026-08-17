"""L3 tool: restructuring_screen

Fail-closed agent-facing interface for the 24 Distressed & Restructuring
flagship. Does not invent math: every number in the output comes from
24_Distressed_Restructuring/_template_RESTRUCTURING.xlsx's formulas,
recalculated for real by LibreOffice, never computed in this file. This
tool's own job is narrowly the six steps in docs/AGENT_TOOL_CONTRACT.md:
validate provenance, call the governed instance-generation path
(tools/model_instance_release.py), recalculate, read Decision & Checks,
and report status -- never re-implement the recovery waterfall, the
13-week liquidity roll, or the liquidation-vs-reorg NPV here.

Third tool on docs/AGENT_TOOL_CONTRACT.md's implementation order. Its
input surface is genuinely different from private_credit_underwrite and
lbo_underwrite, not just relabeled: this template has no single
"Assumptions" sheet. Real inputs are scattered across four sheets
(Recovery Waterfall, 13-Week Liquidity, New Money, Liquidation vs Reorg),
and the domain's own two existing real public cases
(distressed-public-hertz-2021-reorganization,
distressed-public-bbby-2022-liquidity) each source only 1-2 cells total --
the rest stay at illustrative template defaults. That sparsity is
real, not a shortcut this tool introduces: pre-petition or early-case
information about a distressed company's capital structure and recovery
prospects is genuinely hard to source with real provenance, and this tool
follows the same honest-gap convention rather than inventing a denser
input surface those two cases didn't attempt.

facts is therefore sheet-scoped: dict[sheet_name, dict[label, value]],
always targeting column C on each of the four input sheets (matching what
both existing real cases actually did). Liquidation vs Reorg's
Reorganization-side column (D) is out of scope for this tool -- neither
real case sources it, and reorg-side recovery figures are typically the
harder side to source pre-emergence.

Usage (CLI):
  python tools/agents/restructuring_screen.py --demo
  python tools/agents/restructuring_screen.py --use-hertz-fixture
  python tools/agents/restructuring_screen.py --inputs path/to/inputs.json

Contract: docs/AGENT_TOOL_CONTRACT.md
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DOMAIN = ROOT / "24_Distressed_Restructuring"
TEMPLATE = DOMAIN / "_template_RESTRUCTURING.xlsx"
INPUT_SHEETS = ("Recovery Waterfall", "13-Week Liquidity", "New Money", "Liquidation vs Reorg")

sys.path.insert(0, str(ROOT / "tools"))
sys.path.insert(0, str(ROOT))
from model_instance_release import apply_manifest  # noqa: E402
from recalc import recalc  # noqa: E402


@dataclass
class Provenance:
    source_name: str
    as_of_date: str
    retrieval_date: str
    source_url: str | None = None
    transformation: str = "none"


@dataclass
class ToolInput:
    instance_slug: str
    situation_type: str
    as_of: str
    # facts is sheet-scoped: {sheet_name: {row_label: value}}. Label-driven,
    # not coordinate-driven, matching the other two agent tools -- a
    # template row reorder breaks loudly (KeyError on an unknown label)
    # rather than silently writing the wrong cell.
    facts: dict[str, dict[str, float]] = field(default_factory=dict)
    provenance: dict[str, dict[str, Provenance]] = field(default_factory=dict)


@dataclass
class ToolOutput:
    ok: bool
    checks_status: str  # PASS | REVIEW | BREACH | NOT_RUN
    workbook_path: str | None
    headline: dict[str, Any]
    sources_written: list[dict[str, str]]
    message: str
    refresh_log_entry: str | None = None


def _input_rows(sheet_name: str) -> dict[str, int]:
    """Label -> row number for a given input sheet, read from the live
    template rather than hardcoded. Starts at row 3 rather than row 4:
    unlike the other three input sheets (whose first real label follows a
    row-4 or row-5 table header), 13-Week Liquidity's first two real
    inputs ("Initial unrestricted liquidity", "Minimum operating
    liquidity") sit at rows 3-4, directly under the row-2 sheet title with
    no header row between. Scanning from row 3 picks up a few header-only
    cells (e.g. "Tranche") on the other sheets too, which is harmless --
    nothing ever looks those up by label."""
    workbook = openpyxl.load_workbook(TEMPLATE, data_only=False)
    sheet = workbook[sheet_name]
    rows: dict[str, int] = {}
    for row in range(3, sheet.max_row + 1):
        label = sheet.cell(row, 2).value
        if label:
            rows[str(label).strip()] = row
    return rows


def validate_provenance(inp: ToolInput) -> list[str]:
    errors = []
    if not inp.situation_type or not inp.situation_type.strip():
        errors.append("missing required fact: situation_type")
    if not inp.facts or not any(inp.facts.values()):
        errors.append("at least one material fact on an input sheet is required")
        return errors
    for sheet_name, sheet_facts in inp.facts.items():
        if sheet_name not in INPUT_SHEETS:
            errors.append(
                f"unknown input sheet: {sheet_name!r} (must be one of {INPUT_SHEETS})"
            )
            continue
        known_labels = _input_rows(sheet_name)
        sheet_provenance = inp.provenance.get(sheet_name, {})
        for label in sheet_facts:
            if label not in known_labels:
                errors.append(
                    f"unknown row label on {sheet_name!r}: {label!r} "
                    f"(not present on {TEMPLATE.relative_to(ROOT)})"
                )
            if label not in sheet_provenance:
                errors.append(f"missing provenance for material fact: {sheet_name}::{label}")
    return errors


def _build_manifest(inp: ToolInput, instances_dir: Path) -> dict[str, Any]:
    inputs = []
    for sheet_name, sheet_facts in inp.facts.items():
        rows = _input_rows(sheet_name)
        for label, value in sheet_facts.items():
            prov = inp.provenance[sheet_name][label]
            inputs.append({
                "sheet": sheet_name,
                "cell": f"C{rows[label]}",
                "value": value,
                "source": {
                    "name": prov.source_name,
                    "url": prov.source_url,
                    "as_of": prov.as_of_date,
                    "notes": prov.transformation,
                },
            })
    return {
        "schema_version": "1.0",
        "id": inp.instance_slug,
        # Agent-drafted, not yet independently reviewed -- must never be
        # silently counted as M4 evidence regardless of whether the
        # underlying facts are real or a demo fixture. Human review
        # (Issue #7) is required to promote a draft instance further.
        "classification": "agent_tool_draft",
        "counts_toward_M4": False,
        "template": str(TEMPLATE.relative_to(ROOT)),
        "output": str((instances_dir / f"{inp.instance_slug}.xlsx").relative_to(ROOT)),
        "as_of": inp.as_of,
        "cover": {
            "Situation type:": inp.situation_type,
        },
        "inputs": inputs,
        "refresh": {
            "date": inp.as_of,
            "trigger": "L3 agent tool: restructuring_screen",
            "what_changed": f"Applied agent-submitted facts for {inp.situation_type}",
            "reviewer_notes": "Generated by tools/agents/restructuring_screen.py -- not yet human-reviewed",
            "next_check": "On next material fact refresh",
        },
    }


def _read_checks(workbook_path: Path) -> tuple[str, dict[str, str]]:
    """Decision & Checks status lives in column D here (not C, unlike the
    other two agent tools' Checks sheets), and the overall cell is
    labeled "Overall model status" in B14, not "Overall"."""
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook["Decision & Checks"]
    statuses: dict[str, str] = {}
    overall = "NOT_RUN"
    for row in range(5, 13):
        label = sheet.cell(row, 2).value
        status = sheet.cell(row, 4).value
        if not label:
            continue
        statuses[str(label)] = str(status)
    overall_cell = sheet["C14"].value
    if overall_cell is not None:
        overall = str(overall_cell)
    return overall, statuses


def _read_headline(workbook_path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook["Decision & Checks"]
    labels = {
        str(sheet.cell(row, 2).value).strip(): row
        for row in range(17, sheet.max_row + 1)
        if sheet.cell(row, 2).value
    }

    def outcome(label: str) -> Any:
        row = labels.get(label)
        return sheet.cell(row, 3).value if row else None

    return {
        "enterprise_value_available": outcome("Enterprise value available"),
        "fulcrum_security": outcome("Fulcrum security"),
        "liquidation_npv": outcome("Liquidation NPV"),
        "reorganization_npv": outcome("Reorganization NPV"),
        "peak_funding_need": outcome("Peak funding need"),
        "total_new_money_claim": outcome("Total new-money claim"),
        "first_liquidity_breach_week": outcome("First liquidity breach week"),
    }


def run(inp: ToolInput, *, instances_dir: Path | None = None) -> ToolOutput:
    """Execute the tool: validate -> governed instance write -> real
    LibreOffice recalc -> read Decision & Checks. Fails closed on missing
    provenance or on the recalculated workbook showing BREACH from a
    literal recalc failure (a BREACH verdict from real inputs is a valid,
    honest outcome -- see distressed-public-bbby-2022-liquidity, which
    correctly shows distress -- only a recalculation that did not succeed
    cleanly is treated as a tool failure).

    instances_dir defaults to 24_Distressed_Restructuring/instances/
    (production). Tests must override it to a scratch directory under the
    repo root -- never write agent-drafted or demo instances into the real
    evidence corpus directory, which is reserved for source-addressed
    public cases (see tests/test_real_data_only.py)."""
    instances_dir = instances_dir or (DOMAIN / "instances")
    errors = validate_provenance(inp)
    if errors:
        return ToolOutput(
            ok=False,
            checks_status="FAIL",
            workbook_path=None,
            headline={},
            sources_written=[],
            message="fail-closed: " + "; ".join(errors),
        )

    manifest = _build_manifest(inp, instances_dir)
    manifest_path = instances_dir / f"{inp.instance_slug}.manifest.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

    receipt = apply_manifest(manifest_path, ROOT)
    output_path = ROOT / manifest["output"]

    recalc_result = recalc(str(output_path), timeout=60)
    if recalc_result.get("status") != "success" or recalc_result.get("total_errors", 0):
        return ToolOutput(
            ok=False,
            checks_status="FAIL",
            workbook_path=str(output_path.relative_to(ROOT)),
            headline={},
            sources_written=[
                {"field": f"{sheet}::{label}", **asdict(prov)}
                for sheet, sheet_prov in inp.provenance.items()
                for label, prov in sheet_prov.items()
            ],
            message=f"fail-closed: recalculation did not succeed cleanly: {recalc_result}",
        )

    overall, statuses = _read_checks(output_path)
    headline = _read_headline(output_path)
    headline["checks_detail"] = statuses

    return ToolOutput(
        ok=overall not in ("FAIL", "NOT_RUN"),
        checks_status=overall,
        workbook_path=str(output_path.relative_to(ROOT)),
        headline=headline,
        sources_written=[
            {"field": f"{sheet}::{label}", **asdict(prov)}
            for sheet, sheet_prov in inp.provenance.items()
            for label, prov in sheet_prov.items()
        ],
        message=(
            "instance generated and recalculated via the governed builder path; "
            f"Decision & Checks overall model status = {overall}. A BREACH or "
            "REVIEW verdict is a valid, honest reading of the submitted facts, "
            "not a tool failure. Not a client deliverable until stakeholder "
            "sign-off is recorded (Issue #7)."
        ),
        refresh_log_entry=receipt.get("as_of"),
    )


SCRATCH_DIR = ROOT / ".agent-tool-scratch" / "24_Distressed_Restructuring"


def demo() -> ToolOutput:
    """Illustrative run with a fully fictional situation and no real
    source URLs. Writes to SCRATCH_DIR, not the real
    24_Distressed_Restructuring/instances/ corpus."""
    today = date.today().isoformat()
    facts = {
        "New Money": {"New-money commitment": 50.0},
        "13-Week Liquidity": {"Initial unrestricted liquidity": 30.0},
        "Recovery Waterfall": {"Total enterprise value available for distribution ($)": 400.0},
    }
    provenance = {
        sheet: {
            label: Provenance(
                source_name="demo_fixture",
                as_of_date=today,
                retrieval_date=today,
                source_url=None,
                transformation="demo only -- not for client use",
            )
            for label in labels
        }
        for sheet, labels in facts.items()
    }
    inp = ToolInput(
        instance_slug="demo_restructuring_stub",
        situation_type="Demo Co. hypothetical Chapter 11",
        as_of=today,
        facts=facts,
        provenance=provenance,
    )
    return run(inp, instances_dir=SCRATCH_DIR)


HERTZ_MANIFEST_PATH = ROOT / "standards" / "public_cases" / "distressed-public-hertz-2021-reorganization.json"


def hertz_fixture() -> ToolInput:
    """Real reference instance: reuses this repo's own already-sourced
    distressed-public-hertz-2021-reorganization public case -- Hertz's real
    $7.5bn Chapter 11 exit financing commitment, with real provenance.
    Deliberately sparse, matching what the existing real case itself
    sources: only the New Money commitment, not a full capital structure
    or recovery waterfall, since this tool must not invent facts the
    original case did not source just to make the demonstration denser.
    """
    if not HERTZ_MANIFEST_PATH.exists():
        raise FileNotFoundError(f"{HERTZ_MANIFEST_PATH.relative_to(ROOT)} not found")
    manifest = json.loads(HERTZ_MANIFEST_PATH.read_text(encoding="utf-8"))
    retrieval_date = date.today().isoformat()
    source = manifest["sources"][0]

    facts: dict[str, dict[str, float]] = {}
    provenance: dict[str, dict[str, Provenance]] = {}
    for item in manifest["inputs"]:
        sheet_name = item["sheet"]
        if sheet_name not in INPUT_SHEETS:
            continue
        rows = {v: k for k, v in _input_rows(sheet_name).items()}
        row = int(item["cell"][1:])
        label = rows.get(row)
        if not label:
            continue
        facts.setdefault(sheet_name, {})[label] = item["value"]
        provenance.setdefault(sheet_name, {})[label] = Provenance(
            source_name=source["name"],
            as_of_date=manifest["as_of"],
            retrieval_date=retrieval_date,
            source_url=source["url"],
            transformation=(
                f"Reused from {HERTZ_MANIFEST_PATH.relative_to(ROOT)} ({item['input_kind']})."
            ),
        )

    return ToolInput(
        instance_slug="public_hertz_restructuring_proxy",
        situation_type="Hertz Chapter 11 emergence financing and recovery (reused public case)",
        as_of=manifest["as_of"],
        facts=facts,
        provenance=provenance,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--demo", action="store_true", help="fictional situation, writes to .agent-tool-scratch/")
    parser.add_argument(
        "--use-hertz-fixture", action="store_true",
        help="real Hertz Chapter 11 exit-financing fact reused from the existing public case, "
        "writes to 24_Distressed_Restructuring/instances/public_hertz_restructuring_proxy.xlsx",
    )
    parser.add_argument("--inputs", type=Path, help="JSON file matching ToolInput shape")
    parser.add_argument(
        "--output-dir", type=Path, default=None,
        help="override instance output directory (default: 24_Distressed_Restructuring/instances/ for --inputs/--use-hertz-fixture)",
    )
    args = parser.parse_args()

    if args.demo:
        out = demo()
    elif args.use_hertz_fixture:
        out = run(hertz_fixture(), instances_dir=args.output_dir)
    elif args.inputs:
        raw = json.loads(args.inputs.read_text())
        provenance = {
            sheet: {
                label: Provenance(**value) if isinstance(value, dict) else value
                for label, value in sheet_prov.items()
            }
            for sheet, sheet_prov in raw.get("provenance", {}).items()
        }
        inp = ToolInput(
            instance_slug=raw["instance_slug"],
            situation_type=raw["situation_type"],
            as_of=raw["as_of"],
            facts=raw.get("facts", {}),
            provenance=provenance,
        )
        out = run(inp, instances_dir=args.output_dir)
    else:
        parser.error("provide --demo, --use-hertz-fixture, or --inputs")
        return 2

    print(json.dumps(asdict(out), indent=2))
    return 0 if out.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
