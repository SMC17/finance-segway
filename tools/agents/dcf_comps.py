"""L3 tool: dcf_comps

Fail-closed agent-facing interface for the 01 Investment Banking flagship's
DCF and Comps sheets. Does not invent math: every number in the output
comes from 01_Investment_Banking/_template_BASE.xlsx's formulas,
recalculated for real by LibreOffice, never computed in this file. This
tool's own job is narrowly the six steps in docs/AGENT_TOOL_CONTRACT.md:
validate provenance, call the governed instance-generation path
(tools/model_instance_release.py), recalculate, read the DCF/Comps outputs,
and report status -- never re-implement the unlevered FCF build, terminal
value, or comps multiples here.

Fourth and last tool on docs/AGENT_TOOL_CONTRACT.md's original
implementation order. Unlike the other three, this domain's two existing
real public cases (ib-public-hp-autonomy-2012-stress,
ib-public-microsoft-linkedin-2016) source only the Transaction Analysis
and Accretion Dilution sheets -- neither touches DCF, Comps, or
Assumptions at all, so there is no existing real case to reuse for a
fixture the way the other three tools' fixtures reused
pe-public-home-depot-2023 / distressed-public-hertz-2021-reorganization.

The real fixture here (--use-adobe-dcf-fixture) instead reuses the raw
XBRL series already fetched for software-public-adobe-fy2025
(tools/data_fabric/out/ADBE_facts_annual_series.json) and derives real
DCF Assumptions from it: FY1 revenue growth, gross margin, opex %,
effective tax rate, D&A-to-capex ratio, capex %, diluted shares, and net
debt are all real, computed facts. WACC and terminal growth are NOT
disclosed by any issuer, ever -- they stay at this repo's own established
precedent (0.09 / 0.025, the same values ib-public-microsoft-linkedin-2016
already used as modeler_assumption discount/growth inputs for a large-cap
tech valuation), submitted here as modeler_assumption, not observed.

Comps peer data (ticker, price, market cap, EV, EV/Revenue for multiple
comparable companies) requires a fresh multi-company market-data pull this
tool does not attempt to source for its shipped fixture -- Comps!C5:F12
stays at template defaults. That is a real, stated gap, not a rounding
choice: the tool's --inputs path fully supports submitting real Comps
facts once sourced; only the shipped fixture doesn't attempt it.

Usage (CLI):
  python tools/agents/dcf_comps.py --demo
  python tools/agents/dcf_comps.py --use-adobe-dcf-fixture
  python tools/agents/dcf_comps.py --inputs path/to/inputs.json

Contract: docs/AGENT_TOOL_CONTRACT.md
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import asdict, dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
DOMAIN = ROOT / "01_Investment_Banking"
TEMPLATE = DOMAIN / "_template_BASE.xlsx"
ADBE_ANNUAL_SERIES = ROOT / "tools" / "data_fabric" / "out" / "ADBE_facts_annual_series.json"

sys.path.insert(0, str(ROOT / "tools"))
sys.path.insert(0, str(ROOT))
from model_instance_release import apply_manifest  # noqa: E402
from recalc import recalc  # noqa: E402


@dataclass
class Provenance:
    source_name: str
    as_of_date: str
    retrieval_date: str
    source_url: str | None = None
    transformation: str = "none"


@dataclass
class ToolInput:
    instance_slug: str
    company_name: str
    as_of: str
    # Assumptions facts: {row_label: value}, always written to column C
    # (FY1) -- this tool does not attempt multi-year Assumptions overrides,
    # matching the sparse-but-honest convention the domain's own real
    # cases already use. Comps facts: {row_number(5-12): {"ticker":...,
    # "price":..., "mkt_cap":..., "ev":..., "ev_rev":...}}, all real fields
    # optional per row so a partial peer set can be submitted honestly.
    assumptions_facts: dict[str, float] = field(default_factory=dict)
    comps_facts: dict[int, dict[str, float | str]] = field(default_factory=dict)
    provenance: dict[str, Provenance] = field(default_factory=dict)


@dataclass
class ToolOutput:
    ok: bool
    checks_status: str  # PASS | REVIEW | FAIL | NOT_RUN (structural only -- see run())
    workbook_path: str | None
    headline: dict[str, Any]
    sources_written: list[dict[str, str]]
    message: str
    refresh_log_entry: str | None = None


def _assumption_rows() -> dict[str, int]:
    workbook = openpyxl.load_workbook(TEMPLATE, data_only=False)
    sheet = workbook["Assumptions"]
    rows: dict[str, int] = {}
    for row in range(5, sheet.max_row + 1):
        label = sheet.cell(row, 2).value
        if label:
            rows[str(label).strip()] = row
    return rows


COMPS_COLUMNS = {"ticker": "B", "price": "C", "mkt_cap": "D", "ev": "E", "ev_rev": "F"}

# Assumptions!C13 ("WACC %") and C14 ("Terminal growth %") exist as labeled
# rows but no formula anywhere in the template reads them -- confirmed by
# grepping every sheet's formulas for "Assumptions!C13"/"C14": zero hits.
# DCF!I5/I6 are the actual, independent live control cells the discount-
# factor and terminal-value formulas reference ($I$5, $I$6 absolute refs).
# Writing only to Assumptions!C13/C14 would be a silent no-op on the real
# calculation -- found the hard way by checking, not assumed. These two
# labels route to DCF directly instead.
#
# "Base revenue (FY0A, $mm)" and "Net debt ($mm)" have no Assumptions row
# at all: IS!F5 (FY1E revenue) is "=E5*(1+Assumptions!C5)" -- it grows
# forward from IS!E5 (FY0A actual revenue), a raw input cell with no
# Assumptions equivalent. Omitting it doesn't error, it silently zeroes
# every projected year (0 * (1+growth) = 0 forever), cascading through
# EBIT, unlevered FCF, and enterprise value all the way to an implied
# value/share of exactly 0 -- caught by CI's real recalc after this
# session's sandbox-blocked recalc couldn't run far enough to show it.
# DCF!I11 ("Less: net debt") is a DCF-sheet-local input in the same vein
# as I5/I6/I13; it defaults to 0 (equivalent to assuming no net debt),
# which recalculates cleanly but silently, so it's included here as an
# explicit real fact rather than left at a default that happens not to
# error.
_DIRECT_CELLS = {
    "WACC %": ("DCF", "I5"),
    "Terminal growth %": ("DCF", "I6"),
    "Base revenue (FY0A, $mm)": ("IS", "E5"),
    "Net debt ($mm)": ("DCF", "I11"),
    "Diluted shares (mm)": ("DCF", "I13"),
}


def validate_provenance(inp: ToolInput) -> list[str]:
    errors = []
    if not inp.company_name or not inp.company_name.strip():
        errors.append("missing required fact: company_name")
    if not inp.assumptions_facts and not inp.comps_facts:
        errors.append("at least one Assumptions or Comps fact is required")
        return errors
    known_labels = _assumption_rows()
    for label in inp.assumptions_facts:
        key = f"Assumptions::{label}"
        if label not in known_labels and label not in _DIRECT_CELLS:
            errors.append(
                f"unknown Assumptions row label: {label!r} "
                f"(not present on {TEMPLATE.relative_to(ROOT)})"
            )
        if key not in inp.provenance:
            errors.append(f"missing provenance for material fact: {key}")
    for row, fields_ in inp.comps_facts.items():
        if not (5 <= row <= 12):
            errors.append(f"Comps row {row} is out of range (template supports rows 5-12)")
            continue
        for field_name in fields_:
            key = f"Comps[{row}]::{field_name}"
            if field_name not in COMPS_COLUMNS:
                errors.append(f"unknown Comps field: {field_name!r} (must be one of {list(COMPS_COLUMNS)})")
            if key not in inp.provenance:
                errors.append(f"missing provenance for material fact: {key}")
    return errors


def _build_manifest(inp: ToolInput, instances_dir: Path) -> dict[str, Any]:
    rows = _assumption_rows()
    inputs = []
    for label, value in inp.assumptions_facts.items():
        prov = inp.provenance[f"Assumptions::{label}"]
        sheet, cell = _DIRECT_CELLS.get(label, ("Assumptions", f"C{rows.get(label)}"))
        inputs.append({
            "sheet": sheet,
            "cell": cell,
            "value": value,
            "source": {"name": prov.source_name, "url": prov.source_url, "as_of": prov.as_of_date, "notes": prov.transformation},
        })
    for row, fields_ in inp.comps_facts.items():
        for field_name, value in fields_.items():
            prov = inp.provenance[f"Comps[{row}]::{field_name}"]
            inputs.append({
                "sheet": "Comps",
                "cell": f"{COMPS_COLUMNS[field_name]}{row}",
                "value": value,
                "source": {"name": prov.source_name, "url": prov.source_url, "as_of": prov.as_of_date, "notes": prov.transformation},
            })
    return {
        "schema_version": "1.0",
        "id": inp.instance_slug,
        # Agent-drafted, not yet independently reviewed -- must never be
        # silently counted as M4 evidence regardless of whether the
        # underlying facts are real or a demo fixture. Human review
        # (Issue #7) is required to promote a draft instance further.
        "classification": "agent_tool_draft",
        "counts_toward_M4": False,
        "template": str(TEMPLATE.relative_to(ROOT)),
        "output": str((instances_dir / f"{inp.instance_slug}.xlsx").relative_to(ROOT)),
        "as_of": inp.as_of,
        "cover": {"Title:": f"{inp.company_name} -- DCF/Comps valuation (agent-drafted)"},
        "inputs": inputs,
        "refresh": {
            "date": inp.as_of,
            "trigger": "L3 agent tool: dcf_comps",
            "what_changed": f"Applied agent-submitted DCF/Comps facts for {inp.company_name}",
            "reviewer_notes": "Generated by tools/agents/dcf_comps.py -- not yet human-reviewed",
            "next_check": "On next material fact refresh",
        },
    }


def _read_headline(workbook_path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    dcf = workbook["DCF"]
    comps = workbook["Comps"]
    return {
        "wacc": dcf["I5"].value,
        "terminal_growth": dcf["I6"].value,
        "enterprise_value": dcf["I10"].value,
        "equity_value": dcf["I12"].value,
        "implied_value_per_share": dcf["I14"].value,
        "comps_median_ev_revenue": comps["F13"].value,
        "comps_median_price": comps["C13"].value,
    }


def _check_workbook_structurally_sound(workbook_path: Path) -> list[str]:
    """No dedicated DCF/Comps row exists in Decision & Checks -- that
    sheet's checks are entirely Transaction-Analysis/Accretion-Dilution
    (M&A deal mechanics), which this tool does not touch. Sanity-check the
    DCF output directly instead of claiming a Checks status this template
    was never wired to compute for a standalone DCF/Comps run."""
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    dcf = workbook["DCF"]
    errors = []
    implied_value = dcf["I14"].value
    if not isinstance(implied_value, (int, float)):
        errors.append(f"implied value/share is not numeric: {implied_value!r}")
    elif implied_value < 0:
        errors.append(f"implied value/share is negative: {implied_value}")
    ev = dcf["I10"].value
    if not isinstance(ev, (int, float)):
        errors.append(f"enterprise value is not numeric: {ev!r}")
    return errors


def run(inp: ToolInput, *, instances_dir: Path | None = None) -> ToolOutput:
    """Execute the tool: validate -> governed instance write -> real
    LibreOffice recalc -> read DCF/Comps outputs. Fails closed on missing
    provenance, a recalculation that doesn't succeed cleanly, or a
    structurally nonsensical DCF result (negative/non-numeric implied
    value or enterprise value).

    instances_dir defaults to 01_Investment_Banking/instances/
    (production). Tests must override it to a scratch directory under the
    repo root -- never write agent-drafted or demo instances into the real
    evidence corpus directory, which is reserved for source-addressed
    public cases (see tests/test_real_data_only.py)."""
    instances_dir = instances_dir or (DOMAIN / "instances")
    errors = validate_provenance(inp)
    if errors:
        return ToolOutput(
            ok=False, checks_status="FAIL", workbook_path=None, headline={},
            sources_written=[], message="fail-closed: " + "; ".join(errors),
        )

    manifest = _build_manifest(inp, instances_dir)
    manifest_path = instances_dir / f"{inp.instance_slug}.manifest.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

    receipt = apply_manifest(manifest_path, ROOT)
    output_path = ROOT / manifest["output"]

    recalc_result = recalc(str(output_path), timeout=60)
    sources_written = [
        {"field": key, **asdict(prov)} for key, prov in inp.provenance.items()
    ]
    if recalc_result.get("status") != "success" or recalc_result.get("total_errors", 0):
        return ToolOutput(
            ok=False, checks_status="FAIL", workbook_path=str(output_path.relative_to(ROOT)),
            headline={}, sources_written=sources_written,
            message=f"fail-closed: recalculation did not succeed cleanly: {recalc_result}",
        )

    structural_errors = _check_workbook_structurally_sound(output_path)
    headline = _read_headline(output_path)

    return ToolOutput(
        ok=not structural_errors,
        checks_status="PASS" if not structural_errors else "FAIL",
        workbook_path=str(output_path.relative_to(ROOT)),
        headline=headline,
        sources_written=sources_written,
        message=(
            "instance generated and recalculated via the governed builder path; "
            "DCF/Comps outputs read back from the recalculated workbook. "
            + ("; ".join(structural_errors) if structural_errors else "No structural DCF issues found.")
            + " Not a client deliverable until stakeholder sign-off is recorded (Issue #7)."
        ),
        refresh_log_entry=receipt.get("as_of"),
    )


SCRATCH_DIR = ROOT / ".agent-tool-scratch" / "01_Investment_Banking"


def demo() -> ToolOutput:
    """Illustrative run with a fully fictional company and no real source
    URLs. Writes to SCRATCH_DIR, never the real
    01_Investment_Banking/instances/ corpus."""
    today = date.today().isoformat()
    facts = {
        "Base revenue (FY0A, $mm)": 1000.0,
        "Revenue growth %": 0.10,
        "Gross margin %": 0.60,
        "Opex % of revenue": 0.35,
        "Tax rate %": 0.21,
        "Diluted shares (mm)": 100.0,
        "WACC %": 0.09,
        "Terminal growth %": 0.025,
    }
    inp = ToolInput(
        instance_slug="demo_dcf_stub",
        company_name="Demo Target Inc.",
        as_of=today,
        assumptions_facts=facts,
        provenance={
            f"Assumptions::{label}": Provenance(
                source_name="demo_fixture", as_of_date=today, retrieval_date=today,
                source_url=None, transformation="demo only -- not for client use",
            )
            for label in facts
        },
    )
    return run(inp, instances_dir=SCRATCH_DIR)


def adobe_dcf_fixture() -> ToolInput:
    """Real DCF Assumptions derived from Adobe's already-fetched XBRL
    annual series (same source as software-public-adobe-fy2025 -- no new
    fetch). Base FY0A revenue, FY1 growth, margin, opex, tax rate,
    D&A/capex ratio, capex %, diluted shares, and net debt are all real,
    computed facts. WACC and terminal growth are not disclosed by any
    issuer and are submitted as modeler_assumption at this repo's own
    established precedent (0.09 / 0.025, same as
    ib-public-microsoft-linkedin-2016's discount/growth inputs).

    Comps peer data is deliberately NOT submitted here -- see this
    module's docstring for why.
    """
    if not ADBE_ANNUAL_SERIES.exists():
        raise FileNotFoundError(f"{ADBE_ANNUAL_SERIES.relative_to(ROOT)} not found")
    series = json.loads(ADBE_ANNUAL_SERIES.read_text(encoding="utf-8"))
    by_concept = {c["concept"]: c["observations"][-1] for c in series["concepts"]}
    by_concept_prior = {c["concept"]: c["observations"][-2] for c in series["concepts"]}

    revenue = by_concept["Revenues"]["value"]
    revenue_prior = by_concept_prior["Revenues"]["value"]
    gross_profit = by_concept["GrossProfit"]["value"]
    operating_income = by_concept["OperatingIncomeLoss"]["value"]
    da = by_concept["DepreciationDepletionAndAmortization"]["value"]
    capex = by_concept["PaymentsToAcquirePropertyPlantAndEquipment"]["value"]
    net_income = by_concept["NetIncomeLoss"]["value"]
    tax = by_concept["IncomeTaxExpenseBenefit"]["value"]
    shares = by_concept["WeightedAverageNumberOfSharesOutstandingBasic"]["value"]
    ltd = by_concept["LongTermDebt"]["value"]
    cash = by_concept["CashAndCashEquivalentsAtCarryingValue"]["value"]
    as_of = by_concept["Revenues"]["end"]

    ebitda = operating_income + da
    opex = gross_profit - ebitda
    pretax_income = net_income + tax
    net_debt = ltd - cash

    facts = {
        "Base revenue (FY0A, $mm)": round(revenue / 1_000_000, 1),
        "Revenue growth %": round(revenue / revenue_prior - 1, 4),
        "Gross margin %": round(gross_profit / revenue, 4),
        "Opex % of revenue": round(opex / revenue, 4),
        "Tax rate %": round(tax / pretax_income, 4),
        "D&A % of capex": round(da / capex, 4),
        "Capex % of revenue": round(capex / revenue, 4),
        "Shares outstanding (mm)": round(shares / 1_000_000, 1),
        "Diluted shares (mm)": round(shares / 1_000_000, 1),
        "Net debt ($mm)": round(net_debt / 1_000_000, 1),
        "WACC %": 0.09,
        "Terminal growth %": 0.025,
    }
    real_labels = {k for k in facts if k not in ("WACC %", "Terminal growth %")}
    retrieval_date = date.today().isoformat()
    xbrl_source = Provenance(
        source_name="Adobe Inc. XBRL annual fact series (SEC EDGAR company facts, CIK 0000796343)",
        as_of_date=as_of,
        retrieval_date=retrieval_date,
        source_url="https://data.sec.gov/api/xbrl/companyfacts/CIK0000796343.json",
        transformation="Reused from tools/data_fabric/out/ADBE_facts_annual_series.json (same source as software-public-adobe-fy2025); ratios computed here from disclosed dollar figures.",
    )
    net_debt_source = Provenance(
        source_name=xbrl_source.source_name, as_of_date=as_of, retrieval_date=retrieval_date,
        source_url=xbrl_source.source_url,
        transformation="Net debt = LongTermDebt - CashAndCashEquivalentsAtCarryingValue, both from the same XBRL series.",
    )
    precedent_source = Provenance(
        source_name="Repo precedent: ib-public-microsoft-linkedin-2016 discount/growth assumptions",
        as_of_date=as_of,
        retrieval_date=retrieval_date,
        source_url=None,
        transformation="No issuer discloses WACC or terminal growth; reused this repo's own prior modeler_assumption values for a large-cap tech valuation rather than choosing a fresh unsourced number.",
    )
    provenance = {
        f"Assumptions::{label}": (
            precedent_source if label in ("WACC %", "Terminal growth %")
            else net_debt_source if label == "Net debt ($mm)"
            else xbrl_source
        )
        for label in facts
    }

    return ToolInput(
        instance_slug="public_adobe_dcf_proxy",
        company_name="Adobe Inc. (NASDAQ: ADBE), CIK 0000796343",
        as_of=as_of,
        assumptions_facts=facts,
        provenance=provenance,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--demo", action="store_true", help="fictional company, writes to .agent-tool-scratch/")
    parser.add_argument(
        "--use-adobe-dcf-fixture", action="store_true",
        help="real Adobe DCF assumptions derived from already-fetched XBRL data (no new fetch), "
        "writes to 01_Investment_Banking/instances/public_adobe_dcf_proxy.xlsx",
    )
    parser.add_argument("--inputs", type=Path, help="JSON file matching ToolInput shape")
    parser.add_argument(
        "--output-dir", type=Path, default=None,
        help="override instance output directory (default: 01_Investment_Banking/instances/ for --inputs/--use-adobe-dcf-fixture)",
    )
    args = parser.parse_args()

    if args.demo:
        out = demo()
    elif args.use_adobe_dcf_fixture:
        out = run(adobe_dcf_fixture(), instances_dir=args.output_dir)
    elif args.inputs:
        raw = json.loads(args.inputs.read_text())
        provenance = {
            key: Provenance(**value) if isinstance(value, dict) else value
            for key, value in raw.get("provenance", {}).items()
        }
        inp = ToolInput(
            instance_slug=raw["instance_slug"],
            company_name=raw["company_name"],
            as_of=raw["as_of"],
            assumptions_facts=raw.get("assumptions_facts", {}),
            comps_facts={int(k): v for k, v in raw.get("comps_facts", {}).items()},
            provenance=provenance,
        )
        out = run(inp, instances_dir=args.output_dir)
    else:
        parser.error("provide --demo, --use-adobe-dcf-fixture, or --inputs")
        return 2

    print(json.dumps(asdict(out), indent=2))
    return 0 if out.ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
