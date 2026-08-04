"""Validate the institutional LBO workbook contract."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from workbook_engineering import audit_workbook

REQUIRED_SHEETS = {
    "Cover", "Assumptions", "Sources & Uses", "Operating Model",
    "Debt Schedule", "Covenants", "Management Equity", "Returns Waterfall",
    "Sensitivity", "Checks", "Sources", "RefreshLog",
}
REQUIRED_LABELS = {
    "Transaction Sources & Uses",
    "Seven-Year Operating, Tax, and Cash Flow Model",
    "Multi-Tranche Debt and Cash Sweep",
    "Revolver draw / (repayment)",
    "Second-lien PIK",
    "Management Equity and Incentive Pool",
    "Sponsor IRR",
    "Sponsor IRR Sensitivity",
}


def validate(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    errors: list[str] = []
    missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing:
        errors.append(f"missing sheets: {missing}")
    values = {
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row if cell.value is not None
    }
    missing_labels = sorted(REQUIRED_LABELS - values)
    if missing_labels:
        errors.append(f"missing labels: {missing_labels}")
    formulas = [
        cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    if len(formulas) < 300:
        errors.append(f"formula depth below contract: {len(formulas)} < 300")
    debt = workbook["Debt Schedule"] if "Debt Schedule" in workbook.sheetnames else None
    if debt is not None:
        expected = {
            "C14": "='Sources & Uses'!F5",
            "C17": "='Sources & Uses'!F6",
            "C22": "='Sources & Uses'!F7",
            "C27": "=SUM(C16,C21,C26)",
            "C28": "=C27-C13",
        }
        for address, formula in expected.items():
            if debt[address].value != formula:
                errors.append(f"close-state debt identity mismatch at {address}: {debt[address].value}")
        for column in range(4, 11):
            if debt.cell(27, column).value is None or debt.cell(28, column).value is None:
                errors.append(f"missing total/net debt formulas in column {column}")
    summary, findings = audit_workbook(path)
    if summary["external_links"]:
        errors.append("external links present")
    errors.extend(
        f"{finding.code}:{finding.sheet}:{finding.cell}:{finding.message}"
        for finding in findings if finding.severity == "error"
    )
    checks = workbook["Checks"] if "Checks" in workbook.sheetnames else None
    if checks is not None:
        check_formulas = [
            cell.value for row in checks.iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if len(check_formulas) < 9:
            errors.append("insufficient LBO checks")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    errors = validate(args.workbook)
    if errors:
        for error in errors:
            print(f"ERROR: {error}")
        return 1
    print(f"validated {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
