"""Build the canonical project-finance and infrastructure decision model."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    formula_cell, header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "Construction", "Sources & Uses", "Operations",
    "Debt Sculpting", "DSRA", "Coverage", "Sensitivity", "Checks",
    "Sources", "RefreshLog",
]


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[PROJECT] — Project Finance Model", [
        ("Project / concession:", "[Name]"),
        ("Sector / jurisdiction:", "Power / transport / digital / PPP"),
        ("Financial close / COD:", "[dates]"),
        ("Last refreshed:", "[date]"),
        ("Refresh cadence:", "Weekly construction; monthly operations"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Project Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("EPC / hard construction cost", 800.0, 880.0, "$mm", CUR),
        ("Development and owner costs", 80.0, 95.0, "$mm", CUR),
        ("Contingency / hard cost", 0.08, 0.12, "%", PCT),
        ("Debt share of pre-IDC cost", 0.70, 0.65, "%", PCT),
        ("Construction quarters", 8.0, 10.0, "quarters", "0"),
        ("Annual construction debt rate", 0.075, 0.090, "%", PCT2),
        ("Year 1 contracted / merchant revenue", 180.0, 150.0, "$mm", CUR),
        ("Annual revenue growth", 0.025, 0.000, "%", PCT),
        ("Operating cost / revenue", 0.28, 0.36, "%", PCT),
        ("Maintenance capex / revenue", 0.035, 0.050, "%", PCT),
        ("Cash tax / EBITDA", 0.15, 0.10, "%", PCT),
        ("Target DSCR", 1.35, 1.45, "x", MULT),
        ("Minimum DSCR covenant", 1.20, 1.20, "x", MULT),
        ("Annual operating debt rate", 0.065, 0.080, "%", PCT2),
        ("Debt tenor", 10.0, 10.0, "years", "0"),
        ("DSRA months of forward debt service", 6.0, 9.0, "months", "0.0"),
        ("Discount rate for LLCR / PLCR", 0.070, 0.085, "%", PCT2),
        ("Residual / decommissioning reserve per year", 2.0, 4.0, "$mm", CUR),
        ("Availability / volume factor", 0.97, 0.85, "%", PCT),
        ("Inflation escalation", 0.020, 0.030, "%", PCT),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        formula_cell(
            sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})'),
            number_format,
            cross_sheet=True,
        )
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 46, "C": 15, "D": 15, "E": 15, "F": 32})
    sheet.freeze_panes = "A5"

    sheet = workbook["Construction"]
    title(sheet, "B2:M2", "Quarterly Construction Funding and IDC")
    headers = ["$mm", *[f"Q{period}" for period in range(1, 11)], "Total"]
    header(sheet, 4, 2, headers)
    labels = [
        "Spend curve", "Hard construction cost", "Development / owner costs",
        "Contingency", "Total pre-IDC spend", "Equity draw", "Debt draw",
        "Beginning construction debt", "Interest during construction",
        "Ending construction debt", "Cumulative project spend",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    spend_curve = [0.05, 0.10, 0.15, 0.18, 0.18, 0.14, 0.12, 0.08, 0.0, 0.0]
    for column, weight in enumerate(spend_curve, start=3):
        input_cell(sheet.cell(5, column, weight), PCT)
        letter = get_column_letter(column)
        period = column - 2
        sheet.cell(6, column, f'=IF({period}<=Assumptions!$E$9,Assumptions!$E$5*{letter}$5,0)')
        sheet.cell(7, column, f'=IF({period}<=Assumptions!$E$9,Assumptions!$E$6/Assumptions!$E$9,0)')
        sheet.cell(8, column, f"={letter}6*Assumptions!$E$7")
        sheet.cell(9, column, f"=SUM({letter}6:{letter}8)")
        sheet.cell(10, column, f"={letter}9*(1-Assumptions!$E$8)")
        sheet.cell(11, column, f"={letter}9*Assumptions!$E$8")
        if column == 3:
            sheet.cell(12, column, 0.0)
            sheet.cell(15, column, f"={letter}9")
        else:
            previous = get_column_letter(column - 1)
            sheet.cell(12, column, f"={previous}14")
            sheet.cell(15, column, f"={previous}15+{letter}9")
        sheet.cell(13, column, f"=({letter}12+0.5*{letter}11)*Assumptions!$E$10/4")
        sheet.cell(14, column, f"={letter}12+{letter}11+{letter}13")
        for row in range(6, 16):
            sheet.cell(row, column).number_format = CUR
    total_column = 13
    sheet.cell(5, total_column, "=SUM(C5:L5)")
    sheet.cell(5, total_column).number_format = PCT
    for row in range(6, 14):
        sheet.cell(row, total_column, f"=SUM(C{row}:L{row})")
        sheet.cell(row, total_column).number_format = CUR
    sheet.cell(14, total_column, "=L14")
    sheet.cell(15, total_column, "=L15")
    sheet.cell(14, total_column).number_format = CUR
    sheet.cell(15, total_column).number_format = CUR
    for row in (9, 14, 15):
        total_row(sheet, row, 2, 13, CUR)
    set_widths(sheet, {"A": 4, "B": 36, **{get_column_letter(column): 12 for column in range(3, 14)}})
    sheet.freeze_panes = "C5"

    sheet = workbook["Sources & Uses"]
    title(sheet, "B2:G2", "Sources & Uses at Financial Close")
    header(sheet, 4, 2, ["Uses", "Amount", "", "Sources", "Amount", ""])
    uses = [
        ("Hard construction cost", "=Construction!M6"),
        ("Development / owner costs", "=Construction!M7"),
        ("Contingency", "=Construction!M8"),
        ("Interest during construction", "=Construction!M13"),
        ("Initial DSRA funding", "='DSRA'!C7"),
        ("Total uses", "=SUM(C5:C9)"),
    ]
    sources = [
        ("Construction / term debt", "=Construction!M14"),
        ("Sponsor equity", "=MAX(0,C10-F5)"),
        ("Grants / subordinated support", 0.0),
        ("Total sources", "=SUM(F5:F7)"),
        ("Sources less uses", "=F8-C10"),
    ]
    for row, (label, value) in enumerate(uses, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, value)
        sheet.cell(row, 3).number_format = CUR
    for row, (label, value) in enumerate(sources, start=5):
        sheet.cell(row, 5, label)
        cell = sheet.cell(row, 6, value)
        if not isinstance(value, str):
            input_cell(cell, CUR)
        else:
            cell.number_format = CUR
    total_row(sheet, 10, 2, 3, CUR)
    total_row(sheet, 8, 5, 6, CUR)
    set_widths(sheet, {"A": 4, "B": 34, "C": 18, "D": 5, "E": 34, "F": 18, "G": 5})

    sheet = workbook["Operations"]
    title(sheet, "B2:L2", "Ten-Year Operating Forecast and CFADS")
    header(sheet, 4, 2, ["$mm", *[f"Year {year}" for year in range(1, 11)]])
    labels = [
        "Nominal contracted / merchant revenue", "Availability-adjusted revenue",
        "Operating cost", "EBITDA", "Cash taxes", "Maintenance capex",
        "Decommissioning / lifecycle reserve", "CFADS",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 13):
        letter = get_column_letter(column)
        year = column - 2
        if column == 3:
            sheet.cell(5, column, "=Assumptions!E11")
        else:
            previous = get_column_letter(column - 1)
            sheet.cell(5, column, f"={previous}5*(1+Assumptions!$E$12+Assumptions!$E$24)")
        sheet.cell(6, column, f"={letter}5*Assumptions!$E$23")
        sheet.cell(7, column, f"={letter}6*Assumptions!$E$13")
        sheet.cell(8, column, f"={letter}6-{letter}7")
        sheet.cell(9, column, f"=MAX(0,{letter}8*Assumptions!$E$15)")
        sheet.cell(10, column, f"={letter}6*Assumptions!$E$14")
        sheet.cell(11, column, "=Assumptions!$E$22")
        sheet.cell(12, column, f"={letter}8-{letter}9-{letter}10-{letter}11")
        for row in range(5, 13):
            sheet.cell(row, column).number_format = CUR
    for row in (8, 12):
        total_row(sheet, row, 2, 12, CUR)
    set_widths(sheet, {"A": 4, "B": 42, **{get_column_letter(column): 13 for column in range(3, 13)}})
    sheet.freeze_panes = "C5"

    sheet = workbook["Debt Sculpting"]
    title(sheet, "B2:L2", "Sculpted Term Debt Schedule")
    header(sheet, 4, 2, ["$mm / x", *[f"Year {year}" for year in range(1, 11)]])
    labels = [
        "Beginning debt", "Cash interest", "Maximum debt service at target DSCR",
        "Scheduled principal", "Ending debt", "Actual debt service", "CFADS",
        "DSCR", "Discount factor", "PV of CFADS", "LLCR from each year",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 13):
        letter = get_column_letter(column)
        operations_column = letter
        if column == 3:
            sheet.cell(5, column, "=Construction!M14")
        else:
            previous = get_column_letter(column - 1)
            sheet.cell(5, column, f"={previous}9")
        sheet.cell(6, column, f"={letter}5*Assumptions!$E$18")
        sheet.cell(7, column, f"=Operations!{operations_column}12/Assumptions!$E$16")
        maturity_payment = f"{letter}5" if column == 12 else "0"
        sheet.cell(8, column, f"=MIN({letter}5,MAX({maturity_payment},MAX(0,{letter}7-{letter}6)))")
        sheet.cell(9, column, f"=MAX(0,{letter}5-{letter}8)")
        sheet.cell(10, column, f"={letter}6+{letter}8")
        sheet.cell(11, column, f"=Operations!{operations_column}12")
        sheet.cell(12, column, f"=IFERROR({letter}11/{letter}10,0)")
        sheet.cell(13, column, f"=1/(1+Assumptions!$E$21)^{column-2}")
        sheet.cell(14, column, f"={letter}11*{letter}13")
        sheet.cell(15, column, f"=IFERROR(SUM({letter}14:$L$14)/{letter}5,0)")
        for row in range(5, 12):
            sheet.cell(row, column).number_format = CUR
        sheet.cell(12, column).number_format = MULT
        sheet.cell(13, column).number_format = "0.0000"
        sheet.cell(14, column).number_format = CUR
        sheet.cell(15, column).number_format = MULT
    for row in (9, 10):
        total_row(sheet, row, 2, 12, CUR)
    set_widths(sheet, {"A": 4, "B": 44, **{get_column_letter(column): 13 for column in range(3, 13)}})
    sheet.freeze_panes = "C5"

    sheet = workbook["DSRA"]
    title(sheet, "B2:L2", "Debt Service Reserve Account")
    header(sheet, 4, 2, ["$mm", *[f"Year {year}" for year in range(1, 11)]])
    labels = [
        "Forward debt service", "Required DSRA", "Beginning DSRA",
        "Funding / (release)", "Ending DSRA", "DSRA funding at close",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 13):
        letter = get_column_letter(column)
        next_letter = get_column_letter(column + 1) if column < 12 else letter
        sheet.cell(5, column, f"='Debt Sculpting'!{next_letter}10")
        sheet.cell(6, column, f"={letter}5*Assumptions!$E$20/12")
        if column == 3:
            sheet.cell(7, column, f"={letter}6")
        else:
            previous = get_column_letter(column - 1)
            sheet.cell(7, column, f"={previous}9")
        sheet.cell(8, column, f"={letter}6-{letter}7")
        sheet.cell(9, column, f"={letter}7+{letter}8")
        sheet.cell(10, column, f"=IF(COLUMN()=3,{letter}7,0)")
        for row in range(5, 11):
            sheet.cell(row, column).number_format = CUR
    set_widths(sheet, {"A": 4, "B": 36, **{get_column_letter(column): 13 for column in range(3, 13)}})

    sheet = workbook["Coverage"]
    title(sheet, "B2:L2", "Coverage, Liquidity, and Tail Risk")
    header(sheet, 4, 2, ["Metric", *[f"Year {year}" for year in range(1, 11)]])
    labels = [
        "DSCR", "Minimum DSCR", "DSCR headroom", "LLCR", "PLCR",
        "Debt / EBITDA", "DSRA coverage", "Covenant status",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 13):
        letter = get_column_letter(column)
        sheet.cell(5, column, f"='Debt Sculpting'!{letter}12")
        sheet.cell(6, column, "=Assumptions!$E$17")
        sheet.cell(7, column, f"={letter}5-{letter}6")
        sheet.cell(8, column, f"='Debt Sculpting'!{letter}15")
        sheet.cell(9, column, f"=IFERROR(SUM('Debt Sculpting'!{letter}14:$L$14)/'Debt Sculpting'!{letter}5,0)")
        sheet.cell(10, column, f"=IFERROR('Debt Sculpting'!{letter}9/Operations!{letter}8,0)")
        sheet.cell(11, column, f"=IFERROR(DSRA!{letter}9/DSRA!{letter}6,0)")
        sheet.cell(12, column, f'=IF(AND({letter}7>=0,{letter}8>=1,{letter}11>=1),"PASS","BREACH")')
        for row in range(5, 12):
            sheet.cell(row, column).number_format = MULT
    add_status_rules(sheet, "C12:L12")
    set_widths(sheet, {"A": 4, "B": 32, **{get_column_letter(column): 13 for column in range(3, 13)}})

    sheet = workbook["Sensitivity"]
    title(sheet, "B2:G2", "Minimum DSCR Sensitivity")
    header(sheet, 4, 2, ["Revenue haircut / Operating cost", 0.22, 0.28, 0.34, 0.40, 0.46])
    for column in range(3, 8):
        sheet.cell(4, column).number_format = PCT
    for row, haircut in enumerate((0.00, 0.05, 0.10, 0.15, 0.20), start=5):
        sheet.cell(row, 2, haircut)
        sheet.cell(row, 2).number_format = PCT
        for column in range(3, 8):
            letter = get_column_letter(column)
            sheet.cell(row, column, f"=MIN((Operations!C5:L5*(1-$B{row})*(1-{letter}$4)-Operations!C9:L11)/'Debt Sculpting'!C10:L10)")
            sheet.cell(row, column).number_format = MULT
    sheet.conditional_formatting.add(
        "C5:G9",
        ColorScaleRule(
            start_type="min", start_color="FCE4D6",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="E2F0D9",
        ),
    )
    set_widths(sheet, {"A": 4, "B": 34, **{get_column_letter(column): 13 for column in range(3, 8)}})

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("Construction spend curve totals 100%", '=IF(ABS(Construction!M5-1)<0.000001,"PASS","FAIL")'),
        ("Sources equal uses", '=IF(ABS(\'Sources & Uses\'!F9)<0.000001,"PASS","FAIL")'),
        ("Debt never negative", '=IF(MIN(\'Debt Sculpting\'!C9:L9)>=0,"PASS","FAIL")'),
        ("No DSCR covenant breach", '=IF(MIN(Coverage!C7:L7)>=0,"PASS","REVIEW")'),
        ("DSRA fully funded", '=IF(MIN(Coverage!C11:L11)>=1,"PASS","REVIEW")'),
        ("LLCR positive", '=IF(MIN(Coverage!C8:L8)>0,"PASS","FAIL")'),
        ("Debt repaid or balloon disclosed", '=IF(\'Debt Sculpting\'!L9<0.000001,"PASS","REVIEW")'),
        ("Overall model status", '=IF(COUNTIF(C5:C11,"FAIL")=0,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    add_status_rules(sheet, "C5:C12")
    set_widths(sheet, {"A": 4, "B": 46, "C": 18})

    add_sources(workbook, [
        ("EPC and construction budget", "[data room / contract URL]", "[date]", "Cost, timing, contingency, liquidated damages"),
        ("Concession / offtake agreement", "[public filing / contract URL]", "[date]", "Tariff, availability, escalation, termination"),
        ("Debt term sheet and common terms", "[data room URL]", "[date]", "Rates, tenor, sculpting, covenants, reserves"),
        ("Independent engineer report", "[advisor URL]", "[date]", "Construction and operating assumptions"),
        ("Market / resource study", "[consultant or public source]", "[date]", "Demand, irradiation, wind, traffic, throughput"),
        ("Discount and benchmark rates", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates", "[date]", "Document currency and tenor mapping"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("PROJECT_FINANCE_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
