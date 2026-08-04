import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[COMPANY] — Venture Capital Model", [
    ("Sector:", "[fill in]"),
    ("Current stage:", "Seed / A / B / C / Growth"),
    ("Last refreshed:", "[date]"),
    ("Next round expected:", "[date]"),
    ("Refresh cadence:", "Weekly"),
])

# ---------------- CAP TABLE ----------------
ws = wb.create_sheet("Cap Table")
set_col_widths(ws, [4, 20, 14, 14, 14, 14, 14])
ws["B2"] = "Capitalization Table"; ws["B2"].font = TITLE
headers = ["", "Holder / Class", "Shares", "% Fully Diluted", "Price paid/sh", "Invested $", "Class"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

rows = ["Founders (common)", "Employee option pool", "SAFE holders (pre-conversion)",
        "Seed preferred", "Series A preferred", "Series B preferred"]
r = 5
for label in rows:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c_sh = ws.cell(row=r, column=3, value=0); c_sh.font = BLUE; c_sh.number_format = NUM; c_sh.border = BORDER
    c_price = ws.cell(row=r, column=5, value=0); c_price.font = BLUE; c_price.number_format = CUR2; c_price.border = BORDER
    c_inv = ws.cell(row=r, column=6, value=f"=C{r}*E{r}"); c_inv.number_format = CUR; c_inv.border = BORDER
    ws.cell(row=r, column=7, value="Common/Pref").font = BLUE
    r += 1
total_row = r
ws.cell(row=total_row, column=2, value="Total").font = BOLD
ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})").font = BOLD
ws.cell(row=total_row, column=3).number_format = NUM
ws.cell(row=total_row, column=6, value=f"=SUM(F5:F{total_row-1})").font = BOLD
ws.cell(row=total_row, column=6).number_format = CUR
for r2 in range(5, total_row):
    cell = ws.cell(row=r2, column=4, value=f"=IFERROR(C{r2}/$C${total_row},\"-\")")
    cell.number_format = PCT
ws.sheet_view.showGridLines = False

# ---------------- ROUND MODELING ----------------
ws = wb.create_sheet("Round Modeling")
set_col_widths(ws, [4, 26, 16, 4, 26, 16])
ws["B2"] = "New Round Modeling"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [("Pre-money valuation", 0, CUR), ("New money raised", 0, CUR),
          ("New option pool top-up %", 0, PCT), ("Existing fully diluted shares", 0, NUM)]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.number_format = fmt
    c.fill = YELLOW_FILL; c.border = BORDER
    r += 1

ws["E4"] = "Outputs"; ws["E4"].font = BOLD; ws["E4"].fill = GRAY_FILL
ws["E5"] = "Post-money valuation"; ws["F5"] = "=C5+C6"; ws["F5"].number_format = CUR
ws["E6"] = "Price per share"; ws["F6"] = "=IFERROR(C5/C8,\"-\")"; ws["F6"].number_format = CUR2
ws["E7"] = "New shares issued (investor)"; ws["F7"] = "=IFERROR(C6/F6,\"-\")"; ws["F7"].number_format = NUM
ws["E8"] = "New pool shares (top-up)"; ws["F8"] = "=IFERROR(C7*(C8+F7),\"-\")"; ws["F8"].number_format = NUM
ws["E9"] = "Total post-round shares"; ws["F9"] = "=IFERROR(C8+F7+F8,\"-\")"; ws["F9"].font = BOLD; ws["F9"].number_format = NUM
ws["E10"] = "New investor ownership %"; ws["F10"] = "=IFERROR(F7/F9,\"-\")"; ws["F10"].number_format = PCT
ws["E11"] = "Existing holder dilution %"; ws["F11"] = "=IFERROR(1-C8/F9,\"-\")"; ws["F11"].number_format = PCT
for row in range(5, 12):
    ws.cell(row=row, column=6).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- SAFE CONVERSION ----------------
ws = wb.create_sheet("SAFE Conversion")
set_col_widths(ws, [4, 28, 14, 4, 28, 16])
ws["B2"] = "SAFE / Convertible Note Conversion"; ws["B2"].font = TITLE
ws["B4"] = "SAFE Terms"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
terms = [("SAFE investment amount", 0, CUR), ("Valuation cap", 0, CUR), ("Discount %", 0, PCT)]
r = 5
for label, default, fmt in terms:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["E4"] = "At Priced Round"; ws["E4"].font = BOLD; ws["E4"].fill = GRAY_FILL
ws["E5"] = "Priced round price/share"; ws["F5"] = "='Round Modeling'!F6"; ws["F5"].font = GREEN; ws["F5"].number_format = CUR2
ws["E6"] = "Cap price/share (cap / pre-round FD shares)"
ws["F6"] = "=IFERROR(C6/'Round Modeling'!C8,\"-\")"; ws["F6"].number_format = CUR2
ws["E7"] = "Discount price/share"
ws["F7"] = "=IFERROR(F5*(1-C7),\"-\")"; ws["F7"].number_format = CUR2
ws["E8"] = "SAFE conversion price (lower of cap/discount/round)"
ws["F8"] = '=IFERROR(MIN(IF(ISNUMBER(F5),F5,9E99),IF(ISNUMBER(F6),F6,9E99),IF(ISNUMBER(F7),F7,9E99)),"-")'; ws["F8"].font = BOLD; ws["F8"].number_format = CUR2
ws["E9"] = "SAFE shares issued"
ws["F9"] = "=IFERROR(C5/F8,\"-\")"; ws["F9"].font = BOLD; ws["F9"].number_format = NUM
for row in range(5, 10):
    ws.cell(row=row, column=6).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- EXIT WATERFALL ----------------
ws = wb.create_sheet("Exit Waterfall")
set_col_widths(ws, [4, 26, 14, 14, 14, 14, 16, 16])
ws["B2"] = "Exit Proceeds Waterfall (1x non-participating pref)"; ws["B2"].font = TITLE
headers = ["", "Class", "Invested $", "Liq. pref (1x)", "As-converted %",
           "Pref-stack scenario", "As-converted scenario", "Actual proceeds"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 7)
ws["B15"] = "Total exit proceeds"; ws["C15"] = 0; ws["C15"].font = BLUE; ws["C15"].fill = YELLOW_FILL
ws["C15"].number_format = CUR

# Cap Table rows: 5 Founders, 6 Employee pool, 7 SAFE, 8 Seed pref, 9 Series A pref, 10 Series B pref.
# Exit Waterfall is seniority order (most senior first) — map each class to ITS OWN row, not by
# list position (an earlier version of this tab mapped by index and pulled the wrong rows entirely).
class_rows = [
    ("Series B preferred", "='Cap Table'!F10", "=IFERROR('Cap Table'!D10,\"-\")"),
    ("Series A preferred", "='Cap Table'!F9", "=IFERROR('Cap Table'!D9,\"-\")"),
    ("Seed preferred", "='Cap Table'!F8", "=IFERROR('Cap Table'!D8,\"-\")"),
    ("Common (founders + pool)", "='Cap Table'!F5+'Cap Table'!F6",
     "=IFERROR('Cap Table'!D5+'Cap Table'!D6,\"-\")"),
]
r = 5
for cls, invested_formula, pct_formula in class_rows:
    ws.cell(row=r, column=2, value=cls).font = BLACK
    ws.cell(row=r, column=3, value=invested_formula).font = GREEN
    ws.cell(row=r, column=3).number_format = CUR
    ws.cell(row=r, column=4, value=f"=C{r}").number_format = CUR  # 1x pref = invested amount
    ws.cell(row=r, column=5, value=pct_formula).font = GREEN
    ws.cell(row=r, column=5).number_format = PCT
    for c in range(3, 6):
        ws.cell(row=r, column=c).border = BORDER
    r += 1

# Pref-stack scenario: senior-to-junior cascade, same cascading MIN/MAX as a
# recovery waterfall. Common has no liquidation preference — it gets whatever
# preferred didn't claim, which is what keeps this scenario self-consistent
# (rows 5-8 always sum to exactly total proceeds, by construction).
ws["F5"] = "=MIN($C$15,D5)"
ws["F6"] = "=MIN(MAX($C$15-F5,0),D6)"
ws["F7"] = "=MIN(MAX($C$15-F5-F6,0),D7)"
ws["F8"] = "=MAX(0,$C$15-F5-F6-F7)"
for r2 in (5, 6, 7, 8):
    ws.cell(row=r2, column=6).number_format = CUR
    ws.cell(row=r2, column=6).border = BORDER

# As-converted scenario: pro-rata by fully-diluted %, for all four rows —
# also self-consistent (the %s sum to ~100%, so this sums to total proceeds).
# Guarded: E-column % can be text ("-") on a blank cap table, and text*number
# throws #VALUE! rather than quietly returning zero.
for r2 in range(5, 9):
    ws.cell(row=r2, column=7, value=f"=IFERROR($C$15*E{r2},\"-\")")
    ws.cell(row=r2, column=7).number_format = CUR
    ws.cell(row=r2, column=7).border = BORDER

# Actual: pick ONE scenario for the WHOLE table based on a single global
# test, rather than letting each class independently choose MAX(its own
# scenario values) -- that looks tempting but breaks conservation. Proof by
# example: if Series B and A both take their full pref (consuming all of a
# small exit) while Seed independently "converts" into what its fully-diluted
# % implies from the *original* total, Seed's slice was already spoken for by
# B and A -- the three payouts sum to MORE than total proceeds. Neither
# per-class scenario has that problem on its own (each is a self-contained,
# fully-allocated cap table split), so switching between them wholesale for
# every class keeps the total exact. The real-world outcome is often a mix
# (junior classes convert while senior ones hold pref) — this simplification
# picks the safer of the two extremes rather than guessing the exact mix.
ws["B10"] = "Regime: as-converted pays more fund-wide than the full pref stack?"
ws["C10"] = "=IF($C$15>SUM(D5:D7),\"YES — use as-converted\",\"NO — use pref-stack\")"
ws["C10"].font = BOLD
for r2 in range(5, 9):
    ws.cell(row=r2, column=8, value=f"=IF($C$15>SUM($D$5:$D$7),G{r2},F{r2})")
    ws.cell(row=r2, column=8).font = BOLD
    ws.cell(row=r2, column=8).number_format = CUR
    ws.cell(row=r2, column=8).border = BORDER

ws["B11"] = "Total distributed (check — should equal total exit proceeds)"
ws["C11"] = "=SUM(H5:H8)"; ws["C11"].font = BOLD; ws["C11"].number_format = CUR
ws["C11"].border = BORDER

ws["B13"] = ("Simplification: switches the WHOLE cap table between the pref-stack and as-converted "
             "scenario based on one global test (total proceeds vs. sum of all preferences), rather "
             "than letting each class elect independently. Real waterfalls often have junior classes "
             "convert while senior ones keep their pref, in the middle range between these two "
             "extremes — that requires testing conversion elections class by class (a small solver, "
             "not a single formula). This always ties to total proceeds and is exactly right at both "
             "extremes; treat the boundary/middle range as a case for a proper waterfall tool or "
             "counsel, not this template.")
ws["B13"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

# ---------------- COMPARABLE FINANCINGS ----------------
ws = wb.create_sheet("Comparable Financings")
set_col_widths(ws, [4, 20, 14, 16, 16, 16, 20])
ws["B2"] = "Comparable Financings"; ws["B2"].font = TITLE
for i, h in enumerate(["", "Company", "Round", "Date", "Pre-money", "Amount raised", "Lead investor"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(["Company","Round","Date","Pre-money","Amount raised","Lead investor"]))
for r in range(5, 13):
    for c in range(2, 8):
        cell = ws.cell(row=r, column=c, value="[fill in]" if c != 5 and c != 6 else 0)
        cell.font = BLUE
        cell.border = BORDER
        if c in (5, 6):
            cell.number_format = CUR
ws.sheet_view.showGridLines = False

add_refresh_log(wb)

out_path = "VC_template.xlsx"
wb.save(out_path)
print("saved", out_path)
