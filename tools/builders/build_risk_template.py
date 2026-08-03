import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[PORTFOLIO] — Risk Management Model", [
    ("Portfolio:", "[fill in]"),
    ("Base currency:", "USD"),
    ("Last refreshed:", "[date]"),
    ("Risk review cadence:", "Weekly (daily in stress periods)"),
])

# ---------------- VAR ----------------
ws = wb.create_sheet("VaR")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Value at Risk — Parametric (Variance-Covariance)"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Portfolio value ($)", 0, CUR),
    ("Daily volatility (sigma, %)", 0.015, PCT2),
    ("Confidence level (e.g. 0.95, 0.99)", 0.95, PCT),
    ("Holding period (days)", 1, NUM),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B10"] = "Outputs"; ws["B10"].font = BOLD; ws["B10"].fill = GRAY_FILL
ws["B11"] = "Z-score for confidence level"
ws["C11"] = "=NORMSINV(C7)"; ws["C11"].number_format = '0.0000'
ws["B12"] = "1-day VaR ($)"
ws["C12"] = "=C5*C6*C11"; ws["C12"].number_format = CUR
ws["B13"] = f"N-day VaR ($, scaled by sqrt(time))"
ws["C13"] = "=C12*SQRT(C8)"; ws["C13"].font = BOLD; ws["C13"].number_format = CUR
ws["D13"] = "Square-root-of-time scaling assumes i.i.d. returns — a simplification, not exact for fat tails"
ws["D13"].font = ITALIC_GRAY
ws["B14"] = "VaR as % of portfolio"
ws["C14"] = "=IFERROR(C13/C5,\"-\")"; ws["C14"].number_format = PCT
ws["B15"] = "Expected Shortfall / CVaR (1-day, $)"
ws["C15"] = "=IFERROR(C5*C6*NORMDIST(C11,0,1,FALSE)/(1-C7),\"-\")"
ws["C15"].font = BOLD; ws["C15"].number_format = CUR
ws["D15"] = "Analytic ES under normality: sigma x phi(Z) / (1-confidence). Always >= VaR — it's the average loss GIVEN that VaR is breached, not just the threshold."
ws["D15"].font = ITALIC_GRAY
for r2 in range(11, 16):
    ws.cell(row=r2, column=3).border = BORDER

ws["B18"] = "Historical VaR (alternative method)"; ws["B18"].font = BOLD; ws["B18"].fill = GRAY_FILL
ws["B19"] = "Enter trailing daily P&L observations below (min 100 recommended)"
ws["B19"].font = ITALIC_GRAY
ws["B20"] = "Historical VaR = PERCENTILE of P&L distribution at (1-confidence)"
ws["B20"].font = ITALIC_GRAY
ws["B21"] = "Historical VaR ($)"
ws["C21"] = "=IFERROR(-PERCENTILE(E23:E122,1-C7),\"-\")"
ws["C21"].font = BOLD; ws["C21"].number_format = CUR
ws["C21"].border = BORDER
ws["B22"] = "Historical CVaR / ES ($, avg of losses beyond the VaR percentile)"
ws["C22"] = ("=IFERROR(-AVERAGEIF(E23:E122,\"<=\"&PERCENTILE(E23:E122,1-C7)),\"-\")")
ws["C22"].number_format = CUR
ws["C22"].border = BORDER
ws["E22"] = "Daily P&L ($)"; ws["E22"].font = BOLD; ws["E22"].fill = GRAY_FILL
for r2 in range(23, 123):
    c = ws.cell(row=r2, column=5, value=None)
    c.font = BLUE
    c.number_format = CUR
    c.border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- STRESS SCENARIOS ----------------
ws = wb.create_sheet("Stress Scenarios")
set_col_widths(ws, [4, 26, 14, 14, 40])
ws["B2"] = "Stress Scenarios"; ws["B2"].font = TITLE
headers = ["", "Scenario", "Shock applied", "Portfolio impact ($)", "Notes"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers)-1)
scenarios = [
    ("2008-style equity crash", -0.40),
    ("Rate shock +200bp", -0.10),
    ("Credit spread widening +300bp", -0.08),
    ("FX shock (base currency -15%)", -0.15),
    ("Liquidity event (haircut on illiquids)", -0.25),
]
r = 5
for label, shock in scenarios:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c_shock = ws.cell(row=r, column=3, value=shock); c_shock.font = BLUE; c_shock.number_format = PCT; c_shock.border = BORDER
    c_impact = ws.cell(row=r, column=4, value=f"='VaR'!$C$5*C{r}")
    c_impact.number_format = CUR; c_impact.border = BORDER
    ws.cell(row=r, column=5, value="[calibrate shock to actual portfolio beta/sensitivity]").font = ITALIC_GRAY
    r += 1
ws.sheet_view.showGridLines = False

# ---------------- SENSITIVITY ----------------
ws = wb.create_sheet("Sensitivity")
set_col_widths(ws, [4, 20] + [12]*5)
ws["B2"] = "Sensitivity — VaR by Vol & Confidence"; ws["B2"].font = TITLE
ws["B4"] = "Daily vol \\ Confidence"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
confs = [0.90, 0.95, 0.975, 0.99, 0.995]
vols = [0.01, 0.015, 0.02, 0.025, 0.03]
for i, cf in enumerate(confs, start=3):
    c = ws.cell(row=4, column=i, value=cf); c.font = BOLD; c.fill = GRAY_FILL; c.number_format = PCT
for j, v in enumerate(vols, start=5):
    c = ws.cell(row=j, column=2, value=v); c.font = BOLD; c.fill = GRAY_FILL; c.number_format = PCT2
    for i in range(3, 8):
        col = get_column_letter(i)
        cell = ws.cell(row=j, column=i, value=f"='VaR'!$C$5*B{j}*NORMSINV({col}4)")
        cell.number_format = CUR
        cell.border = BORDER
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "RISK_template.xlsx"
wb.save(out_path)
print("saved", out_path)
