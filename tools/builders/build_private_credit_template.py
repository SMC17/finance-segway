"""Build the canonical private-credit underwriting archetype."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2, BPS,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    formula_cell, header, input_cell, section, set_widths, title, total_row,
)

SHEETS = ["Cover", "Assumptions", "Operating Case", "Debt Schedule", "Covenants",
          "Yield & Spread", "Recovery", "Sensitivity", "Checks", "Sources", "RefreshLog"]


def build(output: Path) -> None:
    wb = Workbook(); wb.remove(wb.active)
    for name in SHEETS: wb.create_sheet(name)
    add_cover(wb, "[BORROWER] — Private Credit Underwriting", [
        ("Borrower / issuer:", "[Name]"), ("Facility:", "Unitranche / TLB / revolver / notes"),
        ("Last refreshed:", "[date]"), ("Next test / maturity:", "[date]"),
        ("Refresh cadence:", "Weekly"), ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    ws = wb["Assumptions"]; title(ws, "B2:F2", "Underwriting Assumptions")
    header(ws, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Revenue (LTM)", 500.0, 500.0, "$mm", CUR), ("EBITDA margin", .20, .16, "%", PCT),
        ("Revenue growth", .04, -.05, "%", PCT), ("Maintenance capex / revenue", .03, .035, "%", PCT),
        ("Cash taxes / EBITDA", .15, .10, "%", PCT), ("Change in NWC / revenue change", .08, .10, "%", PCT),
        ("Opening cash", 25.0, 20.0, "$mm", CUR), ("Opening gross debt", 350.0, 350.0, "$mm", CUR),
        ("Base rate", .045, .055, "%", PCT), ("Cash spread", .055, .070, "%", PCT),
        ("OID / upfront fee", .02, .02, "%", PCT), ("PIK spread", 0.0, .02, "%", PCT),
        ("Mandatory amortization", .01, 0.0, "% opening debt", PCT),
        ("Cash sweep -- Tier 1 (leverage >= 4.0x)", .50, .25, "% excess cash", PCT),
        ("Minimum cash", 15.0, 20.0, "$mm", CUR), ("Maximum leverage", 6.0, 5.5, "x", MULT),
        ("Minimum interest coverage", 1.75, 1.50, "x", MULT), ("Minimum DSCR", 1.0, 1.0, "x", MULT),
        ("Recovery multiple", 5.0, 4.0, "x", MULT), ("EV haircut", .20, .35, "%", PCT),
        ("Maturity", 7.0, 7.0, "years", "0.0"),
    ]
    for row, (label, base, down, units, fmt) in enumerate(assumptions, 5):
        ws.cell(row, 2, label); input_cell(ws.cell(row, 3, base), fmt); input_cell(ws.cell(row, 4, down), fmt)
        formula_cell(ws.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})'), fmt, cross_sheet=True)
        ws.cell(row, 6, units)
    set_widths(ws, {"A":4,"B":40,"C":15,"D":15,"E":15,"F":28}); ws.freeze_panes="A5"

    ws = wb["Operating Case"]; title(ws, "B2:H2", "Operating Case & CFADS")
    header(ws, 4, 2, ["$mm", "LTM", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
    labels = ["Revenue","Growth","EBITDA","EBITDA margin","Cash taxes","Maintenance capex","Change in NWC",
              "CFADS before interest","Cash interest","CFADS after interest"]
    for r,l in enumerate(labels,5): ws.cell(r,2,l)
    formula_cell(ws.cell(5,3,"=Assumptions!E5"),CUR,cross_sheet=True)
    formula_cell(ws.cell(7,3,"=C5*Assumptions!E6"),CUR,cross_sheet=True)
    ws["C8"]="=IFERROR(C7/C5,0)"; ws["C8"].number_format=PCT
    for col in range(4,9):
        L,P=get_column_letter(col),get_column_letter(col-1)
        fs={5:f"={P}5*(1+Assumptions!$E$7)",6:f"=IFERROR({L}5/{P}5-1,0)",7:f"={L}5*Assumptions!$E$6",
            8:f"=IFERROR({L}7/{L}5,0)",9:f"={L}7*Assumptions!$E$9",10:f"={L}5*Assumptions!$E$8",
            11:f"=MAX(0,({L}5-{P}5)*Assumptions!$E$10)",12:f"={L}7-{L}9-{L}10-{L}11",
            13:f"='Debt Schedule'!{L}13",14:f"={L}12-{L}13"}
        for r,f in fs.items(): ws.cell(r,col,f).number_format=PCT if r in (6,8) else CUR
    for r in (12,14): total_row(ws,r,2,8,CUR)
    set_widths(ws,{"A":4,"B":36,**{get_column_letter(c):13 for c in range(3,9)}}); ws.freeze_panes="A5"

    ws=wb["Debt Schedule"]; title(ws,"B2:H2","Debt & Cash Schedule")
    header(ws,4,2,["$mm","Close","Year 1","Year 2","Year 3","Year 4","Year 5"])
    labels=["Beginning cash","CFADS after interest","Mandatory amortization","Cash before sweep","Cash sweep","Ending cash",
            "Beginning debt","Cash interest rate","Cash interest expense","PIK accrual","Ending debt","Average debt",
            "PIK interest expense","Total interest expense"]
    for r,l in enumerate(labels,5): ws.cell(r,2,l)
    formula_cell(ws.cell(5,3,"=Assumptions!E11"),CUR,cross_sheet=True); ws["C10"]="=C5"; ws["C10"].number_format=CUR
    formula_cell(ws.cell(11,3,"=Assumptions!E12"),CUR,cross_sheet=True); ws["C15"]="=C11"; ws["C16"]="=C11"
    formula_cell(ws.cell(12,3,"=Assumptions!E13+Assumptions!E14"),PCT,cross_sheet=True)
    for col in range(4,9):
        L,P=get_column_letter(col),get_column_letter(col-1)
        fs={5:f"={P}10",6:f"='Operating Case'!{L}14",7:f"=MIN({L}11,Assumptions!$E$12*Assumptions!$E$17)",
            8:f"={L}5+{L}6-{L}7",9:f"=MIN(MAX(0,{L}11-{L}7),MAX(0,{L}8-Assumptions!$E$19)*{L}27)",
            10:f"={L}8-{L}9",11:f"={P}15",12:"=Assumptions!$E$13+Assumptions!$E$14",13:f"={L}11*{L}12",
            14:f"=MAX(0,{L}11-{L}7-{L}9)*Assumptions!$E$16",15:f"=MAX(0,{L}11-{L}7-{L}9+{L}14)",
            16:f"=AVERAGE({L}11,{L}15)",17:f"={L}11*Assumptions!$E$16",18:f"={L}13+{L}17"}
        for r,f in fs.items(): ws.cell(r,col,f).number_format=PCT if r==12 else CUR
    for r in (10,15,18): total_row(ws,r,2,8,CUR)

    # Excess-cash-flow sweep: a leverage-based step-down grid, the actual
    # provision leveraged-loan credit agreements use, not a flat percentage.
    # Tier 1 is the scenario-toggled Assumptions!E18 rate (Base/Downside);
    # tiers 2-4 step down proportionally from it as beginning-of-period
    # leverage falls through 4.0x/3.0x/2.0x, so the whole grid rescales with
    # the same lever the rest of the model already exposes rather than
    # adding new hardcoded magic numbers. Leverage is computed on the
    # BEGINNING-of-period debt balance against the prior period's EBITDA
    # (never this period's own ending balance), the same non-circular
    # convention every debt schedule in this repo uses.
    section(ws,"B19:H19","Excess cash flow sweep -- leverage-based step-down grid")
    ws["D20"]="Breakpoint (x)"; ws["E20"]="Sweep %"
    grid=[(21,"Tier 1: leverage >= 4.0x",4.0,"=Assumptions!$E$18"),
          (22,"Tier 2: 3.0x <= leverage < 4.0x",3.0,"=$E$21*2/3"),
          (23,"Tier 3: 2.0x <= leverage < 3.0x",2.0,"=$E$21*1/3"),
          (24,"Tier 4: leverage < 2.0x",None,0.0)]
    for row,label,breakpoint_,pct_formula in grid:
        ws.cell(row,2,label)
        if breakpoint_ is not None:
            ws.cell(row,4,breakpoint_).number_format=MULT
        if isinstance(pct_formula,str) and pct_formula.startswith("="):
            formula_cell(ws.cell(row,5,pct_formula),PCT,cross_sheet=True)
        else:
            input_cell(ws.cell(row,5,pct_formula),PCT)
    ws["B26"]="Beginning-of-period gross leverage (x)"
    ws["B27"]="Applicable sweep % (step-down grid lookup)"
    for col in range(4,9):
        L,P=get_column_letter(col),get_column_letter(col-1)
        ws.cell(26,col,f"=IFERROR({L}11/'Operating Case'!{P}7,0)").number_format=MULT
        ws.cell(27,col,f"=IF({L}26>=$D$21,$E$21,IF({L}26>=$D$22,$E$22,IF({L}26>=$D$23,$E$23,$E$24)))").number_format=PCT
    set_widths(ws,{"A":4,"B":40,**{get_column_letter(c):13 for c in range(3,9)}}); ws.freeze_panes="A5"

    ws=wb["Covenants"]; title(ws,"B2:G2","Covenant Compliance")
    header(ws,4,2,["Metric","Year 1","Year 2","Year 3","Year 4","Year 5"])
    labels=["Gross leverage","Maximum leverage","Leverage headroom","Interest coverage","Minimum coverage","Coverage headroom",
            "DSCR","Minimum DSCR","DSCR headroom","Status"]
    for r,l in enumerate(labels,5): ws.cell(r,2,l)
    for col in range(3,8):
        L=get_column_letter(col); M=get_column_letter(col+1)
        fs={5:f"=IFERROR('Debt Schedule'!{M}15/'Operating Case'!{M}7,0)",6:"=Assumptions!$E$20",7:f"={L}6-{L}5",
            8:f"=IFERROR('Operating Case'!{M}7/'Debt Schedule'!{M}18,0)",9:"=Assumptions!$E$21",10:f"={L}8-{L}9",
            11:f"=IFERROR('Operating Case'!{M}12/('Debt Schedule'!{M}7+'Debt Schedule'!{M}13),0)",
            12:"=Assumptions!$E$22",13:f"={L}11-{L}12",14:f'=IF(AND({L}7>=0,{L}10>=0,{L}13>=0),"PASS","BREACH")'}
        for r,f in fs.items(): ws.cell(r,col,f).number_format=MULT if r<14 else "General"
    add_status_rules(ws,"C14:G14"); set_widths(ws,{"A":4,"B":38,**{get_column_letter(c):13 for c in range(3,8)}})

    ws=wb["Yield & Spread"]; title(ws,"B2:E2","Lender Yield & Spread")
    header(ws,4,2,["Metric","Value","Units","Method"])
    rows=[("Cash coupon","=Assumptions!E13+Assumptions!E14","%","Base rate + spread",PCT),
          ("PIK rate","=Assumptions!E16","%","PIK spread",PCT),("Issue price","=100*(1-Assumptions!E15)","per 100","100 less OID","0.00"),
          ("Maturity","=Assumptions!E25","years","Contractual","0.0"),
          ("Approx. cash YTM","=IFERROR((C5*100+(100-C7)/C8)/((100+C7)/2),0)","%","Bond approximation",PCT2),
          ("Approx. all-in yield","=C9+C6","%","Cash YTM + PIK",PCT2),
          ("Cash spread","=Assumptions!E14*10000","bps","Spread × 10,000",BPS),
          ("OID accretion / year","=IFERROR(Assumptions!E15/Assumptions!E25*10000,0)","bps","OID ÷ maturity",BPS)]
    for r,(l,f,u,n,fmt) in enumerate(rows,5): ws.cell(r,2,l); formula_cell(ws.cell(r,3,f),fmt,cross_sheet=True); ws.cell(r,4,u); ws.cell(r,5,n)
    set_widths(ws,{"A":4,"B":30,"C":16,"D":14,"E":46})

    ws=wb["Recovery"]; title(ws,"B2:D2","Recovery & Loss Analysis"); header(ws,4,2,["Bridge","Base","Downside"])
    labels=["Stressed EBITDA","Recovery multiple","Gross enterprise value","EV haircut","Net distributable value","Debt claim","Recovery value","Recovery rate","LGD"]
    base=["='Operating Case'!H7","=Assumptions!C23","=C5*C6","=C7*Assumptions!C24","=C7-C8","='Debt Schedule'!H15","=MIN(C9,C10)","=IFERROR(C11/C10,0)","=1-C12"]
    down=["='Operating Case'!H7*(1+Assumptions!D7)","=Assumptions!D23","=D5*D6","=D7*Assumptions!D24","=D7-D8","='Debt Schedule'!H15","=MIN(D9,D10)","=IFERROR(D11/D10,0)","=1-D12"]
    for r,l in enumerate(labels,5): ws.cell(r,2,l); ws.cell(r,3,base[r-5]); ws.cell(r,4,down[r-5])
    for r in (5,7,8,9,10,11):
        ws.cell(r,3).number_format=CUR; ws.cell(r,4).number_format=CUR
    for c in (3,4): ws.cell(6,c).number_format=MULT; ws.cell(12,c).number_format=PCT; ws.cell(13,c).number_format=PCT
    total_row(ws,11,2,4,CUR); set_widths(ws,{"A":4,"B":36,"C":18,"D":18})

    ws=wb["Sensitivity"]; title(ws,"B2:G2","Recovery Sensitivity"); header(ws,4,2,["EBITDA haircut / multiple",3.,4.,5.,6.,7.])
    for r,h in enumerate((0,.1,.2,.3,.4),5):
        ws.cell(r,2,h).number_format=PCT
        for c in range(3,8):
            L=get_column_letter(c); ws.cell(r,c,f"=IFERROR(MIN(1,MAX(0,('Operating Case'!$H$7*(1-$B{r})*{L}$4)/'Debt Schedule'!$H$15)),1)"); ws.cell(r,c).number_format=PCT
    ws.conditional_formatting.add("C5:G9",ColorScaleRule(start_type="min",start_color="FCE4D6",mid_type="percentile",mid_value=50,mid_color="FFF2CC",end_type="max",end_color="E2F0D9"))
    set_widths(ws,{"A":4,"B":34,**{get_column_letter(c):12 for c in range(3,8)}})

    ws=wb["Checks"]; title(ws,"B2:C2","Model Checks"); header(ws,4,2,["Check","Status"])
    checks=[("Debt nonnegative",'=IF(MIN(\'Debt Schedule\'!D15:H15)>=0,"PASS","FAIL")'),
            ("Cash above minimum",'=IF(MIN(\'Debt Schedule\'!D10:H10)>=Assumptions!E19,"PASS","REVIEW")'),
            ("Recovery bounded",'=IF(AND(MIN(Recovery!C12:D12)>=0,MAX(Recovery!C12:D12)<=1),"PASS","FAIL")'),
            ("No covenant breach",'=IF(COUNTIF(Covenants!C14:G14,"BREACH")=0,"PASS","REVIEW")'),
            ("Overall",'=IF(COUNTIF(C5:C8,"FAIL")+COUNTIF(C5:C8,"REVIEW")=0,"PASS","REVIEW")')]
    for r,(l,f) in enumerate(checks,5): ws.cell(r,2,l); ws.cell(r,3,f)
    add_status_rules(ws,"C5:C9"); set_widths(ws,{"A":4,"B":42,"C":18})

    add_sources(wb,[
        ("Financial statements","SEC filing, audited report, or data room","[period]","Reconcile EBITDA, cash, debt and CFADS"),
        ("Base rate / benchmark","Treasury, central bank, or contract reference","[date]","Freeze curve and floor assumptions"),
        ("Credit agreement / indenture","Public filing or reviewed transaction document","[date]","Covenants, baskets, amortization and collateral"),
        ("Recovery methodology","Documented valuation and waterfall assumptions","[date]","Bridge EV to claim-level recovery"),
    ])
    add_refresh_log(wb); finalize(wb,output)


if __name__ == "__main__":
    p=argparse.ArgumentParser(); p.add_argument("--output",type=Path,default=Path("CREDIT_template.xlsx")); a=p.parse_args(); build(a.output); print(f"saved {a.output}")
