"""Build an integrated sovereign and municipal public-finance credit model."""
from __future__ import annotations
import argparse, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (CUR, INTFMT, LIGHT_GREEN, LIGHT_RED, MULT, PCT, add_cover, add_refresh_log,
    add_sources, add_status_rules, finalize, formula_cell, header, input_cell, set_widths, title, total_row)  # noqa: E402

SHEETS=["Cover","Assumptions","Debt Sustainability","Revenue & Expenditure","Debt Service","Coverage","Scenarios","Sensitivity","Checks","Sources","RefreshLog"]

def build(output:Path)->None:
    wb=Workbook(); wb.remove(wb.active)
    for s in SHEETS: wb.create_sheet(s)
    add_cover(wb,"[ISSUER] — Public Finance Credit Model",[("Issuer:","[Sovereign / state / municipality / authority]"),("Instrument:","GO / revenue bond / sovereign note"),
        ("Last refreshed:","[date]"),("Next payment / issuance:","[date]"),("Refresh cadence:","Weekly"),("Active scenario:","Base"),("Units:","$ in millions unless noted")])
    ws=wb["Assumptions"]; title(ws,"B2:F2","Public Finance Assumptions"); header(ws,4,2,["Assumption","Base","Downside","Active","Units / note"])
    rows=[("Population / service base",1_000_000,950_000,INTFMT),("Recurring revenue",1200.,1080.,CUR),("Recurring expenditure",1050.,1090.,CUR),
          ("Capital expenditure",180.,220.,CUR),("Cash & unrestricted reserves",300.,220.,CUR),("Gross debt",2500.,2500.,CUR),
          ("Annual scheduled principal",125.,125.,CUR),("Pledged revenue",420.,360.,CUR),("Intergovernmental transfers",150.,120.,CUR),
          ("Revenue growth",.03,-.02,PCT),("Expenditure growth",.035,.05,PCT),("Effective interest rate (r)",.045,.06,PCT),
          ("Maximum debt / revenue",2.5,2.5,MULT),("Minimum DSCR",1.5,1.5,MULT),("Minimum days cash",90.,90.,"0.0"),
          ("Pension / OPEB contribution",90.,110.,CUR),("Nominal revenue-base growth (g)",.03,0.,PCT),
          ("Current primary balance",.01,-.03,PCT),("Debt-service reserve fund",50.,40.,CUR)]
    for r,(l,b,d,fmt) in enumerate(rows,5): ws.cell(r,2,l); input_cell(ws.cell(r,3,b),fmt); input_cell(ws.cell(r,4,d),fmt); formula_cell(ws.cell(r,5,f'=IF(Cover!$C$9="Downside",D{r},C{r})'),fmt,cross_sheet=True)
    set_widths(ws,{"A":4,"B":42,"C":15,"D":15,"E":15,"F":30}); ws.freeze_panes="A5"

    ws=wb["Debt Sustainability"]; title(ws,"B2:F2","Debt Sustainability Analysis"); header(ws,4,2,["Metric","Active","Downside","Units","Interpretation"])
    dsa=[("Debt ratio","=IFERROR(Assumptions!E10/Assumptions!E6,0)","=IFERROR(Assumptions!D10/Assumptions!D6,0)",MULT,"Debt / recurring revenue or GDP proxy"),
         ("Effective rate (r)","=Assumptions!E16","=Assumptions!D16",PCT,"Weighted effective debt cost"),("Nominal growth (g)","=Assumptions!E21","=Assumptions!D21",PCT,"Nominal revenue-base growth"),
         ("r - g","=C6-C7","=D6-D7",PCT,"Positive differential raises stabilization burden"),("Current primary balance","=Assumptions!E22","=Assumptions!D22",PCT,"Surplus positive"),
         ("Debt-stabilizing primary balance","=IFERROR((C6-C7)/(1+C7)*C5,0)","=IFERROR((D6-D7)/(1+D7)*D5,0)",PCT,"Required balance to stabilize debt ratio"),
         ("Primary-balance gap","=C9-C10","=D9-D10",PCT,"Positive is stronger than stabilizing"),
         ("One-year projected debt ratio","=IFERROR((1+C6)/(1+C7)*C5-C9,0)","=IFERROR((1+D6)/(1+D7)*D5-D9,0)",MULT,"Standard debt-dynamics identity"),
         ("Trajectory",'=IF(C12<=C5,"STABILIZING / FALLING","RISING")','=IF(D12<=D5,"STABILIZING / FALLING","RISING")',None,"Screen excludes stock-flow adjustments")]
    for r,(l,a,d,fmt,n) in enumerate(dsa,5):
        ws.cell(r,2,l); ws.cell(r,3,a); ws.cell(r,4,d); ws.cell(r,5,fmt or "status"); ws.cell(r,6,n)
        if fmt:
            ws.cell(r,3).number_format=fmt; ws.cell(r,4).number_format=fmt
    ws.conditional_formatting.add("C13:D13",FormulaRule(formula=['C13="STABILIZING / FALLING"'],fill=PatternFill("solid",fgColor=LIGHT_GREEN)))
    ws.conditional_formatting.add("C13:D13",FormulaRule(formula=['C13="RISING"'],fill=PatternFill("solid",fgColor=LIGHT_RED)))
    set_widths(ws,{"A":4,"B":40,"C":18,"D":18,"E":16,"F":48})

    ws=wb["Revenue & Expenditure"]; title(ws,"B2:H2","Operating Forecast"); header(ws,4,2,["$mm","Current","Year 1","Year 2","Year 3","Year 4","Year 5"])
    labels=["Recurring revenue","Revenue growth","Recurring expenditure","Expenditure growth","Operating surplus","Operating margin","Pension / OPEB contribution","Net operating result","Capital expenditure","FCF before debt service"]
    for r,l in enumerate(labels,5): ws.cell(r,2,l)
    formula_cell(ws.cell(5,3,"=Assumptions!E6"),CUR,cross_sheet=True); formula_cell(ws.cell(7,3,"=Assumptions!E7"),CUR,cross_sheet=True)
    ws["C9"]="=C5-C7"; ws["C9"].number_format=CUR; ws["C10"]="=IFERROR(C9/C5,0)"; ws["C10"].number_format=PCT
    formula_cell(ws.cell(11,3,"=Assumptions!E20"),CUR,cross_sheet=True); ws["C12"]="=C9-C11"; ws["C12"].number_format=CUR
    formula_cell(ws.cell(13,3,"=Assumptions!E8"),CUR,cross_sheet=True); ws["C14"]="=C12-C13"; ws["C14"].number_format=CUR
    for c in range(4,9):
        L,P=get_column_letter(c),get_column_letter(c-1)
        fs={5:f"={P}5*(1+Assumptions!$E$14)",6:f"=IFERROR({L}5/{P}5-1,0)",7:f"={P}7*(1+Assumptions!$E$15)",8:f"=IFERROR({L}7/{P}7-1,0)",
            9:f"={L}5-{L}7",10:f"=IFERROR({L}9/{L}5,0)",11:"=Assumptions!$E$20",12:f"={L}9-{L}11",13:"=Assumptions!$E$8",14:f"={L}12-{L}13"}
        for r,f in fs.items(): ws.cell(r,c,f).number_format=PCT if r in (6,8,10) else CUR
    for r in (9,12,14): total_row(ws,r,2,8,CUR)
    set_widths(ws,{"A":4,"B":40,**{get_column_letter(c):13 for c in range(3,9)}}); ws.freeze_panes="A5"

    ws=wb["Debt Service"]; title(ws,"B2:G2","Debt Service, Borrowing & Reserves"); header(ws,4,2,["$mm","Year 1","Year 2","Year 3","Year 4","Year 5"])
    labels=["Beginning debt","Scheduled principal","Cash interest","Total debt service","FCF before debt service","New borrowing / funding gap","Ending debt","Beginning reserves","Ending reserves","Pledged revenue","Revenue-bond DSCR"]
    for r,l in enumerate(labels,5): ws.cell(r,2,l)
    for c in range(3,8):
        L,P=get_column_letter(c),get_column_letter(c-1); O=get_column_letter(c+1); idx=c-2
        fs={5:"=Assumptions!E10" if c==3 else f"={P}11",6:f"=MIN({L}5,Assumptions!$E$11)",7:f"={L}5*Assumptions!$E$16",8:f"={L}6+{L}7",
            9:f"='Revenue & Expenditure'!{O}14",10:f"=MAX(0,{L}8-{L}9)",11:f"={L}5-{L}6+{L}10",12:"=Assumptions!E9" if c==3 else f"={P}13",
            13:f"=MAX(0,{L}12+{L}9-{L}8+{L}10)",14:f"=Assumptions!$E$12*(1+Assumptions!$E$14)^{idx}",15:f"=IFERROR({L}14/{L}8,0)"}
        for r,f in fs.items(): ws.cell(r,c,f).number_format=MULT if r==15 else CUR
    for r in (8,11,13): total_row(ws,r,2,7,CUR)
    set_widths(ws,{"A":4,"B":42,**{get_column_letter(c):13 for c in range(3,8)}}); ws.freeze_panes="A5"

    ws=wb["Coverage"]; title(ws,"B2:G2","Coverage, Liquidity & Debt Burden"); header(ws,4,2,["Metric","Year 1","Year 2","Year 3","Year 4","Year 5"])
    labels=["Debt / recurring revenue","Maximum debt / revenue","Debt burden headroom","Revenue-bond DSCR","Minimum DSCR","DSCR headroom","Days cash on hand","Minimum days cash","Liquidity headroom","Pension burden / revenue","Status"]
    for r,l in enumerate(labels,5): ws.cell(r,2,l)
    for c in range(3,8):
        L=get_column_letter(c); O=get_column_letter(c+1)
        fs={5:f"=IFERROR('Debt Service'!{L}11/'Revenue & Expenditure'!{O}5,0)",6:"=Assumptions!E17",7:f"={L}6-{L}5",8:f"='Debt Service'!{L}15",9:"=Assumptions!E18",10:f"={L}8-{L}9",
            11:f"=IFERROR('Debt Service'!{L}13/('Revenue & Expenditure'!{O}7/365),0)",12:"=Assumptions!E19",13:f"={L}11-{L}12",14:f"=IFERROR('Revenue & Expenditure'!{O}11/'Revenue & Expenditure'!{O}5,0)",
            15:f'=IF(AND({L}7>=0,{L}10>=0,{L}13>=0),"PASS","BREACH")'}
        for r,f in fs.items(): ws.cell(r,c,f).number_format=MULT if r<=10 else ("0.0" if r<=13 else (PCT if r==14 else "General"))
    add_status_rules(ws,"C15:G15"); set_widths(ws,{"A":4,"B":42,**{get_column_letter(c):13 for c in range(3,8)}})

    ws=wb["Scenarios"]; title(ws,"B2:E2","Scenario Summary"); header(ws,4,2,["Metric","Base","Downside","Threshold / interpretation"])
    scen=[("Operating surplus","=Assumptions!C6-Assumptions!C7-Assumptions!C20","=Assumptions!D6-Assumptions!D7-Assumptions!D20",CUR),
          ("Debt / revenue","=IFERROR(Assumptions!C10/Assumptions!C6,0)","=IFERROR(Assumptions!D10/Assumptions!D6,0)",MULT),
          ("Revenue-bond DSCR","=IFERROR(Assumptions!C12/(Assumptions!C11+Assumptions!C10*Assumptions!C16),0)","=IFERROR(Assumptions!D12/(Assumptions!D11+Assumptions!D10*Assumptions!D16),0)",MULT),
          ("Days cash","=IFERROR(Assumptions!C9/(Assumptions!C7/365),0)","=IFERROR(Assumptions!D9/(Assumptions!D7/365),0)","0.0"),
          ("Debt-stabilizing primary balance","=IFERROR((Assumptions!C16-Assumptions!C21)/(1+Assumptions!C21)*(Assumptions!C10/Assumptions!C6),0)","=IFERROR((Assumptions!D16-Assumptions!D21)/(1+Assumptions!D21)*(Assumptions!D10/Assumptions!D6),0)",PCT),
          ("Overall",'=IF(AND(C5>0,C6<=Assumptions!C17,C7>=Assumptions!C18,C8>=Assumptions!C19),"PASS","REVIEW")','=IF(AND(D5>0,D6<=Assumptions!D17,D7>=Assumptions!D18,D8>=Assumptions!D19),"PASS","REVIEW")',None)]
    for r,(l,b,d,fmt) in enumerate(scen,5): ws.cell(r,2,l); ws.cell(r,3,b); ws.cell(r,4,d); ws.cell(r,5,"Decision screen")
    for r,(_,_,_,fmt) in enumerate(scen,5):
        if fmt: ws.cell(r,3).number_format=fmt; ws.cell(r,4).number_format=fmt
    add_status_rules(ws,"C10:D10"); set_widths(ws,{"A":4,"B":40,"C":18,"D":18,"E":32})

    ws=wb["Sensitivity"]; title(ws,"B2:G2","Debt Ratio Sensitivity — One Year"); header(ws,4,2,["Primary balance / r-g",-.02,0,.02,.04,.06])
    for c in range(3,8): ws.cell(4,c).number_format=PCT
    for r,pb in enumerate((-.05,-.025,0,.025,.05),5):
        ws.cell(r,2,pb).number_format=PCT
        for c in range(3,8): L=get_column_letter(c); ws.cell(r,c,f"='Debt Sustainability'!$C$5*(1+{L}$4)-$B{r}").number_format=MULT
    ws.conditional_formatting.add("C5:G9",ColorScaleRule(start_type="min",start_color="E2F0D9",mid_type="percentile",mid_value=50,mid_color="FFF2CC",end_type="max",end_color="FCE4D6"))
    set_widths(ws,{"A":4,"B":34,**{get_column_letter(c):12 for c in range(3,8)}})

    ws=wb["Checks"]; title(ws,"B2:C2","Model Checks"); header(ws,4,2,["Check","Status"])
    checks=[("Debt nonnegative",'=IF(MIN(\'Debt Service\'!C11:G11)>=0,"PASS","FAIL")'),("Reserves nonnegative",'=IF(MIN(\'Debt Service\'!C13:G13)>=0,"PASS","FAIL")'),
            ("Coverage finite",'=IF(AND(MAX(Coverage!C5:G14)<1000,MIN(Coverage!C5:G14)>-1000),"PASS","FAIL")'),
            ("No coverage breach",'=IF(COUNTIF(Coverage!C15:G15,"BREACH")=0,"PASS","REVIEW")'),
            ("Debt dynamics bounded",'=IF(AND(\'Debt Sustainability\'!C12>=0,\'Debt Sustainability\'!C12<10),"PASS","REVIEW")'),
            ("Overall",'=IF(COUNTIF(C5:C9,"FAIL")+COUNTIF(C5:C9,"REVIEW")=0,"PASS","REVIEW")')]
    for r,(l,f) in enumerate(checks,5): ws.cell(r,2,l); ws.cell(r,3,f)
    add_status_rules(ws,"C5:C10"); set_widths(ws,{"A":4,"B":44,"C":18})
    add_sources(wb,[("Official statements / disclosures","EMMA, finance ministry, issuer or audited report","[period]","Revenue, expenditure, debt and reserves"),
                    ("Accounting standards","GASB, IPSAS, or applicable national standards","[date]","Map reporting consistently"),
                    ("Economic and demographic data","Official statistics, FRED, or multilateral dataset","[period]","Freeze series and release vintage"),
                    ("Benchmark rates","Treasury or sovereign curve","[date]","Refinancing and relative-value assumptions"),
                    ("Debt sustainability method","IMF / World Bank public methodology","[date]","Document debt-dynamics and stress choices")])
    add_refresh_log(wb); finalize(wb,output)

if __name__=="__main__":
    p=argparse.ArgumentParser(); p.add_argument("--output",type=Path,default=Path("PUBLIC_FINANCE_template.xlsx")); a=p.parse_args(); build(a.output); print(f"saved {a.output}")
