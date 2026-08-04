import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[DEAL] — Structured Finance / Securitization Model", [
    ("Asset class:", "RMBS / ABS / CLO"),
    ("Closing date:", "[date]"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly (monthly at distribution date)"),
])

# ---------------- COLLATERAL POOL ----------------
ws = wb.create_sheet("Collateral Pool")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Collateral Pool Summary"; ws["B2"].font = TITLE
inputs = [
    ("Aggregate pool balance ($)", 0, CUR),
    ("Weighted avg coupon (WAC, %)", 0.06, PCT),
    ("Weighted avg maturity (WAM, months)", 360, NUM),
    ("Conditional prepayment rate assumption (CPR, %)", 0.08, PCT),
    ("Cumulative default rate assumption (%)", 0.03, PCT),
    ("Recovery rate on defaults (%)", 0.50, PCT),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B12"] = "Single Monthly Mortality (SMM) from CPR"
ws["C12"] = "=1-(1-C8)^(1/12)"; ws["C12"].number_format = PCT2
ws["D12"] = "SMM = 1-(1-CPR)^(1/12) — standard CPR-to-SMM conversion"
ws["D12"].font = ITALIC_GRAY
ws["C12"].border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- TRANCHE WATERFALL ----------------
ws = wb.create_sheet("Waterfall")
set_col_widths(ws, [4, 18, 14, 14, 14, 16, 40])
ws["B2"] = "Tranche Structure & Credit Enhancement"; ws["B2"].font = TITLE
headers = ["", "Tranche", "Face ($)", "% of pool", "Coupon", "Credit enhancement %", "Notes"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers)-1)
tranches = ["Class A (senior)", "Class B (mezz)", "Class C (mezz)", "Class D (subordinate)", "Equity / residual"]
r = 5
for t in tranches:
    ws.cell(row=r, column=2, value=t).font = BLACK
    c_face = ws.cell(row=r, column=3, value=0); c_face.font = BLUE; c_face.number_format = CUR; c_face.border = BORDER
    c_cpn = ws.cell(row=r, column=5, value=0); c_cpn.font = BLUE; c_cpn.number_format = PCT; c_cpn.border = BORDER
    r += 1
total_row = r
ws.cell(row=total_row, column=2, value="Total").font = BOLD
ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})").font = BOLD
ws.cell(row=total_row, column=3).number_format = CUR
for rr in range(5, total_row):
    ws.cell(row=rr, column=4, value=f"=IFERROR(C{rr}/$C${total_row},\"-\")").number_format = PCT
    # CE = sum of face value of all tranches junior to this one, as % of pool
    ws.cell(row=rr, column=6, value=f"=IFERROR(SUM(C{rr+1}:C{total_row-1})/$C${total_row},\"-\")").number_format = PCT
ws["B" + str(total_row+2)] = "CE% for a tranche = subordinate tranche cushion that absorbs losses before it does"
ws["B" + str(total_row+2)].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

# ---------------- WAL & CPR ----------------
ws = wb.create_sheet("WAL & CPR")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Weighted Average Life"; ws["B2"].font = TITLE
ws["B4"] = "Principal paydown schedule (months 1-6 sample, extend as needed)"
ws["B4"].font = ITALIC_GRAY
headers2 = ["Month", "Beg. balance", "Scheduled prin.", "Prepayment (SMM x bal)", "End balance"]
for i, h in enumerate(headers2, start=2):
    ws.cell(row=5, column=i, value=h)
style_header_row(ws, 5, 5, start_col=2)
r = 6
for m in range(1, 7):
    ws.cell(row=r, column=2, value=m)
    if m == 1:
        ws.cell(row=r, column=3, value="='Collateral Pool'!C5")
    else:
        ws.cell(row=r, column=3, value=f"=F{r-1}")
    ws.cell(row=r, column=4, value=0).font = BLUE  # scheduled principal, user fills or links amort schedule
    ws.cell(row=r, column=5, value=f"=C{r}*'Collateral Pool'!$C$12")
    ws.cell(row=r, column=6, value=f"=C{r}-D{r}-E{r}")
    for c in range(3, 7):
        ws.cell(row=r, column=c).number_format = CUR
        ws.cell(row=r, column=c).border = BORDER
    r += 1
ws["B13"] = "Total principal paid (sample months)"
ws["C13"] = "=SUM(D6:E11)"; ws["C13"].number_format = CUR; ws["C13"].border = BORDER
ws["B14"] = "WAL (years) = sum(month x principal paid that month) / total principal / 12"
ws["C14"] = "=IFERROR(SUMPRODUCT(B6:B11,D6:D11+E6:E11)/C13/12,\"-\")"
ws["C14"].font = BOLD; ws["C14"].number_format = '0.00'; ws["C14"].border = BORDER
ws["D14"] = ("Computed on this 6-month sample schedule only — extend rows 6-11 to the deal's actual "
             "WAM for a real WAL, not just a demo of the mechanics")
ws["D14"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "SECURITIZATION_template.xlsx"
wb.save(out_path)
print("saved", out_path)
