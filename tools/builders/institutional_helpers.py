"""Shared controls and styling for institutional finance workbook builders."""
from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "1F4E78"
DARK = "111827"
WHITE = "FFFFFF"
BLUE = "0000FF"
GREEN = "008000"
YELLOW = "FFF2CC"
LIGHT_GREEN = "E2F0D9"
LIGHT_RED = "FCE4D6"
LIGHT_YELLOW = "FFF2CC"
MED_GRAY = "BFBFBF"

THIN = Side(style="thin", color=MED_GRAY)
TOP = Side(style="thin", color=DARK)
BORDER_BOTTOM = Border(bottom=THIN)
BORDER_TOP = Border(top=TOP)

CUR = '$#,##0.0;[Red]($#,##0.0);-'
CUR0 = '$#,##0;[Red]($#,##0);-'
PCT = '0.0%;[Red](0.0%);-'
PCT2 = '0.00%;[Red](0.00%);-'
MULT = '0.0x;[Red](0.0x);-'
NUM = '#,##0.0;[Red](#,##0.0);-'
INTFMT = '#,##0;[Red](#,##0);-'
BPS = '0 "bps"'


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def title(ws, cell_range: str, text: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cell.row].height = 28


def header(ws, row: int, start_col: int, values: Sequence[object]) -> None:
    for offset, value in enumerate(values):
        cell = ws.cell(row=row, column=start_col + offset, value=value)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_BOTTOM


def section(ws, cell_range: str, text: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)


def input_cell(cell, number_format: str | None = None) -> None:
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.font = Font(name="Arial", size=10, color=BLUE)
    cell.border = BORDER_BOTTOM
    if number_format:
        cell.number_format = number_format


def formula_cell(cell, number_format: str | None = None, *, cross_sheet: bool = False, bold: bool = False) -> None:
    cell.font = Font(name="Arial", size=10, color=GREEN if cross_sheet else DARK, bold=bold)
    if number_format:
        cell.number_format = number_format


def total_row(ws, row: int, start_col: int, end_col: int, number_format: str | None = None) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = BORDER_TOP
        cell.font = Font(name="Arial", size=10, bold=True, color=DARK)
        if number_format and col > start_col:
            cell.number_format = number_format


def add_status_rules(ws, cell_range: str) -> None:
    first = cell_range.split(":")[0]
    for label, color in (("PASS", LIGHT_GREEN), ("FAIL", LIGHT_RED), ("REVIEW", LIGHT_YELLOW), ("BREACH", LIGHT_RED)):
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(formula=[f'{first}="{label}"'], fill=PatternFill("solid", fgColor=color)),
        )


def add_cover(wb: Workbook, title_text: str, fields: Sequence[tuple[str, object]], *, scenario_cell: str | None = "C9"):
    ws = wb["Cover"] if "Cover" in wb.sheetnames else wb.create_sheet("Cover")
    title(ws, "B2:F2", title_text)
    for row, (label, value) in enumerate(fields, start=4):
        ws.cell(row=row, column=2, value=label).font = Font(name="Arial", size=10, bold=True, color=DARK)
        input_cell(ws.cell(row=row, column=3, value=value))
    if scenario_cell:
        validation = DataValidation(type="list", formula1='"Base,Downside"', allow_blank=False)
        ws.add_data_validation(validation)
        validation.add(ws[scenario_cell])
    section(ws, "B13:F13", "Modeling conventions")
    legend = (
        ("Blue text / yellow fill", "Hardcoded input or selected scenario"),
        ("Black text", "Formula / same-sheet calculation"),
        ("Green text", "Cross-sheet formula / link"),
        ("Checks", "PASS / REVIEW / FAIL must remain visible"),
    )
    for row, (label, meaning) in enumerate(legend, start=14):
        ws.cell(row=row, column=2, value=label).font = Font(name="Arial", size=10, bold=True, color=DARK)
        ws.cell(row=row, column=3, value=meaning).font = Font(name="Arial", size=10, color=DARK)
    ws["B14"].font = Font(name="Arial", size=10, bold=True, color=BLUE)
    ws["B14"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["B16"].font = Font(name="Arial", size=10, bold=True, color=GREEN)
    set_widths(ws, {"A": 4, "B": 36, "C": 46, "D": 12, "E": 12, "F": 12})
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    return ws


def add_sources(wb: Workbook, rows: Sequence[tuple[str, str, str, str]]) -> None:
    ws = wb["Sources"] if "Sources" in wb.sheetnames else wb.create_sheet("Sources")
    title(ws, "B2:E2", "Source Register")
    header(ws, 4, 2, ["Input / dataset", "Source URL or document", "As-of", "Notes / transformation"])
    for row, values in enumerate(rows, start=5):
        for col, value in enumerate(values, start=2):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(name="Arial", size=10, color=BLUE)
            cell.border = BORDER_BOTTOM
    set_widths(ws, {"A": 4, "B": 34, "C": 58, "D": 16, "E": 48})
    ws.freeze_panes = "A5"


def add_refresh_log(wb: Workbook) -> None:
    ws = wb["RefreshLog"] if "RefreshLog" in wb.sheetnames else wb.create_sheet("RefreshLog")
    title(ws, "B2:G2", "Refresh Log")
    header(ws, 4, 2, ["Date", "Trigger", "Source snapshot", "What changed", "Reviewer / challenge", "Next check"])
    for row in range(5, 30):
        for col in range(2, 8):
            ws.cell(row=row, column=col).border = BORDER_BOTTOM
    set_widths(ws, {"A": 4, "B": 14, "C": 22, "D": 28, "E": 38, "F": 34, "G": 18})
    ws.freeze_panes = "A5"


def finalize(wb: Workbook, output: Path) -> None:
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
