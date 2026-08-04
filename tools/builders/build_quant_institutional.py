"""Build the canonical quantitative and systematic research decision model."""
from __future__ import annotations

import argparse
import math
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    formula_cell, header, input_cell, set_widths, title, total_row,
)

PERIODS = 60

SHEETS = [
    "Cover", "Assumptions", "Backtest", "Performance", "Walk Forward",
    "Costs & Capacity", "Risk", "Stress", "Checks", "Sources", "RefreshLog",
]


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[STRATEGY] — Systematic Research Model", [
        ("Strategy / universe:", "[Name and investable universe]"),
        ("Data vintage / point-in-time policy:", "[policy]"),
        ("Last refreshed:", "[date]"),
        ("Next rebalance / review:", "[date]"),
        ("Refresh cadence:", "Daily or per rebalance"),
        ("Active scenario:", "Base"),
        ("Units:", "Periodic returns unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Backtest, Cost, and Capacity Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Periods per year", 12.0, 12.0, "periods", "0"),
        ("Risk-free return per period", 0.002, 0.002, "%", PCT2),
        ("Linear transaction cost", 8.0, 15.0, "bps per unit turnover", "0.0"),
        ("Impact coefficient", 0.002, 0.008, "return units", "0.0000"),
        ("Capital deployed", 50.0, 150.0, "$mm", CUR),
        ("Average daily volume", 500.0, 250.0, "$mm", CUR),
        ("Maximum participation", 0.10, 0.08, "% ADV", PCT),
        ("Training periods", 24.0, 24.0, "periods", "0"),
        ("Testing periods", 6.0, 6.0, "periods", "0"),
        ("Walk-forward step", 6.0, 6.0, "periods", "0"),
        ("VaR confidence", 0.95, 0.99, "%", PCT),
        ("Target volatility", 0.12, 0.10, "% annual", PCT),
        ("Maximum drawdown limit", -0.20, -0.15, "%", PCT),
        ("Signal lag", 1.0, 1.0, "periods", "0"),
        ("Data publication lag enforced", 1.0, 1.0, "1=yes", "0"),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        formula_cell(sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})'), number_format, cross_sheet=True)
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 42, "C": 15, "D": 15, "E": 15, "F": 34})
    sheet.freeze_panes = "A5"

    sheet = workbook["Backtest"]
    title(sheet, "B2:L2", "Point-in-Time Backtest")
    header(sheet, 4, 2, [
        "Period", "Gross return", "Turnover", "Signal available?", "Linear cost",
        "Impact cost", "Net return", "Cumulative wealth", "Running peak",
        "Drawdown", "Participation",
    ])
    for index in range(PERIODS):
        row = 5 + index
        period = index + 1
        sheet.cell(row, 2, period)
        gross = 0.012 * math.sin(period / 3.0) + 0.006 * math.cos(period / 7.0) + (0.002 if period % 5 else -0.008)
        turnover = 0.25 + 0.15 * abs(math.sin(period / 5.0))
        input_cell(sheet.cell(row, 3, gross), PCT2)
        input_cell(sheet.cell(row, 4, turnover), PCT2)
        sheet.cell(row, 5, '=IF(AND(Assumptions!$E$18>=1,Assumptions!$E$19=1),1,0)')
        sheet.cell(row, 6, f"=ABS(D{row})*Assumptions!$E$7/10000")
        sheet.cell(row, 7, f"=Assumptions!$E$8*(ABS(D{row})*Assumptions!$E$9/Assumptions!$E$10)^1.5")
        sheet.cell(row, 8, f"=IF(E{row}=1,C{row}-F{row}-G{row},0)")
        if row == 5:
            sheet.cell(row, 9, f"=1+H{row}")
            sheet.cell(row, 10, f"=I{row}")
        else:
            sheet.cell(row, 9, f"=I{row-1}*(1+H{row})")
            sheet.cell(row, 10, f"=MAX(J{row-1},I{row})")
        sheet.cell(row, 11, f"=IFERROR(I{row}/J{row}-1,0)")
        sheet.cell(row, 12, f"=ABS(D{row})*Assumptions!$E$9/Assumptions!$E$10")
        for column in range(3, 13):
            sheet.cell(row, column).number_format = PCT2 if column not in (9, 10) else "0.0000x"
    set_widths(sheet, {"A": 4, "B": 12, "C": 14, "D": 14, "E": 16, "F": 14, "G": 14, "H": 14, "I": 16, "J": 16, "K": 14, "L": 14})
    sheet.freeze_panes = "B5"

    sheet = workbook["Performance"]
    title(sheet, "B2:E2", "Gross, Net, and Risk-Adjusted Performance")
    header(sheet, 4, 2, ["Metric", "Gross", "Net", "Interpretation"])
    last_row = 4 + PERIODS
    metrics = [
        ("Annualized arithmetic return", f"=AVERAGE(Backtest!C5:C{last_row})*Assumptions!E5", f"=AVERAGE(Backtest!H5:H{last_row})*Assumptions!E5", "mean periodic return annualized"),
        ("Annualized volatility", f"=STDEV.P(Backtest!C5:C{last_row})*SQRT(Assumptions!E5)", f"=STDEV.P(Backtest!H5:H{last_row})*SQRT(Assumptions!E5)", "population volatility"),
        ("Sharpe ratio", f"=IFERROR((AVERAGE(Backtest!C5:C{last_row})-Assumptions!E6)/STDEV.P(Backtest!C5:C{last_row})*SQRT(Assumptions!E5),0)", f"=IFERROR((AVERAGE(Backtest!H5:H{last_row})-Assumptions!E6)/STDEV.P(Backtest!H5:H{last_row})*SQRT(Assumptions!E5),0)", "excess return / volatility"),
        ("Sortino ratio", f"=IFERROR((AVERAGE(Backtest!C5:C{last_row})-Assumptions!E6)/SQRT(SUMPRODUCT((Backtest!C5:C{last_row}<Assumptions!E6)*(Backtest!C5:C{last_row}-Assumptions!E6)^2)/COUNT(Backtest!C5:C{last_row}))*SQRT(Assumptions!E5),0)", f"=IFERROR((AVERAGE(Backtest!H5:H{last_row})-Assumptions!E6)/SQRT(SUMPRODUCT((Backtest!H5:H{last_row}<Assumptions!E6)*(Backtest!H5:H{last_row}-Assumptions!E6)^2)/COUNT(Backtest!H5:H{last_row}))*SQRT(Assumptions!E5),0)", "downside-risk adjusted"),
        ("Maximum drawdown", "=MIN(Backtest!K5:K64)", "=MIN(Backtest!K5:K64)", "peak-to-trough"),
        ("Hit rate", f'=COUNTIF(Backtest!C5:C{last_row},">0")/COUNT(Backtest!C5:C{last_row})', f'=COUNTIF(Backtest!H5:H{last_row},">0")/COUNT(Backtest!H5:H{last_row})', "positive periods"),
        ("Ending wealth multiple", f"=PRODUCT(1+Backtest!C5:C{last_row})", f"=Backtest!I{last_row}", "growth of one unit"),
        ("Annual cost drag", "=C5-D5", "=C5-D5", "gross less net annualized return"),
    ]
    for row, (label, gross_formula, net_formula, note) in enumerate(metrics, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, gross_formula)
        sheet.cell(row, 4, net_formula)
        sheet.cell(row, 5, note)
        if row in (7, 8):
            sheet.cell(row, 3).number_format = MULT
            sheet.cell(row, 4).number_format = MULT
        elif row == 11:
            sheet.cell(row, 3).number_format = "0.000x"
            sheet.cell(row, 4).number_format = "0.000x"
        else:
            sheet.cell(row, 3).number_format = PCT2
            sheet.cell(row, 4).number_format = PCT2
    set_widths(sheet, {"A": 4, "B": 34, "C": 18, "D": 18, "E": 44})

    sheet = workbook["Walk Forward"]
    title(sheet, "B2:J2", "Walk-Forward Validation Folds")
    header(sheet, 4, 2, [
        "Fold", "Train start", "Train end", "Test start", "Test end",
        "Train Sharpe", "Test Sharpe", "Test net return", "Degradation",
    ])
    for fold in range(1, 7):
        row = 4 + fold
        train_start = 1 + (fold - 1) * 6
        train_end = train_start + 23
        test_start = train_end + 1
        test_end = test_start + 5
        sheet.cell(row, 2, fold)
        sheet.cell(row, 3, train_start)
        sheet.cell(row, 4, train_end)
        sheet.cell(row, 5, test_start)
        sheet.cell(row, 6, test_end)
        train_range = f"INDEX(Backtest!$H$5:$H$64,C{row}):INDEX(Backtest!$H$5:$H$64,D{row})"
        test_range = f"INDEX(Backtest!$H$5:$H$64,E{row}):INDEX(Backtest!$H$5:$H$64,F{row})"
        sheet.cell(row, 7, f"=IFERROR(AVERAGE({train_range})/STDEV.P({train_range})*SQRT(Assumptions!$E$5),0)")
        sheet.cell(row, 8, f"=IFERROR(AVERAGE({test_range})/STDEV.P({test_range})*SQRT(Assumptions!$E$5),0)")
        sheet.cell(row, 9, f"=PRODUCT(1+{test_range})-1")
        sheet.cell(row, 10, f"=H{row}-G{row}")
        sheet.cell(row, 7).number_format = MULT
        sheet.cell(row, 8).number_format = MULT
        sheet.cell(row, 9).number_format = PCT2
        sheet.cell(row, 10).number_format = MULT
    set_widths(sheet, {"A": 4, "B": 10, "C": 13, "D": 13, "E": 13, "F": 13, "G": 16, "H": 16, "I": 16, "J": 16})

    sheet = workbook["Costs & Capacity"]
    title(sheet, "B2:G2", "Cost and Capacity Sensitivity")
    header(sheet, 4, 2, ["Capital / Turnover", 0.10, 0.25, 0.50, 0.75, 1.00])
    for column in range(3, 8):
        sheet.cell(4, column).number_format = PCT2
    for row, capital in enumerate((10, 25, 50, 100, 250), start=5):
        sheet.cell(row, 2, capital)
        sheet.cell(row, 2).number_format = CUR
        for column in range(3, 8):
            letter = get_column_letter(column)
            sheet.cell(row, column, f"={letter}$4*Assumptions!$E$7/10000+Assumptions!$E$8*({letter}$4*$B{row}/Assumptions!$E$10)^1.5")
            sheet.cell(row, column).number_format = PCT2
    sheet.conditional_formatting.add(
        "C5:G9",
        ColorScaleRule(
            start_type="min", start_color="E2F0D9",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="FCE4D6",
        ),
    )
    sheet["B12"] = "Current maximum participation"
    sheet["C12"] = "=MAX(Backtest!L5:L64)"
    sheet["C12"].number_format = PCT2
    sheet["B13"] = "Participation limit"
    sheet["C13"] = "=Assumptions!E11"
    sheet["C13"].number_format = PCT2
    sheet["B14"] = "Capacity status"
    sheet["C14"] = '=IF(C12<=C13,"PASS","BREACH")'
    add_status_rules(sheet, "C14")
    set_widths(sheet, {"A": 4, "B": 28, **{get_column_letter(column): 14 for column in range(3, 8)}})

    sheet = workbook["Risk"]
    title(sheet, "B2:E2", "Distribution and Tail Risk")
    header(sheet, 4, 2, ["Metric", "Value", "Units", "Method"])
    risk_metrics = [
        ("Historical VaR", f"=-PERCENTILE.INC(Backtest!H5:H{last_row},1-Assumptions!E15)", "%", "empirical left-tail quantile", PCT2),
        ("Historical expected shortfall", f'=IFERROR(-AVERAGEIF(Backtest!H5:H{last_row},"<="&-C5,Backtest!H5:H{last_row}),0)', "%", "mean return beyond VaR", PCT2),
        ("Parametric VaR", f"=-(AVERAGE(Backtest!H5:H{last_row})+NORM.S.INV(1-Assumptions!E15)*STDEV.P(Backtest!H5:H{last_row}))", "%", "normal approximation", PCT2),
        ("Maximum drawdown", "=Performance!D9", "%", "realized peak-to-trough", PCT2),
        ("Drawdown limit", "=ABS(Assumptions!E17)", "%", "approved risk limit", PCT2),
        ("Drawdown status", '=IF(C8<=C9,"PASS","BREACH")', "status", "limit test", "General"),
        ("Annualized net volatility", "=Performance!D6", "%", "realized", PCT2),
        ("Volatility target", "=Assumptions!E16", "%", "approved target", PCT2),
        ("Volatility scaling", "=IFERROR(C12/C11,0)", "x", "target / realized", MULT),
    ]
    for row, (label, formula, units, note, number_format) in enumerate(risk_metrics, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
        sheet.cell(row, 3).number_format = number_format
        sheet.cell(row, 4, units)
        sheet.cell(row, 5, note)
    add_status_rules(sheet, "C10")
    set_widths(sheet, {"A": 4, "B": 32, "C": 18, "D": 14, "E": 40})

    sheet = workbook["Stress"]
    title(sheet, "B2:G2", "Strategy Stress Scenarios")
    header(sheet, 4, 2, ["Scenario", "Return shock", "Turnover multiplier", "Liquidity multiplier", "Cost shock", "Estimated P&L"])
    scenarios = [
        ("Equity selloff / beta shock", -0.12, 1.5, 0.6),
        ("Volatility spike", -0.06, 2.0, 0.5),
        ("Crowded unwind", -0.10, 3.0, 0.3),
        ("Liquidity freeze", -0.04, 1.5, 0.15),
        ("Model decay", -0.03, 1.0, 1.0),
    ]
    for row, (name, shock, turnover, liquidity) in enumerate(scenarios, start=5):
        sheet.cell(row, 2, name)
        input_cell(sheet.cell(row, 3, shock), PCT2)
        input_cell(sheet.cell(row, 4, turnover), MULT)
        input_cell(sheet.cell(row, 5, liquidity), MULT)
        sheet.cell(row, 6, f"=AVERAGE(Backtest!D5:D64)*D{row}*Assumptions!$E$7/10000+Assumptions!$E$8*(AVERAGE(Backtest!D5:D64)*D{row}*Assumptions!$E$9/(Assumptions!$E$10*E{row}))^1.5")
        sheet.cell(row, 6).number_format = PCT2
        sheet.cell(row, 7, f"=C{row}-F{row}")
        sheet.cell(row, 7).number_format = PCT2
    set_widths(sheet, {"A": 4, "B": 32, "C": 16, "D": 18, "E": 18, "F": 16, "G": 16})

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Research and Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("Signal lag is positive", '=IF(Assumptions!E18>=1,"PASS","FAIL")'),
        ("Publication lag control enabled", '=IF(Assumptions!E19=1,"PASS","FAIL")'),
        ("Transaction costs included", '=IF(AND(Assumptions!E7>0,Assumptions!E8>0),"PASS","FAIL")'),
        ("No participation breach", '=IF(\'Costs & Capacity\'!C14="PASS","PASS","BREACH")'),
        ("Gross return exceeds or equals net", '=IF(Performance!C5>=Performance!D5,"PASS","FAIL")'),
        ("Walk-forward tests populated", '=IF(COUNT(\'Walk Forward\'!H5:H10)=6,"PASS","FAIL")'),
        ("No drawdown breach", '=IF(Risk!C10="PASS","PASS","REVIEW")'),
        ("All returns finite", '=IF(COUNT(Backtest!H5:H64)=60,"PASS","FAIL")'),
        ("Overall", '=IF(COUNTIF(C5:C12,"FAIL")=0,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    add_status_rules(sheet, "C5:C13")
    set_widths(sheet, {"A": 4, "B": 46, "C": 18})

    add_sources(workbook, [
        ("Point-in-time market and fundamental data", "[dataset URL / manifest]", "[as-of]", "Document revisions, survivorship, and publication lag"),
        ("Corporate actions and universe membership", "[dataset URL]", "[as-of]", "Split, dividend, delisting, and constituent history"),
        ("Execution and transaction-cost data", "[broker / venue source]", "[period]", "Spreads, fees, impact, and borrow"),
        ("Risk-free benchmark", "https://fred.stlouisfed.org/", "[period]", "Record series ID and compounding"),
        ("Live-monitoring comparison", "[production report URL]", "[period]", "Backtest-to-live degradation and incidents"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("QUANT_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
