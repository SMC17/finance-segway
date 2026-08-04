import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[COMPANY] — Distressed / Restructuring Model", [
    ("Situation type:", "Chapter 11 / Out-of-court / Chapter 7"),
    ("Petition/filing date:", "[date]"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly (daily near key dates)"),
])

# ---------------- CAPITAL STRUCTURE & RECOVERY WATERFALL ----------------
ws = wb.create_sheet("Recovery Waterfall")
set_col_widths(ws, [4, 24, 14, 14, 14, 14, 16])
ws["B2"] = "Capital Structure & Recovery Waterfall"; ws["B2"].font = TITLE
headers = ["", "Tranche", "Face claim ($)", "Seniority rank", "Recovery ($)", "Recovery %", "Fulcrum?"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers)-1)
tranches = ["DIP / super-priority", "First lien secured", "Second lien secured",
            "Senior unsecured notes", "Subordinated debt", "Equity"]
r = 5
for i, t in enumerate(tranches):
    ws.cell(row=r, column=2, value=t).font = BLACK
    c_face = ws.cell(row=r, column=3, value=0); c_face.font = BLUE; c_face.number_format = CUR; c_face.border = BORDER
    ws.cell(row=r, column=4, value=i+1).font = BLUE
    r += 1
last_tranche_row = r - 1

ws["B12"] = "Total enterprise value available for distribution ($)"
ws["C12"] = 0; ws["C12"].font = BLUE; ws["C12"].fill = YELLOW_FILL; ws["C12"].number_format = CUR

# waterfall: each tranche gets min(remaining value, its face claim), cascading by seniority
ws["B14"] = "Waterfall (senior to junior, absolute priority)"
ws["B14"].font = BOLD; ws["B14"].fill = GRAY_FILL
r = 5
for i in range(len(tranches)):
    remaining_formula = "C12" if i == 0 else f"MAX(C12-SUM(E5:E{4+i}),0)"
    ws.cell(row=r, column=5, value=f"=MIN({remaining_formula},C{r})")
    ws.cell(row=r, column=5).number_format = CUR
    ws.cell(row=r, column=5).border = BORDER
    ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/C{r},\"-\")")
    ws.cell(row=r, column=6).number_format = PCT
    ws.cell(row=r, column=6).border = BORDER
    # Fulcrum = the first tranche (top-down) whose recovery drops below 100%,
    # given the tranche immediately senior to it was paid in full. Guarded on
    # ISNUMBER so a blank template (all "-") never falsely flags a fulcrum.
    if i == 0:
        fulcrum_formula = f'=IF(AND(ISNUMBER(F{r}),F{r}<1),"FULCRUM","")'
    else:
        fulcrum_formula = (f'=IF(AND(ISNUMBER(F{r}),F{r}<1,ISNUMBER(F{r-1}),F{r-1}>=1),'
                            f'"FULCRUM","")')
    ws.cell(row=r, column=7, value=fulcrum_formula)
    ws.cell(row=r, column=7).font = BOLD
    ws.cell(row=r, column=7).border = BORDER
    r += 1
ws["B20"] = ("Fulcrum security = the tranche where recovery % first drops below 100% "
             "— that class controls the reorg (converts to new equity)")
ws["B20"].font = ITALIC_GRAY
ws["B21"] = "Fulcrum security"
ws["C21"] = (f'=IFERROR(INDEX(B5:B{last_tranche_row},'
             f'MATCH("FULCRUM",G5:G{last_tranche_row},0)),"-")')
ws["C21"].font = BOLD; ws["C21"].border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- LIQUIDATION VS REORG ----------------
ws = wb.create_sheet("Liquidation vs Reorg")
set_col_widths(ws, [4, 30, 16, 16, 40])
ws["B2"] = "Liquidation (Ch. 7) vs. Reorganization (Ch. 11) NPV"; ws["B2"].font = TITLE
headers = ["", "", "Liquidation", "Reorg", "Notes"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 3, start_col=3)

ws["B5"] = "Gross asset value / going-concern EV"
ws["C5"] = 0; ws["C5"].font = BLUE; ws["C5"].number_format = CUR
ws["D5"] = 0; ws["D5"].font = BLUE; ws["D5"].number_format = CUR
ws["B6"] = "Less: liquidation discount / distress costs"
ws["C6"] = 0; ws["C6"].font = BLUE; ws["C6"].number_format = CUR
ws["B7"] = "Less: administrative/professional fees"
ws["C7"] = 0; ws["C7"].font = BLUE; ws["C7"].number_format = CUR
ws["D7"] = 0; ws["D7"].font = BLUE; ws["D7"].number_format = CUR
ws["B8"] = "Time to distribution (months)"
ws["C8"] = 6; ws["C8"].font = BLUE; ws["C8"].number_format = NUM
ws["D8"] = 18; ws["D8"].font = BLUE; ws["D8"].number_format = NUM
ws["B9"] = "Discount rate (annual, for NPV of delayed recovery)"
ws["C9"] = 0.15; ws["C9"].font = BLUE; ws["C9"].fill = YELLOW_FILL; ws["C9"].number_format = PCT

ws["B11"] = "Net proceeds"
ws["C11"] = "=C5-C6-C7"; ws["C11"].number_format = CUR
ws["D11"] = "=D5-D7"; ws["D11"].number_format = CUR
ws["B12"] = "NPV of proceeds (discounted for time to distribution)"
ws["C12"] = "=C11/(1+$C$9)^(C8/12)"; ws["C12"].font = BOLD; ws["C12"].number_format = CUR
ws["D12"] = "=D11/(1+$C$9)^(D8/12)"; ws["D12"].font = BOLD; ws["D12"].number_format = CUR
ws["E12"] = "Compare NPVs — higher wins for creditors as a class, though individual tranche outcomes differ"
ws["E12"].font = ITALIC_GRAY
for r2 in range(5, 13):
    for c in (3, 4):
        ws.cell(row=r2, column=c).border = BORDER
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "RESTRUCTURING_template.xlsx"
wb.save(out_path)
print("saved", out_path)
