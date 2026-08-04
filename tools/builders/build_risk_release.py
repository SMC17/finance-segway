"""Canonical release wrapper for the institutional risk-management builder.

Adds position-level standalone and Euler component VaR, stress contribution,
and liquidity-adjusted VaR to the core portfolio risk workbook.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_risk_institutional import build as build_layout  # noqa: E402
from institutional_helpers import CUR, PCT2, add_status_rules, finalize, header, set_widths, title, total_row  # noqa: E402

FACTOR_POSITION_COLUMNS = [14, 15, 16, 17, 18]  # N:R on Positions
FACTOR_ROWS = [5, 6, 7, 8, 9]                   # Equity through Volatility
POSITION_ROWS = range(5, 17)


def _standalone_variance_formula(row: int) -> str:
    terms: list[str] = []
    for i, position_column_i in enumerate(FACTOR_POSITION_COLUMNS):
        position_letter_i = get_column_letter(position_column_i)
        factor_row_i = FACTOR_ROWS[i]
        for j, position_column_j in enumerate(FACTOR_POSITION_COLUMNS):
            position_letter_j = get_column_letter(position_column_j)
            factor_row_j = FACTOR_ROWS[j]
            correlation_column = get_column_letter(5 + j)
            terms.append(
                f"(Positions!$E${row}*Positions!${position_letter_i}${row})*"
                f"(Positions!$E${row}*Positions!${position_letter_j}${row})*"
                f"Factors!$D${factor_row_i}*Factors!$D${factor_row_j}*"
                f"Factors!${correlation_column}${factor_row_i}"
            )
    return "+".join(terms)


def _portfolio_covariance_formula(row: int) -> str:
    terms: list[str] = []
    for i, position_column in enumerate(FACTOR_POSITION_COLUMNS):
        position_letter = get_column_letter(position_column)
        factor_row_i = FACTOR_ROWS[i]
        for j, factor_row_j in enumerate(FACTOR_ROWS):
            correlation_column = get_column_letter(5 + j)
            terms.append(
                f"(Positions!$E${row}*Positions!${position_letter}${row})*"
                f"Factors!$D${factor_row_i}*Factors!$C${factor_row_j}*"
                f"Factors!$D${factor_row_j}*Factors!${correlation_column}${factor_row_i}"
            )
    return "+".join(terms)


def build(output: Path) -> None:
    build_layout(output)
    workbook = load_workbook(output)
    if "Risk Contributions" in workbook.sheetnames:
        del workbook["Risk Contributions"]
    position = workbook.sheetnames.index("VaR & ES") + 1
    sheet = workbook.create_sheet("Risk Contributions", position)
    title(sheet, "B2:J2", "Position-Level VaR, Stress, and Liquidity Contributions")
    header(sheet, 4, 2, [
        "Instrument", "Gross notional", "Standalone VaR", "Component VaR",
        "% portfolio VaR", "Worst stress loss", "Liquidation loss",
        "Liquidity-adjusted risk", "Risk concentration",
    ])

    for output_row, source_row in enumerate(POSITION_ROWS, start=5):
        sheet.cell(output_row, 2, f"=Positions!B{source_row}")
        sheet.cell(output_row, 3, f"=ABS(Positions!E{source_row})")
        sheet.cell(output_row, 3).number_format = CUR

        standalone_variance = _standalone_variance_formula(source_row)
        sheet.cell(
            output_row,
            4,
            f"='VaR & ES'!$C$7*SQRT(MAX(0,({standalone_variance})/Assumptions!$E$6))",
        )
        sheet.cell(output_row, 4).number_format = CUR

        portfolio_covariance = _portfolio_covariance_formula(source_row)
        sheet.cell(
            output_row,
            5,
            f"=IFERROR('VaR & ES'!$C$7*({portfolio_covariance})/"
            f"(Factors!$C$13*SQRT(Assumptions!$E$6)),0)",
        )
        sheet.cell(output_row, 5).number_format = CUR
        sheet.cell(output_row, 6, f"=IFERROR(E{output_row}/'VaR & ES'!$C$8,0)")
        sheet.cell(output_row, 6).number_format = PCT2

        stress_terms = []
        for scenario_row in range(5, 11):
            factor_terms = []
            for factor_index, position_column in enumerate(FACTOR_POSITION_COLUMNS):
                position_letter = get_column_letter(position_column)
                stress_column = get_column_letter(3 + factor_index)
                factor_terms.append(
                    f"Positions!$E${source_row}*Positions!${position_letter}${source_row}*"
                    f"Stress!${stress_column}${scenario_row}"
                )
            stress_terms.append("(" + "+".join(factor_terms) + ")")
        sheet.cell(output_row, 7, "=-MIN(" + ",".join(stress_terms) + ")")
        sheet.cell(output_row, 7).number_format = CUR
        sheet.cell(output_row, 8, f"=Liquidity!G{source_row}")
        sheet.cell(output_row, 8).number_format = CUR
        sheet.cell(output_row, 9, f"=ABS(E{output_row})+G{output_row}+H{output_row}")
        sheet.cell(output_row, 9).number_format = CUR
        sheet.cell(output_row, 10, f"=IFERROR(I{output_row}/SUM($I$5:$I$16),0)")
        sheet.cell(output_row, 10).number_format = PCT2

    total_row(sheet, 18, 2, 10)
    sheet["B18"] = "Portfolio total"
    for column in range(3, 10):
        letter = get_column_letter(column)
        sheet.cell(18, column, f"=SUM({letter}5:{letter}16)")
        sheet.cell(18, column).number_format = PCT2 if column == 6 else CUR
    sheet["J18"] = "=SUM(J5:J16)"
    sheet["J18"].number_format = PCT2

    sheet["B20"] = "Component VaR reconciliation"
    sheet["C20"] = "=E18-'VaR & ES'!C8"
    sheet["C20"].number_format = CUR
    sheet["D20"] = '=IF(ABS(C20)<0.01,"PASS","REVIEW")'
    add_status_rules(sheet, "D20")
    sheet["B21"] = "Largest risk concentration"
    sheet["C21"] = "=MAX(J5:J16)"
    sheet["C21"].number_format = PCT2
    sheet["B22"] = "Largest contributor"
    sheet["C22"] = "=INDEX(B5:B16,MATCH(MAX(J5:J16),J5:J16,0))"

    set_widths(sheet, {
        "A": 4, "B": 28, "C": 16, "D": 16, "E": 16, "F": 16,
        "G": 18, "H": 16, "I": 20, "J": 18,
    })
    sheet.freeze_panes = "B5"
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("RISK_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
