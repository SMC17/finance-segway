"""Build the canonical fixed-income and rates decision model."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, CUR2, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "Zero Curve", "Bond Analytics", "Portfolio",
    "Key Rate Risk", "Carry & Roll", "Scenarios", "P&L Explain", "Checks",
    "Sources", "RefreshLog",
]


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[PORTFOLIO / SECURITY] — Fixed Income & Rates Model", [
        ("Portfolio / security:", "[Name]"),
        ("Currency / curve:", "[Currency and benchmark]"),
        ("Last refreshed:", "[timestamp]"),
        ("Next coupon / auction / event:", "[date]"),
        ("Refresh cadence:", "Daily"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions; yields in percent"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Pricing and Risk Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Face value", 1000.0, 1000.0, "currency", CUR2),
        ("Coupon rate", 0.050, 0.050, "%", PCT2),
        ("Yield to maturity", 0.055, 0.070, "%", PCT2),
        ("Maturity", 10.0, 10.0, "years", "0.0"),
        ("Coupon frequency", 2.0, 2.0, "per year", "0"),
        ("Parallel yield shock", 0.010, 0.020, "%", PCT2),
        ("One-basis-point bump", 0.0001, 0.0001, "%", "0.0000%"),
        ("Holding period", 1.0, 1.0, "years", "0.0"),
        ("Financing / repo rate", 0.045, 0.060, "%", PCT2),
        ("Position market value", 50.0, 50.0, "$mm", CUR),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 36, "C": 15, "D": 15, "E": 15, "F": 30})

    sheet = workbook["Zero Curve"]
    title(sheet, "B2:G2", "Zero Curve and Discount Factors")
    header(sheet, 4, 2, ["Tenor", "Base zero", "Downside zero", "Active zero", "Discount factor", "1Y forward"])
    tenors = [0.25, 0.50, 1.00, 2.00, 3.00, 5.00, 7.00, 10.00, 20.00, 30.00]
    base_rates = [0.048, 0.047, 0.045, 0.043, 0.042, 0.043, 0.045, 0.047, 0.049, 0.050]
    downside_rates = [rate + (0.005 if tenor <= 3 else 0.012) for tenor, rate in zip(tenors, base_rates)]
    for row, (tenor, base, downside) in enumerate(zip(tenors, base_rates, downside_rates), start=5):
        input_cell(sheet.cell(row, 2, tenor), "0.00")
        input_cell(sheet.cell(row, 3, base), PCT2)
        input_cell(sheet.cell(row, 4, downside), PCT2)
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = PCT2
        sheet.cell(row, 6, f"=EXP(-E{row}*B{row})")
        sheet.cell(row, 6).number_format = "0.000000"
        if row == 5:
            sheet.cell(row, 7, "=E5")
        else:
            sheet.cell(row, 7, f"=IFERROR(-LN(F{row}/F{row-1})/(B{row}-B{row-1}),0)")
        sheet.cell(row, 7).number_format = PCT2
    set_widths(sheet, {"A": 4, "B": 14, "C": 16, "D": 16, "E": 16, "F": 18, "G": 16})

    sheet = workbook["Bond Analytics"]
    title(sheet, "B2:E2", "Bond Price, Duration, and Convexity")
    header(sheet, 4, 2, ["Metric", "Value", "Units", "Method"])
    metrics = [
        ("Price", "=-PV(Assumptions!E7/Assumptions!E9,Assumptions!E8*Assumptions!E9,Assumptions!E5*Assumptions!E6/Assumptions!E9,Assumptions!E5)", "currency", "coupon annuity plus principal", CUR2),
        ("Price +1bp", "=-PV((Assumptions!E7+Assumptions!E11)/Assumptions!E9,Assumptions!E8*Assumptions!E9,Assumptions!E5*Assumptions!E6/Assumptions!E9,Assumptions!E5)", "currency", "finite difference", CUR2),
        ("Price -1bp", "=-PV((Assumptions!E7-Assumptions!E11)/Assumptions!E9,Assumptions!E8*Assumptions!E9,Assumptions!E5*Assumptions!E6/Assumptions!E9,Assumptions!E5)", "currency", "finite difference", CUR2),
        ("Numerical modified duration", "=IFERROR((C7-C6)/(2*C5*Assumptions!E11),0)", "years", "central difference", "0.0000"),
        ("Numerical convexity", "=IFERROR((C6+C7-2*C5)/(C5*Assumptions!E11^2),0)", "years^2", "second difference", "0.0000"),
        ("DV01", "=(C7-C6)/2", "currency", "price value of 1bp", CUR2),
        ("Current yield", "=Assumptions!E5*Assumptions!E6/C5", "%", "annual coupon / price", PCT2),
        ("Price under parallel shock", "=-PV((Assumptions!E7+Assumptions!E10)/Assumptions!E9,Assumptions!E8*Assumptions!E9,Assumptions!E5*Assumptions!E6/Assumptions!E9,Assumptions!E5)", "currency", "full repricing", CUR2),
        ("Duration-convexity estimate", "=C5*(1-C8*Assumptions!E10+0.5*C9*Assumptions!E10^2)", "currency", "second-order approximation", CUR2),
        ("Approximation residual", "=C12-C13", "currency", "full price less estimate", CUR2),
    ]
    for row, (label, formula, units, note, number_format) in enumerate(metrics, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
        sheet.cell(row, 3).number_format = number_format
        sheet.cell(row, 4, units)
        sheet.cell(row, 5, note)
    set_widths(sheet, {"A": 4, "B": 34, "C": 18, "D": 16, "E": 44})

    sheet = workbook["Portfolio"]
    title(sheet, "B2:K2", "Fixed-Income Portfolio Analytics")
    header(sheet, 4, 2, ["Security", "Market value", "Coupon", "Yield", "Maturity", "Price", "Mod duration", "Convexity", "DV01", "Spread duration"])
    securities = [
        ("2Y Treasury", 30, 0.045, 0.044, 2.0, 0.0),
        ("5Y Treasury", 40, 0.043, 0.045, 5.0, 0.0),
        ("10Y Treasury", 50, 0.050, 0.055, 10.0, 0.0),
        ("IG Corporate", 45, 0.060, 0.068, 7.0, 5.5),
        ("HY Corporate", 25, 0.085, 0.105, 5.0, 3.8),
        ("Municipal", 20, 0.040, 0.046, 12.0, 7.5),
    ]
    for row, values in enumerate(securities, start=5):
        for column, value in enumerate(values, start=2):
            input_cell(sheet.cell(row, column, value), CUR if column == 3 else (PCT2 if column in (4, 5) else "0.0"))
        sheet.cell(row, 8, f"=-PV(E{row}/2,F{row}*2,100*D{row}/2,100)")
        sheet.cell(row, 8).number_format = CUR2
        up = f"-PV((E{row}+0.0001)/2,F{row}*2,100*D{row}/2,100)"
        down = f"-PV((E{row}-0.0001)/2,F{row}*2,100*D{row}/2,100)"
        sheet.cell(row, 9, f"=IFERROR((({down})-({up}))/(2*H{row}*0.0001),0)")
        sheet.cell(row, 9).number_format = "0.0000"
        sheet.cell(row, 10, f"=IFERROR((({up})+({down})-2*H{row})/(H{row}*0.0001^2),0)")
        sheet.cell(row, 10).number_format = "0.0000"
        sheet.cell(row, 11, f"=C{row}*I{row}*0.0001")
        sheet.cell(row, 11).number_format = CUR
        sheet.cell(row, 12, f"=G{row}")
        sheet.cell(row, 12).number_format = "0.0000"
    total_row(sheet, 12, 2, 12)
    sheet["B12"] = "Portfolio total / weighted"
    sheet["C12"] = "=SUM(C5:C10)"
    sheet["C12"].number_format = CUR
    for column in (9, 10, 12):
        letter = get_column_letter(column)
        sheet.cell(12, column, f"=SUMPRODUCT(C5:C10,{letter}5:{letter}10)/SUM(C5:C10)")
        sheet.cell(12, column).number_format = "0.0000"
    sheet["K12"] = "=SUM(K5:K10)"
    sheet["K12"].number_format = CUR
    set_widths(sheet, {"A": 4, "B": 24, "C": 16, "D": 13, "E": 13, "F": 13, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 16})

    sheet = workbook["Key Rate Risk"]
    title(sheet, "B2:H2", "Key-Rate DV01 Approximation")
    header(sheet, 4, 2, ["Security", "2Y", "5Y", "10Y", "20Y", "30Y", "Total DV01"])
    key_tenors = [2, 5, 10, 20, 30]
    for row in range(5, 11):
        sheet.cell(row, 2, f"=Portfolio!B{row}")
        maturity = f"Portfolio!F{row}"
        for index, tenor in enumerate(key_tenors, start=3):
            letter = get_column_letter(index)
            sheet.cell(row, index, f"=Portfolio!K{row}*MAX(0,1-ABS({maturity}-{tenor})/MAX(2,{tenor}))")
            sheet.cell(row, index).number_format = CUR
        sheet.cell(row, 8, f"=SUM(C{row}:G{row})")
        sheet.cell(row, 8).number_format = CUR
    total_row(sheet, 12, 2, 8, CUR)
    sheet["B12"] = "Portfolio key-rate DV01"
    for column in range(3, 9):
        letter = get_column_letter(column)
        sheet.cell(12, column, f"=SUM({letter}5:{letter}10)")
        sheet.cell(12, column).number_format = CUR
    set_widths(sheet, {"A": 4, "B": 24, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 16})

    sheet = workbook["Carry & Roll"]
    title(sheet, "B2:G2", "One-Year Carry and Roll-Down")
    header(sheet, 4, 2, ["Security", "Coupon carry", "Financing cost", "Yield after roll", "Roll P&L", "Total carry + roll"])
    for row in range(5, 11):
        sheet.cell(row, 2, f"=Portfolio!B{row}")
        sheet.cell(row, 3, f"=Portfolio!C{row}*Portfolio!D{row}")
        sheet.cell(row, 4, f"=Portfolio!C{row}*Assumptions!E13")
        sheet.cell(row, 5, f"=MAX(0,Portfolio!E{row}-0.002)")
        sheet.cell(row, 6, f"=Portfolio!C{row}*Portfolio!I{row}*(Portfolio!E{row}-E{row})")
        sheet.cell(row, 7, f"=C{row}-D{row}+F{row}")
        for column in (3, 4, 6, 7):
            sheet.cell(row, column).number_format = CUR
        sheet.cell(row, 5).number_format = PCT2
    total_row(sheet, 12, 2, 7, CUR)
    sheet["B12"] = "Portfolio total"
    for column in (3, 4, 6, 7):
        letter = get_column_letter(column)
        sheet.cell(12, column, f"=SUM({letter}5:{letter}10)")
        sheet.cell(12, column).number_format = CUR
    set_widths(sheet, {"A": 4, "B": 24, "C": 16, "D": 16, "E": 16, "F": 16, "G": 18})

    sheet = workbook["Scenarios"]
    title(sheet, "B2:H2", "Curve and Spread Scenarios")
    header(sheet, 4, 2, ["Scenario", "2Y", "5Y", "10Y", "20Y", "30Y", "Credit spread", "Estimated P&L"])
    scenarios = [
        ("Parallel +100bp", 0.010, 0.010, 0.010, 0.010, 0.010, 0.000),
        ("Parallel -100bp", -0.010, -0.010, -0.010, -0.010, -0.010, 0.000),
        ("Bear steepener", 0.005, 0.008, 0.012, 0.015, 0.018, 0.002),
        ("Bull flattener", -0.012, -0.010, -0.008, -0.006, -0.005, -0.001),
        ("Credit widening", 0.002, 0.003, 0.004, 0.005, 0.005, 0.015),
        ("Risk-off rally", -0.008, -0.010, -0.012, -0.010, -0.008, 0.020),
    ]
    for row, scenario in enumerate(scenarios, start=5):
        sheet.cell(row, 2, scenario[0])
        for column, shock in enumerate(scenario[1:], start=3):
            input_cell(sheet.cell(row, column, shock), PCT2)
        sheet.cell(row, 9, f"=-SUMPRODUCT('Key Rate Risk'!C12:G12,C{row}:G{row})/0.0001-SUMPRODUCT(Portfolio!C5:C10,Portfolio!L5:L10)*H{row}")
        sheet.cell(row, 9).number_format = CUR
    set_widths(sheet, {"A": 4, "B": 24, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13, "H": 16, "I": 18})

    sheet = workbook["P&L Explain"]
    title(sheet, "B2:F2", "Rates Portfolio P&L Explain")
    header(sheet, 4, 2, ["Driver", "Exposure", "Move", "P&L", "Notes"])
    rows = [
        ("Carry", "='Carry & Roll'!C12-'Carry & Roll'!D12", 1.0, "coupon less financing"),
        ("Roll-down", "='Carry & Roll'!F12", 1.0, "curve roll"),
        ("Parallel rates", "=-Portfolio!K12/0.0001", 0.0005, "DV01 x move"),
        ("Curve shape", "=-('Key Rate Risk'!C12*0.0002+'Key Rate Risk'!D12*0.0001-'Key Rate Risk'!F12*0.0001)/0.0001", 1.0, "key-rate attribution"),
        ("Credit spread", "=-SUMPRODUCT(Portfolio!C5:C10,Portfolio!L5:L10)", 0.0015, "spread duration"),
        ("Convexity", "=0.5*SUMPRODUCT(Portfolio!C5:C10,Portfolio!J5:J10)", 0.0005**2, "second order"),
    ]
    for row, (driver, exposure, move, note) in enumerate(rows, start=5):
        sheet.cell(row, 2, driver)
        sheet.cell(row, 3, exposure)
        sheet.cell(row, 3).number_format = CUR
        input_cell(sheet.cell(row, 4, move), PCT2 if abs(move) < 1 else "0.0000")
        sheet.cell(row, 5, f"=C{row}*D{row}")
        sheet.cell(row, 5).number_format = CUR
        sheet.cell(row, 6, note)
    sheet["B13"] = "Explained P&L"
    sheet["E13"] = "=SUM(E5:E10)"
    sheet["E13"].number_format = CUR
    input_cell(sheet["E14"], CUR)
    sheet["B14"] = "Actual P&L"
    sheet["E14"] = 1.2
    sheet["B15"] = "Unexplained P&L"
    sheet["E15"] = "=E14-E13"
    sheet["E15"].number_format = CUR
    set_widths(sheet, {"A": 4, "B": 22, "C": 18, "D": 16, "E": 18, "F": 38})

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Fixed-Income Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("Price positive", '=IF(\'Bond Analytics\'!C5>0,"PASS","FAIL")'),
        ("Price decreases as yield rises", '=IF(\'Bond Analytics\'!C6<\'Bond Analytics\'!C5,"PASS","FAIL")'),
        ("Duration positive", '=IF(\'Bond Analytics\'!C8>0,"PASS","FAIL")'),
        ("Convexity positive", '=IF(\'Bond Analytics\'!C9>0,"PASS","FAIL")'),
        ("Key-rate DV01 reconciles", '=IF(ABS(\'Key Rate Risk\'!H12-Portfolio!K12)<0.10,"PASS","REVIEW")'),
        ("Curve discount factors decreasing", '=IF(MAX(\'Zero Curve\'!F6:F14-\'Zero Curve\'!F5:F13)<=0,"PASS","REVIEW")'),
        ("P&L explain populated", '=IF(COUNT(\'P&L Explain\'!E5:E10)=6,"PASS","FAIL")'),
        ("Overall", '=IF(COUNTIF(C5:C11,"FAIL")=0,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    add_status_rules(sheet, "C5:C12")
    set_widths(sheet, {"A": 4, "B": 44, "C": 18})

    add_sources(workbook, [
        ("Security terms and cash flows", "[prospectus / terms URL]", "[date]", "Coupon, maturity, callability, indexation"),
        ("Benchmark and zero curves", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates", "[timestamp]", "Document interpolation and compounding"),
        ("Credit spreads and prices", "[market data URL]", "[timestamp]", "Bid/ask, liquidity, source hierarchy"),
        ("Financing and repo", "[broker / central bank source]", "[date]", "Haircuts, specials, term, and currency basis"),
        ("Portfolio positions", "[OMS / custody source]", "[timestamp]", "Reconcile quantities, accrued interest, and market value"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("FIXED_INCOME_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
