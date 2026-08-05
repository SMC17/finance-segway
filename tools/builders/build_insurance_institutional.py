"""Build the canonical insurance and actuarial decision model."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "Paid Triangle", "Chain Ladder",
    "Bornhuetter-Ferguson", "Underwriting", "Embedded Value", "Capital",
    "Stress", "Checks", "Sources", "RefreshLog",
]

ACCIDENT_YEARS = 10
DEVELOPMENT_PERIODS = 10


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[INSURER / BOOK] — Insurance & Actuarial Model", [
        ("Entity / line of business:", "[Name / LOB]"),
        ("Valuation date:", "[date]"),
        ("Last refreshed:", "[date]"),
        ("Next reserve / capital review:", "[date]"),
        ("Refresh cadence:", "Quarterly and on material claims events"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Actuarial, Underwriting, and Capital Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Expected loss ratio", 0.62, 0.78, "%", PCT),
        ("Expense ratio", 0.28, 0.32, "%", PCT),
        ("Annual premium growth", 0.05, -0.05, "%", PCT),
        ("Claims inflation", 0.035, 0.075, "%", PCT),
        ("Discount rate", 0.045, 0.035, "%", PCT2),
        ("Risk margin / reserve", 0.08, 0.15, "%", PCT),
        ("Opening earned premium", 500.0, 500.0, "$mm", CUR),
        ("Opening invested assets", 900.0, 900.0, "$mm", CUR),
        ("Investment yield", 0.040, 0.025, "%", PCT2),
        ("Opening available capital", 350.0, 350.0, "$mm", CUR),
        ("Premium risk factor", 0.18, 0.24, "% premium", PCT),
        ("Reserve risk factor", 0.12, 0.18, "% reserve", PCT),
        ("Asset risk factor", 0.08, 0.14, "% assets", PCT),
        ("Minimum capital coverage", 1.50, 1.50, "x", MULT),
        ("Cost of capital", 0.10, 0.12, "%", PCT),
        ("Projection horizon", 5.0, 5.0, "years", "0"),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 40, "C": 15, "D": 15, "E": 15, "F": 30})
    sheet.freeze_panes = "A5"

    sheet = workbook["Paid Triangle"]
    title(sheet, "B2:L2", "Cumulative Paid Loss Triangle")
    header(sheet, 4, 2, ["Accident year", *[f"{12 * dev}m" for dev in range(1, DEVELOPMENT_PERIODS + 1)]])
    base_ultimate = [230, 245, 255, 270, 285, 300, 320, 335, 355, 375]
    development = [0.28, 0.48, 0.63, 0.73, 0.81, 0.87, 0.91, 0.94, 0.97, 1.00]
    for ay in range(ACCIDENT_YEARS):
        row = 5 + ay
        sheet.cell(row, 2, 2017 + ay)
        observed = DEVELOPMENT_PERIODS - ay
        for dev in range(observed):
            column = 3 + dev
            value = round(base_ultimate[ay] * development[dev], 2)
            input_cell(sheet.cell(row, column, value), CUR)
    set_widths(sheet, {"A": 4, "B": 16, **{get_column_letter(column): 13 for column in range(3, 13)}})
    sheet.freeze_panes = "C5"

    sheet = workbook["Chain Ladder"]
    title(sheet, "B2:N2", "Chain-Ladder Reserve Development")
    header(sheet, 4, 2, ["Metric", *[f"{12 * dev}m" for dev in range(1, DEVELOPMENT_PERIODS + 1)], "Ultimate", "Reserve"])
    sheet["B5"] = "Selected age-to-age factor"
    for dev in range(1, DEVELOPMENT_PERIODS):
        column = 2 + dev
        letter = get_column_letter(column)
        next_letter = get_column_letter(column + 1)
        last_observed_row = 14 - dev
        sheet.cell(5, column, f"=IFERROR(SUM('Paid Triangle'!{next_letter}5:{next_letter}{last_observed_row})/SUM('Paid Triangle'!{letter}5:{letter}{last_observed_row}),1)")
        sheet.cell(5, column).number_format = "0.0000x"
    sheet.cell(5, 12, 1.0)
    sheet.cell(5, 12).number_format = "0.0000x"
    sheet["B6"] = "Cumulative development factor"
    for dev in range(DEVELOPMENT_PERIODS):
        column = 3 + dev
        letter = get_column_letter(column)
        sheet.cell(6, column, f"=PRODUCT({letter}$5:$K$5)")
        sheet.cell(6, column).number_format = "0.0000x"
    for ay in range(ACCIDENT_YEARS):
        row = 8 + ay
        triangle_row = 5 + ay
        observed = DEVELOPMENT_PERIODS - ay
        latest_column = 2 + observed
        latest_letter = get_column_letter(latest_column)
        sheet.cell(row, 2, f"='Paid Triangle'!B{triangle_row}")
        for dev in range(DEVELOPMENT_PERIODS):
            column = 3 + dev
            if dev < observed:
                letter = get_column_letter(column)
                sheet.cell(row, column, f"='Paid Triangle'!{letter}{triangle_row}")
                sheet.cell(row, column).number_format = CUR
        cdf_column = 2 + observed
        cdf_letter = get_column_letter(cdf_column)
        sheet.cell(row, 13, f"='Paid Triangle'!{latest_letter}{triangle_row}*{cdf_letter}$6")
        sheet.cell(row, 14, f"=M{row}-'Paid Triangle'!{latest_letter}{triangle_row}")
        sheet.cell(row, 13).number_format = CUR
        sheet.cell(row, 14).number_format = CUR
    total_row(sheet, 18, 2, 14, CUR)
    sheet["B18"] = "Total"
    sheet["M18"] = "=SUM(M8:M17)"
    sheet["N18"] = "=SUM(N8:N17)"
    set_widths(sheet, {"A": 4, "B": 26, **{get_column_letter(column): 13 for column in range(3, 15)}})
    sheet.freeze_panes = "C8"

    sheet = workbook["Bornhuetter-Ferguson"]
    title(sheet, "B2:H2", "Bornhuetter-Ferguson Reserve Estimate")
    header(sheet, 4, 2, ["Accident year", "Earned premium", "Expected loss", "% paid", "Paid to date", "BF ultimate", "BF reserve"])
    for ay in range(ACCIDENT_YEARS):
        row = 5 + ay
        chain_row = 8 + ay
        observed = DEVELOPMENT_PERIODS - ay
        latest_column = 2 + observed
        latest_letter = get_column_letter(latest_column)
        cdf_letter = get_column_letter(latest_column)
        sheet.cell(row, 2, f"='Chain Ladder'!B{chain_row}")
        premium = 400 + ay * 25
        input_cell(sheet.cell(row, 3, premium), CUR)
        sheet.cell(row, 4, f"=C{row}*Assumptions!$E$5*(1+Assumptions!$E$8)^{ay}")
        sheet.cell(row, 5, f"=1/'Chain Ladder'!{cdf_letter}$6")
        sheet.cell(row, 6, f"='Paid Triangle'!{latest_letter}{5+ay}")
        sheet.cell(row, 7, f"=F{row}+D{row}*(1-E{row})")
        sheet.cell(row, 8, f"=G{row}-F{row}")
        for column in (3, 4, 6, 7, 8):
            sheet.cell(row, column).number_format = CUR
        sheet.cell(row, 5).number_format = PCT2
    total_row(sheet, 15, 2, 8, CUR)
    sheet["B15"] = "Total"
    for column in (3, 4, 6, 7, 8):
        letter = get_column_letter(column)
        sheet.cell(15, column, f"=SUM({letter}5:{letter}14)")
    set_widths(sheet, {"A": 4, "B": 16, "C": 18, "D": 18, "E": 14, "F": 18, "G": 18, "H": 18})

    sheet = workbook["Underwriting"]
    title(sheet, "B2:G2", "Five-Year Underwriting Forecast")
    header(sheet, 4, 2, ["$mm / %", *[f"Year {year}" for year in range(1, 6)]])
    labels = [
        "Earned premium", "Incurred losses", "Underwriting expenses",
        "Underwriting result", "Loss ratio", "Expense ratio", "Combined ratio",
        "Investment income", "Pre-tax operating income",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 8):
        letter = get_column_letter(column)
        year = column - 2
        if column == 3:
            sheet.cell(5, column, "=Assumptions!E11")
        else:
            previous = get_column_letter(column - 1)
            sheet.cell(5, column, f"={previous}5*(1+Assumptions!$E$7)")
        sheet.cell(6, column, f"={letter}5*Assumptions!$E$5*(1+Assumptions!$E$8)^{year-1}")
        sheet.cell(7, column, f"={letter}5*Assumptions!$E$6")
        sheet.cell(8, column, f"={letter}5-{letter}6-{letter}7")
        sheet.cell(9, column, f"=IFERROR({letter}6/{letter}5,0)")
        sheet.cell(10, column, f"=IFERROR({letter}7/{letter}5,0)")
        sheet.cell(11, column, f"={letter}9+{letter}10")
        sheet.cell(12, column, f"=Assumptions!$E$12*Assumptions!$E$13")
        sheet.cell(13, column, f"={letter}8+{letter}12")
        for row in (5, 6, 7, 8, 12, 13):
            sheet.cell(row, column).number_format = CUR
        for row in (9, 10, 11):
            sheet.cell(row, column).number_format = PCT2
    for row in (8, 13):
        total_row(sheet, row, 2, 7, CUR)
    set_widths(sheet, {"A": 4, "B": 32, **{get_column_letter(column): 15 for column in range(3, 8)}})

    sheet = workbook["Embedded Value"]
    title(sheet, "B2:G2", "Simplified Embedded Value")
    header(sheet, 4, 2, ["Metric", *[f"Year {year}" for year in range(1, 6)]])
    labels = [
        "Pre-tax operating income", "Tax at 25%", "After-tax distributable earnings",
        "Risk margin charge", "Economic earnings", "Discount factor", "PV economic earnings",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 8):
        letter = get_column_letter(column)
        sheet.cell(5, column, f"=Underwriting!{letter}13")
        sheet.cell(6, column, f"=MAX(0,{letter}5*0.25)")
        sheet.cell(7, column, f"={letter}5-{letter}6")
        sheet.cell(8, column, f"='Bornhuetter-Ferguson'!$H$15*Assumptions!$E$10*Assumptions!$E$19")
        sheet.cell(9, column, f"={letter}7-{letter}8")
        sheet.cell(10, column, f"=1/(1+Assumptions!$E$9)^{column-2}")
        sheet.cell(11, column, f"={letter}9*{letter}10")
        for row in (5, 6, 7, 8, 9, 11):
            sheet.cell(row, column).number_format = CUR
        sheet.cell(10, column).number_format = "0.0000"
    sheet["B13"] = "Adjusted net asset value"
    sheet["C13"] = "=Assumptions!E14-'Bornhuetter-Ferguson'!H15*(1+Assumptions!E10)"
    sheet["C13"].number_format = CUR
    sheet["B14"] = "Present value of future economic earnings"
    sheet["C14"] = "=SUM(C11:G11)"
    sheet["C14"].number_format = CUR
    sheet["B15"] = "Embedded value"
    sheet["C15"] = "=C13+C14"
    sheet["C15"].number_format = CUR
    set_widths(sheet, {"A": 4, "B": 42, **{get_column_letter(column): 15 for column in range(3, 8)}})

    sheet = workbook["Capital"]
    title(sheet, "B2:E2", "Economic and Regulatory Capital Screen")
    header(sheet, 4, 2, ["Capital component", "Exposure", "Factor", "Requirement"])
    capital_rows = [
        ("Premium risk", "=Underwriting!C5", "=Assumptions!E15"),
        ("Reserve risk", "='Bornhuetter-Ferguson'!H15", "=Assumptions!E16"),
        ("Asset risk", "=Assumptions!E12", "=Assumptions!E17"),
    ]
    for row, (label, exposure, factor) in enumerate(capital_rows, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, exposure)
        sheet.cell(row, 3).number_format = CUR
        sheet.cell(row, 4, factor)
        sheet.cell(row, 4).number_format = PCT
        sheet.cell(row, 5, f"=C{row}*D{row}")
        sheet.cell(row, 5).number_format = CUR
    sheet["B9"] = "Diversified capital requirement"
    sheet["E9"] = "=SQRT(SUMSQ(E5:E7)+2*0.25*(E5*E6+E5*E7+E6*E7))"
    sheet["E9"].number_format = CUR
    sheet["B10"] = "Available capital"
    sheet["E10"] = "=Assumptions!E14"
    sheet["E10"].number_format = CUR
    sheet["B11"] = "Capital coverage"
    sheet["E11"] = "=IFERROR(E10/E9,0)"
    sheet["E11"].number_format = MULT
    sheet["B12"] = "Minimum coverage"
    sheet["E12"] = "=Assumptions!E18"
    sheet["E12"].number_format = MULT
    sheet["B13"] = "Capital status"
    sheet["E13"] = '=IF(E11>=E12,"PASS","BREACH")'
    add_status_rules(sheet, "E13")
    set_widths(sheet, {"A": 4, "B": 34, "C": 18, "D": 16, "E": 18})

    sheet = workbook["Stress"]
    title(sheet, "B2:G2", "Reserve and Capital Stress Matrix")
    header(sheet, 4, 2, ["Loss ratio / Asset shock", -0.20, -0.10, 0.00, 0.10, 0.20])
    for column in range(3, 8):
        sheet.cell(4, column).number_format = PCT
    for row, loss_ratio_shock in enumerate((0.00, 0.05, 0.10, 0.15, 0.20), start=5):
        sheet.cell(row, 2, loss_ratio_shock)
        sheet.cell(row, 2).number_format = PCT
        for column in range(3, 8):
            letter = get_column_letter(column)
            sheet.cell(row, column, f"=(Assumptions!$E$14*(1+{letter}$4)-'Bornhuetter-Ferguson'!$H$15*(1+$B{row}))/Capital!$E$9")
            sheet.cell(row, column).number_format = MULT
    sheet.conditional_formatting.add(
        "C5:G9",
        ColorScaleRule(
            start_type="min", start_color="FCE4D6",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="E2F0D9",
        ),
    )
    set_widths(sheet, {"A": 4, "B": 28, **{get_column_letter(column): 14 for column in range(3, 8)}})

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Actuarial Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    # A loss-development triangle is deliberately ragged -- the most recent
    # accident years only have their earliest development periods observed,
    # everything beyond that is blank until future refreshes fill it in
    # (same "observed = DEVELOPMENT_PERIODS - ay" shape used to populate the
    # triangle itself above). A blanket MIN(D5:L14 - C5:K14) treats every
    # blank cell as 0, so "next period (blank=0) minus this period (real
    # value)" is large and negative for every row except the single oldest
    # one that happens to have all 10 periods filled -- failing this check
    # unconditionally for any real triangle, regardless of the data.
    # Compare only adjacent *observed* periods per row instead.
    paid_triangle_deltas = []
    for ay in range(ACCIDENT_YEARS):
        row = 5 + ay
        observed = DEVELOPMENT_PERIODS - ay
        for dev in range(1, observed):
            column = 3 + dev
            letter = get_column_letter(column)
            prev_letter = get_column_letter(column - 1)
            paid_triangle_deltas.append(
                f"'Paid Triangle'!{letter}{row}-'Paid Triangle'!{prev_letter}{row}"
            )
    paid_triangle_formula = f'=IF(MIN({",".join(paid_triangle_deltas)})>=0,"PASS","FAIL")'
    checks = [
        ("Paid triangle cumulative", paid_triangle_formula),
        ("Development factors at least one", '=IF(MIN(\'Chain Ladder\'!C5:L5)>=1,"PASS","FAIL")'),
        ("Chain-ladder reserve nonnegative", '=IF(MIN(\'Chain Ladder\'!N8:N17)>=0,"PASS","FAIL")'),
        ("BF reserve nonnegative", '=IF(MIN(\'Bornhuetter-Ferguson\'!H5:H14)>=0,"PASS","FAIL")'),
        ("Combined ratio finite", '=IF(AND(MIN(Underwriting!C11:G11)>=0,MAX(Underwriting!C11:G11)<3),"PASS","FAIL")'),
        ("Capital requirement positive", '=IF(Capital!E9>0,"PASS","FAIL")'),
        ("Capital coverage", '=IF(Capital!E13="PASS","PASS","REVIEW")'),
        ("Embedded value finite", '=IF(ABS(\'Embedded Value\'!C15)<1000000,"PASS","FAIL")'),
        ("Overall", '=IF(COUNTIF(C5:C12,"FAIL")=0,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    add_status_rules(sheet, "C5:C13")
    set_widths(sheet, {"A": 4, "B": 46, "C": 18})

    add_sources(workbook, [
        ("Actuarial loss data", "[claims system / public filing]", "[valuation date]", "Paid, incurred, case reserve, claim counts"),
        ("Premium and exposure data", "[policy administration source]", "[period]", "Earned premium, rate, mix, exposure"),
        ("Reinsurance contracts", "[treaty / facultative documentation]", "[date]", "Limits, attachment, reinstatements, credit risk"),
        ("Capital methodology", "[regulator / internal methodology URL]", "[effective date]", "Risk factors, diversification, management actions"),
        ("Investment portfolio", "[asset ledger / filing]", "[date]", "Yield, duration, credit, liquidity, unrealized gains"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("INSURANCE_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
