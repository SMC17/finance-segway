"""Build a governed, data-grounded credit memo deck for a real public
Private Credit instance.

Second domain on the IC-memo/pptx system, proving pptx_helpers.py and the
data-grounding discipline generalize past the LBO/PE flagship: every number
on a slide is read from the instance workbook after a real LibreOffice
recalculation (tools.recalc), or from the case's own governed manifest
(standards/public_cases/<case_id>.json), which records which inputs are
real, sourced observations (input_kind "observed"/"derived") versus
modeler-chosen assumptions. Nothing on any slide is invented here -- this
script only reads and renders.

Usage:
    python tools/builders/build_credit_memo.py --case-id credit-public-ares-2024
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "tools"))

import pptx_helpers as ph  # noqa: E402
from recalc import recalc  # noqa: E402

import openpyxl  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
INDEX = ROOT / "standards" / "public_cases" / "index.json"


def _load_case(case_id: str) -> dict[str, Any]:
    index = json.loads(INDEX.read_text(encoding="utf-8"))
    for item in index["cases"]:
        if item["case_id"] == case_id:
            return item
    raise KeyError(f"unknown case_id {case_id!r}")


def _recalculated_workbook(output_rel: str) -> openpyxl.Workbook:
    src = ROOT / output_rel
    with tempfile.TemporaryDirectory(dir=ROOT) as tmp:
        scratch = Path(tmp) / src.name
        shutil.copyfile(src, scratch)
        result = recalc(str(scratch), timeout=90)
        if result.get("status") != "success" or result.get("total_errors", 1):
            raise RuntimeError(f"recalculation did not succeed cleanly: {result}")
        return openpyxl.load_workbook(scratch, data_only=True)


def _manifest_sources(case_id: str) -> dict[str, Any]:
    manifest_path = ROOT / "standards" / "public_cases" / f"{case_id}.json"
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def _real_inputs(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    return [item for item in manifest.get("inputs", []) if item.get("input_kind") in {"observed", "derived"}]


def _fmt_usd_mm(value: float) -> str:
    return f"${value:,.0f}mm"


def _fmt_pct(value: float, decimals: int = 1) -> str:
    return f"{value * 100:.{decimals}f}%"


def _fmt_x(value: float, decimals: int = 2) -> str:
    return f"{value:.{decimals}f}x"


def build_credit_memo(case_id: str, output: Path) -> None:
    case = _load_case(case_id)
    manifest = _manifest_sources(case_id)
    real_inputs = _real_inputs(manifest)
    real_source_name = real_inputs[0]["source"]["name"] if real_inputs else "unsourced"
    real_source_url = real_inputs[0]["source"]["url"] if real_inputs else ""
    real_cells = {item["cell"] for item in real_inputs}

    wb = _recalculated_workbook(case["output"])
    portfolio = wb["Portfolio & Concentration"]
    covenants = wb["Covenants"]
    yield_sheet = wb["Yield & Spread"]
    recovery = wb["Recovery"]
    checks = wb["Checks"]

    as_of = case["receipt"]["as_of"]

    borrower_rows = [
        (portfolio.cell(r, 2).value, portfolio.cell(r, 3).value, portfolio.cell(r, 5).value)
        for r in range(5, 10)
    ]
    total_exposure = portfolio["C12"].value
    top1_concentration = portfolio["C13"].value
    weighted_leverage = portfolio["C14"].value
    max_concentration = portfolio["C15"].value

    years = [f"Year {c - 2}" for c in range(3, 8)]
    leverage_headroom = [covenants.cell(7, c).value for c in range(3, 8)]
    coverage_headroom = [covenants.cell(10, c).value for c in range(3, 8)]
    dscr_headroom = [covenants.cell(13, c).value for c in range(3, 8)]
    covenant_status_y1 = covenants["C14"].value
    gross_leverage_y1 = covenants["C5"].value
    max_leverage = covenants["C6"].value

    cash_coupon = yield_sheet["C5"].value
    ytm = yield_sheet["C9"].value
    cash_spread_bps = yield_sheet["C11"].value
    maturity_years = yield_sheet["C8"].value

    recovery_rate_base_val = recovery["C12"].value
    recovery_rate_downside_val = recovery["D12"].value
    lgd_base = recovery["C13"].value
    lgd_downside = recovery["D13"].value
    recovery_multiple_base = recovery["C6"].value
    recovery_multiple_downside = recovery["D6"].value

    overall_status = checks["C9"].value

    prs = ph.new_presentation()

    # 1. Title
    ph.add_title_slide(
        prs,
        "Ares Capital Corporation — Portfolio Credit Memo",
        "Real FY2024 portfolio composition and exposure, illustrative structuring, "
        "covenant, and recovery terms. Not Ares Capital's own underwriting or covenant package.",
        f"As of {as_of}  ·  Source: {real_source_name}  ·  Prepared from a governed, recalculated model instance",
    )

    # 2. Provenance
    slide = ph.add_content_slide(prs, "Reading this deck", kicker="Provenance")
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(1.5), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.2),
        [
            f"Real, sourced: the five largest borrower/segment exposures ({', '.join(sorted(real_cells))}) "
            f"are Ares Capital's own disclosed FY2024 portfolio fair values, from {real_source_name}.",
            "Illustrative: per-borrower leverage multiples, covenant thresholds, yield/spread terms, and "
            "the recovery bridge are template defaults chosen for this exercise, not Ares Capital's actual "
            "underwriting, pricing, or covenant package on any of these positions.",
            "Every figure past this slide is either read directly from the recalculated workbook or "
            "explicitly labeled illustrative — nothing on this deck is invented independent of that model.",
        ],
        size=16,
    )
    ph.add_footer(slide, f"Model checks: Overall {overall_status}  ·  standards/public_cases/{case_id}.json")

    # 3. Portfolio snapshot (real)
    slide = ph.add_content_slide(prs, "Portfolio snapshot", kicker="Real, SEC-sourced")
    ph.add_stat_row(
        slide, ph.Inches(1.5),
        [
            ph.Stat(_fmt_usd_mm(total_exposure), "Total portfolio exposure", tag=real_source_name, color=ph.NAVY),
            ph.Stat(_fmt_pct(top1_concentration), "Top-1 borrower concentration", tag=real_source_name,
                    color=ph.NEGATIVE if top1_concentration > max_concentration else ph.NAVY),
            ph.Stat(_fmt_x(weighted_leverage), "Weighted portfolio leverage", tag="Illustrative", color=ph.ILLUSTRATIVE_TAG),
        ],
    )
    ph.add_table(
        slide, ph.MARGIN, ph.Inches(3.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.9),
        ["Borrower / segment", "Exposure ($mm)", "Leverage (illustrative)"],
        [
            [name, f"{exposure:,.1f}", f"{leverage:.1f}x"]
            for name, exposure, leverage in borrower_rows
        ],
        col_widths=[2.2, 1, 1],
    )
    ph.add_footer(slide, f"Source: {real_source_name} — {real_source_url}")

    # 4. Covenant headroom trend
    slide = ph.add_content_slide(prs, "Covenant headroom", kicker="Computed, illustrative covenant package")
    ph.add_stat_row(
        slide, ph.Inches(1.5),
        [
            ph.Stat(_fmt_x(gross_leverage_y1), "Gross leverage (Year 1)", tag="Illustrative", color=ph.NAVY),
            ph.Stat(_fmt_x(max_leverage), "Maximum leverage covenant", tag="Illustrative", color=ph.ILLUSTRATIVE_TAG),
            ph.Stat(covenant_status_y1, "Year 1 covenant status", tag="Computed",
                    color=ph.POSITIVE if covenant_status_y1 == "PASS" else ph.NEGATIVE),
        ],
    )
    ph.add_line_chart(
        slide, ph.MARGIN, ph.Inches(3.3), ph.Inches(6.0), ph.Inches(3.1),
        "Leverage & coverage headroom (turns)", years,
        [("Leverage headroom", leverage_headroom), ("Coverage headroom", coverage_headroom)],
    )
    ph.add_line_chart(
        slide, ph.Inches(7.0), ph.Inches(3.3), ph.Inches(6.0), ph.Inches(3.1),
        "DSCR headroom (turns)", years, [("DSCR headroom", dscr_headroom)],
    )

    # 5. Yield & spread economics
    slide = ph.add_content_slide(prs, "Lender yield & spread", kicker="Illustrative structuring")
    ph.add_stat_row(
        slide, ph.Inches(1.6),
        [
            ph.Stat(_fmt_pct(cash_coupon), "Cash coupon", tag="Illustrative", color=ph.ILLUSTRATIVE_TAG),
            ph.Stat(f"{cash_spread_bps:,.0f} bps", "Cash spread", tag="Illustrative", color=ph.ILLUSTRATIVE_TAG),
            ph.Stat(_fmt_pct(ytm), "Approx. all-in yield to maturity", tag="Computed", color=ph.NAVY),
        ],
    )
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(3.4), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.4),
        f"Illustrative {maturity_years:.0f}-year term priced at a {cash_spread_bps:,.0f}bps cash spread with "
        "no PIK component and a modest OID -- a template default direct-lending structure, not Ares "
        "Capital's actual pricing on any position in this portfolio.",
        size=14, color=ph.CHARCOAL,
    )
    ph.add_disclosure_note(
        slide,
        "Yield, spread, and OID terms are template defaults applied to real portfolio exposure figures -- "
        "not Ares Capital's disclosed pricing on these or any other positions.",
    )

    # 6. Recovery & loss analysis
    slide = ph.add_content_slide(prs, "Recovery & loss analysis", kicker="Illustrative stress bridge")
    ph.add_table(
        slide, ph.MARGIN, ph.Inches(1.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.8),
        ["Bridge", "Base", "Downside"],
        [
            ["Recovery multiple (x EBITDA)", f"{recovery_multiple_base:.1f}x", f"{recovery_multiple_downside:.1f}x"],
            ["Net distributable value ($mm)", f"{recovery['C9'].value:,.1f}", f"{recovery['D9'].value:,.1f}"],
            ["Debt claim ($mm)", f"{recovery['C10'].value:,.1f}", f"{recovery['D10'].value:,.1f}"],
            ["Recovery rate", _fmt_pct(recovery_rate_base_val, 0), _fmt_pct(recovery_rate_downside_val, 0)],
            ["Loss given default (LGD)", _fmt_pct(lgd_base, 0), _fmt_pct(lgd_downside, 0)],
        ],
        col_widths=[2.4, 1, 1],
    )
    ph.add_disclosure_note(
        slide,
        "Recovery multiple, EV haircut, and resulting recovery/LGD are illustrative stress-case "
        "mechanics on the template's debt claim -- not a real workout outcome or Ares Capital's own "
        "loss experience on any position.",
    )

    # 7. Decision / checks
    slide = ph.add_content_slide(prs, "Portfolio credit view", kicker="Decision")
    verdict_color = ph.POSITIVE if top1_concentration <= max_concentration and overall_status == "PASS" else ph.NEGATIVE
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(1.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.2),
        f"Real FY2024 exposure of {_fmt_usd_mm(total_exposure)} across the five largest positions carries "
        f"{_fmt_pct(top1_concentration)} top-1 concentration against a {_fmt_pct(max_concentration)} "
        "internal limit, under the template's illustrative structuring and covenant package.",
        size=17, bold=True, color=verdict_color,
    )
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(2.9), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.6),
        [
            f"Weighted portfolio leverage of {_fmt_x(weighted_leverage)} (illustrative, per-borrower) "
            f"against a {_fmt_x(max_leverage)} covenant ceiling leaves headroom that widens every "
            "projected year as the illustrative amortization schedule delevers.",
            "Concentration, not leverage, is the binding real-data risk signal here: the top borrower's "
            "real disclosed exposure alone is close to the internal concentration limit.",
            "A real credit decision would need each borrower's actual covenant package and current "
            "financial performance — neither is asserted here; both are illustrative template levers.",
        ],
        size=15,
    )
    ph.add_footer(
        slide,
        f"Model: 05_Private_Credit/_template_CREDIT.xlsx  ·  Instance: {case['output']}  ·  "
        f"Declared maturity: M2  ·  Overall checks: {overall_status}",
    )

    ph.save(prs, output)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-id", default="credit-public-ares-2024")
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    output = args.output or ROOT / "05_Private_Credit" / "presentations" / f"{args.case_id}-credit-memo.pptx"
    build_credit_memo(args.case_id, output)
    print(f"saved {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
