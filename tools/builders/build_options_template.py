import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[UNDERLYING] — Options Model", [
    ("Underlying:", "[fill in]"),
    ("Option type covered:", "Equity / index / FX / commodity options"),
    ("Last refreshed:", "[date]"),
    ("Next expiry to watch:", "[date]"),
    ("Refresh cadence:", "Weekly (daily near expiry)"),
])

# ---------------- BLACK-SCHOLES PRICER ----------------
ws = wb.create_sheet("BS Pricer")
set_col_widths(ws, [4, 26, 14, 4, 26, 16])
ws["B2"] = "Black-Scholes Pricer"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Spot price (S)", 100, CUR2),
    ("Strike price (K)", 100, CUR2),
    ("Time to expiry (yrs, T)", 0.25, '0.0000'),
    ("Risk-free rate (r)", 0.045, PCT2),
    ("Dividend yield (q)", 0.0, PCT2),
    ("Implied volatility (sigma)", 0.30, PCT),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

# Named-ish cell refs: S=C5, K=C6, T=C7, r=C8, q=C9, sigma=C10
ws["E4"] = "Intermediate"; ws["E4"].font = BOLD; ws["E4"].fill = GRAY_FILL
ws["E5"] = "d1"
ws["F5"] = "=(LN(C5/C6)+(C8-C9+0.5*C10^2)*C7)/(C10*SQRT(C7))"
ws["E6"] = "d2"
ws["F6"] = "=F5-C10*SQRT(C7)"
ws["E7"] = "N(d1)"
ws["F7"] = "=NORMSDIST(F5)"
ws["E8"] = "N(d2)"
ws["F8"] = "=NORMSDIST(F6)"
ws["E9"] = "N(-d1)"
ws["F9"] = "=NORMSDIST(-F5)"
ws["E10"] = "N(-d2)"
ws["F10"] = "=NORMSDIST(-F6)"
for row in range(5, 11):
    ws.cell(row=row, column=6).number_format = '0.0000'
    ws.cell(row=row, column=6).border = BORDER

ws["B12"] = "Outputs"; ws["B12"].font = BOLD; ws["B12"].fill = GRAY_FILL
ws["B13"] = "Call price"
ws["C13"] = "=C5*EXP(-C9*C7)*F7-C6*EXP(-C8*C7)*F8"
ws["B14"] = "Put price"
ws["C14"] = "=C6*EXP(-C8*C7)*F10-C5*EXP(-C9*C7)*F9"
for row in (13, 14):
    ws.cell(row=row, column=3).font = BOLD
    ws.cell(row=row, column=3).number_format = CUR2
    ws.cell(row=row, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- GREEKS ----------------
ws = wb.create_sheet("Greeks")
set_col_widths(ws, [4, 20, 16, 16, 40])
ws["B2"] = "Greeks (linked to BS Pricer inputs)"; ws["B2"].font = TITLE
for i, h in enumerate(["", "Greek", "Call", "Put", "Interpretation"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 4)

S, K, T, rr, q, sig = "'BS Pricer'!$C$5", "'BS Pricer'!$C$6", "'BS Pricer'!$C$7", \
                       "'BS Pricer'!$C$8", "'BS Pricer'!$C$9", "'BS Pricer'!$C$10"
d1, d2 = "'BS Pricer'!$F$5", "'BS Pricer'!$F$6"
Nd1, Nd2, Nnd1, Nnd2 = "'BS Pricer'!$F$7", "'BS Pricer'!$F$8", "'BS Pricer'!$F$9", "'BS Pricer'!$F$10"

greek_rows = [
    ("Delta", f"=EXP(-{q}*{T})*{Nd1}", f"=-EXP(-{q}*{T})*{Nnd1}", "Price sensitivity to $1 move in underlying"),
    ("Gamma", f"=EXP(-{q}*{T})*(EXP(-({d1})^2/2)/SQRT(2*PI()))/({S}*{sig}*SQRT({T}))",
              f"=EXP(-{q}*{T})*(EXP(-({d1})^2/2)/SQRT(2*PI()))/({S}*{sig}*SQRT({T}))", "Rate of change of delta"),
    ("Vega (per 1% vol)", f"={S}*EXP(-{q}*{T})*(EXP(-({d1})^2/2)/SQRT(2*PI()))*SQRT({T})/100",
                          f"={S}*EXP(-{q}*{T})*(EXP(-({d1})^2/2)/SQRT(2*PI()))*SQRT({T})/100", "Sensitivity to 1pt vol change"),
    ("Theta (per day)", f"=(-{S}*(EXP(-({d1})^2/2)/SQRT(2*PI()))*{sig}*EXP(-{q}*{T})/(2*SQRT({T}))-{rr}*{K}*EXP(-{rr}*{T})*{Nd2}+{q}*{S}*EXP(-{q}*{T})*{Nd1})/365",
                        f"=(-{S}*(EXP(-({d1})^2/2)/SQRT(2*PI()))*{sig}*EXP(-{q}*{T})/(2*SQRT({T}))+{rr}*{K}*EXP(-{rr}*{T})*{Nnd2}-{q}*{S}*EXP(-{q}*{T})*{Nnd1})/365",
                        "Time decay per calendar day"),
    ("Rho (per 1%)", f"={K}*{T}*EXP(-{rr}*{T})*{Nd2}/100", f"=-{K}*{T}*EXP(-{rr}*{T})*{Nnd2}/100", "Sensitivity to 1pt rate change"),
]
r = 5
for label, call_f, put_f, note in greek_rows:
    ws.cell(row=r, column=2, value=label).font = BLACK
    ws.cell(row=r, column=3, value=call_f).number_format = '0.0000'
    ws.cell(row=r, column=4, value=put_f).number_format = '0.0000'
    ws.cell(row=r, column=5, value=note).font = ITALIC_GRAY
    for c in (3, 4):
        ws.cell(row=r, column=c).border = BORDER
    r += 1
ws.sheet_view.showGridLines = False

# ---------------- STRATEGY PAYOFFS ----------------
ws = wb.create_sheet("Strategy Payoffs")
set_col_widths(ws, [4, 14] + [12]*9)
ws["B2"] = "Payoff at Expiry — Long Straddle Example"; ws["B2"].font = TITLE
ws["B4"] = "Underlying price at expiry ->"; ws["B4"].font = BOLD
prices = list(range(70, 131, 10))
for i, p in enumerate(prices, start=3):
    c = ws.cell(row=4, column=i, value=p); c.font = BOLD; c.fill = GRAY_FILL; c.number_format = CUR

ws["B5"] = "Long call payoff"; ws["B6"] = "Long put payoff"; ws["B7"] = "Net payoff (straddle)"
ws["B8"] = "Less: premium paid"; ws["B9"] = "Net P&L"
for i, p in enumerate(prices, start=3):
    col = get_column_letter(i)
    ws.cell(row=5, column=i, value=f"=MAX({col}4-'BS Pricer'!$C$6,0)").number_format = CUR
    ws.cell(row=6, column=i, value=f"=MAX('BS Pricer'!$C$6-{col}4,0)").number_format = CUR
    ws.cell(row=7, column=i, value=f"={col}5+{col}6").number_format = CUR
    ws.cell(row=8, column=i, value="='BS Pricer'!$C$13+'BS Pricer'!$C$14").number_format = CUR
    ws.cell(row=9, column=i, value=f"={col}7-{col}8").font = BOLD
    ws.cell(row=9, column=i).number_format = CUR
for row in range(5, 10):
    for c in range(3, 12):
        ws.cell(row=row, column=c).border = BORDER
ws["B11"] = "Note: swap the payoff formulas per leg to model spreads, collars, condors, etc."
ws["B11"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

add_refresh_log(wb)

out_path = "/home/claude/model_shop/OPTIONS_template.xlsx"
wb.save(out_path)
print("saved", out_path)
