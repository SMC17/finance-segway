import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[COUNTERPARTY] — Trade Finance Model", [
    ("Counterparty:", "[fill in]"),
    ("Trade corridor:", "[e.g. exporter-importer country pair]"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly"),
])

# ---------------- WORKING CAPITAL CYCLE ----------------
ws = wb.create_sheet("Working Capital Cycle")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Cash Conversion Cycle"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Revenue ($, annual)", 0, CUR),
    ("COGS ($, annual)", 0, CUR),
    ("Average accounts receivable ($)", 0, CUR),
    ("Average inventory ($)", 0, CUR),
    ("Average accounts payable ($)", 0, CUR),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B11"] = "Outputs"; ws["B11"].font = BOLD; ws["B11"].fill = GRAY_FILL
ws["B12"] = "DSO (Days Sales Outstanding)"
ws["C12"] = "=IFERROR(C7/C5*365,\"-\")"; ws["C12"].number_format = '0.0'
ws["B13"] = "DIO (Days Inventory Outstanding)"
ws["C13"] = "=IFERROR(C8/C6*365,\"-\")"; ws["C13"].number_format = '0.0'
ws["B14"] = "DPO (Days Payable Outstanding)"
ws["C14"] = "=IFERROR(C9/C6*365,\"-\")"; ws["C14"].number_format = '0.0'
ws["B15"] = "Cash Conversion Cycle (DSO+DIO-DPO)"
ws["C15"] = "=IFERROR(C12+C13-C14,\"-\")"; ws["C15"].font = BOLD; ws["C15"].number_format = '0.0'
ws["D15"] = "Days cash tied up per cycle — lower/negative is better (supplier financing your growth)"
ws["D15"].font = ITALIC_GRAY
ws["B16"] = "Working capital required ($, approx.)"
ws["C16"] = "=IFERROR(C15/365*C6,\"-\")"; ws["C16"].number_format = CUR
for r2 in range(12, 17):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- LC & FACTORING COST ----------------
ws = wb.create_sheet("LC & Factoring Cost")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Letter of Credit & Factoring Cost Model"; ws["B2"].font = TITLE
ws["B4"] = "Letter of Credit"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
lc_inputs = [
    ("LC face value ($)", 0, CUR),
    ("Issuance fee %", 0.015, PCT),
    ("Confirmation fee %", 0.01, PCT),
    ("Tenor (days)", 90, NUM),
]
r = 5
for label, default, fmt in lc_inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B9"] = "Total LC cost ($)"
ws["C9"] = "=C5*(C6+C7)*(C8/365)"; ws["C9"].font = BOLD; ws["C9"].number_format = CUR
ws["B10"] = "Annualized LC cost %"
ws["C10"] = "=IFERROR(C9/C5/(C8/365),\"-\")"; ws["C10"].number_format = PCT
ws.cell(row=9, column=3).border = BORDER
ws.cell(row=10, column=3).border = BORDER

ws["B13"] = "Factoring / Invoice Discounting"; ws["B13"].font = BOLD; ws["B13"].fill = GRAY_FILL
fact_inputs = [
    ("Invoice face value ($)", 0, CUR),
    ("Advance rate %", 0.85, PCT),
    ("Discount/factor fee %", 0.02, PCT),
    ("Days to collection", 60, NUM),
]
r = 14
for label, default, fmt in fact_inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B18"] = "Cash advanced immediately ($)"
ws["C18"] = "=C14*C15"; ws["C18"].number_format = CUR
ws["B19"] = "Factoring fee ($)"
ws["C19"] = "=C14*C16"; ws["C19"].number_format = CUR
ws["B20"] = "Net proceeds after collection ($)"
ws["C20"] = "=C14-C19"; ws["C20"].number_format = CUR
ws["B21"] = "Effective annualized cost of factoring %"
ws["C21"] = "=IFERROR(C16/(C17/365),\"-\")"; ws["C21"].font = BOLD; ws["C21"].number_format = PCT
ws["D21"] = "Compare vs. bank credit line rate to decide financing channel"
ws["D21"].font = ITALIC_GRAY
for r2 in range(18, 22):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- FINANCING COST COMPARISON ----------------
ws = wb.create_sheet("Financing Cost Comparison")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Which Financing Channel Is Cheapest?"; ws["B2"].font = TITLE

ws["B4"] = "Revolving credit line (alternative)"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Draw amount ($)", 0, CUR),
    ("Base rate + spread, all-in (%)", 0.08, PCT2),
    ("Commitment fee on undrawn (%)", 0.005, PCT2),
    ("Undrawn facility amount ($)", 0, CUR),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B9"] = "Revolver interest cost ($)"
ws["C9"] = "=C5*C6"; ws["C9"].number_format = CUR; ws["C9"].border = BORDER
ws["B10"] = "Revolver commitment fee ($)"
ws["C10"] = "=C7*C8"; ws["C10"].number_format = CUR; ws["C10"].border = BORDER
ws["B11"] = "Revolver annualized all-in cost %"
ws["C11"] = "=IFERROR((C9+C10)/C5,\"-\")"; ws["C11"].number_format = PCT; ws["C11"].border = BORDER

ws["B13"] = "Comparison"; ws["B13"].font = BOLD; ws["B13"].fill = GRAY_FILL
for i, h in enumerate(["", "Channel", "Annualized cost %", ""], start=1):
    ws.cell(row=14, column=i, value=h)
style_header_row(ws, 14, 2, start_col=2)
ws["B15"] = "Letter of credit"
ws["C15"] = "='LC & Factoring Cost'!C10"; ws["C15"].font = GREEN; ws["C15"].number_format = PCT
ws["B16"] = "Factoring / invoice discounting"
ws["C16"] = "='LC & Factoring Cost'!C21"; ws["C16"].font = GREEN; ws["C16"].number_format = PCT
ws["B17"] = "Revolving credit line"
ws["C17"] = "=C11"; ws["C17"].font = GREEN; ws["C17"].number_format = PCT
for r2 in (15, 16, 17):
    ws.cell(row=r2, column=3).border = BORDER

ws["B19"] = "Cheapest channel"
ws["C19"] = '=IF(COUNT(C15:C17)<3,"-",INDEX(B15:B17,MATCH(MIN(C15:C17),C15:C17,0)))'
ws["C19"].font = BOLD; ws["C19"].border = BORDER
ws["B20"] = "Cheapest annualized cost %"
ws["C20"] = "=IF(COUNT(C15:C17)<3,\"-\",MIN(C15:C17))"; ws["C20"].font = BOLD; ws["C20"].number_format = PCT
ws["C20"].border = BORDER
ws["D19"] = "Only declares a winner once all three channels have real numbers — one live rate shouldn't beat two blanks by default"
ws["D19"].font = ITALIC_GRAY
ws["D20"] = "Compares only financing cost — also weigh speed, collateral/covenant burden, and counterparty credit risk transfer (LC shifts payment risk to the issuing bank; factoring can be non-recourse)"
ws["D20"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "TRADE_FINANCE_template.xlsx"
wb.save(out_path)
print("saved", out_path)
