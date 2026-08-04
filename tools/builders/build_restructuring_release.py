"""Release-grade restructuring workbook with 13-week liquidity and priority controls."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl.utils import get_column_letter

try:
    from tools.builders.legacy_release_adapter import build_release
    from tools.builders.template_helpers import (
        BLUE,
        BOLD,
        BORDER,
        CUR,
        GRAY_FILL,
        ITALIC_GRAY,
        MULT,
        NUM,
        PCT,
        TITLE,
        YELLOW_FILL,
        set_col_widths,
        style_header_row,
    )
except ModuleNotFoundError:
    from legacy_release_adapter import build_release
    from template_helpers import (
        BLUE,
        BOLD,
        BORDER,
        CUR,
        GRAY_FILL,
        ITALIC_GRAY,
        MULT,
        NUM,
        PCT,
        TITLE,
        YELLOW_FILL,
        set_col_widths,
        style_header_row,
    )


def _overall(check_range: str) -> str:
    return f'=IF(COUNTIF({check_range},"FAIL")+COUNTIF({check_range},"BREACH")>0,"BREACH",IF(COUNTIF({check_range},"REVIEW")>0,"REVIEW","PASS"))'


def enrich(workbook) -> None:
    for name in ("13-Week Liquidity", "New Money", "Decision & Checks"):
        if name in workbook.sheetnames:
            del workbook[name]

    recovery = workbook["Recovery Waterfall"]
    liquidity_position = workbook.sheetnames.index("Recovery Waterfall") + 1
    liquidity = workbook.create_sheet("13-Week Liquidity", liquidity_position)
    set_col_widths(liquidity, [4, 12, 16, 16, 18, 14, 18, 16, 16, 16, 14])
    liquidity["B2"] = "13-Week Cash Flow, Minimum Liquidity, and Funding Need"
    liquidity["B2"].font = TITLE
    liquidity["B3"] = "Initial unrestricted liquidity"
    liquidity["C3"] = 20.0
    liquidity["C3"].font = BLUE
    liquidity["C3"].fill = YELLOW_FILL
    liquidity["C3"].number_format = CUR
    liquidity["C3"].border = BORDER
    liquidity["B4"] = "Minimum operating liquidity"
    liquidity["C4"] = 5.0
    liquidity["C4"].font = BLUE
    liquidity["C4"].fill = YELLOW_FILL
    liquidity["C4"].number_format = CUR
    liquidity["C4"].border = BORDER
    headers = ["Week", "Beginning cash", "Receipts", "Operating disbursements", "Interest", "Professional fees", "New money", "Ending cash", "Headroom", "Status"]
    for column, value in enumerate(headers, start=2):
        liquidity.cell(6, column, value)
    style_header_row(liquidity, 6, len(headers), start_col=2)
    for week in range(1, 14):
        row = 6 + week
        liquidity.cell(row, 2, week)
        liquidity.cell(row, 3, "=$C$3" if week == 1 else f"=I{row-1}")
        for column, value in (
            (4, 8.0),
            (5, 7.0),
            (6, 0.75),
            (7, 0.50),
            (8, 0.0),
        ):
            cell = liquidity.cell(row, column, value)
            cell.font = BLUE
            cell.fill = YELLOW_FILL
            cell.number_format = CUR
            cell.border = BORDER
        liquidity.cell(row, 9, f"=C{row}+D{row}+H{row}-E{row}-F{row}-G{row}")
        liquidity.cell(row, 10, f"=I{row}-$C$4")
        liquidity.cell(row, 11, f'=IF(J{row}>=0,"PASS","BREACH")')
        for column in (3, 9, 10):
            liquidity.cell(row, column).number_format = CUR
            liquidity.cell(row, column).border = BORDER
        liquidity.cell(row, 11).border = BORDER
    liquidity["B22"] = "Minimum ending cash"
    liquidity["C22"] = "=MIN(I7:I19)"
    liquidity["C22"].number_format = CUR
    liquidity["C22"].border = BORDER
    liquidity["B23"] = "Peak funding need"
    liquidity["C23"] = "=MAX(0,-MIN(J7:J19))"
    liquidity["C23"].number_format = CUR
    liquidity["C23"].border = BORDER
    liquidity["B24"] = "First breach week"
    liquidity["C24"] = '=IFERROR(MATCH("BREACH",K7:K19,0),"-")'
    liquidity["C24"].border = BORDER
    liquidity["B26"] = "A 13-week schedule is an operating control: replace the illustrative weekly flows with treasury-owned receipts, disbursements, financing availability, and variance explanations."
    liquidity["B26"].font = ITALIC_GRAY
    liquidity.freeze_panes = "B7"
    liquidity.sheet_view.showGridLines = False

    new_money = workbook.create_sheet("New Money", liquidity_position + 1)
    set_col_widths(new_money, [4, 40, 18, 18, 44])
    new_money["B2"] = "DIP / Rescue Financing Economics and Priority"
    new_money["B2"].font = TITLE
    inputs = [
        ("New-money commitment", 25.0, CUR, "committed financing"),
        ("Expected draw", "='13-Week Liquidity'!C23", CUR, "peak funding need"),
        ("Upfront fee", 0.02, PCT, "% of drawn amount"),
        ("Annual cash interest rate", 0.12, PCT, "contractual cash coupon"),
        ("Tenor", 1.0, "0.0", "years"),
        ("Exit fee", 0.03, PCT, "% of drawn amount"),
        ("Super-priority / administrative claim", 1.0, MULT, "1.0x if fully senior to prepetition claims"),
    ]
    for row, (label, value, number_format, note) in enumerate(inputs, start=5):
        new_money.cell(row, 2, label)
        cell = new_money.cell(row, 3, value)
        cell.font = BLUE if not isinstance(value, str) else cell.font
        if not isinstance(value, str):
            cell.fill = YELLOW_FILL
        cell.number_format = number_format
        cell.border = BORDER
        new_money.cell(row, 4, note)
    new_money["B14"] = "Undrawn commitment"
    new_money["C14"] = "=MAX(0,C5-C6)"
    new_money["B15"] = "Upfront fee amount"
    new_money["C15"] = "=C6*C7"
    new_money["B16"] = "Cash interest"
    new_money["C16"] = "=C6*C8*C9"
    new_money["B17"] = "Exit fee amount"
    new_money["C17"] = "=C6*C10"
    new_money["B18"] = "Total new-money claim at exit"
    new_money["C18"] = "=C6+C15+C16+C17"
    new_money["B19"] = "Commitment sufficiency"
    new_money["C19"] = '=IF(C5>=C6,"PASS","BREACH")'
    for row in range(14, 20):
        new_money["C" + str(row)].border = BORDER
        if row < 19:
            new_money["C" + str(row)].number_format = CUR
    new_money["B21"] = "Priority, roll-up, milestones, collateral, avoidance risk, and adequate-protection terms require legal review and cannot be inferred from economic seniority alone."
    new_money["B21"].font = ITALIC_GRAY
    new_money.sheet_view.showGridLines = False

    checks_position = workbook.sheetnames.index("Liquidation vs Reorg") + 1
    checks = workbook.create_sheet("Decision & Checks", checks_position)
    set_col_widths(checks, [4, 40, 18, 18, 50])
    checks["B2"] = "Distressed / Restructuring Decision Dashboard and Independent Checks"
    checks["B2"].font = TITLE
    for column, value in enumerate(["Check / decision", "Metric", "Status", "Interpretation / action"], start=2):
        checks.cell(4, column, value)
    style_header_row(checks, 4, 4, start_col=2)
    rows = [
        (5, "Waterfall conservation residual", "='Recovery Waterfall'!C12-SUM('Recovery Waterfall'!E5:E10)", '=IF(ABS(C5)<0.01,"PASS","FAIL")', "Distributed value must equal enterprise value available under the modeled waterfall."),
        (6, "Priority ordering", "=MIN('Recovery Waterfall'!F5-'Recovery Waterfall'!F6,'Recovery Waterfall'!F6-'Recovery Waterfall'!F7,'Recovery Waterfall'!F7-'Recovery Waterfall'!F8,'Recovery Waterfall'!F8-'Recovery Waterfall'!F9)", '=IF(C6>=-0.000001,"PASS","FAIL")', "Junior recoveries may not exceed an impaired senior class under the absolute-priority model."),
        (7, "Fulcrum security", "='Recovery Waterfall'!C21", '=IF(C7="-","REVIEW","PASS")', "The impaired class immediately below fully paid senior claims is the modeled fulcrum."),
        (8, "Minimum 13-week liquidity", "='13-Week Liquidity'!C22", '=IF(C8>=\'13-Week Liquidity\'!C4,"PASS","BREACH")', "Escalate a minimum-liquidity breach before the modeled cash-out date."),
        (9, "Peak funding need", "='13-Week Liquidity'!C23", '=IF(C9<=\'New Money\'!C5,"PASS","BREACH")', "New-money commitment must cover peak need plus execution contingency."),
        (10, "Reorganization NPV uplift", "='Liquidation vs Reorg'!D12-'Liquidation vs Reorg'!C12", '=IF(C10>=0,"PASS","REVIEW")', "Reorganization should not be preferred without a supportable value advantage and feasible plan."),
        (11, "New-money commitment status", "='New Money'!C19", '=IF(C11="PASS","PASS","BREACH")', "Insufficient rescue financing requires a revised operating plan, asset sale, or liquidation path."),
        (12, "Equity recovery", "='Recovery Waterfall'!E10", '=IF(C12>0,"REVIEW","PASS")', "Equity value while a creditor class is impaired requires explicit exception analysis."),
    ]
    for row, label, formula, status, action in rows:
        checks.cell(row, 2, label)
        checks.cell(row, 3, formula)
        checks.cell(row, 3).border = BORDER
        checks.cell(row, 4, status)
        checks.cell(row, 4).border = BORDER
        checks.cell(row, 5, action)
    checks["B14"] = "Overall model status"
    checks["B14"].font = BOLD
    checks["C14"] = _overall("D5:D12")
    checks["C14"].font = BOLD
    checks["C14"].border = BORDER
    checks["B16"] = "Primary decision outputs"
    checks["B16"].font = BOLD
    checks["B16"].fill = GRAY_FILL
    outputs = [
        (17, "Enterprise value available", "='Recovery Waterfall'!C12", CUR),
        (18, "Fulcrum security", "='Recovery Waterfall'!C21", "General"),
        (19, "Liquidation NPV", "='Liquidation vs Reorg'!C12", CUR),
        (20, "Reorganization NPV", "='Liquidation vs Reorg'!D12", CUR),
        (21, "Peak funding need", "='13-Week Liquidity'!C23", CUR),
        (22, "Total new-money claim", "='New Money'!C18", CUR),
        (23, "First liquidity breach week", "='13-Week Liquidity'!C24", NUM),
    ]
    for row, label, formula, number_format in outputs:
        checks.cell(row, 2, label)
        checks.cell(row, 3, formula)
        checks.cell(row, 3).number_format = number_format
        checks.cell(row, 3).border = BORDER
    checks.freeze_panes = "B5"
    checks.sheet_view.showGridLines = False


def build(output: Path) -> None:
    build_release("build_restructuring_template.py", output, enrich)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("RESTRUCTURING_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
