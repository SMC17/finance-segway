"""Build the canonical software / SaaS operating and unit-economics model.

This is a *sector* model, not another generic archetype. The point of the
coverage-group taxonomy is that a software business and an industrial
business are not the same model with different inputs -- their value drivers
differ structurally, and a generic three-statement template silently hides
that.

What a software model must do that the BASE archetype does not:

  - Roll ARR forward as a cohort balance (beginning + new + expansion -
    contraction - churn), because revenue is a *consequence* of that
    balance rather than an independent growth assumption. Net revenue
    retention falls out of the roll-forward instead of being asserted.
  - Separate subscription from services revenue: they carry very different
    gross margins, and blended margin moves with mix even when neither
    component changes.
  - Treat sales & marketing as an investment with a payback period (CAC
    payback, magic number) rather than as an operating expense ratio. A
    company can look unprofitable on GAAP while generating strong returns
    on customer acquisition, and the model has to be able to show that.
  - Report stock-based compensation explicitly. SBC is routinely 10-25% of
    revenue for listed software companies -- large enough that a model
    which buries it inside opex is not decision-useful.
  - Compute Rule of 40, the sector's standard growth-versus-profitability
    trade-off, from the model rather than as a quoted statistic.
  - Roll remaining performance obligations forward alongside ARR. RPO is the
    one forward-revenue quantity software companies actually file under XBRL
    (RevenueRemainingPerformanceObligation), so it is the only piece of the
    contracted-future-revenue picture that can be sourced rather than
    assumed. Bookings fall out of the RPO roll-forward as a residual, and
    splitting RPO against the contract liability separates the billed
    portion (deferred revenue, on balance sheet) from the unbilled portion
    (contracted but not yet invoiced, off balance sheet) -- a distinction
    that materially changes how much of "backlog" is real cash timing.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    header, input_cell, section, set_widths, title, total_row,
)

YEARS = 5
SHEETS = [
    "Cover", "Assumptions", "ARR Rollforward", "Operating Model",
    "RPO & Bookings", "Unit Economics", "Rule of 40", "Checks",
    "Sources", "RefreshLog",
]


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[COMPANY] — Software / SaaS Operating and Unit-Economics Model", [
        ("Company / ticker:", "[Name]"),
        ("Subsector:", "[application | infrastructure | vertical SaaS]"),
        ("Fiscal year end:", "[date]"),
        ("Last refreshed:", "[date]"),
        ("Next filing / refresh:", "[date]"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Software Operating Drivers")
    header(sheet, 4, 2, ["Driver", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Beginning ARR", 1000.0, 1000.0, "$mm", CUR),
        ("New ARR added (% of beginning ARR)", 0.220, 0.120, "%", PCT),
        ("Expansion ARR (% of beginning ARR)", 0.140, 0.080, "%", PCT),
        ("Contraction ARR (% of beginning ARR)", 0.040, 0.070, "%", PCT),
        ("Gross churn (% of beginning ARR)", 0.080, 0.130, "%", PCT),
        ("Services revenue (% of subscription revenue)", 0.100, 0.100, "%", PCT),
        ("Subscription gross margin", 0.800, 0.760, "%", PCT),
        ("Services gross margin", 0.150, 0.050, "%", PCT),
        ("Sales & marketing (% of revenue)", 0.380, 0.420, "%", PCT),
        ("Research & development (% of revenue)", 0.220, 0.240, "%", PCT),
        ("General & administrative (% of revenue)", 0.120, 0.130, "%", PCT),
        ("Stock-based compensation (% of revenue)", 0.180, 0.200, "%", PCT),
        ("Capex (% of revenue)", 0.030, 0.030, "%", PCT),
        ("Cash tax rate", 0.210, 0.210, "%", PCT),
        ("Gross margin on new ARR (for CAC payback)", 0.800, 0.760, "%", PCT),
        ("Beginning remaining performance obligation (RPO)", 950.0, 950.0, "$mm", CUR),
        ("RPO coverage of revenue (ending RPO / revenue)", 0.950, 0.850, "x", MULT),
        ("Current portion of RPO (recognised within 12 months)", 0.640, 0.600, "%", PCT),
        ("Contract liability / deferred revenue (% of revenue)", 0.300, 0.300, "%", PCT),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 48, "C": 15, "D": 15, "E": 15, "F": 32})
    sheet.freeze_panes = "A5"

    # --- ARR roll-forward -------------------------------------------------
    sheet = workbook["ARR Rollforward"]
    title(sheet, "B2:H2", "Annual Recurring Revenue Roll-Forward")
    header(sheet, 4, 2, ["$mm", *[f"Year {year}" for year in range(1, YEARS + 1)]])
    labels = [
        "Beginning ARR", "+ New ARR", "+ Expansion ARR", "- Contraction ARR",
        "- Gross churn", "Ending ARR", "Net new ARR",
        "Net revenue retention", "Gross revenue retention", "ARR growth",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)

    for column in range(3, 3 + YEARS):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)
        if column == 3:
            sheet.cell(5, column, "=Assumptions!$E$5")
        else:
            sheet.cell(5, column, f"={previous}10")
        sheet.cell(6, column, f"={letter}5*Assumptions!$E$6")
        sheet.cell(7, column, f"={letter}5*Assumptions!$E$7")
        sheet.cell(8, column, f"={letter}5*Assumptions!$E$8")
        sheet.cell(9, column, f"={letter}5*Assumptions!$E$9")
        sheet.cell(10, column, f"={letter}5+{letter}6+{letter}7-{letter}8-{letter}9")
        sheet.cell(11, column, f"={letter}10-{letter}5")
        # NRR and GRR are computed from the roll-forward, not asserted:
        # NRR excludes new logos by construction, GRR excludes expansion too.
        sheet.cell(12, column, f"=IFERROR(({letter}5+{letter}7-{letter}8-{letter}9)/{letter}5,0)")
        sheet.cell(13, column, f"=IFERROR(({letter}5-{letter}9)/{letter}5,0)")
        sheet.cell(14, column, f"=IFERROR({letter}10/{letter}5-1,0)")
        for row in range(5, 12):
            sheet.cell(row, column).number_format = CUR
        for row in (12, 13, 14):
            sheet.cell(row, column).number_format = PCT
    total_row(sheet, 10, 2, 2 + YEARS, CUR)
    set_widths(sheet, {"A": 4, "B": 34, **{get_column_letter(c): 14 for c in range(3, 3 + YEARS)}})
    sheet.freeze_panes = "C5"

    # --- Operating model --------------------------------------------------
    sheet = workbook["Operating Model"]
    title(sheet, "B2:H2", "Revenue, Margin, and Cash Flow")
    header(sheet, 4, 2, ["$mm", *[f"Year {year}" for year in range(1, YEARS + 1)]])
    labels = [
        "Subscription revenue", "Services revenue", "Total revenue",
        "Subscription gross profit", "Services gross profit", "Total gross profit",
        "Blended gross margin", "Sales & marketing", "Research & development",
        "General & administrative", "Total operating expense", "GAAP operating income",
        "GAAP operating margin", "Stock-based compensation",
        "Non-GAAP operating income (ex-SBC)", "Non-GAAP operating margin",
        "Cash taxes", "Capex", "Free cash flow", "FCF margin",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)

    for column in range(3, 3 + YEARS):
        letter = get_column_letter(column)
        # Subscription revenue approximates the average ARR balance over the
        # year rather than the ending balance: recognising ending ARR as
        # revenue would overstate every growing year.
        sheet.cell(5, column, f"=AVERAGE('ARR Rollforward'!{letter}5,'ARR Rollforward'!{letter}10)")
        sheet.cell(6, column, f"={letter}5*Assumptions!$E$10")
        sheet.cell(7, column, f"={letter}5+{letter}6")
        sheet.cell(8, column, f"={letter}5*Assumptions!$E$11")
        sheet.cell(9, column, f"={letter}6*Assumptions!$E$12")
        sheet.cell(10, column, f"={letter}8+{letter}9")
        sheet.cell(11, column, f"=IFERROR({letter}10/{letter}7,0)")
        sheet.cell(12, column, f"={letter}7*Assumptions!$E$13")
        sheet.cell(13, column, f"={letter}7*Assumptions!$E$14")
        sheet.cell(14, column, f"={letter}7*Assumptions!$E$15")
        sheet.cell(15, column, f"={letter}12+{letter}13+{letter}14")
        sheet.cell(16, column, f"={letter}10-{letter}15")
        sheet.cell(17, column, f"=IFERROR({letter}16/{letter}7,0)")
        sheet.cell(18, column, f"={letter}7*Assumptions!$E$16")
        sheet.cell(19, column, f"={letter}16+{letter}18")
        sheet.cell(20, column, f"=IFERROR({letter}19/{letter}7,0)")
        sheet.cell(21, column, f"=MAX(0,{letter}16)*Assumptions!$E$18")
        sheet.cell(22, column, f"={letter}7*Assumptions!$E$17")
        # SBC is added back as a non-cash charge; it is still dilution, which
        # is why it stays visible on its own line above.
        sheet.cell(23, column, f"={letter}16+{letter}18-{letter}21-{letter}22")
        sheet.cell(24, column, f"=IFERROR({letter}23/{letter}7,0)")
        for row in (5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 18, 19, 21, 22, 23):
            sheet.cell(row, column).number_format = CUR
        for row in (11, 17, 20, 24):
            sheet.cell(row, column).number_format = PCT
    for row in (7, 10, 15, 23):
        total_row(sheet, row, 2, 2 + YEARS, CUR)
    set_widths(sheet, {"A": 4, "B": 38, **{get_column_letter(c): 14 for c in range(3, 3 + YEARS)}})
    sheet.freeze_panes = "C5"

    # --- RPO and bookings -------------------------------------------------
    # ARR is a management metric: no company files it, and the definitions
    # differ enough between issuers that comparing two companies' ARR is
    # often meaningless. RPO is the opposite -- it is a required disclosure
    # under ASC 606 with a fixed definition, and it is tagged in XBRL. So
    # this sheet is the part of the forward-revenue picture that an instance
    # can actually ground in a filing, and it is deliberately driven off the
    # roll-forward identity rather than off an assumed bookings number:
    #   ending RPO = beginning RPO + gross bookings - revenue recognised
    # Ending RPO is set from a coverage ratio and BOOKINGS is the residual,
    # because coverage is observable from consecutive filings while gross
    # bookings is never disclosed by anyone.
    sheet = workbook["RPO & Bookings"]
    title(sheet, "B2:H2", "Remaining Performance Obligations and Implied Bookings")
    header(sheet, 4, 2, ["$mm", *[f"Year {year}" for year in range(1, YEARS + 1)]])
    labels = [
        "Revenue recognised", "Beginning RPO", "Ending RPO",
        "Implied gross bookings (residual)", "Book-to-bill", "RPO growth",
        "Current RPO (recognised within 12 months)", "Non-current RPO",
        "Contract liability (deferred revenue)", "Unbilled RPO",
        "Unbilled share of RPO", "Current RPO / following-year revenue",
        "Ending RPO / ending ARR",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)

    last_column = 2 + YEARS
    for column in range(3, 3 + YEARS):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)
        sheet.cell(5, column, f"='Operating Model'!{letter}7")
        if column == 3:
            sheet.cell(6, column, "=Assumptions!$E$20")
        else:
            sheet.cell(6, column, f"={previous}7")
        sheet.cell(7, column, f"={letter}5*Assumptions!$E$21")
        # The residual. If a company's RPO grows faster than it recognises
        # revenue, it booked more than it burned through -- that is the whole
        # signal, and stating it as a residual keeps it honest.
        sheet.cell(8, column, f"={letter}7-{letter}6+{letter}5")
        sheet.cell(9, column, f"=IFERROR({letter}8/{letter}5,0)")
        sheet.cell(10, column, f"=IFERROR({letter}7/{letter}6-1,0)")
        sheet.cell(11, column, f"={letter}7*Assumptions!$E$22")
        sheet.cell(12, column, f"={letter}7-{letter}11")
        sheet.cell(13, column, f"={letter}5*Assumptions!$E$23")
        # Unbilled RPO is contracted revenue the company has NOT yet
        # invoiced, so it sits nowhere on the balance sheet. Deferred revenue
        # is the billed half. Reporting only total RPO conflates the two.
        sheet.cell(14, column, f"={letter}7-{letter}13")
        sheet.cell(15, column, f"=IFERROR({letter}14/{letter}7,0)")
        if column < last_column:
            following = get_column_letter(column + 1)
            sheet.cell(16, column, f"=IFERROR({letter}11/'Operating Model'!{following}7,0)")
            sheet.cell(16, column).number_format = MULT
        else:
            sheet.cell(16, column, "n/a")
        sheet.cell(17, column, f"=IFERROR({letter}7/'ARR Rollforward'!{letter}10,0)")
        for row in (5, 6, 7, 8, 11, 12, 13, 14):
            sheet.cell(row, column).number_format = CUR
        for row in (9, 17):
            sheet.cell(row, column).number_format = MULT
        for row in (10, 15):
            sheet.cell(row, column).number_format = PCT
    total_row(sheet, 7, 2, 2 + YEARS, CUR)
    set_widths(sheet, {"A": 4, "B": 44, **{get_column_letter(c): 14 for c in range(3, 3 + YEARS)}})
    sheet.freeze_panes = "C5"

    # --- Unit economics ---------------------------------------------------
    sheet = workbook["Unit Economics"]
    title(sheet, "B2:H2", "Customer Acquisition Efficiency")
    header(sheet, 4, 2, ["Metric", *[f"Year {year}" for year in range(1, YEARS + 1)]])
    labels = [
        "S&M spend", "New + expansion ARR", "Magic number",
        "CAC payback (months)", "Gross-profit-weighted CAC payback (months)",
        "S&M as % of new ARR",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 3 + YEARS):
        letter = get_column_letter(column)
        sheet.cell(5, column, f"='Operating Model'!{letter}12")
        sheet.cell(6, column, f"='ARR Rollforward'!{letter}6+'ARR Rollforward'!{letter}7")
        # Magic number: incremental ARR won per dollar of S&M. Above ~0.75 is
        # conventionally read as efficient enough to keep spending.
        sheet.cell(7, column, f"=IFERROR({letter}6/{letter}5,0)")
        sheet.cell(8, column, f"=IFERROR({letter}5/{letter}6*12,0)")
        # The gross-profit-weighted version is the one that actually matters:
        # payback should be measured on the margin the new ARR carries, not
        # on its top line.
        sheet.cell(9, column, f"=IFERROR({letter}5/({letter}6*Assumptions!$E$19)*12,0)")
        sheet.cell(10, column, f"=IFERROR({letter}5/{letter}6,0)")
        for row in (5, 6):
            sheet.cell(row, column).number_format = CUR
        sheet.cell(7, column).number_format = MULT
        for row in (8, 9):
            sheet.cell(row, column).number_format = '0.0'
        sheet.cell(10, column).number_format = PCT
    set_widths(sheet, {"A": 4, "B": 44, **{get_column_letter(c): 14 for c in range(3, 3 + YEARS)}})
    sheet.freeze_panes = "C5"

    # --- Rule of 40 -------------------------------------------------------
    sheet = workbook["Rule of 40"]
    title(sheet, "B2:H2", "Growth versus Profitability Trade-off")
    header(sheet, 4, 2, ["Metric", *[f"Year {year}" for year in range(1, YEARS + 1)]])
    labels = [
        "Revenue growth", "FCF margin", "Rule of 40 (growth + FCF margin)",
        "Non-GAAP operating margin", "Rule of 40 (growth + operating margin)",
        "Status",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 3 + YEARS):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)
        if column == 3:
            sheet.cell(5, column, "='ARR Rollforward'!C14")
        else:
            sheet.cell(5, column, f"=IFERROR('Operating Model'!{letter}7/'Operating Model'!{previous}7-1,0)")
        sheet.cell(6, column, f"='Operating Model'!{letter}24")
        sheet.cell(7, column, f"={letter}5+{letter}6")
        sheet.cell(8, column, f"='Operating Model'!{letter}20")
        sheet.cell(9, column, f"={letter}5+{letter}8")
        sheet.cell(10, column, f'=IF({letter}7>=0.4,"PASS",IF({letter}7>=0.3,"REVIEW","BELOW"))')
        for row in (5, 6, 7, 8, 9):
            sheet.cell(row, column).number_format = PCT
    add_status_rules(sheet, f"C10:{get_column_letter(2 + YEARS)}10")
    set_widths(sheet, {"A": 4, "B": 44, **{get_column_letter(c): 14 for c in range(3, 3 + YEARS)}})

    # --- Checks -----------------------------------------------------------
    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Software Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    last = get_column_letter(2 + YEARS)
    checks = [
        ("ARR roll-forward reconciles (ending = beginning + new + expansion - contraction - churn)",
         f'=IF(MAX(ABS(\'ARR Rollforward\'!C10-(\'ARR Rollforward\'!C5+\'ARR Rollforward\'!C6+\'ARR Rollforward\'!C7-\'ARR Rollforward\'!C8-\'ARR Rollforward\'!C9)))<0.000001,"PASS","FAIL")'),
        ("ARR balances never negative",
         f'=IF(MIN(\'ARR Rollforward\'!C10:{last}10)>=0,"PASS","FAIL")'),
        ("Gross revenue retention does not exceed 100%",
         f'=IF(MAX(\'ARR Rollforward\'!C13:{last}13)<=1.000001,"PASS","FAIL")'),
        ("Net revenue retention exceeds gross revenue retention",
         f'=IF(MIN(\'ARR Rollforward\'!C12:{last}12-\'ARR Rollforward\'!C13:{last}13)>=-0.000001,"PASS","FAIL")'),
        ("Total revenue equals subscription plus services",
         f'=IF(MAX(ABS(\'Operating Model\'!C7:{last}7-\'Operating Model\'!C5:{last}5-\'Operating Model\'!C6:{last}6))<0.000001,"PASS","FAIL")'),
        ("Blended gross margin sits between services and subscription margin",
         f'=IF(AND(MIN(\'Operating Model\'!C11:{last}11)>=Assumptions!E12,MAX(\'Operating Model\'!C11:{last}11)<=Assumptions!E11),"PASS","REVIEW")'),
        ("Non-GAAP operating income exceeds GAAP (SBC added back)",
         f'=IF(MIN(\'Operating Model\'!C19:{last}19-\'Operating Model\'!C16:{last}16)>=0,"PASS","FAIL")'),
        ("Magic number positive",
         f'=IF(MIN(\'Unit Economics\'!C7:{last}7)>0,"PASS","REVIEW")'),
        ("Gross-profit CAC payback within 36 months",
         f'=IF(MAX(\'Unit Economics\'!C9:{last}9)<=36,"PASS","REVIEW")'),
        ("RPO roll-forward reconciles (ending = beginning + bookings - revenue)",
         f'=IF(MAX(ABS(\'RPO & Bookings\'!C7:{last}7-(\'RPO & Bookings\'!C6:{last}6+\'RPO & Bookings\'!C8:{last}8-\'RPO & Bookings\'!C5:{last}5)))<0.000001,"PASS","FAIL")'),
        ("Current RPO does not exceed total RPO",
         f'=IF(MIN(\'RPO & Bookings\'!C12:{last}12)>=-0.000001,"PASS","FAIL")'),
        ("Contract liability does not exceed total RPO (unbilled RPO non-negative)",
         f'=IF(MIN(\'RPO & Bookings\'!C14:{last}14)>=-0.000001,"PASS","REVIEW")'),
        ("Implied gross bookings positive (no negative-bookings year)",
         f'=IF(MIN(\'RPO & Bookings\'!C8:{last}8)>0,"PASS","REVIEW")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    overall = 5 + len(checks)
    sheet.cell(overall, 2, "Overall")
    sheet.cell(
        overall, 3,
        f'=IF(COUNTIF(C5:C{overall - 1},"FAIL")>0,"FAIL",'
        f'IF(COUNTIF(C5:C{overall - 1},"REVIEW")>0,"REVIEW","PASS"))',
    )
    add_status_rules(sheet, f"C5:C{overall}")
    set_widths(sheet, {"A": 4, "B": 72, "C": 18})

    add_sources(workbook, [
        ("ARR and retention disclosures", "[10-K / 10-Q / shareholder letter]", "[period]", "Beginning/ending ARR, NRR, customer counts"),
        ("Revenue disaggregation", "[10-K revenue note]", "[period]", "Subscription versus services split and gross margin by type"),
        ("Stock-based compensation", "[10-K equity note]", "[period]", "SBC by expense line"),
        ("Remaining performance obligations", "[10-K revenue note / XBRL RevenueRemainingPerformanceObligation]", "[period]", "Total RPO and the portion expected to be recognised within twelve months"),
        ("Contract liabilities", "[10-K balance sheet / XBRL ContractWithCustomerLiability]", "[period]", "Deferred revenue, used to split billed from unbilled RPO"),
        ("Industry cost of capital and multiples", "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/data.html", "[annual]", "Software industry margins, betas, and multiple ranges for driver grounding"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("SOFTWARE_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
