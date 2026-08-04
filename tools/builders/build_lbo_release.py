"""Canonical release wrapper for the institutional LBO builder.

The underlying builder creates the complete workbook layout. This release layer
sets the close-state balances and rewrites the debt schedule with explicit row
identities, avoiding row-offset mistakes and circular interest calculations.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_lbo_institutional import build as build_layout  # noqa: E402
from institutional_helpers import CUR, PCT, MULT, finalize  # noqa: E402

YEARS = 7


def build(output: Path) -> None:
    build_layout(output)
    workbook = load_workbook(output)
    debt = workbook["Debt Schedule"]
    operating = workbook["Operating Model"]
    returns = workbook["Returns Waterfall"]
    sensitivity = workbook["Sensitivity"]

    # Close-state balances.
    debt["C5"] = "=Assumptions!E16"          # opening / minimum cash
    debt["C13"] = "=C5"                      # ending cash at close
    debt["C14"] = "='Sources & Uses'!F5"     # beginning revolver
    debt["C16"] = "=C14"                     # ending revolver
    debt["C17"] = "='Sources & Uses'!F6"     # beginning TLB
    debt["C21"] = "=C17"                     # ending TLB
    debt["C22"] = "='Sources & Uses'!F7"     # beginning second lien
    debt["C26"] = "=C22"                     # ending second lien
    debt["C27"] = "=SUM(C16,C21,C26)"
    debt["C28"] = "=C27-C13"
    for row in (5, 13, 14, 16, 17, 21, 22, 26, 27, 28):
        debt.cell(row, 3).number_format = CUR

    for column in range(4, 4 + YEARS):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)

        debt.cell(5, column, f"={previous}13")
        debt.cell(6, column, f"='Operating Model'!{letter}19")

        debt.cell(14, column, f"={previous}16")
        debt.cell(15, column, f"={letter}14*Assumptions!$E$18")

        debt.cell(17, column, f"={previous}21")
        debt.cell(18, column, f"={letter}17*Assumptions!$E$20")
        debt.cell(19, column, f"=MIN({letter}17,'Sources & Uses'!$F$6*Assumptions!$E$21)")

        debt.cell(22, column, f"={previous}26")
        debt.cell(23, column, f"={letter}22*Assumptions!$E$23")
        debt.cell(24, column, f"={letter}22*Assumptions!$E$24")

        debt.cell(7, column, f"=SUM({letter}15,{letter}18,{letter}23)")
        debt.cell(8, column, f"={letter}19")
        debt.cell(9, column, f"={letter}5+{letter}6-{letter}7-{letter}8")
        debt.cell(10, column, f"=IF({letter}9<Assumptions!$E$16,MIN(Assumptions!$E$17-{letter}14,Assumptions!$E$16-{letter}9),-MIN({letter}14,MAX(0,{letter}9-Assumptions!$E$16)))")
        debt.cell(11, column, f"=MIN(MAX(0,{letter}17-{letter}19),MAX(0,{letter}9+{letter}10-Assumptions!$E$16)*Assumptions!$E$25)")
        debt.cell(12, column, f"=MIN(MAX(0,{letter}22+{letter}24),MAX(0,{letter}9+{letter}10-{letter}11-Assumptions!$E$16)*Assumptions!$E$25)")
        debt.cell(13, column, f"={letter}9+{letter}10-{letter}11-{letter}12")
        debt.cell(16, column, f"=MAX(0,{letter}14+{letter}10)")
        debt.cell(20, column, f"={letter}11")
        debt.cell(21, column, f"=MAX(0,{letter}17-{letter}19-{letter}20)")
        debt.cell(25, column, f"={letter}12")
        debt.cell(26, column, f"=MAX(0,{letter}22+{letter}24-{letter}25)")
        debt.cell(27, column, f"=SUM({letter}16,{letter}21,{letter}26)")
        debt.cell(28, column, f"={letter}27-{letter}13")
        for row in range(5, 29):
            debt.cell(row, column).number_format = CUR

        operating.cell(11, column, f"='Debt Schedule'!{letter}7")
        operating.cell(11, column).number_format = CUR

        return_column = get_column_letter(column - 1)
        returns.cell(8, column - 1, f"='Debt Schedule'!{letter}28")
        returns.cell(8, column - 1).number_format = CUR

    # Correct the selected-exit debt reference in the five-year sensitivity.
    for row in range(5, 10):
        for column in range(3, 8):
            letter = get_column_letter(column)
            sensitivity.cell(row, column, f"=IFERROR(((('Operating Model'!H8*{letter}$4-'Debt Schedule'!G28)*(1-Assumptions!$E$30))/MAX(0.01,('Operating Model'!C8*$B{row}+Assumptions!$E$15+'Sources & Uses'!C7+Assumptions!$E$16-'Sources & Uses'!F6-'Sources & Uses'!F7)))^(1/5)-1,0)")
            sensitivity.cell(row, column).number_format = PCT

    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("LBO_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
