"""Build the canonical structured-finance and securitization decision model."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    formula_cell, header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "Collateral", "Loss Curves", "Waterfall",
    "Triggers", "Tranche Analytics", "Sensitivity", "Checks", "Sources",
    "RefreshLog",
]

MONTHS = 24


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[TRANSACTION] — Structured Credit Model", [
        ("Transaction / collateral:", "ABS / RMBS / CMBS / CLO / specialty"),
        ("Cut-off / payment date:", "[dates]"),
        ("Last refreshed:", "[date]"),
        ("Next distribution / trigger date:", "[date]"),
        ("Refresh cadence:", "Monthly and on servicer reports"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Collateral and Capital Structure Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Opening collateral balance", 1000.0, 1000.0, "$mm", CUR),
        ("Collateral gross coupon", 0.080, 0.075, "%", PCT2),
        ("Annual CPR", 0.120, 0.060, "%", PCT2),
        ("Annual CDR", 0.025, 0.080, "%", PCT2),
        ("Recovery rate", 0.450, 0.250, "%", PCT2),
        ("Recovery lag", 3.0, 6.0, "months", "0"),
        ("Servicing and senior fees", 0.010, 0.012, "% annual", PCT2),
        ("Senior tranche opening balance", 650.0, 650.0, "$mm", CUR),
        ("Senior tranche coupon", 0.050, 0.060, "%", PCT2),
        ("Mezz tranche opening balance", 220.0, 220.0, "$mm", CUR),
        ("Mezz tranche coupon", 0.080, 0.100, "%", PCT2),
        ("Reserve account", 20.0, 20.0, "$mm", CUR),
        ("Minimum OC ratio", 1.08, 1.08, "x", MULT),
        ("Minimum senior IC ratio", 1.50, 1.50, "x", MULT),
        ("Minimum mezz IC ratio", 1.15, 1.15, "x", MULT),
        ("Payment frequency", 12.0, 12.0, "per year", "0"),
        ("Model horizon", float(MONTHS), float(MONTHS), "months", "0"),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        formula_cell(sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})'), number_format, cross_sheet=True)
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 42, "C": 15, "D": 15, "E": 15, "F": 30})
    sheet.freeze_panes = "A5"

    sheet = workbook["Loss Curves"]
    title(sheet, "B2:Z2", "Monthly Prepayment, Default, and Recovery Curves")
    header(sheet, 4, 2, ["Rate", *[f"M{month}" for month in range(1, MONTHS + 1)]])
    labels = ["SMM", "MDR", "Recovery rate", "Cumulative default rate", "Cumulative loss rate"]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 3 + MONTHS):
        letter = get_column_letter(column)
        sheet.cell(5, column, "=1-(1-Assumptions!$E$7)^(1/Assumptions!$E$20)")
        sheet.cell(6, column, "=1-(1-Assumptions!$E$8)^(1/Assumptions!$E$20)")
        sheet.cell(7, column, "=Assumptions!$E$9")
        if column == 3:
            sheet.cell(8, column, f"=Collateral!{letter}9/Assumptions!$E$5")
            sheet.cell(9, column, f"=Collateral!{letter}9*(1-{letter}7)/Assumptions!$E$5")
        else:
            previous = get_column_letter(column - 1)
            sheet.cell(8, column, f"={previous}8+Collateral!{letter}9/Assumptions!$E$5")
            sheet.cell(9, column, f"={previous}9+Collateral!{letter}9*(1-{letter}7)/Assumptions!$E$5")
        for row in range(5, 10):
            sheet.cell(row, column).number_format = PCT2
    set_widths(sheet, {"A": 4, "B": 28, **{get_column_letter(column): 10 for column in range(3, 3 + MONTHS)}})
    sheet.freeze_panes = "C5"

    sheet = workbook["Collateral"]
    title(sheet, "B2:Z2", "Monthly Collateral Cash Flows")
    header(sheet, 4, 2, ["$mm", *[f"M{month}" for month in range(1, MONTHS + 1)]])
    labels = [
        "Beginning collateral", "Scheduled principal", "Defaults", "Recoveries",
        "Prepayments", "Interest collections", "Servicing / senior fees",
        "Net interest collections", "Principal collections", "Ending collateral",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 3 + MONTHS):
        letter = get_column_letter(column)
        curve_letter = letter
        remaining = MONTHS - (column - 3)
        if column == 3:
            sheet.cell(5, column, "=Assumptions!E5")
        else:
            previous = get_column_letter(column - 1)
            sheet.cell(5, column, f"={previous}14")
        sheet.cell(6, column, f"=MIN({letter}5,{letter}5/{remaining})")
        sheet.cell(7, column, f"=MAX(0,{letter}5-{letter}6)*'Loss Curves'!{curve_letter}6")
        lag = 3
        if column - lag >= 3:
            lag_letter = get_column_letter(column - lag)
            sheet.cell(8, column, f"={lag_letter}7*Assumptions!$E$9")
        else:
            sheet.cell(8, column, 0.0)
        sheet.cell(9, column, f"=MAX(0,{letter}5-{letter}6-{letter}7)*'Loss Curves'!{curve_letter}5")
        sheet.cell(10, column, f"={letter}5*Assumptions!$E$6/Assumptions!$E$20")
        sheet.cell(11, column, f"={letter}5*Assumptions!$E$11/Assumptions!$E$20")
        sheet.cell(12, column, f"=MAX(0,{letter}10-{letter}11)")
        sheet.cell(13, column, f"={letter}6+{letter}8+{letter}9")
        sheet.cell(14, column, f"=MAX(0,{letter}5-{letter}6-{letter}7-{letter}9)")
        for row in range(5, 15):
            sheet.cell(row, column).number_format = CUR
    for row in (12, 13, 14):
        total_row(sheet, row, 2, 2 + MONTHS, CUR)
    set_widths(sheet, {"A": 4, "B": 32, **{get_column_letter(column): 10 for column in range(3, 3 + MONTHS)}})
    sheet.freeze_panes = "C5"

    sheet = workbook["Waterfall"]
    title(sheet, "B2:Z2", "Monthly Priority of Payments")
    header(sheet, 4, 2, ["$mm", *[f"M{month}" for month in range(1, MONTHS + 1)]])
    labels = [
        "Available interest", "Available principal", "Senior beginning balance",
        "Senior interest due", "Senior interest paid", "Senior interest shortfall",
        "Senior principal paid", "Senior ending balance", "Mezz beginning balance",
        "Mezz interest due", "Mezz interest paid", "Mezz interest shortfall",
        "Mezz principal paid", "Mezz ending balance", "Residual cash",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 3 + MONTHS):
        letter = get_column_letter(column)
        sheet.cell(5, column, f"=Collateral!{letter}12+Collateral!{letter}8")
        sheet.cell(6, column, f"=Collateral!{letter}13")
        if column == 3:
            sheet.cell(7, column, "=Assumptions!E12")
            sheet.cell(13, column, "=Assumptions!E14")
            senior_shortfall_previous = "0"
            mezz_shortfall_previous = "0"
        else:
            previous = get_column_letter(column - 1)
            sheet.cell(7, column, f"={previous}12")
            sheet.cell(13, column, f"={previous}18")
            senior_shortfall_previous = f"{previous}10"
            mezz_shortfall_previous = f"{previous}16"
        sheet.cell(8, column, f"={letter}7*Assumptions!$E$13/Assumptions!$E$20+{senior_shortfall_previous}")
        sheet.cell(9, column, f"=MIN({letter}5,{letter}8)")
        sheet.cell(10, column, f"={letter}8-{letter}9")
        sheet.cell(11, column, f"=MIN({letter}7,{letter}6)")
        sheet.cell(12, column, f"=MAX(0,{letter}7-{letter}11)")
        sheet.cell(14, column, f"={letter}13*Assumptions!$E$15/Assumptions!$E$20+{mezz_shortfall_previous}")
        sheet.cell(15, column, f"=MIN(MAX(0,{letter}5-{letter}9),{letter}14)")
        sheet.cell(16, column, f"={letter}14-{letter}15")
        sheet.cell(17, column, f"=MIN({letter}13,MAX(0,{letter}6-{letter}11))")
        sheet.cell(18, column, f"=MAX(0,{letter}13-{letter}17)")
        sheet.cell(19, column, f"=MAX(0,{letter}5-{letter}9-{letter}15)+MAX(0,{letter}6-{letter}11-{letter}17)")
        for row in range(5, 20):
            sheet.cell(row, column).number_format = CUR
    for row in (12, 18, 19):
        total_row(sheet, row, 2, 2 + MONTHS, CUR)
    set_widths(sheet, {"A": 4, "B": 34, **{get_column_letter(column): 10 for column in range(3, 3 + MONTHS)}})
    sheet.freeze_panes = "C5"

    sheet = workbook["Triggers"]
    title(sheet, "B2:Z2", "OC / IC Triggers and Cash Diversion")
    header(sheet, 4, 2, ["Metric", *[f"M{month}" for month in range(1, MONTHS + 1)]])
    labels = [
        "OC ratio", "Minimum OC ratio", "Senior interest coverage",
        "Minimum senior IC", "Mezz interest coverage", "Minimum mezz IC",
        "Trigger status",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 3 + MONTHS):
        letter = get_column_letter(column)
        sheet.cell(5, column, f"=IFERROR(Collateral!{letter}14/(Waterfall!{letter}12+Waterfall!{letter}18),0)")
        sheet.cell(6, column, "=Assumptions!$E$17")
        sheet.cell(7, column, f"=IFERROR(Collateral!{letter}12/Waterfall!{letter}8,0)")
        sheet.cell(8, column, "=Assumptions!$E$18")
        sheet.cell(9, column, f"=IFERROR(MAX(0,Collateral!{letter}12-Waterfall!{letter}8)/Waterfall!{letter}14,0)")
        sheet.cell(10, column, "=Assumptions!$E$19")
        sheet.cell(11, column, f'=IF(AND({letter}5>={letter}6,{letter}7>={letter}8,{letter}9>={letter}10),"PASS","TRIGGER")')
        for row in range(5, 11):
            sheet.cell(row, column).number_format = MULT
    add_status_rules(sheet, f"C11:{get_column_letter(2 + MONTHS)}11")
    set_widths(sheet, {"A": 4, "B": 32, **{get_column_letter(column): 10 for column in range(3, 3 + MONTHS)}})
    sheet.freeze_panes = "C5"

    sheet = workbook["Tranche Analytics"]
    title(sheet, "B2:F2", "Tranche Yield, WAL, and Loss Analytics")
    header(sheet, 4, 2, ["Metric", "Senior", "Mezz", "Residual", "Notes"])
    last = get_column_letter(2 + MONTHS)
    analytics = [
        ("Opening balance", "=Assumptions!E12", "=Assumptions!E14", "=Assumptions!E5-Assumptions!E12-Assumptions!E14", "capital structure"),
        ("Ending balance", f"=Waterfall!{last}12", f"=Waterfall!{last}18", "=0", "modeled horizon"),
        ("Principal paid", "=SUM(Waterfall!C11:Z11)", "=SUM(Waterfall!C17:Z17)", "=0", "cash principal"),
        ("Interest paid", "=SUM(Waterfall!C9:Z9)", "=SUM(Waterfall!C15:Z15)", "=SUM(Waterfall!C19:Z19)", "cash interest / residual"),
        ("WAL", "=SUMPRODUCT(COLUMN(Waterfall!C11:Z11)-COLUMN(Waterfall!B11),Waterfall!C11:Z11)/Assumptions!E20/SUM(Waterfall!C11:Z11)", "=SUMPRODUCT(COLUMN(Waterfall!C17:Z17)-COLUMN(Waterfall!B17),Waterfall!C17:Z17)/Assumptions!E20/SUM(Waterfall!C17:Z17)", "=0", "years"),
        ("Interest shortfall", "=SUM(Waterfall!C10:Z10)", "=SUM(Waterfall!C16:Z16)", "=0", "cumulative"),
        ("Principal loss / unpaid balance", f"=Waterfall!{last}12", f"=Waterfall!{last}18", "=MAX(0,Collateral!Z14-Waterfall!Z12-Waterfall!Z18)", "end-horizon exposure"),
    ]
    for row, values in enumerate(analytics, start=5):
        for column, value in enumerate(values, start=2):
            sheet.cell(row, column, value)
            if column in (3, 4, 5):
                sheet.cell(row, column).number_format = MULT if row == 9 else CUR
    set_widths(sheet, {"A": 4, "B": 34, "C": 18, "D": 18, "E": 18, "F": 38})

    sheet = workbook["Sensitivity"]
    title(sheet, "B2:G2", "Senior Ending Balance Sensitivity")
    header(sheet, 4, 2, ["CDR / CPR", 0.00, 0.06, 0.12, 0.18, 0.24])
    for column in range(3, 8):
        sheet.cell(4, column).number_format = PCT
    for row, cdr in enumerate((0.00, 0.03, 0.06, 0.09, 0.12), start=5):
        sheet.cell(row, 2, cdr)
        sheet.cell(row, 2).number_format = PCT
        for column in range(3, 8):
            letter = get_column_letter(column)
            sheet.cell(row, column, f"=MAX(0,Assumptions!$E$12-Assumptions!$E$5*(1-(1-$B{row})^(2))*(1-Assumptions!$E$9)-Assumptions!$E$5*(1-(1-{letter}$4)^2))")
            sheet.cell(row, column).number_format = CUR
    sheet.conditional_formatting.add(
        "C5:G9",
        ColorScaleRule(
            start_type="min", start_color="E2F0D9",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="FCE4D6",
        ),
    )
    set_widths(sheet, {"A": 4, "B": 22, **{get_column_letter(column): 14 for column in range(3, 8)}})

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("Collateral balance nonnegative", '=IF(MIN(Collateral!C14:Z14)>=0,"PASS","FAIL")'),
        ("Senior balance nonnegative", '=IF(MIN(Waterfall!C12:Z12)>=0,"PASS","FAIL")'),
        ("Mezz balance nonnegative", '=IF(MIN(Waterfall!C18:Z18)>=0,"PASS","FAIL")'),
        ("Collateral roll-forward", '=IF(MAX(ABS(Collateral!C5:Z5-Collateral!C6:Z6-Collateral!C7:Z7-Collateral!C9:Z9-Collateral!C14:Z14))<0.000001,"PASS","FAIL")'),
        ("Interest waterfall conserved", '=IF(MAX(ABS(Collateral!C12:Z12+Collateral!C8:Z8-Waterfall!C9:Z9-Waterfall!C15:Z15-(Waterfall!C19:Z19-MAX(0,Waterfall!C6:Z6-Waterfall!C11:Z11-Waterfall!C17:Z17))))<0.000001,"PASS","REVIEW")'),
        ("No trigger breach", '=IF(COUNTIF(Triggers!C11:Z11,"TRIGGER")=0,"PASS","REVIEW")'),
        ("Recovery and loss rates bounded", '=IF(AND(MIN(\'Loss Curves\'!C7:Z9)>=0,MAX(\'Loss Curves\'!C7:Z9)<=1),"PASS","FAIL")'),
        ("Overall", '=IF(COUNTIF(C5:C11,"FAIL")=0,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    add_status_rules(sheet, "C5:C12")
    set_widths(sheet, {"A": 4, "B": 48, "C": 18})

    add_sources(workbook, [
        ("Offering memorandum / indenture", "[public filing / data room URL]", "[date]", "Capital structure, priority of payments, triggers"),
        ("Servicer report", "[trustee / servicer URL]", "[payment date]", "Pool balance, delinquencies, defaults, recoveries"),
        ("Collateral data tape", "[restricted source]", "[cut-off date]", "Loan-level fields and transformations"),
        ("Rating methodology", "[agency methodology URL]", "[date]", "Stress assumptions and legal criteria"),
        ("Market benchmarks", "[market data URL]", "[date]", "Spreads, CPR/CDR expectations, relative value"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("SECURITIZATION_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
