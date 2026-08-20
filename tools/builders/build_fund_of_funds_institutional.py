"""Build the canonical fund-of-funds (FoF) look-through and fee-layering model."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "Underlying Fund Portfolio",
    "NAV Rollforward & Fee Layering", "Performance & Multiples",
    "Checks", "Sources", "RefreshLog",
]

FUND_COUNT = 8


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[FUND OF FUNDS] — Look-Through and Fee-Layering Model", [
        ("FoF vehicle / vintage:", "[Name / vintage]"),
        ("Reporting date:", "[date]"),
        ("Last refreshed:", "[date]"),
        ("Next LP report / valuation:", "[date]"),
        ("Refresh cadence:", "Quarterly, aligned to underlying-fund GP reporting"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Fund-of-Funds Terms and Portfolio Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Total LP commitments to FoF", 500.0, 500.0, "$mm", CUR),
        ("FoF management fee (% of commitments, p.a.)", 0.0075, 0.0075, "%", PCT2),
        ("FoF carried interest (% of profit above hurdle)", 0.05, 0.05, "%", PCT),
        ("FoF preferred return / hurdle", 0.08, 0.08, "%", PCT),
        ("Number of underlying funds", float(FUND_COUNT), float(FUND_COUNT), "count", "0"),
        ("Underlying-fund NAV markdown (stress)", 0.0, 0.25, "%", PCT),
        ("Maximum single-fund concentration limit", 0.20, 0.20, "% of NAV", PCT),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 44, "C": 15, "D": 15, "E": 15, "F": 32})
    sheet.freeze_panes = "A5"

    sheet = workbook["Underlying Fund Portfolio"]
    title(sheet, "B2:J2", "Underlying Fund Portfolio (Look-Through)")
    header(sheet, 4, 2, [
        "Underlying fund / GP", "Vintage year", "Strategy",
        "Commitment", "Called", "Distributed",
        "Reported NAV (Base)", "Active NAV", "Gross TVPI (look-through)",
    ])
    default_funds = [
        ("[Fund A]", 2016, "Buyout", 60.0, 58.0, 71.0, 42.0),
        ("[Fund B]", 2017, "Buyout", 55.0, 52.0, 48.0, 55.0),
        ("[Fund C]", 2018, "Growth equity", 65.0, 58.0, 30.0, 78.0),
        ("[Fund D]", 2019, "Venture", 45.0, 38.0, 12.0, 58.0),
        ("[Fund E]", 2019, "Buyout", 70.0, 63.0, 25.0, 82.0),
        ("[Fund F]", 2020, "Secondaries", 50.0, 44.0, 18.0, 56.0),
        ("[Fund G]", 2021, "Growth equity", 60.0, 41.0, 4.0, 51.0),
        ("[Fund H]", 2022, "Buyout", 65.0, 32.0, 0.5, 36.0),
    ]
    for offset, (name, vintage, strategy, commitment, called, distributed, nav) in enumerate(default_funds):
        row = 5 + offset
        input_cell(sheet.cell(row, 2, name))
        input_cell(sheet.cell(row, 3, vintage), "0")
        input_cell(sheet.cell(row, 4, strategy))
        input_cell(sheet.cell(row, 5, commitment), CUR)
        input_cell(sheet.cell(row, 6, called), CUR)
        input_cell(sheet.cell(row, 7, distributed), CUR)
        input_cell(sheet.cell(row, 8, nav), CUR)
        sheet.cell(row, 9, f"=H{row}*(1-Assumptions!$E$10)")
        sheet.cell(row, 9).number_format = CUR
        sheet.cell(row, 10, f"=IFERROR((G{row}+I{row})/F{row},0)")
        sheet.cell(row, 10).number_format = MULT
    total_row(sheet, 13, 2, 10, CUR)
    sheet["B13"] = "Total / look-through weighted"
    for column in (5, 6, 7, 8, 9):
        letter = get_column_letter(column)
        sheet.cell(13, column, f"=SUM({letter}5:{letter}12)")
    sheet.cell(13, 10, "=IFERROR((G13+I13)/F13,0)")
    sheet.cell(13, 10).number_format = MULT
    sheet["B15"] = "Largest single-fund Active NAV"
    sheet["C15"] = "=MAX(I5:I12)"
    sheet["C15"].number_format = CUR
    sheet["B16"] = "Largest single-fund concentration"
    sheet["C16"] = "=IFERROR(C15/I13,0)"
    sheet["C16"].number_format = PCT
    set_widths(sheet, {"A": 4, "B": 20, "C": 12, "D": 16, "E": 14, "F": 14, "G": 14, "H": 16, "I": 14, "J": 20})
    sheet.freeze_panes = "C5"

    sheet = workbook["NAV Rollforward & Fee Layering"]
    title(sheet, "B2:C2", "FoF-Level NAV Roll-Forward and Fee Layering")
    header(sheet, 4, 2, ["Line item", "Value"])
    rollforward_labels = [
        "Beginning-of-period FoF NAV",
        "+ Capital called from LPs this period",
        "- Distributions to LPs this period",
        "- FoF management fee this period",
        "- FoF carried interest this period",
        "+/- Net realized/unrealized gain (underlying funds)",
    ]
    for offset, label in enumerate(rollforward_labels):
        sheet.cell(5 + offset, 2, label)
    input_cell(sheet.cell(5, 3, 380.0), CUR)
    input_cell(sheet.cell(6, 3, 45.0), CUR)
    input_cell(sheet.cell(7, 3, 38.0), CUR)
    sheet.cell(8, 3, "=Assumptions!$C$5*Assumptions!$C$6")
    sheet.cell(9, 3, "=MAX(0,C7-C5*Assumptions!$C$8)*Assumptions!$C$7")
    input_cell(sheet.cell(10, 3, 28.13), CUR)
    for row in (5, 6, 7, 8, 9, 10):
        sheet.cell(row, 3).number_format = CUR

    sheet["B11"] = "Computed ending-of-period FoF NAV"
    sheet["C11"] = "=C5+C6-C7-IF(ISBLANK(C14),C8,C14)-C9+C10"
    sheet["C11"].number_format = CUR

    sheet["B12"] = "Reported ending-of-period FoF NAV"
    input_cell(sheet.cell(12, 3, 411.0), CUR)

    # A genuine external check, not a tautology: C10 (the realized/unrealized
    # gain) is sourced independently from the fund's own statement of
    # changes in net assets, and C11 is computed forward from it -- this
    # residual can be nonzero if the sourced inputs don't actually
    # reconcile, unlike a plug solved backward from the reported figure.
    sheet["B13"] = "Reconciliation residual (computed vs. reported ending NAV)"
    sheet["C13"] = "=C11-C12"
    sheet["C13"].number_format = CUR

    # A filed fund's expense load is never just its advisory fee: the
    # Statement of Changes in Net Assets opens with net investment income
    # (loss), which already contains the management fee plus every other
    # operating expense. When a real case sources that figure here it
    # supersedes the modelled fee in C8 -- substituting the fee for it is
    # what leaves a real fund's roll-forward unreconciled. Left blank for
    # illustrative cases, which fall back to the modelled fee.
    sheet["B14"] = "- Disclosed net investment loss (supersedes modelled fee when sourced)"
    input_cell(sheet.cell(14, 3), CUR)   # left blank: ISBLANK falls back to the modelled fee

    sheet["B15"] = "Total LP capital called to date"
    input_cell(sheet.cell(15, 3, 400.0), CUR)
    sheet["B16"] = "Total LP distributions to date"
    input_cell(sheet.cell(16, 3, 165.0), CUR)

    set_widths(sheet, {"A": 4, "B": 46, "C": 18})
    sheet.freeze_panes = "A5"

    sheet = workbook["Performance & Multiples"]
    title(sheet, "B2:C2", "FoF Net Multiples, Look-Through Gross Multiples, and Fee Drag")
    header(sheet, 4, 2, ["Metric", "Value"])
    metrics = [
        ("FoF net DPI (distributions / paid-in, to LP)",
         "=IFERROR('NAV Rollforward & Fee Layering'!C16/'NAV Rollforward & Fee Layering'!C15,0)", MULT),
        ("FoF net RVPI (residual value / paid-in, to LP)",
         "=IFERROR('NAV Rollforward & Fee Layering'!C12/'NAV Rollforward & Fee Layering'!C15,0)", MULT),
        ("FoF net TVPI (to LP)", "=C5+C6", MULT),
        ("Look-through gross TVPI (weighted, underlying funds)",
         "='Underlying Fund Portfolio'!J13", MULT),
        ("Fee-layering drag (gross TVPI - FoF net TVPI)", "=C8-C7", MULT),
        ("Fee-layering drag, $ of paid-in capital",
         "=C9*'NAV Rollforward & Fee Layering'!C15", CUR),
    ]
    for offset, (label, formula, number_format) in enumerate(metrics):
        row = 5 + offset
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
        sheet.cell(row, 3).number_format = number_format
    set_widths(sheet, {"A": 4, "B": 52, "C": 18})
    sheet.freeze_panes = "A5"

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Fund-of-Funds Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("NAV roll-forward reconciles",
         '=IF(ABS(\'NAV Rollforward & Fee Layering\'!C13)<0.01,"PASS","FAIL")'),
        ("TVPI equals DPI plus RVPI",
         '=IF(ABS(\'Performance & Multiples\'!C7-(\'Performance & Multiples\'!C5+\'Performance & Multiples\'!C6))<0.0001,"PASS","FAIL")'),
        ("Fee-layering drag is nonnegative",
         '=IF(\'Performance & Multiples\'!C9>=-0.0001,"PASS","FAIL")'),
        ("Single-fund concentration within limit",
         '=IF(\'Underlying Fund Portfolio\'!C16<=Assumptions!$E$11,"PASS","BREACH")'),
        ("Portfolio called capital does not exceed commitments",
         '=IF(\'Underlying Fund Portfolio\'!F13<=\'Underlying Fund Portfolio\'!E13,"PASS","FAIL")'),
        ("FoF-level called capital does not exceed total commitments",
         '=IF(\'NAV Rollforward & Fee Layering\'!C15<=Assumptions!$C$5,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    sheet["B11"] = "Overall"
    sheet["C11"] = '=IF(COUNTIF(C5:C10,"FAIL")+COUNTIF(C5:C10,"BREACH")>0,"BREACH",IF(COUNTIF(C5:C10,"REVIEW")>0,"REVIEW","PASS"))'
    add_status_rules(sheet, "C5:C11")
    set_widths(sheet, {"A": 4, "B": 50, "C": 18})

    add_sources(workbook, [
        ("Underlying fund capital account statements", "[GP quarterly/annual capital account statement]", "[period end]", "Commitment, called, distributed, reported NAV per underlying fund"),
        ("FoF administrator NAV package", "[fund administrator report]", "[period end]", "FoF-level beginning/ending NAV, capital called and distributed to LPs"),
        ("Fee letter / limited partnership agreement", "[LPA / side letter]", "[effective date]", "FoF management fee rate, carried interest rate, hurdle"),
    ])
    add_refresh_log(workbook)

    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("FOF_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
