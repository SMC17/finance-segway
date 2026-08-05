"""Build the canonical private-equity and merchant-banking LBO model."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    header, input_cell, set_widths, title, total_row,
)

YEARS = 7
SHEETS = [
    "Cover", "Assumptions", "Sources & Uses", "Operating Model",
    "Debt Schedule", "Covenants", "Management Equity", "Returns Waterfall",
    "Sensitivity", "Checks", "Sources", "RefreshLog",
]


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[TARGET] — LBO Underwriting Model", [
        ("Target / transaction:", "[Name]"),
        ("Sponsor / principal:", "[Fund / merchant bank]"),
        ("Entry / close date:", "[date]"),
        ("Last refreshed:", "[date]"),
        ("Next IC / financing event:", "[date]"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Transaction, Operating, and Financing Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("LTM revenue", 600.0, 600.0, "$mm", CUR),
        ("LTM EBITDA margin", 0.20, 0.17, "%", PCT),
        ("Annual revenue growth", 0.06, -0.02, "%", PCT),
        ("Annual EBITDA margin expansion", 0.005, -0.003, "% points", PCT2),
        ("D&A / revenue", 0.025, 0.025, "%", PCT),
        ("Capex / revenue", 0.035, 0.045, "%", PCT),
        ("NWC / revenue", 0.080, 0.100, "%", PCT),
        ("Cash tax rate", 0.25, 0.20, "%", PCT),
        ("Entry EBITDA multiple", 10.0, 9.0, "x", MULT),
        ("Transaction fees / enterprise value", 0.020, 0.025, "%", PCT),
        ("Existing net debt refinanced", 100.0, 100.0, "$mm", CUR),
        ("Minimum cash", 20.0, 25.0, "$mm", CUR),
        ("Revolver commitment", 75.0, 75.0, "$mm", CUR),
        ("Revolver cash rate", 0.090, 0.115, "%", PCT2),
        ("TLB opening leverage", 4.0, 3.5, "x EBITDA", MULT),
        ("TLB cash rate", 0.085, 0.105, "%", PCT2),
        ("TLB annual amortization", 0.010, 0.010, "% original", PCT),
        ("Second-lien opening leverage", 1.5, 1.0, "x EBITDA", MULT),
        ("Second-lien cash rate", 0.110, 0.130, "%", PCT2),
        ("Second-lien PIK rate", 0.030, 0.050, "%", PCT2),
        ("Cash sweep", 0.75, 0.50, "% excess cash", PCT),
        ("Maximum leverage covenant", 7.0, 6.5, "x", MULT),
        ("Minimum interest coverage", 1.75, 1.50, "x", MULT),
        ("Exit EBITDA multiple", 10.0, 8.0, "x", MULT),
        ("Exit year", 5.0, 5.0, "year", "0"),
        ("Management equity pool", 0.10, 0.10, "% fully diluted", PCT),
        ("Sponsor preferred return", 0.08, 0.08, "%", PCT),
        ("Management catch-up above hurdle", 0.20, 0.20, "% incremental", PCT),
        ("Revolver commitment fee (undrawn balance)", 0.00375, 0.005, "% p.a.", PCT2),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 46, "C": 15, "D": 15, "E": 15, "F": 32})
    sheet.freeze_panes = "A5"

    sheet = workbook["Sources & Uses"]
    title(sheet, "B2:G2", "Transaction Sources & Uses")
    header(sheet, 4, 2, ["Uses", "Amount", "", "Sources", "Amount", ""])
    uses = [
        ("Purchase enterprise value", "=Assumptions!E5*Assumptions!E6*Assumptions!E13"),
        ("Refinance existing net debt", "=Assumptions!E15"),
        ("Transaction fees", "=C5*Assumptions!E14"),
        ("Minimum cash funded", "=Assumptions!E16"),
        ("Total uses", "=SUM(C5:C8)"),
    ]
    for row, (label, formula) in enumerate(uses, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
        sheet.cell(row, 3).number_format = CUR
    sources = [
        ("Revolver at close", 0.0),
        ("Term loan B", "=Assumptions!E5*Assumptions!E6*Assumptions!E19"),
        ("Second-lien / holdco debt", "=Assumptions!E5*Assumptions!E6*Assumptions!E22"),
        ("Sponsor equity", "=C9-SUM(F5:F7)"),
        ("Total sources", "=SUM(F5:F8)"),
        ("Sources less uses", "=F9-C9"),
    ]
    for row, (label, value) in enumerate(sources, start=5):
        sheet.cell(row, 5, label)
        cell = sheet.cell(row, 6, value)
        if isinstance(value, str):
            cell.number_format = CUR
        else:
            input_cell(cell, CUR)
    total_row(sheet, 9, 2, 3, CUR)
    total_row(sheet, 9, 5, 6, CUR)
    set_widths(sheet, {"A": 4, "B": 34, "C": 18, "D": 5, "E": 34, "F": 18, "G": 5})

    sheet = workbook["Operating Model"]
    title(sheet, "B2:I2", "Seven-Year Operating, Tax, and Cash Flow Model")
    header(sheet, 4, 2, ["$mm / %", "LTM", *[f"Year {year}" for year in range(1, YEARS + 1)]])
    labels = [
        "Revenue", "Revenue growth", "EBITDA margin", "EBITDA", "D&A", "EBIT",
        "Cash interest", "EBT", "Cash taxes", "Net income", "Capex", "NWC",
        "Change in NWC", "FCF before financing", "FCF before interest",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    sheet["C5"] = "=Assumptions!E5"
    sheet["C7"] = "=Assumptions!E6"
    sheet["C8"] = "=C5*C7"
    sheet["C9"] = "=C5*Assumptions!E9"
    sheet["C10"] = "=C8-C9"
    sheet["C16"] = "=C5*Assumptions!E11"
    for row in (5, 8, 9, 10, 16):
        sheet.cell(row, 3).number_format = CUR
    sheet["C7"].number_format = PCT
    for column in range(4, 4 + YEARS):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)
        debt_column = get_column_letter(column - 1)
        sheet.cell(5, column, f"={previous}5*(1+Assumptions!$E$7)")
        sheet.cell(6, column, f"=IFERROR({letter}5/{previous}5-1,0)")
        sheet.cell(7, column, f"=MAX(0,MIN(0.60,{previous}7+Assumptions!$E$8))")
        sheet.cell(8, column, f"={letter}5*{letter}7")
        sheet.cell(9, column, f"={letter}5*Assumptions!$E$9")
        sheet.cell(10, column, f"={letter}8-{letter}9")
        sheet.cell(11, column, f"='Debt Schedule'!{debt_column}12")
        sheet.cell(12, column, f"={letter}10-{letter}11")
        sheet.cell(13, column, f"=MAX(0,{letter}12*Assumptions!$E$12)")
        sheet.cell(14, column, f"={letter}12-{letter}13")
        sheet.cell(15, column, f"={letter}5*Assumptions!$E$10")
        sheet.cell(16, column, f"={letter}5*Assumptions!$E$11")
        sheet.cell(17, column, f"={letter}16-{previous}16")
        sheet.cell(18, column, f"={letter}14+{letter}9-{letter}15-{letter}17")
        sheet.cell(19, column, f"={letter}8-{letter}13-{letter}15-{letter}17")
        for row in (5, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19):
            sheet.cell(row, column).number_format = CUR
        for row in (6, 7):
            sheet.cell(row, column).number_format = PCT
    for row in (8, 18, 19):
        total_row(sheet, row, 2, 3 + YEARS, CUR)
    set_widths(sheet, {"A": 4, "B": 36, **{get_column_letter(column): 13 for column in range(3, 4 + YEARS)}})
    sheet.freeze_panes = "D5"

    sheet = workbook["Debt Schedule"]
    title(sheet, "B2:I2", "Multi-Tranche Debt and Cash Sweep")
    header(sheet, 4, 2, ["$mm / %", "Close", *[f"Year {year}" for year in range(1, YEARS + 1)]])
    labels = [
        "Beginning cash", "FCF before interest", "Cash interest", "Mandatory TLB amortization",
        "Cash before revolver / sweep", "Revolver draw / (repayment)", "Cash sweep to TLB",
        "Cash sweep to second lien", "Ending cash", "Beginning revolver", "Revolver interest",
        "Ending revolver", "Beginning TLB", "TLB interest", "TLB amortization",
        "TLB cash sweep", "Ending TLB", "Beginning second lien", "Second-lien cash interest",
        "Second-lien PIK", "Second-lien cash sweep", "Ending second lien", "Total debt",
        "Net debt", "Revolver commitment fee (undrawn balance)",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    sheet["C5"] = "=Assumptions!E16"
    sheet["C13"] = "='Sources & Uses'!F5"
    sheet["C16"] = "='Sources & Uses'!F6"
    sheet["C20"] = "='Sources & Uses'!F7"
    sheet["C25"] = "=SUM(C16,C20,C13)"
    sheet["C26"] = "=C25-C5"
    for row in (5, 13, 16, 20, 25, 26):
        sheet.cell(row, 3).number_format = CUR
    for column in range(4, 4 + YEARS):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)
        operating_column = letter
        sheet.cell(5, column, f"={previous}13")
        sheet.cell(6, column, f"='Operating Model'!{operating_column}19")
        sheet.cell(14, column, f"={previous}15")
        sheet.cell(15, column, f"={letter}14*Assumptions!$E$18")
        sheet.cell(16, column, f"={previous}20")
        sheet.cell(17, column, f"={letter}16*Assumptions!$E$20")
        sheet.cell(18, column, f"=MIN({letter}16,'Sources & Uses'!$F$6*Assumptions!$E$21)")
        sheet.cell(21, column, f"={previous}24")
        sheet.cell(22, column, f"={letter}21*Assumptions!$E$23")
        sheet.cell(23, column, f"={letter}21*Assumptions!$E$24")
        sheet.cell(7, column, f"={letter}15+{letter}17+{letter}22")
        sheet.cell(8, column, f"={letter}16*0+{letter}18")
        sheet.cell(9, column, f"={letter}5+{letter}6-{letter}7-{letter}8")
        sheet.cell(10, column, f"=IF({letter}9<Assumptions!$E$16,MIN(Assumptions!$E$17-{letter}14,Assumptions!$E$16-{letter}9),-MIN({letter}14,MAX(0,{letter}9-Assumptions!$E$16)))")
        sheet.cell(11, column, f"=MIN(MAX(0,{letter}16-{letter}18),MAX(0,{letter}9+{letter}10-Assumptions!$E$16)*Assumptions!$E$25)")
        sheet.cell(12, column, f"=MIN(MAX(0,{letter}21+{letter}23),MAX(0,{letter}9+{letter}10-{letter}11-Assumptions!$E$16)*Assumptions!$E$25)")
        sheet.cell(13, column, f"={letter}9+{letter}10-{letter}11-{letter}12")
        sheet.cell(19, column, f"={letter}11")
        sheet.cell(20, column, f"=MAX(0,{letter}16-{letter}18-{letter}19)")
        sheet.cell(24, column, f"=MAX(0,{letter}21+{letter}23-{letter}12)")
        sheet.cell(25, column, f"=SUM({letter}15,{letter}20,{letter}24)")
        sheet.cell(26, column, f"={letter}25-{letter}13")
        for row in range(5, 27):
            sheet.cell(row, column).number_format = CUR
    for row in (13, 20, 24, 25, 26):
        total_row(sheet, row, 2, 3 + YEARS, CUR)
    set_widths(sheet, {"A": 4, "B": 42, **{get_column_letter(column): 13 for column in range(3, 4 + YEARS)}})
    sheet.freeze_panes = "D5"

    sheet = workbook["Covenants"]
    title(sheet, "B2:I2", "Lender Covenant and Liquidity Dashboard")
    header(sheet, 4, 2, ["Metric", *[f"Year {year}" for year in range(1, YEARS + 1)]])
    labels = [
        "Gross leverage", "Maximum leverage", "Leverage headroom",
        "Interest coverage", "Minimum coverage", "Coverage headroom",
        "Minimum cash", "Cash headroom", "Status",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 3 + YEARS):
        letter = get_column_letter(column)
        debt_column = get_column_letter(column + 1)
        operating_column = debt_column
        sheet.cell(5, column, f"=IFERROR('Debt Schedule'!{debt_column}25/'Operating Model'!{operating_column}8,0)")
        sheet.cell(6, column, "=Assumptions!$E$26")
        sheet.cell(7, column, f"={letter}6-{letter}5")
        sheet.cell(8, column, f"=IFERROR('Operating Model'!{operating_column}8/'Debt Schedule'!{debt_column}7,0)")
        sheet.cell(9, column, "=Assumptions!$E$27")
        sheet.cell(10, column, f"={letter}8-{letter}9")
        sheet.cell(11, column, "=Assumptions!$E$16")
        sheet.cell(12, column, f"='Debt Schedule'!{debt_column}13-{letter}11")
        sheet.cell(13, column, f'=IF(AND({letter}7>=0,{letter}10>=0,{letter}12>=0),"PASS","BREACH")')
        for row in range(5, 11):
            sheet.cell(row, column).number_format = MULT
        sheet.cell(11, column).number_format = CUR
        sheet.cell(12, column).number_format = CUR
    add_status_rules(sheet, f"C13:{get_column_letter(2 + YEARS)}13")
    set_widths(sheet, {"A": 4, "B": 34, **{get_column_letter(column): 13 for column in range(3, 3 + YEARS)}})

    sheet = workbook["Management Equity"]
    title(sheet, "B2:E2", "Management Equity and Incentive Pool")
    header(sheet, 4, 2, ["Metric", "Value", "Units", "Notes"])
    rows = [
        ("Management fully diluted pool", "=Assumptions!E30", "%", "granted / reserved pool", PCT),
        ("Sponsor ownership before pool", "=1-C5", "%", "fully diluted", PCT),
        ("Sponsor invested equity", "='Sources & Uses'!F8", "$mm", "entry equity", CUR),
        ("Preferred return at exit", "=C7*(1+Assumptions!E31)^Assumptions!E29", "$mm", "sponsor hurdle", CUR),
        ("Management catch-up rate", "=Assumptions!E32", "% incremental", "above preferred return", PCT),
    ]
    for row, (label, formula, units, note, number_format) in enumerate(rows, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
        sheet.cell(row, 3).number_format = number_format
        sheet.cell(row, 4, units)
        sheet.cell(row, 5, note)
    set_widths(sheet, {"A": 4, "B": 34, "C": 18, "D": 16, "E": 42})

    sheet = workbook["Returns Waterfall"]
    title(sheet, "B2:J2", "Exit Equity and Sponsor / Management Waterfall")
    header(sheet, 4, 2, ["Metric", *[f"Year {year}" for year in range(1, YEARS + 1)], "Selected exit"])
    labels = [
        "Exit EBITDA", "Exit multiple", "Exit enterprise value", "Net debt",
        "Gross equity value", "Sponsor preferred return", "Equity above preferred",
        "Management proceeds", "Sponsor proceeds", "Sponsor MOIC", "Sponsor IRR",
    ]
    for row, label in enumerate(labels, start=5):
        sheet.cell(row, 2, label)
    for column in range(3, 3 + YEARS):
        letter = get_column_letter(column)
        operating_column = get_column_letter(column + 1)
        debt_column = operating_column
        year = column - 2
        sheet.cell(5, column, f"='Operating Model'!{operating_column}8")
        sheet.cell(6, column, "=Assumptions!$E$28")
        sheet.cell(7, column, f"={letter}5*{letter}6")
        sheet.cell(8, column, f"='Debt Schedule'!{debt_column}26")
        sheet.cell(9, column, f"=MAX(0,{letter}7-{letter}8)")
        sheet.cell(10, column, f"='Management Equity'!$C$7*(1+Assumptions!$E$31)^{year}")
        sheet.cell(11, column, f"=MAX(0,{letter}9-{letter}10)")
        sheet.cell(12, column, f"=MIN({letter}9,{letter}9*Assumptions!$E$30+{letter}11*Assumptions!$E$32)")
        sheet.cell(13, column, f"=MAX(0,{letter}9-{letter}12)")
        sheet.cell(14, column, f"=IFERROR({letter}13/'Sources & Uses'!$F$8,0)")
        sheet.cell(15, column, f"=IFERROR(({letter}13/'Sources & Uses'!$F$8)^(1/{year})-1,0)")
        for row in range(5, 14):
            sheet.cell(row, column).number_format = CUR if row != 6 else MULT
        sheet.cell(14, column).number_format = MULT
        sheet.cell(15, column).number_format = PCT2
    selected_column = 10
    for row in range(5, 16):
        sheet.cell(row, selected_column, f"=INDEX(C{row}:I{row},1,Assumptions!E29)")
        sheet.cell(row, selected_column).number_format = PCT2 if row == 15 else (MULT if row in (6, 14) else CUR)
    set_widths(sheet, {"A": 4, "B": 34, **{get_column_letter(column): 13 for column in range(3, 11)}})

    sheet = workbook["Sensitivity"]
    title(sheet, "B2:G2", "Sponsor IRR Sensitivity")
    header(sheet, 4, 2, ["Entry / Exit multiple", 7.0, 8.0, 9.0, 10.0, 11.0])
    for column in range(3, 8):
        sheet.cell(4, column).number_format = MULT
    for row, entry_multiple in enumerate((8.0, 9.0, 10.0, 11.0, 12.0), start=5):
        sheet.cell(row, 2, entry_multiple)
        sheet.cell(row, 2).number_format = MULT
        for column in range(3, 8):
            letter = get_column_letter(column)
            sheet.cell(row, column, f"=IFERROR(((('Operating Model'!H8*{letter}$4-'Debt Schedule'!G26)*(1-Assumptions!$E$30))/MAX(0.01,('Operating Model'!C8*$B{row}+Assumptions!$E$15+'Sources & Uses'!C7+Assumptions!$E$16-'Sources & Uses'!F6-'Sources & Uses'!F7)))^(1/5)-1,0)")
            sheet.cell(row, column).number_format = PCT2
    sheet.conditional_formatting.add(
        "C5:G9",
        ColorScaleRule(
            start_type="min", start_color="FCE4D6",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="E2F0D9",
        ),
    )
    set_widths(sheet, {"A": 4, "B": 28, **{get_column_letter(column): 14 for column in range(3, 8)}})

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "LBO Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("Sources equal uses", '=IF(ABS(\'Sources & Uses\'!F10)<0.000001,"PASS","FAIL")'),
        ("Sponsor equity positive", '=IF(\'Sources & Uses\'!F8>0,"PASS","FAIL")'),
        ("Debt balances nonnegative", '=IF(MIN(\'Debt Schedule\'!D15:J25)>=0,"PASS","FAIL")'),
        ("Cash at or above minimum", '=IF(MIN(\'Debt Schedule\'!D13:J13)>=Assumptions!E16,"PASS","REVIEW")'),
        ("No covenant breaches", '=IF(COUNTIF(Covenants!C13:I13,"BREACH")=0,"PASS","REVIEW")'),
        ("Exit equity nonnegative", '=IF(MIN(\'Returns Waterfall\'!C9:I9)>=0,"PASS","FAIL")'),
        ("Waterfall conserves equity", '=IF(MAX(ABS(\'Returns Waterfall\'!C9:I9-\'Returns Waterfall\'!C12:I12-\'Returns Waterfall\'!C13:I13))<0.000001,"PASS","FAIL")'),
        ("Selected exit year valid", '=IF(AND(Assumptions!E29>=1,Assumptions!E29<=7),"PASS","FAIL")'),
        ("Overall", '=IF(COUNTIF(C5:C12,"FAIL")=0,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    add_status_rules(sheet, "C5:C13")
    set_widths(sheet, {"A": 4, "B": 46, "C": 18})

    add_sources(workbook, [
        ("Historical financial statements", "[filing / data room URL]", "[period]", "Revenue, margins, capex, working capital, taxes"),
        ("Quality of earnings and diligence", "[advisor report URL]", "[date]", "EBITDA adjustments, one-time items, risks"),
        ("Debt term sheets and documentation", "[data room URL]", "[date]", "Rates, PIK, amortization, sweeps, covenants, baskets"),
        ("Transaction terms", "[purchase agreement / process letter]", "[date]", "Price, fees, minimum cash, rollover, management pool"),
        ("Market valuation references", "[public market / transaction source]", "[date]", "Entry and exit multiple support"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("LBO_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
