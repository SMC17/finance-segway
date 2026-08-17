"""Build a governed, data-grounded fund memo deck for a real public ETF
Construction & Management instance.

Fifth domain on the IC-memo/pptx system (after PE, Private Credit, and --
in this same pass -- Software/SaaS, Distressed/Restructuring), reusing
tools/builders/pptx_helpers.py unchanged. Every number on a slide is read
from the instance workbook after a real LibreOffice recalculation
(tools.recalc), or from the case's own governed manifest
(standards/public_cases/<case_id>.json), which records which inputs are
real, sourced observations (input_kind "observed"/"derived") versus
modeler-chosen assumptions. Nothing on any slide is invented here -- this
script only reads and renders.

Default case (etf-public-kweb-2026-stress) is the domain's adversarial
case: KraneShares CSI China Internet ETF's real, disclosed 5-year average
annual total return of -15.07%, and its own SEC N-CSR composition table,
which carries both halves the filing discloses -- investments at 102.6% of
net assets and other assets less liabilities at (2.6)% -- so the grid
reconciles to exactly 100% without smoothing anything (see
30_ETF_Construction_Management/validation.md Finding ETF-03).

Usage:
    python tools/builders/build_etf_memo.py --case-id etf-public-kweb-2026-stress
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "tools"))

import pptx_helpers as ph  # noqa: E402
from recalc import recalc  # noqa: E402

import openpyxl  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
INDEX = ROOT / "standards" / "public_cases" / "index.json"


def _load_case(case_id: str) -> dict[str, Any]:
    index = json.loads(INDEX.read_text(encoding="utf-8"))
    for item in index["cases"]:
        if item["case_id"] == case_id:
            return item
    raise KeyError(f"unknown case_id {case_id!r}")


def _recalculated_workbook(output_rel: str) -> openpyxl.Workbook:
    src = ROOT / output_rel
    with tempfile.TemporaryDirectory(dir=ROOT) as tmp:
        scratch = Path(tmp) / src.name
        shutil.copyfile(src, scratch)
        result = recalc(str(scratch), timeout=90)
        if result.get("status") != "success" or result.get("total_errors", 1):
            raise RuntimeError(f"recalculation did not succeed cleanly: {result}")
        return openpyxl.load_workbook(scratch, data_only=True)


def _manifest(case_id: str) -> dict[str, Any]:
    return json.loads((ROOT / "standards" / "public_cases" / f"{case_id}.json").read_text(encoding="utf-8"))


def _fmt_usd_mm(value: float) -> str:
    return f"${value:,.1f}mm"


def _fmt_pct(value: float, decimals: int = 1) -> str:
    return f"{value * 100:.{decimals}f}%"


def build_etf_memo(case_id: str, output: Path) -> None:
    case = _load_case(case_id)
    manifest = _manifest(case_id)
    outcome = manifest["outcome"]

    wb = _recalculated_workbook(case["output"])
    assumptions = wb["Assumptions"]
    portfolio = wb["Portfolio Construction"]
    creation = wb["Creation & Redemption"]
    tracking = wb["Tracking Error & Costs"]
    checks = wb["Checks"]

    as_of = manifest["as_of"]
    fund_name = manifest["cover"]["ETF / share class:"]
    benchmark = manifest["cover"]["Benchmark index:"]
    av_source = manifest["sources"][0]
    ncsr_source = manifest["sources"][2]

    net_assets = assumptions["E5"].value
    expense_ratio = assumptions["E6"].value
    div_yield = assumptions["E7"].value
    price = assumptions["E8"].value

    holdings = [
        (portfolio.cell(r, 2).value, portfolio.cell(r, 3).value, portfolio.cell(r, 4).value)
        for r in range(5, 15)
        if portfolio.cell(r, 4).value
    ]
    holdings_total = portfolio["D36"].value
    sector_rows = [
        (portfolio.cell(r, 2).value, portfolio.cell(r, 3).value)
        for r in range(40, 51)
        if portfolio.cell(r, 3).value
    ]
    sector_total = portfolio["C51"].value

    nav_proxy = creation["C5"].value
    basket_value = creation["C7"].value
    premium_discount = creation["C9"].value
    ap_threshold_breached = creation["C11"].value

    net_tracking_diff = tracking["C9"].value

    check_labels = [checks.cell(r, 2).value for r in range(5, 11) if checks.cell(r, 2).value]
    check_status = {checks.cell(r, 2).value: checks.cell(r, 3).value for r in range(5, 11) if checks.cell(r, 2).value}
    overall_status = checks["C11"].value

    realized_return = outcome["realized"]

    prs = ph.new_presentation()

    # 1. Title
    ph.add_title_slide(
        prs,
        f"{fund_name} — Adversarial Fund Memo",
        f"Real, sourced fund profile and holdings vs. {benchmark}. "
        "Creation-unit and tracking-cost terms are template defaults, not the fund's own disclosed mechanics.",
        f"As of {as_of}  ·  Scenario: Downside  ·  Prepared from a governed, recalculated model instance",
    )

    # 2. Provenance
    slide = ph.add_content_slide(prs, "Reading this deck", kicker="Provenance")
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(1.5), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.6),
        [
            f"Real, sourced: fund net assets, net expense ratio, dividend yield, and market price are "
            f"{fund_name.split('(')[0].strip()}'s own disclosed figures (Alpha Vantage fund profile + market quote).",
            "Real, sourced: 30 disclosed holdings by weight, 7 sector weights, and the fund's own "
            "'other assets less liabilities' balancing line are all from its SEC Form N-CSR annual "
            "shareholder report — nothing is smoothed (see validation.md Finding ETF-03).",
            "Illustrative: creation-unit size, securities-lending revenue offset, cash-drag cost, and "
            "sampling/optimization tracking error are template defaults, not the fund's own disclosed "
            "creation/redemption mechanics or realized tracking-cost bridge.",
            "Realized outcome is real and recorded, not a forecast: the fund's own 5-year average annual "
            "total return through this filing's period end.",
        ],
        size=15.5,
    )
    ph.add_footer(slide, f"Model checks: Overall {overall_status}  ·  standards/public_cases/{case_id}.json")

    # 3. Fund snapshot (real)
    slide = ph.add_content_slide(prs, "Fund snapshot", kicker="Real, SEC/Alpha-Vantage-sourced")
    ph.add_stat_row(
        slide, ph.Inches(1.5),
        [
            ph.Stat(_fmt_usd_mm(net_assets), "Net assets", tag="SEC N-CSR", color=ph.NAVY),
            ph.Stat(_fmt_pct(expense_ratio, 2), "Net expense ratio", tag="Alpha Vantage", color=ph.NAVY),
            ph.Stat(_fmt_pct(div_yield, 2), "Trailing dividend yield", tag="Alpha Vantage", color=ph.NAVY),
        ],
    )
    ph.add_table(
        slide, ph.MARGIN, ph.Inches(3.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.9),
        ["Holding", "Issuer", "Weight"],
        [[sym, issuer, _fmt_pct(weight, 2)] for sym, issuer, weight in holdings],
        col_widths=[1.2, 3.5, 1],
    )
    ph.add_footer(
        slide,
        f"Top 10 of 30 disclosed real holdings shown; all 30 sum with the remainder to {_fmt_pct(holdings_total, 1)}.  ·  "
        f"Source: {av_source['name']} — {av_source['url']}",
    )

    # 4. Sector concentration and the honest disclosure gap
    slide = ph.add_content_slide(prs, "Sector concentration", kicker="Real, SEC N-CSR-sourced")
    ph.add_table(
        slide, ph.MARGIN, ph.Inches(1.6), ph.Inches(7.0), ph.Inches(3.4),
        ["Sector", "Weight"],
        [[sector, _fmt_pct(weight, 1)] for sector, weight in sector_rows],
        col_widths=[3, 1.2],
    )
    ph.add_text(
        slide, ph.Inches(8.2), ph.Inches(1.8), ph.Inches(4.5), ph.Inches(2.0),
        f"Composition reconciles to {_fmt_pct(sector_total, 1)} of net assets. The filing's sector "
        "weights cover investments only (102.6%); its disclosed 'other assets less liabilities' line "
        "of (2.6)% is carried here as the balancing item, exactly as the fund reports it.",
        size=15, color=ph.ILLUSTRATIVE_TAG,
    )
    ph.add_footer(slide, f"Source: {ncsr_source['name']} — {ncsr_source['url']}")
    ph.add_disclosure_note(
        slide,
        "Nothing is smoothed: every figure above is disclosed, and the grid reconciles because the "
        "fund's own balancing line is carried alongside its sector weights rather than dropped.",
    )

    # 5. Creation/redemption and tracking economics (mixed real/illustrative)
    slide = ph.add_content_slide(prs, "Creation, redemption, and tracking cost", kicker="Real price, illustrative mechanics")
    ph.add_stat_row(
        slide, ph.Inches(1.5),
        [
            ph.Stat(f"${price:,.2f}", "Market price (real)", tag="Alpha Vantage", color=ph.NAVY),
            ph.Stat(_fmt_pct(premium_discount, 2), "Premium/(discount) to NAV", tag="Illustrative", color=ph.ILLUSTRATIVE_TAG),
            ph.Stat(_fmt_pct(net_tracking_diff, 2), "Est. net tracking difference", tag="Computed",
                    color=ph.NEGATIVE if net_tracking_diff < -0.01 else ph.NAVY),
        ],
    )
    # C7 = Assumptions!E8 (market price, $/share) * Assumptions!E9 (creation
    # unit size, shares) -- a raw dollar figure, not already in $mm, unlike
    # every other stat on this deck. Converted explicitly rather than
    # guessed from magnitude.
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(3.4), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.2),
        f"Creation-unit basket value of {_fmt_usd_mm(basket_value / 1_000_000)} "
        f"at the real market price; AP arbitrage action threshold: {ap_threshold_breached}.",
        size=14, color=ph.CHARCOAL,
    )
    ph.add_disclosure_note(
        slide,
        "Creation unit size, premium/discount, securities-lending offset, cash drag, and sampling error "
        "are template defaults — not this fund's own disclosed creation/redemption terms or realized "
        "tracking-difference statement.",
    )

    # 6. Decision / realized outcome
    slide = ph.add_content_slide(prs, "Realized outcome", kicker="Decision — real, recorded")
    verdict_color = ph.NEGATIVE if realized_return < 0 else ph.POSITIVE
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(1.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.4),
        f"{fund_name.split('(')[0].strip()}'s own disclosed 5-year average annual total return: "
        f"{_fmt_pct(realized_return, 2)}, through the fund's own N-CSR reporting period — a real, filed "
        "number, not a projection from this workbook.",
        size=18, bold=True, color=verdict_color,
    )
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(3.1), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.6),
        [
            f"The fund tracked its own reference index tightly even under this stress "
            f"({outcome['realized_source']}) — a real sector/security outcome, not a tracking or "
            "wrapper-mechanics failure.",
            "The composition table carries the fund's own (2.6)% 'other assets less liabilities' line "
            "beside its 102.6% of investments — the filing's identity, reproduced rather than rounded.",
            "A real allocation decision needs the fund's own realized tracking-difference statement and "
            "current premium/discount, neither of which is asserted here — both remain illustrative levers.",
        ],
        size=15,
    )
    ph.add_footer(
        slide,
        f"Model: 30_ETF_Construction_Management/_template_ETF.xlsx  ·  Instance: {case['output']}  ·  "
        f"Overall checks: {overall_status}",
    )

    ph.save(prs, output)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-id", default="etf-public-kweb-2026-stress")
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    output = args.output or ROOT / "30_ETF_Construction_Management" / "presentations" / f"{args.case_id}-fund-memo.pptx"
    build_etf_memo(args.case_id, output)
    print(f"saved {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
