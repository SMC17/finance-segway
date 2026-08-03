import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[COMMODITY] — Commodities Model", [
    ("Commodity:", "[e.g. WTI Crude, Corn, Copper]"),
    ("Exchange:", "[NYMEX / CME / ICE]"),
    ("Last refreshed:", "[date]"),
    ("Next roll date:", "[date]"),
    ("Refresh cadence:", "Weekly (daily near roll/expiry)"),
])

# ---------------- FUTURES CURVE ----------------
ws = wb.create_sheet("Futures Curve")
set_col_widths(ws, [4, 16, 14, 14, 14, 14, 20])
ws["B2"] = "Futures Curve — Contango / Backwardation"; ws["B2"].font = TITLE
headers = ["", "Contract month", "Price", "Days to expiry", "Annualized basis %", "vs. front month", "Curve shape"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers)-1)

months = ["M1 (front)", "M2", "M3", "M6", "M12"]
r = 5
first_row = r
for m in months:
    ws.cell(row=r, column=2, value=m).font = BLACK
    c_price = ws.cell(row=r, column=3, value=0); c_price.font = BLUE; c_price.number_format = CUR2; c_price.border = BORDER
    c_days = ws.cell(row=r, column=4, value=0); c_days.font = BLUE; c_days.number_format = NUM; c_days.border = BORDER
    r += 1
last_row = r - 1
for rr in range(first_row, last_row + 1):
    ws.cell(row=rr, column=5,
            value=f"=IFERROR((C{rr}/$C${first_row}-1)*(365/D{rr}),\"-\")").number_format = PCT
    ws.cell(row=rr, column=5).border = BORDER
    ws.cell(row=rr, column=6,
            value=f"=IFERROR(C{rr}/$C${first_row}-1,\"-\")").number_format = PCT
    ws.cell(row=rr, column=6).border = BORDER
    ws.cell(row=rr, column=7,
            value=f'=IF(C{rr}>$C${first_row},"Contango",IF(C{rr}<$C${first_row},"Backwardation","Flat"))')
    ws.cell(row=rr, column=7).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- ROLL YIELD ----------------
ws = wb.create_sheet("Roll Yield")
set_col_widths(ws, [4, 26, 16, 16, 40])
ws["B2"] = "Roll Yield"; ws["B2"].font = TITLE
ws["B4"] = "Expiring contract price"; ws["C4"] = "='Futures Curve'!C5"; ws["C4"].font = GREEN; ws["C4"].number_format = CUR2
ws["B5"] = "Next contract price"; ws["C5"] = "='Futures Curve'!C6"; ws["C5"].font = GREEN; ws["C5"].number_format = CUR2
ws["B6"] = "Roll yield (%)"
ws["C6"] = "=IFERROR(C4/C5-1,\"-\")"; ws["C6"].font = BOLD; ws["C6"].number_format = PCT
ws["D6"] = "Positive = backwardation roll gain (long-only benefits). Negative = contango roll cost."
ws["D6"].font = ITALIC_GRAY
ws["B8"] = "Annual roll cost/gain estimate (12 rolls, illustrative)"
ws["C8"] = "=IFERROR(C6*12,\"-\")"; ws["C8"].number_format = PCT
ws.sheet_view.showGridLines = False

# ---------------- HEDGING ----------------
ws = wb.create_sheet("Hedging")
set_col_widths(ws, [4, 28, 16, 16, 40])
ws["B2"] = "Producer / Consumer Hedge Model"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Physical exposure (units)", 0, NUM),
    ("Futures contract size (units)", 1000, NUM),
    ("Spot price", 0, CUR2),
    ("Futures price (hedge contract)", 0, CUR2),
    ("Hedge ratio (beta-adjusted)", 1.0, '0.00'),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B11"] = "Outputs"; ws["B11"].font = BOLD; ws["B11"].fill = GRAY_FILL
ws["B12"] = "Contracts needed"
ws["C12"] = "=IFERROR(ROUND(C5*C9/C6,0),\"-\")"; ws["C12"].font = BOLD; ws["C12"].number_format = NUM
ws["B13"] = "Notional hedged"
ws["C13"] = "=C12*C6*C8"; ws["C13"].number_format = CUR
ws["B14"] = "Unhedged basis exposure"
ws["C14"] = "=C5*C7-C13"; ws["C14"].number_format = CUR
for r2 in (12, 13, 14):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- SENSITIVITY ----------------
ws = wb.create_sheet("Sensitivity")
set_col_widths(ws, [4, 20] + [12]*5)
ws["B2"] = "Sensitivity — P&L by Spot Move"; ws["B2"].font = TITLE
ws["B4"] = "Spot move %"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
moves = [-0.20, -0.10, 0.0, 0.10, 0.20]
for i, m in enumerate(moves, start=3):
    c = ws.cell(row=4, column=i, value=m); c.font = BOLD; c.fill = GRAY_FILL; c.number_format = PCT
ws["B5"] = "Unhedged P&L"; ws["B6"] = "Hedged P&L"
for i, m in enumerate(moves, start=3):
    col = get_column_letter(i)
    ws.cell(row=5, column=i, value=f"='Hedging'!$C$5*'Hedging'!$C$7*{col}4").number_format = CUR
    ws.cell(row=6, column=i, value=f"=({col}4*'Hedging'!$C$5*'Hedging'!$C$7)-({col}4*'Hedging'!$C$13)").number_format = CUR
    ws.cell(row=5, column=i).border = BORDER
    ws.cell(row=6, column=i).border = BORDER
ws.sheet_view.showGridLines = False

add_refresh_log(wb)

out_path = "/home/claude/model_shop/COMMODITIES_template.xlsx"
wb.save(out_path)
print("saved", out_path)
