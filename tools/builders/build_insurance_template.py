import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[INSURER] — Insurance / Actuarial Model", [
    ("Line of business:", "P&C / Life / Health / Reinsurance"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly (quarterly for reserves)"),
])

# ---------------- UNDERWRITING RATIOS ----------------
ws = wb.create_sheet("Underwriting Ratios")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Underwriting Performance"; ws["B2"].font = TITLE
inputs = [
    ("Earned premium ($)", 0, CUR),
    ("Incurred losses ($)", 0, CUR),
    ("Loss adjustment expenses ($)", 0, CUR),
    ("Underwriting expenses ($)", 0, CUR),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B10"] = "Loss ratio"; ws["C10"] = "=IFERROR((C6+C7)/C5,\"-\")"; ws["C10"].number_format = PCT
ws["B11"] = "Expense ratio"; ws["C11"] = "=IFERROR(C8/C5,\"-\")"; ws["C11"].number_format = PCT
ws["B12"] = "Combined ratio"; ws["C12"] = "=IFERROR(C10+C11,\"-\")"; ws["C12"].font = BOLD; ws["C12"].number_format = PCT
ws["D12"] = "<100% = underwriting profit; >100% = underwriting loss (offset by investment income)"
ws["D12"].font = ITALIC_GRAY
for r2 in (10, 11, 12):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- LOSS RESERVE TRIANGLE ----------------
ws = wb.create_sheet("Loss Reserve Triangle")
set_col_widths(ws, [4, 16] + [12]*6 + [14, 14])
ws["B2"] = "Loss Development Triangle — chain-ladder (cumulative paid losses, $)"; ws["B2"].font = TITLE
ws["B4"] = "Accident Yr \\ Dev Yr"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
dev_years = [f"Dev {i}" for i in range(1, 7)]
for i, dy in enumerate(dev_years, start=3):
    c = ws.cell(row=4, column=i, value=dy); c.font = BOLD; c.fill = GRAY_FILL
ws.cell(row=4, column=9, value="Ultimate Loss"); ws.cell(row=4, column=10, value="IBNR")
style_header_row(ws, 4, 2, start_col=9)
acc_years = [f"AY{2020+i}" for i in range(6)]
for j, ay in enumerate(acc_years, start=5):
    c = ws.cell(row=j, column=2, value=ay); c.font = BOLD; c.fill = GRAY_FILL
    # only fill lower-triangle cells that would actually exist (older years have more dev periods)
    n_periods = 6 - (j - 5)
    for i in range(3, 3 + n_periods):
        cell = ws.cell(row=j, column=i, value=0)
        cell.font = BLUE
        cell.fill = YELLOW_FILL
        cell.number_format = CUR
        cell.border = BORDER

# Volume-weighted age-to-age (link) factors: sum of column k+1 / sum of column k,
# over only the accident years that have data in both columns (upper-left of the triangle).
ws["B14"] = "Age-to-age factor"; ws["B14"].font = BOLD
link_factors = [
    ("D14", "=IFERROR(SUM(D5:D9)/SUM(C5:C9),\"-\")"),   # Dev1 -> Dev2, AY2020-2024
    ("E14", "=IFERROR(SUM(E5:E8)/SUM(D5:D8),\"-\")"),   # Dev2 -> Dev3, AY2020-2023
    ("F14", "=IFERROR(SUM(F5:F7)/SUM(E5:E7),\"-\")"),   # Dev3 -> Dev4, AY2020-2022
    ("G14", "=IFERROR(SUM(G5:G6)/SUM(F5:F6),\"-\")"),   # Dev4 -> Dev5, AY2020-2021
    ("H14", "=IFERROR(SUM(H5:H5)/SUM(G5:G5),\"-\")"),   # Dev5 -> Dev6, AY2020 only
]
for coord, formula in link_factors:
    ws[coord] = formula; ws[coord].number_format = '0.000'; ws[coord].border = BORDER

# Cumulative development factor (CDF) to ultimate at each dev period: product of every
# link factor from that period through Dev6. Dev6 is fully developed, CDF = 1.0.
ws["B15"] = "CDF to ultimate"; ws["B15"].font = BOLD
ws["H15"] = 1.0
ws["G15"] = "=IFERROR(H14*H15,\"-\")"
ws["F15"] = "=IFERROR(G14*G15,\"-\")"
ws["E15"] = "=IFERROR(F14*F15,\"-\")"
ws["D15"] = "=IFERROR(E14*E15,\"-\")"
ws["C15"] = "=IFERROR(D14*D15,\"-\")"
for col in range(3, 9):
    ws.cell(row=15, column=col).number_format = '0.000'
    ws.cell(row=15, column=col).border = BORDER

# Ultimate loss = latest diagonal (most recent paid figure for that AY) x CDF at that
# dev period. IBNR = Ultimate - latest paid (the reserve still to be recognized).
ultimate_rows = [
    (5, "H5", "H15"), (6, "G6", "G15"), (7, "F7", "F15"),
    (8, "E8", "E15"), (9, "D9", "D15"), (10, "C10", "C15"),
]
for row, latest_cell, cdf_cell in ultimate_rows:
    ws.cell(row=row, column=9, value=f"=IFERROR({latest_cell}*{cdf_cell},\"-\")").number_format = CUR
    ws.cell(row=row, column=10,
            value=f"=IFERROR(I{row}-{latest_cell},\"-\")").number_format = CUR
    ws.cell(row=row, column=9).border = BORDER
    ws.cell(row=row, column=10).border = BORDER

ws["B17"] = "Total ultimate / IBNR"; ws["B17"].font = BOLD
ws["I17"] = "=SUM(I5:I10)"; ws["I17"].font = BOLD; ws["I17"].number_format = CUR; ws["I17"].border = BORDER
ws["J17"] = "=SUM(J5:J10)"; ws["J17"].font = BOLD; ws["J17"].number_format = CUR; ws["J17"].border = BORDER
ws["B18"] = "Standard volume-weighted chain-ladder. Link factors use only accident years with data in both adjacent columns (upper-left triangle) — that's why each factor's denominator shrinks moving right."
ws["B18"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

# ---------------- EMBEDDED VALUE ----------------
ws = wb.create_sheet("Embedded Value")
set_col_widths(ws, [4, 32, 16, 40])
ws["B2"] = "Embedded Value (life/health simplified)"; ws["B2"].font = TITLE
inputs = [
    ("Adjusted net asset value (ANAV, $)", 0, CUR),
    ("PV of future profits on in-force business ($)", 0, CUR),
    ("Cost of holding required capital ($)", 0, CUR),
    ("Risk margin / cost of non-hedgeable risk ($)", 0, CUR),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B10"] = "Embedded Value = ANAV + PVFP - CoC - Risk margin"
ws["C10"] = "=C5+C6-C7-C8"; ws["C10"].font = BOLD; ws["C10"].number_format = CUR
ws["C10"].border = BORDER
ws["B12"] = "Value of new business (VNB) this period"; ws["C12"] = 0; ws["C12"].font = BLUE
ws["C12"].number_format = CUR; ws["C12"].fill = YELLOW_FILL; ws["C12"].border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- RESERVE SUMMARY ----------------
ws = wb.create_sheet("Reserve Summary")
set_col_widths(ws, [4, 34, 16, 40])
ws["B2"] = "Reserve Adequacy"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
res_inputs = [
    ("Case reserves (carried, $)", 0, CUR),
    ("Total policyholder surplus / equity ($)", 0, CUR),
]
r = 5
for label, default, fmt in res_inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B9"] = "Outputs"; ws["B9"].font = BOLD; ws["B9"].fill = GRAY_FILL
ws["B10"] = "IBNR (from Loss Reserve Triangle)"
ws["C10"] = "='Loss Reserve Triangle'!J17"; ws["C10"].font = GREEN; ws["C10"].number_format = CUR
ws["B11"] = "Total reserves (case + IBNR)"
ws["C11"] = "=IFERROR(C5+C10,\"-\")"; ws["C11"].font = BOLD; ws["C11"].number_format = CUR
ws["B12"] = "IBNR as % of total reserves"
ws["C12"] = "=IFERROR(C10/C11,\"-\")"; ws["C12"].number_format = PCT
ws["D12"] = "A rising share signals either faster growth or slower-than-assumed reporting/settlement — worth a reserve review either way"
ws["D12"].font = ITALIC_GRAY
ws["B13"] = "Reserves-to-surplus ratio"
ws["C13"] = "=IFERROR(C11/C6,\"-\")"; ws["C13"].font = BOLD; ws["C13"].number_format = MULT
ws["C13"].fill = YELLOW_FILL
ws["D13"] = "NAIC rule of thumb: >3.0x is a leverage warning sign for a P&C carrier"
ws["D13"].font = ITALIC_GRAY
for r2 in (10, 11, 12, 13):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- PROFITABILITY ----------------
ws = wb.create_sheet("Profitability")
set_col_widths(ws, [4, 32, 16, 40])
ws["B2"] = "Total Profitability (underwriting + investment)"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
prof_inputs = [
    ("Net investment income ($)", 0, CUR),
    ("Average policyholder surplus / equity ($)", 0, CUR),
]
r = 5
for label, default, fmt in prof_inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B9"] = "Outputs"; ws["B9"].font = BOLD; ws["B9"].fill = GRAY_FILL
ws["B10"] = "Underwriting result (earned premium x (1 - combined ratio))"
ws["C10"] = "=IFERROR('Underwriting Ratios'!C5*(1-'Underwriting Ratios'!C12),\"-\")"
ws["C10"].font = GREEN; ws["C10"].number_format = CUR
ws["B11"] = "Total profit (underwriting result + net investment income)"
ws["C11"] = "=IFERROR(C10+C5,\"-\")"; ws["C11"].font = BOLD; ws["C11"].number_format = CUR
ws["B12"] = "Return on equity (ROE)"
ws["C12"] = "=IFERROR(C11/C6,\"-\")"; ws["C12"].font = BOLD; ws["C12"].number_format = PCT
ws["D12"] = "Why combined ratio alone can mislead: a carrier can run >100% combined and still be profitable if investment income (float x yield) covers the gap"
ws["D12"].font = ITALIC_GRAY
for r2 in (10, 11, 12):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "INSURANCE_template.xlsx"
wb.save(out_path)
print("saved", out_path)
