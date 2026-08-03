import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[TOKEN] — Crypto / Digital Asset Model", [
    ("Protocol:", "[fill in]"),
    ("Chain / L1-L2:", "[fill in]"),
    ("Last refreshed:", "[date]"),
    ("Next unlock/emission event:", "[date]"),
    ("Refresh cadence:", "Weekly (daily around unlocks)"),
])

# ---------------- TOKENOMICS ----------------
ws = wb.create_sheet("Tokenomics")
set_col_widths(ws, [4, 26, 14, 14, 14, 30])
ws["B2"] = "Tokenomics — Supply Schedule"; ws["B2"].font = TITLE
headers = ["", "Allocation", "Tokens", "% of max supply", "Unlock start", "Vesting notes"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 3)

allocs = ["Public / community", "Team & advisors", "Investors (seed/private)",
          "Treasury / foundation", "Ecosystem incentives / staking rewards"]
r = 5
for a in allocs:
    ws.cell(row=r, column=2, value=a).font = BLACK
    c = ws.cell(row=r, column=3, value=0); c.font = BLUE; c.number_format = NUM; c.border = BORDER
    ws.cell(row=r, column=5, value="[date]").font = BLUE
    ws.cell(row=r, column=6, value="[cliff / linear vesting terms]").font = ITALIC_GRAY
    r += 1
total_row = r
ws.cell(row=total_row, column=2, value="Max supply").font = BOLD
ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})").font = BOLD
ws.cell(row=total_row, column=3).number_format = NUM
for rr in range(5, total_row):
    c = ws.cell(row=rr, column=4, value=f"=IFERROR(C{rr}/$C${total_row},\"-\")")
    c.number_format = PCT

ws.cell(row=total_row+2, column=2, value="Circulating supply (today)").font = BOLD
ws.cell(row=total_row+2, column=3, value=0).font = BLUE
ws.cell(row=total_row+2, column=3).fill = YELLOW_FILL
ws.cell(row=total_row+2, column=3).number_format = NUM
ws.cell(row=total_row+3, column=2, value="Circulating / max supply %").font = BLACK
ws.cell(row=total_row+3, column=3,
        value=f"=IFERROR(C{total_row+2}/C{total_row},\"-\")").number_format = PCT
ws.sheet_view.showGridLines = False
circ_row = total_row + 2
max_row = total_row

# ---------------- VALUATION MULTIPLES ----------------
ws = wb.create_sheet("Valuation")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "On-Chain Valuation Multiples"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Token price", 0, CUR2),
    ("Total value locked (TVL, $)", 0, CUR),
    ("Annualized protocol revenue/fees ($)", 0, CUR),
    ("Daily transaction volume ($, trailing 30d avg)", 0, CUR),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B10"] = "Outputs"; ws["B10"].font = BOLD; ws["B10"].fill = GRAY_FILL
ws["B11"] = "Market cap (circulating)"
ws["C11"] = f"=C5*Tokenomics!C{circ_row}"; ws["C11"].number_format = CUR
ws["B12"] = "Fully diluted valuation (FDV)"
ws["C12"] = f"=C5*Tokenomics!C{max_row}"; ws["C12"].number_format = CUR
ws["B13"] = "Market cap / TVL"
ws["C13"] = "=IFERROR(C11/C6,\"-\")"; ws["C13"].number_format = MULT
ws["B14"] = "FDV / TVL"
ws["C14"] = "=IFERROR(C12/C6,\"-\")"; ws["C14"].number_format = MULT
ws["B15"] = "P/S equivalent (Mkt cap / annualized revenue)"
ws["C15"] = "=IFERROR(C11/C7,\"-\")"; ws["C15"].number_format = MULT
ws["B16"] = "NVT ratio (Mkt cap / daily volume, annualized)"
ws["C16"] = "=IFERROR(C11/(C8*365),\"-\")"; ws["C16"].number_format = MULT
ws["D16"] = "High NVT = valuation rich relative to network usage (P/E-style read)"
ws["D16"].font = ITALIC_GRAY
for r2 in range(11, 17):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- STAKING / YIELD ----------------
ws = wb.create_sheet("Staking Yield")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Staking / Yield Model"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("% of supply staked", 0.0, PCT),
    ("Annual token emissions (new supply, tokens)", 0, NUM),
    ("Protocol fee revenue distributed to stakers ($)", 0, CUR),
    ("Token price (for $ yield calc)", 0, CUR2),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B10"] = "Outputs"; ws["B10"].font = BOLD; ws["B10"].fill = GRAY_FILL
ws["B11"] = "Tokens staked"
ws["C11"] = f"=C5*Tokenomics!C{circ_row}"; ws["C11"].number_format = NUM
ws["B12"] = "Inflationary yield (emissions / staked)"
ws["C12"] = "=IFERROR(C6/C11,\"-\")"; ws["C12"].number_format = PCT
ws["B13"] = "Real yield (fee revenue / staked value)"
ws["C13"] = "=IFERROR(C7/(C11*C8),\"-\")"; ws["C13"].number_format = PCT
ws["B14"] = "Net staking yield (real - inflation dilution)"
ws["C14"] = "=IFERROR(C13-C12,\"-\")"; ws["C14"].font = BOLD; ws["C14"].number_format = PCT
ws["D14"] = "Negative = staking rewards are pure dilution, not real cash flow"
ws["D14"].font = ITALIC_GRAY
for r2 in range(11, 15):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- COMPARABLE PROTOCOLS ----------------
ws = wb.create_sheet("Comparable Protocols")
set_col_widths(ws, [4, 18, 14, 14, 14, 14, 12, 12, 12])
ws["B2"] = "Comparable Protocols"; ws["B2"].font = TITLE
headers = ["", "Protocol", "Mkt cap", "FDV", "TVL", "Daily volume", "Mkt/TVL", "FDV/TVL", "NVT"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 8)

# Row 5 auto-pulls this protocol's own stats from the Valuation tab so it's
# always the first, always-current comp — everything below is peer data the
# user fills in by hand.
ws["B5"] = "This protocol"; ws["B5"].font = BOLD
ws["C5"] = "=Valuation!C11"; ws["D5"] = "=Valuation!C12"
ws["E5"] = "=Valuation!C6"; ws["F5"] = "=Valuation!C8"
for c in (3, 4, 5, 6):
    ws.cell(row=5, column=c).font = GREEN
    ws.cell(row=5, column=c).number_format = CUR
    ws.cell(row=5, column=c).border = BORDER

for r in range(6, 12):
    ws.cell(row=r, column=2, value="[fill in]").font = BLUE
    ws.cell(row=r, column=2).border = BORDER
    for c in (3, 4, 5, 6):
        cell = ws.cell(row=r, column=c, value=0)
        cell.font = BLUE; cell.number_format = CUR; cell.border = BORDER

for r in range(5, 12):
    ws.cell(row=r, column=7, value=f"=IFERROR(C{r}/E{r},\"-\")").number_format = MULT
    ws.cell(row=r, column=8, value=f"=IFERROR(D{r}/E{r},\"-\")").number_format = MULT
    ws.cell(row=r, column=9, value=f"=IFERROR(C{r}/(F{r}*365),\"-\")").number_format = MULT
    for c in (7, 8, 9):
        ws.cell(row=r, column=c).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- SUPPLY EMISSION & DILUTION ----------------
ws = wb.create_sheet("Supply Emission Schedule")
set_col_widths(ws, [4, 32, 14, 14, 14, 14, 14, 40])
ws["B2"] = "Forward Supply Emission & Dilution"; ws["B2"].font = TITLE
ws["B4"] = "Annual emission rate (% of max supply issued/yr)"
ws["C4"] = 0.05; ws["C4"].font = BLUE; ws["C4"].fill = YELLOW_FILL; ws["C4"].number_format = PCT
ws["C4"].border = BORDER
ws["D4"] = "Simplified: assumes a flat schedule. Real unlocks are usually cliff + linear per allocation — treat this as a planning approximation, not a substitute for the actual unlock table."
ws["D4"].font = ITALIC_GRAY

for i, h in enumerate(["", "", "Period 0", "Period 1", "Period 2", "Period 3", "Period 4"], start=1):
    ws.cell(row=6, column=i, value=h)
style_header_row(ws, 6, 5, start_col=3)
ws["B7"] = "Circulating supply"
ws["C7"] = f"=Tokenomics!C{circ_row}"; ws["C7"].font = GREEN
for col in range(4, 8):
    prev = get_column_letter(col - 1)
    ws.cell(row=7, column=col, value=f"=MIN(Tokenomics!$C${max_row},{prev}7+Tokenomics!$C${max_row}*$C$4)")
for col in range(3, 8):
    ws.cell(row=7, column=col).number_format = NUM
    ws.cell(row=7, column=col).border = BORDER

ws["B8"] = "Cumulative dilution vs. today"
for col in range(3, 8):
    letter = get_column_letter(col)
    ws.cell(row=8, column=col, value=f"=IFERROR(1-$C$7/{letter}7,\"-\")")
    ws.cell(row=8, column=col).number_format = PCT
    ws.cell(row=8, column=col).border = BORDER
ws["B9"] = "Implied mkt cap at today's price (if price held flat)"
for col in range(3, 8):
    letter = get_column_letter(col)
    ws.cell(row=9, column=col, value=f"=Valuation!$C$5*{letter}7")
    ws.cell(row=9, column=col).number_format = CUR
    ws.cell(row=9, column=col).border = BORDER
ws["B10"] = ("A holder who doesn't add to their position loses this much ownership share to new "
             "issuance over time -- the price needs to grow faster than dilution just to keep the "
             "position's dollar value from falling.")
ws["B10"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

add_refresh_log(wb)

out_path = "CRYPTO_template.xlsx"
wb.save(out_path)
print("saved", out_path)
