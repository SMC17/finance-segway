"""Canonical release wrapper for the institutional quantitative builder."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_quant_institutional import build as build_layout  # noqa: E402
from formula_compatibility import normalize_workbook_formulas  # noqa: E402
from institutional_helpers import PCT2, finalize  # noqa: E402


def build(output: Path) -> None:
    build_layout(output)
    workbook = load_workbook(output)
    walk_forward = workbook["Walk Forward"]
    # PRODUCT(1+range) is an array expression in some spreadsheet engines.
    # Each test fold is six periods, so enumerate the six scalar observations.
    for row in range(5, 11):
        factors = [
            f"(1+INDEX(Backtest!$H$5:$H$64,E{row}+{offset}))"
            for offset in range(6)
        ]
        walk_forward.cell(row, 9, "=" + "*".join(factors) + "-1")
        walk_forward.cell(row, 9).number_format = PCT2
    normalize_workbook_formulas(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("QUANT_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
