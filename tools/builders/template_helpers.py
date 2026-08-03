"""Shared styling helpers — imported by every build_*_template.py script."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
BOLD_WHITE = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=16, bold=True)
ITALIC_GRAY = Font(name="Arial", size=8, italic=True, color="808080")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
MULT = '0.0x'
NUM = '#,##0;(#,##0);"-"'


def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_refresh_log(wb):
    ws = wb.create_sheet("RefreshLog")
    set_col_widths(ws, [4, 14, 16, 30, 30, 16])
    ws["B2"] = "Refresh Log"; ws["B2"].font = TITLE
    for i, h in enumerate(["", "Date", "Trigger", "What changed", "Reviewer notes", "Next check"], start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 5, start_col=2)
    for r in range(5, 15):
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = BORDER
    ws.sheet_view.showGridLines = False
    return ws


def add_cover(wb, title, fields):
    """fields: list of (label, default_value) tuples"""
    ws = wb.create_sheet("Cover")
    set_col_widths(ws, [4, 30, 40, 4])
    ws["B2"] = title; ws["B2"].font = TITLE
    r = 4
    for label, default in fields:
        ws.cell(row=r, column=2, value=label).font = BOLD
        c = ws.cell(row=r, column=3, value=default)
        c.font = BLUE
        r += 1
    ws.sheet_view.showGridLines = False
    return ws
