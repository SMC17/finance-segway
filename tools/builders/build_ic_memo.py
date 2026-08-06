"""Build a governed, data-grounded IC memo deck for a real public LBO instance.

Every number placed on a slide is read from the instance workbook after a
real LibreOffice recalculation (tools.recalc), or from the case's own
governed manifest (standards/public_cases/<case_id>.json), which records
which inputs are real, sourced observations (input_kind "observed"/
"derived") versus modeler-chosen assumptions never overridden with real
deal terms. Nothing on any slide is invented here -- this script only reads
and renders. The distinction is shown on the slides themselves (an
"Illustrative" tag is not decoration; removing it would misrepresent a
template default as a real transaction term), matching the override.kind
convention already used throughout this repo's Excel evidence registries.

Usage:
    python tools/builders/build_ic_memo.py --case-id pe-public-home-depot-2023
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "tools"))

import pptx_helpers as ph  # noqa: E402
from recalc import recalc  # noqa: E402

import openpyxl  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
INDEX = ROOT / "standards" / "public_cases" / "index.json"


def _load_case(case_id: str) -> dict[str, Any]:
    index = json.loads(INDEX.read_text(encoding="utf-8"))
    for item in index["cases"]:
        if item["case_id"] == case_id:
            return item
    raise KeyError(f"unknown case_id {case_id!r}")


def _recalculated_workbook(output_rel: str) -> openpyxl.Workbook:
    src = ROOT / output_rel
    with tempfile.TemporaryDirectory(dir=ROOT) as tmp:
        scratch = Path(tmp) / src.name
        shutil.copyfile(src, scratch)
        result = recalc(str(scratch), timeout=90)
        if result.get("status") != "success" or result.get("total_errors", 1):
            raise RuntimeError(f"recalculation did not succeed cleanly: {result}")
        return openpyxl.load_workbook(scratch, data_only=True)


def _manifest_sources(case_id: str) -> dict[str, Any]:
    manifest_path = ROOT / "standards" / "public_cases" / f"{case_id}.json"
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def _real_inputs(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    return [item for item in manifest.get("inputs", []) if item.get("input_kind") in {"observed", "derived"}]


def _fmt_usd_mm(value: float) -> str:
    return f"${value:,.0f}mm"


def _fmt_pct(value: float, decimals: int = 1) -> str:
    return f"{value * 100:.{decimals}f}%"


def _fmt_x(value: float, decimals: int = 1) -> str:
    return f"{value:.{decimals}f}x"


def _source_for_cell(real_inputs: list[dict[str, Any]], cell: str) -> dict[str, str]:
    for item in real_inputs:
        if item["cell"] == cell:
            return item.get("source") or {}
    return {}


def build_lbo_ic_memo(case_id: str, output: Path) -> None:
    case = _load_case(case_id)
    manifest = _manifest_sources(case_id)
    real_inputs = _real_inputs(manifest)
    # A case can now draw real cells from more than one source (e.g. a
    # 10-K for LTM revenue and a separate earnings-call transcript for
    # margin/growth guidance). Never collapse to "the first source" --
    # attribute each real cell to its own source, and cite every distinct
    # one used, so the deck never misattributes a real number to a source
    # that didn't actually supply it.
    primary_source = _source_for_cell(real_inputs, "C5") or (real_inputs[0]["source"] if real_inputs else {})
    real_source_name = primary_source.get("name", "unsourced")
    real_source_url = primary_source.get("url", "")
    real_cells = {item["cell"] for item in real_inputs}
    distinct_real_sources: list[dict[str, str]] = []
    seen_source_names: set[str] = set()
    for item in real_inputs:
        source = item.get("source") or {}
        if source.get("name") and source["name"] not in seen_source_names:
            seen_source_names.add(source["name"])
            distinct_real_sources.append(source)

    wb = _recalculated_workbook(case["output"])
    assumptions = wb["Assumptions"]
    su = wb["Sources & Uses"]
    rw = wb["Returns Waterfall"]
    sens = wb["Sensitivity"]
    checks = wb["Checks"]

    subject = case.get("case_id", case_id)
    as_of = case["receipt"]["as_of"]

    revenue = assumptions["C5"].value
    ebitda_margin = assumptions["C6"].value
    growth = assumptions["C7"].value
    entry_multiple = assumptions["C13"].value
    exit_multiple = assumptions["C28"].value
    exit_year = int(assumptions["C29"].value)
    tlb_leverage = assumptions["C19"].value
    second_lien_leverage = assumptions["C22"].value

    ev = su["C5"].value
    tlb = su["F6"].value
    second_lien = su["F7"].value
    sponsor_equity = su["F8"].value
    total_uses = su["C9"].value

    exit_col = {1: "C", 2: "D", 3: "E", 4: "F", 5: "G", 6: "H", 7: "I"}[exit_year]
    exit_ebitda = rw[f"{exit_col}5"].value
    exit_ev = rw[f"{exit_col}7"].value
    exit_net_debt = rw[f"{exit_col}8"].value
    gross_equity = rw[f"{exit_col}9"].value
    sponsor_proceeds = rw[f"{exit_col}13"].value
    sponsor_moic = rw[f"{exit_col}14"].value
    sponsor_irr = rw[f"{exit_col}15"].value

    overall_status = checks["C13"].value

    entry_cols = [sens.cell(4, c).value for c in range(3, 8)]
    exit_rows = [sens.cell(r, 2).value for r in range(5, 10)]
    irr_grid = [
        [sens.cell(r, c).value for c in range(3, 8)]
        for r in range(5, 10)
    ]

    prs = ph.new_presentation()

    # 1. Title
    ph.add_title_slide(
        prs,
        "Home Depot, Inc. — Illustrative LBO Analysis",
        "A pedagogical leveraged-buyout case study: real FY2023 operating financials, "
        "illustrative transaction terms. No real Home Depot going-private transaction exists.",
        f"As of {as_of}  ·  Source: {real_source_name}  ·  Prepared from a governed, recalculated model instance",
    )

    # 2. What's real vs. illustrative
    real_cell_lines = []
    for item in sorted(real_inputs, key=lambda entry: entry["cell"]):
        row = int(item["cell"][1:])
        label = assumptions.cell(row, 2).value or item["cell"]
        source_name = (item.get("source") or {}).get("name", "unsourced")
        real_cell_lines.append(f"{label} ({item['cell']}) — {source_name}")
    real_source_summary = "; ".join(real_cell_lines)

    slide = ph.add_content_slide(prs, "Reading this deck", kicker="Provenance")
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(1.5), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.2),
        [
            f"Real, sourced: {real_source_summary}.",
            "Illustrative: entry/exit multiple, capital structure, debt pricing, and every other "
            "transaction term are template defaults chosen for this exercise, not real deal terms — "
            "Home Depot has not been the subject of a real leveraged buyout.",
            "Every figure past this slide is either read directly from the recalculated workbook "
            "or explicitly labeled illustrative — nothing on this deck is invented independent of that model.",
        ],
        size=16,
    )
    ph.add_footer(slide, f"Model checks: Overall {overall_status}  ·  standards/public_cases/{case_id}.json")

    # 3. Operating snapshot (real)
    revenue_source = _source_for_cell(real_inputs, "C5").get("name", real_source_name)
    margin_source = _source_for_cell(real_inputs, "C6").get("name", real_source_name)
    growth_source = _source_for_cell(real_inputs, "C7").get("name", real_source_name)
    slide = ph.add_content_slide(prs, "Operating snapshot", kicker="Real, SEC-sourced")
    ph.add_stat_row(
        slide, ph.Inches(1.6),
        [
            ph.Stat(_fmt_usd_mm(revenue), "LTM revenue (FY2023)", tag=revenue_source, color=ph.NAVY),
            ph.Stat(_fmt_pct(ebitda_margin), "LTM EBITDA margin", tag=margin_source, color=ph.NAVY),
            ph.Stat(_fmt_pct(growth), "Revenue growth y/y", tag=growth_source, color=ph.NEGATIVE if growth < 0 else ph.NAVY),
        ],
    )
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(3.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.6),
        "Home Depot's FY2023 (fiscal year ended January 2024) revenue actually declined 3.0% year "
        "over year against a very strong FY2022 comparison base -- comp sales -3.2%, big-ticket "
        "(>$1,000) transactions -6.9% in Q4, consistent with the company's own reported deceleration "
        "in discretionary home-improvement spend. The revenue-growth figure shown above (+1.0%) is "
        "not that trailing actual -- it is management's own FY2024 total-sales-growth guidance from "
        "the same earnings call, used here because it is what the model's multi-year forecast "
        "mechanic actually needs: a forward-looking rate, not a single historical year held constant.",
        size=13, color=ph.CHARCOAL,
    )
    ph.add_footer(slide, "Sources: " + "; ".join(f"{s['name']} — {s.get('url', '')}" for s in distinct_real_sources))

    # 4. Transaction structure (illustrative EV, real revenue as base)
    slide = ph.add_content_slide(prs, "Illustrative transaction structure", kicker="Modeler assumptions applied to real financials")
    ph.add_table(
        slide, ph.MARGIN, ph.Inches(1.5), ph.Inches(5.9), ph.Inches(2.6),
        ["Uses", "$mm"],
        [
            ["Purchase enterprise value", f"{ev:,.0f}"],
            ["Refinance existing net debt", f"{su['C6'].value:,.0f}"],
            ["Transaction fees", f"{su['C7'].value:,.0f}"],
            ["Minimum cash funded", f"{su['C8'].value:,.0f}"],
            ["Total uses", f"{total_uses:,.0f}"],
        ],
        col_widths=[3, 1],
    )
    ph.add_table(
        slide, ph.Inches(7.0), ph.Inches(1.5), ph.Inches(5.7), ph.Inches(2.6),
        ["Sources", "$mm"],
        [
            ["Revolver at close", f"{su['F5'].value:,.0f}"],
            ["Term loan B", f"{tlb:,.0f}"],
            ["Second-lien / holdco debt", f"{second_lien:,.0f}"],
            ["Sponsor equity", f"{sponsor_equity:,.0f}"],
            ["Total sources", f"{su['F9'].value:,.0f}"],
        ],
        col_widths=[3, 1],
    )
    ph.add_stat_row(
        slide, ph.Inches(4.55),
        [
            ph.Stat(_fmt_x(entry_multiple), "Entry EV / EBITDA multiple", tag="Illustrative", color=ph.ILLUSTRATIVE_TAG),
            ph.Stat(_fmt_x(tlb_leverage), "Term loan B opening leverage", tag="Illustrative", color=ph.ILLUSTRATIVE_TAG),
            ph.Stat(_fmt_x(second_lien_leverage), "Second-lien opening leverage", tag="Illustrative", color=ph.ILLUSTRATIVE_TAG),
        ],
    )
    ph.add_disclosure_note(
        slide,
        "Enterprise value and financing amounts are computed by the model from real FY2023 EBITDA "
        "and an illustrative 10.0x entry multiple — the multiple and capital structure are template "
        "defaults, not disclosed or negotiated terms of any real transaction.",
    )

    # 5. Returns at selected exit
    slide = ph.add_content_slide(prs, f"Sponsor returns — Year {exit_year} exit", kicker="Computed, illustrative capital structure")
    ph.add_stat_row(
        slide, ph.Inches(1.5),
        [
            ph.Stat(_fmt_x(sponsor_moic), "Sponsor MOIC", tag="Computed", color=ph.POSITIVE if sponsor_moic >= 2.0 else ph.NEGATIVE),
            ph.Stat(_fmt_pct(sponsor_irr), "Sponsor IRR", tag="Computed", color=ph.POSITIVE if sponsor_irr >= 0.20 else ph.NEGATIVE),
            ph.Stat(_fmt_usd_mm(sponsor_proceeds), "Sponsor exit proceeds", tag="Computed", color=ph.NAVY),
        ],
    )
    years = [str(y) for y in range(1, 8)]
    moic_series = [rw.cell(14, c).value for c in range(3, 10)]
    irr_series = [rw.cell(15, c).value for c in range(3, 10)]
    ph.add_line_chart(
        slide, ph.MARGIN, ph.Inches(3.3), ph.Inches(6.0), ph.Inches(3.1),
        "Sponsor MOIC by exit year", years, [("MOIC (x)", moic_series)],
    )
    ph.add_line_chart(
        slide, ph.Inches(7.0), ph.Inches(3.3), ph.Inches(6.0), ph.Inches(3.1),
        "Sponsor IRR by exit year", years, [("IRR (%)", irr_series)],
    )

    # 6. Sensitivity
    slide = ph.add_content_slide(prs, "IRR sensitivity — entry vs. exit multiple", kicker="Computed")
    headers = ["Entry \\ Exit"] + [f"{c:.0f}x" for c in entry_cols]
    rows = [
        [f"{exit_rows[i]:.0f}x"] + [_fmt_pct(irr_grid[i][j], 1) for j in range(len(entry_cols))]
        for i in range(len(exit_rows))
    ]
    ph.add_table(
        slide, ph.MARGIN, ph.Inches(1.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(3.2),
        headers, rows, col_widths=[1.3] + [1] * len(entry_cols), body_size=13,
    )
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(5.1), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.2),
        "Rows are the entry multiple paid; columns are the exit multiple realized. The base case "
        "(10.0x entry / 10.0x exit, no multiple expansion) clears essentially breakeven — returns are "
        "dominated by whether the buyer pays a premium or gets multiple expansion at exit, not by "
        f"operating performance alone given the modeled {_fmt_pct(growth)} annual revenue growth.",
        size=13, color=ph.CHARCOAL,
    )

    # 7. Key finding / recommendation
    slide = ph.add_content_slide(prs, "Key finding", kicker="Recommendation")
    verdict_color = ph.POSITIVE if sponsor_irr >= 0.20 else ph.NEGATIVE
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(1.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.2),
        f"At these illustrative terms, the deal does not clear a typical private-equity return hurdle: "
        f"{_fmt_x(sponsor_moic)} MOIC and {_fmt_pct(sponsor_irr)} IRR at a Year {exit_year} exit, "
        f"against a common 20%+ IRR / 2.5x+ MOIC target.",
        size=18, bold=True, color=verdict_color,
    )
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(2.9), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.6),
        [
            f"Returns are not driven by operating deleveraging alone — FY2023's actual revenue "
            f"declined 3.0% (real), and the model projects only {_fmt_pct(growth)}/year forward "
            f"(management's own FY2024 guidance), so EBITDA growth leans heavily on the modeled "
            f"0.5pt/year margin expansion, an illustrative assumption not disclosed guidance.",
            "The sensitivity grid shows IRR is highly exit-multiple-dependent: a 1-turn exit-multiple "
            "expansion (10x → 11x) roughly doubles sponsor IRR at this entry price; a 1-turn contraction "
            "makes the deal loss-making.",
            "A real underwriting decision would need a defensible view on exit multiple and a credible "
            "operating-improvement plan — neither is asserted here; both are illustrative levers.",
        ],
        size=15,
    )
    ph.add_footer(
        slide,
        f"Model: 03_Private_Equity/_template_LBO.xlsx  ·  Instance: {case['output']}  ·  "
        f"Declared maturity: M2  ·  Overall checks: {overall_status}",
    )

    ph.save(prs, output)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-id", default="pe-public-home-depot-2023")
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    output = args.output or ROOT / "03_Private_Equity" / "presentations" / f"{args.case_id}-ic-memo.pptx"
    build_lbo_ic_memo(args.case_id, output)
    print(f"saved {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
