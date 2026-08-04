"""Release-grade commodities workbook with carry, physical balance, and controls."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl.utils import get_column_letter

try:
    from tools.builders.legacy_release_adapter import build_release
    from tools.builders.template_helpers import (
        BLACK,
        BLUE,
        BOLD,
        BORDER,
        CUR,
        CUR2,
        GRAY_FILL,
        HEADER_FILL,
        ITALIC_GRAY,
        MULT,
        NUM,
        PCT,
        PCT2,
        TITLE,
        YELLOW_FILL,
        set_col_widths,
        style_header_row,
    )
except ModuleNotFoundError:
    from legacy_release_adapter import build_release
    from template_helpers import (
        BLACK,
        BLUE,
        BOLD,
        BORDER,
        CUR,
        CUR2,
        GRAY_FILL,
        HEADER_FILL,
        ITALIC_GRAY,
        MULT,
        NUM,
        PCT,
        PCT2,
        TITLE,
        YELLOW_FILL,
        set_col_widths,
        style_header_row,
    )


def _status_formula(check_range: str) -> str:
    return f'=IF(COUNTIF({check_range},"FAIL")+COUNTIF({check_range},"BREACH")>0,"BREACH",IF(COUNTIF({check_range},"REVIEW")>0,"REVIEW","PASS"))'


def enrich(workbook) -> None:
    if "Physical Balance & Carry" in workbook.sheetnames:
        del workbook["Physical Balance & Carry"]
    if "Decision & Checks" in workbook.sheetnames:
        del workbook["Decision & Checks"]

    insert_at = workbook.sheetnames.index("Sensitivity")
    sheet = workbook.create_sheet("Physical Balance & Carry", insert_at)
    set_col_widths(sheet, [4, 38, 18, 18, 18, 34])
    sheet["B2"] = "Physical Balance, Cost of Carry, and Basis"
    sheet["B2"].font = TITLE
    headers = ["Metric", "Base", "Downside", "Active", "Units / source"]
    for column, value in enumerate(headers, start=2):
        sheet.cell(4, column, value)
    style_header_row(sheet, 4, len(headers), start_col=2)

    inputs = [
        ("Spot price", "='Hedging'!C7", "='Hedging'!C7*0.70", CUR2, "linked to hedge spot"),
        ("Observed futures price", "='Hedging'!C8", "='Hedging'!C8*1.10", CUR2, "linked to hedge contract"),
        ("Risk-free rate", 0.04, 0.07, PCT2, "annual continuously compounded"),
        ("Storage and financing rate", 0.02, 0.20, PCT2, "annual cost of carry"),
        ("Convenience yield", 0.03, 0.00, PCT2, "annual non-cash benefit"),
        ("Time to delivery", 0.50, 0.25, "0.00", "years"),
        ("Beginning inventory", 100.0, 20.0, NUM, "physical units"),
        ("Production", 50.0, 0.0, NUM, "physical units"),
        ("Purchases", 20.0, 10.0, NUM, "physical units"),
        ("Sales", 80.0, 25.0, NUM, "physical units"),
        ("Internal consumption / shrink", 30.0, 5.0, NUM, "physical units"),
        ("Reported ending inventory", 60.0, 0.0, NUM, "physical units"),
        ("Basis warning threshold", 10.0, 5.0, CUR2, "absolute spot-futures difference"),
        ("Maximum hedge ratio", 1.25, 1.10, MULT, "absolute futures units / physical units"),
    ]
    for row, (label, base, downside, number_format, note) in enumerate(inputs, start=5):
        sheet.cell(row, 2, label).font = BLACK
        for column, value in ((3, base), (4, downside)):
            cell = sheet.cell(row, column, value)
            cell.font = BLUE
            cell.fill = YELLOW_FILL
            cell.number_format = number_format
            cell.border = BORDER
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 5).border = BORDER
        sheet.cell(row, 6, note)

    sheet["B21"] = "Derived outputs"
    sheet["B21"].font = BOLD
    sheet["B21"].fill = GRAY_FILL
    outputs = [
        (22, "Net carry rate", "=E7+E8-E9", PCT2),
        (23, "Cost-of-carry fair forward", "=E5*EXP(E22*E10)", CUR2),
        (24, "Observed less fair forward", "=E6-E23", CUR2),
        (25, "Spot-futures basis", "=E5-E6", CUR2),
        (26, "Front-to-next roll yield", "='Roll Yield'!C6", PCT2),
        (27, "Calculated ending inventory", "=E11+E12+E13-E14-E15", NUM),
        (28, "Physical balance residual", "=E27-E16", NUM),
        (29, "Absolute hedge ratio", "=IFERROR(ABS('Hedging'!C12*'Hedging'!C6/'Hedging'!C5),0)", MULT),
        (30, "Unhedged basis exposure", "='Hedging'!C14", CUR),
    ]
    for row, label, formula, number_format in outputs:
        sheet.cell(row, 2, label)
        sheet.cell(row, 5, formula)
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 5).border = BORDER
    sheet["B32"] = "The cost-of-carry result is a reference relationship, not a claim that convenience yield or storage is directly observable."
    sheet["B32"].font = ITALIC_GRAY
    sheet.freeze_panes = "B5"
    sheet.sheet_view.showGridLines = False

    checks = workbook.create_sheet("Decision & Checks", insert_at + 1)
    set_col_widths(checks, [4, 36, 18, 18, 46])
    checks["B2"] = "Commodities Decision Dashboard and Independent Checks"
    checks["B2"].font = TITLE
    for column, value in enumerate(["Check / decision", "Metric", "Status", "Interpretation / action"], start=2):
        checks.cell(4, column, value)
    style_header_row(checks, 4, 4, start_col=2)
    rows = [
        (5, "Cost-of-carry residual", "='Physical Balance & Carry'!E24", '=IF(ABS(C5)<=MAX(0.01,ABS(\'Physical Balance & Carry\'!E23)*0.05),"PASS","REVIEW")', "Review storage, funding, convenience yield, delivery option, and contract alignment."),
        (6, "Physical balance residual", "='Physical Balance & Carry'!E28", '=IF(ABS(C6)<0.000001,"PASS","FAIL")', "Inventory must conserve across beginning stock, flows, shrink, and reported ending stock."),
        (7, "Absolute basis", "=ABS('Physical Balance & Carry'!E25)", '=IF(C7<=\'Physical Balance & Carry\'!E17,"PASS","BREACH")', "Escalate material location, grade, quality, or timing basis risk."),
        (8, "Hedge ratio", "='Physical Balance & Carry'!E29", '=IF(C8<=\'Physical Balance & Carry\'!E18,"PASS","BREACH")', "Review over-hedging, contract multiplier, and beta-adjustment assumptions."),
        (9, "Roll yield", "='Physical Balance & Carry'!E26", '=IF(C9<0,"REVIEW","PASS")', "Negative roll yield is expected in contango but must be included in return and hedge economics."),
        (10, "Ending inventory", "='Physical Balance & Carry'!E27", '=IF(C10>=0,"PASS","BREACH")', "Negative physical inventory indicates an impossible operating plan or missing purchase."),
    ]
    for row, label, metric, status, action in rows:
        checks.cell(row, 2, label)
        checks.cell(row, 3, metric)
        checks.cell(row, 3).border = BORDER
        checks.cell(row, 4, status)
        checks.cell(row, 4).border = BORDER
        checks.cell(row, 5, action)
    checks["B12"] = "Overall model status"
    checks["B12"].font = BOLD
    checks["C12"] = _status_formula("D5:D10")
    checks["C12"].font = BOLD
    checks["C12"].border = BORDER
    checks["B14"] = "Primary decision outputs"
    checks["B14"].font = BOLD
    checks["B14"].fill = GRAY_FILL
    decisions = [
        (15, "Fair forward", "='Physical Balance & Carry'!E23", CUR2),
        (16, "Observed futures", "='Physical Balance & Carry'!E6", CUR2),
        (17, "Basis", "='Physical Balance & Carry'!E25", CUR2),
        (18, "Contracts needed", "='Hedging'!C12", NUM),
        (19, "Unhedged exposure", "='Hedging'!C14", CUR),
    ]
    for row, label, formula, number_format in decisions:
        checks.cell(row, 2, label)
        checks.cell(row, 3, formula)
        checks.cell(row, 3).number_format = number_format
        checks.cell(row, 3).border = BORDER
    checks.freeze_panes = "B5"
    checks.sheet_view.showGridLines = False



def build(output: Path) -> None:
    build_release("build_commodities_template.py", output, enrich)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("COMMODITIES_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
