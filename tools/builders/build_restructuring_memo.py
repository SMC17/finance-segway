"""Build a governed, data-grounded restructuring memo deck for a real
public Distressed & Restructuring instance.

Seventh domain on the IC-memo/pptx system, reusing
tools/builders/pptx_helpers.py unchanged. Every number on a slide is read
from the instance workbook after a real LibreOffice recalculation
(tools.recalc), or from the case's own governed manifest
(standards/public_cases/<case_id>.json). Nothing on any slide is invented
here -- this script only reads and renders.

Default case (distressed-public-hertz-2021-reorganization) sources
exactly ONE real fact -- Hertz's actual $7.5bn Chapter 11 exit financing
commitment -- matching what the case itself sources, no more. This deck
is deliberately shorter and more sparse than the other domains' memos:
Recovery Waterfall, 13-Week Liquidity, and Liquidation-vs-Reorg are
entirely untouched template defaults for this case (no real capital
structure or cash-flow data was sourced), so a deck that dressed those
sheets' outputs up as "analysis" would misrepresent illustrative zeros as
decision-grade numbers. The deck says so explicitly rather than hiding it
behind computed-looking slides.

Usage:
    python tools/builders/build_restructuring_memo.py --case-id distressed-public-hertz-2021-reorganization
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "tools"))

import pptx_helpers as ph  # noqa: E402
from recalc import recalc  # noqa: E402

import openpyxl  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
INDEX = ROOT / "standards" / "public_cases" / "index.json"


def _load_case(case_id: str) -> dict[str, Any]:
    index = json.loads(INDEX.read_text(encoding="utf-8"))
    for item in index["cases"]:
        if item["case_id"] == case_id:
            return item
    raise KeyError(f"unknown case_id {case_id!r}")


def _recalculated_workbook(output_rel: str) -> openpyxl.Workbook:
    src = ROOT / output_rel
    with tempfile.TemporaryDirectory(dir=ROOT) as tmp:
        scratch = Path(tmp) / src.name
        shutil.copyfile(src, scratch)
        result = recalc(str(scratch), timeout=90)
        if result.get("status") != "success" or result.get("total_errors", 1):
            raise RuntimeError(f"recalculation did not succeed cleanly: {result}")
        return openpyxl.load_workbook(scratch, data_only=True)


def _manifest(case_id: str) -> dict[str, Any]:
    return json.loads((ROOT / "standards" / "public_cases" / f"{case_id}.json").read_text(encoding="utf-8"))


def _fmt_usd_mm(value: float) -> str:
    return f"${value:,.1f}mm"


def build_restructuring_memo(case_id: str, output: Path) -> None:
    case = _load_case(case_id)
    manifest = _manifest(case_id)
    outcome = manifest["outcome"]
    real_inputs = [item for item in manifest["inputs"] if item.get("input_kind") in {"observed", "derived"}]
    situation = manifest["cover"]["Situation type:"]

    wb = _recalculated_workbook(case["output"])
    new_money = wb["New Money"]
    checks = wb["Decision & Checks"]

    as_of = manifest["as_of"]
    commitment = new_money["C5"].value
    total_claim_at_exit = new_money["C18"].value
    commitment_status = new_money["C19"].value

    check_status = {checks.cell(r, 2).value: checks.cell(r, 4).value for r in range(5, 13) if checks.cell(r, 2).value}
    overall_status = checks["C14"].value

    prs = ph.new_presentation()

    # 1. Title
    ph.add_title_slide(
        prs,
        f"{situation} — Restructuring Memo",
        "This case sources exactly one real fact: the disclosed exit-financing commitment. "
        "Capital structure, liquidity, and recovery mechanics are untouched template defaults for this case.",
        f"As of {as_of}  ·  Prepared from a governed, recalculated model instance",
    )

    # 2. Provenance -- deliberately blunt about sparsity
    slide = ph.add_content_slide(prs, "Reading this deck", kicker="Provenance — sparse by design, not by accident")
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(1.5), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.8),
        [
            f"Real, sourced: exactly one fact — {_fmt_usd_mm(commitment)}, {manifest['sources'][0]['name']} — "
            "the actual disclosed Chapter 11 exit-financing commitment.",
            "NOT populated for this case: Recovery Waterfall (capital-structure face claims), 13-Week "
            "Liquidity (cash-flow schedule), and Liquidation-vs-Reorg (asset value / NPV bridge) remain "
            "entirely at template defaults — pre-petition capital-structure and liquidity detail were not "
            "sourced with real provenance for this case, and this deck does not manufacture that detail.",
            "The fee, interest-rate, and tenor terms below are illustrative mechanics applied to the real "
            "commitment ceiling — not the real facility's actual disclosed pricing.",
            "This is the same sparsity discipline this case's own manifest already enforces: better an "
            "honest single real fact than a denser deck built on invented capital-structure detail.",
        ],
        size=15,
    )
    ph.add_footer(slide, f"Model checks: Overall {overall_status}  ·  standards/public_cases/{case_id}.json")

    # 3. Real fact + realized outcome
    slide = ph.add_content_slide(prs, "Exit financing and outcome", kicker="Real, disclosed")
    ph.add_stat_row(
        slide, ph.Inches(1.6),
        [
            ph.Stat(_fmt_usd_mm(commitment), "Exit-financing commitment", tag="Form 10-K", color=ph.NAVY),
            ph.Stat("Completed" if outcome["realized"] == 1.0 else "Not completed",
                    "Chapter 11 emergence", tag="Form 10-K", color=ph.POSITIVE),
        ],
    )
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(3.4), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.6),
        f"{situation.split('Chapter')[0].strip()} emerged from Chapter 11 with committed exit financing of "
        f"{_fmt_usd_mm(commitment)} — a real, disclosed, completed outcome, not a projection from this "
        "workbook's own mechanics.",
        size=15, color=ph.CHARCOAL,
    )
    ph.add_footer(slide, f"Source: {outcome['realized_source']}")

    # 4. Illustrative financing mechanics
    slide = ph.add_content_slide(prs, "Modeled financing mechanics", kicker="Illustrative terms on a real ceiling")
    ph.add_table(
        slide, ph.MARGIN, ph.Inches(1.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.6),
        ["Metric", "Value"],
        [
            ["Commitment ceiling (real)", _fmt_usd_mm(commitment)],
            ["Total new-money claim at exit (illustrative terms)", _fmt_usd_mm(total_claim_at_exit)],
            ["Commitment sufficiency check", commitment_status],
        ],
        col_widths=[4, 2],
    )
    ph.add_disclosure_note(
        slide,
        "Upfront fee, cash interest rate, tenor, and exit fee are template defaults, not this facility's "
        "actual disclosed pricing. The expected draw feeding this calculation comes from the 13-Week "
        "Liquidity sheet's own template-default cash-flow schedule, not a real liquidity forecast.",
    )

    # 5. Decision — honest about what's missing
    slide = ph.add_content_slide(prs, "What a real recovery decision would still need", kicker="Decision")
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(1.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.2),
        f"The only decision-grade fact this case supports today: {_fmt_usd_mm(commitment)} of exit "
        "financing was real, committed, and the case completed Chapter 11 emergence.",
        size=17, bold=True, color=ph.NAVY,
    )
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(2.9), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.8),
        [
            "A real capital-structure recovery view needs each tranche's actual face claim and seniority "
            "— not sourced here; Recovery Waterfall stays at template defaults for this case.",
            "A real liquidity-runway view needs the company's actual 13-week cash-flow forecast — not "
            "sourced here.",
            "A real liquidation-vs-reorganization NPV comparison needs actual asset values and "
            "distress-cost estimates — not sourced here.",
            "None of these gaps are hidden inside a computed-looking number: the workbook's own "
            "Decision & Checks status reflects that most checks are running on template defaults, "
            "not real inputs, for this specific case.",
        ],
        size=15,
    )
    ph.add_footer(
        slide,
        f"Model: 24_Distressed_Restructuring/_template_RESTRUCTURING.xlsx  ·  Instance: {case['output']}  ·  "
        f"Overall checks: {overall_status}",
    )

    ph.save(prs, output)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-id", default="distressed-public-hertz-2021-reorganization")
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    output = args.output or ROOT / "24_Distressed_Restructuring" / "presentations" / f"{args.case_id}-memo.pptx"
    build_restructuring_memo(args.case_id, output)
    print(f"saved {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
