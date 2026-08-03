"""
Builds _template.xlsx — the master model template.
Tabs: Cover, Assumptions, IS, BS, CF, DCF, Comps, Sensitivity, RefreshLog
Convention: blue=hardcode input, black=formula, green=cross-sheet link, yellow fill=key assumption
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
BOLD_WHITE = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=16, bold=True)
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
MULT = '0.0x'
YR = '@'

wb = openpyxl.Workbook()
wb.remove(wb.active)

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------------- COVER ----------------
ws = wb.create_sheet("Cover")
set_col_widths(ws, [4, 30, 40, 4])
ws["B2"] = "[TICKER] — Company Name"
ws["B2"].font = TITLE
ws["B4"] = "Sector:"; ws["B4"].font = BOLD; ws["C4"] = "[fill in]"; ws["C4"].font = BLUE
ws["B5"] = "Coverage started:"; ws["B5"].font = BOLD; ws["C5"] = "[date]"; ws["C5"].font = BLUE
ws["B6"] = "Last refreshed:"; ws["B6"].font = BOLD; ws["C6"] = "[date]"; ws["C6"].font = BLUE
ws["B7"] = "Next earnings date:"; ws["B7"].font = BOLD; ws["C7"] = "[date]"; ws["C7"].font = BLUE
ws["B8"] = "Refresh cadence:"; ws["B8"].font = BOLD; ws["C8"] = "Weekly"; ws["C8"].font = BLACK
ws["B10"] = "Thesis (2-3 sentences):"; ws["B10"].font = BOLD
ws["B11"] = "[fill in]"; ws["B11"].font = BLUE
ws.merge_cells("B11:D14")
ws["B16"] = "COLOR LEGEND"; ws["B16"].font = BOLD
legend = [
    ("Blue text", "Hardcoded input / scenario lever — edit freely", BLUE),
    ("Black text", "Formula — do not overwrite", BLACK),
    ("Green text", "Link to another sheet in this workbook", GREEN),
    ("Yellow fill", "Key assumption — review every refresh", None),
]
r = 17
for label, desc, font in legend:
    ws.cell(row=r, column=2, value=label).font = font if font else BOLD
    if label == "Yellow fill":
        ws.cell(row=r, column=2).fill = YELLOW_FILL
    ws.cell(row=r, column=3, value=desc).font = BLACK
    r += 1
ws.sheet_view.showGridLines = False

# ---------------- ASSUMPTIONS ----------------
ws = wb.create_sheet("Assumptions")
set_col_widths(ws, [4, 30, 14, 14, 14, 14, 14, 40])
ws["B2"] = "Assumptions"; ws["B2"].font = TITLE
headers = ["", "Driver", "FY1", "FY2", "FY3", "FY4", "FY5", "Source / notes"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

rows = [
    ("Revenue growth %", PCT, "yellow"),
    ("Gross margin %", PCT, "yellow"),
    ("Opex % of revenue", PCT, "yellow"),
    ("Tax rate %", PCT, "yellow"),
    ("D&A % of capex", PCT, None),
    ("Capex % of revenue", PCT, None),
    ("NWC % of revenue chg", PCT, None),
    ("Shares outstanding (mm)", CUR, None),
    ("WACC %", PCT, "yellow"),
    ("Terminal growth %", PCT, "yellow"),
]
r = 5
for label, fmt, flag in rows:
    ws.cell(row=r, column=2, value=label).font = BLACK
    for c in range(3, 8):
        cell = ws.cell(row=r, column=c, value=0.0)
        cell.font = BLUE
        cell.number_format = fmt
        if flag == "yellow":
            cell.fill = YELLOW_FILL
        cell.border = BORDER
    ws.cell(row=r, column=8, value="Source: [10-K/10-Q cite or user estimate]").font = Font(name="Arial", size=9, italic=True, color="808080")
    r += 1
ws.sheet_view.showGridLines = False

# ---------------- INCOME STATEMENT ----------------
ws = wb.create_sheet("IS")
set_col_widths(ws, [4, 26, 13, 13, 13, 13, 13, 13, 13])
ws["B2"] = "Income Statement ($mm)"; ws["B2"].font = TITLE
per_headers = ["", "Line item", "FY-2A", "FY-1A", "FY0A", "FY1E", "FY2E", "FY3E", "FY4E"]
for i, h in enumerate(per_headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(per_headers))

is_rows = ["Revenue", "  Revenue growth %", "COGS", "Gross profit", "  Gross margin %",
           "Opex", "EBITDA", "  EBITDA margin %", "D&A", "EBIT", "Interest expense",
           "Pre-tax income", "Tax", "Net income", "Diluted shares", "Diluted EPS"]
r = 5
first_data_row = r
for label in is_rows:
    is_pct = "%" in label
    ws.cell(row=r, column=2, value=label.strip()).font = BOLD if label in ("Revenue", "Gross profit", "EBITDA", "EBIT", "Net income") else BLACK
    for c in range(3, 10):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        cell.number_format = PCT if is_pct else CUR
        cell.font = BLACK
    r += 1
last_data_row = r - 1

# Sample formula wiring (illustrative — real model should link every period consistently)
rev_row, growth_row, cogs_row, gp_row, gm_row = first_data_row, first_data_row+1, first_data_row+2, first_data_row+3, first_data_row+4
for c in range(4, 10):  # FY-1A onward references prior col growth
    col_letter = get_column_letter(c)
    prev_letter = get_column_letter(c-1)
    ws.cell(row=growth_row, column=c, value=f"=IFERROR({col_letter}{rev_row}/{prev_letter}{rev_row}-1,\"-\")")
    ws.cell(row=gp_row, column=c, value=f"={col_letter}{rev_row}-{col_letter}{cogs_row}")
    ws.cell(row=gm_row, column=c, value=f"=IFERROR({col_letter}{gp_row}/{col_letter}{rev_row},\"-\")")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "C5"

# ---------------- BALANCE SHEET ----------------
ws = wb.create_sheet("BS")
set_col_widths(ws, [4, 26, 13, 13, 13, 13, 13])
ws["B2"] = "Balance Sheet ($mm)"; ws["B2"].font = TITLE
for i, h in enumerate(["", "Line item", "FY-1A", "FY0A", "FY1E", "FY2E", "FY3E"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 6)
bs_rows = ["Cash & equivalents", "Accounts receivable", "Inventory", "Total current assets",
           "PP&E net", "Goodwill & intangibles", "Total assets",
           "Accounts payable", "Debt (current)", "Total current liabilities",
           "Long-term debt", "Total liabilities", "Total equity"]
r = 5
for label in bs_rows:
    ws.cell(row=r, column=2, value=label).font = BOLD if "Total" in label else BLACK
    for c in range(3, 8):
        cell = ws.cell(row=r, column=c); cell.number_format = CUR; cell.border = BORDER; cell.font = BLACK
    r += 1
ws.sheet_view.showGridLines = False

# ---------------- CASH FLOW ----------------
ws = wb.create_sheet("CF")
set_col_widths(ws, [4, 26, 13, 13, 13, 13, 13])
ws["B2"] = "Cash Flow ($mm)"; ws["B2"].font = TITLE
for i, h in enumerate(["", "Line item", "FY-1A", "FY0A", "FY1E", "FY2E", "FY3E"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 6)
cf_rows = ["Net income", "+ D&A", "+/- Change in NWC", "Cash flow from operations",
           "Capex", "Free cash flow", "Debt issuance/(repayment)", "Dividends/buybacks",
           "Net change in cash"]
r = 5
for label in cf_rows:
    ws.cell(row=r, column=2, value=label).font = BOLD if label in ("Cash flow from operations", "Free cash flow") else BLACK
    for c in range(3, 8):
        cell = ws.cell(row=r, column=c); cell.number_format = CUR; cell.border = BORDER; cell.font = BLACK
    r += 1
ws.sheet_view.showGridLines = False

# ---------------- DCF ----------------
ws = wb.create_sheet("DCF")
set_col_widths(ws, [4, 26, 13, 13, 13, 13, 13, 4, 22, 14])
ws["B2"] = "DCF Valuation"; ws["B2"].font = TITLE
for i, h in enumerate(["", "", "FY1E", "FY2E", "FY3E", "FY4E", "FY5E"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 5, start_col=3)
ws["B5"] = "Unlevered FCF"
for c in range(3, 8):
    ws.cell(row=5, column=c).number_format = CUR
    ws.cell(row=5, column=c).font = GREEN  # link to CF sheet
ws["B6"] = "Discount factor"
ws["B7"] = "PV of FCF"
for c in range(3, 8):
    col = get_column_letter(c)
    ws.cell(row=6, column=c, value=f"=1/(1+$I$5)^({c-2})").number_format = '0.000'
    ws.cell(row=7, column=c, value=f"={col}5*{col}6").number_format = CUR

ws["H4"] = "Key outputs"; ws["H4"].font = BOLD
ws["H5"] = "WACC"; ws["I5"] = 0.10; ws["I5"].font = BLUE; ws["I5"].fill = YELLOW_FILL; ws["I5"].number_format = PCT
ws["H6"] = "Terminal growth"; ws["I6"] = 0.025; ws["I6"].font = BLUE; ws["I6"].fill = YELLOW_FILL; ws["I6"].number_format = PCT
ws["H7"] = "Terminal value"; ws["I7"] = "=G5*(1+I6)/(I5-I6)"; ws["I7"].number_format = CUR
ws["H8"] = "PV of terminal value"; ws["I8"] = "=I7*G6"; ws["I8"].number_format = CUR
ws["H9"] = "Sum PV of FCF"; ws["I9"] = "=SUM(C7:G7)"; ws["I9"].number_format = CUR
ws["H10"] = "Enterprise value"; ws["I10"] = "=I8+I9"; ws["I10"].font = BOLD; ws["I10"].number_format = CUR
ws["H11"] = "Less: net debt"; ws["I11"] = 0; ws["I11"].font = BLUE; ws["I11"].number_format = CUR
ws["H12"] = "Equity value"; ws["I12"] = "=I10-I11"; ws["I12"].font = BOLD; ws["I12"].number_format = CUR
ws["H13"] = "Diluted shares (mm)"; ws["I13"] = 1; ws["I13"].font = BLUE; ws["I13"].number_format = CUR
ws["H14"] = "Implied value/share"; ws["I14"] = "=I12/I13"; ws["I14"].font = BOLD; ws["I14"].number_format = CUR
ws.sheet_view.showGridLines = False

# ---------------- COMPS ----------------
ws = wb.create_sheet("Comps")
set_col_widths(ws, [4, 22, 12, 12, 12, 12, 12, 12])
ws["B2"] = "Comparable Companies"; ws["B2"].font = TITLE
comp_headers = ["", "Ticker", "Price", "Mkt Cap", "EV", "EV/Rev", "EV/EBITDA", "P/E"]
for i, h in enumerate(comp_headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(comp_headers))
for r in range(5, 12):
    ws.cell(row=r, column=2, value="[TICKER]").font = BLUE
    for c in range(3, 9):
        cell = ws.cell(row=r, column=c, value=0)
        cell.font = BLUE
        cell.number_format = MULT if c >= 6 else CUR
        cell.border = BORDER
ws.cell(row=13, column=2, value="Median").font = BOLD
for c in range(3, 9):
    col = get_column_letter(c)
    ws.cell(row=13, column=c, value=f"=MEDIAN({col}5:{col}12)").font = BOLD
    ws.cell(row=13, column=c).number_format = MULT if c >= 6 else CUR
ws.sheet_view.showGridLines = False

# ---------------- SENSITIVITY ----------------
ws = wb.create_sheet("Sensitivity")
set_col_widths(ws, [4, 20] + [12]*6)
ws["B2"] = "Sensitivity — Implied Value/Share"; ws["B2"].font = TITLE
ws["B4"] = "WACC \\ Term. growth"; ws["B4"].font = BOLD
ws["B4"].fill = GRAY_FILL
term_g = [0.015, 0.02, 0.025, 0.03, 0.035]
waccs = [0.08, 0.09, 0.10, 0.11, 0.12]
for i, g in enumerate(term_g, start=3):
    c = ws.cell(row=4, column=i, value=g); c.number_format = PCT; c.font = BOLD; c.fill = GRAY_FILL
for j, w in enumerate(waccs, start=5):
    c = ws.cell(row=j, column=2, value=w); c.number_format = PCT; c.font = BOLD; c.fill = GRAY_FILL
    for i in range(3, 8):
        cell = ws.cell(row=j, column=i, value="[data table — Excel What-If Analysis]")
        cell.font = Font(name="Arial", size=8, italic=True, color="808080")
ws.sheet_view.showGridLines = False

# ---------------- REFRESH LOG ----------------
ws = wb.create_sheet("RefreshLog")
set_col_widths(ws, [4, 14, 16, 30, 30, 16])
ws["B2"] = "Refresh Log"; ws["B2"].font = TITLE
log_headers = ["", "Date", "Trigger", "What changed", "Reviewer notes", "Next check"]
for i, h in enumerate(log_headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(log_headers))
for r in range(5, 15):
    for c in range(2, 7):
        ws.cell(row=r, column=c).border = BORDER
ws.sheet_view.showGridLines = False

wb.move_sheet("Cover", offset=-len(wb.sheetnames))
out_path = "/home/claude/model_shop/_template.xlsx"
wb.save(out_path)
print("saved", out_path)
