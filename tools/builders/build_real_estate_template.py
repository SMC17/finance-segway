import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[PROPERTY/REIT] — Real Estate Model", [
    ("Property type:", "Office / Multifamily / Industrial / Retail"),
    ("Location:", "[fill in]"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly"),
])

# ---------------- PROPERTY PRO FORMA ----------------
ws = wb.create_sheet("Property Pro Forma")
set_col_widths(ws, [4, 30, 14, 14, 14])
ws["B2"] = "Property Pro Forma (annual, $)"; ws["B2"].font = TITLE
rows = ["Gross potential rent", "- Vacancy & credit loss", "Effective gross income",
        "+ Other income (parking, fees)", "Total revenue",
        "- Operating expenses (taxes, insurance, mgmt, R&M)", "Net Operating Income (NOI)",
        "- Capex reserve", "- Debt service (P&I)", "Cash flow before tax"]
r = 5
for label in rows:
    is_bold = label in ("Net Operating Income (NOI)", "Cash flow before tax", "Total revenue")
    ws.cell(row=r, column=2, value=label).font = BOLD if is_bold else BLACK
    c = ws.cell(row=r, column=3, value=0)
    c.font = BLUE; c.number_format = CUR; c.border = BORDER
    r += 1
# wire formulas: EGI = gross - vacancy; total rev = EGI + other; NOI = total rev - opex; CF = NOI - capex - debt service
ws["C7"] = "=C5+C6"; ws["C7"].font = BLACK  # effective gross income
ws["C9"] = "=C7+C8"; ws["C9"].font = BOLD  # total revenue
ws["C11"] = "=C9+C10"; ws["C11"].font = BOLD  # NOI (opex entered negative)
ws["C14"] = "=C11+C12+C13"; ws["C14"].font = BOLD  # CF before tax
ws.sheet_view.showGridLines = False

# ---------------- CAP RATE & VALUATION ----------------
ws = wb.create_sheet("Cap Rate & Valuation")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Cap Rate Valuation"; ws["B2"].font = TITLE
ws["B4"] = "NOI (from pro forma)"; ws["C4"] = "='Property Pro Forma'!C11"; ws["C4"].font = GREEN; ws["C4"].number_format = CUR
ws["B5"] = "Market cap rate %"; ws["C5"] = 0.06; ws["C5"].font = BLUE; ws["C5"].fill = YELLOW_FILL; ws["C5"].number_format = PCT
ws["B6"] = "Implied value (NOI / cap rate)"; ws["C6"] = "=IFERROR(C4/C5,\"-\")"; ws["C6"].font = BOLD; ws["C6"].number_format = CUR
ws["B8"] = "Purchase price ($)"; ws["C8"] = 0; ws["C8"].font = BLUE; ws["C8"].number_format = CUR
ws["B9"] = "Going-in cap rate (NOI / price)"; ws["C9"] = "=IFERROR(C4/C8,\"-\")"; ws["C9"].number_format = PCT
ws["B10"] = "Debt amount"; ws["C10"] = 0; ws["C10"].font = BLUE; ws["C10"].number_format = CUR
ws["B11"] = "Cash-on-cash return"; ws["C11"] = "=IFERROR('Property Pro Forma'!C14/(C8-C10),\"-\")"; ws["C11"].number_format = PCT
for r2 in [4,5,6,8,9,10,11]:
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- REIT FFO/AFFO ----------------
ws = wb.create_sheet("REIT FFO-AFFO")
set_col_widths(ws, [4, 34, 16, 40])
ws["B2"] = "REIT FFO / AFFO"; ws["B2"].font = TITLE
ws["B4"] = "Net income (GAAP)"; ws["C4"] = 0; ws["C4"].font = BLUE; ws["C4"].number_format = CUR
ws["B5"] = "+ Real estate depreciation & amortization"; ws["C5"] = 0; ws["C5"].font = BLUE; ws["C5"].number_format = CUR
ws["B6"] = "- Gains on property sales"; ws["C6"] = 0; ws["C6"].font = BLUE; ws["C6"].number_format = CUR
ws["B7"] = "FFO (Funds From Operations)"; ws["C7"] = "=C4+C5+C6"; ws["C7"].font = BOLD; ws["C7"].number_format = CUR
ws["B9"] = "- Recurring capex / leasing costs"; ws["C9"] = 0; ws["C9"].font = BLUE; ws["C9"].number_format = CUR
ws["B10"] = "- Straight-line rent adjustment"; ws["C10"] = 0; ws["C10"].font = BLUE; ws["C10"].number_format = CUR
ws["B11"] = "AFFO (Adjusted FFO)"; ws["C11"] = "=C7+C9+C10"; ws["C11"].font = BOLD; ws["C11"].number_format = CUR
ws["D11"] = "AFFO is the better proxy for sustainable dividend-paying capacity"
ws["D11"].font = ITALIC_GRAY
ws["B13"] = "Shares/units outstanding"; ws["C13"] = 1; ws["C13"].font = BLUE; ws["C13"].number_format = NUM
ws["B14"] = "FFO / share"; ws["C14"] = "=IFERROR(C7/C13,\"-\")"; ws["C14"].number_format = CUR2
ws["B15"] = "AFFO / share"; ws["C15"] = "=IFERROR(C11/C13,\"-\")"; ws["C15"].number_format = CUR2
for r2 in [4,5,6,7,9,10,11,13,14,15]:
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- 5-YEAR HOLD & LEVERED IRR ----------------
ws = wb.create_sheet("5-Year Hold & IRR")
set_col_widths(ws, [4, 30, 12, 12, 12, 12, 12, 12])
ws["B2"] = "5-Year Hold — Levered Cash Flow & IRR"; ws["B2"].font = TITLE
ws["B4"] = "Assumptions"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
ws["B5"] = "NOI growth rate (%/yr)"
ws["C5"] = 0.02; ws["C5"].font = BLUE; ws["C5"].fill = YELLOW_FILL; ws["C5"].number_format = PCT
ws["B6"] = "Exit cap rate (%)"
ws["C6"] = 0.065; ws["C6"].font = BLUE; ws["C6"].fill = YELLOW_FILL; ws["C6"].number_format = PCT
ws["B7"] = "Selling costs (% of exit value)"
ws["C7"] = 0.02; ws["C7"].font = BLUE; ws["C7"].number_format = PCT
ws["B8"] = "Annual debt service ($, assumed constant / interest-only)"
ws["C8"] = "=IFERROR(ABS('Property Pro Forma'!C13),\"-\")"; ws["C8"].font = GREEN; ws["C8"].number_format = CUR
ws["B9"] = "Initial equity (purchase price - debt)"
ws["C9"] = "=IFERROR('Cap Rate & Valuation'!C8-'Cap Rate & Valuation'!C10,\"-\")"
ws["C9"].font = GREEN; ws["C9"].number_format = CUR
for r2 in (5, 6, 7, 8, 9):
    ws.cell(row=r2, column=3).border = BORDER

ws["B12"] = ""
for i, h in enumerate(["", "", "Yr1", "Yr2", "Yr3", "Yr4", "Yr5"], start=1):
    ws.cell(row=12, column=i, value=h)
style_header_row(ws, 12, 5, start_col=3)
ws["B13"] = "NOI"
ws["C13"] = "='Property Pro Forma'!C11"; ws["C13"].font = GREEN
for col in range(4, 8):
    prev = get_column_letter(col - 1)
    ws.cell(row=13, column=col, value=f"={prev}13*(1+$C$5)")
ws["B14"] = "Debt service"
for col in range(3, 8):
    ws.cell(row=14, column=col, value="=$C$8")
ws["B15"] = "Levered cash flow"
for col in range(3, 8):
    letter = get_column_letter(col)
    ws.cell(row=15, column=col, value=f"={letter}13-{letter}14")
ws["B16"] = "DSCR (NOI / debt service)"
for col in range(3, 8):
    letter = get_column_letter(col)
    ws.cell(row=16, column=col, value=f"=IFERROR({letter}13/{letter}14,\"-\")")
    ws.cell(row=16, column=col).number_format = MULT
    ws.cell(row=16, column=col).fill = YELLOW_FILL
for row in (13, 14, 15):
    for col in range(3, 8):
        ws.cell(row=row, column=col).number_format = CUR
for row in (13, 14, 15, 16):
    for col in range(3, 8):
        ws.cell(row=row, column=col).border = BORDER

ws["B18"] = "Exit (end of Yr5)"; ws["B18"].font = BOLD; ws["B18"].fill = GRAY_FILL
ws["B19"] = "Forward NOI (Yr6, for exit cap)"
ws["C19"] = "=G13*(1+$C$5)"; ws["C19"].number_format = CUR; ws["C19"].border = BORDER
ws["B20"] = "Exit value (forward NOI / exit cap rate)"
ws["C20"] = "=IFERROR(C19/C6,\"-\")"; ws["C20"].number_format = CUR; ws["C20"].border = BORDER
ws["B21"] = "Less: selling costs"
ws["C21"] = "=-C20*C7"; ws["C21"].number_format = CUR; ws["C21"].border = BORDER
ws["B22"] = "Less: debt payoff (assumed interest-only, balance unchanged)"
ws["C22"] = "=-'Cap Rate & Valuation'!C10"; ws["C22"].font = GREEN; ws["C22"].number_format = CUR
ws["C22"].border = BORDER
ws["B23"] = "Net exit equity proceeds"
ws["C23"] = "=C20+C21+C22"; ws["C23"].font = BOLD; ws["C23"].number_format = CUR; ws["C23"].border = BORDER

ws["B25"] = "Levered Returns"; ws["B25"].font = BOLD; ws["B25"].fill = GRAY_FILL
for i, h in enumerate(["", "", "Yr0", "Yr1", "Yr2", "Yr3", "Yr4", "Yr5"], start=1):
    ws.cell(row=25, column=i, value=h)
style_header_row(ws, 25, 6, start_col=3)
ws["B26"] = "Equity cash flow"
ws["C26"] = "=-C9"; ws["C26"].number_format = CUR; ws["C26"].border = BORDER
# Row 26 runs Yr0..Yr5 (columns C-H); the NOI/CF block above runs Yr1..Yr5
# (columns C-G, no Yr0) — row26's column N maps to that block's column N-1.
for col in range(4, 8):
    source = get_column_letter(col - 1)
    ws.cell(row=26, column=col, value=f"={source}15")
    ws.cell(row=26, column=col).number_format = CUR
    ws.cell(row=26, column=col).border = BORDER
ws.cell(row=26, column=8, value="=G15+C23")
ws.cell(row=26, column=8).number_format = CUR
ws.cell(row=26, column=8).border = BORDER
ws["B27"] = "Levered IRR"
ws["C27"] = "=IFERROR(IRR(C26:H26),\"-\")"; ws["C27"].font = BOLD; ws["C27"].number_format = PCT
ws["C27"].fill = YELLOW_FILL; ws["C27"].border = BORDER
ws["B28"] = "Equity multiple (MOIC)"
ws["C28"] = "=IFERROR(SUM(D26:H26)/C9,\"-\")"; ws["C28"].font = BOLD; ws["C28"].number_format = MULT
ws["C28"].border = BORDER
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "REAL_ESTATE_template.xlsx"
wb.save(out_path)
print("saved", out_path)
