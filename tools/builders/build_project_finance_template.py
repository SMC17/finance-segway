import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[PROJECT] — Project Finance Model", [
    ("Sector:", "Infrastructure / Energy / PPP"),
    ("Financial close date:", "[date]"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly during construction; monthly during ops"),
])

# ---------------- CONSTRUCTION BUDGET ----------------
ws = wb.create_sheet("Construction Budget")
set_col_widths(ws, [4, 30, 16, 16, 40])
ws["B2"] = "Construction Budget & Drawdown"; ws["B2"].font = TITLE
inputs = [
    ("Total project cost ($)", 0, CUR),
    ("Debt / equity ratio (e.g. 0.70 = 70% debt)", 0.70, PCT),
    ("Construction period (months)", 24, NUM),
    ("Interest during construction rate (%)", 0.06, PCT),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B11"] = "Debt drawn"; ws["C11"] = "=C5*C6"; ws["C11"].number_format = CUR
ws["B12"] = "Equity drawn"; ws["C12"] = "=C5*(1-C6)"; ws["C12"].number_format = CUR
ws["B13"] = "Interest during construction (IDC, simple approx.)"
ws["C13"] = "=C11*C8*(C7/12)/2"; ws["C13"].number_format = CUR
ws["D13"] = "Approximation: avg drawn balance x rate x period, /2 for linear drawdown"
ws["D13"].font = ITALIC_GRAY
ws["B14"] = "Total debt at COD (incl. capitalized IDC)"
ws["C14"] = "=C11+C13"; ws["C14"].font = BOLD; ws["C14"].number_format = CUR
for r2 in (11, 12, 13, 14):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- OPERATING CASH FLOW ----------------
ws = wb.create_sheet("Operating Cash Flow")
set_col_widths(ws, [4, 26, 13, 13, 13, 13, 13])
ws["B2"] = "Operating Period Cash Flow ($mm)"; ws["B2"].font = TITLE
for i, h in enumerate(["", "Line item", "Yr1", "Yr2", "Yr3", "Yr4", "Yr5"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 5, start_col=3)
rows = ["Contracted/merchant revenue", "- Operating costs", "EBITDA",
        "- Maintenance capex reserve", "Cash flow available for debt service (CFADS)"]
r = 5
for label in rows:
    ws.cell(row=r, column=2, value=label).font = BOLD if label in ("EBITDA", "Cash flow available for debt service (CFADS)") else BLACK
    for c in range(3, 8):
        cell = ws.cell(row=r, column=c, value=0)
        cell.font = BLUE
        cell.number_format = CUR
        cell.border = BORDER
    r += 1
for c in range(3, 8):
    col = get_column_letter(c)
    ws.cell(row=7, column=c, value=f"={col}5+{col}6")  # EBITDA = rev - opcosts(neg)
    ws.cell(row=9, column=c, value=f"={col}7+{col}8")  # CFADS = EBITDA - capex(neg)
ws.sheet_view.showGridLines = False

# ---------------- DSCR & DEBT SIZING ----------------
ws = wb.create_sheet("DSCR & Debt Sizing")
set_col_widths(ws, [4, 26, 13, 13, 13, 13, 13])
ws["B2"] = "DSCR & Debt Sizing"; ws["B2"].font = TITLE
for i, h in enumerate(["", "", "Yr1", "Yr2", "Yr3", "Yr4", "Yr5"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 5, start_col=3)
ws["B5"] = "CFADS"
ws["B6"] = "Debt service (P&I)"
ws["B7"] = "DSCR (CFADS / Debt service)"
for c in range(3, 8):
    col = get_column_letter(c)
    ws.cell(row=5, column=c, value=f"='Operating Cash Flow'!{col}9")
    ws.cell(row=5, column=c).font = GREEN
    ws.cell(row=6, column=c, value=0).font = BLUE
    ws.cell(row=7, column=c, value=f"=IFERROR({col}5/{col}6,\"-\")")
    for r2 in (5, 6):
        ws.cell(row=r2, column=c).number_format = CUR
        ws.cell(row=r2, column=c).border = BORDER
    ws.cell(row=7, column=c).number_format = '0.00x'
    ws.cell(row=7, column=c).fill = YELLOW_FILL
    ws.cell(row=7, column=c).border = BORDER

ws["B9"] = "Minimum DSCR covenant"; ws["C9"] = 1.20; ws["C9"].font = BLUE; ws["C9"].fill = YELLOW_FILL
ws["C9"].number_format = '0.00x'; ws["C9"].border = BORDER
ws["B10"] = "Debt interest rate (for sizing PV)"
ws["C10"] = 0.06; ws["C10"].font = BLUE; ws["C10"].fill = YELLOW_FILL
ws["C10"].number_format = PCT2; ws["C10"].border = BORDER

ws["B12"] = "Debt sizing (sculpted to the DSCR covenant)"; ws["B12"].font = BOLD; ws["B12"].fill = GRAY_FILL
ws["B13"] = "Max annual debt service (CFADS / target DSCR)"
ws["B14"] = "PV of max debt service (discounted at debt rate)"
for c in range(3, 8):
    col = get_column_letter(c)
    year = c - 2
    ws.cell(row=13, column=c, value=f"=IFERROR({col}5/$C$9,\"-\")")
    ws.cell(row=13, column=c).number_format = CUR
    ws.cell(row=13, column=c).border = BORDER
    ws.cell(row=14, column=c, value=f"=IFERROR({col}13/(1+$C$10)^{year},\"-\")")
    ws.cell(row=14, column=c).number_format = CUR
    ws.cell(row=14, column=c).border = BORDER
ws["B16"] = "Max sculpted debt (sum of PVs of annual capacity)"
ws["C16"] = "=SUM(C14:G14)"
ws["C16"].font = BOLD; ws["C16"].number_format = CUR; ws["C16"].border = BORDER
ws["D16"] = ("This is capacity, not a repayment schedule — a real deal still needs an amortization "
             "profile (sculpted or level) that fits within it and keeps DSCR above covenant every "
             "period, not just on average")
ws["D16"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "PROJECT_FINANCE_template.xlsx"
wb.save(out_path)
print("saved", out_path)
