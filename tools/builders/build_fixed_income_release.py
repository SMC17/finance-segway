"""Canonical release wrapper for the institutional fixed-income builder."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_fixed_income_institutional import build as build_layout  # noqa: E402
from institutional_helpers import CUR, finalize  # noqa: E402


def build(output: Path) -> None:
    build_layout(output)
    workbook = load_workbook(output)
    portfolio = workbook["Portfolio"]
    scenarios = workbook["Scenarios"]
    pnl = workbook["P&L Explain"]

    # The raw builder writes spread duration into column G and duplicates it in
    # an unlabeled column L. Make the data model explicit and remove the duplicate.
    headers = [
        "Security", "Market value", "Coupon", "Yield", "Maturity",
        "Spread duration", "Price", "Mod duration", "Convexity", "DV01",
    ]
    for offset, value in enumerate(headers, start=2):
        portfolio.cell(4, offset, value)
    portfolio.merge_cells("B2:K2")
    portfolio["B2"] = "Fixed-Income Portfolio Analytics"
    for row in range(5, 11):
        portfolio.cell(row, 12).value = None
    portfolio["G12"] = "=SUMPRODUCT(C5:C10,G5:G10)/SUM(C5:C10)"
    portfolio["G12"].number_format = "0.0000"
    portfolio["K12"] = "=SUM(K5:K10)"
    portfolio["K12"].number_format = CUR
    portfolio["L12"].value = None

    # Credit-spread P&L uses the explicit spread-duration column G.
    for row in range(5, 11):
        scenarios.cell(row, 9, f"=-SUMPRODUCT('Key Rate Risk'!C12:G12,C{row}:G{row})/0.0001-SUMPRODUCT(Portfolio!C5:C10,Portfolio!G5:G10)*H{row}")
        scenarios.cell(row, 9).number_format = CUR
    pnl["C9"] = "=-SUMPRODUCT(Portfolio!C5:C10,Portfolio!G5:G10)"
    pnl["C9"].number_format = CUR

    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("FIXED_INCOME_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
