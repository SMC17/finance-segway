"""Release-grade enrichers for the six legacy M2 finance models."""
from __future__ import annotations

from pathlib import Path
from typing import Callable

from openpyxl.utils import get_column_letter

try:
    from tools.builders.legacy_release_adapter import build_release
    from tools.builders.template_helpers import (
        BLUE,
        BOLD,
        BORDER,
        CUR,
        GRAY_FILL,
        ITALIC_GRAY,
        MULT,
        NUM,
        PCT,
        PCT2,
        TITLE,
        YELLOW_FILL,
        set_col_widths,
        style_header_row,
    )
except ModuleNotFoundError:
    from legacy_release_adapter import build_release
    from template_helpers import (
        BLUE,
        BOLD,
        BORDER,
        CUR,
        GRAY_FILL,
        ITALIC_GRAY,
        MULT,
        NUM,
        PCT,
        PCT2,
        TITLE,
        YELLOW_FILL,
        set_col_widths,
        style_header_row,
    )


def _delete(workbook, *names: str) -> None:
    for name in names:
        if name in workbook.sheetnames:
            del workbook[name]


def _sheet(workbook, name: str, title_text: str, headers: list[str], widths: list[int]):
    ws = workbook.create_sheet(name)
    set_col_widths(ws, widths)
    ws["B2"] = title_text
    ws["B2"].font = TITLE
    for column, value in enumerate(headers, start=2):
        ws.cell(4, column, value)
    style_header_row(ws, 4, len(headers), start_col=2)
    ws.sheet_view.showGridLines = False
    return ws


def _input(cell, value, number_format: str) -> None:
    cell.value = value
    cell.font = BLUE
    cell.fill = YELLOW_FILL
    cell.border = BORDER
    cell.number_format = number_format


def _formula(cell, value: str, number_format: str = NUM) -> None:
    cell.value = value
    cell.border = BORDER
    cell.number_format = number_format


def _overall(status_range: str) -> str:
    return (
        f'=IF(COUNTIF({status_range},"FAIL")+COUNTIF({status_range},"BREACH")>0,'
        f'"BREACH",IF(COUNTIF({status_range},"REVIEW")>0,"REVIEW","PASS"))'
    )


def _decision_sheet(workbook, title_text: str, rows: list[tuple[str, str, str, str]]):
    ws = _sheet(
        workbook,
        "Decision & Checks",
        title_text,
        ["Check / decision", "Metric", "Status", "Interpretation / action"],
        [4, 42, 18, 18, 54],
    )
    for row, (label, metric, status, action) in enumerate(rows, start=5):
        ws.cell(row, 2, label)
        _formula(ws.cell(row, 3), metric)
        _formula(ws.cell(row, 4), status, "General")
        ws.cell(row, 5, action)
    overall_row = 5 + len(rows) + 1
    ws.cell(overall_row, 2, "Overall model status").font = BOLD
    _formula(ws.cell(overall_row, 3), _overall(f"D5:D{4 + len(rows)}"), "General")
    ws.cell(overall_row, 3).font = BOLD
    return ws


def enrich_investment_banking(workbook) -> None:
    _delete(
        workbook,
        "Transaction Analysis",
        "Accretion Dilution",
        "Valuation Reconciliation",
        "Decision & Checks",
    )
    tx = _sheet(
        workbook,
        "Transaction Analysis",
        "Transaction Valuation, Offer Premium, and DCF Reconciliation",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 44],
    )
    inputs = [
        ("Enterprise value", 1000.0, 1000.0, CUR, "transaction enterprise value"),
        ("Debt", 200.0, 200.0, CUR, "debt assumed or refinanced"),
        ("Cash", 100.0, 100.0, CUR, "cash retained in equity bridge"),
        ("Diluted shares", 100.0, 100.0, NUM, "fully diluted shares"),
        ("Offer price", 9.5, 15.0, CUR, "consideration per share"),
        ("Discount rate", 0.10, 0.05, PCT, "WACC / required return"),
        ("Terminal growth", 0.03, 0.06, PCT, "must remain below discount rate"),
        ("FCF Year 1", 70.0, 70.0, CUR, "unlevered free cash flow"),
        ("FCF Year 2", 75.0, 65.0, CUR, "unlevered free cash flow"),
        ("FCF Year 3", 80.0, 60.0, CUR, "unlevered free cash flow"),
        ("FCF Year 4", 85.0, 55.0, CUR, "unlevered free cash flow"),
        ("FCF Year 5", 90.0, 50.0, CUR, "unlevered free cash flow"),
        ("Offer-premium warning", 0.30, 0.30, PCT, "maximum premium screen"),
        ("Valuation-dispersion warning", 0.50, 0.50, PCT, "maximum DCF / bridge dispersion"),
    ]
    for row, (label, base, adverse, fmt, note) in enumerate(inputs, start=5):
        tx.cell(row, 2, label)
        _input(tx.cell(row, 3), base, fmt)
        _input(tx.cell(row, 4), adverse, fmt)
        tx.cell(row, 5, note)
    labels = [
        "Equity value",
        "Bridge value / share",
        "Offer premium",
        "PV of forecast FCF",
        "Terminal value",
        "PV of terminal value",
        "DCF enterprise value",
        "DCF equity value",
        "DCF value / share",
        "DCF / bridge dispersion",
    ]
    for row, label in enumerate(labels, start=21):
        tx.cell(row, 2, label)
    for column in (3, 4):
        L = get_column_letter(column)
        formulas = {
            21: f"={L}5-{L}6+{L}7",
            22: f"=IFERROR({L}21/{L}8,0)",
            23: f"=IFERROR({L}9/{L}22-1,0)",
            24: f"={L}12/(1+{L}10)+{L}13/(1+{L}10)^2+{L}14/(1+{L}10)^3+{L}15/(1+{L}10)^4+{L}16/(1+{L}10)^5",
            25: f'=IF({L}10<={L}11,"-",{L}16*(1+{L}11)/({L}10-{L}11))',
            26: f'=IF(ISNUMBER({L}25),{L}25/(1+{L}10)^5,"-")',
            27: f'=IF(AND(ISNUMBER({L}24),ISNUMBER({L}26)),{L}24+{L}26,"-")',
            28: f'=IF(ISNUMBER({L}27),{L}27-{L}6+{L}7,"-")',
            29: f'=IF(ISNUMBER({L}28),{L}28/{L}8,"-")',
            30: f'=IF(ISNUMBER({L}29),ABS({L}29-{L}22)/MAX(ABS({L}22),1),"-")',
        }
        for row, formula in formulas.items():
            fmt = PCT if row in (23, 30) else CUR
            _formula(tx.cell(row, column), formula, fmt)
    tx["B32"] = "Terminal value is invalid whenever terminal growth is greater than or equal to the discount rate."
    tx["B32"].font = ITALIC_GRAY

    acc = _sheet(
        workbook,
        "Accretion Dilution",
        "Pro Forma EPS, Synergy, Financing, and Share Issuance",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 44],
    )
    acc_inputs = [
        ("Buyer earnings", 100.0, 100.0, CUR),
        ("Buyer diluted shares", 100.0, 100.0, NUM),
        ("Target earnings", 10.0, 5.0, CUR),
        ("After-tax synergies", 5.0, 0.0, CUR),
        ("Incremental financing cost", 3.0, 20.0, CUR),
        ("Shares issued", 10.0, 50.0, NUM),
        ("Minimum accretion", 0.0, 0.0, PCT),
    ]
    for row, (label, base, adverse, fmt) in enumerate(acc_inputs, start=5):
        acc.cell(row, 2, label)
        _input(acc.cell(row, 3), base, fmt)
        _input(acc.cell(row, 4), adverse, fmt)
    for row, label in enumerate(
        ["Standalone EPS", "Pro forma earnings", "Pro forma shares", "Pro forma EPS", "EPS accretion / dilution"],
        start=14,
    ):
        acc.cell(row, 2, label)
    for column in (3, 4):
        L = get_column_letter(column)
        formulas = {
            14: f"=IFERROR({L}5/{L}6,0)",
            15: f"={L}5+{L}7+{L}8-{L}9",
            16: f"={L}6+{L}10",
            17: f"=IFERROR({L}15/{L}16,0)",
            18: f"=IFERROR({L}17/{L}14-1,0)",
        }
        for row, formula in formulas.items():
            _formula(acc.cell(row, column), formula, PCT if row == 18 else CUR)

    rec = _sheet(
        workbook,
        "Valuation Reconciliation",
        "Bridge, DCF, Offer, and Transaction Decision Reconciliation",
        ["Valuation method", "Base", "Adversarial", "Decision use"],
        [4, 36, 18, 18, 48],
    )
    rec_rows = [
        ("Enterprise-to-equity bridge / share", "='Transaction Analysis'!C22", "='Transaction Analysis'!D22", "capital-structure bridge"),
        ("DCF value / share", "='Transaction Analysis'!C29", "='Transaction Analysis'!D29", "intrinsic value and terminal-value sensitivity"),
        ("Offer price", "='Transaction Analysis'!C9", "='Transaction Analysis'!D9", "consideration and premium"),
        ("EPS accretion / dilution", "='Accretion Dilution'!C18", "='Accretion Dilution'!D18", "buyer shareholder impact"),
    ]
    for row, (label, base, adverse, note) in enumerate(rec_rows, start=5):
        rec.cell(row, 2, label)
        _formula(rec.cell(row, 3), base, PCT if row == 8 else CUR)
        _formula(rec.cell(row, 4), adverse, PCT if row == 8 else CUR)
        rec.cell(row, 5, note)

    _decision_sheet(
        workbook,
        "Investment Banking Decision Dashboard and Independent Checks",
        [
            ("Enterprise-to-equity bridge residual", "='Transaction Analysis'!C5-'Transaction Analysis'!C6+'Transaction Analysis'!C7-'Transaction Analysis'!C21", '=IF(ABS(C5)<0.01,"PASS","FAIL")', "Enterprise value must reconcile to equity value through debt and cash."),
            ("Per-share residual", "='Transaction Analysis'!C21-'Transaction Analysis'!C22*'Transaction Analysis'!C8", '=IF(ABS(C6)<0.01,"PASS","FAIL")', "Equity value must equal diluted shares times per-share value."),
            ("Terminal-value validity", "='Transaction Analysis'!D10-'Transaction Analysis'!D11", '=IF(C7>0,"PASS","BREACH")', "Discount rate must exceed terminal growth."),
            ("Adversarial offer premium", "='Transaction Analysis'!D23", '=IF(C8<=\'Transaction Analysis\'!D17,"PASS","BREACH")', "Escalate premiums above the approved transaction threshold."),
            ("Adversarial EPS accretion", "='Accretion Dilution'!D18", '=IF(C9>=\'Accretion Dilution\'!D11,"PASS","BREACH")', "Transaction economics must not dilute buyer EPS beyond the approved boundary."),
            ("Adversarial valuation dispersion", "='Transaction Analysis'!D30", '=IF(AND(ISNUMBER(C10),C10<=\'Transaction Analysis\'!D18),"PASS","BREACH")', "Bridge and DCF values require reconciliation before committee use."),
        ],
    )


def enrich_corporate_finance(workbook) -> None:
    _delete(workbook, "Treasury & Liquidity", "Capital Allocation", "Credit Metrics", "Decision & Checks")
    treasury = _sheet(
        workbook,
        "Treasury & Liquidity",
        "Cash, Debt, Minimum Liquidity, and Financing Capacity",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 46],
    )
    inputs = [
        ("Opening cash", 100.0, 50.0, CUR),
        ("Operating cash flow", 150.0, 30.0, CUR),
        ("Capital expenditure", 50.0, 60.0, CUR),
        ("Dividends", 20.0, 20.0, CUR),
        ("Buybacks", 10.0, 30.0, CUR),
        ("Debt issuance", 0.0, 80.0, CUR),
        ("Debt repayment", 30.0, 0.0, CUR),
        ("Opening debt", 200.0, 300.0, CUR),
        ("EBITDA", 120.0, 60.0, CUR),
        ("Cash interest", 10.0, 40.0, CUR),
        ("Minimum cash", 50.0, 80.0, CUR),
        ("Maximum net leverage", 3.5, 3.5, MULT),
        ("Minimum interest coverage", 2.0, 2.0, MULT),
    ]
    for row, (label, base, adverse, fmt) in enumerate(inputs, start=5):
        treasury.cell(row, 2, label)
        _input(treasury.cell(row, 3), base, fmt)
        _input(treasury.cell(row, 4), adverse, fmt)
    outputs = [
        "Ending cash",
        "Ending debt",
        "Net debt",
        "Net leverage",
        "Interest coverage",
        "Financing gap",
        "Post-capex free cash flow",
        "Shareholder distributions",
    ]
    for row, label in enumerate(outputs, start=21):
        treasury.cell(row, 2, label)
    for column in (3, 4):
        L = get_column_letter(column)
        formulas = {
            21: f"={L}5+{L}6-{L}7-{L}8-{L}9+{L}10-{L}11",
            22: f"={L}12+{L}10-{L}11",
            23: f"={L}22-{L}21",
            24: f"=IFERROR({L}23/{L}13,0)",
            25: f"=IFERROR({L}13/{L}14,0)",
            26: f"=MAX(0,{L}15-{L}21)",
            27: f"={L}6-{L}7",
            28: f"={L}8+{L}9",
        }
        for row, formula in formulas.items():
            _formula(treasury.cell(row, column), formula, MULT if row in (24, 25) else CUR)

    allocation = _sheet(
        workbook,
        "Capital Allocation",
        "Capital Allocation Sources, Uses, and Distribution Funding",
        ["Capital use", "Base", "Adversarial", "Decision rule"],
        [4, 40, 18, 18, 48],
    )
    allocation_rows = [
        ("Operating cash flow", "='Treasury & Liquidity'!C6", "='Treasury & Liquidity'!D6", "internal funding source"),
        ("Capital expenditure", "='Treasury & Liquidity'!C7", "='Treasury & Liquidity'!D7", "maintenance and growth investment"),
        ("Dividends", "='Treasury & Liquidity'!C8", "='Treasury & Liquidity'!D8", "fund only after minimum liquidity"),
        ("Buybacks", "='Treasury & Liquidity'!C9", "='Treasury & Liquidity'!D9", "fund only after minimum liquidity"),
        ("Debt repayment", "='Treasury & Liquidity'!C11", "='Treasury & Liquidity'!D11", "deleveraging use"),
        ("Debt issuance", "='Treasury & Liquidity'!C10", "='Treasury & Liquidity'!D10", "external funding source"),
    ]
    for row, (label, base, adverse, note) in enumerate(allocation_rows, start=5):
        allocation.cell(row, 2, label)
        _formula(allocation.cell(row, 3), base, CUR)
        _formula(allocation.cell(row, 4), adverse, CUR)
        allocation.cell(row, 5, note)
    allocation["B13"] = "Distribution funding residual"
    _formula(allocation["C13"], "=MAX(0,C7+C8-C5+C6)", CUR)
    _formula(allocation["D13"], "=MAX(0,D7+D8-D5+D6)", CUR)
    allocation["B14"] = "Cash rollforward residual"
    _formula(allocation["C14"], "='Treasury & Liquidity'!C21-('Treasury & Liquidity'!C5+'Treasury & Liquidity'!C6-'Treasury & Liquidity'!C7-'Treasury & Liquidity'!C8-'Treasury & Liquidity'!C9+'Treasury & Liquidity'!C10-'Treasury & Liquidity'!C11)", CUR)
    _formula(allocation["D14"], "='Treasury & Liquidity'!D21-('Treasury & Liquidity'!D5+'Treasury & Liquidity'!D6-'Treasury & Liquidity'!D7-'Treasury & Liquidity'!D8-'Treasury & Liquidity'!D9+'Treasury & Liquidity'!D10-'Treasury & Liquidity'!D11)", CUR)

    credit = _sheet(
        workbook,
        "Credit Metrics",
        "Leverage, Coverage, Liquidity, and Distribution Capacity",
        ["Metric", "Base", "Adversarial", "Threshold / interpretation"],
        [4, 40, 18, 18, 48],
    )
    credit_rows = [
        ("Ending cash", "='Treasury & Liquidity'!C21", "='Treasury & Liquidity'!D21", "must exceed minimum cash"),
        ("Net leverage", "='Treasury & Liquidity'!C24", "='Treasury & Liquidity'!D24", "must remain below maximum"),
        ("Interest coverage", "='Treasury & Liquidity'!C25", "='Treasury & Liquidity'!D25", "must remain above minimum"),
        ("Financing gap", "='Treasury & Liquidity'!C26", "='Treasury & Liquidity'!D26", "must equal zero"),
        ("Distribution funding residual", "='Capital Allocation'!C13", "='Capital Allocation'!D13", "must equal zero"),
    ]
    for row, (label, base, adverse, note) in enumerate(credit_rows, start=5):
        credit.cell(row, 2, label)
        _formula(credit.cell(row, 3), base, MULT if row in (6, 7) else CUR)
        _formula(credit.cell(row, 4), adverse, MULT if row in (6, 7) else CUR)
        credit.cell(row, 5, note)

    _decision_sheet(
        workbook,
        "Corporate Finance Decision Dashboard and Independent Checks",
        [
            ("Cash rollforward residual", "='Capital Allocation'!D14", '=IF(ABS(C5)<0.01,"PASS","FAIL")', "Cash sources and uses must conserve."),
            ("Debt rollforward residual", "='Treasury & Liquidity'!D22-('Treasury & Liquidity'!D12+'Treasury & Liquidity'!D10-'Treasury & Liquidity'!D11)", '=IF(ABS(C6)<0.01,"PASS","FAIL")', "Ending debt must reconcile to issuance and repayment."),
            ("Minimum cash", "='Treasury & Liquidity'!D21-'Treasury & Liquidity'!D15", '=IF(C7>=0,"PASS","BREACH")', "Escalate funding if ending cash falls below the minimum buffer."),
            ("Net leverage", "='Treasury & Liquidity'!D24", '=IF(C8<=\'Treasury & Liquidity\'!D16,"PASS","BREACH")', "Leverage must remain within the approved capital-structure limit."),
            ("Interest coverage", "='Treasury & Liquidity'!D25", '=IF(C9>=\'Treasury & Liquidity\'!D17,"PASS","BREACH")', "Cash interest must remain covered by operating earnings."),
            ("Distribution funding", "='Capital Allocation'!D13", '=IF(C10<0.01,"PASS","BREACH")', "Dividends and buybacks cannot rely on uncovered liquidity."),
        ],
    )


def enrich_private_credit(workbook) -> None:
    _delete(workbook, "Portfolio & Concentration", "Amendment Economics", "Decision & Checks")
    portfolio = _sheet(
        workbook,
        "Portfolio & Concentration",
        "Portfolio Exposure, Concentration, and Weighted Credit Metrics",
        ["Borrower / segment", "Base exposure", "Downside exposure", "Base leverage", "Downside leverage"],
        [4, 34, 18, 18, 18, 18],
    )
    rows = [
        ("Borrower A", 35.0, 45.0, 4.0, 7.0),
        ("Borrower B", 25.0, 30.0, 4.5, 6.5),
        ("Borrower C", 20.0, 15.0, 3.5, 5.0),
        ("Borrower D", 12.0, 7.0, 3.0, 4.0),
        ("Borrower E", 8.0, 3.0, 2.5, 3.5),
    ]
    for row, values in enumerate(rows, start=5):
        portfolio.cell(row, 2, values[0])
        for column, value in enumerate(values[1:], start=3):
            _input(portfolio.cell(row, column), value, MULT if column >= 5 else CUR)
    portfolio["B12"] = "Total exposure"
    _formula(portfolio["C12"], "=SUM(C5:C9)", CUR)
    _formula(portfolio["D12"], "=SUM(D5:D9)", CUR)
    portfolio["B13"] = "Top-1 concentration"
    _formula(portfolio["C13"], "=IFERROR(MAX(C5:C9)/C12,0)", PCT)
    _formula(portfolio["D13"], "=IFERROR(MAX(D5:D9)/D12,0)", PCT)
    portfolio["B14"] = "Weighted leverage"
    _formula(portfolio["C14"], "=IFERROR(SUMPRODUCT(C5:C9,E5:E9)/C12,0)", MULT)
    _formula(portfolio["D14"], "=IFERROR(SUMPRODUCT(D5:D9,F5:F9)/D12,0)", MULT)
    portfolio["B15"] = "Maximum concentration"
    _input(portfolio["C15"], 0.40, PCT)
    _input(portfolio["D15"], 0.40, PCT)

    amendment = _sheet(
        workbook,
        "Amendment Economics",
        "Debt Rollforward, PIK, DSCR, Leverage, and Recovery",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 46],
    )
    inputs = [
        ("Opening debt", 350.0, 350.0, CUR),
        ("Mandatory amortization", 10.0, 0.0, CUR),
        ("Cash sweep", 40.0, 0.0, CUR),
        ("PIK interest", 0.0, 30.0, CUR),
        ("CFADS", 80.0, 20.0, CUR),
        ("Cash interest", 25.0, 40.0, CUR),
        ("EBITDA", 100.0, 50.0, CUR),
        ("Recovery enterprise value", 500.0, 180.0, CUR),
        ("Senior claims", 50.0, 50.0, CUR),
        ("Lender claim", 300.0, 380.0, CUR),
        ("Minimum DSCR", 1.0, 1.0, MULT),
        ("Maximum leverage", 6.0, 6.0, MULT),
        ("Minimum recovery", 0.60, 0.60, PCT),
        ("Amendment fee", 0.01, 0.03, PCT),
        ("Extension years", 0.0, 2.0, NUM),
    ]
    for row, (label, base, adverse, fmt) in enumerate(inputs, start=5):
        amendment.cell(row, 2, label)
        _input(amendment.cell(row, 3), base, fmt)
        _input(amendment.cell(row, 4), adverse, fmt)
    outputs = ["Ending debt", "Debt service", "DSCR", "Leverage", "Recovery value", "Recovery rate", "LGD", "Amendment fee dollars"]
    for row, label in enumerate(outputs, start=23):
        amendment.cell(row, 2, label)
    for column in (3, 4):
        L = get_column_letter(column)
        formulas = {
            23: f"=MAX(0,{L}5-{L}6-{L}7+{L}8)",
            24: f"={L}6+{L}10",
            25: f"=IFERROR({L}9/{L}24,0)",
            26: f"=IFERROR({L}23/{L}11,0)",
            27: f"=MIN({L}14,MAX(0,{L}12-{L}13))",
            28: f"=IFERROR({L}27/{L}14,1)",
            29: f"=1-{L}28",
            30: f"={L}5*{L}18",
        }
        for row, formula in formulas.items():
            fmt = PCT if row in (28, 29) else (MULT if row in (25, 26) else CUR)
            _formula(amendment.cell(row, column), formula, fmt)

    _decision_sheet(
        workbook,
        "Private Credit Decision Dashboard and Independent Checks",
        [
            ("Debt rollforward residual", "='Amendment Economics'!D23-MAX(0,'Amendment Economics'!D5-'Amendment Economics'!D6-'Amendment Economics'!D7+'Amendment Economics'!D8)", '=IF(ABS(C5)<0.01,"PASS","FAIL")', "Ending debt must reconcile to amortization, sweep, and PIK."),
            ("Recovery / LGD identity", "='Amendment Economics'!D28+'Amendment Economics'!D29-1", '=IF(ABS(C6)<0.000001,"PASS","FAIL")', "Recovery rate plus LGD must equal one."),
            ("Downside DSCR", "='Amendment Economics'!D25", '=IF(C7>=\'Amendment Economics\'!D15,"PASS","BREACH")', "CFADS must cover cash interest and mandatory amortization."),
            ("Downside leverage", "='Amendment Economics'!D26", '=IF(C8<=\'Amendment Economics\'!D16,"PASS","BREACH")', "Debt must remain within approved leverage."),
            ("Downside recovery", "='Amendment Economics'!D28", '=IF(C9>=\'Amendment Economics\'!D17,"PASS","BREACH")', "Escalate impairment and amendment strategy."),
            ("Debt compounding", "='Amendment Economics'!D8-'Amendment Economics'!D6-'Amendment Economics'!D7", '=IF(C10<=0,"PASS","BREACH")', "PIK cannot exceed deleveraging without explicit escalation."),
            ("Portfolio concentration", "='Portfolio & Concentration'!D13", '=IF(C11<=\'Portfolio & Concentration\'!D15,"PASS","BREACH")', "Single-name concentration must remain below the approved limit."),
        ],
    )


def enrich_debt_finance(workbook) -> None:
    _delete(workbook, "Refinancing & Rates", "Decision & Checks")
    ladder = workbook["Maturity Ladder"]
    ladder["B15"] = "Near-term maturities (Years 1-2)"
    ladder["C15"] = "=SUM(C10:D10)"
    ladder["C15"].number_format = CUR
    ladder["B16"] = "Available cash and committed lines"
    ladder["C16"] = 150.0
    ladder["C16"].font = BLUE
    ladder["C16"].fill = YELLOW_FILL
    ladder["C16"].number_format = CUR
    ladder["B17"] = "Refinancing gap"
    ladder["C17"] = "=MAX(0,C15-C16)"
    ladder["C17"].number_format = CUR
    ladder["B18"] = "Largest maturity / total"
    ladder["C18"] = "=IFERROR(MAX(C10:H10)/SUM(C10:H10),0)"
    ladder["C18"].number_format = PCT

    rates = _sheet(
        workbook,
        "Refinancing & Rates",
        "Debt Rollforward, Maturity Wall, Weighted Cost, and Coverage",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 46],
    )
    inputs = [
        ("Opening debt", 500.0, 500.0, CUR),
        ("Issuance", 100.0, 300.0, CUR),
        ("Repayments", 150.0, 50.0, CUR),
        ("Year 1 maturity", 50.0, 300.0, CUR),
        ("Year 2 maturity", 75.0, 250.0, CUR),
        ("Year 3 maturity", 100.0, 100.0, CUR),
        ("Year 4 maturity", 100.0, 50.0, CUR),
        ("Year 5 maturity", 125.0, 50.0, CUR),
        ("Liquidity", 100.0, 50.0, CUR),
        ("Committed lines", 50.0, 50.0, CUR),
        ("Tranche 1 amount", 300.0, 500.0, CUR),
        ("Tranche 1 rate", 0.06, 0.12, PCT),
        ("Tranche 2 amount", 150.0, 250.0, CUR),
        ("Tranche 2 rate", 0.07, 0.15, PCT),
        ("EBITDA", 150.0, 80.0, CUR),
        ("Maximum maturity concentration", 0.40, 0.35, PCT),
        ("Minimum interest coverage", 2.0, 2.0, MULT),
        ("Maximum weighted cost", 0.10, 0.10, PCT),
    ]
    for row, (label, base, adverse, fmt) in enumerate(inputs, start=5):
        rates.cell(row, 2, label)
        _input(rates.cell(row, 3), base, fmt)
        _input(rates.cell(row, 4), adverse, fmt)
    outputs = ["Ending debt", "Near-term maturities", "Refinancing gap", "Maturity concentration", "Weighted cost", "Cash interest", "Interest coverage"]
    for row, label in enumerate(outputs, start=26):
        rates.cell(row, 2, label)
    for column in (3, 4):
        L = get_column_letter(column)
        formulas = {
            26: f"=MAX(0,{L}5+{L}6-{L}7)",
            27: f"={L}8+{L}9",
            28: f"=MAX(0,{L}27-{L}13-{L}14)",
            29: f"=IFERROR(MAX({L}8:{L}12)/SUM({L}8:{L}12),0)",
            30: f"=IFERROR(({L}15*{L}16+{L}17*{L}18)/({L}15+{L}17),0)",
            31: f"={L}26*{L}30",
            32: f"=IFERROR({L}19/{L}31,0)",
        }
        for row, formula in formulas.items():
            fmt = PCT if row in (29, 30) else (MULT if row == 32 else CUR)
            _formula(rates.cell(row, column), formula, fmt)

    _decision_sheet(
        workbook,
        "Debt Finance Decision Dashboard and Independent Checks",
        [
            ("Debt rollforward residual", "='Refinancing & Rates'!D26-MAX(0,'Refinancing & Rates'!D5+'Refinancing & Rates'!D6-'Refinancing & Rates'!D7)", '=IF(ABS(C5)<0.01,"PASS","FAIL")', "Debt must reconcile to issuance and repayment."),
            ("Weighted-cost residual", "='Refinancing & Rates'!D30*('Refinancing & Rates'!D15+'Refinancing & Rates'!D17)-('Refinancing & Rates'!D15*'Refinancing & Rates'!D16+'Refinancing & Rates'!D17*'Refinancing & Rates'!D18)", '=IF(ABS(C6)<0.01,"PASS","FAIL")', "Weighted pricing must reconcile to tranche balances and rates."),
            ("Refinancing gap", "='Refinancing & Rates'!D28", '=IF(C7=0,"PASS","BREACH")', "Near-term maturities must be covered by liquidity and committed facilities."),
            ("Maturity concentration", "='Refinancing & Rates'!D29", '=IF(C8<=\'Refinancing & Rates\'!D20,"PASS","BREACH")', "Escalate concentrated maturity walls."),
            ("Interest coverage", "='Refinancing & Rates'!D32", '=IF(C9>=\'Refinancing & Rates\'!D21,"PASS","BREACH")', "EBITDA must cover pro forma cash interest."),
            ("Weighted refinancing cost", "='Refinancing & Rates'!D30", '=IF(C10<=\'Refinancing & Rates\'!D22,"PASS","BREACH")', "Escalate expensive issuance and rate-reset risk."),
        ],
    )


def enrich_public_finance(workbook) -> None:
    _delete(workbook, "Revenue Stress", "Decision & Checks")
    stress = _sheet(
        workbook,
        "Revenue Stress",
        "Debt Dynamics, Revenue Coverage, Reserves, and Fiscal Stress",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 48],
    )
    inputs = [
        ("Opening debt ratio", 0.60, 1.20, MULT),
        ("Nominal interest rate", 0.05, 0.10, PCT),
        ("Nominal growth rate", 0.04, 0.01, PCT),
        ("Primary balance ratio", 0.015, -0.03, PCT),
        ("Pledged revenue", 120.0, 50.0, CUR),
        ("Debt service", 80.0, 70.0, CUR),
        ("Unrestricted reserves", 20.0, 5.0, CUR),
        ("Operating expenditure", 100.0, 100.0, CUR),
        ("Maximum debt ratio", 1.0, 1.0, MULT),
        ("Minimum DSCR", 1.20, 1.20, MULT),
        ("Minimum reserve coverage", 0.10, 0.10, PCT),
    ]
    for row, (label, base, adverse, fmt) in enumerate(inputs, start=5):
        stress.cell(row, 2, label)
        _input(stress.cell(row, 3), base, fmt)
        _input(stress.cell(row, 4), adverse, fmt)
    outputs = ["Projected debt ratio", "Debt-stabilizing primary balance", "Primary-balance gap", "DSCR", "Reserve coverage"]
    for row, label in enumerate(outputs, start=19):
        stress.cell(row, 2, label)
    for column in (3, 4):
        L = get_column_letter(column)
        formulas = {
            19: f"=((1+{L}6)/(1+{L}7))*{L}5-{L}8",
            20: f"=(({L}6-{L}7)/(1+{L}7))*{L}5",
            21: f"={L}8-{L}20",
            22: f"=IFERROR({L}9/{L}10,0)",
            23: f"=IFERROR({L}11/{L}12,0)",
        }
        for row, formula in formulas.items():
            _formula(stress.cell(row, column), formula, PCT if row in (20, 21, 23) else MULT)

    _decision_sheet(
        workbook,
        "Public Finance Decision Dashboard and Independent Checks",
        [
            ("Debt-ratio identity residual", "='Revenue Stress'!D19-(((1+'Revenue Stress'!D6)/(1+'Revenue Stress'!D7))*'Revenue Stress'!D5-'Revenue Stress'!D8)", '=IF(ABS(C5)<0.000001,"PASS","FAIL")', "Projected debt must follow the debt-dynamics identity."),
            ("Stabilizing-balance residual", "='Revenue Stress'!D20-((('Revenue Stress'!D6-'Revenue Stress'!D7)/(1+'Revenue Stress'!D7))*'Revenue Stress'!D5)", '=IF(ABS(C6)<0.000001,"PASS","FAIL")', "Required primary balance must reconcile to r-g and opening debt."),
            ("Projected debt ratio", "='Revenue Stress'!D19", '=IF(C7<=\'Revenue Stress\'!D13,"PASS","BREACH")', "Escalate debt trajectories above the approved ceiling."),
            ("Primary-balance gap", "='Revenue Stress'!D21", '=IF(C8>=0,"PASS","BREACH")', "A negative gap implies a destabilizing fiscal stance."),
            ("Debt-service coverage", "='Revenue Stress'!D22", '=IF(C9>=\'Revenue Stress\'!D14,"PASS","BREACH")', "Pledged revenue must cover scheduled debt service."),
            ("Reserve coverage", "='Revenue Stress'!D23", '=IF(C10>=\'Revenue Stress\'!D15,"PASS","BREACH")', "Unrestricted reserves must cover the approved expenditure buffer."),
        ],
    )


def enrich_venture_capital(workbook) -> None:
    _delete(workbook, "Ownership & Dilution", "Reserves & Follow-ons", "Exit Waterfall", "Decision & Checks")
    ownership = _sheet(
        workbook,
        "Ownership & Dilution",
        "Post-Money Ownership, Option Pool, and Round Pricing",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 46],
    )
    inputs = [
        ("Pre-money value", 80.0, 100.0, CUR),
        ("Investment", 20.0, 20.0, CUR),
        ("Existing shares", 8.0, 10.0, NUM),
        ("New investor shares", 2.0, 4.0, NUM),
        ("Option-pool expansion", 0.0, 5.0, NUM),
        ("Maximum investor ownership", 0.40, 0.40, PCT),
        ("Maximum option-pool dilution", 0.20, 0.20, PCT),
    ]
    for row, (label, base, adverse, fmt) in enumerate(inputs, start=5):
        ownership.cell(row, 2, label)
        _input(ownership.cell(row, 3), base, fmt)
        _input(ownership.cell(row, 4), adverse, fmt)
    outputs = ["Post-money value", "Total post-money shares", "Investor ownership", "Founder ownership", "Option-pool ownership", "Round price / share", "Pre-money implied price / share"]
    for row, label in enumerate(outputs, start=15):
        ownership.cell(row, 2, label)
    for column in (3, 4):
        L = get_column_letter(column)
        formulas = {
            15: f"={L}5+{L}6",
            16: f"={L}7+{L}8+{L}9",
            17: f"=IFERROR({L}8/{L}16,0)",
            18: f"=IFERROR({L}7/{L}16,0)",
            19: f"=IFERROR({L}9/{L}16,0)",
            20: f"=IFERROR({L}6/{L}8,0)",
            21: f"=IFERROR({L}5/{L}7,0)",
        }
        for row, formula in formulas.items():
            _formula(ownership.cell(row, column), formula, PCT if row in (17, 18, 19) else CUR)

    reserves = _sheet(
        workbook,
        "Reserves & Follow-ons",
        "Follow-On Requirements, Reserves, and Ownership Defense",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 46],
    )
    reserve_inputs = [
        ("Follow-on required", 10.0, 30.0, CUR),
        ("Available reserves", 15.0, 5.0, CUR),
        ("Pro rata ownership target", 0.20, 0.20, PCT),
        ("Next-round size", 40.0, 60.0, CUR),
    ]
    for row, (label, base, adverse, fmt) in enumerate(reserve_inputs, start=5):
        reserves.cell(row, 2, label)
        _input(reserves.cell(row, 3), base, fmt)
        _input(reserves.cell(row, 4), adverse, fmt)
    reserves["B11"] = "Reserve gap"
    _formula(reserves["C11"], "=MAX(0,C5-C6)", CUR)
    _formula(reserves["D11"], "=MAX(0,D5-D6)", CUR)
    reserves["B12"] = "Pro rata capital required"
    _formula(reserves["C12"], "=C7*C8", CUR)
    _formula(reserves["D12"], "=D7*D8", CUR)
    reserves["B13"] = "Ownership-defense shortfall"
    _formula(reserves["C13"], "=MAX(0,C12-C6)", CUR)
    _formula(reserves["D13"], "=MAX(0,D12-D6)", CUR)

    waterfall = _sheet(
        workbook,
        "Exit Waterfall",
        "Exit Proceeds, Investor MOIC, and Return Thresholds",
        ["Metric", "Base", "Adversarial", "Units / control"],
        [4, 42, 18, 18, 46],
    )
    waterfall_inputs = [
        ("Exit equity value", 500.0, 80.0, CUR),
        ("Investment cost", 20.0, 20.0, CUR),
        ("Minimum gross MOIC", 3.0, 3.0, MULT),
    ]
    for row, (label, base, adverse, fmt) in enumerate(waterfall_inputs, start=5):
        waterfall.cell(row, 2, label)
        _input(waterfall.cell(row, 3), base, fmt)
        _input(waterfall.cell(row, 4), adverse, fmt)
    waterfall["B11"] = "Investor ownership"
    _formula(waterfall["C11"], "='Ownership & Dilution'!C17", PCT)
    _formula(waterfall["D11"], "='Ownership & Dilution'!D17", PCT)
    waterfall["B12"] = "Investor exit proceeds"
    _formula(waterfall["C12"], "=C5*C11", CUR)
    _formula(waterfall["D12"], "=D5*D11", CUR)
    waterfall["B13"] = "Gross MOIC"
    _formula(waterfall["C13"], "=IFERROR(C12/C6,0)", MULT)
    _formula(waterfall["D13"], "=IFERROR(D12/D6,0)", MULT)

    # ---- Full cap-table liquidation-preference waterfall ----
    # The section above answers "what does ONE new investor's slice return."
    # It says nothing about liquidation preferences, seniority, or what
    # happens to the REST of the cap table at exit -- which is the actual
    # "waterfall" a VC cap table needs (1x non-participating pref, senior
    # to junior, with a convert-to-common test), driven off Cap Table's
    # real share/price data rather than a single fully-diluted percentage.
    # Reuses C5/D5 (Exit equity value) above as total exit proceeds for
    # both scenarios -- no separate/redundant total-proceeds input.
    waterfall["B16"] = "Liquidation Preference Waterfall (Full Cap Table, 1x Non-Participating)"
    waterfall["B16"].font = BOLD
    waterfall["B16"].fill = GRAY_FILL
    waterfall["B17"] = (
        "Seniority order Series B > Series A > Seed > common (founders + option pool); each class "
        "takes the greater of its 1x liquidation preference (cascaded senior-to-junior against total "
        "exit proceeds) or its as-converted pro-rata share, decided by a single fund-wide test (does "
        "converting the WHOLE cap table pay more than the full preference stack) rather than each "
        "class electing independently -- switching class-by-class can pay out more than total proceeds "
        "exist (see tools/builders/build_vc_template.py for the derivation). SAFE holders pre-conversion "
        "are excluded from this cascade -- their claim depends on deal-specific conversion mechanics at "
        "exit, which this waterfall does not model."
    )
    waterfall["B17"].font = ITALIC_GRAY
    set_col_widths(waterfall, [4, 42, 16, 16, 14, 14, 14, 14, 14, 14, 14])

    cascade_headers = ["Class", "Invested $", "Liq. pref (1x)", "As-conv. %",
                        "Pref-stack: Base", "Pref-stack: Adv.", "As-conv.: Base",
                        "As-conv.: Adv.", "Actual: Base", "Actual: Adv."]
    for column, value in enumerate(cascade_headers, start=2):
        waterfall.cell(19, column, value)
    style_header_row(waterfall, 19, len(cascade_headers), start_col=2)

    class_rows = [
        ("Series B preferred", "='Cap Table'!F10", "=IFERROR('Cap Table'!D10,\"-\")"),
        ("Series A preferred", "='Cap Table'!F9", "=IFERROR('Cap Table'!D9,\"-\")"),
        ("Seed preferred", "='Cap Table'!F8", "=IFERROR('Cap Table'!D8,\"-\")"),
        ("Common (founders + pool)", "='Cap Table'!F5+'Cap Table'!F6",
         "=IFERROR('Cap Table'!D5+'Cap Table'!D6,\"-\")"),
    ]
    first_class_row = 20
    for offset, (cls, invested_formula, pct_formula) in enumerate(class_rows):
        row = first_class_row + offset
        waterfall.cell(row, 2, cls)
        _formula(waterfall.cell(row, 3), invested_formula, CUR)
        _formula(waterfall.cell(row, 4), f"=C{row}", CUR)  # 1x pref = invested amount
        _formula(waterfall.cell(row, 5), pct_formula, PCT)
    last_class_row = first_class_row + len(class_rows) - 1  # 23

    # Pref-stack cascade: senior-to-junior, common has no preference (gets
    # the residual) -- self-consistent by construction, rows always sum to
    # exactly the scenario's total proceeds.
    b, a1, seed, common = (first_class_row, first_class_row + 1, first_class_row + 2, first_class_row + 3)
    for col, total_cell in ((6, "$C$5"), (7, "$D$5")):  # F=Base, G=Adversarial
        L = get_column_letter(col)
        _formula(waterfall.cell(b, col), f"=MIN({total_cell},D{b})", CUR)
        _formula(waterfall.cell(a1, col), f"=MIN(MAX({total_cell}-{L}{b},0),D{a1})", CUR)
        _formula(waterfall.cell(seed, col), f"=MIN(MAX({total_cell}-{L}{b}-{L}{a1},0),D{seed})", CUR)
        _formula(waterfall.cell(common, col), f"=MAX(0,{total_cell}-{L}{b}-{L}{a1}-{L}{seed})", CUR)

    # As-converted: pro-rata by fully-diluted %, guarded against the
    # blank-template "-" text placeholder on column E.
    for row in range(first_class_row, last_class_row + 1):
        _formula(waterfall.cell(row, 8), f"=IFERROR($C$5*E{row},\"-\")", CUR)   # H = Base
        _formula(waterfall.cell(row, 9), f"=IFERROR($D$5*E{row},\"-\")", CUR)   # I = Adversarial

    regime_row = last_class_row + 1  # 24
    waterfall.cell(regime_row, 2, "Regime: as-converted pays more fund-wide than the full pref stack?")
    _formula(waterfall.cell(regime_row, 3),
             f'=IF($C$5>SUM(D{first_class_row}:D{last_class_row - 1}),"YES","NO")', "General")
    _formula(waterfall.cell(regime_row, 4),
             f'=IF($D$5>SUM(D{first_class_row}:D{last_class_row - 1}),"YES","NO")', "General")

    # Actual: switch the WHOLE table between scenarios based on one global
    # test (see the note above row 19) rather than a per-class election.
    for row in range(first_class_row, last_class_row + 1):
        _formula(waterfall.cell(row, 10),
                 f'=IFERROR(IF($C$5>SUM($D${first_class_row}:$D${last_class_row - 1}),H{row},F{row}),"-")', CUR)
        _formula(waterfall.cell(row, 11),
                 f'=IFERROR(IF($D$5>SUM($D${first_class_row}:$D${last_class_row - 1}),I{row},G{row}),"-")', CUR)
        waterfall.cell(row, 10).font = BOLD
        waterfall.cell(row, 11).font = BOLD

    total_base_row = regime_row + 2   # 26
    total_adv_row = total_base_row + 1  # 27
    waterfall.cell(total_base_row, 2, "Total distributed, Base (check -- should equal total exit proceeds)")
    _formula(waterfall.cell(total_base_row, 3), f"=SUM(J{first_class_row}:J{last_class_row})", CUR)
    waterfall.cell(total_adv_row, 2, "Total distributed, Adversarial (check -- should equal total exit proceeds)")
    _formula(waterfall.cell(total_adv_row, 3), f"=SUM(K{first_class_row}:K{last_class_row})", CUR)

    _decision_sheet(
        workbook,
        "Venture Capital Decision Dashboard and Independent Checks",
        [
            ("Post-money identity residual", "='Ownership & Dilution'!D15-'Ownership & Dilution'!D5-'Ownership & Dilution'!D6", '=IF(ABS(C5)<0.01,"PASS","FAIL")', "Post-money value must equal pre-money value plus new investment."),
            ("Ownership conservation residual", "='Ownership & Dilution'!D17+'Ownership & Dilution'!D18+'Ownership & Dilution'!D19-1", '=IF(ABS(C6)<0.000001,"PASS","FAIL")', "Investor, founder, and option-pool ownership must sum to one."),
            ("Follow-on reserve gap", "='Reserves & Follow-ons'!D11", '=IF(C7=0,"PASS","BREACH")', "Reserves must cover approved follow-on requirements."),
            ("Investor ownership", "='Ownership & Dilution'!D17", '=IF(C8<=\'Ownership & Dilution\'!D10,"PASS","BREACH")', "Escalate ownership concentration and governance terms."),
            ("Option-pool dilution", "='Ownership & Dilution'!D19", '=IF(C9<=\'Ownership & Dilution\'!D11,"PASS","BREACH")', "Pool expansion must remain within the approved dilution boundary."),
            ("Round pricing", "='Ownership & Dilution'!D20-'Ownership & Dilution'!D21", '=IF(C10>=0,"PASS","BREACH")', "A negative price delta indicates a down round."),
            ("Gross MOIC", "='Exit Waterfall'!D13", '=IF(C11>=\'Exit Waterfall\'!D7,"PASS","BREACH")', "Exit proceeds must clear the approved return threshold."),
            ("Liquidation waterfall conservation, Base", f"='Exit Waterfall'!C{total_base_row}-'Exit Waterfall'!C5", '=IF(ABS(C12)<0.01,"PASS","FAIL")', "Total distributed across the full cap table must equal total exit proceeds."),
            ("Liquidation waterfall conservation, Adversarial", f"='Exit Waterfall'!C{total_adv_row}-'Exit Waterfall'!D5", '=IF(ABS(C13)<0.01,"PASS","FAIL")', "Total distributed across the full cap table must equal total exit proceeds."),
        ],
    )


ENRICHERS: dict[str, tuple[str, Callable]] = {
    "01": ("build_template.py", enrich_investment_banking),
    "02": ("build_template.py", enrich_corporate_finance),
    "05": ("build_private_credit_template.py", enrich_private_credit),
    "06": ("build_debt_finance_template.py", enrich_debt_finance),
    "07": ("build_public_finance_template.py", enrich_public_finance),
    "13": ("build_vc_template.py", enrich_venture_capital),
}


def build_model(model_id: str, output: Path) -> None:
    if model_id not in ENRICHERS:
        raise KeyError(f"unsupported legacy model {model_id}")
    script_name, enrich = ENRICHERS[model_id]
    build_release(script_name, output, enrich)
