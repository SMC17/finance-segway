"""Build the integrated private-credit / debt-finance underwriting archetype.

The model combines lender yield and OID analysis with active-scenario operating
forecasts, debt and cash schedules, covenant headroom, recovery/LGD, sensitivity,
source discipline, and explicit model checks.

Usage:
    python tools/builders/build_credit_template.py --output CREDIT_template.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from institutional_helpers import (
    BLUE, CUR, DARK, GREEN, INTFMT, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_status_rules, finalize,
    formula_cell, header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "Operating Case", "Debt Schedule", "Covenants",
    "Yield & Spread", "Recovery", "Sensitivity", "Checks", "Sources", "RefreshLog",
]


def build(output: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEETS:
        wb.create_sheet(name)

    add_cover(wb, "[BORROWER] — Credit Underwriting Model", [
        ("Borrower / issuer:", "[Name]"),
        ("Facility type:", "Unitranche / TLB / revolver / notes"),
        ("Last refreshed:", "[date]"),
        ("Next covenant / maturity date:", "[date]"),
        ("Refresh cadence:", "Weekly"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    # Assumptions -----------------------------------------------------------------
    ws = wb["Assumptions"]
    title(ws, "B2:F2", "Underwriting Assumptions")
    header(ws, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / source"])
    assumptions = [
        ("Revenue (LTM)", 500.0, 500.0, "$mm", CUR),
        ("EBITDA margin", 0.20, 0.16, "%", PCT),
        ("Revenue growth", 0.04, -0.05, "%", PCT),
        ("Maintenance capex / revenue", 0.03, 0.035, "%", PCT),
        ("Cash taxes / EBT", 0.24, 0.20, "%", PCT),
        ("Change in NWC / revenue growth", 0.08, 0.10, "%", PCT),
        ("Opening cash", 25.0, 20.0, "$mm", CUR),
        ("Opening gross debt", 350.0, 350.0, "$mm", CUR),
        ("SOFR / base rate", 0.045, 0.055, "%", PCT),
        ("Cash spread", 0.055, 0.070, "%", PCT),
        ("OID / upfront fee", 0.02, 0.02, "%", PCT),
        ("PIK spread", 0.00, 0.02, "%", PCT),
        ("Mandatory amortization", 0.01, 0.00, "% opening debt", PCT),
        ("Cash sweep", 0.50, 0.25, "% excess cash", PCT),
        ("Minimum cash", 15.0, 20.0, "$mm", CUR),
        ("Maximum leverage covenant", 6.00, 5.50, "x", MULT),
        ("Minimum interest coverage covenant", 1.75, 1.50, "x", MULT),
        ("Minimum DSCR", 1.00, 1.00, "x", MULT),
        ("Recovery multiple on EBITDA", 5.0, 4.0, "x", MULT),
        ("Enterprise value haircut", 0.20, 0.35, "%", PCT),
        ("Maturity", 7.0, 7.0, "years", "0.0"),
    ]
    for row, (label, base, downside, units, fmt) in enumerate(assumptions, start=5):
        ws.cell(row=row, column=2, value=label)
        input_cell(ws.cell(row=row, column=3, value=base), fmt)
        input_cell(ws.cell(row=row, column=4, value=downside), fmt)
        active = ws.cell(row=row, column=5, value=f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        formula_cell(active, fmt, cross_sheet=True)
        ws.cell(row=row, column=6, value=units)
    set_widths(ws, {"A": 4, "B": 38, "C": 15, "D": 15, "E": 15, "F": 30})
    ws.freeze_panes = "A5"

    # Operating case --------------------------------------------------------------
    ws = wb["Operating Case"]
    title(ws, "B2:H2", "Operating Case & CFADS")
    header(ws, 4, 2, ["$mm", "LTM", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
    labels = [
        "Revenue", "Growth", "EBITDA", "EBITDA margin", "Less: cash taxes",
        "Less: maintenance capex", "Less: change in NWC", "CFADS before interest",
        "Less: cash interest", "CFADS after interest",
    ]
    for row, label in enumerate(labels, start=5):
        ws.cell(row=row, column=2, value=label)
    for row in (5, 7, 12, 14):
        ws.cell(row=row, column=2).font = Font(name="Arial", size=10, bold=True, color=DARK)
    formula_cell(ws.cell(5, 3, "=Assumptions!E5"), CUR, cross_sheet=True)
    formula_cell(ws.cell(7, 3, "=C5*Assumptions!E6"), CUR, cross_sheet=True)
    ws["C8"] = "=IFERROR(C7/C5,0)"; ws["C8"].number_format = PCT
    for col in range(4, 9):
        letter, prev = get_column_letter(col), get_column_letter(col - 1)
        formulas = {
            5: f"={prev}5*(1+Assumptions!$E$7)",
            6: f"=IFERROR({letter}5/{prev}5-1,0)",
            7: f"={letter}5*Assumptions!$E$6",
            8: f"=IFERROR({letter}7/{letter}5,0)",
            9: f"=MAX(0,{letter}7*Assumptions!$E$9)",
            10: f"={letter}5*Assumptions!$E$8",
            11: f"=MAX(0,({letter}5-{prev}5)*Assumptions!$E$10)",
            12: f"={letter}7-{letter}9-{letter}10-{letter}11",
            13: f"='Debt Schedule'!{letter}13",
            14: f"={letter}12-{letter}13",
        }
        for row, formula in formulas.items():
            cell = ws.cell(row=row, column=col, value=formula)
            cell.number_format = PCT if row in (6, 8) else CUR
            if row == 13:
                formula_cell(cell, CUR, cross_sheet=True)
    for row in (12, 14):
        total_row(ws, row, 2, 8, CUR)
    set_widths(ws, {"A": 4, "B": 38, **{get_column_letter(c): 13 for c in range(3, 9)}})
    ws.freeze_panes = "A5"

    # Debt and cash schedule -------------------------------------------------------
    ws = wb["Debt Schedule"]
    title(ws, "B2:H2", "Debt & Cash Schedule")
    header(ws, 4, 2, ["$mm", "Close", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
    labels = [
        "Beginning cash", "CFADS after interest", "Mandatory amortization", "Cash before sweep",
        "Cash sweep", "Ending cash", "Beginning debt", "Cash interest rate",
        "Cash interest expense", "PIK accrual", "Ending debt", "Average debt",
        "PIK interest expense", "Total interest expense",
    ]
    for row, label in enumerate(labels, start=5):
        ws.cell(row=row, column=2, value=label)
    for row in (10, 15, 18):
        ws.cell(row=row, column=2).font = Font(name="Arial", size=10, bold=True, color=DARK)
    formula_cell(ws.cell(5, 3, "=Assumptions!E11"), CUR, cross_sheet=True)
    ws["C10"] = "=C5"; ws["C10"].number_format = CUR
    formula_cell(ws.cell(11, 3, "=Assumptions!E12"), CUR, cross_sheet=True)
    ws["C15"] = "=C11"; ws["C15"].number_format = CUR
    ws["C16"] = "=C11"; ws["C16"].number_format = CUR
    formula_cell(ws.cell(12, 3, "=Assumptions!E13+Assumptions!E14"), PCT, cross_sheet=True)
    for col in range(4, 9):
        letter, prev = get_column_letter(col), get_column_letter(col - 1)
        formulas = {
            5: f"={prev}10",
            6: f"='Operating Case'!{letter}14",
            7: f"=MIN({letter}11,Assumptions!$E$12*Assumptions!$E$17)",
            8: f"={letter}5+{letter}6-{letter}7",
            9: f"=MIN(MAX(0,{letter}11-{letter}7),MAX(0,{letter}8-Assumptions!$E$19)*Assumptions!$E$18)",
            10: f"={letter}8-{letter}9",
            11: f"={prev}15",
            12: "=Assumptions!$E$13+Assumptions!$E$14",
            13: f"={letter}11*{letter}12",
            14: f"=MAX(0,{letter}11-{letter}7-{letter}9)*Assumptions!$E$16",
            15: f"=MAX(0,{letter}11-{letter}7-{letter}9+{letter}14)",
            16: f"=AVERAGE({letter}11,{letter}15)",
            17: f"={letter}11*Assumptions!$E$16",
            18: f"={letter}13+{letter}17",
        }
        for row, formula in formulas.items():
            cell = ws.cell(row=row, column=col, value=formula)
            cell.number_format = PCT if row == 12 else CUR
            if row == 6:
                formula_cell(cell, CUR, cross_sheet=True)
    for row in (10, 15, 18):
        total_row(ws, row, 2, 8, CUR)
    set_widths(ws, {"A": 4, "B": 40, **{get_column_letter(c): 13 for c in range(3, 9)}})
    ws.freeze_panes = "A5"

    # Covenants -------------------------------------------------------------------------
    ws = wb["Covenants"]
    title(ws, "B2:G2", "Covenant Compliance")
    header(ws, 4, 2, ["Metric", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"])
    labels = [
        "Gross leverage", "Maximum leverage covenant", "Leverage headroom",
        "Interest coverage", "Minimum coverage covenant", "Coverage headroom",
        "Debt service coverage ratio", "Minimum DSCR", "DSCR headroom", "Covenant status",
    ]
    for row, label in enumerate(labels, start=5):
        ws.cell(row=row, column=2, value=label)
    for row in (7, 10, 13, 14):
        ws.cell(row=row, column=2).font = Font(name="Arial", size=10, bold=True, color=DARK)
    for col in range(3, 8):
        letter = get_column_letter(col)
        model_col = get_column_letter(col + 1)
        formulas = {
            5: f"=IFERROR('Debt Schedule'!{model_col}15/'Operating Case'!{model_col}7,0)",
            6: "=Assumptions!$E$20",
            7: f"={letter}6-{letter}5",
            8: f"=IFERROR('Operating Case'!{model_col}7/'Debt Schedule'!{model_col}18,0)",
            9: "=Assumptions!$E$21",
            10: f"={letter}8-{letter}9",
            11: f"=IFERROR('Operating Case'!{model_col}12/('Debt Schedule'!{model_col}7+'Debt Schedule'!{model_col}13),0)",
            12: "=Assumptions!$E$22",
            13: f"={letter}11-{letter}12",
            14: f'=IF(AND({letter}7>=0,{letter}10>=0,{letter}13>=0),"PASS","FAIL")',
        }
        for row, formula in formulas.items():
            ws.cell(row=row, column=col, value=formula)
            if row < 14:
                ws.cell(row=row, column=col).number_format = MULT
    add_status_rules(ws, "C14:G14")
    set_widths(ws, {"A": 4, "B": 40, **{get_column_letter(c): 13 for c in range(3, 8)}})
    ws.freeze_panes = "A5"

    # Yield & spread ---------------------------------------------------------------
    ws = wb["Yield & Spread"]
    title(ws, "B2:E2", "Lender Yield & Spread")
    header(ws, 4, 2, ["Metric", "Value", "Units", "Method / note"])
    rows = [
        ("Cash coupon rate", "=Assumptions!E13+Assumptions!E14", "%", "Base rate + cash spread", PCT),
        ("PIK rate", "=Assumptions!E16", "%", "PIK spread", PCT),
        ("Issue price", "=100*(1-Assumptions!E15)", "per 100", "100 minus OID / upfront fee", "0.00"),
        ("Maturity", "=Assumptions!E25", "years", "Contractual maturity", "0.0"),
        ("Approx. cash YTM", "=IFERROR((C5*100+(100-C7)/C8)/((100+C7)/2),0)", "%", "Coupon plus OID accretion over average price", PCT2),
        ("Approx. all-in yield", "=C9+C6", "%", "Cash YTM + PIK rate", PCT2),
        ("Cash spread", "=Assumptions!E14*10000", "bps", "Cash spread × 10,000", '0 "bps"'),
        ("OID accretion / year", "=IFERROR(Assumptions!E15/Assumptions!E25*10000,0)", "bps", "OID ÷ maturity × 10,000", '0 "bps"'),
    ]
    for row, (label, formula, units, note, fmt) in enumerate(rows, start=5):
        ws.cell(row=row, column=2, value=label)
        formula_cell(ws.cell(row=row, column=3, value=formula), fmt, cross_sheet=row in (5, 6, 7, 8, 11, 12))
        ws.cell(row=row, column=4, value=units)
        ws.cell(row=row, column=5, value=note)
    set_widths(ws, {"A": 4, "B": 32, "C": 16, "D": 14, "E": 52})

    # Recovery and sensitivity -----------------------------------------------------
    ws = wb["Recovery"]
    title(ws, "B2:D2", "Recovery & Loss Analysis")
    header(ws, 4, 2, ["Recovery bridge", "Base", "Downside"])
    labels = [
        "Stressed EBITDA", "Recovery multiple", "Gross enterprise value", "Less: EV haircut",
        "Net distributable value", "Debt claim", "Recovery value", "Recovery rate", "Loss given default",
    ]
    base = [
        "='Operating Case'!H7", "=Assumptions!C23", "=C5*C6", "=C7*Assumptions!C24",
        "=C7-C8", "='Debt Schedule'!H15", "=MIN(C9,C10)", "=IFERROR(C11/C10,0)", "=1-C12",
    ]
    downside = [
        "='Operating Case'!H7*(1+Assumptions!D7)", "=Assumptions!D23", "=D5*D6", "=D7*Assumptions!D24",
        "=D7-D8", "='Debt Schedule'!H15", "=MIN(D9,D10)", "=IFERROR(D11/D10,0)", "=1-D12",
    ]
    for row, label in enumerate(labels, start=5):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=base[row - 5])
        ws.cell(row=row, column=4, value=downside[row - 5])
    for row in (5, 7, 8, 9, 10, 11):
        for col in (3, 4): ws.cell(row=row, column=col).number_format = CUR
    for col in (3, 4):
        ws.cell(row=6, column=col).number_format = MULT
        ws.cell(row=12, column=col).number_format = PCT
        ws.cell(row=13, column=col).number_format = PCT
    total_row(ws, 11, 2, 4, CUR)
    set_widths(ws, {"A": 4, "B": 38, "C": 18, "D": 18})

    ws = wb["Sensitivity"]
    title(ws, "B2:G2", "Recovery Sensitivity")
    header(ws, 4, 2, ["EBITDA haircut / recovery multiple", 3.0, 4.0, 5.0, 6.0, 7.0])
    for row, haircut in enumerate((0.0, 0.1, 0.2, 0.3, 0.4), start=5):
        ws.cell(row=row, column=2, value=haircut).number_format = PCT
        for col in range(3, 8):
            letter = get_column_letter(col)
            ws.cell(row=row, column=col, value=f"=IFERROR(MIN(1,MAX(0,('Operating Case'!$H$7*(1-$B{row})*{letter}$4)/'Debt Schedule'!$H$15)),1)")
            ws.cell(row=row, column=col).number_format = PCT
    ws.conditional_formatting.add("C5:G9", ColorScaleRule(start_type="min", start_color="E2F0D9", mid_type="percentile", mid_value=50, mid_color="FFF2CC", end_type="max", end_color="FCE4D6"))
    set_widths(ws, {"A": 4, "B": 36, **{get_column_letter(c): 12 for c in range(3, 8)}})

    # Checks and sources -----------------------------------------------------------
    ws = wb["Checks"]
    title(ws, "B2:C2", "Model Checks")
    header(ws, 4, 2, ["Check", "Status"])
    checks = [
        ("Debt never below zero", '=IF(MIN(\'Debt Schedule\'!D15:H15)>=0,"PASS","FAIL")'),
        ("Ending cash above minimum", '=IF(MIN(\'Debt Schedule\'!D10:H10)>=Assumptions!E19,"PASS","REVIEW")'),
        ("Recovery rate bounded", '=IF(AND(MIN(Recovery!C12:D12)>=0,MAX(Recovery!C12:D12)<=1),"PASS","FAIL")'),
        ("No covenant failures", '=IF(COUNTIF(Covenants!C14:G14,"FAIL")=0,"PASS","REVIEW")'),
        ("Overall model status", '=IF(COUNTIF(C5:C8,"FAIL")+COUNTIF(C5:C8,"REVIEW")=0,"PASS","REVIEW")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=formula)
    add_status_rules(ws, "C5:C9")
    set_widths(ws, {"A": 4, "B": 44, "C": 18})

    ws = wb["Sources"]
    title(ws, "B2:E2", "Source Register")
    header(ws, 4, 2, ["Input / dataset", "Source URL", "As-of", "Notes"])
    sources = [
        ("Financial statements / filings", "https://www.sec.gov/edgar/sec-api-documentation", "[period]", "Reconcile to audited statements"),
        ("Base rates / Treasury yields", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates", "[date]", "Document selected reference rate"),
        ("Loan documentation", "[data room or public filing URL]", "[date]", "Covenants, baskets, amortization, collateral"),
        ("Recovery methodology", "https://www.ivsc.org/standards/", "[date]", "Document EV and recovery bridge"),
    ]
    for row, values in enumerate(sources, start=5):
        for col, value in enumerate(values, start=2):
            ws.cell(row=row, column=col, value=value).font = Font(name="Arial", size=10, color=BLUE)
    set_widths(ws, {"A": 4, "B": 32, "C": 58, "D": 16, "E": 42})

    add_refresh_log(wb)
    finalize(wb, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("CREDIT_template.xlsx"))
    args = parser.parse_args()
    build(args.output)
    print(f"saved {args.output}")
