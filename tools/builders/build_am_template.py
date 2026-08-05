import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[FUND] — Asset Management Model", [
    ("Fund structure:", "Hedge fund / PE fund / mutual fund"),
    ("Vintage / inception:", "[date]"),
    ("Last refreshed:", "[date]"),
    ("Next capital call/distribution:", "[date]"),
    ("Refresh cadence:", "Weekly (monthly NAV strike)"),
])

# ---------------- FUND NAV ----------------
ws = wb.create_sheet("Fund NAV")
set_col_widths(ws, [4, 26, 14, 14, 14, 14])
ws["B2"] = "Fund NAV Build"; ws["B2"].font = TITLE
headers = ["", "", "Period 0", "Period 1", "Period 2", "Period 3"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 4, start_col=3)
rows = ["Beginning NAV", "+ Capital contributions", "+ Realized/unrealized gains",
        "- Fees & expenses", "- Distributions", "Ending NAV"]
r = 5
for label in rows:
    ws.cell(row=r, column=2, value=label).font = BOLD if label in ("Beginning NAV", "Ending NAV") else BLACK
    for c in range(3, 7):
        cell = ws.cell(row=r, column=c, value=0 if label != "Ending NAV" else None)
        cell.number_format = CUR
        cell.border = BORDER
        cell.font = BLUE if label != "Ending NAV" else BLACK
    r += 1
end_row = r - 1
begin_row = 5
for c in range(3, 7):
    col = get_column_letter(c)
    ws.cell(row=end_row, column=c,
            value=f"={col}{begin_row}+{col}{begin_row+1}+{col}{begin_row+2}-{col}{begin_row+3}-{col}{begin_row+4}")
    ws.cell(row=end_row, column=c).font = BOLD
    ws.cell(row=end_row, column=c).number_format = CUR
# chain beginning NAV of period n+1 = ending NAV of period n
for c in range(4, 7):
    col = get_column_letter(c)
    prev = get_column_letter(c-1)
    cell = ws.cell(row=begin_row, column=c, value=f"={prev}{end_row}")
    cell.font = BLACK  # formula now, not the hardcoded-0 input it replaced
ws.sheet_view.showGridLines = False

# ---------------- FEE WATERFALL ----------------
ws = wb.create_sheet("Fee Waterfall")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Fee Waterfall — Mgmt Fee + Carry w/ Hurdle"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Committed capital", 0, CUR),
    ("Management fee %", 0.02, PCT),
    ("Preferred return / hurdle %", 0.08, PCT),
    ("Carry / promote %", 0.20, PCT),
    ("GP catch-up %", 1.00, PCT),
    ("Gross fund profit (this period)", 0, CUR),
    ("Capital invested (basis for hurdle)", 0, CUR),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B13"] = "Waterfall"; ws["B13"].font = BOLD; ws["B13"].fill = GRAY_FILL
ws["B14"] = "1. Return of capital"
ws["C14"] = "=MIN(C10,C11)"; ws["C14"].number_format = CUR
ws["B15"] = "2. Preferred return (hurdle) to LPs"
ws["C15"] = "=MIN(MAX(C10-C14,0),C11*C7)"; ws["C15"].number_format = CUR
ws["B16"] = "3. GP catch-up"
ws["C16"] = "=MIN(MAX(C10-C14-C15,0),C15*C8/(1-C8))"; ws["C16"].number_format = CUR
ws["B17"] = "4. Remaining profit split (LP/GP per carry%)"
ws["C17"] = "=MAX(C10-C14-C15-C16,0)"; ws["C17"].number_format = CUR
ws["B18"] = "   GP share of remainder"
ws["C18"] = "=C17*C8"; ws["C18"].number_format = CUR
ws["B19"] = "   LP share of remainder"
ws["C19"] = "=C17*(1-C8)"; ws["C19"].number_format = CUR
ws["B21"] = "Total GP take (carry + catch-up)"
ws["C21"] = "=C16+C18"; ws["C21"].font = BOLD; ws["C21"].number_format = CUR
ws["B22"] = "Total LP take"
ws["C22"] = "=C14+C15+C19"; ws["C22"].font = BOLD; ws["C22"].number_format = CUR
ws["B23"] = "Annual management fee (separate from carry)"
ws["C23"] = "=C5*C6"; ws["C23"].number_format = CUR
for r2 in list(range(14, 20)) + [21, 22, 23]:
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- PERFORMANCE ATTRIBUTION ----------------
ws = wb.create_sheet("Performance Attribution")
set_col_widths(ws, [4, 24, 14, 14, 14, 40])
ws["B2"] = "Performance Attribution"; ws["B2"].font = TITLE
headers = ["", "Position/Strategy", "Contribution ($)", "Contribution (%)", "Weight (%)", "Notes"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers)-1)
for r in range(5, 12):
    ws.cell(row=r, column=2, value="[fill in]").font = BLUE
    c1 = ws.cell(row=r, column=3, value=0); c1.font = BLUE; c1.number_format = CUR; c1.border = BORDER
    c2 = ws.cell(row=r, column=4, value=0); c2.font = BLUE; c2.number_format = PCT; c2.border = BORDER
    c3 = ws.cell(row=r, column=5, value=0); c3.font = BLUE; c3.number_format = PCT; c3.border = BORDER
total_r = 12
ws.cell(row=total_r, column=2, value="Total fund return").font = BOLD
ws.cell(row=total_r, column=3, value="=SUM(C5:C11)").font = BOLD
ws.cell(row=total_r, column=3).number_format = CUR
ws.sheet_view.showGridLines = False

# ---------------- FUND PERFORMANCE (TVPI / DPI / NET IRR) ----------------
ws = wb.create_sheet("Fund Performance")
set_col_widths(ws, [4, 30, 14, 14, 14, 14, 40])
ws["B2"] = "Fund Performance — LP Reporting Metrics"; ws["B2"].font = TITLE
ws["B4"] = "Uses the period cash flows and ending NAV from the Fund NAV tab."
ws["B4"].font = ITALIC_GRAY

for i, h in enumerate(["", "", "Period 0", "Period 1", "Period 2", "Period 3"], start=1):
    ws.cell(row=6, column=i, value=h)
style_header_row(ws, 6, 4, start_col=3)
ws["B7"] = "LP net cash flow (distributions - contributions)"
for col in range(3, 7):
    letter = get_column_letter(col)
    ws.cell(row=7, column=col, value=f"='Fund NAV'!{letter}9-'Fund NAV'!{letter}6")
    ws.cell(row=7, column=col).number_format = CUR
    ws.cell(row=7, column=col).border = BORDER
ws["B8"] = "  + terminal NAV added to final period"
ws["C8"] = "='Fund NAV'!F10"; ws["C8"].font = GREEN; ws["C8"].number_format = CUR
ws["C8"].border = BORDER
ws["B9"] = "LP cash flow incl. terminal value (for IRR)"
for col in range(3, 6):
    letter = get_column_letter(col)
    ws.cell(row=9, column=col, value=f"={letter}7")
    ws.cell(row=9, column=col).number_format = CUR
    ws.cell(row=9, column=col).border = BORDER
ws.cell(row=9, column=6, value="=F7+C8")
ws.cell(row=9, column=6).number_format = CUR
ws.cell(row=9, column=6).border = BORDER

ws["B12"] = "Outputs"; ws["B12"].font = BOLD; ws["B12"].fill = GRAY_FILL
ws["B13"] = "Cumulative capital called"
ws["C13"] = "=SUM('Fund NAV'!C6:F6)"; ws["C13"].font = GREEN; ws["C13"].number_format = CUR
ws["B14"] = "Cumulative distributions"
ws["C14"] = "=SUM('Fund NAV'!C9:F9)"; ws["C14"].font = GREEN; ws["C14"].number_format = CUR
ws["B15"] = "Current NAV (residual value)"
ws["C15"] = "='Fund NAV'!F10"; ws["C15"].font = GREEN; ws["C15"].number_format = CUR
ws["B16"] = "DPI (Distributions to Paid-In)"
ws["C16"] = "=IFERROR(C14/C13,\"-\")"; ws["C16"].font = BOLD; ws["C16"].number_format = MULT
ws["B17"] = "RVPI (Residual Value to Paid-In)"
ws["C17"] = "=IFERROR(C15/C13,\"-\")"; ws["C17"].font = BOLD; ws["C17"].number_format = MULT
ws["B18"] = "TVPI (Total Value to Paid-In = DPI + RVPI)"
ws["C18"] = "=IFERROR(C16+C17,\"-\")"; ws["C18"].font = BOLD; ws["C18"].number_format = MULT
ws["D18"] = "Headline LP metric — total value (realized + unrealized) per dollar called"
ws["D18"].font = ITALIC_GRAY
ws["B19"] = "Net IRR to LPs (periodic, undated)"
ws["C19"] = "=IFERROR(IRR(C9:F9),\"-\")"; ws["C19"].font = BOLD; ws["C19"].number_format = PCT
ws["C19"].fill = YELLOW_FILL
ws["D19"] = "Treats each period as evenly spaced — use XIRR with real dates for called capital at irregular intervals"
ws["D19"].font = ITALIC_GRAY
for r2 in (13, 14, 15, 16, 17, 18, 19):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "AM_template.xlsx"
wb.save(out_path)
print("saved", out_path)
