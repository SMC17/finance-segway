"""Build the canonical options and derivatives decision model.

The workbook extends the original Black-Scholes template with an explicit
Base/Downside selector, Newton-Raphson implied-volatility iterations, a simple
volatility surface, portfolio Greeks, scenario P&L, and independent model
checks. It remains formula-driven and requires no macros.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, CUR2, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    formula_cell, header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "European Pricer", "Implied Vol", "Greeks",
    "Vol Surface", "Portfolio", "Scenario P&L", "Checks", "Sources",
    "RefreshLog",
]


def bs_call(spot: str, strike: str, maturity: str, rate: str,
            dividend: str, volatility: str) -> str:
    d1 = (
        f"(LN({spot}/{strike})+({rate}-{dividend}+0.5*{volatility}^2)*"
        f"{maturity})/({volatility}*SQRT({maturity}))"
    )
    d2 = f"({d1})-{volatility}*SQRT({maturity})"
    return (
        f"={spot}*EXP(-{dividend}*{maturity})*NORM.S.DIST({d1},TRUE)-"
        f"{strike}*EXP(-{rate}*{maturity})*NORM.S.DIST({d2},TRUE)"
    )


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[UNDERLYING] — Options & Derivatives Model", [
        ("Underlying:", "[Ticker / index / FX / commodity]"),
        ("Use case:", "Pricing / volatility / portfolio risk"),
        ("Last refreshed:", "[date]"),
        ("Next expiry / event:", "[date]"),
        ("Refresh cadence:", "Daily near expiry; weekly otherwise"),
        ("Active scenario:", "Base"),
        ("Units:", "Per option unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Pricing and Risk Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Spot price", 100.0, 80.0, "currency", CUR2),
        ("Strike price", 100.0, 100.0, "currency", CUR2),
        ("Time to expiry", 0.50, 0.50, "years", "0.0000"),
        ("Risk-free rate", 0.040, 0.030, "%", PCT2),
        ("Dividend yield", 0.010, 0.010, "%", PCT2),
        ("Implied volatility", 0.250, 0.450, "%", PCT2),
        ("Observed call price", 8.25, 15.00, "currency", CUR2),
        ("Contract multiplier", 100.0, 100.0, "units", "0"),
        ("Position count", 10.0, 10.0, "contracts", "0"),
        ("Strike skew slope", -0.20, -0.30, "vol per moneyness", "0.000"),
        ("Term-vol slope", 0.020, 0.040, "vol per log-year", "0.000"),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        formula_cell(
            sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})'),
            number_format,
            cross_sheet=True,
        )
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 34, "C": 15, "D": 15, "E": 15, "F": 30})
    sheet.freeze_panes = "A5"

    sheet = workbook["European Pricer"]
    title(sheet, "B2:F2", "European Black-Scholes Pricer")
    header(sheet, 4, 2, ["Metric", "Value", "Units", "Formula role", "Control"])
    rows = [
        ("d1", "=(LN(Assumptions!E5/Assumptions!E6)+(Assumptions!E8-Assumptions!E9+0.5*Assumptions!E10^2)*Assumptions!E7)/(Assumptions!E10*SQRT(Assumptions!E7))", "z", "standardized moneyness", ""),
        ("d2", "=C5-Assumptions!E10*SQRT(Assumptions!E7)", "z", "d1 less volatility-time", ""),
        ("Call price", "=Assumptions!E5*EXP(-Assumptions!E9*Assumptions!E7)*NORM.S.DIST(C5,TRUE)-Assumptions!E6*EXP(-Assumptions!E8*Assumptions!E7)*NORM.S.DIST(C6,TRUE)", "currency", "discounted expected payoff", ""),
        ("Put price", "=Assumptions!E6*EXP(-Assumptions!E8*Assumptions!E7)*NORM.S.DIST(-C6,TRUE)-Assumptions!E5*EXP(-Assumptions!E9*Assumptions!E7)*NORM.S.DIST(-C5,TRUE)", "currency", "discounted expected payoff", ""),
        ("Call intrinsic value", "=MAX(0,Assumptions!E5-Assumptions!E6)", "currency", "spot intrinsic", ""),
        ("Put intrinsic value", "=MAX(0,Assumptions!E6-Assumptions!E5)", "currency", "spot intrinsic", ""),
        ("Call time value", "=C7-C9", "currency", "price less intrinsic", ""),
        ("Put time value", "=C8-C10", "currency", "price less intrinsic", ""),
        ("Put-call parity residual", "=C7-C8-(Assumptions!E5*EXP(-Assumptions!E9*Assumptions!E7)-Assumptions!E6*EXP(-Assumptions!E8*Assumptions!E7))", "currency", "must equal zero", '=IF(ABS(C13)<0.000001,"PASS","FAIL")'),
    ]
    for row, (label, formula, units, note, control) in enumerate(rows, start=5):
        sheet.cell(row, 2, label)
        formula_cell(sheet.cell(row, 3, formula), "0.000000" if row in (5, 6, 13) else CUR2, cross_sheet=True)
        sheet.cell(row, 4, units)
        sheet.cell(row, 5, note)
        if control:
            sheet.cell(row, 6, control)
    add_status_rules(sheet, "F13")
    set_widths(sheet, {"A": 4, "B": 30, "C": 18, "D": 14, "E": 36, "F": 14})

    sheet = workbook["Implied Vol"]
    title(sheet, "B2:G2", "Newton-Raphson Implied Volatility")
    header(sheet, 4, 2, ["Iteration", "Volatility", "Model call", "Observed call", "Price error", "Vega"])
    sheet["B5"] = 0
    input_cell(sheet["C5"], PCT2)
    sheet["C5"] = 0.20
    for row in range(5, 15):
        if row > 5:
            sheet.cell(row, 2, row - 5)
            sheet.cell(row, 3, f"=MAX(0.0001,C{row-1}-F{row-1}/MAX(0.000001,G{row-1}))")
            sheet.cell(row, 3).number_format = PCT2
        volatility = f"C{row}"
        sheet.cell(row, 4, bs_call("Assumptions!$E$5", "Assumptions!$E$6", "Assumptions!$E$7", "Assumptions!$E$8", "Assumptions!$E$9", volatility))
        sheet.cell(row, 4).number_format = CUR2
        formula_cell(sheet.cell(row, 5, "=Assumptions!$E$11"), CUR2, cross_sheet=True)
        sheet.cell(row, 6, f"=D{row}-E{row}")
        sheet.cell(row, 6).number_format = CUR2
        d1 = f"(LN(Assumptions!$E$5/Assumptions!$E$6)+(Assumptions!$E$8-Assumptions!$E$9+0.5*C{row}^2)*Assumptions!$E$7)/(C{row}*SQRT(Assumptions!$E$7))"
        sheet.cell(row, 7, f"=Assumptions!$E$5*EXP(-Assumptions!$E$9*Assumptions!$E$7)*NORM.S.DIST({d1},FALSE)*SQRT(Assumptions!$E$7)")
        sheet.cell(row, 7).number_format = "0.0000"
    sheet["B17"] = "Converged implied volatility"
    sheet["C17"] = "=C14"
    sheet["C17"].number_format = PCT2
    sheet["D17"] = "Final absolute price error"
    sheet["E17"] = "=ABS(F14)"
    sheet["E17"].number_format = CUR2
    set_widths(sheet, {"A": 4, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16})

    sheet = workbook["Greeks"]
    title(sheet, "B2:E2", "European Greeks")
    header(sheet, 4, 2, ["Greek", "Call", "Put", "Interpretation"])
    d1 = "'European Pricer'!$C$5"
    d2 = "'European Pricer'!$C$6"
    spot = "Assumptions!$E$5"
    strike = "Assumptions!$E$6"
    maturity = "Assumptions!$E$7"
    rate = "Assumptions!$E$8"
    dividend = "Assumptions!$E$9"
    volatility = "Assumptions!$E$10"
    greek_rows = [
        ("Delta", f"=EXP(-{dividend}*{maturity})*NORM.S.DIST({d1},TRUE)", f"=EXP(-{dividend}*{maturity})*(NORM.S.DIST({d1},TRUE)-1)", "first-order spot exposure"),
        ("Gamma", f"=EXP(-{dividend}*{maturity})*NORM.S.DIST({d1},FALSE)/({spot}*{volatility}*SQRT({maturity}))", f"=C6", "delta curvature"),
        ("Vega", f"={spot}*EXP(-{dividend}*{maturity})*NORM.S.DIST({d1},FALSE)*SQRT({maturity})", f"=C7", "P&L per 100 vol points"),
        ("Theta / year", f"=-{spot}*EXP(-{dividend}*{maturity})*NORM.S.DIST({d1},FALSE)*{volatility}/(2*SQRT({maturity}))-{rate}*{strike}*EXP(-{rate}*{maturity})*NORM.S.DIST({d2},TRUE)+{dividend}*{spot}*EXP(-{dividend}*{maturity})*NORM.S.DIST({d1},TRUE)", f"=-{spot}*EXP(-{dividend}*{maturity})*NORM.S.DIST({d1},FALSE)*{volatility}/(2*SQRT({maturity}))+{rate}*{strike}*EXP(-{rate}*{maturity})*NORM.S.DIST(-{d2},TRUE)-{dividend}*{spot}*EXP(-{dividend}*{maturity})*NORM.S.DIST(-{d1},TRUE)", "annual time decay"),
        ("Rho", f"={maturity}*{strike}*EXP(-{rate}*{maturity})*NORM.S.DIST({d2},TRUE)", f"=-{maturity}*{strike}*EXP(-{rate}*{maturity})*NORM.S.DIST(-{d2},TRUE)", "rate exposure"),
    ]
    for row, (label, call_formula, put_formula, note) in enumerate(greek_rows, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, call_formula)
        sheet.cell(row, 4, put_formula)
        sheet.cell(row, 5, note)
        sheet.cell(row, 3).number_format = "0.000000"
        sheet.cell(row, 4).number_format = "0.000000"
    set_widths(sheet, {"A": 4, "B": 22, "C": 18, "D": 18, "E": 40})

    sheet = workbook["Vol Surface"]
    title(sheet, "B2:G2", "Illustrative Volatility Surface")
    header(sheet, 4, 2, ["Maturity / Strike", 80, 90, 100, 110, 120])
    maturities = [0.25, 0.50, 1.00, 2.00]
    for row, maturity_value in enumerate(maturities, start=5):
        sheet.cell(row, 2, maturity_value)
        sheet.cell(row, 2).number_format = "0.00"
        for column in range(3, 8):
            strike_cell = f"{get_column_letter(column)}$4"
            sheet.cell(row, column, f"=MAX(0.01,Assumptions!$E$10+Assumptions!$E$14*({strike_cell}/Assumptions!$E$5-1)+Assumptions!$E$15*LN($B{row}/Assumptions!$E$7))")
            sheet.cell(row, column).number_format = PCT2
    sheet.conditional_formatting.add(
        "C5:G8",
        ColorScaleRule(
            start_type="min", start_color="E2F0D9",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="FCE4D6",
        ),
    )
    set_widths(sheet, {"A": 4, "B": 22, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13})

    sheet = workbook["Portfolio"]
    title(sheet, "B2:M2", "Options Portfolio Greeks")
    header(sheet, 4, 2, [
        "Leg", "Type", "Quantity", "Strike", "Maturity", "Vol", "Price",
        "Delta", "Gamma", "Vega", "Theta", "Market value",
    ])
    legs = [
        ("Core call", "Call", 10, 100, 0.50, 0.25),
        ("Protective put", "Put", 10, 90, 0.50, 0.30),
        ("Short call", "Call", -5, 110, 0.50, 0.24),
        ("Long-dated call", "Call", 4, 105, 1.00, 0.27),
    ]
    for row, values in enumerate(legs, start=5):
        for column, value in enumerate(values, start=2):
            cell = sheet.cell(row, column, value)
            if column in (4, 5, 6, 7):
                input_cell(cell, PCT2 if column == 7 else (CUR2 if column == 5 else "0.00"))
        type_test = f'$C{row}="Call"'
        spot_ref = "Assumptions!$E$5"
        strike_ref = f"E{row}"
        maturity_ref = f"F{row}"
        vol_ref = f"G{row}"
        rate_ref = "Assumptions!$E$8"
        div_ref = "Assumptions!$E$9"
        d1_local = f"(LN({spot_ref}/{strike_ref})+({rate_ref}-{div_ref}+0.5*{vol_ref}^2)*{maturity_ref})/({vol_ref}*SQRT({maturity_ref}))"
        d2_local = f"({d1_local})-{vol_ref}*SQRT({maturity_ref})"
        call_price = f"{spot_ref}*EXP(-{div_ref}*{maturity_ref})*NORM.S.DIST({d1_local},TRUE)-{strike_ref}*EXP(-{rate_ref}*{maturity_ref})*NORM.S.DIST({d2_local},TRUE)"
        put_price = f"{strike_ref}*EXP(-{rate_ref}*{maturity_ref})*NORM.S.DIST(-({d2_local}),TRUE)-{spot_ref}*EXP(-{div_ref}*{maturity_ref})*NORM.S.DIST(-({d1_local}),TRUE)"
        sheet.cell(row, 8, f"=IF({type_test},{call_price},{put_price})")
        sheet.cell(row, 8).number_format = CUR2
        sheet.cell(row, 9, f"=IF({type_test},EXP(-{div_ref}*{maturity_ref})*NORM.S.DIST({d1_local},TRUE),EXP(-{div_ref}*{maturity_ref})*(NORM.S.DIST({d1_local},TRUE)-1))*D{row}")
        sheet.cell(row, 10, f"=EXP(-{div_ref}*{maturity_ref})*NORM.S.DIST({d1_local},FALSE)/({spot_ref}*{vol_ref}*SQRT({maturity_ref}))*D{row}")
        sheet.cell(row, 11, f"={spot_ref}*EXP(-{div_ref}*{maturity_ref})*NORM.S.DIST({d1_local},FALSE)*SQRT({maturity_ref})*D{row}")
        sheet.cell(row, 12, f"=-{spot_ref}*EXP(-{div_ref}*{maturity_ref})*NORM.S.DIST({d1_local},FALSE)*{vol_ref}/(2*SQRT({maturity_ref}))*D{row}")
        sheet.cell(row, 13, f"=H{row}*D{row}*Assumptions!$E$12")
        sheet.cell(row, 13).number_format = CUR
    total_row(sheet, 10, 2, 13)
    sheet["B10"] = "Portfolio total"
    for column in range(9, 14):
        letter = get_column_letter(column)
        sheet.cell(10, column, f"=SUM({letter}5:{letter}8)")
        sheet.cell(10, column).number_format = CUR if column == 13 else "0.000000"
    set_widths(sheet, {"A": 4, "B": 22, "C": 12, "D": 12, "E": 13, "F": 13, "G": 13, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 16})

    sheet = workbook["Scenario P&L"]
    title(sheet, "B2:H2", "Delta-Gamma-Vega Scenario P&L")
    header(sheet, 4, 2, ["Spot shock / Vol shock", -0.10, -0.05, 0.00, 0.05, 0.10])
    for column in range(3, 8):
        sheet.cell(4, column).number_format = PCT
    spot_shocks = [-0.20, -0.10, 0.00, 0.10, 0.20]
    for row, spot_shock in enumerate(spot_shocks, start=5):
        sheet.cell(row, 2, spot_shock)
        sheet.cell(row, 2).number_format = PCT
        for column in range(3, 8):
            letter = get_column_letter(column)
            sheet.cell(row, column, f"=Portfolio!$I$10*(Assumptions!$E$5*$B{row})+0.5*Portfolio!$J$10*(Assumptions!$E$5*$B{row})^2+Portfolio!$K$10*{letter}$4")
            sheet.cell(row, column).number_format = CUR
    sheet.conditional_formatting.add(
        "C5:G9",
        ColorScaleRule(
            start_type="min", start_color="FCE4D6",
            mid_type="percentile", mid_value=50, mid_color="FFF2CC",
            end_type="max", end_color="E2F0D9",
        ),
    )
    set_widths(sheet, {"A": 4, "B": 24, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14})

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("Put-call parity", '=IF(ABS(\'European Pricer\'!C13)<0.000001,"PASS","FAIL")'),
        ("Implied volatility convergence", '=IF(\'Implied Vol\'!E17<0.0001,"PASS","REVIEW")'),
        ("Gamma positive", '=IF(AND(Greeks!C6>0,Greeks!D6>0),"PASS","FAIL")'),
        ("Call price within bounds", '=IF(AND(\'European Pricer\'!C7>=MAX(0,Assumptions!E5*EXP(-Assumptions!E9*Assumptions!E7)-Assumptions!E6*EXP(-Assumptions!E8*Assumptions!E7)),\'European Pricer\'!C7<=Assumptions!E5*EXP(-Assumptions!E9*Assumptions!E7)),"PASS","FAIL")'),
        ("Portfolio formulas populated", '=IF(COUNT(Portfolio!H5:M8)=24,"PASS","FAIL")'),
        ("Overall", '=IF(COUNTIF(C5:C9,"FAIL")=0,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    add_status_rules(sheet, "C5:C10")
    set_widths(sheet, {"A": 4, "B": 42, "C": 18})

    add_sources(workbook, [
        ("Option contract specification", "[exchange / broker URL]", "[date]", "Multiplier, expiry, settlement, exercise style"),
        ("Underlying and option market data", "[market data URL]", "[timestamp]", "Spot, bid/ask, implied volatility"),
        ("Risk-free curve", "https://home.treasury.gov/resource-center/data-chart-center/interest-rates", "[date]", "Document interpolation and currency basis"),
        ("Dividend / carry assumption", "[issuer or market source]", "[date]", "Expected dividends or foreign rates"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("OPTIONS_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
