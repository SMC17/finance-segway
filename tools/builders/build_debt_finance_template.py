"""Build a distinct debt-finance issuance, refinancing, and maturity model."""
from __future__ import annotations
import argparse, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (CUR, MULT, PCT, PCT2, BPS, add_cover, add_refresh_log, add_sources,
    add_status_rules, finalize, formula_cell, header, input_cell, set_widths, title, total_row)  # noqa: E402

SHEETS=["Cover","Assumptions","Capital Structure","Maturity Ladder","Refinancing","Interest Rate Risk","Covenants","Recovery","Checks","Sources","RefreshLog"]

def build(output:Path)->None:
    wb=Workbook(); wb.remove(wb.active)
    for s in SHEETS: wb.create_sheet(s)
    add_cover(wb,"[ISSUER] — Debt Finance & Refinancing",[("Issuer:","[Name]"),("Transaction:","Refinancing / acquisition / general corporate"),
        ("Last refreshed:","[date]"),("Next maturity / launch:","[date]"),("Refresh cadence:","Weekly"),("Active scenario:","Base"),("Units:","$ in millions")])
    ws=wb["Assumptions"]; title(ws,"B2:F2","Issuance & Capital Structure Assumptions"); header(ws,4,2,["Assumption","Base","Downside","Active","Note"])
    rows=[("EBITDA",120.,95.,CUR),("Cash",40.,25.,CUR),("Revolver drawn",25.,75.,CUR),("Term loan",300.,300.,CUR),("Senior notes",250.,250.,CUR),
          ("Other debt",50.,50.,CUR),("Base rate",.045,.06,PCT),("Revolver spread",.025,.04,PCT),("Term loan spread",.04,.06,PCT),
          ("Senior notes coupon",.055,.07,PCT),("New issue spread",.045,.075,PCT),("OID / fees",.02,.03,PCT),("Minimum cash",25.,30.,CUR),
          ("Max leverage",5.5,5.0,MULT),("Min interest coverage",2.0,1.75,MULT),("Recovery multiple",5.,4.,MULT),("EV haircut",.2,.35,PCT)]
    for r,(l,b,d,fmt) in enumerate(rows,5): ws.cell(r,2,l); input_cell(ws.cell(r,3,b),fmt); input_cell(ws.cell(r,4,d),fmt); formula_cell(ws.cell(r,5,f'=IF(Cover!$C$9="Downside",D{r},C{r})'),fmt,cross_sheet=True)
    set_widths(ws,{"A":4,"B":36,"C":15,"D":15,"E":15,"F":30})

    ws=wb["Capital Structure"]; title(ws,"B2:G2","Current Capital Structure"); header(ws,4,2,["Instrument","Balance","Rate","Annual interest","Seniority","Secured?"])
    instruments=[("Revolver","=Assumptions!E7","=Assumptions!E12","=C5*D5","1","Yes"),("Term loan","=Assumptions!E8","=Assumptions!E11+Assumptions!E13","=C6*D6","2","Yes"),
                 ("Senior notes","=Assumptions!E9","=Assumptions!E14","=C7*D7","3","No"),("Other debt","=Assumptions!E10","=Assumptions!E14","=C8*D8","4","Varies")]
    for r,row in enumerate(instruments,5):
        for c,v in enumerate(row,2): ws.cell(r,c,v)
        ws.cell(r,3).number_format=CUR; ws.cell(r,4).number_format=PCT; ws.cell(r,5).number_format=CUR
    ws["B10"]="Gross debt"; ws["C10"]="=SUM(C5:C8)"; ws["C10"].number_format=CUR
    ws["B11"]="Net debt"; ws["C11"]="=C10-Assumptions!E6"; ws["C11"].number_format=CUR
    ws["B12"]="Cash interest"; ws["C12"]="=SUM(E5:E8)"; ws["C12"].number_format=CUR
    ws["B13"]="Gross leverage"; ws["C13"]="=IFERROR(C10/Assumptions!E5,0)"; ws["C13"].number_format=MULT
    ws["B14"]="Net leverage"; ws["C14"]="=IFERROR(C11/Assumptions!E5,0)"; ws["C14"].number_format=MULT
    ws["B15"]="Interest coverage"; ws["C15"]="=IFERROR(Assumptions!E5/C12,0)"; ws["C15"].number_format=MULT
    for r in (10,11,12,13,14,15): total_row(ws,r,2,3,ws.cell(r,3).number_format)
    set_widths(ws,{"A":4,"B":28,"C":16,"D":14,"E":18,"F":12,"G":14})

    ws=wb["Maturity Ladder"]; title(ws,"B2:H2","Contractual Maturity Ladder"); header(ws,4,2,["Instrument","Year 1","Year 2","Year 3","Year 4","Year 5","Beyond"])
    balances=[("Revolver",0.,0.,25.,0.,0.,0.),("Term loan",3.,3.,3.,3.,288.,0.),("Senior notes",0.,0.,0.,0.,0.,250.),("Other debt",10.,10.,10.,10.,10.,0.)]
    for r,row in enumerate(balances,5):
        ws.cell(r,2,row[0])
        for c,v in enumerate(row[1:],3): input_cell(ws.cell(r,c,v),CUR)
    ws["B10"]="Total maturities"
    for c in range(3,9): ws.cell(10,c,f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}8)").number_format=CUR
    ws["B12"]="Cumulative maturities / gross debt"
    for c in range(3,9): ws.cell(12,c,f"=IFERROR(SUM($C$10:{get_column_letter(c)}10)/'Capital Structure'!$C$10,0)").number_format=PCT
    total_row(ws,10,2,8,CUR); set_widths(ws,{"A":4,"B":26,**{get_column_letter(c):13 for c in range(3,9)}})

    ws=wb["Refinancing"]; title(ws,"B2:G2","Refinancing Sources & Uses"); header(ws,4,2,["Uses","Amount","Sources","Amount","Pricing / note","Status"])
    uses=[("Refinance current debt","='Capital Structure'!C10"),("Minimum cash funding","=MAX(0,Assumptions!E17-Assumptions!E6)"),("Fees / OID","=C5*Assumptions!E16")]
    sources=[("New term debt","=SUM(C5:C7)"),("Existing cash","=MAX(0,Assumptions!E6-Assumptions!E17)"),("Equity / other","=MAX(0,C8-E5-E6)")]
    for r,(l,f) in enumerate(uses,5): ws.cell(r,2,l); ws.cell(r,3,f).number_format=CUR
    for r,(l,f) in enumerate(sources,5): ws.cell(r,4,l); ws.cell(r,5,f).number_format=CUR
    ws["B8"]="Total uses"; ws["C8"]="=SUM(C5:C7)"; ws["C8"].number_format=CUR
    ws["D8"]="Total sources"; ws["E8"]="=SUM(E5:E7)"; ws["E8"].number_format=CUR
    ws["F5"]="=Assumptions!E15+Assumptions!E16/7"; ws["F5"].number_format=PCT2
    ws["G8"]='=IF(ABS(C8-E8)<0.01,"PASS","FAIL")'; add_status_rules(ws,"G8")
    ws["B10"]="Pro forma gross leverage"; ws["C10"]="=IFERROR(E5/Assumptions!E5,0)"; ws["C10"].number_format=MULT
    ws["B11"]="Pro forma annual cash interest"; ws["C11"]="=E5*F5"; ws["C11"].number_format=CUR
    ws["B12"]="Pro forma interest coverage"; ws["C12"]="=IFERROR(Assumptions!E5/C11,0)"; ws["C12"].number_format=MULT
    set_widths(ws,{"A":4,"B":28,"C":16,"D":26,"E":16,"F":20,"G":14})

    ws=wb["Interest Rate Risk"]; title(ws,"B2:G2","Interest Rate & Spread Sensitivity"); header(ws,4,2,["Base-rate shock / spread shock",-100,0,100,200,300])
    for c in range(3,8): ws.cell(4,c).number_format=BPS
    for r,s in enumerate((-100,0,100,200,300),5):
        ws.cell(r,2,s).number_format=BPS
        for c in range(3,8):
            L=get_column_letter(c); ws.cell(r,c,f"='Capital Structure'!$C$5*MAX(0,Assumptions!$E$11+Assumptions!$E$12+($B{r}+{L}$4)/10000)+'Capital Structure'!$C$6*MAX(0,Assumptions!$E$11+Assumptions!$E$13+($B{r}+{L}$4)/10000)+'Capital Structure'!$E$7+'Capital Structure'!$E$8"); ws.cell(r,c).number_format=CUR
    ws.conditional_formatting.add("C5:G9",ColorScaleRule(start_type="min",start_color="E2F0D9",mid_type="percentile",mid_value=50,mid_color="FFF2CC",end_type="max",end_color="FCE4D6"))
    set_widths(ws,{"A":4,"B":32,**{get_column_letter(c):14 for c in range(3,8)}})

    ws=wb["Covenants"]; title(ws,"B2:E2","Covenant & Rating Screens"); header(ws,4,2,["Metric","Actual","Threshold","Headroom / status"])
    rows=[("Gross leverage","='Capital Structure'!C13","=Assumptions!E18","=D5-C5"),("Interest coverage","='Capital Structure'!C15","=Assumptions!E19","=C6-D6"),
          ("Near-term maturities / debt","=IFERROR(SUM('Maturity Ladder'!C10:E10)/'Capital Structure'!C10,0)",.35,"=D7-C7"),
          ("Refinancing status","=Refinancing!G8","PASS",'=IF(C8=D8,"PASS","FAIL")')]
    for r,row in enumerate(rows,5):
        for c,v in enumerate(row,2): ws.cell(r,c,v)
    for r in (5,6):
        for c in (3,4,5): ws.cell(r,c).number_format=MULT
    for c in (3,4,5): ws.cell(7,c).number_format=PCT
    add_status_rules(ws,"E8"); set_widths(ws,{"A":4,"B":32,"C":18,"D":18,"E":22})

    ws=wb["Recovery"]; title(ws,"B2:E2","Claim Recovery Waterfall"); header(ws,4,2,["Metric","Base","Downside","Note"])
    labels=["Stressed EBITDA","Recovery multiple","Enterprise value","EV haircut","Distributable value","Gross debt","Recovery value","Recovery rate","LGD"]
    for r,l in enumerate(labels,5): ws.cell(r,2,l)
    base=["=Assumptions!C5","=Assumptions!C20","=C5*C6","=C7*Assumptions!C21","=C7-C8","='Capital Structure'!C10","=MIN(C9,C10)","=IFERROR(C11/C10,0)","=1-C12"]
    down=["=Assumptions!D5","=Assumptions!D20","=D5*D6","=D7*Assumptions!D21","=D7-D8","='Capital Structure'!C10","=MIN(D9,D10)","=IFERROR(D11/D10,0)","=1-D12"]
    for r in range(5,14): ws.cell(r,3,base[r-5]); ws.cell(r,4,down[r-5])
    for r in (5,7,8,9,10,11): ws.cell(r,3).number_format=CUR; ws.cell(r,4).number_format=CUR
    for c in (3,4): ws.cell(6,c).number_format=MULT; ws.cell(12,c).number_format=PCT; ws.cell(13,c).number_format=PCT
    set_widths(ws,{"A":4,"B":32,"C":18,"D":18,"E":42})

    ws=wb["Checks"]; title(ws,"B2:C2","Model Checks"); header(ws,4,2,["Check","Status"])
    checks=[("Sources equal uses",'=IF(Refinancing!G8="PASS","PASS","FAIL")'),("Debt nonnegative",'=IF(\'Capital Structure\'!C10>=0,"PASS","FAIL")'),
            ("Maturities reconcile",'=IF(ABS(SUM(\'Maturity Ladder\'!C10:H10)-\'Capital Structure\'!C10)<0.01,"PASS","REVIEW")'),
            ("Recovery bounded",'=IF(AND(MIN(Recovery!C12:D12)>=0,MAX(Recovery!C12:D12)<=1),"PASS","FAIL")'),
            ("Overall",'=IF(COUNTIF(C5:C8,"FAIL")+COUNTIF(C5:C8,"REVIEW")=0,"PASS","REVIEW")')]
    for r,(l,f) in enumerate(checks,5): ws.cell(r,2,l); ws.cell(r,3,f)
    add_status_rules(ws,"C5:C9"); set_widths(ws,{"A":4,"B":42,"C":18})
    add_sources(wb,[("Debt instruments","Credit agreements, indentures, or filings","[date]","Balance, rate, maturity, collateral and covenants"),
                    ("Market pricing","Treasury / swap curve and comparable new issues","[date]","Freeze base-rate and spread assumptions"),
                    ("Financial statements","Audited statements or filing","[period]","EBITDA, cash and debt reconciliation"),
                    ("Recovery methodology","Documented enterprise-value and priority assumptions","[date]","Claim-level recovery bridge")])
    add_refresh_log(wb); finalize(wb,output)

if __name__=="__main__":
    p=argparse.ArgumentParser(); p.add_argument("--output",type=Path,default=Path("DEBT_FINANCE_template.xlsx")); a=p.parse_args(); build(a.output); print(f"saved {a.output}")
