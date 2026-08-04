"""Build the canonical multi-asset risk-management decision model."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "Positions", "Factors", "VaR & ES", "Stress",
    "Liquidity", "P&L Explain", "Limits", "Checks", "Sources", "RefreshLog",
]

FACTORS = ["Equity", "Rates", "Credit", "FX", "Volatility"]


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[PORTFOLIO] — Multi-Asset Risk Model", [
        ("Portfolio / desk:", "[Name]"),
        ("Risk horizon:", "1-day and stressed liquidation"),
        ("Last refreshed:", "[timestamp]"),
        ("Next limit review:", "[date]"),
        ("Refresh cadence:", "Intraday / daily"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Risk and Liquidity Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("VaR confidence", 0.99, 0.995, "%", PCT),
        ("Trading days per year", 252.0, 252.0, "days", "0"),
        ("Equity annual volatility", 0.22, 0.35, "%", PCT),
        ("Rates annual volatility", 0.012, 0.025, "yield change", PCT2),
        ("Credit annual volatility", 0.015, 0.035, "spread change", PCT2),
        ("FX annual volatility", 0.12, 0.20, "%", PCT),
        ("Volatility-factor annual volatility", 0.30, 0.50, "%", PCT),
        ("Maximum portfolio VaR", 12.0, 10.0, "$mm", CUR),
        ("Maximum stress loss", 40.0, 30.0, "$mm", CUR),
        ("Maximum illiquid NAV", 0.25, 0.20, "%", PCT),
        ("Maximum single-name concentration", 0.20, 0.15, "%", PCT),
        ("Maximum gross leverage", 4.0, 3.0, "x", MULT),
        ("Liquidity impact coefficient", 0.015, 0.030, "loss coefficient", PCT2),
        ("Available capital / NAV", 250.0, 250.0, "$mm", CUR),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 6, units)
    set_widths(sheet, {"A": 4, "B": 40, "C": 15, "D": 15, "E": 15, "F": 30})
    sheet.freeze_panes = "A5"

    sheet = workbook["Positions"]
    title(sheet, "B2:S2", "Position-Level Risk Inventory")
    headers = [
        "Instrument", "Asset class", "Market value", "Gross notional", "Delta",
        "Beta", "Duration", "Spread duration", "DV01", "Vega", "FX delta",
        "Days to liquidate", *FACTORS, "Weighted risk exposure",
    ]
    header(sheet, 4, 2, headers)
    positions = [
        ("Equity Index Future", "Equity", 45, 120, 1.00, 1.00, 0, 0, 0, 0, 0, 0.2, 1.0, 0, 0, 0, 0),
        ("Single Name A", "Equity", 35, 35, 1.00, 1.20, 0, 0, 0, 0, 0, 1.0, 1.2, 0, 0.1, 0, 0.1),
        ("Single Name Put", "Option", 8, 25, -0.35, -0.40, 0, 0, 0, 0.18, 0, 1.5, -0.4, 0, 0, 0, 0.8),
        ("10Y Treasury", "Rates", 50, 50, 1.00, 0, 8.2, 0, 0.041, 0, 0, 0.5, 0, -8.2, 0, 0, 0),
        ("IG Credit ETF", "Credit", 32, 50, 1.00, 0.35, 5.5, 5.0, 0.018, 0, 0, 1.0, 0.35, -2.0, -5.0, 0, 0),
        ("HY Bond", "Credit", 28, 40, 1.00, 0.55, 4.0, 3.8, 0.011, 0, 0, 4.0, 0.55, -1.0, -3.8, 0, 0),
        ("EURUSD Forward", "FX", 5, 80, 1.00, 0, 0, 0, 0, 0, 80, 0.3, 0, 0, 0, 1.0, 0),
        ("Commodity Future", "Commodity", 20, 65, 1.00, 0.25, 0, 0, 0, 0, 0, 0.5, 0.25, 0, 0.2, 0.1, 0.1),
        ("Variance Swap", "Volatility", 6, 30, 0, 0, 0, 0, 0, 0.45, 0, 2.0, 0, 0, 0, 0, 1.0),
        ("Private Credit Position", "Private", 40, 40, 1.00, 0.20, 2.0, 2.5, 0.008, 0, 0, 20.0, 0.20, -0.5, -2.5, 0, 0),
        ("Cash", "Cash", 25, 25, 1.00, 0, 0, 0, 0, 0, 0, 0.0, 0, 0, 0, 0, 0),
        ("Short Equity Basket", "Equity", -18, 60, -1.00, -0.90, 0, 0, 0, 0, 0, 0.5, -0.9, 0, -0.1, 0, 0),
    ]
    for row, values in enumerate(positions, start=5):
        for column, value in enumerate(values, start=2):
            cell = sheet.cell(row, column, value)
            if column >= 4:
                input_cell(cell, PCT2 if column in (6, 7, 14, 15, 16, 17, 18) else (CUR if column in (4, 5, 10, 11, 12) else "0.00"))
        sheet.cell(row, 19, f"=SUMPRODUCT(C{row}:Q{row},C{row}:Q{row})^0.5")
        sheet.cell(row, 19).number_format = CUR
    total_row(sheet, 18, 2, 19)
    sheet["B18"] = "Portfolio total"
    for column in (4, 5, 10, 11, 12, 19):
        letter = get_column_letter(column)
        sheet.cell(18, column, f"=SUM({letter}5:{letter}16)")
        sheet.cell(18, column).number_format = CUR
    set_widths(sheet, {"A": 4, "B": 28, "C": 16, **{get_column_letter(column): 13 for column in range(4, 20)}})
    sheet.freeze_panes = "D5"

    sheet = workbook["Factors"]
    title(sheet, "B2:H2", "Factor Exposure and Covariance")
    header(sheet, 4, 2, ["Factor", "Exposure", "Annual vol", *FACTORS])
    correlations = [
        [1.00, -0.20, 0.35, 0.10, 0.45],
        [-0.20, 1.00, -0.25, 0.05, -0.10],
        [0.35, -0.25, 1.00, 0.10, 0.30],
        [0.10, 0.05, 0.10, 1.00, 0.10],
        [0.45, -0.10, 0.30, 0.10, 1.00],
    ]
    vol_rows = [7, 8, 9, 10, 11]
    exposure_columns = [14, 15, 16, 17, 18]
    for index, factor in enumerate(FACTORS):
        row = 5 + index
        sheet.cell(row, 2, factor)
        source_column = get_column_letter(exposure_columns[index])
        sheet.cell(row, 3, f"=SUMPRODUCT(Positions!$D$5:$D$16,Positions!${source_column}$5:${source_column}$16)")
        sheet.cell(row, 3).number_format = CUR
        sheet.cell(row, 4, f"=Assumptions!E{vol_rows[index]}")
        sheet.cell(row, 4).number_format = PCT2
        for offset, correlation in enumerate(correlations[index], start=5):
            input_cell(sheet.cell(row, offset, correlation), "0.00")
    sheet["B12"] = "Annual portfolio variance"
    terms = []
    for i in range(5):
        for j in range(5):
            exposure_i = f"$C${5+i}"
            exposure_j = f"$C${5+j}"
            vol_i = f"$D${5+i}"
            vol_j = f"$D${5+j}"
            corr = f"{get_column_letter(5+j)}{5+i}"
            terms.append(f"{exposure_i}*{exposure_j}*{vol_i}*{vol_j}*{corr}")
    sheet["C12"] = "=" + "+".join(terms)
    sheet["C12"].number_format = CUR
    sheet["B13"] = "Annual portfolio volatility"
    sheet["C13"] = "=SQRT(MAX(0,C12))"
    sheet["C13"].number_format = CUR
    set_widths(sheet, {"A": 4, "B": 20, "C": 18, "D": 16, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12})

    sheet = workbook["VaR & ES"]
    title(sheet, "B2:E2", "Parametric VaR and Expected Shortfall")
    header(sheet, 4, 2, ["Metric", "Value", "Units", "Method"])
    metrics = [
        ("Annual portfolio volatility", "=Factors!C13", "$mm", "factor covariance", CUR),
        ("Daily portfolio volatility", "=C5/SQRT(Assumptions!E6)", "$mm", "annual / sqrt(days)", CUR),
        ("Normal quantile", "=-NORM.S.INV(1-Assumptions!E5)", "z", "confidence quantile", "0.000"),
        ("One-day VaR", "=C6*C7", "$mm", "sigma x z", CUR),
        ("One-day expected shortfall", "=C6*NORM.S.DIST(C7,FALSE)/(1-Assumptions!E5)", "$mm", "normal tail expectation", CUR),
        ("VaR / NAV", "=C8/Assumptions!E18", "%", "capital consumption", PCT2),
        ("Diversification benefit", "=1-C5/SUMPRODUCT(ABS(Factors!C5:C9),Factors!D5:D9)", "%", "1 - portfolio / standalone", PCT2),
    ]
    for row, (label, formula, units, note, number_format) in enumerate(metrics, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
        sheet.cell(row, 3).number_format = number_format
        sheet.cell(row, 4, units)
        sheet.cell(row, 5, note)
    set_widths(sheet, {"A": 4, "B": 30, "C": 18, "D": 14, "E": 42})

    sheet = workbook["Stress"]
    title(sheet, "B2:I2", "Portfolio Stress Matrix")
    header(sheet, 4, 2, ["Scenario", "Equity", "Rates", "Credit", "FX", "Vol", "Liquidity", "P&L"])
    scenarios = [
        ("Global risk-off", -0.18, -0.015, -0.025, -0.08, 0.40, 0.50),
        ("Inflation / rates shock", -0.08, 0.025, 0.012, 0.05, 0.20, 0.70),
        ("Credit crisis", -0.12, -0.010, -0.050, -0.04, 0.35, 0.35),
        ("Dollar squeeze", -0.07, 0.010, -0.015, 0.12, 0.25, 0.45),
        ("Volatility normalization", 0.05, 0.005, 0.005, 0.00, -0.30, 1.00),
        ("Idiosyncratic concentration", -0.25, 0.000, -0.010, 0.00, 0.20, 0.30),
    ]
    for row, scenario in enumerate(scenarios, start=5):
        sheet.cell(row, 2, scenario[0])
        for column, shock in enumerate(scenario[1:], start=3):
            input_cell(sheet.cell(row, column, shock), PCT2 if column < 8 else MULT)
        sheet.cell(row, 9, f"=SUMPRODUCT(Factors!$C$5:$C$9,C{row}:G{row})-SUMPRODUCT(ABS(Positions!$D$5:$D$16),MAX(0,Positions!$M$5:$M$16-1))*Assumptions!$E$17*(1-H{row})")
        sheet.cell(row, 9).number_format = CUR
    set_widths(sheet, {"A": 4, "B": 30, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13, "H": 13, "I": 16})

    sheet = workbook["Liquidity"]
    title(sheet, "B2:H2", "Liquidity and Liquidation Cost")
    header(sheet, 4, 2, ["Instrument", "Gross notional", "Days", "ADV participation proxy", "Impact loss", "Liquidation loss", "Illiquid?"])
    for source_row in range(5, 17):
        row = source_row
        sheet.cell(row, 2, f"=Positions!B{source_row}")
        sheet.cell(row, 3, f"=ABS(Positions!E{source_row})")
        sheet.cell(row, 3).number_format = CUR
        sheet.cell(row, 4, f"=Positions!M{source_row}")
        sheet.cell(row, 5, f"=MIN(1,C{row}/MAX(1,Assumptions!$E$18)/MAX(1,D{row}))")
        sheet.cell(row, 5).number_format = PCT2
        sheet.cell(row, 6, f"=Assumptions!$E$17*E{row}^1.5")
        sheet.cell(row, 6).number_format = PCT2
        sheet.cell(row, 7, f"=C{row}*F{row}")
        sheet.cell(row, 7).number_format = CUR
        sheet.cell(row, 8, f'=IF(D{row}>5,"YES","NO")')
    sheet["B19"] = "Illiquid gross notional"
    sheet["C19"] = '=SUMIF(H5:H16,"YES",C5:C16)'
    sheet["C19"].number_format = CUR
    sheet["B20"] = "Illiquid NAV ratio"
    sheet["C20"] = "=C19/Assumptions!E18"
    sheet["C20"].number_format = PCT2
    sheet["B21"] = "Total modeled liquidation loss"
    sheet["C21"] = "=SUM(G5:G16)"
    sheet["C21"].number_format = CUR
    set_widths(sheet, {"A": 4, "B": 28, "C": 16, "D": 12, "E": 20, "F": 14, "G": 16, "H": 12})

    sheet = workbook["P&L Explain"]
    title(sheet, "B2:H2", "Daily P&L Explain")
    header(sheet, 4, 2, ["Driver", "Exposure", "Move", "First-order P&L", "Convexity / residual", "Explained P&L", "Notes"])
    rows = [
        ("Equity", "=Factors!C5", -0.012, 0.5, "delta/beta"),
        ("Rates", "=Factors!C6", 0.0015, 0.0, "duration"),
        ("Credit", "=Factors!C7", 0.0020, 0.0, "spread duration"),
        ("FX", "=Factors!C8", -0.006, 0.0, "currency delta"),
        ("Volatility", "=Factors!C9", 0.025, 0.8, "vega / convexity"),
    ]
    for row, (driver, exposure, move, residual, note) in enumerate(rows, start=5):
        sheet.cell(row, 2, driver)
        sheet.cell(row, 3, exposure)
        sheet.cell(row, 3).number_format = CUR
        input_cell(sheet.cell(row, 4, move), PCT2)
        sheet.cell(row, 5, f"=C{row}*D{row}")
        sheet.cell(row, 5).number_format = CUR
        input_cell(sheet.cell(row, 6, residual), CUR)
        sheet.cell(row, 7, f"=E{row}+F{row}")
        sheet.cell(row, 7).number_format = CUR
        sheet.cell(row, 8, note)
    sheet["B12"] = "Explained P&L"
    sheet["G12"] = "=SUM(G5:G9)"
    sheet["G12"].number_format = CUR
    sheet["B13"] = "Actual P&L"
    input_cell(sheet["G13"], CUR)
    sheet["G13"] = -1.5
    sheet["B14"] = "Unexplained P&L"
    sheet["G14"] = "=G13-G12"
    sheet["G14"].number_format = CUR
    set_widths(sheet, {"A": 4, "B": 18, "C": 16, "D": 14, "E": 18, "F": 18, "G": 18, "H": 32})

    sheet = workbook["Limits"]
    title(sheet, "B2:F2", "Risk Limit Dashboard")
    header(sheet, 4, 2, ["Limit", "Actual", "Threshold", "Headroom", "Status"])
    limits = [
        ("One-day VaR", "='VaR & ES'!C8", "=Assumptions!E12", "maximum", CUR),
        ("Worst stress loss", "=-MIN(Stress!I5:I10)", "=Assumptions!E13", "maximum", CUR),
        ("Illiquid NAV", "=Liquidity!C20", "=Assumptions!E14", "maximum", PCT2),
        ("Single-name concentration", "=MAX(ABS(Positions!D5:D16))/Assumptions!E18", "=Assumptions!E15", "maximum", PCT2),
        ("Gross leverage", "=SUM(ABS(Positions!E5:E16))/Assumptions!E18", "=Assumptions!E16", "maximum", MULT),
        ("Unexplained P&L", "=ABS('P&L Explain'!G14)", 5.0, "maximum", CUR),
    ]
    for row, (label, actual, threshold, direction, number_format) in enumerate(limits, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, actual)
        sheet.cell(row, 4, threshold)
        if not isinstance(threshold, str):
            input_cell(sheet.cell(row, 4), number_format)
        sheet.cell(row, 3).number_format = number_format
        sheet.cell(row, 4).number_format = number_format
        sheet.cell(row, 5, f"=D{row}-C{row}")
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 6, f'=IF(C{row}<=D{row},"PASS","BREACH")')
    add_status_rules(sheet, "F5:F10")
    set_widths(sheet, {"A": 4, "B": 30, "C": 18, "D": 18, "E": 18, "F": 14})

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "Risk Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("Correlation diagonal equals one", '=IF(AND(Factors!E5=1,Factors!F6=1,Factors!G7=1,Factors!H8=1,Factors!I9=1),"PASS","FAIL")'),
        ("Correlation matrix bounded", '=IF(AND(MIN(Factors!E5:I9)>=-1,MAX(Factors!E5:I9)<=1),"PASS","FAIL")'),
        ("Portfolio variance nonnegative", '=IF(Factors!C12>=0,"PASS","FAIL")'),
        ("VaR and ES positive", '=IF(AND(\'VaR & ES\'!C8>=0,\'VaR & ES\'!C9>=\'VaR & ES\'!C8),"PASS","FAIL")'),
        ("No risk limit breaches", '=IF(COUNTIF(Limits!F5:F10,"BREACH")=0,"PASS","REVIEW")'),
        ("Liquidity loss nonnegative", '=IF(Liquidity!C21>=0,"PASS","FAIL")'),
        ("Position inventory populated", '=IF(COUNTA(Positions!B5:B16)=12,"PASS","FAIL")'),
        ("P&L explain complete", '=IF(COUNT(\'P&L Explain\'!G5:G9)=5,"PASS","FAIL")'),
        ("Overall", '=IF(COUNTIF(C5:C12,"FAIL")=0,"PASS","FAIL")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    add_status_rules(sheet, "C5:C13")
    set_widths(sheet, {"A": 4, "B": 44, "C": 18})

    add_sources(workbook, [
        ("Position and market-value feed", "[OMS / risk source]", "[timestamp]", "Reconcile to books and records"),
        ("Factor and covariance calibration", "[model dataset URL]", "[window]", "Document decay, winsorization, and regime treatment"),
        ("Curve, spread, volatility, and FX data", "[market data URL]", "[timestamp]", "Record interpolation and stale-data policy"),
        ("Liquidity and ADV data", "[venue / vendor source]", "[period]", "Document liquidation horizon and impact calibration"),
        ("Risk limit policy", "[approved policy URL]", "[effective date]", "Thresholds, escalation, exceptions, and ownership"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("RISK_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
