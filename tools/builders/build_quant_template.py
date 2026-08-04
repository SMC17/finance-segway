import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[STRATEGY] — Quantitative / Systematic Model", [
    ("Strategy:", "[fill in]"),
    ("Asset class / universe:", "[fill in]"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly (daily if live-traded)"),
])

# ---------------- RETURNS & SHARPE ----------------
ws = wb.create_sheet("Returns & Sharpe")
set_col_widths(ws, [4, 30, 16, 40, 14, 14, 14, 14])
ws["B2"] = "Risk-Adjusted Return Statistics"; ws["B2"].font = TITLE
ws["B4"] = "Monthly return series (enter below, extend as needed)"
ws["B4"].font = ITALIC_GRAY
ws["B5"] = "Month"; ws["C5"] = "Return %"
ws["E5"] = "Benchmark %"; ws["F5"] = "Wealth idx"; ws["G5"] = "Peak"; ws["H5"] = "Drawdown"
style_header_row(ws, 5, 2, start_col=2)
style_header_row(ws, 5, 4, start_col=5)
r = 6
for m in range(1, 25):
    ws.cell(row=r, column=2, value=f"M{m}")
    c = ws.cell(row=r, column=3, value=0.0)
    c.font = BLUE; c.number_format = PCT2; c.border = BORDER
    e = ws.cell(row=r, column=5, value=0.0)
    e.font = BLUE; e.number_format = PCT2; e.border = BORDER

    f = ws.cell(row=r, column=6,
                value=f"=1*(1+C{r})" if r == 6 else f"=F{r-1}*(1+C{r})")
    f.number_format = '0.0000'; f.border = BORDER
    g = ws.cell(row=r, column=7, value=f"=MAX($F$6:F{r})")
    g.number_format = '0.0000'; g.border = BORDER
    h = ws.cell(row=r, column=8, value=f"=IFERROR(F{r}/G{r}-1,\"-\")")
    h.number_format = PCT; h.border = BORDER
    r += 1
last_data_row = r - 1

ws["B32"] = "Risk-free rate (annual, %)"; ws["C32"] = 0.045; ws["C32"].font = BLUE
ws["C32"].fill = YELLOW_FILL; ws["C32"].number_format = PCT

ws["B34"] = "Outputs"; ws["B34"].font = BOLD; ws["B34"].fill = GRAY_FILL
ws["B35"] = "Avg monthly return"
ws["C35"] = f"=AVERAGE(C6:C{last_data_row})"; ws["C35"].number_format = PCT2
ws["B36"] = "Monthly std dev"
ws["C36"] = f"=STDEV(C6:C{last_data_row})"; ws["C36"].number_format = PCT2
ws["B37"] = "Annualized return"
ws["C37"] = "=(1+C35)^12-1"; ws["C37"].number_format = PCT
ws["B38"] = "Annualized volatility"
ws["C38"] = "=C36*SQRT(12)"; ws["C38"].number_format = PCT
ws["B39"] = "Sharpe ratio"
ws["C39"] = "=IFERROR((C37-C32)/C38,\"-\")"; ws["C39"].font = BOLD; ws["C39"].number_format = '0.00'
ws["B40"] = "Downside deviation (returns < 0 only)"
ws["C40"] = f'=IFERROR(SQRT(SUMPRODUCT((C6:C{last_data_row}<0)*(C6:C{last_data_row})^2)/COUNTIF(C6:C{last_data_row},"<0"))*SQRT(12),"-")'
ws["C40"].number_format = PCT
ws["B41"] = "Sortino ratio"
ws["C41"] = "=IFERROR((C37-C32)/C40,\"-\")"; ws["C41"].font = BOLD; ws["C41"].number_format = '0.00'
ws["B42"] = "Max drawdown"
ws["C42"] = f"=MIN(H6:H{last_data_row})"
ws["C42"].font = BOLD; ws["C42"].number_format = PCT
ws["D42"] = "Trough of the wealth-index drawdown series (columns F-H) — worst peak-to-trough decline in the sample"
ws["D42"].font = ITALIC_GRAY
for r2 in range(35, 43):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- BENCHMARK & FACTOR EXPOSURE ----------------
ws2 = wb["Returns & Sharpe"]
ws2["B44"] = "Benchmark & Factor Exposure (single-factor / CAPM)"; ws2["B44"].font = BOLD; ws2["B44"].fill = GRAY_FILL
ws2["B45"] = "Avg monthly benchmark return"
ws2["C45"] = f"=AVERAGE(E6:E{last_data_row})"; ws2["C45"].number_format = PCT2
ws2["B46"] = "Annualized benchmark return"
ws2["C46"] = "=(1+C45)^12-1"; ws2["C46"].number_format = PCT
ws2["B47"] = "Beta (vs benchmark)"
ws2["C47"] = f"=IFERROR(SLOPE(C6:C{last_data_row},E6:E{last_data_row}),\"-\")"
ws2["C47"].font = BOLD; ws2["C47"].number_format = '0.00'
ws2["B48"] = "R-squared"
ws2["C48"] = f"=IFERROR(RSQ(C6:C{last_data_row},E6:E{last_data_row}),\"-\")"; ws2["C48"].number_format = PCT
ws2["B49"] = "Jensen's alpha (annualized) = Rp - [Rf + Beta x (Rm - Rf)]"
ws2["C49"] = "=IFERROR(C37-(C32+C47*(C46-C32)),\"-\")"; ws2["C49"].font = BOLD; ws2["C49"].number_format = PCT
ws2["D49"] = "Positive = generating return beyond what beta exposure to the benchmark would explain"
ws2["D49"].font = ITALIC_GRAY
for r2 in range(45, 50):
    ws2.cell(row=r2, column=3).border = BORDER

# ---------------- POSITION SIZING ----------------
ws = wb.create_sheet("Position Sizing")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Position Sizing (Kelly & Fixed Fractional)"; ws["B2"].font = TITLE
inputs = [
    ("Win rate (%)", 0.55, PCT),
    ("Avg win / avg loss ratio (payoff ratio)", 1.5, '0.00'),
    ("Account equity ($)", 0, CUR),
    ("Max risk per trade (fixed-fractional, %)", 0.01, PCT),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B10"] = "Kelly fraction = W - (1-W)/R"
ws["C10"] = "=C5-(1-C5)/C6"; ws["C10"].font = BOLD; ws["C10"].number_format = PCT
ws["D10"] = "Full Kelly is aggressive — most practitioners use 1/4 to 1/2 Kelly"
ws["D10"].font = ITALIC_GRAY
ws["B11"] = "Half-Kelly position size ($)"
ws["C11"] = "=IFERROR(C7*C10/2,\"-\")"; ws["C11"].number_format = CUR
ws["B12"] = "Fixed-fractional position size ($)"
ws["C12"] = "=C7*C8"; ws["C12"].number_format = CUR
for r2 in (10, 11, 12):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "QUANT_template.xlsx"
wb.save(out_path)
print("saved", out_path)
