import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[MFI] — Microfinance Model", [
    ("Institution:", "[fill in]"),
    ("Region:", "[fill in]"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly"),
])

# ---------------- LOAN PORTFOLIO ----------------
ws = wb.create_sheet("Loan Portfolio")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Loan Portfolio Overview"; ws["B2"].font = TITLE
inputs = [
    ("Gross loan portfolio (GLP, $)", 0, CUR),
    ("Number of active borrowers", 0, NUM),
    ("Portfolio at risk >30 days ($)", 0, CUR),
    ("Portfolio at risk >90 days ($)", 0, CUR),
    ("Write-offs this period ($)", 0, CUR),
    ("Average loan balance ($)", 0, CUR),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B12"] = "Portfolio Quality"; ws["B12"].font = BOLD; ws["B12"].fill = GRAY_FILL
ws["B13"] = "PAR 30 %"; ws["C13"] = "=IFERROR(C7/C5,\"-\")"; ws["C13"].number_format = PCT
ws["B14"] = "PAR 90 %"; ws["C14"] = "=IFERROR(C8/C5,\"-\")"; ws["C14"].number_format = PCT
ws["B15"] = "Write-off ratio (annualized)"; ws["C15"] = "=IFERROR(C9/C5,\"-\")"; ws["C15"].number_format = PCT
ws["B16"] = "Avg loan / borrower check"; ws["C16"] = "=IFERROR(C5/C6,\"-\")"; ws["C16"].number_format = CUR
for r2 in range(13, 17):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- SUSTAINABILITY RATIOS ----------------
ws = wb.create_sheet("Sustainability")
set_col_widths(ws, [4, 34, 16, 40])
ws["B2"] = "Operational & Financial Self-Sufficiency (OSS/FSS)"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Financial revenue (interest + fees, $)", 0, CUR),
    ("Financial expense (cost of funds, $)", 0, CUR),
    ("Loan loss provision expense ($)", 0, CUR),
    ("Operating expense ($)", 0, CUR),
    ("Cost of capital at market rate (imputed, $)", 0, CUR),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B11"] = "Outputs"; ws["B11"].font = BOLD; ws["B11"].fill = GRAY_FILL
ws["B12"] = "Operational Self-Sufficiency (OSS) = Rev / (Fin exp + Loss prov + Opex)"
ws["C12"] = "=IFERROR(C5/(C6+C7+C8),\"-\")"; ws["C12"].font = BOLD; ws["C12"].number_format = PCT
ws["D12"] = ">100% = covering costs from operations, not subsidy-dependent"
ws["D12"].font = ITALIC_GRAY
ws["B13"] = "Financial Self-Sufficiency (FSS) = Rev / (Fin exp + Loss prov + Opex + imputed cost of capital)"
ws["C13"] = "=IFERROR(C5/(C6+C7+C8+C9),\"-\")"; ws["C13"].font = BOLD; ws["C13"].number_format = PCT
ws["D13"] = "Stricter than OSS — adjusts for subsidized funding cost vs. commercial rate"
ws["D13"].font = ITALIC_GRAY
ws["B14"] = "Portfolio yield (Rev / avg GLP)"
ws["C14"] = "=IFERROR(C5/'Loan Portfolio'!C5,\"-\")"; ws["C14"].number_format = PCT
for r2 in range(12, 15):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- LOAN LOSS PROVISIONING ----------------
ws = wb.create_sheet("Provisioning")
set_col_widths(ws, [4, 26, 16, 14, 16, 40])
ws["B2"] = "Loan Loss Reserve Adequacy (days-past-due tiering)"; ws["B2"].font = TITLE
for i, h in enumerate(["", "Aging bucket", "Outstanding balance ($)", "Reserve rate",
                        "Required reserve ($)", ""], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 4, start_col=2)

tiers = [
    ("Current (0-30 days)", 0.01),
    ("31-90 days", 0.10),
    ("91-180 days", 0.50),
    ("180+ days / written off pending", 1.00),
]
r = 5
for label, rate in tiers:
    ws.cell(row=r, column=2, value=label).font = BLACK
    bal = ws.cell(row=r, column=3, value=0); bal.font = BLUE; bal.fill = YELLOW_FILL
    bal.number_format = CUR; bal.border = BORDER
    rt = ws.cell(row=r, column=4, value=rate); rt.font = BLUE; rt.number_format = PCT; rt.border = BORDER
    req = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); req.number_format = CUR; req.border = BORDER
    r += 1

ws["B9"] = "Total outstanding portfolio"; ws["B9"].font = BOLD
ws["C9"] = "=SUM(C5:C8)"; ws["C9"].font = BOLD; ws["C9"].number_format = CUR; ws["C9"].border = BORDER
ws["B10"] = "Total required reserve (sum of tiers)"; ws["B10"].font = BOLD
ws["E10"] = "=SUM(E5:E8)"; ws["E10"].font = BOLD; ws["E10"].number_format = CUR; ws["E10"].border = BORDER

ws["B12"] = "Actual reserve held ($)"
c = ws.cell(row=12, column=3, value=0); c.font = BLUE; c.fill = YELLOW_FILL; c.number_format = CUR; c.border = BORDER
ws["B13"] = "Reserve adequacy (actual / required)"
ws["C13"] = "=IFERROR(C12/E10,\"-\")"; ws["C13"].font = BOLD; ws["C13"].number_format = PCT
ws["C13"].fill = YELLOW_FILL; ws["C13"].border = BORDER
ws["D13"] = "<100% = under-reserved relative to portfolio risk profile — review provisioning policy"
ws["D13"].font = ITALIC_GRAY
ws["B14"] = "Cross-check vs Loan Portfolio tab (should tie to GLP)"
ws["C14"] = "=IFERROR(C9-'Loan Portfolio'!C5,\"-\")"; ws["C14"].font = GREEN; ws["C14"].number_format = CUR
ws["C14"].border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- GROWTH & PRODUCTIVITY ----------------
ws = wb.create_sheet("Growth & Productivity")
set_col_widths(ws, [4, 34, 16, 40])
ws["B2"] = "Growth & Loan Officer Productivity"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs = [
    ("Active borrowers — prior period", 0, NUM),
    ("Active borrowers — current period", 0, NUM),
    ("New disbursements this period ($)", 0, CUR),
    ("Number of loan officers", 0, NUM),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B10"] = "Outputs"; ws["B10"].font = BOLD; ws["B10"].fill = GRAY_FILL
ws["B11"] = "Borrower growth rate (period over period)"
ws["C11"] = "=IFERROR((C6-C5)/C5,\"-\")"; ws["C11"].number_format = PCT
ws["B12"] = "Borrowers per loan officer (caseload)"
ws["C12"] = "=IFERROR(C6/C8,\"-\")"; ws["C12"].number_format = NUM
ws["D12"] = "Typical sustainable caseload benchmark: 250-400 borrowers/officer, group-lending-dependent"
ws["D12"].font = ITALIC_GRAY
ws["B13"] = "Disbursement per active borrower ($, avg)"
ws["C13"] = "=IFERROR(C7/C6,\"-\")"; ws["C13"].number_format = CUR
for r2 in (11, 12, 13):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "MICROFINANCE_template.xlsx"
wb.save(out_path)
print("saved", out_path)
