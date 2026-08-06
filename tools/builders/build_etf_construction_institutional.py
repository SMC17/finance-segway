"""Build the canonical ETF construction, creation/redemption, and tracking-cost model."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from institutional_helpers import (  # noqa: E402
    CUR, CUR0, CUR2, MULT, PCT, PCT2,
    add_cover, add_refresh_log, add_sources, add_status_rules, finalize,
    header, input_cell, set_widths, title, total_row,
)

SHEETS = [
    "Cover", "Assumptions", "Portfolio Construction", "Creation & Redemption",
    "Tracking Error & Costs", "Checks", "Sources", "RefreshLog",
]

# Illustrative default holdings (template defaults) -- a real instance
# manifest overrides these with a real ETF's disclosed constituents.
DEFAULT_HOLDINGS = [
    ("[Holding 1]", "[Issuer 1]", 0.08),
    ("[Holding 2]", "[Issuer 2]", 0.07),
    ("[Holding 3]", "[Issuer 3]", 0.05),
    ("[Holding 4]", "[Issuer 4]", 0.04),
    ("[Holding 5]", "[Issuer 5]", 0.04),
    ("[Holding 6]", "[Issuer 6]", 0.03),
    ("[Holding 7]", "[Issuer 7]", 0.03),
    ("[Holding 8]", "[Issuer 8]", 0.03),
    ("[Holding 9]", "[Issuer 9]", 0.02),
    ("[Holding 10]", "[Issuer 10]", 0.02),
]
DEFAULT_SECTORS = [
    ("[Sector 1]", 0.30),
    ("[Sector 2]", 0.20),
    ("[Sector 3]", 0.15),
    ("[Sector 4]", 0.10),
    ("[Sector 5]", 0.08),
]
HOLDING_ROWS = 30
SECTOR_ROWS = 11


def build(output: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEETS:
        workbook.create_sheet(name)

    add_cover(workbook, "[ETF] — Construction, Creation/Redemption, and Tracking-Cost Model", [
        ("ETF / share class:", "[Name / ticker]"),
        ("Benchmark index:", "[Index]"),
        ("As of date:", "[date]"),
        ("Last refreshed:", "[date]"),
        ("Next NAV / holdings refresh:", "[date]"),
        ("Active scenario:", "Base"),
        ("Units:", "$ in millions unless noted"),
    ])

    sheet = workbook["Assumptions"]
    title(sheet, "B2:F2", "Fund, Replication, and Creation-Unit Assumptions")
    header(sheet, 4, 2, ["Assumption", "Base", "Downside", "Active", "Units / note"])
    assumptions = [
        ("Fund net assets (AUM)", 10000.0, 10000.0, "$mm", CUR),
        ("Net expense ratio", 0.0050, 0.0050, "% p.a.", PCT2),
        ("Trailing 12-month dividend yield", 0.0150, 0.0150, "% p.a.", PCT2),
        ("Market price (last close)", 100.00, 100.00, "$/share", CUR2),
        ("Standard creation unit size", 50000.0, 50000.0, "shares", "#,##0"),
        ("Current premium / (discount) to NAV", 0.0000, -0.0025, "%", PCT2),
        ("Estimated securities-lending revenue offset", 0.0002, 0.0001, "% of NAV p.a.", PCT2),
        ("Estimated cash-drag cost", 0.0002, 0.0008, "% of NAV p.a.", PCT2),
        ("Estimated sampling / optimization tracking error", 0.0010, 0.0050, "% of NAV p.a.", PCT2),
        ("AP arbitrage action threshold", 0.0010, 0.0025, "% premium/discount", PCT2),
    ]
    for row, (label, base, downside, units, number_format) in enumerate(assumptions, start=5):
        sheet.cell(row, 2, label)
        input_cell(sheet.cell(row, 3, base), number_format)
        input_cell(sheet.cell(row, 4, downside), number_format)
        sheet.cell(row, 5, f'=IF(Cover!$C$9="Downside",D{row},C{row})')
        sheet.cell(row, 5).number_format = number_format
        sheet.cell(row, 6, units)
    sheet["B16"] = "Implied shares outstanding"
    sheet["C16"] = "=E5/E8"
    sheet["C16"].number_format = '#,##0.0'
    sheet["D16"] = "mm shares -- derived: AUM / market price"
    set_widths(sheet, {"A": 4, "B": 48, "C": 15, "D": 15, "E": 15, "F": 34})
    sheet.freeze_panes = "A5"

    sheet = workbook["Portfolio Construction"]
    title(sheet, "B2:E2", "Constituent Holdings (Look-Through)")
    header(sheet, 4, 2, ["Symbol", "Issuer / description", "Weight"])
    for offset, (symbol, issuer, weight) in enumerate(DEFAULT_HOLDINGS):
        row = 5 + offset
        input_cell(sheet.cell(row, 2, symbol))
        input_cell(sheet.cell(row, 3, issuer))
        input_cell(sheet.cell(row, 4, weight), PCT2)
    other_row = 5 + HOLDING_ROWS
    sheet.cell(other_row, 2, "All other constituents")
    sheet.cell(other_row, 3, "[remaining positions]")
    sheet.cell(other_row, 4, f"=MAX(0,1-SUM(D5:D{other_row - 1}))")
    sheet.cell(other_row, 4).number_format = PCT2
    total_row(sheet, other_row + 1, 2, 4, PCT2)
    sheet.cell(other_row + 1, 4, f"=SUM(D5:D{other_row})")
    sheet.cell(other_row + 1, 4).number_format = PCT2

    sector_start = other_row + 4
    title(sheet, f"B{sector_start - 1}:C{sector_start - 1}", "Sector Allocation (as disclosed)")
    header(sheet, sector_start, 2, ["Sector", "Weight"])
    for offset, (sector, weight) in enumerate(DEFAULT_SECTORS):
        row = sector_start + 1 + offset
        input_cell(sheet.cell(row, 2, sector))
        input_cell(sheet.cell(row, 3, weight), PCT2)
    sector_total_row = sector_start + SECTOR_ROWS + 1
    sheet.cell(sector_total_row, 2, "Total disclosed sector weight")
    sheet.cell(sector_total_row, 3, f"=SUM(C{sector_start + 1}:C{sector_start + SECTOR_ROWS})")
    sheet.cell(sector_total_row, 3).number_format = PCT2
    set_widths(sheet, {"A": 4, "B": 30, "C": 34, "D": 14, "E": 4})
    sheet.freeze_panes = "C5"

    sheet = workbook["Creation & Redemption"]
    title(sheet, "B2:C2", "Creation Unit Economics and Authorized-Participant Arbitrage")
    header(sheet, 4, 2, ["Metric", "Value"])
    rows = [
        ("NAV per share (proxy: market price / (1+premium))", "=Assumptions!$E$8/(1+Assumptions!$E$10)", CUR2),
        ("Creation unit basket value (NAV basis)", "=C5*Assumptions!$E$9", CUR0),
        ("Creation unit basket value (market-price basis)", "=Assumptions!$E$8*Assumptions!$E$9", CUR0),
        ("Creation unit implied share of total AUM", "=C6/(Assumptions!$E$5*1000000)", PCT2),
        ("Premium / (discount) at market price vs. NAV", "=Assumptions!$E$10", PCT2),
        ("AP arbitrage profit per creation unit at current premium/discount", "=C9*C7", CUR0),
        ("AP action threshold breached?", '=IF(ABS(C9)>Assumptions!$E$14,"YES -- AP arbitrage expected to compress premium/discount","NO -- within normal trading band")', None),
    ]
    for offset, (label, formula, number_format) in enumerate(rows):
        row = 5 + offset
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
        if number_format:
            sheet.cell(row, 3).number_format = number_format
    set_widths(sheet, {"A": 4, "B": 56, "C": 22})
    sheet.freeze_panes = "A5"

    sheet = workbook["Tracking Error & Costs"]
    title(sheet, "B2:C2", "Fund Return vs. Benchmark: Tracking-Difference Bridge")
    header(sheet, 4, 2, ["Component", "Annualized impact"])
    rows = [
        ("Expense ratio drag", "=-Assumptions!$E$6", PCT2),
        ("Securities-lending revenue offset", "=Assumptions!$E$11", PCT2),
        ("Cash-drag cost", "=-Assumptions!$E$12", PCT2),
        ("Sampling / optimization tracking error (+/-)", "=-Assumptions!$E$13", PCT2),
    ]
    for offset, (label, formula, number_format) in enumerate(rows):
        row = 5 + offset
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
        sheet.cell(row, 3).number_format = number_format
    sheet["B9"] = "Estimated net tracking difference (fund return - benchmark return)"
    sheet["C9"] = "=SUM(C5:C8)"
    sheet["C9"].number_format = PCT2
    sheet["B11"] = "Reference: trailing dividend yield (informational, not part of tracking bridge)"
    sheet["C11"] = "=Assumptions!$E$7"
    sheet["C11"].number_format = PCT2
    set_widths(sheet, {"A": 4, "B": 56, "C": 20})
    sheet.freeze_panes = "A5"

    sheet = workbook["Checks"]
    title(sheet, "B2:C2", "ETF Model Checks")
    header(sheet, 4, 2, ["Check", "Status"])
    checks = [
        ("Holdings weights (shown + other) sum to ~100%",
         f'=IF(ABS(\'Portfolio Construction\'!D{other_row + 1}-1)<0.02,"PASS","REVIEW")'),
        ("Disclosed sector weights sum within a plausible band (90%-101%)",
         f'=IF(AND(\'Portfolio Construction\'!C{sector_total_row}>=0.90,\'Portfolio Construction\'!C{sector_total_row}<=1.01),"PASS","REVIEW")'),
        ("Implied shares outstanding is positive",
         '=IF(Assumptions!C16>0,"PASS","FAIL")'),
        ("Creation unit basket value is positive",
         '=IF(\'Creation & Redemption\'!C6>0,"PASS","FAIL")'),
        ("Net expense ratio is nonnegative and below 3%",
         '=IF(AND(Assumptions!E6>=0,Assumptions!E6<0.03),"PASS","REVIEW")'),
        ("Estimated net tracking difference within a plausible band (-2% to +0.5%)",
         '=IF(AND(\'Tracking Error & Costs\'!C9>=-0.02,\'Tracking Error & Costs\'!C9<=0.005),"PASS","REVIEW")'),
    ]
    for row, (label, formula) in enumerate(checks, start=5):
        sheet.cell(row, 2, label)
        sheet.cell(row, 3, formula)
    sheet["B11"] = "Overall"
    sheet["C11"] = '=IF(COUNTIF(C5:C10,"FAIL")>0,"FAIL",IF(COUNTIF(C5:C10,"REVIEW")>0,"REVIEW","PASS"))'
    add_status_rules(sheet, "C5:C11")
    set_widths(sheet, {"A": 4, "B": 58, "C": 18})

    add_sources(workbook, [
        ("ETF issuer fund profile / holdings disclosure", "[issuer fact sheet or N-CSR/N-PORT filing]", "[date]", "AUM, expense ratio, dividend yield, holdings, sector weights"),
        ("Market data (last price, volume)", "[exchange / market data vendor]", "[date]", "Market price used as a NAV proxy"),
        ("Prospectus / Statement of Additional Information", "[SEC-filed prospectus / SAI]", "[date]", "Creation unit size, in-kind/cash creation mix, authorized participant list"),
    ])
    add_refresh_log(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("ETF_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
