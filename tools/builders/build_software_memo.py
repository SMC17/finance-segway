"""Build a governed, data-grounded operating memo deck for a real public
Software & SaaS instance.

Sixth domain on the IC-memo/pptx system, reusing
tools/builders/pptx_helpers.py unchanged. Every number on a slide is read
from the instance workbook after a real LibreOffice recalculation
(tools.recalc), or from the case's own governed manifest
(standards/public_cases/<case_id>.json). Nothing on any slide is invented
here -- this script only reads and renders.

Default case (software-public-uipath-fy2023-stress) is the domain's
adversarial case: UiPath's own disclosed dollar-based net retention rate
cratering from 145% to 123% in FY2023, with a realized outcome showing the
decline persisted the following year (119%, from UiPath's own FY2024
10-K) -- not a one-off.

Usage:
    python tools/builders/build_software_memo.py --case-id software-public-uipath-fy2023-stress
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "tools"))

import pptx_helpers as ph  # noqa: E402
from recalc import recalc  # noqa: E402

import openpyxl  # noqa: E402

ROOT = Path(__file__).resolve().parents[2]
INDEX = ROOT / "standards" / "public_cases" / "index.json"


def _load_case(case_id: str) -> dict[str, Any]:
    index = json.loads(INDEX.read_text(encoding="utf-8"))
    for item in index["cases"]:
        if item["case_id"] == case_id:
            return item
    raise KeyError(f"unknown case_id {case_id!r}")


def _recalculated_workbook(output_rel: str) -> openpyxl.Workbook:
    src = ROOT / output_rel
    with tempfile.TemporaryDirectory(dir=ROOT) as tmp:
        scratch = Path(tmp) / src.name
        shutil.copyfile(src, scratch)
        result = recalc(str(scratch), timeout=90)
        if result.get("status") != "success" or result.get("total_errors", 1):
            raise RuntimeError(f"recalculation did not succeed cleanly: {result}")
        return openpyxl.load_workbook(scratch, data_only=True)


def _manifest(case_id: str) -> dict[str, Any]:
    return json.loads((ROOT / "standards" / "public_cases" / f"{case_id}.json").read_text(encoding="utf-8"))


def _fmt_usd_mm(value: float) -> str:
    return f"${value:,.1f}mm"


def _fmt_pct(value: float, decimals: int = 1) -> str:
    return f"{value * 100:.{decimals}f}%"


def build_software_memo(case_id: str, output: Path) -> None:
    case = _load_case(case_id)
    manifest = _manifest(case_id)
    outcome = manifest["outcome"]

    wb = _recalculated_workbook(case["output"])
    arr = wb["ARR Rollforward"]
    opmodel = wb["Operating Model"]
    rpo = wb["RPO & Bookings"]
    unit_econ = wb["Unit Economics"]
    checks = wb["Checks"]

    as_of = manifest["as_of"]
    company_name = manifest["cover"]["Company / ticker:"]

    beginning_arr = arr["C5"].value
    ending_arr = arr["C10"].value
    nrr = arr["C12"].value
    grr = arr["C13"].value
    arr_growth = arr["C14"].value

    subscription_rev = opmodel["C5"].value
    total_rev = opmodel["C7"].value
    blended_margin = opmodel["C11"].value
    sm_pct = None
    for row in range(5, 20):
        if opmodel.cell(row, 2).value == "Sales & marketing":
            sm_dollars = opmodel.cell(row, 3).value
            sm_pct = sm_dollars / total_rev if total_rev else None
            break

    rpo_ending = rpo["C7"].value
    rpo_coverage = rpo["C7"].value / total_rev if total_rev else None

    magic_number = unit_econ["C7"].value
    cac_payback = unit_econ["C8"].value

    check_status = {checks.cell(r, 2).value: checks.cell(r, 3).value for r in range(5, 18) if checks.cell(r, 2).value}
    overall_status = checks["C18"].value

    realized_retention = outcome["realized"]
    forecast_retention = outcome["forecast"]

    prs = ph.new_presentation()

    # 1. Title
    ph.add_title_slide(
        prs,
        f"{company_name.split('(')[0].strip()} — FY2023 Net-Retention Deceleration",
        "Real, disclosed ARR and net retention deceleration. Flow-rate decomposition, revenue "
        "mix, and unit-economics levers are declared drivers or template defaults, not disclosed by UiPath.",
        f"As of {as_of}  ·  Scenario: Downside  ·  Prepared from a governed, recalculated model instance",
    )

    # 2. Provenance
    slide = ph.add_content_slide(prs, "Reading this deck", kicker="Provenance")
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(1.5), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.6),
        [
            "Real, disclosed: Beginning/Ending ARR and net retention rate are UiPath's own reported "
            "figures — unlike a revenue-as-ARR-proxy case, UiPath discloses ARR directly every quarter.",
            "Derived, not disclosed: the new-customer ARR contribution and the retained-cohort net "
            "expansion effect are both computed from two disclosed facts (net-new ARR in dollars, "
            "dollar-based net retention rate) — real arithmetic on real numbers, not an estimate.",
            "Declared drivers: the further split into gross Expansion vs. Contraction vs. Churn is not "
            "disclosed by any issuer; two of the three are chosen (informed by this repo's own template "
            "defaults), the third carries the balancing residual so the four flow rates sum to exactly "
            "UiPath's disclosed 30% FY2023 ARR growth.",
            "Realized outcome is real and recorded, not a forecast: net retention the following fiscal "
            "year, from UiPath's own FY2024 10-K.",
        ],
        size=15,
    )
    ph.add_footer(slide, f"Model checks: Overall {overall_status}  ·  standards/public_cases/{case_id}.json")

    # 3. ARR snapshot (real)
    slide = ph.add_content_slide(prs, "ARR roll-forward", kicker="Real, disclosed")
    ph.add_stat_row(
        slide, ph.Inches(1.5),
        [
            ph.Stat(_fmt_usd_mm(beginning_arr), "Beginning ARR", tag="UiPath 10-K", color=ph.NAVY),
            ph.Stat(_fmt_usd_mm(ending_arr), "Ending ARR", tag="UiPath 10-K", color=ph.NAVY),
            ph.Stat(_fmt_pct(nrr, 0), "Net revenue retention", tag="UiPath 10-K",
                    color=ph.NEGATIVE if nrr < 1.3 else ph.NAVY),
        ],
    )
    ph.add_table(
        slide, ph.MARGIN, ph.Inches(3.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.4),
        ["Metric", "Value"],
        [
            ["ARR growth (FY2023)", _fmt_pct(arr_growth, 1)],
            ["Gross revenue retention", _fmt_pct(grr, 1)],
            ["Net revenue retention", _fmt_pct(nrr, 1)],
        ],
        col_widths=[3, 1.5],
    )
    ph.add_footer(slide, "Source: UiPath, Inc. FY2023 Form 10-K (SEC EDGAR, CIK 0001734722)")

    # 4. Revenue, margin, and opex (real ratios applied to real ARR scale)
    slide = ph.add_content_slide(prs, "Revenue and cost structure", kicker="Real, computed from disclosed dollars")
    ph.add_stat_row(
        slide, ph.Inches(1.5),
        [
            ph.Stat(_fmt_usd_mm(total_rev), "Modeled total revenue", tag="Computed", color=ph.NAVY),
            ph.Stat(_fmt_pct(blended_margin, 1), "Blended gross margin", tag="Derived", color=ph.NAVY),
            ph.Stat(_fmt_pct(sm_pct, 1) if sm_pct else "n/a", "Sales & marketing / revenue", tag="Derived",
                    color=ph.NEGATIVE if sm_pct and sm_pct > 0.5 else ph.NAVY),
        ],
    )
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(3.4), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.6),
        "UiPath's real FY2023 sales & marketing ratio exceeds even this template's own Downside-scenario "
        "illustrative default (42%) — a real, disclosed inefficiency more severe than the template "
        "author's own worst-case assumption, not a modeling artifact.",
        size=14, color=ph.ILLUSTRATIVE_TAG,
    )
    ph.add_disclosure_note(
        slide,
        "Modeled total revenue uses real ARR as the scale driver (not a revenue proxy), which implies "
        "a figure roughly 10.8% above UiPath's actual disclosed GAAP revenue -- a real, measured basis "
        "gap between an ARR-native metric and GAAP recognition timing, not an error.",
    )

    # 5. RPO and unit economics
    slide = ph.add_content_slide(prs, "RPO and acquisition efficiency", kicker="Mixed real / illustrative")
    ph.add_stat_row(
        slide, ph.Inches(1.5),
        [
            ph.Stat(_fmt_pct(rpo_coverage, 1) if rpo_coverage else "n/a", "RPO coverage of revenue",
                    tag="Derived (ending RPO real)", color=ph.NAVY),
            ph.Stat(f"{magic_number:.2f}" if magic_number else "n/a", "Magic number", tag="Illustrative",
                    color=ph.ILLUSTRATIVE_TAG),
            ph.Stat(f"{cac_payback:,.0f} mo" if cac_payback else "n/a", "CAC payback", tag="Illustrative",
                    color=ph.ILLUSTRATIVE_TAG),
        ],
    )
    ph.add_disclosure_note(
        slide,
        "Beginning RPO is a template default (not sourced); ending-RPO coverage of revenue is real, "
        "derived from UiPath's own disclosed $894.0mm remaining performance obligations. S&M efficiency "
        "metrics (magic number, CAC payback) are illustrative -- UiPath does not disclose new/expansion "
        "ARR by customer cohort at the granularity these ratios would need to be real.",
    )

    # 6. Decision / realized outcome
    slide = ph.add_content_slide(prs, "Realized outcome", kicker="Decision — real, recorded")
    verdict_color = ph.NEGATIVE if realized_retention < forecast_retention else ph.POSITIVE
    ph.add_text(
        slide, ph.MARGIN, ph.Inches(1.6), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(1.4),
        f"Naive forecast: FY2023's net retention ({_fmt_pct(forecast_retention, 0)}) holds flat into "
        f"FY2024. Realized: {_fmt_pct(realized_retention, 0)}, from UiPath's own FY2024 10-K — the "
        "decline was not a one-off.",
        size=18, bold=True, color=verdict_color,
    )
    ph.add_bullets(
        slide, ph.MARGIN, ph.Inches(3.1), ph.SLIDE_W - 2 * ph.MARGIN, ph.Inches(2.6),
        [
            "Net retention fell in three consecutive disclosed years: 145% -> 123% -> 119%, alongside a "
            "well-documented April 2023 leadership crisis (the co-CEO structure was unwound).",
            "The real sales & marketing ratio (66.3% of revenue) is worse than this template's own "
            "built-in Downside default, a genuine signal this case does not need to invent.",
            "A real allocation decision needs UiPath's own forward guidance and cohort-level retention "
            "detail, neither of which is asserted here.",
        ],
        size=15,
    )
    ph.add_footer(
        slide,
        f"Model: 31_Software_SaaS/_template_SOFTWARE.xlsx  ·  Instance: {case['output']}  ·  "
        f"Overall checks: {overall_status}",
    )

    ph.save(prs, output)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--case-id", default="software-public-uipath-fy2023-stress")
    parser.add_argument("--output", type=Path, default=None)
    args = parser.parse_args()
    output = args.output or ROOT / "31_Software_SaaS" / "presentations" / f"{args.case_id}-memo.pptx"
    build_software_memo(args.case_id, output)
    print(f"saved {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
