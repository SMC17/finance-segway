import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[BOND/CURVE] — Fixed Income Model", [
    ("Instrument:", "[fill in]"),
    ("Issuer / sector:", "[fill in]"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly"),
])

# ---------------- BOND PRICING ----------------
ws = wb.create_sheet("Bond Pricing")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Bond Pricing"; ws["B2"].font = TITLE
inputs = [
    ("Face value ($)", 1000, CUR),
    ("Coupon rate (annual, %)", 0.05, PCT),
    ("Coupon frequency (payments/yr)", 2, NUM),
    ("Years to maturity", 10, NUM),
    ("Yield to maturity (annual, %)", 0.055, PCT),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["B11"] = "Price (using PV formula)"
ws["C11"] = "=-PV(C9/C7,C7*C8,C5*C6/C7,C5)"
ws["C11"].font = BOLD; ws["C11"].number_format = CUR2; ws["C11"].border = BORDER
ws["D11"] = "PV(YTM per period, n periods, coupon per period, face value)"
ws["D11"].font = ITALIC_GRAY
ws["B12"] = "Price as % of par"
ws["C12"] = "=IFERROR(C11/C5,\"-\")"; ws["C12"].number_format = PCT2; ws["C12"].border = BORDER
ws["B13"] = "Current yield (annual coupon / price)"
ws["C13"] = "=IFERROR(C5*C6/C11,\"-\")"; ws["C13"].number_format = PCT2; ws["C13"].border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- DURATION & CONVEXITY ----------------
ws = wb.create_sheet("Duration & Convexity")
set_col_widths(ws, [4, 34, 16, 40])
ws["B2"] = "Duration & Convexity"; ws["B2"].font = TITLE
ws["B4"] = "Modified duration (approx., via price shock)"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
ws["B5"] = "Yield shock (bp) for numerical estimate"; ws["C5"] = 50; ws["C5"].font = BLUE; ws["C5"].number_format = '0'
ws["B6"] = "Price at YTM - shock"
ws["C6"] = "=-PV(('Bond Pricing'!C9-C5/10000)/'Bond Pricing'!C7,'Bond Pricing'!C7*'Bond Pricing'!C8,'Bond Pricing'!C5*'Bond Pricing'!C6/'Bond Pricing'!C7,'Bond Pricing'!C5)"
ws["C6"].number_format = CUR2
ws["B7"] = "Price at YTM + shock"
ws["C7"] = "=-PV(('Bond Pricing'!C9+C5/10000)/'Bond Pricing'!C7,'Bond Pricing'!C7*'Bond Pricing'!C8,'Bond Pricing'!C5*'Bond Pricing'!C6/'Bond Pricing'!C7,'Bond Pricing'!C5)"
ws["C7"].number_format = CUR2
ws["B8"] = "Modified duration = (P- - P+) / (2 x P0 x shock in decimal)"
ws["C8"] = "=IFERROR((C6-C7)/(2*'Bond Pricing'!C11*(C5/10000)),\"-\")"
ws["C8"].font = BOLD; ws["C8"].number_format = '0.00'
ws["B9"] = "Convexity = (P- + P+ - 2xP0) / (P0 x shock^2)"
ws["C9"] = "=IFERROR((C6+C7-2*'Bond Pricing'!C11)/('Bond Pricing'!C11*(C5/10000)^2),\"-\")"
ws["C9"].number_format = '0.00'
ws["B10"] = "Est. price change for 100bp move (duration + convexity)"
ws["C10"] = "=IFERROR(-C8*0.01+0.5*C9*0.01^2,\"-\")"
ws["C10"].number_format = PCT2
for r2 in range(6, 11):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- YIELD CURVE ----------------
ws = wb.create_sheet("Yield Curve")
set_col_widths(ws, [4, 14, 14, 14, 40])
ws["B2"] = "Yield Curve & Spread Analysis"; ws["B2"].font = TITLE
headers = ["", "Tenor", "Benchmark yield", "Instrument spread (bp)", "All-in yield"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers)-1)
tenors = ["3M", "2Y", "5Y", "10Y", "30Y"]
r = 5
for t in tenors:
    ws.cell(row=r, column=2, value=t).font = BLACK
    c_y = ws.cell(row=r, column=3, value=0); c_y.font = BLUE; c_y.number_format = PCT2; c_y.border = BORDER
    c_s = ws.cell(row=r, column=4, value=0); c_s.font = BLUE; c_s.number_format = '0'; c_s.border = BORDER
    c_all = ws.cell(row=r, column=5, value=f"=C{r}+D{r}/10000")
    c_all.number_format = PCT2; c_all.border = BORDER
    r += 1
ws["B11"] = "2s10s spread (bp)"
ws["C11"] = "=(C7-C6)*10000"; ws["C11"].number_format = '0'
ws["D11"] = "Negative = inverted curve"
ws["D11"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "/home/claude/model_shop/FIXED_INCOME_template.xlsx"
wb.save(out_path)
print("saved", out_path)
