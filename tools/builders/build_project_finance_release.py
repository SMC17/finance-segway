"""Canonical release wrapper for the institutional project-finance builder."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_project_finance_institutional import build as build_layout  # noqa: E402
from formula_compatibility import normalize_workbook_formulas  # noqa: E402
from institutional_helpers import MULT, finalize  # noqa: E402


def build(output: Path) -> None:
    build_layout(output)
    workbook = load_workbook(output)
    sensitivity = workbook["Sensitivity"]
    # Replace one range-array formula with ten explicit annual DSCR formulas.
    for row in range(5, 10):
        for column in range(3, 8):
            cost_letter = get_column_letter(column)
            annual = []
            for year_column in range(3, 13):
                year_letter = get_column_letter(year_column)
                annual.append(
                    f"MAX(0,Operations!{year_letter}5*(1-$B{row})*(1-{cost_letter}$4)-"
                    f"Operations!{year_letter}9-Operations!{year_letter}10-Operations!{year_letter}11)/"
                    f"MAX(0.000001,'Debt Sculpting'!{year_letter}10)"
                )
            sensitivity.cell(row, column, "=MIN(" + ",".join(annual) + ")")
            sensitivity.cell(row, column).number_format = MULT
    normalize_workbook_formulas(workbook)
    finalize(workbook, output)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=Path("PROJECT_FINANCE_template.xlsx"))
    arguments = parser.parse_args()
    build(arguments.output)
    print(f"saved {arguments.output}")
