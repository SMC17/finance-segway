import openpyxl
from template_helpers import *

wb = openpyxl.Workbook()
wb.remove(wb.active)

add_cover(wb, "[COMPANY] — Fintech / Payments Model", [
    ("Business model:", "Payments / lending / neobank / infra"),
    ("Last refreshed:", "[date]"),
    ("Refresh cadence:", "Weekly"),
])

# ---------------- UNIT ECONOMICS ----------------
ws = wb.create_sheet("Unit Economics")
set_col_widths(ws, [4, 30, 16, 40])
ws["B2"] = "Unit Economics"; ws["B2"].font = TITLE
inputs = [
    ("Total payment volume, TPV ($/mo)", 0, CUR),
    ("Take rate (%)", 0.025, PCT2),
    ("Interchange/processing cost (% of TPV)", 0.015, PCT2),
    ("CAC ($ per customer)", 0, CUR),
    ("Avg monthly revenue per customer ($)", 0, CUR),
    ("Monthly gross margin per customer (%)", 0.60, PCT),
    ("Monthly churn rate (%)", 0.03, PCT),
]
r = 5
for label, default, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B13"] = "Outputs"; ws["B13"].font = BOLD; ws["B13"].fill = GRAY_FILL
ws["B14"] = "Revenue (take rate x TPV)"
ws["C14"] = "=C5*C6"; ws["C14"].number_format = CUR
ws["B15"] = "Processing cost ($)"
ws["C15"] = "=C5*C7"; ws["C15"].number_format = CUR
ws["B16"] = "Net revenue after processing cost"
ws["C16"] = "=C14-C15"; ws["C16"].font = BOLD; ws["C16"].number_format = CUR
ws["B17"] = "Customer lifetime (months) = 1/churn"
ws["C17"] = "=IFERROR(1/C11,\"-\")"; ws["C17"].number_format = '0.0'
ws["B18"] = "LTV = monthly rev x gross margin x lifetime"
ws["C18"] = "=C9*C10*C17"; ws["C18"].font = BOLD; ws["C18"].number_format = CUR
ws["B19"] = "LTV / CAC ratio"
ws["C19"] = "=IFERROR(C18/C8,\"-\")"; ws["C19"].font = BOLD; ws["C19"].number_format = '0.00x'
ws["D19"] = "Rule of thumb: >3x is healthy, <1x means losing money per customer"
ws["D19"].font = ITALIC_GRAY
ws["B20"] = "CAC payback period (months)"
ws["C20"] = "=IFERROR(C8/(C9*C10),\"-\")"; ws["C20"].number_format = '0.0'
for r2 in range(14, 21):
    ws.cell(row=r2, column=3).border = BORDER
ws.sheet_view.showGridLines = False

# ---------------- COHORT RETENTION ----------------
ws = wb.create_sheet("Cohort Retention")
set_col_widths(ws, [4, 14] + [10]*8)
ws["B2"] = "Cohort Retention (% of cohort still active)"; ws["B2"].font = TITLE
ws["B4"] = "Cohort"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
months = [f"M{i}" for i in range(0, 8)]
for i, m in enumerate(months, start=3):
    c = ws.cell(row=4, column=i, value=m); c.font = BOLD; c.fill = GRAY_FILL
cohorts = ["Jan cohort", "Feb cohort", "Mar cohort", "Apr cohort"]
r = 5
for coh in cohorts:
    ws.cell(row=r, column=2, value=coh).font = BOLD
    for i, m in enumerate(months, start=3):
        val = 1.0 if i == 3 else 0
        c = ws.cell(row=r, column=i, value=val)
        c.font = BLUE; c.number_format = PCT; c.border = BORDER
    r += 1
ws["B9"] = "Average retention"; ws["B9"].font = BOLD
for i in range(3, 11):
    letter = get_column_letter(i)
    c = ws.cell(row=9, column=i, value=f"=AVERAGE({letter}5:{letter}8)")
    c.font = BOLD; c.number_format = PCT; c.border = BORDER
ws["B10"] = "Curve-implied LTV multiplier (sum of avg retention, M0-M7)"
ws["C10"] = "=SUM(C9:J9)"; ws["C10"].font = BOLD; ws["C10"].number_format = '0.00'
ws["C10"].border = BORDER
ws["B11"] = "M0 always = 100% by definition. Fill subsequent months as cohorts age."
ws["B11"].font = ITALIC_GRAY
ws["B12"] = ("Retention-curve LTV = monthly rev x gross margin x this multiplier — compare to the "
             "steady-state 1/churn LTV on Unit Economics; a big gap means churn isn't actually constant "
             "month to month (usually front-loaded), and the curve number is the more honest one.")
ws["B12"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

# ---------------- FRAUD & RISK ----------------
ws = wb.create_sheet("Fraud & Risk")
set_col_widths(ws, [4, 34, 16, 44])
ws["B2"] = "Fraud, Chargeback & Credit Loss"; ws["B2"].font = TITLE
ws["B4"] = "Inputs"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
inputs2 = [
    ("Total payment volume, TPV ($/mo)", 0, CUR),
    ("Fraud loss rate (bps of TPV)", 8, NUM),
    ("Chargeback count (this period)", 0, NUM),
    ("Chargeback cost per incident ($, fee + lost goods)", 25, CUR),
    ("Credit losses — lending book only, if applicable ($)", 0, CUR),
]
r = 5
for label, default, fmt in inputs2:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1

ws["B11"] = "Outputs"; ws["B11"].font = BOLD; ws["B11"].fill = GRAY_FILL
ws["B12"] = "Fraud loss ($)"
ws["C12"] = "=C5*C6/10000"; ws["C12"].number_format = CUR; ws["C12"].border = BORDER
ws["D12"] = "bps convention: 8 bps = 0.08%"
ws["D12"].font = ITALIC_GRAY
ws["B13"] = "Chargeback cost ($)"
ws["C13"] = "=C7*C8"; ws["C13"].number_format = CUR; ws["C13"].border = BORDER
ws["B14"] = "Total risk-related loss ($)"
ws["C14"] = "=C12+C13+C9"; ws["C14"].font = BOLD; ws["C14"].number_format = CUR; ws["C14"].border = BORDER
ws["B15"] = "Total loss as % of net revenue"
ws["C15"] = "=IFERROR(C14/'Unit Economics'!C16,\"-\")"; ws["C15"].font = BOLD; ws["C15"].number_format = PCT
ws["C15"].fill = YELLOW_FILL; ws["C15"].border = BORDER
ws["D15"] = "Card networks typically flag issuers/acquirers above ~90-100bps fraud-to-TPV as a monitoring risk"
ws["D15"].font = ITALIC_GRAY
ws.sheet_view.showGridLines = False

add_refresh_log(wb)
out_path = "FINTECH_template.xlsx"
wb.save(out_path)
print("saved", out_path)
