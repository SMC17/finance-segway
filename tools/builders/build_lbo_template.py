"""
Builds LBO_template.xlsx — private equity / merchant banking archetype.
Distinct mechanics from the IB base template: Sources & Uses, multi-tranche
debt schedule with cash sweep, and an IRR/MOIC returns waterfall.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
BOLD_WHITE = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=16, bold=True)
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
MULT = '0.0x'
ITALIC_GRAY = Font(name="Arial", size=8, italic=True, color="808080")

wb = openpyxl.Workbook()
wb.remove(wb.active)

def style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------------- COVER ----------------
ws = wb.create_sheet("Cover")
set_col_widths(ws, [4, 30, 40, 4])
ws["B2"] = "[TARGET] — LBO Model"; ws["B2"].font = TITLE
fields = [("Sponsor:", ""), ("Deal type:", "LBO / take-private / add-on"),
          ("Entry date:", "[date]"), ("Last refreshed:", "[date]"),
          ("Hold period (yrs):", 5), ("Refresh cadence:", "Weekly")]
r = 4
for label, default in fields:
    ws.cell(row=r, column=2, value=label).font = BOLD
    c = ws.cell(row=r, column=3, value=default); c.font = BLUE
    r += 1
ws.sheet_view.showGridLines = False

# ---------------- SOURCES & USES ----------------
ws = wb.create_sheet("Sources & Uses")
set_col_widths(ws, [4, 26, 14, 6, 26, 14])
ws["B2"] = "Sources & Uses"; ws["B2"].font = TITLE
ws["B4"] = "Sources"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
ws["E4"] = "Uses"; ws["E4"].font = BOLD; ws["E4"].fill = GRAY_FILL

sources = ["Revolver (undrawn at close)", "Term Loan A", "Term Loan B", "Senior Notes",
           "Sponsor equity", "Management rollover"]
uses = ["Purchase of equity (at entry mult.)", "Refinance existing debt",
        "Transaction fees", "Financing fees", "Cash to balance sheet"]

r = 5
for s in sources:
    ws.cell(row=r, column=2, value=s).font = BLACK
    c = ws.cell(row=r, column=3, value=0); c.font = BLUE; c.number_format = CUR; c.border = BORDER
    r += 1
ws.cell(row=r, column=2, value="Total sources").font = BOLD
ws.cell(row=r, column=3, value=f"=SUM(C5:C{r-1})").font = BOLD
ws.cell(row=r, column=3).number_format = CUR
src_total_row = r

r = 5
for u in uses:
    ws.cell(row=r, column=5, value=u).font = BLACK
    c = ws.cell(row=r, column=6, value=0); c.font = BLUE; c.number_format = CUR; c.border = BORDER
    r += 1
ws.cell(row=r, column=5, value="Total uses").font = BOLD
ws.cell(row=r, column=6, value=f"=SUM(F5:F{r-1})").font = BOLD
ws.cell(row=r, column=6).number_format = CUR
uses_total_row = r

ws.cell(row=max(src_total_row, uses_total_row)+2, column=2, value="Check (sources = uses):").font = BOLD
ws.cell(row=max(src_total_row, uses_total_row)+2, column=3,
        value=f"=C{src_total_row}-F{uses_total_row}").number_format = CUR
ws.sheet_view.showGridLines = False

# ---------------- DEBT SCHEDULE ----------------
ws = wb.create_sheet("Debt Schedule")
set_col_widths(ws, [4, 24, 12, 12, 12, 12, 12, 12, 4, 26, 14])
ws["B2"] = "Debt Schedule (with cash sweep)"; ws["B2"].font = TITLE
for i, h in enumerate(["", "", "Yr0", "Yr1", "Yr2", "Yr3", "Yr4", "Yr5"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 6, start_col=3)

# ---- assumptions block (drives every formula below; columns J:K) ----
ws["J4"] = "Assumptions"; ws["J4"].font = BOLD; ws["J4"].fill = GRAY_FILL
ws["K4"] = ""; ws["K4"].fill = GRAY_FILL
assumptions = [
    ("Entry EBITDA growth (%/yr)", 0.05, PCT),
    ("FCF conversion (FCF / EBITDA, %)", 0.50, PCT),
    ("Mandatory amort (% of original TLB/yr)", 0.01, PCT),
    ("Cash sweep (% of excess FCF, after debt svc)", 0.75, PCT),
    ("Term Loan B interest rate (%)", 0.09, PCT2),
]
r = 5
for label, default, fmt in assumptions:
    ws.cell(row=r, column=10, value=label).font = BLACK
    c = ws.cell(row=r, column=11, value=default); c.font = BLUE; c.fill = YELLOW_FILL
    c.number_format = fmt; c.border = BORDER
    r += 1
ws["J10"] = "Original TLB balance (Yr0, $)"
ws["K10"] = "='Sources & Uses'!C7"; ws["K10"].font = GREEN; ws["K10"].number_format = CUR
ws["K10"].border = BORDER
ws["J12"] = "Models Term Loan B only — the amortizing/sweep tranche in a typical"
ws["J13"] = "structure. Revolver stays undrawn at close; notes are bullet/no-amort."
ws["J12"].font = ITALIC_GRAY; ws["J13"].font = ITALIC_GRAY

# ---- EBITDA and FCF, grown off the entry EBITDA on the Returns tab ----
ws["B5"] = "EBITDA"; ws["B5"].font = GREEN
ws["C5"] = "=Returns!C5"; ws["C5"].font = GREEN
for col in range(4, 9):
    prev = get_column_letter(col - 1)
    ws.cell(row=5, column=col, value=f"={prev}5*(1+$K$5)")
ws["B6"] = "Cash flow available for debt paydown (FCF)"; ws["B6"].font = BLACK
for col in range(3, 9):
    letter = get_column_letter(col)
    ws.cell(row=6, column=col, value=f"={letter}5*$K$6")
for row in (5, 6):
    for c in range(3, 9):
        ws.cell(row=row, column=c).number_format = CUR
        ws.cell(row=row, column=c).border = BORDER

ws["B8"] = "Term Loan B"; ws["B8"].font = BOLD
ws["B9"] = "  Beginning balance"
ws["B10"] = "  Mandatory amortization"
ws["B11"] = "  Cash sweep (excess FCF after debt svc)"
ws["B12"] = "  Ending balance"
ws["B13"] = "  Interest expense (rate x beginning balance)"

# Beginning balance: Yr0 = original TLB at close; Yr1+ = prior year's ending.
ws["C9"] = "=$K$10"
for col in range(4, 9):
    prev = get_column_letter(col - 1)
    ws.cell(row=9, column=col, value=f"={prev}12")

# Mandatory amortization: flat % of the ORIGINAL balance, capped at what's left.
for col in range(3, 9):
    letter = get_column_letter(col)
    ws.cell(row=10, column=col, value=f"=MIN($K$10*$K$7,{letter}9)")

# Interest: on the BEGINNING balance only (not average) — keeps the schedule
# acyclic. Sweep can then depend on interest without a circular reference,
# since interest never depends on this period's ending balance.
for col in range(3, 9):
    letter = get_column_letter(col)
    ws.cell(row=13, column=col, value=f"={letter}9*$K$9")

# Cash sweep: sweep% of FCF left over after mandatory amort + interest,
# capped so it can never amortize past what mandatory amort already left.
for col in range(3, 9):
    letter = get_column_letter(col)
    ws.cell(row=11, column=col,
            value=f"=MIN(MAX(0,$K$8*({letter}6-{letter}10-{letter}13)),{letter}9-{letter}10)")

# Ending balance = beginning - mandatory amort - sweep.
for col in range(3, 9):
    letter = get_column_letter(col)
    ws.cell(row=12, column=col, value=f"={letter}9-{letter}10-{letter}11")

for row in [9, 10, 11, 12, 13]:
    for c in range(3, 9):
        cell = ws.cell(row=row, column=c)
        cell.number_format = CUR
        cell.border = BORDER
        cell.font = BLACK

ws["B15"] = "Total debt (end of period)"; ws["B15"].font = BOLD
ws["B16"] = "Total debt / EBITDA (leverage)"; ws["B16"].font = BOLD
for col in range(3, 9):
    letter = get_column_letter(col)
    ws.cell(row=15, column=col, value=f"={letter}12")
    ws.cell(row=15, column=col).number_format = CUR
    ws.cell(row=16, column=col, value=f"=IFERROR({letter}15/{letter}5,\"-\")")
    ws.cell(row=16, column=col).number_format = MULT
    ws.cell(row=16, column=col).fill = YELLOW_FILL
ws.sheet_view.showGridLines = False

# ---------------- RETURNS WATERFALL ----------------
ws = wb.create_sheet("Returns")
set_col_widths(ws, [4, 28, 14, 14, 14, 14])
ws["B2"] = "Returns — IRR / MOIC"; ws["B2"].font = TITLE

ws["B4"] = "Entry"; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
ws["B5"] = "Entry EBITDA"; ws["C5"] = 0; ws["C5"].font = BLUE; ws["C5"].number_format = CUR
ws["B6"] = "Entry multiple"; ws["C6"] = 0; ws["C6"].font = BLUE; ws["C6"].fill = YELLOW_FILL; ws["C6"].number_format = MULT
ws["B7"] = "Entry EV"; ws["C7"] = "=C5*C6"; ws["C7"].number_format = CUR
ws["B8"] = "Sponsor equity check"; ws["C8"] = "='Sources & Uses'!C9"; ws["C8"].font = GREEN; ws["C8"].number_format = CUR

ws["B10"] = "Exit"; ws["B10"].font = BOLD; ws["B10"].fill = GRAY_FILL
ws["B11"] = "Exit EBITDA (Yr5)"; ws["C11"] = "='Debt Schedule'!H5"; ws["C11"].font = GREEN; ws["C11"].number_format = CUR
ws["B12"] = "Exit multiple"; ws["C12"] = 0; ws["C12"].font = BLUE; ws["C12"].fill = YELLOW_FILL; ws["C12"].number_format = MULT
ws["B13"] = "Exit EV"; ws["C13"] = "=C11*C12"; ws["C13"].number_format = CUR
ws["B14"] = "Less: net debt at exit"; ws["C14"] = "='Debt Schedule'!H15"; ws["C14"].font = GREEN; ws["C14"].number_format = CUR
ws["B15"] = "Exit equity value"; ws["C15"] = "=C13-C14"; ws["C15"].font = BOLD; ws["C15"].number_format = CUR

ws["B17"] = "Returns"; ws["B17"].font = BOLD; ws["B17"].fill = GRAY_FILL
ws["B18"] = "MOIC"; ws["C18"] = "=IFERROR(C15/C8,\"-\")"; ws["C18"].font = BOLD; ws["C18"].number_format = MULT
ws["B19"] = "Hold period (yrs)"; ws["C19"] = "=Cover!C8"; ws["C19"].font = GREEN
ws["B20"] = "IRR"; ws["C20"] = "=IFERROR(C18^(1/C19)-1,\"-\")"; ws["C20"].font = BOLD; ws["C20"].number_format = PCT
ws.sheet_view.showGridLines = False

# ---------------- SENSITIVITY ----------------
ws = wb.create_sheet("Sensitivity")
set_col_widths(ws, [4, 20] + [12]*5)
ws["B2"] = "Sensitivity — IRR by Entry/Exit Multiple"; ws["B2"].font = TITLE
ws["B4"] = "Entry \\ Exit mult."; ws["B4"].font = BOLD; ws["B4"].fill = GRAY_FILL
exits = [7, 8, 9, 10, 11]
entries = [6, 7, 8, 9, 10]
for i, e in enumerate(exits, start=3):
    c = ws.cell(row=4, column=i, value=e); c.font = BOLD; c.fill = GRAY_FILL; c.number_format = MULT
for j, en in enumerate(entries, start=5):
    c = ws.cell(row=j, column=2, value=en); c.font = BOLD; c.fill = GRAY_FILL; c.number_format = MULT
    for i in range(3, 8):
        ws.cell(row=j, column=i, value="[data table — Excel What-If Analysis]").font = Font(
            name="Arial", size=8, italic=True, color="808080")
ws.sheet_view.showGridLines = False

# ---------------- REFRESH LOG ----------------
ws = wb.create_sheet("RefreshLog")
set_col_widths(ws, [4, 14, 16, 30, 30, 16])
ws["B2"] = "Refresh Log"; ws["B2"].font = TITLE
for i, h in enumerate(["", "Date", "Trigger", "What changed", "Reviewer notes", "Next check"], start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 5, start_col=2)
ws.sheet_view.showGridLines = False

out_path = "LBO_template.xlsx"
wb.save(out_path)
print("saved", out_path)
