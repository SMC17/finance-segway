"""Build and validate the reconciled Private Credit, Debt Finance, and Public Finance models."""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BUILDERS = ROOT / "tools" / "builders"
ERROR_TOKENS = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!"}


@dataclass(frozen=True)
class Contract:
    name: str
    builder: str
    committed: str
    required_sheets: tuple[str, ...]
    minimum_formulas: int
    required_labels: tuple[str, ...]


CONTRACTS = (
    Contract("Private Credit", "build_private_credit_template.py", "05_Private_Credit/_template_CREDIT.xlsx",
             ("Cover", "Assumptions", "Operating Case", "Debt Schedule", "Covenants", "Yield & Spread", "Recovery", "Sensitivity", "Checks", "Sources", "RefreshLog"),
             220, ("CFADS before interest", "Cash sweep", "Recovery rate", "Approx. all-in yield")),
    Contract("Debt Finance", "build_debt_finance_template.py", "06_Debt_Finance/_template_CREDIT.xlsx",
             ("Cover", "Assumptions", "Capital Structure", "Maturity Ladder", "Refinancing", "Interest Rate Risk", "Covenants", "Recovery", "Checks", "Sources", "RefreshLog"),
             100, ("Total maturities", "Total uses", "Total sources", "Pro forma interest coverage")),
    Contract("Public Finance", "build_public_finance_template.py", "07_Public_Finance/_template_PUBLIC_FINANCE.xlsx",
             ("Cover", "Assumptions", "Debt Sustainability", "Revenue & Expenditure", "Debt Service", "Coverage", "Scenarios", "Sensitivity", "Checks", "Sources", "RefreshLog"),
             220, ("Debt-stabilizing primary balance", "Revenue-bond DSCR", "Days cash on hand", "Primary-balance gap")),
)


def formulas_and_labels(wb: openpyxl.Workbook) -> tuple[list[str], set[str], list[str]]:
    formulas: list[str] = []
    labels: set[str] = set()
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    labels.add(value)
                    if value.startswith("="):
                        formulas.append(value)
                        upper = value.upper()
                        if any(token in upper for token in ERROR_TOKENS):
                            errors.append(f"{ws.title}!{cell.coordinate}: {value}")
                        if "[" in value and "]" in value:
                            errors.append(f"external formula link {ws.title}!{cell.coordinate}: {value}")
                    elif value.upper() in ERROR_TOKENS:
                        errors.append(f"literal error {ws.title}!{cell.coordinate}: {value}")
    return formulas, labels, errors


def validate_sheet_contract(sheetnames: list[str], required_sheets: tuple[str, ...]) -> list[str]:
    """Require every canonical sheet in canonical order while permitting governed extensions."""
    missing = [name for name in required_sheets if name not in sheetnames]
    if missing:
        return [f"missing required sheet(s): {missing}"]
    observed = tuple(name for name in sheetnames if name in required_sheets)
    if observed != required_sheets:
        return [f"required sheet order mismatch: {observed}"]
    return []


def validate(path: Path, contract: Contract) -> dict[str, object]:
    wb = openpyxl.load_workbook(path, data_only=False, keep_links=True)
    formulas, labels, errors = formulas_and_labels(wb)
    if wb._external_links:
        errors.append(f"{len(wb._external_links)} external workbook link(s)")
    errors.extend(validate_sheet_contract(wb.sheetnames, contract.required_sheets))
    if len(formulas) < contract.minimum_formulas:
        errors.append(f"formula depth {len(formulas)} < {contract.minimum_formulas}")
    for label in contract.required_labels:
        if label not in labels:
            errors.append(f"missing required label: {label}")
    check_formulas = [c.value for row in wb["Checks"].iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("=")]
    if not check_formulas:
        errors.append("Checks sheet has no formulas")
    return {"name": contract.name, "path": str(path), "formula_count": len(formulas), "errors": errors}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--committed", action="store_true", help="validate committed workbook paths instead of temporary builder outputs")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    results = []
    if args.committed:
        for contract in CONTRACTS:
            results.append(validate(ROOT / contract.committed, contract))
    else:
        with tempfile.TemporaryDirectory() as tmp:
            tmpdir = Path(tmp)
            for index, contract in enumerate(CONTRACTS):
                out = tmpdir / f"{index}_{Path(contract.committed).name}"
                subprocess.run([sys.executable, str(BUILDERS / contract.builder), "--output", str(out)], check=True)
                results.append(validate(out, contract))
    payload = {"contracts": [asdict(c) for c in CONTRACTS], "results": results}
    if args.report:
        args.report.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    for result in results:
        print(f"{result['name']}: {result['formula_count']} formulas")
        for error in result["errors"]:
            print(f"  ERROR: {error}")
    return 1 if any(r["errors"] for r in results) else 0


if __name__ == "__main__":
    raise SystemExit(main())
