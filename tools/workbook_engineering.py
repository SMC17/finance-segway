"""Workbook build, parity, and model-risk engineering utilities.

These checks operate on every workbook archetype and catch failure modes that
simple "file opens" validation misses: external links, literal Excel errors,
volatile formulas, whole-column scans, hidden calculation sheets, formula-depth
outliers, copied-formula inconsistencies, and builder/workbook drift.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

ERROR_TOKENS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}
VOLATILE_FUNCTIONS = {"OFFSET", "INDIRECT", "NOW", "TODAY", "RAND", "RANDBETWEEN", "CELL", "INFO"}
RISKY_FUNCTIONS = {"VLOOKUP", "HLOOKUP", "LOOKUP"}
FUNCTION_RE = re.compile(r"\b([A-Z][A-Z0-9_.]*)\s*\(", re.IGNORECASE)
WHOLE_COLUMN_RE = re.compile(r"(?<![A-Z0-9_])\$?[A-Z]{1,3}:\$?[A-Z]{1,3}(?![A-Z0-9_])", re.IGNORECASE)
EXTERNAL_LINK_RE = re.compile(r"\[[^\]]+\](?:[^!]+!)?")
MAGIC_NUMBER_RE = re.compile(r"(?<![A-Z0-9_.])(-?\d+(?:\.\d+)?)(?![A-Z0-9_.])", re.IGNORECASE)


@dataclass(frozen=True)
class Finding:
    severity: str
    code: str
    workbook: str
    sheet: str | None
    cell: str | None
    message: str


def _color(value: Any) -> str | None:
    if value is None:
        return None
    color_type = getattr(value, "type", None)
    if color_type == "rgb":
        return getattr(value, "rgb", None)
    if color_type == "theme":
        return f"theme:{getattr(value, 'theme', None)}:{getattr(value, 'tint', None)}"
    if color_type == "indexed":
        return f"indexed:{getattr(value, 'indexed', None)}"
    return None


def _border_side(side: Any) -> tuple[Any, ...]:
    return (getattr(side, "style", None), _color(getattr(side, "color", None)))


def _cell_signature(cell: Any) -> dict[str, Any]:
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        value_kind = "formula"
    elif value is None:
        value_kind = "blank"
    else:
        value_kind = type(value).__name__
    return {
        "coordinate": cell.coordinate,
        "value_kind": value_kind,
        "value": value,
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "color": _color(cell.font.color),
        },
        "fill": {
            "type": cell.fill.fill_type,
            "fg": _color(cell.fill.fgColor),
            "bg": _color(cell.fill.bgColor),
        },
        "border": {
            "left": _border_side(cell.border.left),
            "right": _border_side(cell.border.right),
            "top": _border_side(cell.border.top),
            "bottom": _border_side(cell.border.bottom),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "text_rotation": cell.alignment.text_rotation,
            "indent": cell.alignment.indent,
        },
        "protection": {
            "locked": cell.protection.locked,
            "hidden": cell.protection.hidden,
        },
    }


def workbook_fingerprint(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    sheets: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    cells.append(_cell_signature(cell))
        dimensions = {
            key: {"width": dim.width, "hidden": dim.hidden, "outline_level": dim.outlineLevel}
            for key, dim in sorted(ws.column_dimensions.items())
            if dim.width is not None or dim.hidden or dim.outlineLevel
        }
        rows = {
            str(key): {"height": dim.height, "hidden": dim.hidden, "outline_level": dim.outlineLevel}
            for key, dim in sorted(ws.row_dimensions.items())
            if dim.height is not None or dim.hidden or dim.outlineLevel
        }
        validations = []
        for validation in ws.data_validations.dataValidation:
            validations.append({
                "type": validation.type,
                "formula1": validation.formula1,
                "formula2": validation.formula2,
                "sqref": str(validation.sqref),
                "allow_blank": validation.allowBlank,
            })
        conditional_ranges = sorted(str(key.sqref) for key in ws.conditional_formatting)
        sheets.append({
            "title": ws.title,
            "state": ws.sheet_state,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "show_grid_lines": ws.sheet_view.showGridLines,
            "merged_ranges": sorted(str(item) for item in ws.merged_cells.ranges),
            "column_dimensions": dimensions,
            "row_dimensions": rows,
            "data_validations": sorted(validations, key=lambda item: (item["sqref"], str(item["type"]))),
            "conditional_format_ranges": conditional_ranges,
            "cells": cells,
        })
    defined_names = []
    for name, item in wb.defined_names.items():
        defined_names.append((name, getattr(item, "attr_text", None), getattr(item, "localSheetId", None)))
    payload = {
        "sheet_order": wb.sheetnames,
        "sheets": sheets,
        "defined_names": sorted(defined_names),
        "calculation": {
            "calc_mode": wb.calculation.calcMode,
            "full_calc_on_load": wb.calculation.fullCalcOnLoad,
            "force_full_calc": wb.calculation.forceFullCalc,
        },
        "external_links": len(getattr(wb, "_external_links", [])),
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")
    payload["sha256"] = hashlib.sha256(encoded).hexdigest()
    return payload


def compare_workbooks(generated: Path, committed: Path) -> dict[str, Any]:
    left = workbook_fingerprint(generated)
    right = workbook_fingerprint(committed)
    differences: list[str] = []
    if left["sheet_order"] != right["sheet_order"]:
        differences.append("sheet_order")
    left_sheets = {sheet["title"]: sheet for sheet in left["sheets"]}
    right_sheets = {sheet["title"]: sheet for sheet in right["sheets"]}
    for name in sorted(set(left_sheets) | set(right_sheets)):
        if name not in left_sheets:
            differences.append(f"sheet_missing_generated:{name}")
            continue
        if name not in right_sheets:
            differences.append(f"sheet_missing_committed:{name}")
            continue
        for field in (
            "state", "freeze_panes", "show_grid_lines", "merged_ranges",
            "column_dimensions", "row_dimensions", "data_validations",
            "conditional_format_ranges", "cells",
        ):
            if left_sheets[name][field] != right_sheets[name][field]:
                differences.append(f"{name}:{field}")
    for field in ("defined_names", "calculation", "external_links"):
        if left[field] != right[field]:
            differences.append(field)
    return {
        "generated": str(generated),
        "committed": str(committed),
        "generated_sha256": left["sha256"],
        "committed_sha256": right["sha256"],
        "parity": not differences,
        "differences": differences,
    }


def _formula_functions(formula: str) -> set[str]:
    return {match.group(1).upper() for match in FUNCTION_RE.finditer(formula)}


def _magic_numbers(formula: str) -> list[str]:
    allowed = {"0", "1", "2", "4", "12", "100", "365", "10000"}
    numbers = []
    for match in MAGIC_NUMBER_RE.finditer(formula):
        value = match.group(1)
        if value not in allowed and not value.startswith("0."):
            numbers.append(value)
    return numbers


def audit_workbook(path: Path) -> tuple[dict[str, Any], list[Finding]]:
    wb = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    findings: list[Finding] = []
    formula_count = 0
    literal_error_count = 0
    formula_lengths: list[int] = []
    hidden_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state != "visible"]
    if getattr(wb, "_external_links", []):
        findings.append(Finding("error", "external_link_part", str(path), None, None, "workbook contains external-link parts"))
    for sheet in hidden_sheets:
        findings.append(Finding("warning", "hidden_sheet", str(path), sheet, None, "hidden calculation sheets require explicit review"))
    formula_map: dict[tuple[str, int], list[tuple[str, str]]] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value in ERROR_TOKENS:
                    literal_error_count += 1
                    findings.append(Finding("error", "literal_excel_error", str(path), ws.title, cell.coordinate, value))
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                formula_count += 1
                formula_lengths.append(len(value))
                functions = _formula_functions(value)
                volatile = sorted(functions & VOLATILE_FUNCTIONS)
                risky = sorted(functions & RISKY_FUNCTIONS)
                if volatile:
                    findings.append(Finding("warning", "volatile_formula", str(path), ws.title, cell.coordinate, ", ".join(volatile)))
                if risky:
                    findings.append(Finding("info", "legacy_lookup", str(path), ws.title, cell.coordinate, ", ".join(risky)))
                if EXTERNAL_LINK_RE.search(value):
                    findings.append(Finding("error", "external_formula_reference", str(path), ws.title, cell.coordinate, value))
                if WHOLE_COLUMN_RE.search(value):
                    findings.append(Finding("warning", "whole_column_reference", str(path), ws.title, cell.coordinate, value))
                if value.upper().count("IF(") > 3:
                    findings.append(Finding("warning", "deeply_nested_if", str(path), ws.title, cell.coordinate, value))
                if len(value) > 240:
                    findings.append(Finding("warning", "long_formula", str(path), ws.title, cell.coordinate, f"length={len(value)}"))
                magic = _magic_numbers(value)
                if magic:
                    findings.append(Finding("info", "magic_number", str(path), ws.title, cell.coordinate, ",".join(magic[:8])))
                formula_map.setdefault((ws.title, cell.row), []).append((cell.coordinate, value))
    for (sheet, _row), formulas in formula_map.items():
        if len(formulas) < 3:
            continue
        formulas = sorted(formulas)
        origin = formulas[0][0]
        normalized: list[str] = []
        for coordinate, formula in formulas:
            try:
                normalized.append(Translator(formula, origin=coordinate).translate_formula(origin))
            except Exception:
                normalized.append(formula)
        if len(set(normalized)) > max(2, len(normalized) // 2):
            findings.append(Finding(
                "info", "row_formula_inconsistency", str(path), sheet, origin,
                f"{len(set(normalized))} patterns across {len(normalized)} formulas",
            ))
    summary = {
        "workbook": str(path),
        "sheets": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "hidden_sheets": hidden_sheets,
        "formulas": formula_count,
        "literal_errors": literal_error_count,
        "external_links": len(getattr(wb, "_external_links", [])),
        "max_formula_length": max(formula_lengths, default=0),
        "findings": {
            severity: sum(1 for item in findings if item.severity == severity)
            for severity in ("error", "warning", "info")
        },
    }
    return summary, findings


def audit_library(root: Path) -> dict[str, Any]:
    workbooks = sorted(
        path for path in root.rglob("*.xlsx")
        if not any(part.startswith(".") for part in path.parts)
    )
    summaries = []
    findings: list[Finding] = []
    for workbook in workbooks:
        summary, workbook_findings = audit_workbook(workbook)
        summaries.append(summary)
        findings.extend(workbook_findings)
    return {
        "workbook_count": len(workbooks),
        "formula_count": sum(item["formulas"] for item in summaries),
        "error_count": sum(1 for item in findings if item.severity == "error"),
        "warning_count": sum(1 for item in findings if item.severity == "warning"),
        "info_count": sum(1 for item in findings if item.severity == "info"),
        "workbooks": summaries,
        "findings": [asdict(item) for item in findings],
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("root", nargs="?", type=Path, default=Path("."))
    parser.add_argument("--report", type=Path, default=Path("workbook-engineering-report.json"))
    parser.add_argument("--fail-on-warnings", action="store_true")
    args = parser.parse_args(argv)
    report = audit_library(args.root.resolve())
    args.report.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(json.dumps({key: report[key] for key in (
        "workbook_count", "formula_count", "error_count",
        "warning_count", "info_count",
    )}, indent=2))
    if report["error_count"]:
        return 1
    if args.fail_on_warnings and report["warning_count"]:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
