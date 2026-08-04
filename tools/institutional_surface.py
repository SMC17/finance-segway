"""Institutional operating surface for every finance-segway workbook."""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

SURFACE_SHEETS = ("Institutional Surface", "Challenge Log", "Lineage Map")
NAVY, BLUE, WHITE = "17365D", "D9EAF7", "FFFFFF"
THIN = Side(style="thin", color="B7B7B7")

ANCHORS = {
    "model_risk": ("Federal Reserve SR 26-2 Model Risk Management", "https://www.federalreserve.gov/supervisionreg/srletters/SR2602.htm", "governance, validation, change control, monitoring"),
    "basel": ("Basel Framework", "https://www.bis.org/basel_framework/", "prudential capital, liquidity, credit and market risk"),
    "frtb": ("Basel MAR21 Sensitivities-Based Method", "https://www.bis.org/basel_framework/chapter/MAR/21.htm", "market-risk factors and aggregation"),
    "irrbb": ("Basel SRP31 IRRBB", "https://www.bis.org/basel_framework/chapter/SRP/31.htm", "EVE/NII, basis and option risk"),
    "imf_dsa": ("IMF Debt Sustainability Analysis", "https://www.imf.org/external/pubs/ft/dsa/index.htm", "baseline and debt stress testing"),
    "naic_rbc": ("NAIC Risk-Based Capital", "https://content.naic.org/insurance-topics/risk-based-capital", "risk-sensitive statutory capital"),
    "ifrs17": ("IFRS 17 Insurance Contracts", "https://www.ifrs.org/issued-standards/list-of-standards/ifrs-17-insurance-contracts/", "insurance measurement and risk release"),
    "world_bank_ppp": ("World Bank PPP Reference Guide", "https://ppp.worldbank.org/public-private-partnership/library/ppp-reference-guide-3-0", "project preparation and risk allocation"),
    "sec_fair_value": ("SEC Rule 2a-5", "https://www.sec.gov/files/rules/final/2020/ic-34128.pdf", "fair-value governance and testing"),
    "iosco_liquidity": ("IOSCO Liquidity Risk Recommendations", "https://www.iosco.org/library/pubdocs/pdf/IOSCOPD590.pdf", "fund liquidity and stress testing"),
    "cftc_margin": ("CFTC Uncleared Swap Margin", "https://www.cftc.gov/LawRegulation/FederalRegister/FinalRules/2015-32320.html", "initial and variation margin"),
    "ffiec_payments": ("FFIEC Retail Payment Systems", "https://ithandbook.ffiec.gov/it-booklets/retail-payment-systems.aspx", "payments, settlement and fraud risk"),
    "occ_credit": ("OCC Rating Credit Risk", "https://www.occ.treas.gov/publications-and-resources/publications/comptrollers-handbook/files/rating-credit-risk/pub-ch-rating-credit-risk.pdf", "repayment capacity and risk rating"),
    "fasb_cecl": ("FASB CECL", "https://www.fasb.org/page/PageContent?pageId=/projects/recentlycompleted/credit-losses.html", "lifetime expected credit loss"),
    "sec_crypto": ("SEC SAB 122", "https://www.sec.gov/rules-regulations/staff-guidance/staff-accounting-bulletins/staff-accounting-bulletin-122", "crypto safeguarding accounting"),
}
DEFAULTS = {
    "key_outputs": ["base, downside and break-even outputs", "liquidity / funding requirement and binding constraint"],
    "primary_documents": ["approved model card and assumption register", "independent validation and challenge record"],
    "scenario_families": ["data freshness / source failure", "combined severe-but-plausible downside"],
    "challenge_tests": ["source-to-model lineage is reproducible", "units, signs, dates and legal definitions reconcile"],
    "failure_modes": ["stale vintage presented as current", "decision output disconnected from a binding constraint"],
    "insider_questions": ["Which assumption is owner-approved versus modeler judgement?", "What fails first and who must act?"],
    "source_classes": [
        {"name": "model governance evidence", "cadence": "per release", "control": "version, reviewer, sign-off and change log"},
        {"name": "decision-use snapshot", "cadence": "per decision", "control": "immutable as-of date and source receipt"},
    ],
}


def _split(value: str) -> list[str]:
    return [item.strip() for item in (value or "").split("|") if item.strip()]


def load_inventory(root: Path) -> dict[str, Any]:
    return json.loads((root / "standards/model_inventory.json").read_text())


def load_profiles(root: Path) -> dict[str, Any]:
    profiles = []
    for path in sorted((root / "standards/domain_profiles").glob("*.tsv")):
        for row in csv.DictReader(path.open(encoding="utf-8"), delimiter="\t"):
            sources = []
            for raw in _split(row["sources"]):
                name, cadence, control = [item.strip() for item in raw.split("~", 2)]
                sources.append({"name": name, "cadence": cadence, "control": control})
            keys = ["model_risk", *_split(row["anchors"].replace(",", "|"))]
            profiles.append({
                "id": row["id"], "domain": row["domain"],
                "decision_arena": row["arena"], "committee_artifact": row["artifact"],
                "operating_cadence": row["cadence"],
                "key_outputs": _split(row["outputs"]) + DEFAULTS["key_outputs"],
                "primary_documents": _split(row["documents"]) + DEFAULTS["primary_documents"],
                "scenario_families": _split(row["scenarios"]) + DEFAULTS["scenario_families"],
                "challenge_tests": _split(row["challenges"]) + DEFAULTS["challenge_tests"],
                "failure_modes": _split(row["failures"]) + DEFAULTS["failure_modes"],
                "insider_questions": _split(row["questions"]) + DEFAULTS["insider_questions"],
                "source_classes": sources + DEFAULTS["source_classes"],
                "regulatory_anchors": [
                    {"name": ANCHORS[key][0], "url": ANCHORS[key][1], "scope": ANCHORS[key][2]}
                    for key in dict.fromkeys(keys) if key
                ],
            })
    return {"schema_version": "1.0", "as_of": "2026-08-03", "profiles": profiles}


def profiles_by_id(root: Path) -> dict[str, dict[str, Any]]:
    return {item["id"]: item for item in load_profiles(root)["profiles"]}


def validate_profiles(root: Path) -> list[str]:
    models = {item["id"]: item for item in load_inventory(root)["models"]}
    profiles = load_profiles(root)["profiles"]
    errors, seen = [], set()
    for profile in profiles:
        pid = profile["id"]
        if pid in seen:
            errors.append(f"duplicate_profile:{pid}")
        seen.add(pid)
        if pid not in models:
            errors.append(f"orphan_profile:{pid}")
            continue
        if profile["domain"] != models[pid]["domain"]:
            errors.append(f"{pid}:domain_mismatch")
        for field in ("key_outputs", "primary_documents", "scenario_families", "challenge_tests", "failure_modes", "insider_questions"):
            if len(profile[field]) < 5:
                errors.append(f"{pid}:insufficient:{field}")
        if len(profile["source_classes"]) < 4:
            errors.append(f"{pid}:insufficient:source_classes")
        for anchor in profile["regulatory_anchors"]:
            if not anchor["url"].startswith("https://"):
                errors.append(f"{pid}:invalid_anchor")
    errors.extend(f"missing_profile:{pid}" for pid in sorted(set(models) - seen))
    return sorted(set(errors))


def _title(sheet, text: str, end: int) -> None:
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end)
    cell = sheet.cell(2, 2, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=15)


def _header(sheet, row: int, headers: list[str]) -> None:
    for column, value in enumerate(headers, 2):
        cell = sheet.cell(row, column, value)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = Border(bottom=THIN)


def _section(sheet, row: int, text: str) -> int:
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    cell = sheet.cell(row, 2, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True)
    return row + 1


def _name(workbook, name: str, ref: str) -> None:
    if name in workbook.defined_names:
        del workbook.defined_names[name]
    workbook.defined_names.add(DefinedName(name, attr_text=ref))


def _surface(workbook, model: dict[str, Any], profile: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Institutional Surface", 1)
    sheet.sheet_view.showGridLines = False
    _title(sheet, f"{model['domain']} — Institutional Decision Surface", 6)
    metadata = [
        ("Model ID", model["id"]), ("Domain", model["domain"]),
        ("Declared maturity", model["declared_maturity"]), ("Canonical builder", model["builder"]),
        ("Decision arena", profile["decision_arena"]), ("Committee artifact", profile["committee_artifact"]),
        ("Operating cadence", profile["operating_cadence"]),
    ]
    row = 4
    for label, value in metadata:
        sheet.cell(row, 2, label).font = Font(bold=True)
        sheet.cell(row, 3, value)
        sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        sheet.cell(row, 3).alignment = Alignment(wrap_text=True)
        row += 1
    blocks = [
        ("Key decision outputs", profile["key_outputs"]),
        ("Insider questions", profile["insider_questions"]),
        ("Scenario families", profile["scenario_families"]),
        ("Primary diligence documents", profile["primary_documents"]),
        ("Challenge tests", profile["challenge_tests"]),
        ("Known failure modes", profile["failure_modes"]),
    ]
    for title, values in blocks:
        row += 1
        row = _section(sheet, row, title)
        _header(sheet, row, ["#", "Requirement", "Owner", "Evidence / location", "Status"])
        row += 1
        start = row
        for index, value in enumerate(values, 1):
            for column, item in enumerate((index, value, "", "", "OPEN"), 2):
                sheet.cell(row, column, item)
            row += 1
        validation = DataValidation(type="list", formula1='"OPEN,IN REVIEW,CLOSED,NOT APPLICABLE"')
        sheet.add_data_validation(validation)
        validation.add(f"F{start}:F{row - 1}")
    row += 1
    row = _section(sheet, row, "Regulatory / methodological anchors")
    _header(sheet, row, ["Standard", "Scope", "Source URL", "Applicability", "Review note"])
    row += 1
    for anchor in profile["regulatory_anchors"]:
        for column, value in enumerate((anchor["name"], anchor["scope"], anchor["url"], "Map explicitly", ""), 2):
            sheet.cell(row, column, value)
        row += 1
    for column, width in {"B": 26, "C": 48, "D": 24, "E": 34, "F": 20}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "B5"
    _name(workbook, "FS_MODEL_ID", "'Institutional Surface'!$C$4")
    _name(workbook, "FS_DOMAIN", "'Institutional Surface'!$C$5")
    _name(workbook, "FS_MATURITY", "'Institutional Surface'!$C$6")


def _challenge(workbook, profile: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Challenge Log", 2)
    sheet.sheet_view.showGridLines = False
    _title(sheet, "Independent Challenge, Override and Closure Log", 9)
    _header(sheet, 4, ["Date", "Reviewer", "Challenge", "Owner response", "Evidence", "Impact", "Status", "Closure"])
    row = 5
    for challenge in profile["challenge_tests"]:
        for column, value in enumerate(("", "", challenge, "", "", "", "OPEN", ""), 2):
            sheet.cell(row, column, value)
        row += 1
    for _ in range(45):
        row += 1
    status = DataValidation(type="list", formula1='"OPEN,IN REVIEW,REMEDIATED,ACCEPTED RISK,CLOSED"')
    impact = DataValidation(type="list", formula1='"LOW,MEDIUM,HIGH,CRITICAL"')
    sheet.add_data_validation(status); sheet.add_data_validation(impact)
    status.add("H5:H54"); impact.add("G5:G54")
    for column, width in {"B": 13, "C": 22, "D": 52, "E": 46, "F": 36, "G": 14, "H": 18, "I": 14}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "B5"


def _lineage(workbook, profile: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Lineage Map", 3)
    sheet.sheet_view.showGridLines = False
    _title(sheet, "Source, Transformation, Ownership and Refresh Map", 10)
    _header(sheet, 4, ["Source class", "Cadence", "Control", "System / provider", "As-of", "Owner", "Transform", "Downstream", "Status"])
    row = 5
    for source in profile["source_classes"]:
        values = (source["name"], source["cadence"], source["control"], "", "", "", "", "", "TO MAP")
        for column, value in enumerate(values, 2):
            sheet.cell(row, column, value)
        row += 1
    status = DataValidation(type="list", formula1='"TO MAP,ACTIVE,STALE,EXCEPTION,RETIRED"')
    sheet.add_data_validation(status); status.add("J5:J40")
    for column, width in {"B": 24, "C": 18, "D": 42, "E": 24, "F": 18, "G": 22, "H": 42, "I": 30, "J": 15}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "B5"


def apply_surface(path: Path, model: dict[str, Any], profile: dict[str, Any]) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    for name in SURFACE_SHEETS:
        if name in workbook.sheetnames:
            del workbook[name]
    _surface(workbook, model, profile)
    _challenge(workbook, profile)
    _lineage(workbook, profile)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(path)
    return {"model_id": model["id"], "workbook": str(path), "surface_sheets": list(SURFACE_SHEETS)}


def validate_workbook_surface(path: Path, model: dict[str, Any], profile: dict[str, Any]) -> list[str]:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    errors = [f"{model['id']}:missing_sheet:{name}" for name in SURFACE_SHEETS if name not in workbook.sheetnames]
    if "Institutional Surface" in workbook.sheetnames:
        surface = workbook["Institutional Surface"]
        expected = {"C4": model["id"], "C5": model["domain"], "C6": model["declared_maturity"], "C7": model["builder"], "C8": profile["decision_arena"]}
        errors.extend(f"{model['id']}:surface_mismatch:{cell}" for cell, value in expected.items() if surface[cell].value != value)
    errors.extend(f"{model['id']}:missing_defined_name:{name}" for name in ("FS_MODEL_ID", "FS_DOMAIN", "FS_MATURITY") if name not in workbook.defined_names)
    return errors
