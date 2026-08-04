"""Build the finance-segway institutional control-plane workbook."""
from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from institutional_surface import load_inventory, load_profiles, validate_profiles

NAVY, BLUE, GREEN, YELLOW, RED, WHITE = "17365D", "D9EAF7", "E2F0D9", "FFF2CC", "FCE4D6", "FFFFFF"
THIN = Side(style="thin", color="B7B7B7")
INTEGER = '#,##0;[Red](#,##0);-'


def _read_optional(path: Path, default: Any) -> Any:
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else default


def _title(sheet, text: str, end_column: int) -> None:
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=end_column)
    cell = sheet.cell(2, 2, text)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(color=WHITE, bold=True, size=15)
    cell.alignment = Alignment(horizontal="left")
    sheet.sheet_view.showGridLines = False


def _header(sheet, row: int, headers: list[str]) -> None:
    for column, text in enumerate(headers, 2):
        cell = sheet.cell(row, column, text)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN)


def _body(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))


def _widths(sheet, mapping: dict[str, float]) -> None:
    for column, width in mapping.items():
        sheet.column_dimensions[column].width = width


def build(root: Path, output: Path) -> dict[str, Any]:
    profile_errors = validate_profiles(root)
    if profile_errors:
        raise ValueError(f"invalid institutional profile registry: {profile_errors}")
    inventory = load_inventory(root)
    profile_payload = load_profiles(root)
    profiles = {item["id"]: item for item in profile_payload["profiles"]}
    release = _read_optional(root / "standards/releases/flagship-2.1.0.json", {"models": [], "status": "MISSING"})
    release_by_id = {item["model_id"]: item for item in release.get("models", [])}
    public = _read_optional(
        root / "standards/public_cases/index.json", {"cases": [], "case_count": 0}
    )

    workbook = Workbook()
    workbook.remove(workbook.active)

    dashboard = workbook.create_sheet("Executive Dashboard")
    _title(dashboard, "Finance-Segway Institutional Control Plane", 11)
    for row, (label, value) in enumerate((
        ("As of", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Inventory version", inventory["version"]),
        ("Profile registry", profile_payload["schema_version"]),
        ("Release status", release.get("status", "MISSING")),
    ), 4):
        dashboard.cell(row, 2, label).font = Font(bold=True)
        dashboard.cell(row, 3, value)
    kpis = [
        ("Core models", len(inventory["models"]), "Inventory-controlled archetypes"),
        ("M2 models", sum(m["declared_maturity"] == "M2" for m in inventory["models"]), "Integrated decision models"),
        ("M1 models", sum(m["declared_maturity"] == "M1" for m in inventory["models"]), "Correct skeletons awaiting deeper mechanics"),
        ("Institutional profiles", len(profiles), "Domain-specific decision, diligence and challenge profiles"),
        ("Public evidence cases", public.get("case_count", len(public.get("cases", []))), "Source-addressed historical and adversarial cases"),
        ("Release-evidenced models", len(release_by_id), "Models with SHA-256 release evidence"),
        ("Required engines", sum(len(m["required_engines"]) for m in inventory["models"]), "Declared financial-engine requirements"),
        ("Stakeholder perspectives", sum(len(m["required_perspectives"]) for m in inventory["models"]), "Declared stakeholder views"),
        ("Independent checks", sum(len(m["reference_checks"]) for m in inventory["models"]), "Declared reference checks"),
    ]
    _header(dashboard, 9, ["Metric", "Value", "Definition"])
    for row, values in enumerate(kpis, 10):
        for column, value in enumerate(values, 2):
            dashboard.cell(row, column, value)
        dashboard.cell(row, 3).number_format = INTEGER
    _body(dashboard, 10, 18, 2, 4)
    dashboard["F9"], dashboard["G9"] = "Maturity", "Count"
    for row, maturity in enumerate(("M1", "M2", "M3", "M4"), 10):
        dashboard.cell(row, 6, maturity)
        dashboard.cell(row, 7, sum(model["declared_maturity"] == maturity for model in inventory["models"]))
    chart = BarChart()
    chart.type, chart.style, chart.title = "col", 10, "Model maturity distribution"
    chart.y_axis.title, chart.x_axis.title = "Models", "Maturity"
    chart.add_data(Reference(dashboard, min_col=7, min_row=9, max_row=13), titles_from_data=True)
    chart.set_categories(Reference(dashboard, min_col=6, min_row=10, max_row=13))
    chart.height, chart.width = 7, 12
    dashboard.add_chart(chart, "F15")
    _widths(dashboard, {"A": 3, "B": 30, "C": 16, "D": 58, "F": 16, "G": 12})

    inventory_sheet = workbook.create_sheet("Model Inventory")
    _title(inventory_sheet, "Canonical Inventory and Release Coverage", 16)
    _header(inventory_sheet, 4, ["ID", "Domain", "Archetype", "Maturity", "Target", "Builder", "Workbook", "Decision arena", "Engines", "Perspectives", "Checks", "Scenarios", "Source classes", "Release formulas", "Release status"])
    for row, model in enumerate(inventory["models"], 5):
        profile, evidence = profiles[model["id"]], release_by_id.get(model["id"], {})
        values = [model["id"], model["domain"], model["archetype"], model["declared_maturity"], model["target_maturity"], model["builder"], model["workbook"], profile["decision_arena"], len(model["required_engines"]), len(model["required_perspectives"]), len(model["reference_checks"]), len(profile["scenario_families"]), len(profile["source_classes"]), evidence.get("formula_count"), "EVIDENCED" if evidence else "NOT IN FLAGSHIP RELEASE"]
        for column, value in enumerate(values, 2):
            inventory_sheet.cell(row, column, value)
    last = 4 + len(inventory["models"])
    inventory_sheet.auto_filter.ref, inventory_sheet.freeze_panes = f"B4:P{last}", "B5"
    inventory_sheet.conditional_formatting.add(f"E5:E{last}", ColorScaleRule(start_type="min", start_color=RED, mid_type="percentile", mid_value=50, mid_color=YELLOW, end_type="max", end_color=GREEN))
    _body(inventory_sheet, 5, last, 2, 16)
    _widths(inventory_sheet, {"A": 3, "B": 7, "C": 30, "D": 16, "E": 12, "F": 12, "G": 42, "H": 42, "I": 46, "J": 10, "K": 12, "L": 10, "M": 10, "N": 12, "O": 14, "P": 24})

    engine_sheet = workbook.create_sheet("Engine Coverage")
    _title(engine_sheet, "Financial Engine and Stakeholder Coverage", 8)
    _header(engine_sheet, 4, ["Model ID", "Domain", "Type", "Requirement", "Maturity", "Evidence", "Owner"])
    row = 5
    for model in inventory["models"]:
        for kind, values in (("ENGINE", model["required_engines"]), ("PERSPECTIVE", model["required_perspectives"]), ("REFERENCE CHECK", model["reference_checks"])):
            for requirement in values:
                for column, value in enumerate((model["id"], model["domain"], kind, requirement, model["declared_maturity"], "", ""), 2):
                    engine_sheet.cell(row, column, value)
                row += 1
    engine_sheet.auto_filter.ref, engine_sheet.freeze_panes = f"B4:H{row - 1}", "B5"
    _body(engine_sheet, 5, row - 1, 2, 8)
    _widths(engine_sheet, {"A": 3, "B": 10, "C": 30, "D": 18, "E": 40, "F": 12, "G": 36, "H": 22})

    decision_sheet = workbook.create_sheet("Decision Arenas")
    _title(decision_sheet, "Decision Outputs and Insider Questions", 8)
    _header(decision_sheet, 4, ["Model ID", "Domain", "Decision arena", "Committee artifact", "Item type", "Item", "Status"])
    row = 5
    for profile in profile_payload["profiles"]:
        for item_type, values in (("KEY OUTPUT", profile["key_outputs"]), ("INSIDER QUESTION", profile["insider_questions"]), ("PRIMARY DOCUMENT", profile["primary_documents"]), ("FAILURE MODE", profile["failure_modes"])):
            for item in values:
                for column, value in enumerate((profile["id"], profile["domain"], profile["decision_arena"], profile["committee_artifact"], item_type, item, "OPEN"), 2):
                    decision_sheet.cell(row, column, value)
                row += 1
    decision_sheet.auto_filter.ref, decision_sheet.freeze_panes = f"B4:H{row - 1}", "B5"
    _body(decision_sheet, 5, row - 1, 2, 8)
    _widths(decision_sheet, {"A": 3, "B": 10, "C": 28, "D": 46, "E": 42, "F": 20, "G": 64, "H": 14})

    source_sheet = workbook.create_sheet("Source & Refresh")
    _title(source_sheet, "Data Lineage and Refresh Control", 9)
    _header(source_sheet, 4, ["Model ID", "Domain", "Source class", "Cadence", "Control", "Provider / system", "As-of", "Status"])
    row = 5
    for profile in profile_payload["profiles"]:
        for source in profile["source_classes"]:
            for column, value in enumerate((profile["id"], profile["domain"], source["name"], source["cadence"], source["control"], "", "", "TO MAP"), 2):
                source_sheet.cell(row, column, value)
            row += 1
    source_sheet.auto_filter.ref, source_sheet.freeze_panes = f"B4:I{row - 1}", "B5"
    _body(source_sheet, 5, row - 1, 2, 9)
    _widths(source_sheet, {"A": 3, "B": 10, "C": 28, "D": 26, "E": 18, "F": 54, "G": 28, "H": 16, "I": 14})

    scenario_sheet = workbook.create_sheet("Scenario Library")
    _title(scenario_sheet, "Cross-Domain Scenario and Challenge Library", 8)
    _header(scenario_sheet, 4, ["Model ID", "Domain", "Type", "Scenario / challenge", "Parameterization", "Owner", "Status"])
    row = 5
    for profile in profile_payload["profiles"]:
        for item_type, items in (("SCENARIO", profile["scenario_families"]), ("CHALLENGE", profile["challenge_tests"])):
            for item in items:
                for column, value in enumerate((profile["id"], profile["domain"], item_type, item, "", "", "OPEN"), 2):
                    scenario_sheet.cell(row, column, value)
                row += 1
    scenario_sheet.auto_filter.ref, scenario_sheet.freeze_panes = f"B4:H{row - 1}", "B5"
    _body(scenario_sheet, 5, row - 1, 2, 8)
    _widths(scenario_sheet, {"A": 3, "B": 10, "C": 28, "D": 18, "E": 60, "F": 36, "G": 20, "H": 14})

    benchmark_sheet = workbook.create_sheet("Public Cases")
    _title(benchmark_sheet, "Source-Addressed Public Evidence Cases", 10)
    _header(benchmark_sheet, 4, ["Case ID", "Model", "As of", "Template", "Output", "Manifest", "Applied inputs", "Workbook SHA-256", "M4 credit"])
    row = 5
    for item in public.get("cases", []):
        receipt = item.get("receipt") or {}
        values = [item.get("case_id"), item.get("model_id"), receipt.get("as_of"), receipt.get("template"), item.get("output"), item.get("manifest"), len(receipt.get("applied_inputs", [])), receipt.get("workbook_sha256"), "NO — HISTORICAL EVIDENCE"]
        for column, value in enumerate(values, 2):
            benchmark_sheet.cell(row, column, value)
        row += 1
    if row == 5:
        benchmark_sheet["B5"] = "No committed public-case index found."
    benchmark_sheet.freeze_panes = "B5"
    _body(benchmark_sheet, 5, max(5, row - 1), 2, 10)
    _widths(benchmark_sheet, {"A": 3, "B": 32, "C": 24, "D": 14, "E": 48, "F": 48, "G": 48, "H": 14, "I": 68, "J": 18})

    release_sheet = workbook.create_sheet("Release Evidence")
    _title(release_sheet, "Cryptographic Release Evidence", 11)
    _header(release_sheet, 4, ["Model ID", "Domain", "Maturity", "Workbook", "Builder", "Workbook SHA-256", "Builder SHA-256", "Formulas", "Sheets", "Audit status"])
    row = 5
    for evidence in release.get("models", []):
        audit = evidence.get("audit_findings", {})
        values = [evidence.get("model_id"), evidence.get("domain"), evidence.get("maturity"), evidence.get("workbook"), evidence.get("builder"), evidence.get("workbook_sha256"), evidence.get("builder_sha256"), evidence.get("formula_count"), evidence.get("sheet_count"), "PASS" if not audit.get("error") else "FAIL"]
        for column, value in enumerate(values, 2):
            release_sheet.cell(row, column, value)
        row += 1
    release_sheet.freeze_panes = "B5"
    _body(release_sheet, 5, max(5, row - 1), 2, 11)
    _widths(release_sheet, {"A": 3, "B": 10, "C": 28, "D": 12, "E": 48, "F": 48, "G": 68, "H": 68, "I": 12, "J": 10, "K": 14})

    roadmap = workbook.create_sheet("Maturity Roadmap")
    _title(roadmap, "Evidence-Gated Maturity Roadmap", 10)
    _header(roadmap, 4, ["Model ID", "Domain", "Current", "Next", "Mechanical gap", "External evidence gap", "Independent reviewer", "Target date", "Status"])
    row = 5
    for model in inventory["models"]:
        mechanical_gap = "Promote to M2: independent checks, scenarios and decision mechanics" if model["declared_maturity"] == "M1" else "M2 mechanics complete; do not inflate maturity without evidence"
        evidence_gap = "Public source snapshots; populated model card; independent validation; effective challenge; stakeholder sign-off; externally sourced cases"
        for column, value in enumerate((model["id"], model["domain"], model["declared_maturity"], model["target_maturity"], mechanical_gap, evidence_gap, "", "", "OPEN"), 2):
            roadmap.cell(row, column, value)
        row += 1
    roadmap.freeze_panes = "B5"
    _body(roadmap, 5, row - 1, 2, 10)
    _widths(roadmap, {"A": 3, "B": 10, "C": 28, "D": 12, "E": 12, "F": 58, "G": 72, "H": 24, "I": 16, "J": 14})

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output)
    return {"output": str(output), "models": len(inventory["models"]), "profiles": len(profiles), "public_instances": public.get("case_count", 0), "release_models": len(release_by_id), "sheets": workbook.sheetnames}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path("."))
    parser.add_argument("--output", type=Path, default=Path("00_Control_Plane/Finance_Model_Control_Plane.xlsx"))
    parser.add_argument("--report", type=Path, default=Path("control-plane-report.json"))
    args = parser.parse_args()
    root = args.root.resolve()
    output = args.output if args.output.is_absolute() else root / args.output
    report = build(root, output)
    report_path = args.report if args.report.is_absolute() else root / args.report
    report_path.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
