"""Validate the finance-segway model inventory and maturity claims.

The inventory is deliberately conservative: it distinguishes a correct skeleton
from a decision model, institutional underwriting model, and maintained
production system. CI should fail when the repository claims a maturity level
without the evidence required for that level.

Until this file's history, "evidence required" for M2+ meant a non-empty
reference_checks list of free-text strings on the inventory entry itself --
never cross-checked against anything. A model could declare
reference_checks: ["some_check_name"] whether or not that check existed, ran,
or passed. Three real verification tiers exist elsewhere in the repo and were
never consulted here:

  1. tools/verify_public_case_status.py -- opens every real public-case
     workbook, recalculates it via LibreOffice, and reads back its genuine
     computed Overall/Decision status. Covers all 24 domains
     (standards/public_cases/index.json has at least one case per model_id).
     This is the tier wired in below, since it's both the strongest evidence
     (a real recalculation, not a static claim) and the only one with
     complete, unambiguous model_id coverage.
  2. tools/legacy_engine_oracles.py / tools/domain_hardening_oracles.py --
     pure-Python oracles scoring JSON case inputs. Covers 15/24 model_ids
     with a clean ORACLES[model_id] mapping, but is redundant with tier 1
     for models both cover and doesn't reach the other 9 -- not required
     here, but see check_oracle_coverage() for an informational note.
  3. tools/verify_reference_calcs.py -- six named independent-oracle checks
     (Black-Scholes, bond duration, LBO sources & uses, VC waterfalls x2,
     BASE archetype integration). No clean check-to-model_id mapping exists
     (a "BASE archetype" check spans multiple domains), so it isn't wired
     into per-model gating here either -- it remains a standalone CI check.

Usage:
    python tools/validate_model_inventory.py
    python tools/validate_model_inventory.py --report /tmp/model-report.json
    python tools/validate_model_inventory.py --strict

    # M2+ requires real, recalculated public-case evidence by default (fail
    # closed) -- this needs LibreOffice (soffice) on PATH. For fast local
    # iteration on schema-only questions, skip it explicitly:
    python tools/validate_model_inventory.py --skip-public-case-verification

    # CI that already ran tools/verify_public_case_status.py can reuse its
    # report instead of recalculating a second time:
    python tools/validate_model_inventory.py --public-case-report public-case-status-report.json
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl

try:
    from tools import reference_check_registry
except (ImportError, ModuleNotFoundError):  # script-style execution
    import reference_check_registry

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INVENTORY = ROOT / "standards" / "model_inventory.json"
HARD_FAILURE_STATUSES = {"MISSING_WORKBOOK", "RECALC_FAILED", "NO_STATUS_FOUND"}
EXPECTED_MODEL_COUNT = 25
ALLOWED_MATURITY = {"M0", "M1", "M2", "M3", "M4"}
ALLOWED_HORIZONS = {
    "trading_intraday",
    "trading_daily",
    "short_term_1y",
    "corporate_5y",
    "corporate_10y",
    "long_term_10y",
    "long_term_30y",
    "long_term_40y",
    "perpetual",
}
ERROR_TOKENS = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}


@dataclass
class ModelResult:
    model_id: str
    domain: str
    maturity: str
    workbook: str
    formula_count: int = 0
    sheet_count: int = 0
    instance_count: int = 0
    errors: list[str] | None = None
    warnings: list[str] | None = None
    reference_check_binding: dict | None = None

    def __post_init__(self) -> None:
        self.errors = [] if self.errors is None else self.errors
        self.warnings = [] if self.warnings is None else self.warnings


def load_inventory(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data.get("models"), list):
        raise ValueError("inventory must contain a models list")
    return data


def load_public_case_verification(report_path: Path | None) -> dict[str, list[dict[str, Any]]]:
    """Return real public-case verification results indexed by model_id.

    If report_path is given, load a previously-generated
    tools/verify_public_case_status.py --report JSON (fast, no recalculation
    -- reuses CI that already ran it). Otherwise run verification directly,
    which recalculates every public-case workbook via LibreOffice and
    therefore requires soffice on PATH.
    """
    sys.path.insert(0, str(ROOT / "tools"))
    if report_path is not None:
        report = json.loads(report_path.read_text(encoding="utf-8"))
    else:
        from verify_public_case_status import verify as verify_public_cases

        report = verify_public_cases()
    by_model: dict[str, list[dict[str, Any]]] = {}
    for entry in report["results"]:
        model_id = entry.get("model_id")
        if model_id is not None:
            by_model.setdefault(model_id, []).append(entry)
    return by_model


def count_instances(folder: Path) -> int:
    if not folder.exists():
        return 0
    count = 0
    for path in folder.rglob("*.xlsx"):
        if path.name.startswith("_template") or path.name.startswith("~$"):
            continue
        count += 1
    return count


def inspect_workbook(path: Path, result: ModelResult) -> set[str]:
    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:  # noqa: BLE001
        result.errors.append(f"workbook cannot be opened: {exc}")
        return set()

    sheetnames = set(workbook.sheetnames)
    result.sheet_count = len(workbook.sheetnames)
    if "Cover" not in sheetnames:
        result.errors.append("missing Cover sheet")
    if "RefreshLog" not in sheetnames:
        result.errors.append("missing RefreshLog sheet")
    if getattr(workbook, "_external_links", []):
        result.errors.append("contains external workbook links")

    literal_errors: list[str] = []
    formula_count = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                elif isinstance(value, str) and value.strip() in ERROR_TOKENS:
                    literal_errors.append(f"{sheet.title}!{cell.coordinate}={value}")
    result.formula_count = formula_count
    if literal_errors:
        result.errors.append(f"literal Excel errors: {literal_errors[:5]}")
    return sheetnames


def validate_model(
    model: dict[str, Any],
    public_case_index: dict[str, list[dict[str, Any]]] | None,
) -> ModelResult:
    model_id = str(model.get("id", ""))
    domain = str(model.get("domain", ""))
    maturity = str(model.get("declared_maturity", ""))
    workbook_rel = str(model.get("workbook", ""))
    result = ModelResult(model_id, domain, maturity, workbook_rel)

    required_strings = ("id", "domain", "folder", "archetype", "workbook", "builder", "horizon")
    for field in required_strings:
        if not isinstance(model.get(field), str) or not model[field].strip():
            result.errors.append(f"missing or invalid {field}")

    if maturity not in ALLOWED_MATURITY:
        result.errors.append(f"invalid maturity {maturity!r}")
    if model.get("target_maturity") not in ALLOWED_MATURITY:
        result.errors.append(f"invalid target maturity {model.get('target_maturity')!r}")
    if model.get("horizon") not in ALLOWED_HORIZONS:
        result.errors.append(f"invalid horizon {model.get('horizon')!r}")

    engines = model.get("required_engines")
    perspectives = model.get("required_perspectives")
    reference_checks = model.get("reference_checks")
    if not isinstance(engines, list) or not engines:
        result.errors.append("required_engines must be a non-empty list")
    if not isinstance(perspectives, list) or not perspectives:
        result.errors.append("required_perspectives must be a non-empty list")
    if not isinstance(reference_checks, list):
        result.errors.append("reference_checks must be a list")

    folder = ROOT / str(model.get("folder", ""))
    workbook = ROOT / workbook_rel
    builder = ROOT / str(model.get("builder", ""))
    if not folder.is_dir():
        result.errors.append(f"domain folder missing: {folder.relative_to(ROOT)}")
    sheetnames: set[str] = set()
    if not workbook.is_file():
        result.errors.append(f"workbook missing: {workbook_rel}")
    else:
        sheetnames = inspect_workbook(workbook, result)
    if not builder.is_file():
        result.errors.append(f"builder missing: {builder.relative_to(ROOT)}")

    result.instance_count = count_instances(folder)

    if maturity in {"M1", "M2", "M3", "M4"}:
        if result.formula_count == 0:
            result.errors.append("M1+ requires at least one formula")
        if result.sheet_count < 2:
            result.errors.append("M1+ requires multiple worksheets")

    if maturity in {"M2", "M3", "M4"}:
        if not reference_checks:
            result.errors.append("M2+ requires at least one independent reference check")
        if result.formula_count < 20:
            result.errors.append("M2+ requires at least 20 formula cells")
        if len(engines or []) < 3:
            result.errors.append("M2+ requires at least three documented domain engines")
        if public_case_index is not None:
            cases = public_case_index.get(model_id, [])
            if not cases:
                result.errors.append(
                    "M2+ requires at least one real public case in "
                    "standards/public_cases/index.json with a genuine, "
                    "recalculated Overall/Decision status -- reference_checks "
                    "naming a check is not itself evidence that check ran or passed"
                )
            elif all(case.get("status") in HARD_FAILURE_STATUSES for case in cases):
                result.errors.append(
                    "M2+ requires at least one public case whose workbook actually "
                    f"recalculates cleanly and reports a real status; found: "
                    f"{[case.get('status') for case in cases]}"
                )

    if maturity in {"M3", "M4"}:
        if len(perspectives or []) < 3:
            result.errors.append("M3+ requires at least three stakeholder perspectives")
        if "Sources" not in sheetnames:
            result.errors.append("M3+ requires a Sources sheet")
        if "Checks" not in sheetnames:
            result.errors.append("M3+ requires a Checks sheet")

    if maturity == "M4":
        minimum_instances = int(model.get("minimum_instances_for_M4", 2))
        if result.instance_count < minimum_instances:
            result.errors.append(
                f"M4 requires at least {minimum_instances} populated instances; found {result.instance_count}"
            )

    if model.get("target_maturity") == maturity:
        result.warnings.append("declared maturity already equals target; confirm next promotion target")

    return result


def validate_inventory(
    data: dict[str, Any],
    public_case_index: dict[str, list[dict[str, Any]]] | None = None,
) -> tuple[list[ModelResult], list[str]]:
    models = data["models"]
    inventory_errors: list[str] = []
    if len(models) != EXPECTED_MODEL_COUNT:
        inventory_errors.append(f"expected {EXPECTED_MODEL_COUNT} core model domains; found {len(models)}")

    ids = [str(model.get("id", "")) for model in models]
    folders = [str(model.get("folder", "")) for model in models]
    workbooks = [str(model.get("workbook", "")) for model in models]
    for label, values in (("id", ids), ("folder", folders), ("workbook", workbooks)):
        duplicates = sorted(value for value, count in Counter(values).items() if count > 1)
        if duplicates:
            inventory_errors.append(f"duplicate {label} values: {duplicates}")

    results = [validate_model(model, public_case_index) for model in models]

    # Bind every declared reference_checks token to the oracle that executes
    # it (or record that none does). A bound identity check that fails its
    # oracle blocks the gate; a declared token no oracle produces is a
    # warning - the claim exists but nothing runs it.
    try:
        bindings = reference_check_registry.resolve_reference_checks(models)
    except Exception as exc:  # noqa: BLE001
        inventory_errors.append(f"reference-check binding pass could not run: {exc}")
        bindings = {}
    for result in results:
        binding = bindings.get(result.model_id)
        if binding is None:
            continue
        result.reference_check_binding = binding
        if result.maturity in {"M2", "M3", "M4"}:
            for token in binding["identity_failures"]:
                result.errors.append(
                    f"reference check {token!r} FAILED its oracle identity"
                )
            unbound = sorted(
                token
                for token, status in binding["tokens"].items()
                if status == "unbound"
            )
            if unbound:
                result.warnings.append(
                    "reference checks declared but bound to no oracle: "
                    + ", ".join(unbound)
                )

    return results, inventory_errors


def print_summary(results: list[ModelResult], inventory_errors: list[str]) -> None:
    print("Finance-Segway Model Inventory")
    print("=" * 88)
    print(f"{'ID':<4} {'Domain':<36} {'Mat':<4} {'Sheets':>6} {'Formulas':>9} {'Inst':>5} Status")
    print("-" * 88)
    for result in results:
        status = "FAIL" if result.errors else ("WARN" if result.warnings else "PASS")
        print(
            f"{result.model_id:<4} {result.domain[:36]:<36} {result.maturity:<4} "
            f"{result.sheet_count:>6} {result.formula_count:>9} {result.instance_count:>5} {status}"
        )
        for error in result.errors:
            print(f"     ERROR: {error}")
        for warning in result.warnings:
            print(f"     WARN:  {warning}")
    for error in inventory_errors:
        print(f"INVENTORY ERROR: {error}")

    maturity_counts = Counter(result.maturity for result in results)
    print("-" * 88)
    print("Maturity distribution:", ", ".join(f"{k}={maturity_counts[k]}" for k in sorted(maturity_counts)))

    bindings = {
        result.model_id: result.reference_check_binding
        for result in results
        if result.reference_check_binding is not None
    }
    if bindings:
        coverage = reference_check_registry.coverage_summary(bindings)
        print(
            "Reference-check binding: "
            f"{coverage['identity']} identity-verified, "
            f"{coverage['flag']} flag-exercised, "
            f"{coverage['unbound']} unbound "
            f"of {coverage['declared']} declared tokens; "
            f"{coverage['models_with_oracle']}/24 models oracle-backed, "
            f"{coverage['models_workbook_verified']}/24 workbook-verified"
        )


def write_report(path: Path, results: list[ModelResult], inventory_errors: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    bindings = {
        result.model_id: result.reference_check_binding
        for result in results
        if result.reference_check_binding is not None
    }
    payload = {
        "inventory_errors": inventory_errors,
        "models": [asdict(result) for result in results],
        "summary": {
            "models": len(results),
            "failed": sum(bool(result.errors) for result in results),
            "warnings": sum(bool(result.warnings) for result in results),
            "maturity_distribution": dict(Counter(result.maturity for result in results)),
            "reference_check_coverage": (
                reference_check_registry.coverage_summary(bindings) if bindings else None
            ),
        },
    }
    path.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inventory", type=Path, default=DEFAULT_INVENTORY)
    parser.add_argument("--report", type=Path)
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Treat warnings as failures. Useful for release promotion, not routine PR checks.",
    )
    parser.add_argument(
        "--public-case-report", type=Path, default=None,
        help=(
            "Reuse a JSON report already produced by "
            "tools/verify_public_case_status.py --report instead of recalculating "
            "every public case again."
        ),
    )
    parser.add_argument(
        "--skip-public-case-verification", action="store_true",
        help=(
            "Skip cross-checking M2+ claims against real, recalculated public-case "
            "evidence (falls back to the old reference_checks string-presence check "
            "only). For fast local iteration on schema-only questions -- not for CI."
        ),
    )
    args = parser.parse_args()

    try:
        data = load_inventory(args.inventory)
        public_case_index = (
            None
            if args.skip_public_case_verification
            else load_public_case_verification(args.public_case_report)
        )
        results, inventory_errors = validate_inventory(data, public_case_index)
    except Exception as exc:  # noqa: BLE001
        print(f"inventory validation could not start: {exc}", file=sys.stderr)
        return 2

    print_summary(results, inventory_errors)
    if args.report:
        write_report(args.report, results, inventory_errors)

    failed = bool(inventory_errors) or any(result.errors for result in results)
    if args.strict:
        failed = failed or any(result.warnings for result in results)
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
