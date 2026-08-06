"""Generate, validate, and fingerprint maintained model instances.

An instance manifest is the machine-readable source of truth for workbook inputs,
source provenance, Cover metadata, and RefreshLog entries. The generator refuses
to overwrite formulas unless the manifest explicitly opts in, records every
changed cell, and writes a deterministic release receipt beside the workbook.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

BLUE = "0000FF"
YELLOW = "FFFF00"


@dataclass(frozen=True)
class AppliedInput:
    sheet: str
    cell: str
    previous: Any
    value: Any
    source_url: str | None
    source_as_of: str | None


def _load_manifest(path: Path) -> dict[str, Any]:
    manifest = json.loads(path.read_text(encoding="utf-8"))
    required = {"schema_version", "id", "template", "output", "as_of", "inputs"}
    missing = sorted(required - set(manifest))
    if missing:
        raise ValueError(f"manifest missing required keys: {missing}")
    if not isinstance(manifest["inputs"], list) or not manifest["inputs"]:
        raise ValueError("manifest inputs must be a nonempty list")
    return manifest


# The Cover label that honestly holds a case subject, per model. Domains
# with no label:value identity row (their identity slot is the Cover title
# itself) map to the reserved "Title:" key, which _set_cover writes to the
# title cell; unknown/new models default to "Title:" since every template
# carries a B2 title.
SUBJECT_COVER_LABELS = {
    "03": "Target / transaction:", "04": "Target / transaction:",
    "05": "Borrower / issuer:", "06": "Issuer:", "07": "Issuer:",
    "10": "Counterparty:", "11": "Institution:", "12": "Issuer / security:",
    "14": "Underlying:", "15": "Commodity:", "18": "Entity / line of business:",
    "19": "Transaction / collateral:", "20": "Project / concession:",
    "21": "Portfolio / security:", "22": "Strategy / universe:",
    "24": "Situation type:",
}


def subject_cover(model_id: str, subject: str) -> dict[str, str]:
    """The cover mapping a generated manifest should carry for its subject."""
    return {SUBJECT_COVER_LABELS.get(str(model_id), "Title:"): subject}


def _find_label_row(sheet: Any, label: str, *, label_column: int = 2) -> int | None:
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row, label_column).value or "").strip() == label:
            return row
    return None


def _set_cover(workbook: Any, manifest: dict[str, Any]) -> None:
    if "Cover" not in workbook.sheetnames:
        raise ValueError("template has no Cover sheet")
    cover = workbook["Cover"]
    # Two tiers: infrastructure defaults are best-effort (several templates
    # have no "Active scenario:" row - there is nothing to set and skipping
    # is correct), while keys the MANIFEST asserts stay fail-loud below.
    defaults = {
        "Last refreshed:": manifest["as_of"],
        "Active scenario:": manifest.get("scenario", "Base"),
    }
    asserted = manifest.get("cover", {})
    mapping = {**defaults, **asserted}
    for label, value in mapping.items():
        if label in defaults and label not in asserted and _find_label_row(cover, label) is None:
            continue
        if label == "Title:":
            # Reserved key: several templates (e.g. Asset Management, Risk,
            # Crypto, REIT, Fintech) have no label:value row that honestly
            # holds a case subject - their identity slot is the Cover title
            # itself (row 2, column B, the "[TICKER] - Company Name"
            # placeholder). "Title:" writes that cell in place, keeping its
            # existing title styling, instead of forcing the subject under a
            # semantically wrong label.
            if not str(cover.cell(2, 2).value or "").strip():
                raise ValueError(
                    "manifest cover key 'Title:' but the Cover sheet has no "
                    "title text in B2 to replace"
                )
            cover.cell(2, 2, value)
            continue
        row = _find_label_row(cover, label)
        if row is None:
            # A manifest cover key that matches no row silently drops that
            # value -- the template keeps its literal placeholder (e.g.
            # "[Name]") in a workbook that is otherwise real and sourced.
            # Fail loudly instead: this exact silent-drop class of bug left
            # every public case's Cover sheet showing template placeholder
            # text instead of the real subject, undetected, for the whole
            # life of the public-evidence program.
            raise ValueError(
                f"manifest cover key {label!r} matches no Cover sheet row -- "
                "check the label against the template's actual Cover field text"
            )
        cover.cell(row, 3, value)
        cover.cell(row, 3).font = Font(name="Arial", size=10, color=BLUE)
        cover.cell(row, 3).fill = PatternFill("solid", fgColor=YELLOW)


def _append_sources(workbook: Any, manifest: dict[str, Any]) -> None:
    if "Sources" not in workbook.sheetnames:
        return
    sources = workbook["Sources"]
    rows = list(manifest.get("sources", []))
    for item in manifest.get("inputs", []):
        source = item.get("source") or {}
        if source.get("url"):
            rows.append({
                "name": source.get("name") or f"{item['sheet']}!{item['cell']}",
                "url": source.get("url"),
                "as_of": source.get("as_of") or manifest["as_of"],
                "notes": source.get("notes") or "Manifest input source",
            })
    seen = set()
    unique_rows = []
    for item in rows:
        key = (item.get("name"), item.get("url"), item.get("as_of"))
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append(item)
    row = max(5, sources.max_row + 1)
    for item in unique_rows:
        values = [item.get("name"), item.get("url"), item.get("as_of"), item.get("notes")]
        for column, value in enumerate(values, start=2):
            cell = sources.cell(row, column, value)
            cell.font = Font(name="Arial", size=10, color=BLUE)
        row += 1


def _append_refresh(workbook: Any, manifest: dict[str, Any]) -> None:
    if "RefreshLog" not in workbook.sheetnames:
        return
    refresh = workbook["RefreshLog"]
    entry = manifest.get("refresh") or {
        "date": manifest["as_of"],
        "trigger": "Initial instance generation",
        "what_changed": f"Applied manifest {manifest['id']}",
        "reviewer_notes": "Generated reproducibly from machine-readable inputs",
        "next_check": manifest.get("next_check", "[date]"),
    }
    row = 5
    while row <= max(refresh.max_row + 1, 1000) and any(
        refresh.cell(row, column).value not in (None, "") for column in range(2, 7)
    ):
        row += 1
    values = [
        entry.get("date", manifest["as_of"]),
        entry.get("trigger"),
        entry.get("what_changed"),
        entry.get("reviewer_notes"),
        entry.get("next_check"),
    ]
    for column, value in enumerate(values, start=2):
        refresh.cell(row, column, value)


def apply_manifest(manifest_path: Path, root: Path, *, validate_only: bool = False) -> dict[str, Any]:
    manifest = _load_manifest(manifest_path)
    template = (root / manifest["template"]).resolve()
    output = (root / manifest["output"]).resolve()
    if not template.exists():
        raise FileNotFoundError(template)
    workbook = load_workbook(template, data_only=False, keep_links=True)
    applied: list[AppliedInput] = []
    for item in manifest["inputs"]:
        sheet_name = item["sheet"]
        address = item["cell"]
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"input references missing sheet {sheet_name}")
        cell = workbook[sheet_name][address]
        previous = cell.value
        if isinstance(previous, str) and previous.startswith("=") and not item.get("allow_formula_override", False):
            raise ValueError(f"refusing to overwrite formula {sheet_name}!{address}")
        source = item.get("source") or {}
        applied.append(AppliedInput(
            sheet_name, address, previous, item.get("value"),
            source.get("url"), source.get("as_of"),
        ))
        if not validate_only:
            cell.value = item.get("value")
            cell.font = Font(name="Arial", size=10, color=BLUE)
            cell.fill = PatternFill("solid", fgColor=YELLOW)
    if validate_only:
        return {"manifest": str(manifest_path), "template": str(template), "inputs": len(applied), "valid": True}
    _set_cover(workbook, manifest)
    _append_sources(workbook, manifest)
    _append_refresh(workbook, manifest)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    receipt = {
        "schema_version": "1.0",
        "instance_id": manifest["id"],
        "as_of": manifest["as_of"],
        "manifest": str(manifest_path.relative_to(root)),
        "template": manifest["template"],
        "output": manifest["output"],
        "applied_inputs": [asdict(item) for item in applied],
        "workbook_sha256": hashlib.sha256(output.read_bytes()).hexdigest(),
        "generated_on": date.today().isoformat(),
    }
    receipt_path = output.with_suffix(".receipt.json")
    receipt_path.write_text(json.dumps(receipt, indent=2, default=str) + "\n", encoding="utf-8")
    return receipt


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("manifest", type=Path)
    parser.add_argument("--root", type=Path, default=Path("."))
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args(argv)
    root = args.root.resolve()
    receipt = apply_manifest(args.manifest.resolve(), root, validate_only=args.validate_only)
    print(json.dumps(receipt, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
