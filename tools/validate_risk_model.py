"""Validate the institutional risk-management workbook contract."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from workbook_engineering import audit_workbook

REQUIRED_SHEETS = {
    "Cover", "Assumptions", "Positions", "Factors", "VaR & ES",
    "Risk Contributions", "Stress", "Liquidity", "P&L Explain", "Limits",
    "Checks", "Sources", "RefreshLog",
}
REQUIRED_LABELS = {
    "Position-Level Risk Inventory",
    "Factor Exposure and Covariance",
    "One-day expected shortfall",
    "Position-Level VaR, Stress, and Liquidity Contributions",
    "Component VaR reconciliation",
    "Largest risk concentration",
    "Portfolio Stress Matrix",
    "Total modeled liquidation loss",
    "Unexplained P&L",
    "Risk Limit Dashboard",
}


def validate(path: Path) -> list[str]:
    workbook = load_workbook(path, data_only=False, keep_links=True)
    errors: list[str] = []
    missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing:
        errors.append(f"missing sheets: {missing}")
    values = {
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row if cell.value is not None
    }
    missing_labels = sorted(REQUIRED_LABELS - values)
    if missing_labels:
        errors.append(f"missing labels: {missing_labels}")
    formulas = [
        cell.value
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    if len(formulas) < 300:
        errors.append(f"formula depth below contract: {len(formulas)} < 300")
    contributions = workbook["Risk Contributions"] if "Risk Contributions" in workbook.sheetnames else None
    if contributions is not None:
        if contributions["E18"].value != "=SUM(E5:E16)":
            errors.append("component VaR portfolio total is not explicitly summed")
        if contributions["C20"].value != "=E18-'VaR & ES'!C8":
            errors.append("component VaR reconciliation formula missing")
        if contributions["J18"].value != "=SUM(J5:J16)":
            errors.append("risk concentration does not sum to portfolio total")
    summary, findings = audit_workbook(path)
    if summary["external_links"]:
        errors.append("external links present")
    errors.extend(
        f"{finding.code}:{finding.sheet}:{finding.cell}:{finding.message}"
        for finding in findings if finding.severity == "error"
    )
    if "Checks" in workbook.sheetnames:
        checks = [
            cell.value for row in workbook["Checks"].iter_rows() for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
        if len(checks) < 9:
            errors.append("insufficient risk checks")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    errors = validate(args.workbook)
    if errors:
        for error in errors:
            print(f"ERROR: {error}")
        return 1
    print(f"validated {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
