"""Validate deep builders and the integrity of committed workbook artifacts.

The builder output is the canonical next model release and must satisfy the full
institutional workbook contract. Committed XLSX files are reviewed release
artifacts and must remain genuine, auditable, formula-driven workbooks while
parallel-agent synthesis is in progress.
"""
from __future__ import annotations

import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BUILDERS = ROOT / "tools" / "builders"
ERROR_TOKENS = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}


@dataclass(frozen=True)
class ModelCase:
    name: str
    builder: Path
    committed: tuple[Path, ...]
    required_tabs: tuple[str, ...]
    formula_anchors: frozenset[str]
    minimum_formula_cells: int


CASES = (
    ModelCase(
        "credit",
        BUILDERS / "build_credit_template.py",
        (
            ROOT / "05_Private_Credit" / "_template_CREDIT.xlsx",
            ROOT / "06_Debt_Finance" / "_template_CREDIT.xlsx",
        ),
        (
            "Cover", "Assumptions", "Operating Case", "Debt Schedule", "Covenants",
            "Yield & Spread", "Recovery", "Sensitivity", "Checks", "Sources", "RefreshLog",
        ),
        frozenset({
            "Assumptions!E5", "Operating Case!D5", "Operating Case!H14",
            "Debt Schedule!D15", "Debt Schedule!H18", "Covenants!C14",
            "Yield & Spread!C9", "Recovery!C12", "Sensitivity!G9", "Checks!C9",
        }),
        220,
    ),
    ModelCase(
        "public_finance",
        BUILDERS / "build_public_finance_template.py",
        (ROOT / "07_Public_Finance" / "_template_PUBLIC_FINANCE.xlsx",),
        (
            "Cover", "Assumptions", "Debt Sustainability", "Revenue & Expenditure",
            "Debt Service", "Coverage", "Scenarios", "Sensitivity", "Checks", "Sources", "RefreshLog",
        ),
        frozenset({
            "Assumptions!E5", "Debt Sustainability!C10", "Debt Sustainability!C12",
            "Revenue & Expenditure!D5", "Revenue & Expenditure!H14",
            "Debt Service!C11", "Debt Service!G15", "Coverage!C15",
            "Scenarios!C10", "Sensitivity!G9", "Checks!C10",
        }),
        220,
    ),
)


def scan_formulas(workbook) -> tuple[dict[str, str], list[str]]:
    formulas: dict[str, str] = {}
    errors: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formulas[f"{sheet.title}!{cell.coordinate}"] = value
                elif isinstance(value, str) and value.strip() in ERROR_TOKENS:
                    errors.append(f"{sheet.title}!{cell.coordinate}={value}")
    return formulas, errors


def validate_generated(path: Path, case: ModelCase) -> int:
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    if tuple(workbook.sheetnames) != case.required_tabs:
        raise AssertionError(f"{path}: wrong sheet contract: {workbook.sheetnames}")
    if workbook["Cover"]["C6"].value in (None, ""):
        raise AssertionError(f"{path}: Cover!C6 must hold Last refreshed")
    if workbook["Cover"]["C7"].value in (None, ""):
        raise AssertionError(f"{path}: Cover!C7 must hold Next material date")
    if workbook["Cover"]["C9"].value not in {"Base", "Downside"}:
        raise AssertionError(f"{path}: Cover!C9 must be Base or Downside")
    if getattr(workbook, "_external_links", []):
        raise AssertionError(f"{path}: external workbook links are not allowed")
    formulas, errors = scan_formulas(workbook)
    if errors:
        raise AssertionError(f"{path}: literal Excel errors: {errors[:10]}")
    if len(formulas) < case.minimum_formula_cells:
        raise AssertionError(f"{path}: only {len(formulas)} formulas")
    missing = case.formula_anchors - formulas.keys()
    if missing:
        raise AssertionError(f"{path}: missing formula anchors {sorted(missing)}")
    return len(formulas)


def validate_committed_artifact(path: Path) -> int:
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    missing = {"Cover", "RefreshLog"} - set(workbook.sheetnames)
    if missing:
        raise AssertionError(f"{path}: missing universal tabs {sorted(missing)}")
    if getattr(workbook, "_external_links", []):
        raise AssertionError(f"{path}: external workbook links are not allowed")
    formulas, errors = scan_formulas(workbook)
    if errors:
        raise AssertionError(f"{path}: literal Excel errors: {errors[:10]}")
    if not formulas:
        raise AssertionError(f"{path}: workbook contains no formulas")
    return len(formulas)


def build(builder: Path, output: Path) -> None:
    subprocess.run(
        [sys.executable, str(builder), "--output", str(output)],
        cwd=BUILDERS,
        check=True,
    )


def main() -> int:
    failures: list[str] = []
    with tempfile.TemporaryDirectory(prefix="finance-model-validation-") as temp_dir:
        temp = Path(temp_dir)
        for case in CASES:
            try:
                generated = temp / f"{case.name}.xlsx"
                build(case.builder, generated)
                formula_count = validate_generated(generated, case)
                artifact_counts = [validate_committed_artifact(path) for path in case.committed]
                print(
                    f"PASS {case.name}: generated contract has {formula_count} formulas; "
                    f"committed artifacts are valid ({artifact_counts})"
                )
            except Exception as exc:  # noqa: BLE001 - aggregate all model failures
                failures.append(f"{case.name}: {exc}")

    if failures:
        print("\nIntegrated-model validation failed:")
        for failure in failures:
            print(f"- {failure}")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
