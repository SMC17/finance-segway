"""Load committed, source-addressed LBO/PE workbook outputs into Postgres.

Excel remains the calculation engine. This optional query layer reads cached
``data_only`` values and never recalculates or mutates a workbook.

Usage::

    python tools/postgres_etl.py --dry-run
    python tools/postgres_etl.py --dsn "dbname=finance_segway"
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path
from typing import Any

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[1]
REAL_PUBLIC = "real_public"

# Registry entries must be committed public cases. Synthetic or generated
# benchmark paths are rejected by validate_registry before workbook access.
DEALS: dict[str, dict[str, str]] = {
    "home_depot_2023": {
        "path": "03_Private_Equity/instances/public_home_depot_2023.xlsx",
        "classification": REAL_PUBLIC,
    },
    "macys_2020_adversarial": {
        "path": "03_Private_Equity/instances/public_macys_2020_adversarial.xlsx",
        "classification": REAL_PUBLIC,
    },
    # 04_Merchant_Banking shares 03_Private_Equity's exact LBO archetype
    # (Cover/Assumptions/Sources & Uses/Debt Schedule/Returns Waterfall/
    # Checks, same labels) -- merchant banking is a principal-investing
    # variant of the same engine, not a different schema. Extraction here
    # is label-driven, so these load through the identical extract_* path
    # with no new code.
    "alleghany_2021": {
        "path": "04_Merchant_Banking/instances/public_alleghany_2021.xlsx",
        "classification": REAL_PUBLIC,
    },
    "wework_2022_adversarial": {
        "path": "04_Merchant_Banking/instances/public_wework_2022_adversarial.xlsx",
        "classification": REAL_PUBLIC,
    },
}

REQUIRED_SHEETS = {
    "Cover",
    "Assumptions",
    "Sources & Uses",
    "Debt Schedule",
    "Returns Waterfall",
    "Checks",
}


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def validate_registry(registry: dict[str, dict[str, str]] = DEALS) -> None:
    """Fail closed if the ETL registry is not real-only and source-addressed."""

    if not registry:
        raise ValueError("deal registry must not be empty")
    seen_paths: set[str] = set()
    for deal_id, meta in registry.items():
        if not deal_id or not deal_id.replace("_", "").isalnum():
            raise ValueError(f"invalid deal id: {deal_id!r}")
        if set(meta) != {"path", "classification"}:
            raise ValueError(f"{deal_id}: registry fields must be path and classification")
        path = meta["path"]
        lowered = path.lower()
        if meta["classification"] != REAL_PUBLIC:
            raise ValueError(f"{deal_id}: only {REAL_PUBLIC!r} is accepted")
        if "benchmark" in lowered or "synthetic" in lowered:
            raise ValueError(f"{deal_id}: synthetic lineage is forbidden: {path}")
        if not lowered.endswith(".xlsx") or "/instances/public_" not in lowered:
            raise ValueError(f"{deal_id}: expected a public instance workbook: {path}")
        resolved = (REPO_ROOT / path).resolve()
        if not resolved.is_relative_to(REPO_ROOT):
            raise ValueError(f"{deal_id}: workbook escapes repository root")
        if path in seen_paths:
            raise ValueError(f"duplicate workbook path: {path}")
        seen_paths.add(path)


def read_labeled_table(
    worksheet,
    *,
    header_row: int,
    label_col: int = 2,
    first_data_col: int = 3,
    last_data_col: int | None = None,
) -> dict[str, dict[str, Any]]:
    """Read a row-label x column-header table without fixed data cells."""

    last_data_col = last_data_col or worksheet.max_column
    headers: dict[int, str] = {}
    for column in range(first_data_col, last_data_col + 1):
        value = worksheet.cell(row=header_row, column=column).value
        if value not in (None, ""):
            header = str(value).strip()
            if header in headers.values():
                raise ValueError(f"{worksheet.title}: duplicate header {header!r}")
            headers[column] = header
    if not headers:
        raise ValueError(f"{worksheet.title}: no table headers found")

    table: dict[str, dict[str, Any]] = {}
    for row in range(header_row + 1, worksheet.max_row + 1):
        value = worksheet.cell(row=row, column=label_col).value
        if value in (None, ""):
            continue
        label = str(value).strip()
        if label in table:
            raise ValueError(f"{worksheet.title}: duplicate row label {label!r}")
        table[label] = {
            header: worksheet.cell(row=row, column=column).value
            for column, header in headers.items()
        }
    if not table:
        raise ValueError(f"{worksheet.title}: no labeled rows found")
    return table


def period_order(period_label: Any) -> int:
    if period_label == "Close":
        return 0
    if isinstance(period_label, str) and period_label.startswith("Year "):
        try:
            return int(period_label.split(maxsplit=1)[1])
        except (IndexError, ValueError):
            pass
    return 99


def extract_cover(workbook) -> dict[str, Any]:
    cover = workbook["Cover"]
    deal_name = cover["C4"].value
    if not deal_name:
        raise ValueError("Cover!C4 deal name is required")
    return {
        "deal_name": deal_name,
        "sponsor": cover["C5"].value,
        "entry_date": (
            str(cover["C6"].value) if cover["C6"].value is not None else None
        ),
        "active_scenario": cover["C9"].value,
    }


def extract_overall_check(workbook) -> Any:
    worksheet = workbook["Checks"]
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=2).value == "Overall":
            return worksheet.cell(row=row, column=3).value
    raise ValueError("Checks sheet has no Overall row")


def extract_assumptions(workbook) -> list[dict[str, Any]]:
    table = read_labeled_table(
        workbook["Assumptions"],
        header_row=4,
        label_col=2,
        first_data_col=3,
        last_data_col=6,
    )
    return [
        {
            "assumption": label,
            "base": columns.get("Base"),
            "downside": columns.get("Downside"),
            "active": columns.get("Active"),
            "units": columns.get("Units / note"),
        }
        for label, columns in table.items()
    ]


def extract_sources_uses(workbook) -> list[tuple[str, str, float]]:
    worksheet = workbook["Sources & Uses"]
    header_row = next(
        (
            row
            for row in range(1, worksheet.max_row + 1)
            if worksheet.cell(row=row, column=2).value == "Uses"
            and worksheet.cell(row=row, column=5).value == "Sources"
        ),
        None,
    )
    if header_row is None:
        raise ValueError("Sources & Uses sheet has no paired section header")

    rows: list[tuple[str, str, float]] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        for side, label_column, amount_column in (
            ("Uses", 2, 3),
            ("Sources", 5, 6),
        ):
            label = worksheet.cell(row=row, column=label_column).value
            amount = worksheet.cell(row=row, column=amount_column).value
            lowered = str(label).lower() if label else ""
            excluded = lowered.startswith("total") or lowered.startswith("sources less")
            if label and is_number(amount) and not excluded:
                rows.append((side, str(label), float(amount)))
    if not rows:
        raise ValueError("Sources & Uses sheet yielded no numeric line items")
    return rows


def extract_metric_table(workbook, sheet_name: str) -> list[tuple[str, int, str, float]]:
    table = read_labeled_table(
        workbook[sheet_name], header_row=4, label_col=2, first_data_col=3
    )
    rows: list[tuple[str, int, str, float]] = []
    for metric, columns in table.items():
        for period, value in columns.items():
            if is_number(value):
                rows.append((str(period), period_order(period), metric, float(value)))
    if not rows:
        raise ValueError(f"{sheet_name}: no numeric outputs found")
    return rows


def extract_returns(workbook) -> list[tuple[str, bool, str, float]]:
    rows = extract_metric_table(workbook, "Returns Waterfall")
    return [(period, period == "Selected exit", metric, value) for period, _, metric, value in rows]


def extract_deal(deal_id: str, path: str) -> dict[str, Any]:
    workbook_path = (REPO_ROOT / path).resolve()
    if not workbook_path.is_relative_to(REPO_ROOT) or not workbook_path.is_file():
        raise FileNotFoundError(f"{deal_id}: registered workbook not found: {path}")
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        missing = REQUIRED_SHEETS.difference(workbook.sheetnames)
        if missing:
            raise ValueError(f"{deal_id}: missing sheets: {sorted(missing)}")
        return {
            "deal_id": deal_id,
            "path": path,
            "cover": extract_cover(workbook),
            "overall_check": extract_overall_check(workbook),
            "assumptions": extract_assumptions(workbook),
            "sources_uses": extract_sources_uses(workbook),
            "debt_schedule": extract_metric_table(workbook, "Debt Schedule"),
            "returns": extract_returns(workbook),
        }
    finally:
        workbook.close()


def load_deal(cursor, deal_id: str, payload: dict[str, Any]) -> None:
    cover = payload["cover"]
    cursor.execute(
        """
        INSERT INTO deals (deal_id, deal_name, classification, sponsor, entry_date,
                           active_scenario, overall_check_status, workbook_path,
                           last_synced_at)
        VALUES (%s, %s, 'real_public', %s, %s, %s, %s, %s, now())
        ON CONFLICT (deal_id) DO UPDATE SET
            deal_name = EXCLUDED.deal_name,
            classification = 'real_public',
            sponsor = EXCLUDED.sponsor,
            entry_date = EXCLUDED.entry_date,
            active_scenario = EXCLUDED.active_scenario,
            overall_check_status = EXCLUDED.overall_check_status,
            workbook_path = EXCLUDED.workbook_path,
            last_synced_at = now()
        """,
        (
            deal_id,
            cover["deal_name"],
            cover["sponsor"],
            cover["entry_date"],
            cover["active_scenario"],
            payload["overall_check"],
            payload["path"],
        ),
    )

    cursor.execute("DELETE FROM deal_assumptions WHERE deal_id = %s", (deal_id,))
    for row in payload["assumptions"]:
        cursor.execute(
            """INSERT INTO deal_assumptions
               (deal_id, assumption, base, downside, active, units)
               VALUES (%s, %s, %s, %s, %s, %s)""",
            (
                deal_id,
                row["assumption"],
                row["base"],
                row["downside"],
                row["active"],
                row["units"],
            ),
        )

    cursor.execute("DELETE FROM deal_sources_uses WHERE deal_id = %s", (deal_id,))
    for side, label, amount in payload["sources_uses"]:
        cursor.execute(
            """INSERT INTO deal_sources_uses
               (deal_id, side, line_item, amount) VALUES (%s, %s, %s, %s)""",
            (deal_id, side, label, amount),
        )

    cursor.execute("DELETE FROM deal_debt_schedule WHERE deal_id = %s", (deal_id,))
    for period, order, metric, value in payload["debt_schedule"]:
        cursor.execute(
            """INSERT INTO deal_debt_schedule
               (deal_id, period, period_order, metric, value)
               VALUES (%s, %s, %s, %s, %s)""",
            (deal_id, period, order, metric, value),
        )

    cursor.execute("DELETE FROM deal_returns WHERE deal_id = %s", (deal_id,))
    for period, selected, metric, value in payload["returns"]:
        cursor.execute(
            """INSERT INTO deal_returns
               (deal_id, period, is_selected_exit, metric, value)
               VALUES (%s, %s, %s, %s, %s)""",
            (deal_id, period, selected, metric, value),
        )


def selected_return_summary(payload: dict[str, Any]) -> dict[str, float]:
    return {
        metric: value
        for period, selected, metric, value in payload["returns"]
        if selected
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dsn",
        default=os.environ.get("FINANCE_SEGWAY_PG_DSN", "dbname=finance_segway"),
    )
    parser.add_argument("--dry-run", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    validate_registry()
    payloads = {
        deal_id: extract_deal(deal_id, meta["path"])
        for deal_id, meta in DEALS.items()
    }

    if args.dry_run:
        for deal_id, payload in payloads.items():
            returns = selected_return_summary(payload)
            print(
                f"{deal_id}: sponsor MOIC={returns.get('Sponsor MOIC')}, "
                f"sponsor IRR={returns.get('Sponsor IRR')}, "
                f"overall check={payload['overall_check']}"
            )
        print("Dry run complete (no DB writes).")
        return

    import psycopg2

    with psycopg2.connect(args.dsn) as connection:
        with connection.cursor() as cursor:
            # Delete rows for retired registry entries, including any legacy
            # synthetic rows left by the superseded four-case pilot.
            cursor.execute(
                "DELETE FROM deals WHERE NOT (deal_id = ANY(%s))",
                (list(DEALS),),
            )
            for deal_id, payload in payloads.items():
                load_deal(cursor, deal_id, payload)
    print(f"Loaded {len(payloads)} real-public deals into Postgres.")


if __name__ == "__main__":
    main()
