"""Score every domain against machine-checkable M3 exit criteria.

M3 ("Institutional Underwriting") has until now existed only as prose in
docs/MODEL_GOVERNANCE_STANDARD.md. Prose criteria are not a gate: work
scored against them can *argue* it qualifies, and a label applied that way
is worth less than no label at all to the underwriters this repo is aimed
at. This module turns the standard into an executable scorecard that fails
closed.

Design rules, in priority order:

  1. FAIL CLOSED. Every criterion defaults to FAIL. A criterion passes only
     on positive evidence found by executing something -- never because a
     document asserts it, and never because evidence is merely absent.
  2. COMPOSE, DON'T REIMPLEMENT. Reference-engine agreement, builder
     parity, real-data coverage, and recalculated case status are already
     verified by dedicated tools. This module calls them and reports; it
     does not invent a second, weaker opinion about the same question.
  3. NO AGENT MAY SATISFY THE HUMAN GATE. Criterion G (effective challenge)
     reads governance/signoff.json approvals. No amount of agent work moves
     it. This is deliberate and load-bearing: it is the criterion that
     makes the rest of the scorecard non-gameable, and it is why this tool
     currently reports zero M3 domains.
  4. AN UNDECLARED CLAIM IS A FAILED CLAIM. Criterion A requires the
     inventory to declare, per engine, which workbook sheet implements it
     (`engine_surfaces`). A domain listing twelve required engines without
     saying where any of them live has not built twelve engines; it has
     written a list. Absent mapping = FAIL, with the missing engines named.

Usage:
    python tools/verify_m3_readiness.py
    python tools/verify_m3_readiness.py --model-id 01
    python tools/verify_m3_readiness.py --report m3-readiness.json
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tools"))

INVENTORY = ROOT / "standards" / "model_inventory.json"
PUBLIC_INDEX = ROOT / "standards" / "public_cases" / "index.json"

# Minimum share of a template's real-data-eligible input cells that a
# domain's own public cases must actually source before its source register
# counts as institutional-grade. Deliberately modest: the point is to
# exclude "three real cells wearing a full template," not to demand every
# illustrative structuring lever become a disclosed fact.
MIN_REAL_DATA_COVERAGE = 0.25

CRITERIA = ("A", "B", "C", "D", "E", "F", "G", "H")
CRITERION_NAMES = {
    "A": "Complete engine set (declared and implemented)",
    "B": "Stakeholder lenses (declared and distinct)",
    "C": "Source register and real-data provenance",
    "D": "Visible checks and audit trail",
    "E": "Decision usefulness and adversarial differentiation",
    "F": "Independent reference-engine agreement",
    "G": "Effective challenge (human sign-off)",
    "H": "Reproducible release",
}


def _load(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _result(status: str, evidence: str) -> dict[str, str]:
    return {"status": status, "evidence": evidence}


def check_a_engine_set(model: dict[str, Any]) -> dict[str, str]:
    """Every required engine must name a sheet that exists and has formulas."""
    required = model.get("required_engines", [])
    if not required:
        return _result("FAIL", "no required_engines declared in inventory")
    surfaces = model.get("engine_surfaces") or {}
    undeclared = [engine for engine in required if engine not in surfaces]
    if undeclared:
        return _result(
            "FAIL",
            f"{len(undeclared)}/{len(required)} engines have no declared workbook "
            f"surface: {', '.join(sorted(undeclared))}",
        )
    workbook_path = ROOT / model["workbook"]
    if not workbook_path.exists():
        return _result("FAIL", f"workbook missing: {model['workbook']}")
    workbook = load_workbook(workbook_path, data_only=False)
    missing_sheets: list[str] = []
    empty_sheets: list[str] = []
    for engine in required:
        sheet_name = surfaces[engine]
        if sheet_name not in workbook.sheetnames:
            missing_sheets.append(f"{engine}->{sheet_name}")
            continue
        sheet = workbook[sheet_name]
        formulas = sum(
            1
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
        if formulas == 0:
            empty_sheets.append(f"{engine}->{sheet_name}")
    if missing_sheets or empty_sheets:
        parts = []
        if missing_sheets:
            parts.append(f"declared sheet absent: {', '.join(missing_sheets)}")
        if empty_sheets:
            parts.append(f"declared sheet has no formulas: {', '.join(empty_sheets)}")
        return _result("FAIL", "; ".join(parts))
    return _result("PASS", f"all {len(required)} engines mapped to sheets carrying formulas")


def check_b_stakeholder_lenses(model: dict[str, Any]) -> dict[str, str]:
    """Each declared perspective must map to a surface producing its own outputs."""
    required = model.get("required_perspectives", [])
    if len(required) < 3:
        return _result("FAIL", f"M3 requires >=3 stakeholder perspectives, found {len(required)}")
    surfaces = model.get("perspective_surfaces") or {}
    undeclared = [item for item in required if item not in surfaces]
    if undeclared:
        return _result(
            "FAIL",
            f"{len(undeclared)}/{len(required)} perspectives have no declared surface: "
            f"{', '.join(sorted(undeclared))}",
        )
    workbook_path = ROOT / model["workbook"]
    workbook = load_workbook(workbook_path, data_only=False)
    problems: list[str] = []
    for perspective in required:
        location = surfaces[perspective]
        sheet_name = location.split("!", 1)[0]
        if sheet_name not in workbook.sheetnames:
            problems.append(f"{perspective}->{location} (sheet absent)")
    if problems:
        return _result("FAIL", f"declared perspective surface missing: {', '.join(problems)}")
    distinct_targets = {value.split("!", 1)[0] for value in surfaces.values()}
    if len(distinct_targets) < 2:
        return _result(
            "FAIL",
            "all perspectives point at one sheet -- lenses must produce distinct outputs, "
            "not one number relabelled",
        )
    return _result("PASS", f"{len(required)} perspectives mapped across {len(distinct_targets)} surfaces")


def check_c_source_register(model: dict[str, Any], coverage_by_model: dict[str, float]) -> dict[str, str]:
    register = ROOT / model["folder"] / "sources" / "source_register.csv"
    if not register.exists():
        return _result("FAIL", "no sources/source_register.csv")
    rows = [line for line in register.read_text(encoding="utf-8").splitlines()[1:] if line.strip()]
    if not rows:
        return _result("FAIL", "source_register.csv has no entries")
    coverage = coverage_by_model.get(model["id"])
    if coverage is None:
        return _result("FAIL", "no public case measured for real-data coverage")
    if coverage < MIN_REAL_DATA_COVERAGE:
        return _result(
            "FAIL",
            f"best public case sources {coverage * 100:.1f}% of eligible input cells "
            f"(needs >={MIN_REAL_DATA_COVERAGE * 100:.0f}%); {len(rows)} register entries",
        )
    return _result(
        "PASS",
        f"{len(rows)} register entries; best case sources {coverage * 100:.1f}% of eligible inputs",
    )


def check_d_audit_trail(model: dict[str, Any]) -> dict[str, str]:
    folder = ROOT / model["folder"]
    workbook = load_workbook(ROOT / model["workbook"], data_only=False)
    missing: list[str] = []
    check_sheets = [name for name in workbook.sheetnames if "check" in name.lower()]
    if not check_sheets:
        missing.append("no Checks sheet")
    if "RefreshLog" not in workbook.sheetnames:
        missing.append("no RefreshLog sheet")
    if not (folder / "model_card.md").exists():
        missing.append("no model_card.md")
    if not (folder / "validation.md").exists():
        missing.append("no validation.md")
    if missing:
        return _result("FAIL", "; ".join(missing))
    return _result("PASS", f"checks sheet ({check_sheets[0]}), RefreshLog, model card, validation record")


def check_e_decision_usefulness(model: dict[str, Any], case_status: dict[str, list[dict[str, Any]]]) -> dict[str, str]:
    entries = case_status.get(model["id"], [])
    if len(entries) < 2:
        return _result("FAIL", f"needs a conventional and an adversarial public case, found {len(entries)}")
    by_type = {entry.get("case_type"): entry for entry in entries}
    if "conventional" not in by_type or "adversarial" not in by_type:
        return _result("FAIL", f"case types present: {sorted(by_type)} -- needs both")
    conventional = by_type["conventional"].get("status")
    adversarial = by_type["adversarial"].get("status")
    if conventional == adversarial:
        return _result(
            "FAIL",
            f"conventional and adversarial cases both land on {conventional!r} -- "
            "the stress case does not move the decision",
        )
    return _result("PASS", f"conventional={conventional}, adversarial={adversarial} (differentiated)")


def check_f_reference_agreement(model: dict[str, Any], oracle_backed: set[str]) -> dict[str, str]:
    if model["id"] not in oracle_backed:
        return _result(
            "FAIL",
            "no independent reference-engine check binds to this model "
            "(tools/verify_reference_calcs.py)",
        )
    return _result("PASS", "independent oracle check bound and passing")


def check_g_effective_challenge(model: dict[str, Any]) -> dict[str, str]:
    signoff_path = ROOT / model["folder"] / "governance" / "signoff.json"
    if not signoff_path.exists():
        return _result("FAIL", "no governance/signoff.json")
    signoff = _load(signoff_path)
    approvals = signoff.get("approvals", [])
    required_roles = set(signoff.get("required_roles", []))
    approved_roles = {item.get("role") for item in approvals if isinstance(item, dict)}
    missing = sorted(required_roles - approved_roles)
    if missing:
        return _result(
            "FAIL",
            f"unsigned roles: {', '.join(missing)} -- human-only gate, no agent may satisfy it",
        )
    return _result("PASS", f"all required roles signed: {', '.join(sorted(approved_roles))}")


def check_h_reproducible_release(model: dict[str, Any], parity_ok: set[str]) -> dict[str, str]:
    builder = ROOT / model["builder"]
    if not builder.exists():
        return _result("FAIL", f"builder missing: {model['builder']}")
    if model["domain"] not in parity_ok:
        return _result("FAIL", "builder output does not match the committed workbook (parity failed)")
    return _result("PASS", "builder regenerates the committed workbook (semantic + presentation parity)")


def _coverage_by_model() -> dict[str, float]:
    """Best real-data coverage across each model's public cases."""
    from verify_template_exhaustion import run as coverage_run

    report = coverage_run()
    best: dict[str, float] = {}
    for item in report["results"]:
        model_id = item["model_id"]
        best[model_id] = max(best.get(model_id, 0.0), item["coverage"])
    return best


def _case_status_by_model() -> dict[str, list[dict[str, Any]]]:
    from verify_public_case_status import verify as verify_cases

    report = verify_cases()
    by_model: dict[str, list[dict[str, Any]]] = {}
    for entry in report["results"]:
        by_model.setdefault(entry["model_id"], []).append(entry)
    return by_model


def _oracle_backed() -> set[str]:
    """Model ids with at least one declared check bound to a passing identity.

    A token bound as "identity" was actually recomputed by an independent
    pure-Python oracle and agreed with the workbook. Tokens that are merely
    "flag" (a risk flag was exercised, not an identity verified) or
    "unbound" (declared in the inventory, checked by nothing) do not count
    -- an unbound reference check is a claim, not evidence.
    """
    from reference_check_registry import resolve_reference_checks

    inventory = _load(INVENTORY)
    bindings = resolve_reference_checks(inventory["models"])
    backed: set[str] = set()
    for model_id, binding in bindings.items():
        if binding.get("identity_failures"):
            continue
        if any(status == "identity" for status in binding.get("tokens", {}).values()):
            backed.add(str(model_id))
    return backed


def _parity_ok() -> set[str]:
    from build_all_models import build_and_compare
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        report = build_and_compare(ROOT, Path(tmp))
    return {
        item["domain"]
        for item in report["results"]
        if item.get("semantic_parity") and item.get("presentation_parity")
    }


def score_model(
    model: dict[str, Any],
    coverage_by_model: dict[str, float],
    case_status: dict[str, list[dict[str, Any]]],
    oracle_backed: set[str],
    parity_ok: set[str],
) -> dict[str, Any]:
    scorecard = {
        "A": check_a_engine_set(model),
        "B": check_b_stakeholder_lenses(model),
        "C": check_c_source_register(model, coverage_by_model),
        "D": check_d_audit_trail(model),
        "E": check_e_decision_usefulness(model, case_status),
        "F": check_f_reference_agreement(model, oracle_backed),
        "G": check_g_effective_challenge(model),
        "H": check_h_reproducible_release(model, parity_ok),
    }
    passed = [key for key in CRITERIA if scorecard[key]["status"] == "PASS"]
    blockers = [
        f"{key} ({CRITERION_NAMES[key]}): {scorecard[key]['evidence']}"
        for key in CRITERIA
        if scorecard[key]["status"] != "PASS"
    ]
    return {
        "model_id": model["id"],
        "domain": model["domain"],
        "declared_maturity": model["declared_maturity"],
        "criteria": scorecard,
        "passed_count": len(passed),
        "m3_ready": len(passed) == len(CRITERIA),
        "blockers": blockers,
    }


def run(model_id: str | None = None) -> dict[str, Any]:
    inventory = _load(INVENTORY)
    models = inventory["models"]
    if model_id:
        models = [item for item in models if item["id"] == model_id]
        if not models:
            raise ValueError(f"unknown model id {model_id!r}")

    coverage = _coverage_by_model()
    case_status = _case_status_by_model()
    oracle_backed = _oracle_backed()
    parity_ok = _parity_ok()

    results = [
        score_model(model, coverage, case_status, oracle_backed, parity_ok) for model in models
    ]
    results.sort(key=lambda item: (-item["passed_count"], item["model_id"]))
    ready = [item for item in results if item["m3_ready"]]
    return {
        "models_scored": len(results),
        "m3_ready_count": len(ready),
        "m3_ready": [item["model_id"] for item in ready],
        "mean_criteria_passed": round(
            sum(item["passed_count"] for item in results) / len(results), 2
        )
        if results
        else 0.0,
        "results": results,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--model-id", type=str, default=None)
    parser.add_argument("--report", type=Path, default=None)
    args = parser.parse_args()

    report = run(args.model_id)

    header = f"{'id':<4} {'domain':<36} {'decl':<5} " + " ".join(CRITERIA) + "  passed"
    print(header)
    print("-" * len(header))
    for item in report["results"]:
        marks = " ".join(
            "P" if item["criteria"][key]["status"] == "PASS" else "."
            for key in CRITERIA
        )
        print(
            f"{item['model_id']:<4} {item['domain']:<36} {item['declared_maturity']:<5} "
            f"{marks}  {item['passed_count']}/{len(CRITERIA)}"
        )
    print("-" * len(header))
    print(
        f"Models scored: {report['models_scored']}  "
        f"M3-ready: {report['m3_ready_count']}  "
        f"Mean criteria passed: {report['mean_criteria_passed']}/{len(CRITERIA)}"
    )
    print("\nCriteria: " + "; ".join(f"{key}={CRITERION_NAMES[key]}" for key in CRITERIA))

    if args.model_id and report["results"]:
        print(f"\nBlockers for {args.model_id}:")
        for line in report["results"][0]["blockers"]:
            print(f"  - {line}")

    if args.report:
        args.report.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
