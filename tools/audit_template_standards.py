"""Audit every domain template against docs/MODELING_STANDARDS_REFERENCE.md.

Static checks, run against the canonical `_template_*.xlsx` in each domain
folder (not instances -- instances inherit whatever the template does).
This is advisory: it finds real, checkable deviations from well-established
conventions (color coding, circularity-switch visibility, mid-year
discounting consistency, covenant-headroom presence, hardcodes leaking
outside input sheets) and reports them. It does not auto-fix anything --
template edits get reviewed like any other change to shared infrastructure.

Run: python tools/audit_template_standards.py [--json]
"""
from __future__ import annotations

import argparse
import glob
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]

BLUE = "0000FF"
INPUT_LIKE_SHEETS = {
    "cover", "assumptions", "sources", "refreshlog", "institutional surface",
    "challenge log", "lineage map",
}
# Sheets that are legitimately dense with same-sheet hardcoded reference
# tables (rate curves, triangles) rather than a single "Assumptions" block.
HARDCODE_TABLE_SHEETS = {
    "paid triangle", "zero curve", "futures curve", "vol surface",
}

CIRC_SHEET_KEYWORDS = ("debt schedule",)
CIRC_TEXT_PATTERN = re.compile(r"circ(ularity)?[\s_-]*(breaker|switch|toggle)", re.I)

DCF_SHEET_KEYWORDS = ("dcf",)
MID_YEAR_PATTERN = re.compile(r"mid[\s-]?year", re.I)

COVENANT_DOMAIN_HINTS = ("credit", "debt_finance", "covenants")
HEADROOM_PATTERN = re.compile(r"headroom", re.I)


@dataclass
class ModelAuditResult:
    domain: str
    template_path: str
    findings: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def status(self) -> str:
        return "REVIEW" if self.findings else "PASS"


def _sheet_text_blob(ws) -> str:
    parts = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                parts.append(cell.value)
    return "\n".join(parts)


def _has_debt_schedule_with_sweep(wb) -> bool:
    return any(name.lower() in CIRC_SHEET_KEYWORDS for name in wb.sheetnames)


_SHEET_QUALIFIED_REF = re.compile(r"(?:'[^']+'|[A-Za-z_][\w ]*)!\$?[A-Z]{1,3}\$?\d+")
_CELL_REF = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")


def _col_index(letters: str) -> int:
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def _same_sheet_refs(formula: str) -> list[tuple[str, int, int]]:
    """(coord, col_index, row) for same-sheet cell refs only (strip
    sheet-qualified refs like Assumptions!$E$16 first -- those point at a
    different sheet's static input, never part of an in-sheet circular
    loop)."""
    stripped = _SHEET_QUALIFIED_REF.sub("", formula)
    return [
        (f"{m.group(1)}{m.group(2)}", _col_index(m.group(1)), int(m.group(2)))
        for m in _CELL_REF.finditer(stripped)
    ]


def _resolves_to_strictly_prior_period(ws, coord: str, own_col: int, depth: int = 2) -> bool:
    """True if `coord`'s formula (or the cell it defers to) only ever
    reaches cells in columns strictly before `own_col` -- i.e. it's a
    prior-period link, the WSP/Macabacus non-circular resolution (interest
    on the opening balance), not a same-period average that would need a
    circularity-breaker switch. Cuts off at `depth` hops rather than
    resolving indefinitely; an unresolved chain is treated as fine here,
    since the goal is to catch clear same-period self-reference, not to
    prove non-circularity exhaustively."""
    cell = ws[coord]
    if cell.data_type != "f":
        return True  # a bare value can't be circular
    refs = _same_sheet_refs(str(cell.value))
    if not refs:
        return True
    for ref_coord, ref_col, _ in refs:
        if ref_col == own_col:
            if depth <= 0:
                return False
            if not _resolves_to_strictly_prior_period(ws, ref_coord, own_col, depth - 1):
                return False
        elif ref_col > own_col:
            return False  # forward reference within the sheet -- not a clean opening-balance chain
    return True


def _check_circularity_switch(wb, result: ModelAuditResult) -> None:
    if not _has_debt_schedule_with_sweep(wb):
        return
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() not in CIRC_SHEET_KEYWORDS:
            continue
        ws = wb[sheet_name]
        interest_rows = [
            row[0].row for row in ws.iter_rows(min_col=2, max_col=2)
            if isinstance(row[0].value, str) and "interest" in row[0].value.lower()
        ]
        circular_cells = []
        for row_idx in interest_rows:
            for cell in ws[row_idx]:
                if cell.column <= 2 or cell.data_type != "f":
                    continue
                if not _resolves_to_strictly_prior_period(ws, cell.coordinate, cell.column):
                    circular_cells.append(cell.coordinate)
        if not circular_cells:
            continue  # interest is computed on a strictly-prior-period balance -- the
            # sanctioned non-circular resolution; no switch is needed because
            # there is no circularity.
        blob = "\n".join(
            _sheet_text_blob(wb[name])
            for name in wb.sheetnames
            if name.lower() in CIRC_SHEET_KEYWORDS or name.lower() == "assumptions"
        )
        if not CIRC_TEXT_PATTERN.search(blob):
            result.findings.append(
                f"{sheet_name}: interest formula(s) at "
                f"{', '.join(circular_cells)} resolve to a same-period cell "
                "(true circularity) with no visibly labeled "
                "circularity-breaker switch (WSP/Macabacus convention: a "
                "CHOOSE/toggle between average- and beginning-balance "
                "interest, labeled on Assumptions or the schedule itself)."
            )


def _check_mid_year_convention(wb, result: ModelAuditResult) -> None:
    dcf_sheets = [name for name in wb.sheetnames if name.lower() in DCF_SHEET_KEYWORDS]
    if not dcf_sheets:
        return
    blob = "\n".join(_sheet_text_blob(wb[name]) for name in dcf_sheets)
    if not MID_YEAR_PATTERN.search(blob):
        result.notes.append(
            "DCF sheet present with no explicit mid-year-convention label "
            "found; confirm discounting timing is documented (year-end vs. "
            "mid-year) rather than implicit."
        )


def _check_covenant_headroom(wb, result: ModelAuditResult, domain: str) -> None:
    has_covenants_sheet = any(name.lower() == "covenants" for name in wb.sheetnames)
    if not has_covenants_sheet:
        return
    blob = _sheet_text_blob(wb["Covenants"])
    if not HEADROOM_PATTERN.search(blob):
        result.findings.append(
            "Covenants sheet reports ratios against thresholds but has no "
            "explicit headroom/cushion (%) output -- industry convention "
            "expresses covenant risk as headroom, not just the raw ratio."
        )


def _check_color_convention(wb, result: ModelAuditResult) -> None:
    """Flag formula cells rendered in the input-blue font -- the one
    color-convention violation that's mechanically checkable (a formula in
    blue reads as an editable input to anyone trained on the convention,
    which is the exact failure mode the convention exists to prevent)."""
    offenders = 0
    sample = None
    for name in wb.sheetnames:
        if name.lower() in INPUT_LIKE_SHEETS:
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                color = cell.font.color.rgb if cell.font and cell.font.color else None
                if isinstance(color, str) and color.upper().endswith(BLUE):
                    offenders += 1
                    if sample is None:
                        sample = f"{name}!{cell.coordinate}"
    if offenders:
        result.findings.append(
            f"{offenders} formula cell(s) rendered in input-blue font "
            f"(e.g. {sample}) -- misreads as a hardcoded input under the "
            "blue=input/black=formula/green=cross-sheet convention."
        )


def _check_hardcodes_outside_inputs(wb, result: ModelAuditResult) -> None:
    """Numeric literals on calculation sheets, outside declared input-like
    sheets or known reference-table sheets, are the 'hardcode buried in a
    formula-only sheet' failure the convention exists to catch."""
    offenders = 0
    sample = None
    for name in wb.sheetnames:
        lname = name.lower()
        if lname in INPUT_LIKE_SHEETS or lname in HARDCODE_TABLE_SHEETS:
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type != "n":
                    continue
                if cell.value in (0, 1, None):
                    continue  # 0/1 are near-universally structural, not assumptions
                offenders += 1
                if sample is None:
                    sample = f"{name}!{cell.coordinate}={cell.value!r}"
    if offenders > 0:
        result.notes.append(
            f"{offenders} bare numeric literal(s) on non-input sheet(s) "
            f"(e.g. {sample}) -- confirm these are structural (row/column "
            "indices, period counters) rather than undisclosed hardcoded "
            "assumptions that belong on Assumptions."
        )


def audit_template(path: Path, domain: str) -> ModelAuditResult:
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        template_path = str(path.relative_to(ROOT))
    except ValueError:
        template_path = str(path)
    result = ModelAuditResult(domain=domain, template_path=template_path)
    _check_circularity_switch(wb, result)
    _check_mid_year_convention(wb, result)
    _check_covenant_headroom(wb, result, domain)
    _check_color_convention(wb, result)
    _check_hardcodes_outside_inputs(wb, result)
    return result


def audit_all() -> list[ModelAuditResult]:
    templates = sorted(ROOT.glob("*/_template_*.xlsx"))
    results = []
    for path in templates:
        domain = path.parent.name
        results.append(audit_template(path, domain))
    return results


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()

    results = audit_all()

    if args.json:
        print(json.dumps(
            [
                {
                    "domain": r.domain,
                    "template": r.template_path,
                    "status": r.status,
                    "findings": r.findings,
                    "notes": r.notes,
                }
                for r in results
            ],
            indent=2,
        ))
        return 0

    review_count = sum(1 for r in results if r.status == "REVIEW")
    print(f"Audited {len(results)} templates against docs/MODELING_STANDARDS_REFERENCE.md")
    print(f"{len(results) - review_count} PASS, {review_count} REVIEW\n")
    for r in results:
        print(f"[{r.status}] {r.domain}  ({r.template_path})")
        for f in r.findings:
            print(f"    FINDING: {f}")
        for n in r.notes:
            print(f"    note: {n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
