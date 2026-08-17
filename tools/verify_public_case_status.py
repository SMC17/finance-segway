"""Recalculate every real public-case workbook and read its genuine,
computed decision status back -- the check neither
tools/evidence_receipt_integrity.py (hashes only) nor
tools/final_public_evidence.py (schema and input-count only) performs.

Both of those existing gates can pass while a workbook is financially
broken: a receipt hash matches whatever bytes are committed regardless of
whether those bytes reflect a correct recalculation, and a manifest can
satisfy "at least two observed/derived inputs" while leaving every
valuation-driving assumption on template defaults. This script is what
actually opens each of the 48 real instances, recalculates it for real via
LibreOffice, and reads back the Overall/Decision status row -- the same
manual process that found:

  - a false-negative in Insurance's "Paid triangle cumulative" check that
    failed unconditionally regardless of the underlying data (fixed);
  - roughly a dozen domains where the "conventional" reference case and
    the "adversarial" stress case land on the identical overall status,
    because their manifests only source-address a couple of deal-specific
    cells and leave the assumptions that actually drive the decision
    (WACC, terminal growth, EPS-accretion inputs, ...) on template
    defaults -- not yet fixed, tracked below.

Two tiers of enforcement:

1. Hard failures (always block): workbook missing, recalculation doesn't
   succeed cleanly, or no genuine Overall/Decision status can be found at
   all. These indicate the model is structurally broken, not just
   thin -- exactly the Insurance bug's signature.

2. Tracked failures (block only if not explicitly allowlisted below): a
   conventional case landing on the same concerning status (FAIL/BREACH)
   as its adversarial sibling. This is evidence the manifest is thin
   rather than a compile-time defect, so it's tracked in
   KNOWN_THIN_CASES rather than failing every PR immediately -- but an
   *unlisted* case matching this pattern fails CI. As each domain's
   manifest is sourced with real valuation assumptions (see
   docs/EVIDENCE_STATUS_BOARD.md), remove it from the allowlist so the
   gate ratchets tighter over time instead of silently staying lenient.

Usage:
  python tools/verify_public_case_status.py
  python tools/verify_public_case_status.py --report report.json
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
INDEX = ROOT / "standards" / "public_cases" / "index.json"

sys.path.insert(0, str(ROOT / "tools"))
sys.path.insert(0, str(ROOT))
from recalc import recalc  # noqa: E402

CHECKS_SHEET_NAMES = ("Checks", "Decision & Checks")
CONCERNING_STATUSES = {"FAIL", "BREACH"}

# (model_id, domain) pairs where the conventional and adversarial public
# cases land on the identical concerning status because the manifest only
# source-addresses a handful of deal-specific cells, leaving the
# assumptions that actually drive the decision on template defaults --
# still genuinely thin, not yet sourced. Remove an entry here only once
# real assumptions have been sourced for both of that domain's public
# cases. As of this pass, only Real Estate remains here: real occupancy
# data was added for both WeWork (75%, real) and Realty Income (98.6%,
# real), but both still land on Overall BREACH via a multi-sheet Lease
# Roll -> Debt Schedule -> 5-Year Hold & IRR formula chain that wasn't
# fully reverse-engineered in this pass (the DSCR-driving inputs
# specifically remain unsourced) -- genuine partial progress, not a
# claimed fix.
KNOWN_THIN_CASES: frozenset[str] = frozenset({
    "real-estate-public-wework-2022-stress", "real-estate-public-realty-income-2023",
})

# Pairs verified to carry real, per-company sourced inputs (see each
# manifest's own "source" fields for citations) that happen to land on the
# same concerning status anyway -- because that's genuinely how both real
# cases turned out, or because of a separate, documented structural issue
# unrelated to data sourcing. Unlike KNOWN_THIN_CASES, an entry here is not
# something to "fix" by sourcing more data -- it stays unless the
# underlying real facts or the structural issue change.
#
# - Investment Banking: HP/Autonomy's real 2011/2010 earnings show -1.7%
#   EPS dilution, Microsoft/LinkedIn's real 2016/2015 earnings show -4.5%
#   EPS dilution -- different real magnitudes from different real
#   companies, both happen to be BREACH under the template's
#   zero-tolerance dilution threshold.
# - Venture Capital: Instacart's real 2023 IPO valuation (~$9.9bn, down
#   from a ~$39bn 2021 private peak) correctly BREACHes "Round pricing"
#   while Snowflake's real 2020 IPO does not -- genuinely differentiated
#   -- but both still show Overall BREACH via the liquidation-preference
#   election-solver checks (Base/Adverse holder-election equilibrium),
#   which don't cleanly apply to an IPO (all preferred converts to common
#   at IPO; there is no real liquidation waterfall to source data for).
#   A conceptual mismatch between the case and the template's mechanic,
#   not a data gap.
#
# - Fund of Funds: fof-public-hlpaf-2026 (conventional, Hamilton Lane
#   Private Assets Fund) and fof-public-skybridge-fy2023-stress
#   (adversarial, SkyBridge Multi-Adviser Hedge Fund Portfolios) both land
#   on Overall BREACH, but for different, real, structural reasons -- not
#   a thin manifest. HLPAF's NAV roll-forward residual (~$77.8mm, ~1.3% of
#   NAV) comes from applying this template's simplified two-line fee model
#   to a real fund's actual multi-share-class fee structure (management
#   fee + incentive fee + per-class distribution fees). SkyBridge's is
#   larger (~$250mm residual, plus a genuine single-fund concentration
#   BREACH at ~27.4% vs. the 20% limit) because that fund uses real
#   open-end subscription-redemption mechanics this template has no line
#   for (capital-call mechanics only), on top of the fund's own real,
#   disclosed FY2023 collapse (-30.29% return, FTX position marked to real
#   fair value $0). Two different real vehicles, two different real
#   mismatches -- not fixable by sourcing more data, since the gap is
#   template-mechanics-vs-real-vehicle-structure, not missing inputs.
#
# Equity Finance's AMC/Tesla pair used to be here: both showed Overall
# BREACH via a confirmed template defect (Cap Table & Dilution's derived
# outputs -- rows 18-25 -- only formulaed the Active/E column, leaving C18
# permanently blank, which broke the 'Converted-share reconciliation'
# check for every Base-scenario instance). Fixed at the builder level
# (tools/builders/build_equity_finance_release.py now formulas C/D/E for
# all eight derived-output rows) and reverified via LibreOffice recalc on
# regenerated instances: Tesla now REVIEW (driven by a genuine, unrelated
# convertible anti-dilution signal), AMC still BREACH but now for its own
# real reasons (75.5% existing-holder dilution and 19.6% rights
# participation, both genuinely breach their thresholds) -- differentiated,
# so removed from this set.
KNOWN_REAL_CORRELATED_CASES: frozenset[str] = frozenset({
    "ib-public-hp-autonomy-2012-stress", "ib-public-microsoft-linkedin-2016",
    "vc-public-instacart-2023-down-round", "vc-public-snowflake-2020",
    "fof-public-hlpaf-2026", "fof-public-skybridge-fy2023-stress",
})


def _load(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _find_overall_status(workbook_path: Path) -> tuple[str | None, str | None]:
    """Search every sheet for a row whose label contains "overall"
    (case-insensitive substring, covers both "Overall" and "Overall model
    status" -- the two conventions in use across domains) and return
    (sheet_name, status). Not restricted to CHECKS_SHEET_NAMES: a renamed
    or missing checks sheet should surface as "not found", not a KeyError."""
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in range(1, min(sheet.max_row, 60) + 1):
            for col in range(1, 6):
                label = sheet.cell(row, col).value
                if label and "overall" in str(label).strip().lower():
                    for status_col in range(col + 1, col + 4):
                        status = sheet.cell(row, status_col).value
                        if status is not None:
                            return sheet_name, str(status)
    return None, None


def verify(*, keep_scratch: bool = False) -> dict[str, Any]:
    index = _load(INDEX)
    cases = index["cases"]
    results: list[dict[str, Any]] = []
    errors: list[str] = []

    with tempfile.TemporaryDirectory(dir=ROOT) as tmp:
        tmp_dir = Path(tmp)
        for case in cases:
            case_id = case["case_id"]
            src = ROOT / case["output"]
            if not src.exists():
                errors.append(f"{case_id}: missing workbook {case['output']}")
                results.append({"case_id": case_id, "status": "MISSING_WORKBOOK"})
                continue

            scratch = tmp_dir / src.name
            shutil.copyfile(src, scratch)
            recalc_result = recalc(str(scratch), timeout=90)
            recalc_ok = (
                recalc_result.get("status") == "success"
                and not recalc_result.get("total_errors", 1)
            )
            if not recalc_ok:
                errors.append(f"{case_id}: recalculation did not succeed cleanly: {recalc_result}")
                results.append({"case_id": case_id, "status": "RECALC_FAILED", "recalc": recalc_result})
                continue

            sheet_name, status = _find_overall_status(scratch)
            if status is None:
                errors.append(f"{case_id}: no Overall/Decision status found in any sheet")
                results.append({"case_id": case_id, "status": "NO_STATUS_FOUND"})
                continue

            results.append({
                "case_id": case_id,
                "model_id": case["model_id"],
                "case_type": case["case_type"],
                "sheet": sheet_name,
                "status": status,
            })

        if keep_scratch:
            shutil.copytree(tmp_dir, ROOT / ".verify-public-case-scratch", dirs_exist_ok=True)

    by_model: dict[str, dict[str, dict[str, Any]]] = {}
    for r in results:
        if "model_id" not in r:
            continue
        by_model.setdefault(r["model_id"], {})[r["case_type"]] = r

    thin_matches: list[str] = []
    unexpected_thin: list[str] = []
    for model_id, by_type in by_model.items():
        conventional = by_type.get("conventional")
        adversarial = by_type.get("adversarial")
        if not conventional or not adversarial:
            continue
        if (
            conventional["status"] == adversarial["status"]
            and conventional["status"] in CONCERNING_STATUSES
        ):
            thin_matches.append(conventional["case_id"])
            thin_matches.append(adversarial["case_id"])
            allowed = KNOWN_THIN_CASES | KNOWN_REAL_CORRELATED_CASES
            if conventional["case_id"] not in allowed:
                unexpected_thin.append(conventional["case_id"])
                errors.append(
                    f"{conventional['case_id']}: conventional case matches its adversarial "
                    f"sibling's {conventional['status']} status and is not in KNOWN_THIN_CASES "
                    "or KNOWN_REAL_CORRELATED_CASES -- either this domain regressed to a thin "
                    "manifest, or a real fix landed and the case should be moved into "
                    "KNOWN_REAL_CORRELATED_CASES instead of left unlisted"
                )

    stale_allowlist = sorted(KNOWN_THIN_CASES - set(thin_matches))
    if stale_allowlist:
        errors.append(
            "KNOWN_THIN_CASES lists cases that no longer match the thin-manifest pattern -- "
            f"remove them from the allowlist now that they're fixed: {stale_allowlist}"
        )
    stale_correlated = sorted(KNOWN_REAL_CORRELATED_CASES - set(thin_matches))
    if stale_correlated:
        errors.append(
            "KNOWN_REAL_CORRELATED_CASES lists cases that no longer match on a concerning "
            f"status -- move them out (they may now simply be differentiated): {stale_correlated}"
        )

    return {
        "status": "PASS" if not errors else "FAIL",
        "cases_checked": len(results),
        "known_thin_cases": sorted(KNOWN_THIN_CASES),
        "known_real_correlated_cases": sorted(KNOWN_REAL_CORRELATED_CASES),
        "unexpected_thin_cases": unexpected_thin,
        "stale_allowlist_entries": stale_allowlist,
        "stale_correlated_entries": stale_correlated,
        "results": results,
        "errors": errors,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report", type=Path, default=None)
    parser.add_argument(
        "--keep-scratch", action="store_true",
        help="preserve recalculated copies under .verify-public-case-scratch/ for inspection",
    )
    args = parser.parse_args()

    report = verify(keep_scratch=args.keep_scratch)
    payload = json.dumps(report, indent=2)
    if args.report:
        args.report.write_text(payload + "\n", encoding="utf-8")
    print(payload)
    return 0 if report["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
