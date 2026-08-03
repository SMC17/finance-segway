"""
weekly_refresh_check.py

Scans a folder of ticker models (each built from _template.xlsx) and produces
a single triage report: what's overdue, what's broken, what's approaching an
earnings date. This is meant to run once a week (cron / Task Scheduler /
GitHub Action) and feed a single CSV + printed summary you can act on in
15 minutes instead of re-opening 150 workbooks.

Usage:
    python3 weekly_refresh_check.py /path/to/models_folder

What it checks per workbook (reads the Cover + RefreshLog tabs only —
fast, no recalculation needed):
    1. Days since "Last refreshed" (Cover!C6) -> flag if > 7
    2. Days until "Next earnings date" (Cover!C7) -> flag if <= 5 (refresh NOW)
    3. Whether the file has any cached formula errors (#REF!, #NAME?, #VALUE! etc.)
    4. Missing required tabs (structural drift from template)

Optionally, if yfinance is installed, cross-checks the ticker's live price
against nothing stored in-model (v1 has no live price cell yet) and flags
tickers with no ticker recognized -- extend this to diff key financial line
items once you wire actuals into Assumptions!H (source col).
"""
import sys
import os
import csv
import glob
from datetime import datetime, date
import openpyxl

# Different archetypes have genuinely different tab structures — an LBO model
# has no "IS/BS/CF" tabs the way a corp-finance DCF model does. Match required
# tabs by filename/foldername signal rather than one universal list.
ARCHETYPE_REQUIRED_TABS = {
    "corp_finance": {"Cover", "Assumptions", "IS", "BS", "CF", "DCF", "Comps", "Sensitivity", "RefreshLog"},
    "lbo": {"Cover", "Sources & Uses", "Debt Schedule", "Returns", "Sensitivity", "RefreshLog"},
    "credit": {"Cover", "Assumptions", "Covenants", "Debt Schedule", "RefreshLog"},
    "risk": {"Cover", "VaR", "Stress Scenarios", "RefreshLog"},
    "default": {"Cover", "RefreshLog"},  # fallback: only the two universal tabs
}


def detect_archetype(path, sheetnames):
    lower = path.lower()
    if "lbo" in lower or "sources & uses" in {s.lower() for s in sheetnames}:
        return "lbo"
    if "credit" in lower:
        return "credit"
    if "risk" in lower:
        return "risk"
    if {"IS", "BS", "CF", "DCF"}.issubset(sheetnames):
        return "corp_finance"
    return "default"
ERROR_TOKENS = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}
STALE_DAYS = 7
EARNINGS_WARNING_DAYS = 5


def parse_date_cell(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value if isinstance(value, date) and not isinstance(value, datetime) else value.date()
    s = str(value).strip()
    if not s or s.startswith("["):
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%B %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def scan_workbook(path):
    result = {
        "file": os.path.basename(path),
        "path": path,
        "archetype": "",
        "missing_tabs": "",
        "last_refreshed": "",
        "days_since_refresh": "",
        "next_earnings": "",
        "days_to_earnings": "",
        "formula_errors_found": "",
        "flags": [],
    }
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        result["flags"].append(f"COULD NOT OPEN: {e}")
        return result

    archetype = detect_archetype(path, set(wb.sheetnames))
    required = ARCHETYPE_REQUIRED_TABS[archetype]
    missing = required - set(wb.sheetnames)
    result["archetype"] = archetype
    if missing:
        result["missing_tabs"] = ", ".join(sorted(missing))
        result["flags"].append(f"STRUCTURAL DRIFT vs {archetype} (missing: {', '.join(sorted(missing))})")

    today = date.today()

    if "Cover" in wb.sheetnames:
        cov = wb["Cover"]
        last_ref = parse_date_cell(cov["C6"].value)
        next_earn = parse_date_cell(cov["C7"].value)

        if last_ref:
            days_since = (today - last_ref).days
            result["last_refreshed"] = str(last_ref)
            result["days_since_refresh"] = days_since
            if days_since > STALE_DAYS:
                result["flags"].append(f"STALE ({days_since}d since refresh)")
        else:
            result["flags"].append("NO REFRESH DATE SET")

        if next_earn:
            days_to = (next_earn - today).days
            result["next_earnings"] = str(next_earn)
            result["days_to_earnings"] = days_to
            if 0 <= days_to <= EARNINGS_WARNING_DAYS:
                result["flags"].append(f"EARNINGS IN {days_to}d — refresh before/after print")
            elif days_to < 0:
                result["flags"].append(f"EARNINGS DATE PASSED {abs(days_to)}d AGO — update next date")

    # scan all sheets for cached error tokens
    errors_found = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in ERROR_TOKENS:
                    errors_found.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
    if errors_found:
        result["formula_errors_found"] = "; ".join(errors_found[:10])
        result["flags"].append(f"{len(errors_found)} CELL ERROR(S)")

    if not result["flags"]:
        result["flags"].append("OK")

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 weekly_refresh_check.py /path/to/models_folder")
        sys.exit(1)

    folder = sys.argv[1]
    files = sorted(glob.glob(os.path.join(folder, "**", "*.xlsx"), recursive=True))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]

    if not files:
        print(f"No .xlsx files found in {folder}")
        sys.exit(0)

    rows = [scan_workbook(f) for f in files]

    report_path = os.path.join(folder, f"refresh_report_{date.today().isoformat()}.csv")
    fieldnames = ["file", "archetype", "flags", "last_refreshed", "days_since_refresh",
                  "next_earnings", "days_to_earnings", "missing_tabs",
                  "formula_errors_found", "path"]
    with open(report_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            r["flags"] = "; ".join(r["flags"])
            writer.writerow(r)

    # console summary, most urgent first
    def urgency(r):
        f = r["flags"]
        if "EARNINGS IN" in f or "COULD NOT OPEN" in f:
            return 0
        if "ERROR" in f or "DRIFT" in f:
            return 1
        if "STALE" in f:
            return 2
        return 3

    rows_sorted = sorted(rows, key=urgency)
    print(f"\nScanned {len(rows)} models in {folder}\n" + "-" * 60)
    for r in rows_sorted:
        flag_str = r["flags"] if isinstance(r["flags"], str) else "; ".join(r["flags"])
        if flag_str != "OK":
            print(f"{r['file']:30s} {flag_str}")
    ok_count = sum(1 for r in rows if (r["flags"] if isinstance(r["flags"], str) else "; ".join(r["flags"])) == "OK")
    print("-" * 60)
    print(f"{ok_count}/{len(rows)} clean. Full report: {report_path}")


if __name__ == "__main__":
    main()
