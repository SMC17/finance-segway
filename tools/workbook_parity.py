"""Semantic and presentation parity for generated Excel workbooks.

Semantic parity covers sheet order/state, formulas, hardcoded values, defined
names, and external-link state. Presentation parity separately records number
formats, styles, dimensions, merges, validations, conditional formatting, and
calculation settings. Harmless Excel/LibreOffice serialization variants are
canonicalized before semantic comparison.
"""
from __future__ import annotations

import hashlib
import json
import math
import re
from decimal import Decimal
from numbers import Real
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

_SIMPLE_QUOTED_SHEET = re.compile(r"'([A-Za-z_][A-Za-z0-9_]*)'!")
_BOOLEAN_CALL = re.compile(r"\b(TRUE|FALSE)\(\)", re.IGNORECASE)
_SCIENTIFIC_ZERO_PADDING = re.compile(
    r"(?<![A-Z0-9_.$])(\d+(?:\.\d+)?)[Ee]\+?0*(\d+)(?![A-Z0-9_.$])"
)
# Only decimal literals are normalized here. Integers are left untouched so
# row/column references and explicit integer assumptions remain transparent.
# The reference guards exclude A1, $A$1, identifiers, and decimal fragments.
_DECIMAL_LITERAL = re.compile(
    r"(?<![A-Za-z0-9_.$])((?:\d+\.\d+)|(?:\d+\.))(?![A-Za-z0-9_.$])"
)
_SEMANTIC_SIGNIFICANT_DIGITS = 12


def _normalize_decimal_literal(match: re.Match[str]) -> str:
    value = Decimal(match.group(1))
    normalized = format(value.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return normalized or "0"


def _normalize_formula_segment(segment: str) -> str:
    segment = _SIMPLE_QUOTED_SHEET.sub(r"\1!", segment)
    segment = _BOOLEAN_CALL.sub(lambda match: match.group(1).upper(), segment)
    segment = _SCIENTIFIC_ZERO_PADDING.sub(
        lambda match: f"{_normalize_decimal_literal_from_text(match.group(1))}E{int(match.group(2))}",
        segment,
    )
    segment = _DECIMAL_LITERAL.sub(_normalize_decimal_literal, segment)
    return segment


def _normalize_decimal_literal_from_text(text: str) -> str:
    value = Decimal(text)
    normalized = format(value.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return normalized or "0"


def normalize_formula(formula: str) -> str:
    """Canonicalize equivalent formula serialization without changing logic.

    Formula text inside quoted Excel strings is preserved byte-for-byte. Outside
    strings, the normalizer removes unnecessary sheet quotes, Boolean call
    parentheses, exponent zero padding, and trailing decimal zeros. It does not
    round formula constants, so a material assumption change remains a failure.
    """
    normalized = formula.strip()
    pieces = normalized.split('"')
    for index in range(0, len(pieces), 2):
        pieces[index] = _normalize_formula_segment(pieces[index])
    return '"'.join(pieces)


def _canonical_number(value: Real) -> str:
    """Return deterministic spreadsheet precision for hardcoded numerics.

    Excel and LibreOffice round-trip IEEE-754 values with slightly different
    final digits. Twelve significant digits is stricter than displayed precision
    in the model library while eliminating binary serialization noise. NaN and
    infinities remain explicit and therefore auditable.
    """
    numeric = float(value)
    if math.isnan(numeric):
        return "NaN"
    if math.isinf(numeric):
        return "Infinity" if numeric > 0 else "-Infinity"
    if numeric == 0:
        return "0"
    return format(numeric, f".{_SEMANTIC_SIGNIFICANT_DIGITS}g")


def _color(value: Any) -> str | None:
    if value is None:
        return None
    kind = getattr(value, "type", None)
    if kind == "rgb":
        return getattr(value, "rgb", None)
    if kind == "theme":
        return f"theme:{getattr(value, 'theme', None)}:{getattr(value, 'tint', None)}"
    if kind == "indexed":
        return f"indexed:{getattr(value, 'indexed', None)}"
    return None


def _side(value: Any) -> tuple[Any, ...]:
    return (getattr(value, "style", None), _color(getattr(value, "color", None)))


def _semantic_cell(cell: Any) -> dict[str, Any]:
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        kind = "formula"
        value = normalize_formula(value)
    elif value is None:
        kind = "blank"
    elif isinstance(value, Real) and not isinstance(value, bool):
        kind = "number"
        value = _canonical_number(value)
    else:
        kind = type(value).__name__
    return {"coordinate": cell.coordinate, "kind": kind, "value": value}


def _presentation_cell(cell: Any) -> dict[str, Any]:
    return {
        "coordinate": cell.coordinate,
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "color": _color(cell.font.color),
        },
        "fill": {
            "type": cell.fill.fill_type,
            "fg": _color(cell.fill.fgColor),
            "bg": _color(cell.fill.bgColor),
        },
        "border": {
            "left": _side(cell.border.left),
            "right": _side(cell.border.right),
            "top": _side(cell.border.top),
            "bottom": _side(cell.border.bottom),
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "text_rotation": cell.alignment.text_rotation,
            "indent": cell.alignment.indent,
        },
        "protection": {
            "locked": cell.protection.locked,
            "hidden": cell.protection.hidden,
        },
    }


def workbook_fingerprints(path: Path) -> tuple[dict[str, Any], dict[str, Any]]:
    workbook = load_workbook(path, data_only=False, read_only=False, keep_links=True)
    semantic_sheets = []
    presentation_sheets = []
    for sheet in workbook.worksheets:
        semantic_cells = []
        presentation_cells = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    semantic_cells.append(_semantic_cell(cell))
                if cell.value is not None or cell.has_style:
                    presentation_cells.append(_presentation_cell(cell))
        semantic_sheets.append({
            "title": sheet.title,
            "state": sheet.sheet_state,
            "cells": semantic_cells,
        })
        validations = [{
            "type": item.type,
            "formula1": item.formula1,
            "formula2": item.formula2,
            "sqref": str(item.sqref),
            "allow_blank": item.allowBlank,
        } for item in sheet.data_validations.dataValidation]
        presentation_sheets.append({
            "title": sheet.title,
            "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            "show_grid_lines": sheet.sheet_view.showGridLines,
            "merged_ranges": sorted(str(item) for item in sheet.merged_cells.ranges),
            "column_dimensions": {
                key: {"width": item.width, "hidden": item.hidden, "outline": item.outlineLevel}
                for key, item in sorted(sheet.column_dimensions.items())
                if item.width is not None or item.hidden or item.outlineLevel
            },
            "row_dimensions": {
                str(key): {"height": item.height, "hidden": item.hidden, "outline": item.outlineLevel}
                for key, item in sorted(sheet.row_dimensions.items())
                if item.height is not None or item.hidden or item.outlineLevel
            },
            "data_validations": sorted(validations, key=lambda item: (item["sqref"], str(item["type"]))),
            "conditional_format_ranges": sorted(str(item.sqref) for item in sheet.conditional_formatting),
            "cells": presentation_cells,
        })
    names = sorted(
        (name, getattr(item, "attr_text", None), getattr(item, "localSheetId", None))
        for name, item in workbook.defined_names.items()
    )
    semantic = {
        "sheet_order": workbook.sheetnames,
        "sheets": semantic_sheets,
        "defined_names": names,
        "external_links": len(getattr(workbook, "_external_links", [])),
    }
    presentation = {
        "sheets": presentation_sheets,
        "calculation": {
            "calc_mode": workbook.calculation.calcMode,
            "full_calc_on_load": workbook.calculation.fullCalcOnLoad,
            "force_full_calc": workbook.calculation.forceFullCalc,
        },
    }
    for payload in (semantic, presentation):
        encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")
        payload["sha256"] = hashlib.sha256(encoded).hexdigest()
    return semantic, presentation


def _compare_sheets(left: dict[str, Any], right: dict[str, Any], fields: tuple[str, ...]) -> list[str]:
    differences = []
    left_sheets = {item["title"]: item for item in left["sheets"]}
    right_sheets = {item["title"]: item for item in right["sheets"]}
    for name in sorted(set(left_sheets) | set(right_sheets)):
        if name not in left_sheets:
            differences.append(f"sheet_missing_generated:{name}")
            continue
        if name not in right_sheets:
            differences.append(f"sheet_missing_committed:{name}")
            continue
        for field in fields:
            if left_sheets[name].get(field) != right_sheets[name].get(field):
                differences.append(f"{name}:{field}")
    return differences


def _cell_diff_details(left: dict[str, Any], right: dict[str, Any], limit: int = 100) -> list[dict[str, Any]]:
    details: list[dict[str, Any]] = []
    left_sheets = {item["title"]: item for item in left["sheets"]}
    right_sheets = {item["title"]: item for item in right["sheets"]}
    for sheet_name in sorted(set(left_sheets) & set(right_sheets)):
        left_cells = {item["coordinate"]: item for item in left_sheets[sheet_name]["cells"]}
        right_cells = {item["coordinate"]: item for item in right_sheets[sheet_name]["cells"]}
        for coordinate in sorted(set(left_cells) | set(right_cells)):
            generated = left_cells.get(coordinate)
            committed = right_cells.get(coordinate)
            if generated == committed:
                continue
            details.append({
                "sheet": sheet_name,
                "cell": coordinate,
                "generated": generated,
                "committed": committed,
            })
            if len(details) >= limit:
                return details
    return details


def compare_workbooks(generated: Path, committed: Path) -> dict[str, Any]:
    generated_semantic, generated_presentation = workbook_fingerprints(generated)
    committed_semantic, committed_presentation = workbook_fingerprints(committed)
    semantic_differences = []
    if generated_semantic["sheet_order"] != committed_semantic["sheet_order"]:
        semantic_differences.append("sheet_order")
    semantic_differences.extend(_compare_sheets(
        generated_semantic, committed_semantic, ("state", "cells"),
    ))
    for field in ("defined_names", "external_links"):
        if generated_semantic[field] != committed_semantic[field]:
            semantic_differences.append(field)
    presentation_differences = _compare_sheets(
        generated_presentation,
        committed_presentation,
        (
            "freeze_panes", "show_grid_lines", "merged_ranges",
            "column_dimensions", "row_dimensions", "data_validations",
            "conditional_format_ranges", "cells",
        ),
    )
    if generated_presentation["calculation"] != committed_presentation["calculation"]:
        presentation_differences.append("calculation")
    return {
        "generated": str(generated),
        "committed": str(committed),
        "semantic_sha256_generated": generated_semantic["sha256"],
        "semantic_sha256_committed": committed_semantic["sha256"],
        "presentation_sha256_generated": generated_presentation["sha256"],
        "presentation_sha256_committed": committed_presentation["sha256"],
        "parity": not semantic_differences,
        "semantic_parity": not semantic_differences,
        "presentation_parity": not presentation_differences,
        "differences": semantic_differences,
        "semantic_differences": semantic_differences,
        "semantic_cell_differences": _cell_diff_details(generated_semantic, committed_semantic),
        "presentation_differences": presentation_differences,
    }
