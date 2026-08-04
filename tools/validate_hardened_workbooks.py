"""Build and validate release-grade workbooks for every M1 hardening domain."""
from __future__ import annotations

import argparse
import importlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from tools.workbook_engineering import audit_workbook
except ModuleNotFoundError:
    from workbook_engineering import audit_workbook

ROOT = Path(__file__).resolve().parents[1]

CONTRACTS: dict[str, dict[str, Any]] = {
    "08": {
        "domain": "Asset Management",
        "module": "tools.builders.build_asset_management_release",
        "filename": "ASSET_MANAGEMENT.xlsx",
        "required_sheets": [
            "Fund NAV",
            "Fee Waterfall",
            "Fund Performance",
            "Return Measurement",
            "Exposure & Liquidity",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Fund NAV": ["C10", "D5", "D10", "F10"],
            "Fee Waterfall": ["C16", "C21", "C22", "C24"],
            "Fund Performance": ["C16", "C17", "C18", "C19"],
            "Return Measurement": ["C9", "C12", "C13", "C15"],
            "Exposure & Liquidity": ["C18", "D18", "C20", "D20", "C22", "D22", "C23"],
            "Decision & Checks": ["C5", "D5", "C7", "D7", "C15"],
        },
        "labels": {
            "Decision & Checks": [
                "NAV rollforward residual",
                "Waterfall conservation residual",
                "Time-weighted return",
                "Downside liquidity coverage",
            ]
        },
        "exact_formulas": {
            "Fee Waterfall!C16": "=MIN(MAX(C10-C14-C15,0),C15*C8/(1-C8)*C9)",
            "Fee Waterfall!C24": "=C10-C21-C22",
        },
    },
    "10": {
        "domain": "Trade Finance",
        "module": "tools.builders.build_trade_finance_release",
        "filename": "TRADE_FINANCE.xlsx",
        "required_sheets": [
            "Working Capital Cycle",
            "Credit & Facility",
            "LC & Factoring Cost",
            "Documentary Controls",
            "Financing Cost Comparison",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Working Capital Cycle": ["C12", "C13", "C14", "C15", "C16"],
            "Credit & Facility": ["E17", "E18", "E19", "E20", "E21"],
            "LC & Factoring Cost": ["C9", "C10", "C18", "C19", "C20", "C21"],
            "Documentary Controls": ["D5", "D7", "D10", "D11"],
            "Decision & Checks": ["C5", "D5", "C7", "D7", "C14"],
        },
        "labels": {
            "Decision & Checks": [
                "Cash-conversion identity residual",
                "Expected loss residual",
                "Risk-adjusted margin",
                "Documentary-control exceptions",
            ]
        },
        "exact_formulas": {
            "LC & Factoring Cost!C21": '=IFERROR(C19/C20*365/C17,"-")',
        },
    },
    "11": {
        "domain": "Microfinance",
        "module": "tools.builders.build_microfinance_release",
        "filename": "MICROFINANCE.xlsx",
        "required_sheets": [
            "Loan Portfolio",
            "Portfolio Rollforward",
            "Sustainability",
            "Funding & Liquidity",
            "Client & Conduct",
            "Provisioning",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Portfolio Rollforward": ["E16", "E17", "E18", "E19", "E20", "E21"],
            "Sustainability": ["C12", "C13", "C14"],
            "Funding & Liquidity": ["C14", "D14", "C15", "D15", "C17", "C18"],
            "Client & Conduct": ["D5", "D6", "D9", "D10"],
            "Provisioning": ["C13", "C14"],
            "Decision & Checks": ["C5", "D5", "C7", "D7", "C15"],
        },
        "labels": {
            "Decision & Checks": [
                "Portfolio rollforward residual",
                "Composite credit-risk ratio",
                "Operational self-sufficiency",
                "Downside funding gap",
            ]
        },
    },
    "12": {
        "domain": "Equity Finance",
        "module": "tools.builders.build_equity_finance_release",
        "filename": "EQUITY_FINANCE.xlsx",
        "required_sheets": [
            "Assumptions",
            "Cap Table & Dilution",
            "Rights Offering",
            "Convertible Securities",
            "IS",
            "BS",
            "CF",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Cap Table & Dilution": ["E18", "E19", "E20", "E21", "E22", "E23", "E24", "E25"],
            "Rights Offering": ["C13", "D13", "C14", "D14", "C17", "D17", "C19", "D19"],
            "Convertible Securities": ["C14", "D14", "C15", "D15", "C16", "D16", "C19", "D19"],
            "Assumptions": ["C12", "D12", "E12", "F12", "G12"],
            "Decision & Checks": ["C5", "D5", "C10", "D10", "C16"],
        },
        "labels": {
            "Decision & Checks": [
                "Share-conservation residual",
                "Existing-holder dilution",
                "Rights TERP residual",
                "Incremental anti-dilution shares",
            ]
        },
    },
    "15": {
        "domain": "Commodities",
        "module": "tools.builders.build_commodities_release",
        "filename": "COMMODITIES.xlsx",
        "required_sheets": [
            "Futures Curve",
            "Roll Yield",
            "Hedging",
            "Physical Balance & Carry",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Physical Balance & Carry": ["E22", "E23", "E24", "E25", "E27", "E28", "E29"],
            "Decision & Checks": ["C5", "D5", "C6", "D6", "C12"],
        },
        "labels": {
            "Decision & Checks": ["Overall model status", "Fair forward", "Unhedged exposure"]
        },
    },
    "16": {
        "domain": "Crypto & Digital Assets",
        "module": "tools.builders.build_crypto_release",
        "filename": "CRYPTO.xlsx",
        "required_sheets": [
            "Tokenomics",
            "Supply Rollforward & Unlocks",
            "Valuation",
            "Treasury & Runway",
            "Custody & Liquidity",
            "Staking Yield",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Supply Rollforward & Unlocks": ["E16", "E17", "E18", "E19", "E20", "E21"],
            "Valuation": ["C11", "C12", "C13", "C14", "C15", "C16"],
            "Treasury & Runway": ["C18", "D18", "C19", "D19", "C20", "D20", "C22"],
            "Custody & Liquidity": ["D5", "D6", "D7", "D12"],
            "Staking Yield": ["C11", "C12", "C13", "C14"],
            "Decision & Checks": ["C5", "D5", "C7", "D7", "C15"],
        },
        "labels": {
            "Decision & Checks": [
                "Supply-conservation residual",
                "Unlock rate",
                "Downside operating runway",
                "Custody / liquidity exceptions",
            ]
        },
    },
    "17": {
        "domain": "Real Estate & REIT",
        "module": "tools.builders.build_real_estate_release",
        "filename": "REAL_ESTATE.xlsx",
        "required_sheets": [
            "Property Pro Forma",
            "Lease Roll",
            "Cap Rate & Valuation",
            "Debt Schedule",
            "REIT FFO-AFFO",
            "5-Year Hold & IRR",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Property Pro Forma": ["C7", "C9", "C11", "C13", "C14"],
            "Lease Roll": ["C5", "H5", "I5", "J5", "K5"],
            "Debt Schedule": ["C8", "D8", "E8", "F8", "G8", "I8", "J8"],
            "REIT FFO-AFFO": ["C7", "C11"],
            "Decision & Checks": ["C5", "D5", "C8", "D8", "C14"],
        },
        "labels": {
            "Decision & Checks": [
                "NOI identity residual",
                "Minimum five-year DSCR",
                "Levered IRR",
                "AFFO",
            ]
        },
        "exact_formulas": {
            "Property Pro Forma!C7": "=C5-C6",
            "Property Pro Forma!C11": "=C9-C10",
            "REIT FFO-AFFO!C7": "=C4+C5-C6",
            "REIT FFO-AFFO!C11": "=C7-C9-C10",
        },
    },
    "23": {
        "domain": "Fintech & Payments",
        "module": "tools.builders.build_fintech_release",
        "filename": "FINTECH.xlsx",
        "required_sheets": [
            "Unit Economics",
            "Cohort Retention",
            "Network & Cohorts",
            "Fraud & Risk",
            "Capital & Liquidity",
            "Operational Controls",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Network & Cohorts": ["E16", "E17", "E18", "E20", "E21"],
            "Fraud & Risk": ["C12", "C14", "C15", "C16"],
            "Capital & Liquidity": ["C14", "D14", "C15", "D15", "C18", "C19"],
            "Decision & Checks": ["C5", "D5", "C7", "D7", "C16"],
        },
        "labels": {
            "Decision & Checks": [
                "Revenue identity residual",
                "Contribution after risk losses",
                "Downside capital coverage",
                "Transaction failure rate",
            ]
        },
    },
    "24": {
        "domain": "Distressed & Restructuring",
        "module": "tools.builders.build_restructuring_release",
        "filename": "RESTRUCTURING.xlsx",
        "required_sheets": [
            "Recovery Waterfall",
            "13-Week Liquidity",
            "New Money",
            "Liquidation vs Reorg",
            "Decision & Checks",
        ],
        "formula_cells": {
            "Recovery Waterfall": ["E5", "F5", "G5", "C21"],
            "13-Week Liquidity": ["C7", "I7", "J7", "K7", "C22", "C23", "C24"],
            "New Money": ["C14", "C15", "C16", "C17", "C18", "C19"],
            "Decision & Checks": ["C5", "D5", "C8", "D8", "C14"],
        },
        "labels": {
            "Decision & Checks": [
                "Waterfall conservation residual",
                "Minimum 13-week liquidity",
                "Reorganization NPV uplift",
                "Fulcrum security",
            ]
        },
    },
}


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def build_all(output_dir: Path) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    outputs: dict[str, Path] = {}
    for model_id, contract in CONTRACTS.items():
        module = importlib.import_module(contract["module"])
        output = output_dir / contract["filename"]
        module.build(output)
        outputs[model_id] = output
    return outputs


def validate_workbook(model_id: str, path: Path) -> dict[str, Any]:
    contract = CONTRACTS[model_id]
    errors: list[str] = []
    workbook = load_workbook(path, data_only=False, keep_links=True)
    missing_sheets = [
        name for name in contract["required_sheets"] if name not in workbook.sheetnames
    ]
    if missing_sheets:
        errors.append(f"missing required sheets: {missing_sheets}")
    for sheet_name, cells in contract.get("formula_cells", {}).items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for coordinate in cells:
            if not _is_formula(sheet[coordinate].value):
                errors.append(
                    f"{sheet_name}!{coordinate} is not a formula: {sheet[coordinate].value!r}"
                )
    for address, expected in contract.get("exact_formulas", {}).items():
        sheet_name, coordinate = address.split("!", 1)
        actual = workbook[sheet_name][coordinate].value
        if actual != expected:
            errors.append(f"{address} expected {expected!r}, found {actual!r}")
    for sheet_name, labels in contract.get("labels", {}).items():
        if sheet_name not in workbook.sheetnames:
            continue
        observed = {
            str(cell.value).strip()
            for row in workbook[sheet_name].iter_rows()
            for cell in row
            if cell.value not in (None, "")
        }
        for label in labels:
            if label not in observed:
                errors.append(f"{sheet_name}: missing decision label {label!r}")
    if getattr(workbook, "_external_links", []):
        errors.append("external workbook links are prohibited")
    summary, findings = audit_workbook(path)
    audit_errors = [item.__dict__ for item in findings if item.severity == "error"]
    if audit_errors:
        errors.append(f"workbook engineering errors: {audit_errors}")
    return {
        "model_id": model_id,
        "domain": contract["domain"],
        "path": str(path),
        "sheets": len(workbook.sheetnames),
        "formulas": summary["formulas"],
        "errors": errors,
        "status": "PASS" if not errors else "FAIL",
    }


def validate_directory(directory: Path) -> dict[str, Any]:
    results = []
    errors = []
    for model_id, contract in CONTRACTS.items():
        result = validate_workbook(model_id, directory / contract["filename"])
        results.append(result)
        errors.extend(f"{model_id}: {error}" for error in result["errors"])
    return {
        "schema_version": "1.0",
        "models": len(results),
        "errors": errors,
        "results": results,
        "status": "PASS" if not errors else "FAIL",
        "promotion_statement": (
            "Workbook contract passage is required before inventory maturity or builder "
            "paths are changed."
        ),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--build-dir", type=Path)
    parser.add_argument("--workbook-dir", type=Path)
    parser.add_argument(
        "--report", type=Path, default=ROOT / "hardened-workbook-report.json"
    )
    args = parser.parse_args()
    if bool(args.build_dir) == bool(args.workbook_dir):
        parser.error("provide exactly one of --build-dir or --workbook-dir")
    directory = args.build_dir or args.workbook_dir
    assert directory is not None
    if args.build_dir:
        build_all(directory)
    report = validate_directory(directory)
    args.report.write_text(json.dumps(report, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "models": report["models"],
                "status": report["status"],
                "errors": report["errors"],
            },
            indent=2,
        )
    )
    return 0 if report["status"] == "PASS" else 1


if __name__ == "__main__":
    raise SystemExit(main())
