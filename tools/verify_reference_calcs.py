"""Independent-oracle regression checks for high-risk workbook mechanics.

Each check mutates a copied canonical workbook, recalculates it with
LibreOffice, and compares the resulting cached values with an independent
Python calculation or an explicit accounting identity. The checks deliberately
reference the released institutional workbooks rather than retired skeleton
sheet names.
"""
from __future__ import annotations

import math
import os
import shutil
import sys
import tempfile
from collections.abc import Callable

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(REPO_ROOT, "tools"))

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from recalc import recalc  # noqa: E402

TOL = 1e-4


def close(a, b, tol=TOL):
    if a is None or b is None:
        return False
    if b == 0:
        return abs(a) < 1e-6
    return abs(a - b) / abs(b) < tol


def with_recalc(src_path: str, populate_fn: Callable) -> openpyxl.Workbook:
    """Copy, mutate, recalculate, and return a data-only workbook."""
    temporary = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temporary.close()
    shutil.copy(src_path, temporary.name)
    try:
        workbook = openpyxl.load_workbook(temporary.name)
        populate_fn(workbook)
        workbook.save(temporary.name)
        result = recalc(temporary.name, timeout=90)
        if result.get("status") != "success" or result.get("total_errors", 0):
            raise RuntimeError(f"recalculation failed: {result}")
        return openpyxl.load_workbook(temporary.name, data_only=True)
    finally:
        os.unlink(temporary.name)


# ---------------------------------------------------------------------
# Options: Black-Scholes closed form and put-call parity.
# ---------------------------------------------------------------------
def norm_cdf(x: float) -> float:
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


# A third, independent Black-Scholes implementation (see also
# tools/reference_engines.py and finance_segway/derivatives.py) — kept
# separate deliberately since this one is the oracle checking the workbook,
# not something those should import. Local and single-call-site, so the q/sigma
# order swap that motivated keyword-only args in the other two isn't live here.
def black_scholes(S, K, T, r, q, sigma):
    d1 = (math.log(S / K) + (r - q + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    call = S * math.exp(-q * T) * norm_cdf(d1) - K * math.exp(-r * T) * norm_cdf(d2)
    put = K * math.exp(-r * T) * norm_cdf(-d2) - S * math.exp(-q * T) * norm_cdf(-d1)
    return call, put


def check_black_scholes():
    S, K, T, r, q, sigma = 100.0, 100.0, 0.25, 0.045, 0.0, 0.30
    path = os.path.join(REPO_ROOT, "14_Options_Derivatives", "_template_OPTIONS.xlsx")

    def populate(workbook):
        assumptions = workbook["Assumptions"]
        for cell, value in zip(
            ("C5", "C6", "C7", "C8", "C9", "C10"),
            (S, K, T, r, q, sigma),
        ):
            assumptions[cell] = value
        workbook["Cover"]["C9"] = "Base"

    workbook = with_recalc(path, populate)
    pricer = workbook["European Pricer"]
    sheet_call = pricer["C7"].value
    sheet_put = pricer["C8"].value
    reference_call, reference_put = black_scholes(S, K, T, r, q, sigma)
    parity_lhs = sheet_call - sheet_put
    parity_rhs = S * math.exp(-q * T) - K * math.exp(-r * T)
    ok = (
        close(sheet_call, reference_call, tol=1e-6)
        and close(sheet_put, reference_put, tol=1e-6)
        and close(parity_lhs, parity_rhs, tol=1e-6)
    )
    detail = (
        f"call sheet={sheet_call:.6f} oracle={reference_call:.6f}; "
        f"put sheet={sheet_put:.6f} oracle={reference_put:.6f}; "
        f"parity residual={parity_lhs - parity_rhs:.10f}"
    )
    return "Options: Black-Scholes price and put-call parity", ok, detail


# ---------------------------------------------------------------------
# Fixed income: closed-form price and modified duration.
# ---------------------------------------------------------------------
def bond_price(face, coupon_rate, frequency, years, ytm):
    periods = int(round(frequency * years))
    coupon = face * coupon_rate / frequency
    periodic_yield = ytm / frequency
    return (
        sum(coupon / (1 + periodic_yield) ** period for period in range(1, periods + 1))
        + face / (1 + periodic_yield) ** periods
    )


def closed_form_modified_duration(face, coupon_rate, frequency, years, ytm):
    periods = int(round(frequency * years))
    coupon = face * coupon_rate / frequency
    periodic_yield = ytm / frequency
    price = bond_price(face, coupon_rate, frequency, years, ytm)
    macaulay_periods = sum(
        period * (coupon if period < periods else coupon + face) / (1 + periodic_yield) ** period
        for period in range(1, periods + 1)
    ) / price
    return (macaulay_periods / frequency) / (1 + periodic_yield)


def check_bond_duration():
    face, coupon, frequency, years, ytm = 1000.0, 0.05, 2.0, 10.0, 0.055
    path = os.path.join(REPO_ROOT, "21_Fixed_Income_Rates", "_template_FIXED_INCOME.xlsx")

    def populate(workbook):
        assumptions = workbook["Assumptions"]
        for cell, value in zip(
            ("C5", "C6", "C7", "C8", "C9"),
            (face, coupon, ytm, years, frequency),
        ):
            assumptions[cell] = value
        workbook["Cover"]["C9"] = "Base"

    workbook = with_recalc(path, populate)
    analytics = workbook["Bond Analytics"]
    sheet_price = analytics["C5"].value
    sheet_duration = analytics["C8"].value
    reference_price = bond_price(face, coupon, frequency, years, ytm)
    reference_duration = closed_form_modified_duration(face, coupon, frequency, years, ytm)
    ok = close(sheet_price, reference_price, tol=1e-6) and close(
        sheet_duration, reference_duration, tol=0.001
    )
    detail = (
        f"price sheet={sheet_price:.6f} oracle={reference_price:.6f}; "
        f"modified duration sheet={sheet_duration:.6f} oracle={reference_duration:.6f}"
    )
    return "Fixed income: price and modified duration", ok, detail


# ---------------------------------------------------------------------
# LBO: balanced transaction and every debt/cash roll-forward identity.
# ---------------------------------------------------------------------
def check_lbo_sources_uses_and_debt_schedule():
    path = os.path.join(REPO_ROOT, "03_Private_Equity", "_template_LBO.xlsx")
    workbook = with_recalc(path, lambda _: None)
    sources_uses = workbook["Sources & Uses"]
    debt = workbook["Debt Schedule"]

    sources_total = sources_uses["F9"].value
    uses_total = sources_uses["C9"].value
    balance_check = sources_uses["F10"].value
    ok = close(sources_total, uses_total, tol=1e-8) and close(balance_check, 0.0, tol=1e-8)
    failures = []

    if not close(
        debt["C27"].value,
        debt["C16"].value + debt["C21"].value + debt["C26"].value,
        tol=1e-8,
    ):
        failures.append("close total debt")
    if not close(debt["C28"].value, debt["C27"].value - debt["C13"].value, tol=1e-8):
        failures.append("close net debt")

    for column in range(4, 11):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)
        identities = {
            "begin cash": (debt[f"{letter}5"].value, debt[f"{previous}13"].value),
            "ending cash": (
                debt[f"{letter}13"].value,
                debt[f"{letter}9"].value
                + debt[f"{letter}10"].value
                - debt[f"{letter}11"].value
                - debt[f"{letter}12"].value,
            ),
            "begin revolver": (debt[f"{letter}14"].value, debt[f"{previous}16"].value),
            "ending revolver": (
                debt[f"{letter}16"].value,
                max(0.0, debt[f"{letter}14"].value + debt[f"{letter}10"].value),
            ),
            "begin TLB": (debt[f"{letter}17"].value, debt[f"{previous}21"].value),
            "ending TLB": (
                debt[f"{letter}21"].value,
                max(
                    0.0,
                    debt[f"{letter}17"].value
                    - debt[f"{letter}19"].value
                    - debt[f"{letter}20"].value,
                ),
            ),
            "begin second lien": (debt[f"{letter}22"].value, debt[f"{previous}26"].value),
            "ending second lien": (
                debt[f"{letter}26"].value,
                max(
                    0.0,
                    debt[f"{letter}22"].value
                    + debt[f"{letter}24"].value
                    - debt[f"{letter}25"].value,
                ),
            ),
            "total debt": (
                debt[f"{letter}27"].value,
                debt[f"{letter}16"].value
                + debt[f"{letter}21"].value
                + debt[f"{letter}26"].value,
            ),
            "net debt": (
                debt[f"{letter}28"].value,
                debt[f"{letter}27"].value - debt[f"{letter}13"].value,
            ),
            "revolver commitment fee": (
                debt[f"{letter}29"].value,
                max(0.0, workbook["Assumptions"]["E17"].value - debt[f"{letter}14"].value)
                * workbook["Assumptions"]["E33"].value,
            ),
            "cash interest includes commitment fee": (
                debt[f"{letter}7"].value,
                debt[f"{letter}15"].value
                + debt[f"{letter}18"].value
                + debt[f"{letter}23"].value
                + debt[f"{letter}29"].value,
            ),
        }
        for identity, (actual, expected) in identities.items():
            if not close(actual, expected, tol=1e-8):
                failures.append(f"{letter}:{identity}")
        if min(
            debt[f"{letter}16"].value,
            debt[f"{letter}21"].value,
            debt[f"{letter}26"].value,
        ) < -1e-8:
            failures.append(f"{letter}:negative debt")
        if debt[f"{letter}29"].value <= 0:
            failures.append(f"{letter}:commitment fee not actually charged")

    ok = ok and not failures
    detail = (
        f"sources={sources_total:.6f}, uses={uses_total:.6f}, balance={balance_check:.10f}; "
        f"roll-forward failures={failures or 'none'}; year-7 net debt={debt['J28'].value:.6f}"
    )
    return "LBO: Sources & Uses and seven-year debt roll-forward", ok, detail


# ---------------------------------------------------------------------
# VC ownership, round-pricing, and exit-proceeds identities.
# ---------------------------------------------------------------------
def check_vc_waterfall_conservation():
    path = os.path.join(REPO_ROOT, "13_Venture_Capital", "_template_VC.xlsx")

    scenarios = (
        ("base", 80.0, 20.0, 8.0, 2.0, 0.0, 500.0),
        ("dilutive", 100.0, 20.0, 10.0, 4.0, 5.0, 80.0),
    )

    def populate(workbook, values):
        pre_money, investment, existing, new, option_pool, exit_value = values
        ownership = workbook["Ownership & Dilution"]
        for cell, value in zip(
            ("C5", "C6", "C7", "C8", "C9"),
            (pre_money, investment, existing, new, option_pool),
        ):
            ownership[cell] = value
        waterfall = workbook["Exit Waterfall"]
        waterfall["C5"] = exit_value
        waterfall["C6"] = investment

    ok = True
    details = []
    for scenario, *values in scenarios:
        pre_money, investment, existing, new, option_pool, exit_value = values
        workbook = with_recalc(
            path, lambda item, inputs=values: populate(item, inputs)
        )
        ownership = workbook["Ownership & Dilution"]
        waterfall = workbook["Exit Waterfall"]
        total_shares = existing + new + option_pool
        expected_investor_ownership = new / total_shares
        expected_founder_ownership = existing / total_shares
        expected_pool_ownership = option_pool / total_shares
        expected_proceeds = exit_value * expected_investor_ownership
        expected_moic = expected_proceeds / investment
        observed_ownerships = (
            ownership["C17"].value,
            ownership["C18"].value,
            ownership["C19"].value,
        )
        scenario_ok = all(
            (
                close(ownership["C15"].value, pre_money + investment, tol=1e-8),
                close(ownership["C16"].value, total_shares, tol=1e-8),
                close(observed_ownerships[0], expected_investor_ownership, tol=1e-8),
                close(observed_ownerships[1], expected_founder_ownership, tol=1e-8),
                close(observed_ownerships[2], expected_pool_ownership, tol=1e-8),
                close(sum(observed_ownerships), 1.0, tol=1e-8),
                close(ownership["C20"].value, investment / new, tol=1e-8),
                close(waterfall["C11"].value, expected_investor_ownership, tol=1e-8),
                close(waterfall["C12"].value, expected_proceeds, tol=1e-8),
                close(waterfall["C13"].value, expected_moic, tol=1e-8),
                waterfall["C12"].value <= exit_value,
            )
        )
        ok = ok and scenario_ok
        details.append(
            f"{scenario}: ownership={waterfall['C11'].value}, "
            f"proceeds={waterfall['C12'].value}, exit={exit_value}, "
            f"ownership_sum={sum(observed_ownerships)}"
        )
    return "VC: ownership, pricing, and exit-proceeds identities", ok, "; ".join(details)


def check_vc_holder_election_waterfall():
    """Re-derive the preferred-class election equilibrium independently."""

    path = os.path.join(REPO_ROOT, "13_Venture_Capital", "_template_VC.xlsx")
    class_names = ("Series B", "Series A", "Seed")
    shares = (1_000_000.0, 1_500_000.0, 1_000_000.0)
    preference_claims = (5_000_000.0, 3_000_000.0, 1_000_000.0)
    common_shares = 7_000_000.0
    base_exit = 20_000_000.0
    adverse_exit = 50_000_000.0

    def populate(workbook):
        cap_table = workbook["Cap Table"]
        cap_rows = {
            5: (6_000_000.0, 0.0001, "Common", 0.0, "N/A", 0.0, 0, 1.0),
            6: (1_000_000.0, 0.0001, "Common", 0.0, "N/A", 0.0, 0, 1.0),
            7: (0.0, 0.0, "SAFE — deal-specific", 0.0, "Deal-specific", 0.0, 0, 0.0),
            8: (1_000_000.0, 1.0, "Preferred", 1.0, "Non-participating", 0.0, 1, 1.0),
            9: (1_500_000.0, 2.0, "Preferred", 1.0, "Non-participating", 0.0, 2, 1.0),
            10: (1_000_000.0, 5.0, "Preferred", 1.0, "Non-participating", 0.0, 3, 1.0),
        }
        for row, values in cap_rows.items():
            for column, value in zip((3, 5, 7, 8, 9, 10, 11, 12), values):
                cap_table.cell(row, column, value)
        waterfall = workbook["Exit Waterfall"]
        waterfall["C5"] = base_exit
        waterfall["D5"] = adverse_exit

    def candidate(exit_proceeds, elections):
        remaining = exit_proceeds
        preference_payouts = []
        for converts, claim in zip(elections, preference_claims):
            payout = 0.0 if converts else min(remaining, claim)
            preference_payouts.append(payout)
            remaining = max(0.0, remaining - payout)
        denominator = common_shares + sum(
            share for share, converts in zip(shares, elections) if converts
        )
        payouts = [
            remaining * share / denominator if converts else preference_payout
            for share, converts, preference_payout in zip(
                shares, elections, preference_payouts
            )
        ]
        return payouts + [remaining * common_shares / denominator]

    def equilibrium(exit_proceeds):
        stable = []
        for mask in range(8):
            elections = ((mask >> 2) & 1, (mask >> 1) & 1, mask & 1)
            payouts = candidate(exit_proceeds, elections)
            if all(
                payouts[index] + 1e-7
                >= candidate(
                    exit_proceeds,
                    tuple(
                        1 - election if other == index else election
                        for other, election in enumerate(elections)
                    ),
                )[index]
                for index in range(3)
            ):
                stable.append((mask, elections, payouts))
        if not stable:
            raise AssertionError("independent oracle found no stable election")
        return stable[0]

    workbook = with_recalc(path, populate)
    waterfall = workbook["Exit Waterfall"]
    expected_base = equilibrium(base_exit)
    expected_adverse = equilibrium(adverse_exit)
    observed_base = [waterfall.cell(row, 9).value for row in range(21, 25)]
    observed_adverse = [waterfall.cell(row, 11).value for row in range(21, 25)]
    observed_base_elections = [waterfall.cell(row, 8).value for row in range(21, 24)]
    observed_adverse_elections = [waterfall.cell(row, 10).value for row in range(21, 24)]
    expected_base_elections = [
        "CONVERT" if value else "PREFERENCE" for value in expected_base[1]
    ]
    expected_adverse_elections = [
        "CONVERT" if value else "PREFERENCE" for value in expected_adverse[1]
    ]
    mismatches = []
    for scenario, observed, expected in (
        ("Base", observed_base, expected_base[2]),
        ("Adverse", observed_adverse, expected_adverse[2]),
    ):
        for name, actual, reference in zip((*class_names, "Common"), observed, expected):
            if not close(actual, reference, tol=1e-8):
                mismatches.append(
                    f"{scenario} {name}: sheet={actual}, oracle={reference:.6f}"
                )
    ok = all(
        (
            waterfall["C17"].value == "SUPPORTED",
            observed_base_elections == expected_base_elections,
            observed_adverse_elections == expected_adverse_elections,
            waterfall["I28"].value == expected_base[0],
            waterfall["K28"].value == expected_adverse[0],
            close(waterfall["I26"].value, base_exit, tol=1e-8),
            close(waterfall["K26"].value, adverse_exit, tol=1e-8),
            not mismatches,
        )
    )
    detail = (
        f"Base mask={waterfall['I28'].value} elections={observed_base_elections}; "
        f"Adverse mask={waterfall['K28'].value} elections={observed_adverse_elections}; "
        f"payout mismatches={mismatches or 'none'}"
    )
    return "VC: holder-by-holder liquidation-preference equilibrium", ok, detail


# ---------------------------------------------------------------------
# BASE archetype: integrated income statement and cash-flow linkage.
# ---------------------------------------------------------------------
def check_base_archetype_integration():
    path = os.path.join(REPO_ROOT, "01_Investment_Banking", "_template_BASE.xlsx")
    growth = [0.10, 0.09, 0.08, 0.07, 0.06]
    gross_margin = 0.60
    opex_percent = 0.25
    tax_rate = 0.25
    da_percent = 0.80
    capex_percent = 0.05
    shares = 100
    initial_revenue = 1000.0
    initial_interest = 20.0

    def populate(workbook):
        assumptions = workbook["Assumptions"]
        for index, rate in enumerate(growth):
            assumptions.cell(row=5, column=3 + index, value=rate)
        for column in range(3, 8):
            assumptions.cell(row=6, column=column, value=gross_margin)
            assumptions.cell(row=7, column=column, value=opex_percent)
            assumptions.cell(row=8, column=column, value=tax_rate)
            assumptions.cell(row=9, column=column, value=da_percent)
            assumptions.cell(row=10, column=column, value=capex_percent)
            assumptions.cell(row=12, column=column, value=shares)
        income_statement = workbook["IS"]
        income_statement["E5"] = initial_revenue
        income_statement["E15"] = initial_interest

    workbook = with_recalc(path, populate)
    income_statement = workbook["IS"]
    sheet_revenue = [income_statement.cell(row=5, column=column).value for column in range(6, 10)]
    sheet_net_income = [
        income_statement.cell(row=18, column=column).value for column in range(6, 10)
    ]

    reference_revenue = []
    reference_net_income = []
    revenue = initial_revenue
    for growth_rate in growth[:4]:
        revenue *= 1 + growth_rate
        gross_profit = revenue * gross_margin
        ebitda = gross_profit - revenue * opex_percent
        depreciation = revenue * capex_percent * da_percent
        pretax_income = ebitda - depreciation - initial_interest
        net_income = pretax_income * (1 - tax_rate)
        reference_revenue.append(revenue)
        reference_net_income.append(net_income)

    ok = all(close(actual, expected) for actual, expected in zip(sheet_revenue, reference_revenue))
    ok = ok and all(
        close(actual, expected) for actual, expected in zip(sheet_net_income, reference_net_income)
    )
    cash_flow_net_income = [
        workbook["CF"].cell(row=5, column=column).value for column in range(5, 8)
    ]
    ok = ok and all(
        close(actual, expected)
        for actual, expected in zip(cash_flow_net_income, sheet_net_income[:3])
    )
    detail = (
        f"revenue sheet={sheet_revenue}, oracle={reference_revenue}; "
        f"net income sheet={sheet_net_income}, oracle={reference_net_income}; "
        f"cash-flow linked net income={cash_flow_net_income}"
    )
    return "BASE: projected income statement and cash-flow linkage", ok, detail


# ---------------------------------------------------------------------
# Private Credit: leverage-based excess-cash-flow sweep step-down grid.
# ---------------------------------------------------------------------
def check_credit_ecf_sweep_stepdown():
    path = os.path.join(REPO_ROOT, "05_Private_Credit", "_template_CREDIT.xlsx")
    workbook = with_recalc(path, lambda _: None)
    debt = workbook["Debt Schedule"]

    tier1 = debt["E21"].value
    breakpoints = [debt["D21"].value, debt["D22"].value, debt["D23"].value]
    rates = [tier1, debt["E22"].value, debt["E23"].value, debt["E24"].value]

    failures = []
    if not close(rates[1], tier1 * 2 / 3, tol=1e-8):
        failures.append("tier2 not 2/3 of tier1")
    if not close(rates[2], tier1 / 3, tol=1e-8):
        failures.append("tier3 not 1/3 of tier1")
    if rates[3] != 0:
        failures.append("tier4 not zero")

    tiers_visited = set()
    for column in range(4, 9):
        letter = get_column_letter(column)
        previous = get_column_letter(column - 1)
        beginning_debt = debt[f"{letter}11"].value
        prior_ebitda = workbook["Operating Case"][f"{previous}7"].value
        expected_leverage = beginning_debt / prior_ebitda if prior_ebitda else 0.0
        actual_leverage = debt[f"{letter}26"].value
        if not close(actual_leverage, expected_leverage, tol=1e-6):
            failures.append(f"{letter}:leverage")

        if expected_leverage >= breakpoints[0]:
            expected_rate = rates[0]
        elif expected_leverage >= breakpoints[1]:
            expected_rate = rates[1]
        elif expected_leverage >= breakpoints[2]:
            expected_rate = rates[2]
        else:
            expected_rate = rates[3]
        tiers_visited.add(round(expected_rate, 10))
        actual_rate = debt[f"{letter}27"].value
        if not close(actual_rate, expected_rate, tol=1e-8):
            failures.append(f"{letter}:sweep rate")

        mandatory_amort = debt[f"{letter}7"].value
        cfads_after_interest = debt[f"{letter}6"].value
        cash_before_sweep = debt[f"{letter}8"].value
        minimum_cash = workbook["Assumptions"]["E19"].value
        expected_sweep = min(
            max(0.0, beginning_debt - mandatory_amort),
            max(0.0, cash_before_sweep - minimum_cash) * expected_rate,
        )
        actual_sweep = debt[f"{letter}9"].value
        if not close(actual_sweep, expected_sweep, tol=1e-6):
            failures.append(f"{letter}:sweep amount")

    if len(tiers_visited) < 2:
        failures.append("step-down grid never crosses a tier boundary -- decorative, not mechanically real")

    ok = not failures
    detail = (
        f"tier rates={rates}; breakpoints={breakpoints}; "
        f"tiers visited across schedule={sorted(tiers_visited, reverse=True)}; "
        f"failures={failures or 'none'}"
    )
    return "Private Credit: leverage-based ECF sweep step-down grid", ok, detail


# ---------------------------------------------------------------------
# Software / SaaS: ARR roll-forward, retention ordering, revenue split.
# ---------------------------------------------------------------------
def check_software_arr_rollforward():
    """Recompute the ARR cohort balance and its derived retention metrics.

    Independent of the workbook's own formulas: the expected series is
    generated here in pure Python from the Assumptions drivers, then
    compared cell-for-cell against the recalculated sheet. Retention is the
    part worth checking hardest -- NRR and GRR are easy to define wrongly
    (including new logos in NRR is the classic error, and it inflates the
    number in exactly the direction a seller would like).
    """
    path = os.path.join(REPO_ROOT, "31_Software_SaaS", "_template_SOFTWARE.xlsx")
    workbook = with_recalc(path, lambda _: None)
    assumptions = workbook["Assumptions"]
    arr = workbook["ARR Rollforward"]
    operating = workbook["Operating Model"]

    beginning = assumptions["E5"].value
    new_rate = assumptions["E6"].value
    expansion_rate = assumptions["E7"].value
    contraction_rate = assumptions["E8"].value
    churn_rate = assumptions["E9"].value
    services_rate = assumptions["E10"].value
    subscription_margin = assumptions["E11"].value
    services_margin = assumptions["E12"].value

    failures = []
    balance = beginning
    for index, column in enumerate("CDEFG"):
        new = balance * new_rate
        expansion = balance * expansion_rate
        contraction = balance * contraction_rate
        churn = balance * churn_rate
        ending = balance + new + expansion - contraction - churn
        # NRR: installed base only -- excludes new logos by construction.
        expected_nrr = (balance + expansion - contraction - churn) / balance
        # GRR: churn only -- no credit for expansion.
        expected_grr = (balance - churn) / balance

        if not close(arr[f"{column}5"].value, balance, tol=1e-6):
            failures.append(f"year{index + 1} beginning ARR")
        if not close(arr[f"{column}10"].value, ending, tol=1e-6):
            failures.append(f"year{index + 1} ending ARR")
        if not close(arr[f"{column}12"].value, expected_nrr, tol=1e-9):
            failures.append(f"year{index + 1} NRR")
        if not close(arr[f"{column}13"].value, expected_grr, tol=1e-9):
            failures.append(f"year{index + 1} GRR")
        if expected_nrr < expected_grr - 1e-12:
            failures.append(f"year{index + 1} NRR below GRR")

        # Revenue recognises the average ARR balance, not the ending one.
        subscription = (balance + ending) / 2
        services = subscription * services_rate
        total_revenue = subscription + services
        blended = (
            subscription * subscription_margin + services * services_margin
        ) / total_revenue
        if not close(operating[f"{column}7"].value, total_revenue, tol=1e-6):
            failures.append(f"year{index + 1} total revenue")
        if not close(operating[f"{column}11"].value, blended, tol=1e-9):
            failures.append(f"year{index + 1} blended gross margin")
        if not (services_margin - 1e-9 <= blended <= subscription_margin + 1e-9):
            failures.append(f"year{index + 1} blended margin outside component bounds")

        # SBC add-back must never reduce operating income.
        if operating[f"{column}19"].value < operating[f"{column}16"].value - 1e-9:
            failures.append(f"year{index + 1} non-GAAP below GAAP")

        balance = ending

    ok = not failures
    detail = (
        f"5-year ARR {beginning:.1f} -> {balance:.1f}; "
        f"NRR={arr['C12'].value:.4f} GRR={arr['C13'].value:.4f} "
        f"(NRR excludes new logos by construction); "
        f"blended GM={operating['C11'].value:.4f} within "
        f"[{services_margin:.2f}, {subscription_margin:.2f}]; "
        f"failures={'none' if ok else failures}"
    )
    return "Software: ARR roll-forward, retention ordering, revenue split", ok, detail


def check_software_rpo_rollforward():
    """Recompute the RPO balance and the bookings residual independently.

    The load-bearing property is that gross bookings are a *derived*
    quantity, not an input. No issuer discloses bookings, so a model that
    accepts one has accepted a number nobody can check. Here bookings come
    out of the ASC 606 identity

        ending RPO = beginning RPO + gross bookings - revenue recognised

    which means the check is really "does the workbook's bookings line
    equal the residual implied by two disclosed balances and one disclosed
    flow". It also verifies the billed/unbilled split: contract liability
    (deferred revenue) is on the balance sheet, the rest of RPO is not, and
    conflating them overstates how much of backlog is near-term cash.
    """
    path = os.path.join(REPO_ROOT, "31_Software_SaaS", "_template_SOFTWARE.xlsx")
    workbook = with_recalc(path, lambda _: None)
    assumptions = workbook["Assumptions"]
    operating = workbook["Operating Model"]
    rpo = workbook["RPO & Bookings"]

    beginning_rpo = assumptions["E20"].value
    coverage = assumptions["E21"].value
    current_share = assumptions["E22"].value
    deferred_rate = assumptions["E23"].value

    failures = []
    balance = beginning_rpo
    for index, column in enumerate("CDEFG"):
        revenue = operating[f"{column}7"].value
        ending = revenue * coverage
        bookings = ending - balance + revenue
        current = ending * current_share
        deferred = revenue * deferred_rate
        unbilled = ending - deferred

        if not close(rpo[f"{column}6"].value, balance, tol=1e-6):
            failures.append(f"year{index + 1} beginning RPO")
        if not close(rpo[f"{column}7"].value, ending, tol=1e-6):
            failures.append(f"year{index + 1} ending RPO")
        if not close(rpo[f"{column}8"].value, bookings, tol=1e-6):
            failures.append(f"year{index + 1} implied bookings")
        # The identity itself, recomputed rather than read off the sheet.
        if not close(balance + bookings - revenue, ending, tol=1e-6):
            failures.append(f"year{index + 1} RPO identity")
        if not close(rpo[f"{column}11"].value, current, tol=1e-6):
            failures.append(f"year{index + 1} current RPO")
        if not close(
            rpo[f"{column}11"].value + rpo[f"{column}12"].value, ending, tol=1e-6
        ):
            failures.append(f"year{index + 1} current + non-current != total RPO")
        if not close(rpo[f"{column}14"].value, unbilled, tol=1e-6):
            failures.append(f"year{index + 1} unbilled RPO")
        if unbilled < -1e-9:
            failures.append(f"year{index + 1} deferred revenue exceeds total RPO")
        if bookings <= 0:
            failures.append(f"year{index + 1} non-positive implied bookings")
        # Book-to-bill above 1.0 must coincide with a growing RPO balance;
        # if those two ever disagree the residual has been mis-signed.
        book_to_bill = bookings / revenue
        if (book_to_bill > 1.0 + 1e-9) != (ending > balance + 1e-9):
            failures.append(f"year{index + 1} book-to-bill contradicts RPO growth")

        balance = ending

    ok = not failures
    detail = (
        f"5-year RPO {beginning_rpo:.1f} -> {balance:.1f} at {coverage:.2f}x "
        f"coverage; bookings derived as residual (year1 "
        f"{rpo['C8'].value:.1f}, book-to-bill {rpo['C9'].value:.3f}); "
        f"unbilled share {rpo['C15'].value:.3f} of RPO; "
        f"failures={'none' if ok else failures}"
    )
    return "Software: RPO roll-forward, bookings residual, billed/unbilled split", ok, detail


CHECKS = [
    check_black_scholes,
    check_bond_duration,
    check_lbo_sources_uses_and_debt_schedule,
    check_vc_waterfall_conservation,
    check_vc_holder_election_waterfall,
    check_base_archetype_integration,
    check_credit_ecf_sweep_stepdown,
    check_software_arr_rollforward,
    check_software_rpo_rollforward,
]


def main() -> int:
    print(f"Running {len(CHECKS)} independent-oracle verification checks...\n")
    all_ok = True
    for check in CHECKS:
        try:
            name, ok, detail = check()
        except Exception as error:  # noqa: BLE001 - report every regression coherently.
            name, ok, detail = check.__name__, False, f"raised {type(error).__name__}: {error}"
        print(f"[{'PASS' if ok else 'FAIL'}] {name}")
        print(f"       {detail}\n")
        all_ok = all_ok and ok
    print("=" * 60)
    print("ALL CHECKS PASSED" if all_ok else "SOME CHECKS FAILED")
    return 0 if all_ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
