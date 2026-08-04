"""Render a receipt-verified real-market RAM research workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from research.ram.simple_covariance import (
        equal_weight_risk,
        inverse_vol_weights,
        is_positive_semidefinite,
        portfolio_variance,
    )
except ModuleNotFoundError:  # direct execution from research/ram
    from simple_covariance import (
        equal_weight_risk,
        inverse_vol_weights,
        is_positive_semidefinite,
        portfolio_variance,
    )


ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = Path(__file__).resolve().parent / "data"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_release(as_of: date) -> tuple[dict[str, Any], list[list[float]], list[dict[str, str]], dict[str, Any]]:
    stamp = as_of.isoformat()
    receipt_path = DATA_DIR / f"receipt_{stamp}.json"
    if not receipt_path.is_file():
        raise FileNotFoundError(f"missing source receipt: {receipt_path}")
    receipt = json.loads(receipt_path.read_text(encoding="utf-8"))
    if receipt.get("classification") != "external_historical_market_observation":
        raise ValueError("receipt is not a real historical market observation")
    if receipt.get("as_of") != stamp or not receipt.get("sources"):
        raise ValueError("receipt as-of or source list is invalid")
    artifacts = receipt.get("artifacts", {})
    if len(artifacts) < 4:
        raise ValueError("receipt does not cover the complete release")
    for relative, expected in artifacts.items():
        path = (ROOT / relative).resolve()
        if not path.is_relative_to(ROOT) or not path.is_file():
            raise FileNotFoundError(f"receipted artifact is missing: {relative}")
        actual = sha256_file(path)
        if actual != expected:
            raise ValueError(f"artifact hash mismatch: {relative}")

    metadata_path = DATA_DIR / f"universe_{stamp}.json"
    covariance_path = DATA_DIR / f"covariance_{stamp}.json"
    regime_path = ROOT / f"research/kdb/exports/regime_summary_{stamp.replace('-', '')}.csv"
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    covariance = json.loads(covariance_path.read_text(encoding="utf-8"))
    with regime_path.open(newline="", encoding="utf-8") as handle:
        regimes = list(csv.DictReader(handle))
    if not regimes:
        raise ValueError("regime export is empty")
    return metadata, covariance, regimes, receipt


def header(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)


def build(as_of: date, output: Path) -> None:
    metadata, covariance, regimes, receipt = load_release(as_of)
    tickers = metadata["tickers"]
    volatilities = metadata["volatilities"]
    if len(tickers) != len(covariance) or not is_positive_semidefinite(covariance):
        raise ValueError("receipted covariance is not dimensionally valid PSD data")
    equal = equal_weight_risk(covariance)
    inverse_weights = inverse_vol_weights(volatilities)
    inverse_variance = portfolio_variance(inverse_weights, covariance)

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover["B2"] = "Finance-Segway — receipt-verified RAM research view"
    cover["B2"].font = Font(bold=True, size=14)
    rows = (
        ("Status", "RESEARCH ONLY — not a governed decision model"),
        ("As of", metadata["as_of"]),
        ("Universe", ", ".join(tickers)),
        ("Observations", metadata["n_returns"]),
        ("Methodology", metadata["methodology"]),
        ("Source count", len(receipt["sources"])),
        ("Receipt status", receipt["evidence_status"]),
        ("Maturity contribution", "None"),
    )
    for index, (label, value) in enumerate(rows, start=4):
        cover.cell(index, 2, label).font = Font(bold=True)
        cover.cell(index, 3, value)
    cover.column_dimensions["B"].width = 24
    cover.column_dimensions["C"].width = 100

    universe = workbook.create_sheet("Universe")
    for column, label in enumerate(("Ticker", "Annualized vol", "Equal weight", "Inverse-vol weight"), start=2):
        header(universe.cell(4, column, label))
    for row, ticker in enumerate(tickers, start=5):
        index = row - 5
        universe.cell(row, 2, ticker)
        universe.cell(row, 3, volatilities[index])
        universe.cell(row, 4, equal.weights[index])
        universe.cell(row, 5, inverse_weights[index])
    for column in range(2, 6):
        universe.column_dimensions[get_column_letter(column)].width = 22

    matrix = workbook.create_sheet("Covariance")
    matrix["B2"] = "Annualized covariance — receipt verified"
    matrix["B3"] = "PSD"
    matrix["C3"] = "PASS"
    for column, ticker in enumerate(tickers, start=3):
        header(matrix.cell(5, column, ticker))
    for row, ticker in enumerate(tickers, start=6):
        header(matrix.cell(row, 2, ticker))
        for column, value in enumerate(covariance[row - 6], start=3):
            matrix.cell(row, column, value)

    risk = workbook.create_sheet("Risk Summary")
    for column, label in enumerate(("Metric", "Equal weight", "Inverse volatility"), start=2):
        header(risk.cell(4, column, label))
    for row, values in enumerate(
        (
            ("Variance", equal.variance, inverse_variance),
            ("Volatility", equal.volatility, math.sqrt(inverse_variance)),
            ("Weight sum", sum(equal.weights), sum(inverse_weights)),
        ),
        start=5,
    ):
        for column, value in enumerate(values, start=2):
            risk.cell(row, column, value)

    regime = workbook.create_sheet("Regime Summary")
    fields = list(regimes[0])
    for column, field in enumerate(fields, start=2):
        header(regime.cell(4, column, field))
    for row, observation in enumerate(regimes, start=5):
        for column, field in enumerate(fields, start=2):
            regime.cell(row, column, observation[field])

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--as-of", type=date.fromisoformat, required=True)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    output = args.output or Path(__file__).resolve().parent / "visualizations" / f"universe_{args.as_of}.xlsx"
    build(args.as_of, output)
    print(output)


if __name__ == "__main__":
    main()
