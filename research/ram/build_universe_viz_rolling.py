\"\"\"Rolling vol / correlation visualization builder (research rail only).

Produces research/ram/visualizations/_universe_viz_rolling.xlsx when real
returns are available under research/ram/data/rets_real_10.json.

Not a governed decision model.
\"\"\"
from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

DATA_DIR = Path(__file__).resolve().parent / "data"
OUT_DIR = Path(__file__).resolve().parent / "visualizations"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_PATH = OUT_DIR / "_universe_viz_rolling.xlsx"

WINDOW = 21  # trading days


def rolling_vol(series: list[float], window: int = WINDOW) -> list[float | None]:
    out: list[float | None] = [None] * len(series)
    for i in range(window - 1, len(series)):
        chunk = series[i - window + 1 : i + 1]
        mean = sum(chunk) / window
        var = sum((x - mean) ** 2 for x in chunk) / (window - 1)
        out[i] = math.sqrt(var * 252)
    return out


def build() -> None:
    rets_path = DATA_DIR / "rets_real_10.json"
    meta_path = DATA_DIR / "universe_real_10.json"
    if not rets_path.exists() or not meta_path.exists():
        print("Real returns not found. Run fetch_real_universe.py first.")
        print("(Optional) write rets_real_10.json from the fetcher to enable this viz.")
        return

    meta = json.loads(meta_path.read_text())
    payload = json.loads(rets_path.read_text())
    tickers = payload.get("tickers") or meta["tickers"]
    returns = payload["returns"]  # T x N
    dates = payload.get("dates") or list(range(len(returns)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"
    ws["B2"] = "Research — Rolling Vol Visualization (Stage 0/1)"
    ws["B2"].font = Font(bold=True, size=14)
    ws["B4"] = "Status"
    ws["C4"] = "RESEARCH ONLY — not a governed decision model"
    ws["B5"] = "Window"
    ws["C5"] = f"{WINDOW} trading days, annualized"
    ws["B6"] = "Universe"
    ws["C6"] = ", ".join(tickers)
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 80

    # Rolling vol sheet (last 60 points for readability)
    ws = wb.create_sheet("Rolling Vol")
    ws["A1"] = "date"
    for j, t in enumerate(tickers):
        ws.cell(row=1, column=2 + j, value=t)
        col = [returns[i][j] for i in range(len(returns))]
        rvol = rolling_vol(col)
        start = max(0, len(rvol) - 60)
        for row_i, idx in enumerate(range(start, len(rvol)), start=2):
            if j == 0:
                ws.cell(row=row_i, column=1, value=dates[idx] if idx < len(dates) else idx)
            val = rvol[idx]
            if val is not None:
                ws.cell(row=row_i, column=2 + j, value=round(val, 4))

    ws = wb.create_sheet("Notes")
    ws["B2"] = "Regenerate"
    ws["C2"] = "python research/ram/fetch_real_universe.py && python research/ram/build_universe_viz_rolling.py"
    ws["B4"] = "Governance"
    ws["C4"] = "Outside model inventory. No M-maturity claim."

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
