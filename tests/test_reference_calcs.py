from __future__ import annotations

import unittest
from unittest.mock import patch

from openpyxl import Workbook

from tools import verify_reference_calcs


class VentureCapitalReferenceCalcTests(unittest.TestCase):
    def test_vc_oracle_targets_the_institutional_workbook_contract(self):
        def recalculate(_path, populate):
            workbook = Workbook()
            workbook.remove(workbook.active)
            ownership = workbook.create_sheet("Ownership & Dilution")
            waterfall = workbook.create_sheet("Exit Waterfall")
            populate(workbook)

            pre_money = ownership["C5"].value
            investment = ownership["C6"].value
            existing = ownership["C7"].value
            new = ownership["C8"].value
            option_pool = ownership["C9"].value
            total_shares = existing + new + option_pool
            investor_ownership = new / total_shares
            founder_ownership = existing / total_shares
            pool_ownership = option_pool / total_shares

            ownership["C15"] = pre_money + investment
            ownership["C16"] = total_shares
            ownership["C17"] = investor_ownership
            ownership["C18"] = founder_ownership
            ownership["C19"] = pool_ownership
            ownership["C20"] = investment / new
            waterfall["C11"] = investor_ownership
            waterfall["C12"] = waterfall["C5"].value * investor_ownership
            waterfall["C13"] = waterfall["C12"].value / waterfall["C6"].value
            return workbook

        with patch.object(
            verify_reference_calcs, "with_recalc", side_effect=recalculate
        ):
            name, passed, details = (
                verify_reference_calcs.check_vc_waterfall_conservation()
            )

        self.assertEqual(
            name, "VC: ownership, pricing, and exit-proceeds identities"
        )
        self.assertTrue(passed, details)
        self.assertIn("ownership_sum=1.0", details)

    def test_vc_holder_election_oracle_covers_mixed_conversion_states(self):
        def recalculate(_path, populate):
            workbook = Workbook()
            workbook.remove(workbook.active)
            workbook.create_sheet("Cap Table")
            waterfall = workbook.create_sheet("Exit Waterfall")
            populate(workbook)
            waterfall["C17"] = "SUPPORTED"
            waterfall["I28"] = 1
            waterfall["K28"] = 3
            base_payouts = (5_000_000.0, 3_000_000.0, 1_500_000.0, 10_500_000.0)
            # The explicit values below are the independently enumerated mask-3
            # equilibrium: B retains preference while A and Seed convert.
            adverse_payouts = (
                5_000_000.0,
                45_000_000.0 * 1_500_000.0 / 9_500_000.0,
                45_000_000.0 * 1_000_000.0 / 9_500_000.0,
                45_000_000.0 * 7_000_000.0 / 9_500_000.0,
            )
            for row, payout, election in zip(
                range(21, 24),
                base_payouts[:3],
                ("PREFERENCE", "PREFERENCE", "CONVERT"),
            ):
                waterfall.cell(row, 8, election)
                waterfall.cell(row, 9, payout)
            waterfall.cell(24, 9, base_payouts[3])
            for row, payout, election in zip(
                range(21, 24),
                adverse_payouts[:3],
                ("PREFERENCE", "CONVERT", "CONVERT"),
            ):
                waterfall.cell(row, 10, election)
                waterfall.cell(row, 11, payout)
            waterfall.cell(24, 11, adverse_payouts[3])
            waterfall["I26"] = sum(base_payouts)
            waterfall["K26"] = sum(adverse_payouts)
            return workbook

        with patch.object(
            verify_reference_calcs, "with_recalc", side_effect=recalculate
        ):
            name, passed, details = (
                verify_reference_calcs.check_vc_holder_election_waterfall()
            )

        self.assertEqual(
            name, "VC: holder-by-holder liquidation-preference equilibrium"
        )
        self.assertTrue(passed, details)
        self.assertIn("Adverse mask=3", details)


if __name__ == "__main__":
    unittest.main()
