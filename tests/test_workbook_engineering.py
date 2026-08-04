from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from tools.workbook_engineering import audit_workbook
from tools.workbook_parity import compare_workbooks, normalize_formula


class WorkbookEngineeringTests(unittest.TestCase):
    def create_book(self, path: Path, formula: str = "=B1+C1") -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Model"
        sheet["B1"] = 1
        sheet["C1"] = 2
        sheet["D1"] = formula
        sheet.freeze_panes = "B2"
        workbook.save(path)

    def test_semantic_parity(self):
        with TemporaryDirectory() as directory:
            left = Path(directory) / "left.xlsx"
            right = Path(directory) / "right.xlsx"
            self.create_book(left)
            self.create_book(right)
            result = compare_workbooks(left, right)
            self.assertTrue(result["semantic_parity"])
            self.assertTrue(result["presentation_parity"])

    def test_formula_drift_has_cell_detail(self):
        with TemporaryDirectory() as directory:
            left = Path(directory) / "left.xlsx"
            right = Path(directory) / "right.xlsx"
            self.create_book(left)
            self.create_book(right, "=B1-C1")
            result = compare_workbooks(left, right)
            self.assertFalse(result["semantic_parity"])
            self.assertIn("Model:cells", result["semantic_differences"])
            self.assertEqual(result["semantic_cell_differences"][0]["cell"], "D1")

    def test_equivalent_formula_serialization(self):
        self.assertEqual(normalize_formula("='VaR'!C5*FALSE()*9E+099"), "=VaR!C5*FALSE*9E99")
        self.assertEqual(
            normalize_formula('=MAX(0,MIN(0.60,A1+0.10))+"keep 0.60"'),
            '=MAX(0,MIN(0.6,A1+0.1))+"keep 0.60"',
        )
        with TemporaryDirectory() as directory:
            left = Path(directory) / "left.xlsx"
            right = Path(directory) / "right.xlsx"
            self.create_book(left, "='Model'!B1+9E99+FALSE()+0.60")
            self.create_book(right, "=Model!B1+9E+099+FALSE+0.6")
            self.assertTrue(compare_workbooks(left, right)["semantic_parity"])

    def test_float_round_trip_noise_is_not_semantic_drift(self):
        with TemporaryDirectory() as directory:
            left = Path(directory) / "left.xlsx"
            right = Path(directory) / "right.xlsx"
            self.create_book(left)
            self.create_book(right)
            left_book = load_workbook(left)
            right_book = load_workbook(right)
            left_book["Model"]["E1"] = 0.05499999999999999
            right_book["Model"]["E1"] = 0.055
            left_book.save(left)
            right_book.save(right)
            self.assertTrue(compare_workbooks(left, right)["semantic_parity"])

    def test_material_hardcode_drift_still_fails(self):
        with TemporaryDirectory() as directory:
            left = Path(directory) / "left.xlsx"
            right = Path(directory) / "right.xlsx"
            self.create_book(left)
            self.create_book(right)
            left_book = load_workbook(left)
            right_book = load_workbook(right)
            left_book["Model"]["E1"] = 0.055
            right_book["Model"]["E1"] = 0.056
            left_book.save(left)
            right_book.save(right)
            result = compare_workbooks(left, right)
            self.assertFalse(result["semantic_parity"])
            self.assertEqual(result["semantic_cell_differences"][0]["cell"], "E1")

    def test_presentation_drift_does_not_fail_semantics(self):
        with TemporaryDirectory() as directory:
            left = Path(directory) / "left.xlsx"
            right = Path(directory) / "right.xlsx"
            self.create_book(left)
            self.create_book(right)
            workbook = load_workbook(right)
            workbook["Model"]["D1"].font = Font(bold=True)
            workbook.save(right)
            result = compare_workbooks(left, right)
            self.assertTrue(result["semantic_parity"])
            self.assertFalse(result["presentation_parity"])

    def test_audit_external_and_volatile_formula(self):
        with TemporaryDirectory() as directory:
            path = Path(directory) / "book.xlsx"
            self.create_book(path, "=OFFSET(B1,0,0)+[other.xlsx]Sheet1!A1")
            _, findings = audit_workbook(path)
            codes = {finding.code for finding in findings}
            self.assertIn("volatile_formula", codes)
            self.assertIn("external_formula_reference", codes)


if __name__ == "__main__":
    unittest.main()
