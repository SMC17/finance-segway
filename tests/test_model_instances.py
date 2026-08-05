from pathlib import Path
from tempfile import TemporaryDirectory
import json
import unittest

from openpyxl import Workbook, load_workbook

from tools.model_instances import apply_manifest
from tools.model_instance_release import apply_manifest as apply_release_manifest


class ModelInstanceTests(unittest.TestCase):
    def create_template(self, root: Path) -> None:
        workbook = Workbook()
        cover = workbook.active
        cover.title = "Cover"
        cover["B4"] = "Last refreshed:"
        cover["C4"] = "[date]"
        cover["B9"] = "Active scenario:"
        cover["C9"] = "Base"
        assumptions = workbook.create_sheet("Assumptions")
        assumptions["C5"] = 100
        assumptions["D5"] = "=C5*2"
        sources = workbook.create_sheet("Sources")
        sources.append([None, "Input", "URL", "As-of", "Notes"])
        refresh = workbook.create_sheet("RefreshLog")
        refresh.append([None, "Date", "Trigger", "What changed", "Reviewer notes", "Next check"])
        workbook.save(root / "template.xlsx")

    def write_manifest(self, root: Path, *, cell: str = "C5") -> Path:
        manifest = {
            "schema_version": "1.0",
            "id": "test-instance",
            "template": "template.xlsx",
            "output": "instances/test.xlsx",
            "as_of": "2026-08-03",
            "scenario": "Downside",
            "inputs": [{
                "sheet": "Assumptions",
                "cell": cell,
                "value": 125,
                "source": {
                    "name": "Test source",
                    "url": "https://example.com/source",
                    "as_of": "2026-08-03",
                    "notes": "Unit-test fixture",
                },
            }],
        }
        path = root / "manifest.json"
        path.write_text(json.dumps(manifest), encoding="utf-8")
        return path

    def test_generate_instance_and_receipt(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            self.create_template(root)
            manifest = self.write_manifest(root)
            receipt = apply_manifest(manifest, root)
            output = root / "instances/test.xlsx"
            self.assertTrue(output.exists())
            self.assertEqual(len(receipt["workbook_sha256"]), 64)
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(workbook["Assumptions"]["C5"].value, 125)
            self.assertEqual(workbook["Cover"]["C4"].value, "2026-08-03")
            self.assertEqual(workbook["Cover"]["C9"].value, "Downside")
            self.assertEqual(workbook["Sources"]["C5"].value, "https://example.com/source")
            self.assertTrue(output.with_suffix(".receipt.json").exists())

    def test_formula_override_rejected(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            self.create_template(root)
            manifest = self.write_manifest(root, cell="D5")
            with self.assertRaisesRegex(ValueError, "refusing to overwrite formula"):
                apply_manifest(manifest, root)

    def test_unmapped_cover_key_fails_loudly_instead_of_silently_dropping(self):
        # Regression guard: a manifest "cover" key that matches no Cover
        # sheet row used to be silently skipped, so a workbook could carry
        # real, sourced financial data while its Cover sheet still showed
        # the template's literal placeholder text. Confirmed live across
        # every public case in the repo before this test was added.
        with TemporaryDirectory() as directory:
            root = Path(directory)
            self.create_template(root)
            manifest = self.write_manifest(root)
            payload = json.loads(manifest.read_text(encoding="utf-8"))
            payload["cover"] = {"Subject:": "This label does not exist on Cover"}
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "matches no Cover sheet row"):
                apply_manifest(manifest, root)

    def test_validate_only_does_not_write(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            self.create_template(root)
            manifest = self.write_manifest(root)
            result = apply_manifest(manifest, root, validate_only=True)
            self.assertTrue(result["valid"])
            self.assertFalse((root / "instances/test.xlsx").exists())

    def test_public_release_rejects_synthetic_benchmark_lineage(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            self.create_template(root)
            manifest = self.write_manifest(root)
            payload = json.loads(manifest.read_text(encoding="utf-8"))
            payload["classification"] = "external_historical_case"
            payload["counts_toward_M4"] = False
            payload["inputs"][0]["source"]["url"] = (
                "repo://standards/benchmark_cases/test-instance.json"
            )
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "synthetic benchmark lineage"):
                apply_release_manifest(manifest, root)

    def test_public_release_receipt_carries_claim_boundary(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            self.create_template(root)
            manifest = self.write_manifest(root)
            payload = json.loads(manifest.read_text(encoding="utf-8"))
            payload["classification"] = "external_historical_case"
            payload["counts_toward_M4"] = False
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            receipt = apply_release_manifest(manifest, root)
            self.assertEqual(
                receipt["classification"], "external_historical_case"
            )
            self.assertIs(receipt["counts_toward_M4"], False)


if __name__ == "__main__":
    unittest.main()
