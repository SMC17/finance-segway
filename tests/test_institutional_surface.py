from pathlib import Path
from tempfile import TemporaryDirectory
import json
import unittest

from openpyxl import Workbook, load_workbook

from tools.institutional_surface import (
    SURFACE_SHEETS,
    apply_surface,
    load_inventory,
    profiles_by_id,
    validate_profiles,
    validate_workbook_surface,
)

ROOT = Path(__file__).resolve().parents[1]


class InstitutionalSurfaceTests(unittest.TestCase):
    def test_registry_covers_inventory(self):
        self.assertEqual(validate_profiles(ROOT), [])
        inventory = load_inventory(ROOT)
        profiles = profiles_by_id(ROOT)
        self.assertEqual(set(profiles), {item["id"] for item in inventory["models"]})

    def test_surface_is_idempotent_and_preserves_formulas(self):
        model = load_inventory(ROOT)["models"][0]
        profile = profiles_by_id(ROOT)[model["id"]]
        with TemporaryDirectory() as directory:
            path = Path(directory) / "model.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Cover"
            sheet["A1"] = "=1+1"
            workbook.save(path)
            apply_surface(path, model, profile)
            apply_surface(path, model, profile)
            self.assertEqual(validate_workbook_surface(path, model, profile), [])
            result = load_workbook(path, data_only=False)
            self.assertEqual(result["Cover"]["A1"].value, "=1+1")
            for name in SURFACE_SHEETS:
                self.assertEqual(result.sheetnames.count(name), 1)
            self.assertEqual(result["Institutional Surface"]["C4"].value, model["id"])
            self.assertIn("FS_MODEL_ID", result.defined_names)
            self.assertIn("FS_DOMAIN", result.defined_names)
            self.assertIn("FS_MATURITY", result.defined_names)

    def test_profile_validation_rejects_missing_domain(self):
        with TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "standards/domain_profiles").mkdir(parents=True)
            inventory = {"version": "test", "models": [{
                "id": "01", "domain": "Investment Banking",
                "declared_maturity": "M1", "builder": "builder.py",
                "workbook": "book.xlsx",
            }]}
            (root / "standards/model_inventory.json").write_text(json.dumps(inventory), encoding="utf-8")
            (root / "standards/domain_profiles/empty.tsv").write_text(
                "id\tdomain\tarena\tartifact\tcadence\toutputs\tdocuments\tscenarios\tchallenges\tfailures\tquestions\tsources\tanchors\n",
                encoding="utf-8",
            )
            self.assertIn("missing_profile:01", validate_profiles(root))


if __name__ == "__main__":
    unittest.main()
