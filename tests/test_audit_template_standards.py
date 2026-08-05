"""Verify the standards audit catches real defects and doesn't cry wolf on
the sanctioned non-circular debt-schedule pattern this repo actually uses.

Built on synthetic workbooks constructed here, not on the repo's own
templates, so a bug in the audit logic can't hide behind "the templates
happen to already look clean."
"""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "tools"))

import openpyxl  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

from audit_template_standards import audit_template  # noqa: E402

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")


def _write(wb, path: Path) -> Path:
    wb.save(path)
    return path


class ColorConventionTests(unittest.TestCase):
    def test_flags_formula_rendered_in_input_blue(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Operating Model"
        ws["B5"] = "Revenue"
        c = ws["C5"]
        c.value = "=C4*1.05"
        c.font = BLUE  # bug: a formula, but styled as if it were a hardcode
        with tempfile.TemporaryDirectory() as tmp:
            path = _write(wb, Path(tmp) / "t.xlsx")
            result = audit_template(path, "TEST")
        self.assertTrue(any("input-blue font" in f for f in result.findings))

    def test_does_not_flag_correctly_colored_formula(self) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Operating Model"
        ws["C5"] = "=C4*1.05"
        ws["C5"].font = BLACK
        with tempfile.TemporaryDirectory() as tmp:
            path = _write(wb, Path(tmp) / "t.xlsx")
            result = audit_template(path, "TEST")
        self.assertFalse(any("input-blue font" in f for f in result.findings))


class CircularitySwitchTests(unittest.TestCase):
    def _debt_schedule_wb(self, *, circular: bool):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Debt Schedule")
        ws["B5"] = "Beginning TLB"
        ws["C5"] = 100
        ws["D5"] = "=C7"  # prior period's ending balance -- always non-circular
        ws["B6"] = "TLB interest"
        ws["C6"] = "=C5*0.08"
        ws["D6"] = "=D5*0.08"
        ws["B7"] = "Ending TLB"
        ws["C7"] = "=C5-10"
        if circular:
            # same-period self-reference: interest depends on the average of
            # beginning and ending balance in the *same* column -- genuinely
            # circular, and exactly the pattern that needs a labeled switch.
            ws["D7"] = "=D5-D6-10"
            ws["D6"] = "=AVERAGE(D5,D7)*0.08"
        else:
            ws["D7"] = "=D5-10"
        return wb

    def test_flags_genuine_same_period_circularity_without_switch(self) -> None:
        wb = self._debt_schedule_wb(circular=True)
        with tempfile.TemporaryDirectory() as tmp:
            path = _write(wb, Path(tmp) / "t.xlsx")
            result = audit_template(path, "TEST")
        self.assertTrue(any("circularity-breaker switch" in f for f in result.findings))

    def test_does_not_flag_opening_balance_pattern(self) -> None:
        """This is the pattern this repo's own LBO/credit/REIT debt
        schedules actually use: interest on the strictly-prior-period
        balance. It's a WSP-sanctioned non-circular resolution and must not
        be flagged as if it were missing a switch it doesn't need."""
        wb = self._debt_schedule_wb(circular=False)
        with tempfile.TemporaryDirectory() as tmp:
            path = _write(wb, Path(tmp) / "t.xlsx")
            result = audit_template(path, "TEST")
        self.assertFalse(any("circularity-breaker switch" in f for f in result.findings))

    def test_switch_present_suppresses_finding_even_if_circular(self) -> None:
        wb = self._debt_schedule_wb(circular=True)
        ws = wb["Debt Schedule"]
        ws["B10"] = "Circularity breaker switch"
        with tempfile.TemporaryDirectory() as tmp:
            path = _write(wb, Path(tmp) / "t.xlsx")
            result = audit_template(path, "TEST")
        self.assertFalse(any("circularity-breaker switch" in f for f in result.findings))


class CovenantHeadroomTests(unittest.TestCase):
    def test_flags_covenants_sheet_without_headroom_output(self) -> None:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Covenants")
        ws["B5"] = "Leverage ratio"
        ws["C5"] = "=4.5"
        ws["B6"] = "Maximum leverage covenant"
        ws["C6"] = 6.0
        with tempfile.TemporaryDirectory() as tmp:
            path = _write(wb, Path(tmp) / "t.xlsx")
            result = audit_template(path, "TEST")
        self.assertTrue(any("headroom" in f for f in result.findings))

    def test_does_not_flag_covenants_sheet_with_headroom(self) -> None:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Covenants")
        ws["B5"] = "Leverage ratio"
        ws["C5"] = "=4.5"
        ws["B6"] = "Covenant headroom (%)"
        ws["C6"] = "=(6.0-C5)/6.0"
        with tempfile.TemporaryDirectory() as tmp:
            path = _write(wb, Path(tmp) / "t.xlsx")
            result = audit_template(path, "TEST")
        self.assertFalse(any("headroom" in f for f in result.findings))


class RealTemplateRegressionTests(unittest.TestCase):
    """Confirm the three real bugs this audit found and fixed this session
    stay fixed in the committed templates."""

    def test_all_committed_templates_currently_clean(self) -> None:
        from audit_template_standards import audit_all

        results = audit_all()
        self.assertEqual(len(results), 25)
        failing = {r.domain: r.findings for r in results if r.findings}
        self.assertEqual(failing, {})


if __name__ == "__main__":
    unittest.main()
